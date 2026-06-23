#!/usr/bin/env python3
"""Generate the synthetic Lane-3 freeform cap-table fixture (cap-table-lane3-freeform).

SYNTHETIC ONLY — fictional Cadence / invented founders + numbers; never real founder data.
Models a founder's own freeform workbook with three tabs:
  * "Cap Table" — two founders + common share counts (founders_block; clean, unambiguous
    cap base — no ownership-% column, so no cap-base reconciliation gate fires)
  * "SAFEs" — one complete post-money SAFE (safes_block; all required fields present → no blocker)
  * "Returns" — an exit-scenario modeling tab the structure-detection sub-agent must classify as
    an ignore type (derived_calculation), NOT invent a block_type for

Regenerate: `python gen_lane3_fixture.py`
"""

from __future__ import annotations

import os

from openpyxl import Workbook  # type: ignore[import-untyped]

HERE = os.path.dirname(os.path.abspath(__file__))


def build() -> Workbook:
    wb = Workbook()

    cap = wb.active
    cap.title = "Cap Table"
    cap.append(["Holder", "Common Shares"])
    cap.append(["Maya Gold", 4_000_000])
    cap.append(["Ofer Tal", 4_000_000])

    safes = wb.create_sheet("SAFEs")
    safes.append(["Investor", "Amount", "Post-Money Cap", "Discount", "Date"])
    safes.append(["Foobar Capital LLC", 500_000, 10_000_000, 0.20, "2026-01-15"])

    returns = wb.create_sheet("Returns")
    returns.append(["Exit Value", "Founder Proceeds", "SAFE Proceeds"])
    for exit_v in (20_000_000, 50_000_000, 100_000_000):
        # illustrative computed columns — a modeling tab, not equity holdings
        returns.append([exit_v, round(exit_v * 0.85), round(exit_v * 0.15)])

    return wb


if __name__ == "__main__":
    out = os.path.join(HERE, "freeform_lane3.xlsx")
    build().save(out)
    print(f"wrote {out}")
