#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Extract structured data from Excel (.xlsx) or CSV files.

Usage:
    python extract_model.py --file model.xlsx --pretty
    python extract_model.py --file data.csv -o model_data.json
    echo '{"sheets": [...]}' | python extract_model.py --stdin

Output: JSON with structure:
    {"sheets": [{"name": str, "headers": [str], "rows": [[value]], "detected_type": str|null}]}
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from typing import Any


def _write_output(data: str, output_path: str | None, *, summary: dict[str, Any] | None = None) -> None:
    if output_path:
        abs_path = os.path.abspath(output_path)
        parent = os.path.dirname(abs_path)
        if parent == "/":
            print(f"Error: output path resolves to root directory: {abs_path}", file=sys.stderr)
            sys.exit(1)
        os.makedirs(parent, exist_ok=True)
        with open(abs_path, "w", encoding="utf-8") as f:
            f.write(data)
        receipt: dict[str, Any] = {"ok": True, "path": abs_path, "bytes": len(data.encode("utf-8"))}
        if summary:
            receipt.update(summary)
        sys.stdout.write(json.dumps(receipt, separators=(",", ":")) + "\n")
    else:
        sys.stdout.write(data)


# Tab name heuristics for detecting sheet purpose
_TAB_PATTERNS: dict[str, list[str]] = {
    "assumptions": ["assumption", "input", "driver", "parameter"],
    "revenue": ["revenue", "sales", "arr", "mrr", "income"],
    "expenses": ["expense", "opex", "cost", "headcount", "hiring", "payroll"],
    "cash": ["cash", "runway", "burn", "balance"],
    "pnl": ["p&l", "pnl", "profit", "loss", "income statement"],
    "summary": ["summary", "dashboard", "overview", "kpi"],
    "scenarios": ["scenario", "sensitivity", "case"],
}


def _detect_tab_type(name: str) -> str | None:
    lower = name.lower().strip()
    for tab_type, patterns in _TAB_PATTERNS.items():
        for pat in patterns:
            if pat in lower:
                return tab_type
    return None


def _safe_value(val: Any) -> Any:
    """Convert cell value to JSON-serializable type."""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    return str(val)


def extract_xlsx(file_path: str) -> dict[str, Any]:
    """Extract data from an Excel file."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        print(
            "Error: openpyxl is required for .xlsx files. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows_data: list[list[Any]] = []
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_vals = [_safe_value(c) for c in row]
            if i == 0:
                headers = [str(v) if v is not None else f"col_{j}" for j, v in enumerate(row)]
            else:
                rows_data.append(row_vals)
        sheets.append(
            {
                "name": ws.title,
                "headers": headers,
                "rows": rows_data,
                "detected_type": _detect_tab_type(ws.title),
                "row_count": len(rows_data),
                "col_count": len(headers),
            }
        )
    wb.close()
    return {"sheets": sheets, "source_format": "xlsx", "source_file": os.path.basename(file_path)}


def extract_csv(file_path: str) -> dict[str, Any]:
    """Extract data from a CSV file."""
    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows_raw = list(reader)

    if not rows_raw:
        return {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": [],
                    "rows": [],
                    "detected_type": None,
                    "row_count": 0,
                    "col_count": 0,
                }
            ],
            "source_format": "csv",
            "source_file": os.path.basename(file_path),
        }

    headers = rows_raw[0]
    rows_data = []
    for row in rows_raw[1:]:
        row_vals: list[Any] = []
        for v in row:
            # Try to coerce to number
            try:
                row_vals.append(int(v))
            except ValueError:
                try:
                    row_vals.append(float(v))
                except ValueError:
                    row_vals.append(v if v else None)
        rows_data.append(row_vals)

    name = os.path.splitext(os.path.basename(file_path))[0]
    return {
        "sheets": [
            {
                "name": name,
                "headers": headers,
                "rows": rows_data,
                "detected_type": _detect_tab_type(name),
                "row_count": len(rows_data),
                "col_count": len(headers),
            }
        ],
        "source_format": "csv",
        "source_file": os.path.basename(file_path),
    }


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Extract structured data from financial model files")
    group = p.add_mutually_exclusive_group(required=True)
    group.add_argument("--file", help="Path to .xlsx or .csv file")
    group.add_argument("--stdin", action="store_true", help="Read pre-structured JSON from stdin")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON output")
    p.add_argument("-o", "--output", help="Write output to file instead of stdout")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    if args.stdin:
        if sys.stdin.isatty():
            print("Error: --stdin requires piped input", file=sys.stderr)
            sys.exit(1)
        try:
            data = json.load(sys.stdin)
        except json.JSONDecodeError as e:
            print(f"Error: invalid JSON on stdin: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        file_path = args.file
        if not os.path.isfile(file_path):
            print(f"Error: file not found: {file_path}", file=sys.stderr)
            sys.exit(1)

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xlsx":
            data = extract_xlsx(file_path)
        elif ext == ".csv":
            data = extract_csv(file_path)
        else:
            print(f"Error: unsupported file type '{ext}' (expected .xlsx or .csv)", file=sys.stderr)
            sys.exit(1)

    indent = 2 if args.pretty else None
    out = json.dumps(data, indent=indent) + "\n"
    _write_output(
        out,
        args.output,
        summary={"sheets": len(data.get("sheets", []))},
    )


if __name__ == "__main__":
    main()
