#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Validator + cap-table extraction for Lanes 2/3/4.

Modes:
  * --mode=validate: read existing inputs.json + instruments.json from --dir
    and schema-validate them. Pure check; useful for Lane 4 (founder pastes
    structured JSON) or as a CI gate after extraction.
  * --mode=carta: extract from a Carta XLSX export (Summary Cap Table +
    optional Intermediate / Detailed / per-class Ledgers / Convertible
    Ledger / Equity Incentive Plan sheets). Real Carta column conventions
    verified against a corpus of real exports (see
    references/carta-pulley-mapping.md).
  * --mode=pulley: stub — Pulley column mapping is Phase 1 follow-up
    (no real Pulley exports in the corpus yet to verify against).
  * --mode=freeform: validates Context-A SPREADSHEET_STRUCTURE_DETECTION
    sub-agent output; the actual cell-mapping happens in the sub-agent.

Carta extractor implementation notes (per real-world corpus):
  * Carta puts a banner in rows 2-3; real headers are in row 5.
  * Per-class ledger naming follows `{ClassPrefix} Ledger`
    (CS Ledger, PS1 Ledger, PA1 Ledger, etc.). Larger exports include
    them all; minimal exports have only Summary / Intermediate / Detailed.
  * Convertible Ledger contains SAFEs (security_id prefix "SAFE*") and
    notes in the same sheet; "Note Block Name" groups by SAFE template.
  * Conversion Discount is stored as percent (e.g. 0.2 = 20%);
    normalized to multiplier form per Gotcha #3.
"""

from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys
import warnings
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _artifact_writer import load_schema  # noqa: E402
from _cap_table_schema_validator import validate  # noqa: E402


class CartaFingerprintMismatchError(ValueError):
    """Raised by _carta_extract when --mode=carta is explicit but the workbook
    doesn't match Carta's verified sheet fingerprint. Caller (_mode_carta)
    catches this and emits an E_CARTA_FINGERPRINT_MISMATCH blocker receipt
    with exit code 1."""


# Carta export sheet fingerprints (verified against real exports in
# captable corpus). The "primary" check uses just `Summary Cap Table` for
# minimal exports; the "full" check tightens the match.
CARTA_PRIMARY_FINGERPRINT = {"Summary Cap Table"}
CARTA_FULL_FINGERPRINT = {
    "Summary Cap Table",
    "Intermediate Cap Table",
    "Detailed Cap Table",
}
# OCX (Open Cap Table eXcel) — Carta's optional standardized format
CARTA_OCX_FINGERPRINT = {"Capitalization by Stakeholder", "Voting Details", "Context"}
PULLEY_FINGERPRINT = {"Ownership"}
PULLEY_CONTRACT_TABS = {
    "Shares",
    "SAFEs",
    "Convertible Notes",
    "Stock Options",
    "RSAs",
    "RSUs",
    "Warrants",
}

# Carta convertible-ledger column names (row 5 headers; verified)
CARTA_CONVERTIBLE_COLUMNS = [
    "Formatted Security ID",
    "Security ID",
    "Stakeholder Name",
    "Stakeholder Email",
    "Principal",
    "Other Consideration",
    "Interest",
    "Total",
    "Destination",
    "Issue Date",
    "Board Approval Date",
    "Termination Date",
    "Canceled Date",
    "Cancellation Reason",
    "Converted Date",
    "Maturity Date",
    "Interest Rate",
    "Valuation Cap",
    "Conversion Discount",
    "Change In Control Percent",
    "Conversion Trigger",
    "Note Block Name",
]

# Heuristic: a SAFE has security_id prefix matching this pattern; everything
# else in Convertible Ledger is a convertible note.
_SAFE_PREFIX_RE = re.compile(r"^SAFE\d*-\d+$", re.IGNORECASE)

_SCHEMA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "references",
    "schemas",
)


def _validate_file(path: str, schema_path: str) -> list[str]:
    if not os.path.exists(path):
        return [f"file missing: {path}"]
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    schema = load_schema(schema_path)
    return validate(data, schema)


def _mode_validate(args: argparse.Namespace) -> int:
    """Read inputs.json + instruments.json + cap_state.json from --dir and validate."""
    errors_by_file = {}
    for fname, sname in [
        ("inputs.json", "inputs.schema.json"),
        ("instruments.json", "instruments.schema.json"),
        ("cap_state.json", "cap_state.schema.json"),
    ]:
        path = os.path.join(args.dir, fname)
        if os.path.exists(path):
            errs = _validate_file(path, os.path.join(_SCHEMA_DIR, sname))
            if errs:
                errors_by_file[fname] = errs

    if errors_by_file:
        sys.stderr.write("extract_cap_table.py validate: errors found\n")
        for fname, errs in errors_by_file.items():
            sys.stderr.write(f"  {fname}:\n")
            for e in errs:
                sys.stderr.write(f"    - {e}\n")
        print(json.dumps({"ok": False, "errors": errors_by_file}))
        return 1
    print(
        json.dumps(
            {
                "ok": True,
                "validated": list(
                    set(["inputs.json", "instruments.json", "cap_state.json"]) & set(os.listdir(args.dir))
                ),
            }
        )
    )
    return 0


def _open_xlsx(path: str) -> Any:
    """Open an XLSX via BytesIO bypass (handles .xlsx(N) macOS dupes)."""
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        with open(path, "rb") as fp:
            return openpyxl.load_workbook(io.BytesIO(fp.read()), data_only=True)


def _carta_detect(sheet_names: list[str]) -> str | None:
    """Return 'carta' if Carta-default; 'carta_ocx' if OCX; None otherwise."""
    sheets = {s.strip() for s in sheet_names}
    if CARTA_OCX_FINGERPRINT.issubset(sheets):
        return "carta_ocx"
    if CARTA_FULL_FINGERPRINT.issubset(sheets) or CARTA_PRIMARY_FINGERPRINT.issubset(sheets):
        return "carta"
    return None


def _find_header_row(ws: Any, max_scan: int = 8) -> int:
    """Find the row with the most short-string cells (the header row).

    Carta puts a banner in rows 2-3 and headers in row 5. We don't
    hard-code row 5 because some Carta exports may shift; the heuristic
    matches in practice."""
    best_row = 1
    best_score = -1
    for r_idx in range(1, max_scan + 1):
        try:
            row = next(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True), ())
            strs = [str(c).strip() for c in row if c is not None]
            score = sum(1 for s in strs if 2 <= len(s) <= 50 and len(s.split()) <= 6)
            if score > best_score:
                best_score = score
                best_row = r_idx
        except StopIteration:
            break
    return best_row


def _normalize_discount(d: Any) -> tuple[float | None, str | None]:
    """Per Gotcha #3: convert percent to multiplier form."""
    if d is None or d == "":
        return None, None
    try:
        v = float(d)
    except (TypeError, ValueError):
        return None, f"discount value {d!r} not numeric"
    if 0 < v <= 1.0:
        # Could be already-multiplier (0.80 = 20% discount) OR percent-as-fraction (0.20 = 20% discount)
        # Carta stores 0.20 to mean 20% discount (not 80% multiplier). Convert to multiplier form.
        return 1.0 - v, f"Carta discount {v} (= {v * 100:.0f}%) converted to multiplier {1.0 - v:.4f}"
    if 1 < v <= 100:
        return 1.0 - (v / 100.0), f"Carta discount {v}% converted to multiplier {1.0 - v / 100.0:.4f}"
    return None, f"discount value {v} out of expected range"


def _infer_safe_form(cap: Any, discount: Any) -> str:
    """Infer the SAFE form from cap + discount presence."""
    has_cap = cap is not None and cap != "" and float(cap) > 0
    has_disc = discount is not None and discount != "" and float(discount) > 0
    if has_cap and has_disc:
        return "cap_plus_discount"
    if has_cap and not has_disc:
        return "yc_postmoney_cap"
    if not has_cap and has_disc:
        return "yc_postmoney_discount"
    return "other"


def _extract_convertible_ledger(ws: Any) -> list[dict[str, Any]]:
    """Parse the Convertible Ledger sheet. Returns a list of raw records."""
    header_row = _find_header_row(ws)
    rows = list(ws.iter_rows(values_only=True))
    if header_row > len(rows):
        return []
    headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[header_row - 1])]
    records = []
    for r in rows[header_row:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        rec = dict(zip(headers, r, strict=False))
        records.append(rec)
    return records


def _convertible_record_to_instrument(rec: dict[str, Any], idx: int) -> tuple[str, dict[str, Any], list[str]]:
    """Convert a Carta Convertible Ledger row to instruments.json entry.

    Returns (kind, instrument_dict, warnings) where kind is "safe" or "note".
    """
    warnings_list: list[str] = []
    sec_id = (rec.get("Security ID") or "").strip()
    is_safe = bool(_SAFE_PREFIX_RE.match(sec_id))

    # Skip cancelled / converted records (these are historical)
    converted = rec.get("Converted Date")
    cancelled = rec.get("Canceled Date")
    if converted or cancelled:
        warnings_list.append(f"{sec_id}: skipped (converted={converted!r} cancelled={cancelled!r})")
        return ("skip", {}, warnings_list)

    def _to_iso_date(v: Any) -> str | None:
        """Convert a Carta date cell to ISO 8601 string (or None)."""
        if v is None or v == "":
            return None
        if hasattr(v, "date"):  # datetime
            result: str = v.date().isoformat()
            return result
        if hasattr(v, "isoformat"):  # date
            result_d: str = v.isoformat()
            return result_d
        return str(v)[:10]

    investor_name = (rec.get("Stakeholder Name") or "").strip()
    issue_date = _to_iso_date(rec.get("Issue Date")) or "1900-01-01"

    principal = rec.get("Principal") or 0
    interest_rate = rec.get("Interest Rate") or 0  # decimal already; e.g. 0.06
    valuation_cap = rec.get("Valuation Cap")
    if valuation_cap == 0:
        valuation_cap = None
    discount_raw = rec.get("Conversion Discount")
    discount_mult, disc_warning = _normalize_discount(discount_raw)
    if disc_warning:
        warnings_list.append(f"{sec_id}: {disc_warning}")
    # SAFE has no maturity; use sentinel. Notes may have a real date.
    maturity_date = _to_iso_date(rec.get("Maturity Date")) or "9999-12-31"

    if is_safe:
        form = _infer_safe_form(valuation_cap, discount_raw)
        return (
            "safe",
            {
                "id": f"safe_{idx:03d}",
                "investor_name": investor_name,
                "purchase_amount": float(principal),
                "post_money_valuation_cap": float(valuation_cap) if valuation_cap else None,
                "discount_multiplier": discount_mult,
                "mfn_provision": None,
                "pro_rata_side_letter": None,
                "issuance_date": issue_date,
                "form": form,
                "conversion_price_override": None,
                "source_document": f"carta:{sec_id}",
                "extraction_confidence": "high",
            },
            warnings_list,
        )
    # Convertible note
    return (
        "note",
        {
            "id": f"note_{idx:03d}",
            "investor_name": investor_name,
            "principal": float(principal),
            "annual_interest_rate": float(interest_rate),
            "day_count_basis": 365,
            "compounding_periods_per_year": None,
            "interest_converts_to_shares": True,
            "issuance_date": issue_date,
            "last_interest_event_date": None,
            "valuation_cap": float(valuation_cap) if valuation_cap else None,
            "discount_multiplier": discount_mult,
            "capitalization_denominator": None,
            "capitalization_denominator_policy": "Carta-supplied: confirm with note text",
            "qualified_financing_threshold": 1_000_000.0,  # default; confirm
            "maturity_date": maturity_date,
            "maturity_default_treatment": "convert_at_cap",
            "maturity_conversion_price_override": None,
            "non_qualified_financing_treatment": None,
            "source_document": f"carta:{sec_id}",
            "extraction_confidence": "high",
        },
        warnings_list,
    )


def _carta_extract(xlsx_path: str) -> dict[str, Any]:
    """Extract structured data from a Carta XLSX into our canonical format.

    Returns a dict with `inputs_seed`, `instruments`, `summary_totals`,
    `warnings`, `format`, `sheets_consumed`. Caller persists to JSON.
    """
    wb = _open_xlsx(xlsx_path)
    sheet_names = wb.sheetnames
    format_detected = _carta_detect(sheet_names) or "unknown"
    warnings_list: list[str] = []
    sheets_consumed: list[str] = []

    # When --mode=carta is explicit but the sheet fingerprint doesn't match
    # Carta's verified shape ("Summary Cap Table" + "Convertible Ledger"),
    # fail loudly. The previous behavior returned ok=true, format="unknown"
    # while still running the Convertible Ledger consumer on whatever sheet
    # happened to be named that, producing nonsense (cancelled SAFEs not
    # skipped, discount normalization not applied, both rows classified as
    # notes). Raising E_CARTA_FINGERPRINT_MISMATCH tells the founder the
    # file isn't the expected shape so they can re-export or fall back to
    # --mode=freeform.
    if format_detected == "unknown":
        raise CartaFingerprintMismatchError(
            "E_CARTA_FINGERPRINT_MISMATCH: --mode=carta was specified but the workbook "
            f"sheet names {sheet_names!r} do not match Carta's verified Summary Cap Table + "
            "Convertible Ledger fingerprint. The export may be from a different version of "
            "Carta, a different vendor (Pulley/etc.), or a custom workbook. Re-export the "
            "Carta cap table or use --mode=freeform to dispatch a Context-A "
            "SPREADSHEET_STRUCTURE_DETECTION sub-agent."
        )

    # 1. Read Summary Cap Table for share-class totals
    summary_totals: dict[str, Any] = {}
    if "Summary Cap Table" in sheet_names:
        sheets_consumed.append("Summary Cap Table")
        ws = wb["Summary Cap Table"]
        # Extract company name from row 2 (banner: "{Company} Summary Cap Table")
        try:
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            banner = next((str(c) for c in row2 if c), "")
            company_match = re.match(r"^([^,]+?)\s+Summary Cap Table", banner)
            if company_match:
                summary_totals["company_name"] = company_match.group(1).strip()
        except StopIteration:
            pass
        # Extract as-of date from row 3
        try:
            row3 = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            row3_banner = next((str(c) for c in row3 if c), "")
            as_of_match = re.match(r"^As of (\S+)", row3_banner)
            if as_of_match:
                date_str = as_of_match.group(1)
                # Carta uses M/D/YYYY format; convert to ISO
                from datetime import datetime

                try:
                    parsed = datetime.strptime(date_str, "%m/%d/%Y")
                    summary_totals["as_of_date"] = parsed.date().isoformat()
                except ValueError:
                    summary_totals["as_of_date"] = date_str
        except StopIteration:
            pass

    # 2. Extract convertibles from Convertible Ledger (if present)
    instruments_safes = []
    instruments_notes = []
    if "Convertible Ledger" in sheet_names:
        sheets_consumed.append("Convertible Ledger")
        ws = wb["Convertible Ledger"]
        records = _extract_convertible_ledger(ws)
        for idx, rec in enumerate(records, start=1):
            kind, instrument, rec_warnings = _convertible_record_to_instrument(rec, idx)
            warnings_list.extend(rec_warnings)
            if kind == "safe":
                instruments_safes.append(instrument)
            elif kind == "note":
                instruments_notes.append(instrument)

    instruments = {
        "safes": instruments_safes,
        "convertible_notes": instruments_notes,
        "warrants": [],  # Lane-2 warrant extraction is a follow-up (per-class warrant ledgers)
        "option_grants": [],  # Lane-2 grant extraction is a follow-up (Equity Incentive Plan)
        "metadata": {},
    }

    return {
        "format": format_detected,
        "sheets_in_workbook": sheet_names,
        "sheets_consumed": sheets_consumed,
        "summary_totals": summary_totals,
        "instruments": instruments,
        "warnings": warnings_list,
        "counts": {
            "safes_extracted": len(instruments_safes),
            "notes_extracted": len(instruments_notes),
            "total_sheets": len(sheet_names),
        },
    }


def _mode_carta(args: argparse.Namespace) -> int:
    """Extract from a Carta XLSX export. Writes to --instruments path
    (appending if file already exists) and emits extraction_audit to -o."""
    if not args.xlsx:
        sys.stderr.write("--xlsx required for --mode=carta\n")
        return 1
    if not os.path.exists(args.xlsx):
        sys.stderr.write(f"file not found: {args.xlsx}\n")
        return 1
    try:
        result = _carta_extract(args.xlsx)
    except CartaFingerprintMismatchError as e:
        # Surface the structured blocker; do NOT run partial consumers.
        err_receipt: dict[str, Any] = {
            "ok": False,
            "mode": "carta",
            "blocker": "E_CARTA_FINGERPRINT_MISMATCH",
            "error": str(e)[:600],
            "remedy": (
                "The workbook isn't the Carta shape we extract from. Re-export from Carta, "
                "or fall back to --mode=freeform to dispatch Context-A SPREADSHEET_STRUCTURE_DETECTION."
            ),
        }
        print(json.dumps(err_receipt, indent=2))
        return 1
    except Exception as e:
        err_receipt = {
            "ok": False,
            "mode": "carta",
            "blocker": "carta_extraction_failed",
            "error": f"{type(e).__name__}: {e}"[:300],
            "remedy": "Fall back to --mode=freeform and dispatch Context-A SPREADSHEET_STRUCTURE_DETECTION.",
        }
        print(json.dumps(err_receipt, indent=2))
        return 1

    receipt: dict[str, Any] = {
        "ok": True,
        "mode": "carta",
        "format": result["format"],
        "summary": result["summary_totals"],
        "counts": result["counts"],
        "warnings": result["warnings"][:20],  # cap stderr output
        "sheets_consumed": result["sheets_consumed"],
    }

    # Optional: write extraction_audit.json with full data
    if args.output:
        audit_path = os.path.abspath(args.output)
        os.makedirs(os.path.dirname(audit_path) or ".", exist_ok=True)
        with open(audit_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, default=str)
        receipt["audit_written_to"] = audit_path

    # Optional: write instruments.json
    if hasattr(args, "instruments") and args.instruments:
        inst_path = os.path.abspath(args.instruments)
        os.makedirs(os.path.dirname(inst_path) or ".", exist_ok=True)
        # Merge with existing if present
        existing = {}
        if os.path.exists(inst_path):
            with open(inst_path, encoding="utf-8") as f:
                existing = json.load(f)
        existing_meta = dict(existing.get("metadata") or {})
        existing_meta["schema_version"] = "v0.5.0-instruments"
        merged = {
            "safes": existing.get("safes", []) + result["instruments"]["safes"],
            "convertible_notes": existing.get("convertible_notes", []) + result["instruments"]["convertible_notes"],
            "warrants": existing.get("warrants", []),
            "option_grants": existing.get("option_grants", []),
            "metadata": existing_meta,
        }
        with open(inst_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, indent=2, default=str)
        receipt["instruments_written_to"] = inst_path

    print(json.dumps(receipt, indent=2))
    if args.pretty and result["warnings"]:
        sys.stderr.write("  warnings:\n")
        for w in result["warnings"][:10]:
            sys.stderr.write(f"    {w}\n")
    return 0


def _mode_pulley_stub(args: argparse.Namespace) -> int:
    """Pulley not yet implemented; route to freeform."""
    receipt = {
        "ok": False,
        "mode": "pulley",
        "blocker": "pulley_mapping_not_yet_implemented",
        "remedy": (
            "Pulley extraction is a Phase 1 follow-up — no real Pulley exports "
            "in the test corpus to verify against. Run --mode=freeform and "
            "dispatch the Context A sub-agent SPREADSHEET_STRUCTURE_DETECTION."
        ),
    }
    print(json.dumps(receipt, indent=2))
    return 1


def _mode_freeform_validate(args: argparse.Namespace) -> int:
    """Validate Context-A sub-agent output for spreadsheet structure detection.

    Expects JSON on stdin matching the agent's SPREADSHEET_STRUCTURE_DETECTION
    return shape: {"blocks": [{"block_type", "sheet", "cell_range",
    "column_role_map", "confidence", "evidence"}, ...]}.

    For v0.1: validates the shape, surfaces ambiguities, but does NOT yet
    convert the cell-mapping to instruments.json automatically — that's
    Phase 1 follow-up work. The agent's mapping is captured to
    extraction_audit.json and the founder confirms before commit.
    """
    payload = json.load(sys.stdin)
    if "blocks" not in payload or not isinstance(payload["blocks"], list):
        sys.stderr.write("extract_cap_table.py freeform: expected {blocks: [...]} on stdin\n")
        return 1

    audit = {
        "mode": "freeform",
        "blocks_detected": len(payload["blocks"]),
        "blocks": payload["blocks"],
        "low_confidence_blocks": [b for b in payload["blocks"] if b.get("confidence") in {"low", "medium"}],
        "ambiguities": [a for b in payload["blocks"] for a in b.get("ambiguities", [])],
        "next_action": (
            "Present low_confidence_blocks + ambiguities to the founder via "
            "AskUserQuestion. Once confirmed, map each block into "
            "inputs.json + instruments.json (Phase 1 follow-up: automate the mapping)."
        ),
    }

    if args.output:
        out = os.path.abspath(args.output)
        os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            json.dump(audit, f, indent=2)
        audit["written_to"] = out
    print(json.dumps(audit, indent=2 if args.pretty else None))
    return 0


def _normalize_path(path: str) -> tuple[str, str | None]:
    """Return (normalized_path, warning_message).

    Handles real-world artifacts the corpus test surfaced:
      * macOS duplicate-download suffixes: `file.xlsx(1)`, `file.xlsx(2)` →
        treat as `.xlsx`. (openpyxl's extension check rejects them otherwise.)
      * Trailing whitespace in filenames.
    """
    import re

    warning: str | None = None
    norm = path.strip()
    # Strip macOS dupe suffix: ".xlsx(1)" → ".xlsx", ".pdf(2)" → ".pdf", etc.
    m = re.match(r"^(.*\.(?:xlsx|xls|XLSX|XLS|pdf|PDF|docx))\(\d+\)$", norm)
    if m:
        new_norm = m.group(1)
        warning = (
            f"file path normalized: {os.path.basename(path)!r} → "
            f"{os.path.basename(new_norm)!r} (macOS duplicate-download suffix stripped)"
        )
        norm = new_norm
    return norm, warning


def _check_supported_input_type(path: str) -> tuple[bool, str]:
    """Reject obviously-wrong file types with founder-friendly guidance.

    Returns (ok, message). Corpus test surfaced:
      * `.eml` — email files; user forwarded the email instead of detaching
    """
    lower = path.lower()
    if lower.endswith(".eml"):
        return (
            False,
            "This looks like an email file (.eml). Please attach the cap-table "
            "spreadsheet itself (the XLSX/XLS attachment from the email), not "
            "the email containing it.",
        )
    return True, ""


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--mode",
        required=True,
        choices=["validate", "carta", "pulley", "freeform", "auto"],
        help=(
            "validate: schema-check existing JSON in --dir; carta: extract from "
            "Carta XLSX (--xlsx); pulley: stub; freeform: validate Context-A "
            "sub-agent output (stdin); auto: sniff sheet fingerprint of --xlsx "
            "and dispatch to carta/pulley/freeform."
        ),
    )
    p.add_argument("--dir", help="Required for --mode=validate")
    p.add_argument("--xlsx", help="Path to XLSX file (for carta/pulley/auto)")
    p.add_argument("--instruments", help="(Carta mode) Where to write/append instruments.json")
    p.add_argument("-o", "--output", help="Where to write extraction_audit.json")
    p.add_argument("--pretty", action="store_true")
    args = p.parse_args()

    # Normalize the xlsx path if provided (handle macOS dupe suffixes)
    if args.xlsx:
        normalized, warning = _normalize_path(args.xlsx)
        if warning:
            sys.stderr.write(f"  note: {warning}\n")
        # Reject unsupported types
        ok, msg = _check_supported_input_type(normalized)
        if not ok:
            print(json.dumps({"ok": False, "blocker": "unsupported_input_type", "remedy": msg}))
            return 1
        args.xlsx = normalized

    if args.mode == "validate":
        if not args.dir:
            sys.stderr.write("--dir required for --mode=validate\n")
            return 1
        return _mode_validate(args)
    if args.mode == "auto":
        # Sniff the workbook and dispatch
        if not args.xlsx:
            sys.stderr.write("--xlsx required for --mode=auto\n")
            return 1
        try:
            wb = _open_xlsx(args.xlsx)
        except Exception as e:
            print(json.dumps({"ok": False, "blocker": "load_failed", "error": str(e)}))
            return 1
        detected = _carta_detect(wb.sheetnames)
        if detected in {"carta", "carta_ocx"}:
            args.mode = "carta"
            return _mode_carta(args)
        # Pulley check
        sheets = {s.strip() for s in wb.sheetnames}
        if PULLEY_FINGERPRINT.issubset(sheets) and PULLEY_CONTRACT_TABS & sheets:
            args.mode = "pulley"
            return _mode_pulley_stub(args)
        # Default: freeform — caller must dispatch Context-A
        print(
            json.dumps(
                {
                    "ok": False,
                    "mode": "auto",
                    "detected_format": "freeform",
                    "remedy": (
                        "Workbook does not match Carta or Pulley fingerprints. "
                        "Run --mode=freeform with the Context-A SPREADSHEET_STRUCTURE_DETECTION "
                        "sub-agent output piped on stdin."
                    ),
                    "sheet_names": wb.sheetnames,
                },
                indent=2,
            )
        )
        return 1
    if args.mode == "carta":
        return _mode_carta(args)
    if args.mode == "pulley":
        return _mode_pulley_stub(args)
    return _mode_freeform_validate(args)


if __name__ == "__main__":
    sys.exit(main())
