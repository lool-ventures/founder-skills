#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = []
# ///
"""Regression tests for financial-model-review scripts.

Run:  pytest founder-skills/tests/test_financial_model_review.py -v

All tests use subprocess to exercise the scripts exactly as the agent does.
"""

from __future__ import annotations

import copy
import json
import os
import re
import subprocess
import sys
import tempfile
from typing import Any

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FMR_SCRIPTS_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "skills", "financial-model-review", "scripts")
FIXTURES_DIR = os.path.join(SCRIPT_DIR, "fixtures")


def run_script(
    name: str,
    args: list[str] | None = None,
    stdin_data: str | None = None,
    script_dir: str | None = None,
) -> tuple[int, dict[str, Any], str]:
    """Run a script and return (exit_code, parsed_json, stderr)."""
    base = script_dir or FMR_SCRIPTS_DIR
    cmd = [sys.executable, os.path.join(base, name)]
    if args:
        cmd.extend(args)
    result = subprocess.run(cmd, input=stdin_data, capture_output=True, text=True)
    try:
        data = json.loads(result.stdout) if result.stdout.strip() else {}
    except json.JSONDecodeError:
        data = {}
    return result.returncode, data, result.stderr


def run_script_raw(
    name: str,
    args: list[str] | None = None,
    stdin_data: str | None = None,
    script_dir: str | None = None,
) -> tuple[int, str, str]:
    """Run a script and return (exit_code, raw_stdout, stderr)."""
    base = script_dir or FMR_SCRIPTS_DIR
    cmd = [sys.executable, os.path.join(base, name)]
    if args:
        cmd.extend(args)
    result = subprocess.run(cmd, input=stdin_data, capture_output=True, text=True)
    return result.returncode, result.stdout, result.stderr


# --- extract_model.py tests ---


def test_extract_model_csv() -> None:
    """CSV extraction produces structured JSON."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Month,Revenue,Expenses\n2025-01,50000,80000\n2025-02,55000,82000\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert "sheets" in data
    assert len(data["sheets"]) == 1  # CSV = single sheet


def test_extract_model_xlsx() -> None:
    """XLSX extraction produces structured JSON with multiple sheets."""
    import pytest

    fixture = os.path.join(FIXTURES_DIR, "sample_model.xlsx")
    if not os.path.exists(fixture):
        pytest.skip("sample_model.xlsx fixture not yet created")
    rc, data, stderr = run_script("extract_model.py", ["--file", fixture, "--pretty"])
    assert rc == 0
    assert data is not None
    assert "sheets" in data
    assert len(data["sheets"]) >= 2  # sample has multiple sheets


def test_extract_model_revenue_sheet_first_data_row_not_pre_header() -> None:
    """Revenue sheet first data row (2025-01, 100, 500, 50000) must land in rows[],
    not pre_header_rows.

    Regression: _find_header_row scored the data row higher than the text header
    (Month, Customers, ARPU, MRR) because '2025-01' matched the monthly pattern.
    Fix: rows predominantly numeric/date are disqualified from header candidacy.
    """
    import pytest

    fixture = os.path.join(FIXTURES_DIR, "sample_model.xlsx")
    if not os.path.exists(fixture):
        pytest.skip("sample_model.xlsx fixture not yet created")
    rc, data, stderr = run_script("extract_model.py", ["--file", fixture, "--pretty"])
    assert rc == 0
    assert data is not None
    rev = next((s for s in data["sheets"] if s["name"] == "Revenue"), None)
    assert rev is not None, "Revenue sheet not found in fixture"
    # Correct headers are the text labels
    assert rev["headers"][0] == "Month", (
        f"Expected first header 'Month', got {rev['headers'][0]!r} — data row classified as header"
    )
    # 2025-01 must appear in rows[], not in pre_header_rows
    row_labels = [str(r[0]) for r in rev["rows"] if r]
    assert "2025-01" in row_labels, "First data row (2025-01) must appear in rows[], not pre_header_rows"
    pre_header_labels = [str(r[0]) for r in rev["pre_header_rows"] if r]
    assert "2025-01" not in pre_header_labels, (
        "2025-01 must not appear in pre_header_rows — it is a data row, not banner text"
    )


def test_extract_model_banner_rows_still_captured_as_pre_header() -> None:
    """Text-dominant banner rows before a period header must still go to pre_header_rows.

    Guards against over-correction: a Carta/Pulley export with a company name
    banner (e.g. 'Acme Corp Financial Summary') before the period columns must
    still classify the banner as pre-header and the period row as header.
    """
    from openpyxl import Workbook  # type: ignore[import-untyped]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_path = f.name
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "P&L"
    # Row 1: banner text (text-dominant, not data)
    ws.append(["Acme Corp Financial Summary", None, None, None])
    # Row 2: period header (text cells that match monthly patterns)
    ws.append(["Line Item", "Jan 2025", "Feb 2025", "Mar 2025"])
    # Row 3+: data rows
    ws.append(["Revenue", 50000, 55000, 60000])
    wb.save(xlsx_path)
    wb.close()

    rc, data, stderr = run_script("extract_model.py", ["--file", xlsx_path, "--pretty"])
    os.unlink(xlsx_path)
    assert rc == 0
    sheet = data["sheets"][0]
    # Banner row must be in pre_header_rows
    assert sheet["pre_header_rows"], "Banner row should appear in pre_header_rows"
    assert sheet["pre_header_rows"][0][0] == "Acme Corp Financial Summary"
    # Period row must be selected as the header
    assert "Jan 2025" in sheet["headers"], f"Period row should be the header, got headers={sheet['headers']}"


def test_extract_model_stdin_passthrough() -> None:
    """Stdin JSON passes through as model_data."""
    input_data = json.dumps({"sheets": [{"name": "Manual", "headers": ["A"], "rows": [[1]]}]})
    rc, data, stderr = run_script("extract_model.py", ["--stdin"], stdin_data=input_data)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["name"] == "Manual"


def test_extract_model_pre_header_rows_csv() -> None:
    """CSV extraction includes pre_header_rows field (always empty for CSV)."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Month,Revenue\n2025-01,50000\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert "pre_header_rows" in data["sheets"][0]
    assert data["sheets"][0]["pre_header_rows"] == []


def test_extract_model_nonexistent_file() -> None:
    rc, data, stderr = run_script("extract_model.py", ["--file", "/tmp/nonexistent.xlsx"])
    assert rc == 1


def test_extract_csv_handles_utf8_bom(tmp_path: Any) -> None:
    """Windows-Excel CSV exports are BOM-prefixed; the first header must not
    become '\\ufeffMonth' (silent header-match failure)."""
    p = tmp_path / "model.csv"
    # Real UTF-8 BOM bytes followed by CSV content
    p.write_bytes(b"\xef\xbb\xbfMonth,Revenue\n2026-01,100000\n")
    rc, data, stderr = run_script("extract_model.py", ["--file", str(p), "--pretty"])
    assert rc == 0
    headers = data["sheets"][0]["headers"]
    assert headers[0] == "Month", f"Expected 'Month', got {headers[0]!r}"


def test_extract_model_output_flag() -> None:
    """The -o flag writes to file instead of stdout."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Month,Revenue\n2025-01,50000\n")
        f.flush()
        csv_path = f.name
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as out:
        out_path = out.name
    rc, data, stderr = run_script("extract_model.py", ["--file", csv_path, "-o", out_path])
    os.unlink(csv_path)
    assert rc == 0
    assert data is not None and data["ok"] is True
    with open(out_path) as fh:
        written = json.load(fh)
    os.unlink(out_path)
    assert "sheets" in written


# --- periodicity detection tests ---


def test_extract_periodicity_quarterly_csv() -> None:
    """CSV with Q1/Q2/Q3/Q4 headers detects quarterly periodicity."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Line Item,Q1 2024,Q2 2024,Q3 2024,Q4 2024\nRevenue,100,200,300,400\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "quarterly"
    assert data["periodicity_summary"] == "quarterly"


def test_extract_periodicity_monthly_csv() -> None:
    """CSV with month name headers detects monthly periodicity."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Line Item,Jan 2024,Feb 2024,Mar 2024\nRevenue,100,200,300\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "monthly"
    assert data["periodicity_summary"] == "monthly"


def test_extract_periodicity_iso_monthly_csv() -> None:
    """CSV with YYYY-MM headers detects monthly periodicity."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Line Item,2024-01,2024-02,2024-03\nRevenue,100,200,300\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "monthly"


def test_extract_periodicity_variant_1q24() -> None:
    """CSV with 1Q24-style headers detects quarterly."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Line Item,1Q24,2Q24,3Q24,4Q24\nRevenue,100,200,300,400\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "quarterly"


def test_extract_periodicity_month_range_quarterly() -> None:
    """CSV with Jan-Mar style headers detects quarterly, not monthly."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Line Item,Jan-Mar 2024,Apr-Jun 2024,Jul-Sep 2024,Oct-Dec 2024\nRevenue,100,200,300,400\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "quarterly"


def test_extract_periodicity_annual_csv() -> None:
    """CSV with FY headers detects annual periodicity."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Line Item,FY2024,FY2025,FY2026\nRevenue,1000,2000,3000\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "annual"


def test_extract_periodicity_unknown_csv() -> None:
    """CSV with non-time-series headers returns unknown."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Category,Amount,Notes\nSalaries,50000,Monthly\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data is not None
    assert data["sheets"][0]["periodicity"] == "unknown"
    assert data["periodicity_summary"] == "unknown"


def test_extract_periodicity_stdin_no_periodicity() -> None:
    """Stdin passthrough does not add periodicity (caller's responsibility)."""
    input_data = json.dumps({"sheets": [{"name": "Manual", "headers": ["A"], "rows": [[1]]}]})
    rc, data, stderr = run_script("extract_model.py", ["--stdin"], stdin_data=input_data)
    assert rc == 0
    assert data is not None
    # Stdin passes through as-is — no periodicity added
    assert "periodicity_summary" not in data


def test_extract_periodicity_mixed_xlsx() -> None:
    """XLSX with quarterly and monthly sheets returns mixed summary."""
    from openpyxl import Workbook  # type: ignore[import-untyped]

    wb = Workbook()
    # Sheet 1: quarterly headers
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "P&L"
    ws1.append(["Line Item", "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"])
    ws1.append(["Revenue", 100000, 120000, 140000, 160000])
    # Sheet 2: monthly headers
    ws2 = wb.create_sheet("Revenue")
    ws2.append(["Metric", "Jan 2024", "Feb 2024", "Mar 2024"])
    ws2.append(["MRR", 30000, 32000, 34000])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        tmp_path = f.name
    try:
        rc, data, stderr = run_script("extract_model.py", ["--file", tmp_path, "--pretty"])
        assert rc == 0
        assert data is not None
        periodicities = {s["name"]: s["periodicity"] for s in data["sheets"]}
        assert periodicities["P&L"] == "quarterly"
        assert periodicities["Revenue"] == "monthly"
        assert data["periodicity_summary"] == "mixed"
    finally:
        os.unlink(tmp_path)


# --- cell_refs tests ---


def test_extract_model_cell_refs_csv() -> None:
    """CSV extraction has empty cell_refs."""
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Month,Revenue\n2025-01,50000\n")
        f.flush()
        rc, data, stderr = run_script("extract_model.py", ["--file", f.name, "--pretty"])
    os.unlink(f.name)
    assert rc == 0
    assert data["sheets"][0]["cell_refs"] == []


def test_extract_model_cell_refs_xlsx() -> None:
    """XLSX extraction produces correct cell_refs with coordinates."""
    from openpyxl import Workbook  # type: ignore[import-untyped]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_path = f.name
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "P&L"
    # Row 1: headers
    ws.append(["Line Item", "Jan 2025", "Feb 2025"])
    # Row 2: Revenue
    ws.append(["Revenue", 50000, 55000])
    # Row 3: Expenses
    ws.append(["Expenses", 80000, 82000])
    wb.save(xlsx_path)
    wb.close()

    rc, data, stderr = run_script("extract_model.py", ["--file", xlsx_path, "--pretty"])
    os.unlink(xlsx_path)
    assert rc == 0
    sheet = data["sheets"][0]
    cell_refs = sheet["cell_refs"]
    assert isinstance(cell_refs, list)
    assert len(cell_refs) == 2  # Revenue + Expenses rows

    # Verify Revenue row refs
    rev_ref = next(r for r in cell_refs if r["label"] == "Revenue")
    assert rev_ref["row_index"] == 0  # first data row after header
    assert "Jan 2025" in rev_ref["cols"]
    assert rev_ref["cols"]["Jan 2025"] == "B2"  # Excel row 2, col B
    assert rev_ref["cols"]["Feb 2025"] == "C2"

    # Verify Expenses row refs
    exp_ref = next(r for r in cell_refs if r["label"] == "Expenses")
    assert exp_ref["row_index"] == 1
    assert exp_ref["cols"]["Jan 2025"] == "B3"


def _load_fmr_compose_module() -> Any:
    import importlib.util

    path = os.path.join(FMR_SCRIPTS_DIR, "compose_report.py")
    spec = importlib.util.spec_from_file_location("fmr_compose_report_module", path)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    sys.modules["fmr_compose_report_module"] = mod
    spec.loader.exec_module(mod)
    return mod


def test_coaching_payload_never_carries_pass_items() -> None:
    """Lock: pass-item evidence must never reach coaching_payload (it is built
    from failed/warned items + summary counts only), so capping pass-evidence
    length can never become a coaching regression. A pass item bearing a
    distinctive long evidence string must not surface in the payload."""
    mod = _load_fmr_compose_module()
    marker = "UNIQUE_PASS_EVIDENCE_TOKEN_zzz checked many things in great detail"
    checklist = {
        "summary": {
            "score_pct": 95,
            "overall_status": "pass",
            "total": 46,
            "pass": 46,
            "fail": 0,
            "warn": 0,
            "failed_items": [],
            "warned_items": [],
        },
        "items": [{"id": "STRUCT_01", "status": "pass", "evidence": marker, "notes": None}],
    }
    inputs = {"company": {"company_name": "TestCo"}}
    with tempfile.TemporaryDirectory(prefix="test-coaching-") as d:
        payload = mod._emit_coaching_payload(
            inputs, checklist, [], d, os.path.join(d, "report.md"), "<!--marker-->", None
        )
    assert marker not in json.dumps(payload), (
        "pass-item evidence leaked into coaching_payload — pass brevity would be a coaching regression"
    )


def _load_validate_extraction_module() -> Any:
    import importlib.util

    path = os.path.join(FMR_SCRIPTS_DIR, "validate_extraction.py")
    spec = importlib.util.spec_from_file_location("fmr_validate_extraction_module", path)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    sys.modules["fmr_validate_extraction_module"] = mod
    spec.loader.exec_module(mod)
    return mod


def test_cell_refs_duplicate_headers_resolve_to_own_coordinate() -> None:
    """Duplicate column headers must not collide in cell_refs.cols: a value
    present only in the FIRST duplicate column must resolve to that column's
    coordinate, not the last duplicate's. The last-duplicate-wins collision
    corrupts the best-match provenance that feeds the Step-3.6 founder review."""
    import pytest

    pytest.importorskip("openpyxl")
    from openpyxl import Workbook  # type: ignore[import-untyped]

    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "dup_headers.xlsx")
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "P&L"
        ws.append(["Line Item", "Q1", "Q2", "Q1", "Q2"])  # Q1/Q2 headers duplicated
        ws.append(["Revenue", 11, 22, 33, 44])  # non-empty label -> cell_refs populates
        wb.save(path)
        wb.close()

        rc, data, stderr = run_script("extract_model.py", ["--file", path, "--pretty"])
        assert rc == 0, stderr
        sheet = data["sheets"][0]
        ref = next(r for r in sheet["cell_refs"] if r["label"] == "Revenue")
        cols = ref["cols"]
        # Both Q1 columns (B2 and D2) must survive under distinct keys.
        coords = set(cols.values())
        assert "B2" in coords and "D2" in coords, (
            f"both Q1 columns' coordinates must survive (no last-wins collision): {cols}"
        )
        assert len([k for k in cols if k.startswith("Q1")]) == 2, (
            f"the two Q1 columns must occupy two distinct keys, not collide: {cols}"
        )

        # Consumer: a value present ONLY in the first Q1 column resolves to B2.
        ve = _load_validate_extraction_module()
        got = ve._find_cell_ref(11.0, data)
        assert got is not None and got["ref"] == "P&L!B2", (
            f"value 11 (first Q1, B2) must resolve to B2, not the last Q1 column: {got}"
        )


def test_extract_model_cell_refs_duplicate_labels() -> None:
    """XLSX with duplicate row labels produces separate cell_refs entries."""
    from openpyxl import Workbook  # type: ignore[import-untyped]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_path = f.name
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Line Item", "Jan 2025"])
    ws.append(["Payroll", 50000])  # row 2 in Excel
    ws.append(["Payroll", 80000])  # row 3 — same label, different row
    ws.append(["Payroll", 30000])  # row 4 — third duplicate
    wb.save(xlsx_path)
    wb.close()

    rc, data, stderr = run_script("extract_model.py", ["--file", xlsx_path, "--pretty"])
    os.unlink(xlsx_path)
    assert rc == 0
    cell_refs = data["sheets"][0]["cell_refs"]
    assert len(cell_refs) == 3  # all three rows present
    # Each has a unique row_index
    indices = [r["row_index"] for r in cell_refs]
    assert len(set(indices)) == 3  # no duplicates
    # All labeled "Payroll"
    assert all(r["label"] == "Payroll" for r in cell_refs)
    # Coordinates are distinct
    coords = [r["cols"]["Jan 2025"] for r in cell_refs]
    assert len(set(coords)) == 3  # B2, B3, B4


# --- Checklist IDs and helpers ---

_CHECKLIST_IDS: list[str] = [
    # Structure & Presentation
    "STRUCT_01",
    "STRUCT_02",
    "STRUCT_03",
    "STRUCT_04",
    "STRUCT_05",
    "STRUCT_06",
    "STRUCT_07",
    "STRUCT_08",
    "STRUCT_09",
    # Revenue & Unit Economics
    "UNIT_10",
    "UNIT_11",
    "UNIT_12",
    "UNIT_13",
    "UNIT_14",
    "UNIT_15",
    "UNIT_16",
    "UNIT_17",
    "UNIT_18",
    "UNIT_19",
    # Expenses, Cash & Runway
    "CASH_20",
    "CASH_21",
    "CASH_22",
    "CASH_23",
    "CASH_24",
    "CASH_25",
    "CASH_26",
    "CASH_27",
    "CASH_28",
    "CASH_29",
    "CASH_30",
    "CASH_31",
    "CASH_32",
    # Metrics & Efficiency
    "METRIC_33",
    "METRIC_34",
    "METRIC_35",
    # Fundraising Bridge
    "BRIDGE_36",
    "BRIDGE_37",
    "BRIDGE_38",
    # Sector-Specific
    "SECTOR_39",
    "SECTOR_40",
    "SECTOR_41",
    "SECTOR_42",
    "SECTOR_43",
    "SECTOR_44",
    # Overall
    "OVERALL_45",
    "OVERALL_46",
]


def _make_checklist_items(
    overrides: dict[str, dict[str, str]] | None = None,
    exclude: set[str] | None = None,
) -> list[dict[str, str]]:
    """Build a full 46-item checklist payload. Override specific items by ID."""
    overrides = overrides or {}
    exclude = exclude or set()
    items = []
    for item_id in _CHECKLIST_IDS:
        if item_id in exclude:
            continue
        base = {"id": item_id, "status": "pass", "evidence": f"Evidence for {item_id}"}
        if item_id in overrides:
            base.update(overrides[item_id])
        items.append(base)
    return items


# --- checklist.py tests ---


def test_checklist_all_pass() -> None:
    items = _make_checklist_items()
    payload = json.dumps({"items": items})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["summary"]["overall_status"] == "strong"
    assert data["summary"]["score_pct"] == 100.0
    assert data["summary"]["total"] == 46


def test_checklist_some_fail() -> None:
    items = _make_checklist_items(
        overrides={
            "STRUCT_01": {"status": "fail", "evidence": "Assumptions buried in formulas"},
            "UNIT_11": {"status": "fail", "evidence": "Zero churn assumed"},
            "CASH_23": {"status": "warn", "evidence": "Runway math unclear"},
        }
    )
    payload = json.dumps({"items": items})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["summary"]["fail"] == 2
    assert data["summary"]["warn"] == 1
    assert data["summary"]["score_pct"] < 100.0


def test_failed_items_have_severity() -> None:
    """Every entry in summary.failed_items must carry a severity in {high, medium, low}."""
    items = _make_checklist_items(
        overrides={
            "STRUCT_01": {"status": "fail", "evidence": "Assumptions buried"},
            "UNIT_11": {"status": "fail", "evidence": "Zero churn"},
            "CASH_23": {"status": "fail", "evidence": "Runway unclear"},
        }
    )
    payload = json.dumps({"items": items})
    rc, data, _ = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    failed = data["summary"]["failed_items"]
    assert len(failed) == 3
    for entry in failed:
        assert "severity" in entry, f"failed_item missing severity: {entry}"
        assert entry["severity"] in {"high", "medium", "low"}, f"unexpected severity {entry['severity']!r} in {entry}"


def test_warned_items_have_severity() -> None:
    """Every entry in summary.warned_items must carry a severity in {high, medium, low}."""
    items = _make_checklist_items(
        overrides={
            "STRUCT_01": {"status": "warn", "evidence": "Assumptions partially exposed"},
            "UNIT_11": {"status": "warn", "evidence": "Churn estimated"},
            "CASH_23": {"status": "warn", "evidence": "Runway approximated"},
        }
    )
    payload = json.dumps({"items": items})
    rc, data, _ = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    warned = data["summary"]["warned_items"]
    assert len(warned) == 3
    for entry in warned:
        assert "severity" in entry, f"warned_item missing severity: {entry}"
        assert entry["severity"] in {"high", "medium", "low"}, f"unexpected severity {entry['severity']!r} in {entry}"


def test_severity_distribution_matches_category_mapping() -> None:
    """Severity must match category mapping: STRUCT (low), UNIT (medium), CASH (high)."""
    items = _make_checklist_items(
        overrides={
            "STRUCT_01": {"status": "fail", "evidence": "Assumptions buried"},  # Structure & Presentation -> low
            "UNIT_11": {"status": "fail", "evidence": "Zero churn"},  # Revenue & Unit Economics -> medium
            "CASH_23": {"status": "fail", "evidence": "Runway unclear"},  # Expenses, Cash & Runway -> high
        }
    )
    payload = json.dumps({"items": items})
    rc, data, _ = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    by_id = {entry["id"]: entry for entry in data["summary"]["failed_items"]}
    assert by_id["STRUCT_01"]["severity"] == "low"
    assert by_id["UNIT_11"]["severity"] == "medium"
    assert by_id["CASH_23"]["severity"] == "high"


def test_existing_summary_fields_unchanged() -> None:
    """Adding severity must not remove or rename existing failed_items keys."""
    items = _make_checklist_items(overrides={"STRUCT_01": {"status": "fail", "evidence": "buried"}})
    payload = json.dumps({"items": items})
    rc, data, _ = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    entry = data["summary"]["failed_items"][0]
    # All pre-Phase-1-Task-2 keys still present.
    for key in ("id", "category", "label", "evidence", "notes"):
        assert key in entry, f"existing key {key!r} missing from failed_items entry"


def test_checklist_gating_unknown_sector_warns() -> None:
    """When sector_type is missing, a warning about sector_type is emitted on stderr."""
    company = {"stage": "seed", "geography": "us", "sector": "fintech", "traits": []}
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert "sector_type" in stderr.lower()


def test_checklist_not_applicable_pre_scored() -> None:
    """Backward compat: without company profile, agent-supplied not_applicable is trusted."""
    items = _make_checklist_items(
        overrides={
            "CASH_28": {"status": "not_applicable", "evidence": "Single-currency company"},
            "CASH_29": {"status": "not_applicable", "evidence": "Single entity"},
            "CASH_30": {"status": "not_applicable", "evidence": "Not Israel-based"},
            "CASH_31": {"status": "not_applicable", "evidence": "No IIA grants"},
            "CASH_32": {"status": "not_applicable", "evidence": "No VAT issues"},
            "SECTOR_39": {"status": "not_applicable", "evidence": "Not a marketplace"},
            "SECTOR_41": {"status": "not_applicable", "evidence": "Not hardware"},
            "SECTOR_42": {"status": "not_applicable", "evidence": "Not usage-based"},
            "SECTOR_43": {"status": "not_applicable", "evidence": "Not consumer"},
            "SECTOR_44": {"status": "not_applicable", "evidence": "No deferred revenue"},
        }
    )
    payload = json.dumps({"items": items})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["summary"]["not_applicable"] == 10
    assert data["summary"]["score_pct"] == 100.0


def _load_fmr_checklist_module() -> Any:
    import importlib.util

    path = os.path.join(FMR_SCRIPTS_DIR, "checklist.py")
    spec = importlib.util.spec_from_file_location("fmr_checklist_module", path)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    sys.modules["fmr_checklist_module"] = mod
    spec.loader.exec_module(mod)
    return mod


def test_company_is_taken_from_inputs_not_the_payload() -> None:
    """The payload is LLM-retyped; inputs.json is producer-written. The file wins."""
    mod = _load_fmr_checklist_module()
    company, warnings = mod._resolve_company(
        {"stage": "series-a", "model_format": "spreadsheet"},
        {"company": {"stage": "seed", "model_format": "deck"}},
    )
    assert company is not None
    assert company["stage"] == "seed"
    assert company["model_format"] == "deck"
    assert any("stage" in w for w in warnings)
    assert any("model_format" in w for w in warnings)


def test_company_falls_back_to_payload_when_inputs_absent() -> None:
    mod = _load_fmr_checklist_module()
    company, warnings = mod._resolve_company({"stage": "seed"}, None)
    assert company == {"stage": "seed"}
    assert warnings == []


def test_company_is_none_when_neither_source_has_it() -> None:
    """None must keep meaning 'cannot gate', not 'gate as spreadsheet'."""
    mod = _load_fmr_checklist_module()
    company, warnings = mod._resolve_company(None, {"metadata": {}})
    assert company is None
    assert warnings == []


def test_agreement_produces_no_warning() -> None:
    mod = _load_fmr_checklist_module()
    company, warnings = mod._resolve_company(
        {"stage": "seed", "model_format": "deck"},
        {"company": {"stage": "seed", "model_format": "deck"}},
    )
    assert company == {"stage": "seed", "model_format": "deck"}
    assert warnings == []


def test_inputs_file_profile_beats_the_payload_end_to_end(tmp_path: Any) -> None:
    """Wiring test: the resolved profile must reach gating, not just exist as a helper.

    The payload claims a spreadsheet model, the review inputs say deck. Deck must win, so the
    spreadsheet-only criteria come back not_applicable and the divergence is reported.
    """
    inputs_path = tmp_path / "inputs.json"
    inputs_path.write_text(
        json.dumps(
            {
                "company": {
                    "stage": "seed",
                    "model_format": "deck",
                    "geography": "israel",
                    "revenue_model_type": "saas-sales-led",
                }
            }
        ),
        encoding="utf-8",
    )
    payload = json.dumps(
        {
            "items": _make_checklist_items(),
            "company": {
                "stage": "seed",
                "model_format": "spreadsheet",
                "geography": "israel",
                "revenue_model_type": "saas-sales-led",
            },
        }
    )
    rc, data, stderr = run_script("checklist.py", ["--pretty", "--inputs", str(inputs_path)], stdin_data=payload)
    assert rc == 0
    by_id = {i["id"]: i for i in data["items"]}
    # STRUCT_01 is "Model format: spreadsheet only" — a deck must gate it away.
    assert by_id["STRUCT_01"]["status"] == "not_applicable"
    # UNIT_10 is "Model format: all" — it must survive.
    assert by_id["UNIT_10"]["status"] == "pass"
    assert "company.model_format differs" in stderr


def test_checklist_gating_normalizes_geography() -> None:
    """Free-form geography values are normalized; sector gates use sector_type."""
    company = {
        "stage": "seed",
        "geography": "United States",
        "sector": "B2B SaaS",
        "sector_type": "saas",
        "traits": [],
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    cash30 = next(i for i in data["items"] if i["id"] == "CASH_30")
    assert cash30["status"] == "not_applicable"


def test_checklist_gating_known_hub_geography_no_warning() -> None:
    """Major startup-hub geographies (e.g. India) normalize cleanly without the
    'not in normalization map' stderr warning, and gate the same as any
    non-Israel geography (Israel-specific items auto-gated as not_applicable)."""
    for raw_geo in ("India", "Germany", "France", "Canada", "Singapore", "Australia"):
        company: dict[str, Any] = {
            "stage": "seed",
            "geography": raw_geo,
            "sector": "B2B SaaS",
            "sector_type": "saas",
            "traits": [],
        }
        items = _make_checklist_items()
        payload = json.dumps({"items": items, "company": company})
        rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
        assert rc == 0
        assert "not in normalization map" not in stderr, f"unexpected warning for geography={raw_geo!r}: {stderr}"
        assert data is not None
        cash30 = next(i for i in data["items"] if i["id"] == "CASH_30")
        assert cash30["status"] == "not_applicable", f"{raw_geo} should auto-gate Israel-only CASH_30"


def test_checklist_gating_unrecognized_geography_still_warns_and_proceeds() -> None:
    """A genuinely unrecognized geography still falls through gracefully (warning
    on stderr, raw lowercased value used, no crash) rather than being silently
    swallowed."""
    company = {
        "stage": "seed",
        "geography": "Ruritania",
        "sector": "B2B SaaS",
        "sector_type": "saas",
        "traits": [],
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert "not in normalization map" in stderr
    assert data is not None
    cash30 = next(i for i in data["items"] if i["id"] == "CASH_30")
    assert cash30["status"] == "not_applicable"


def test_checklist_missing_sector_type_warns() -> None:
    """When sector_type is missing, a warning is emitted on stderr."""
    company = {"stage": "seed", "geography": "us", "sector": "saas", "traits": []}
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert "sector_type" in stderr.lower()


def test_checklist_gating_us_saas_company() -> None:
    """With company profile, script auto-gates items whose gates don't match."""
    company = {"stage": "seed", "geography": "us", "sector": "saas", "sector_type": "saas", "traits": []}
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    gated_ids = {
        "CASH_28",
        "CASH_29",
        "CASH_30",
        "CASH_31",
        "CASH_32",
        "SECTOR_39",
        "SECTOR_41",
        "SECTOR_42",
        "SECTOR_43",
        "SECTOR_44",
        "OVERALL_46",
    }
    for item in data["items"]:
        if item["id"] in gated_ids:
            assert item["status"] == "not_applicable", f"{item['id']} should be auto-gated but was {item['status']}"
            # Founder-facing wording: the evidence line reaches report.md and report.html, so it
            # states a reason rather than naming the gate field.
            assert item["evidence"].startswith("Not applicable — ")
    s40 = next(i for i in data["items"] if i["id"] == "SECTOR_40")
    assert s40["status"] == "not_applicable"
    assert data["summary"]["not_applicable"] >= 11


def test_checklist_gating_israel_ai_company() -> None:
    """Israel AI company: Israel items apply, AI items apply, marketplace/hardware don't."""
    company = {
        "stage": "seed",
        "geography": "israel",
        "sector": "ai-native",
        "sector_type": "ai-native",
        "traits": ["multi-currency"],
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    for iid in ("CASH_30", "CASH_31", "CASH_32"):
        item = next(i for i in data["items"] if i["id"] == iid)
        assert item["status"] != "not_applicable", f"{iid} should be applicable for Israel"
    cash28 = next(i for i in data["items"] if i["id"] == "CASH_28")
    assert cash28["status"] != "not_applicable"
    s40 = next(i for i in data["items"] if i["id"] == "SECTOR_40")
    assert s40["status"] != "not_applicable"
    for iid in ("SECTOR_39", "SECTOR_41", "SECTOR_43"):
        item = next(i for i in data["items"] if i["id"] == iid)
        assert item["status"] == "not_applicable", f"{iid} should be gated"


def test_checklist_ai_cost_gate_broadened() -> None:
    """SECTOR_40 should be applicable when expenses.cogs has inference_costs, even for non-AI sector."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
    }
    inputs_with_ai_costs = {
        "expenses": {
            "cogs": {"hosting": 5000, "inference_costs": 3000},
        },
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company, "inputs": inputs_with_ai_costs})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    s40 = next(it for it in data["items"] if it["id"] == "SECTOR_40")
    assert s40["status"] != "not_applicable", "SECTOR_40 should be applicable when AI costs present"


def test_checklist_ai_cost_gate_no_ai_costs() -> None:
    """SECTOR_40 should remain not_applicable for non-AI sector without AI cost keys."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
    }
    inputs_without_ai_costs = {
        "expenses": {
            "cogs": {"hosting": 5000, "support": 2000},
        },
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company, "inputs": inputs_without_ai_costs})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    s40 = next(it for it in data["items"] if it["id"] == "SECTOR_40")
    assert s40["status"] == "not_applicable", "SECTOR_40 should stay gated without AI costs"


def test_checklist_ai_powered_trait_triggers_sector_40() -> None:
    """ai-powered trait triggers SECTOR_40 for SaaS companies."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "Cybersecurity SaaS",
        "revenue_model_type": "saas-sales-led",
        "traits": ["ai-powered"],
        # no sector_type — derives "saas" from revenue_model_type
        # no AI cogs in inputs — trait alone should trigger SECTOR_40
    }
    items = _make_checklist_items(overrides={"SECTOR_40": {"status": "fail", "evidence": "No AI costs shown"}})
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    s40 = next(i for i in data["items"] if i["id"] == "SECTOR_40")
    assert s40["status"] == "fail", "SECTOR_40 should not be auto-gated when ai-powered trait present"


def _assert_validation_errors(data: dict | None, *fragments: str) -> None:
    """Assert data has validation.status == 'invalid' and errors contain all fragments."""
    assert data is not None, "expected JSON output with validation errors"
    assert data["validation"]["status"] == "invalid"
    joined = " ".join(data["validation"]["errors"]).lower()
    for frag in fragments:
        assert frag.lower() in joined, f"expected '{frag}' in validation errors: {data['validation']['errors']}"


def test_checklist_missing_items() -> None:
    items = _make_checklist_items(exclude={"STRUCT_01", "UNIT_10"})
    payload = json.dumps({"items": items})
    rc, data, _ = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 1  # rejected input now exits 1 (loud refusal); the diagnostic still lands on stdout
    _assert_validation_errors(data, "STRUCT_01")


def test_checklist_invalid_status() -> None:
    items = _make_checklist_items(overrides={"STRUCT_01": {"status": "maybe"}})
    payload = json.dumps({"items": items})
    rc, data, _ = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 1  # rejected input now exits 1 (loud refusal); the diagnostic still lands on stdout
    _assert_validation_errors(data, "invalid")


def test_checklist_by_category() -> None:
    items = _make_checklist_items()
    payload = json.dumps({"items": items})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "by_category" in data["summary"]
    cats = data["summary"]["by_category"]
    assert "Structure & Presentation" in cats
    assert "Revenue & Unit Economics" in cats
    assert cats["Structure & Presentation"]["pass"] == 9


def test_checklist_overall_status_thresholds() -> None:
    """Score >= 85 = strong, >= 70 = solid, >= 50 = needs_work, < 50 = major_revision."""
    fail_ids = {f"UNIT_{i}": {"status": "fail", "evidence": "test"} for i in range(10, 19)}
    items = _make_checklist_items(overrides=fail_ids)
    payload = json.dumps({"items": items})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["summary"]["overall_status"] == "solid"


def test_checklist_deck_format_gates_structural_items() -> None:
    """When model_format is 'deck', structural and expense items auto-gate to not_applicable."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
        "model_format": "deck",
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    # All 9 STRUCT items should be not_applicable
    for i in range(1, 10):
        item = next(it for it in data["items"] if it["id"] == f"STRUCT_0{i}")
        assert item["status"] == "not_applicable", f"STRUCT_0{i} should be gated for deck format"
    # CASH_20-27 (non-geo-gated expense items) should be not_applicable
    for i in range(20, 28):
        item = next(it for it in data["items"] if it["id"] == f"CASH_{i}")
        assert item["status"] == "not_applicable", f"CASH_{i} should be gated for deck format"
    # Revenue/Unit Economics items should still be applicable
    unit10 = next(it for it in data["items"] if it["id"] == "UNIT_10")
    assert unit10["status"] != "not_applicable"


def test_checklist_deck_format_sub_scores() -> None:
    """Deck format produces business_quality_pct and model_maturity_pct in summary."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
        "model_format": "deck",
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    summary = data["summary"]
    assert "business_quality_pct" in summary
    assert "model_maturity_pct" in summary
    # business_quality_pct should be 100% (all remaining items pass)
    assert summary["business_quality_pct"] == 100.0
    # model_maturity_pct should be None (all structural items are N/A)
    assert summary["model_maturity_pct"] is None


def test_checklist_spreadsheet_format_no_extra_gating() -> None:
    """When model_format is 'spreadsheet', no extra gating occurs."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
        "model_format": "spreadsheet",
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    struct01 = next(it for it in data["items"] if it["id"] == "STRUCT_01")
    assert struct01["status"] == "pass"


def test_checklist_no_model_format_backward_compat() -> None:
    """When model_format is absent, no extra gating occurs (backward compat)."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    struct01 = next(it for it in data["items"] if it["id"] == "STRUCT_01")
    assert struct01["status"] == "pass"
    # Sub-scores present with same value as score_pct
    summary = data["summary"]
    assert "business_quality_pct" in summary
    assert "model_maturity_pct" in summary


def test_checklist_partial_format_evaluates_all_46_items() -> None:
    """model_format='partial' must not auto-gate any item due to format gating.

    'partial' = incomplete spreadsheet — structure is still assessable.
    Documented contract: partial evaluates all 46 items (same as spreadsheet).
    Only deck/conversational gate STRUCT+CASH items.
    """
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
        "model_format": "partial",
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    # No item should be not_applicable due to model_format_gate='spreadsheet' gating.
    # (Profile-based geography/sector gating unrelated to format may still fire.)
    format_gated = [
        it
        for it in data["items"]
        if it["status"] == "not_applicable" and "model_format_gate" in str(it.get("evidence", "")).lower()
    ]
    assert format_gated == [], (
        f"partial format must not gate any items via model_format_gate, but got: {[it['id'] for it in format_gated]}"
    )
    # STRUCT items must be evaluable (pass, not not_applicable via format gating)
    for i in range(1, 10):
        item = next(it for it in data["items"] if it["id"] == f"STRUCT_0{i}")
        assert item["status"] != "not_applicable" or "model_format_gate" not in str(item.get("evidence", "")), (
            f"STRUCT_0{i} must not be format-gated for partial format"
        )


def test_checklist_deck_format_gates_exactly_22_spreadsheet_items() -> None:
    """deck format must gate all 22 items with model_format_gate='spreadsheet' (9 STRUCT + 13 CASH).

    Pin: 9 STRUCT_01..09 + CASH_20..32 (13 items) = 22 items total.
    Geography-gated CASH items (28, 29, 30, 31, 32) have both geo and format gates;
    with us/saas profile the geo gate fires first, but format gate also applies.
    """
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "saas",
        "sector_type": "saas",
        "traits": [],
        "model_format": "deck",
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    # All 9 STRUCT items must be not_applicable for deck
    for i in range(1, 10):
        item = next(it for it in data["items"] if it["id"] == f"STRUCT_0{i}")
        assert item["status"] == "not_applicable", f"STRUCT_0{i} should be gated for deck"
    # All 13 CASH items (20-32) must be not_applicable for deck (format or geo gate)
    for i in range(20, 33):
        item = next(it for it in data["items"] if it["id"] == f"CASH_{i}")
        assert item["status"] == "not_applicable", f"CASH_{i} should be gated for deck"


def test_checklist_dispatch_shape_propagates_run_id_and_gates() -> None:
    """The exact shape the SKILL.md CHECKLIST dispatch returns must yield a
    checklist.json with metadata.run_id (Context B parity) and engaged
    profile auto-gating (regression: items-only payloads produced neither)."""
    items = _make_checklist_items()
    payload = json.dumps(
        {
            "company": {
                "company_name": "TestCo",
                "slug": "testco",
                "stage": "pre-seed",
                "sector": "B2B SaaS",
                "geography": "US",
                "revenue_model_type": "saas-sales-led",
            },
            "metadata": {"run_id": "20260610T000000Z"},
            "items": items,
        }
    )
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["metadata"]["run_id"] == "20260610T000000Z"
    # pre-seed profile must auto-gate at least one seed+ item (e.g. UNIT_17)
    gated = [
        i
        for i in data["items"]
        if i["status"] == "not_applicable" and str(i.get("evidence", "")).startswith("Not applicable — ")
    ]
    assert gated, "company block present but no profile auto-gating engaged"


# --- Valid inputs fixture ---

_VALID_INPUTS: dict[str, Any] = {
    "company": {
        "company_name": "TestCo",
        "slug": "testco",
        "stage": "seed",
        "sector": "B2B SaaS",
        "geography": "US",
        "revenue_model_type": "saas-sales-led",
    },
    "revenue": {
        "arr": {"value": 600000, "as_of": "2025-12"},
        "mrr": {"value": 50000, "as_of": "2025-12"},
        "growth_rate_monthly": 0.08,
        "churn_monthly": 0.03,
        "nrr": 1.05,
        "grr": 0.95,
    },
    "expenses": {
        "headcount": [
            {"role": "Engineer", "count": 5, "salary_annual": 150000, "geography": "US", "burden_pct": 0.30},
            {"role": "Sales", "count": 2, "salary_annual": 120000, "geography": "US", "burden_pct": 0.25},
        ],
        "cogs": {"hosting": 5000, "support": 2000},
    },
    "cash": {
        "current_balance": 2000000,
        "debt": 0,
        "balance_date": "2025-12",
        "monthly_net_burn": 80000,
    },
    "unit_economics": {
        "cac": {
            "total": 1500,
            "components": {"ad_spend": 500, "sales_salaries": 800, "tools": 200},
            "fully_loaded": True,
        },
        "ltv": {
            "value": 6000,
            "method": "formula",
            "inputs": {"arpu_monthly": 500, "gross_margin": 0.75, "churn_monthly": 0.03},
            "observed_vs_assumed": "assumed",
        },
        "payback_months": 10,
        "gross_margin": 0.75,
    },
    "bridge": {
        "raise_amount": 5000000,
        "runway_target_months": 24,
    },
}


def _ue_metrics(currency: str | None) -> list[dict]:
    payload = copy.deepcopy(_VALID_INPUTS)
    if currency:
        payload["currency"] = currency
    rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(payload))
    assert rc == 0 and data is not None
    metrics: list[dict] = data["metrics"]
    return metrics


def _by_id(metrics: list[dict], mid: str) -> dict | None:
    return next((m for m in metrics if m["id"] == mid), None)


def test_non_usd_preserves_the_withheld_grade_as_a_reference() -> None:
    """With the USD ARR floors correctly suppressed, burn multiple and Rule of 40 both drop to
    `contextual` at once — leaving a non-USD founder numbers with no assessment. The grade the
    comparison already produced is preserved as an explicit REFERENCE so graded signal survives."""
    usd, ils = _ue_metrics(None), _ue_metrics("ILS")
    for mid in ("burn_multiple", "rule_of_40"):
        u, i = _by_id(usd, mid), _by_id(ils, mid)
        if u is None or i is None or u["rating"] not in ("strong", "acceptable", "warning", "fail"):
            continue  # not graded on this fixture — nothing to preserve
        assert i["rating"] == "contextual", f"{mid}: primary rating must stay the reliance boundary"
        assert i.get("benchmark_reference_rating") == u["rating"], (
            f"{mid}: reference grade should match what USD graded ({u['rating']}), "
            f"got {i.get('benchmark_reference_rating')!r}"
        )


def test_non_usd_reference_grade_carries_its_provenance() -> None:
    """A founder shown a reference grade is entitled to see what it was measured against."""
    ils = _ue_metrics("ILS")
    refs = [m for m in ils if m.get("benchmark_reference_rating")]
    assert refs, "expected at least one reference-graded metric for a non-USD model"
    for m in refs:
        assert m.get("benchmark_reference_source"), f"{m['id']}: reference grade without a source"
        assert m.get("benchmark_reference_as_of"), f"{m['id']}: reference grade without an as-of"
        note = m.get("benchmark_reference_note", "")
        assert "reference" in note.lower() and "not a verdict" in note.lower(), (
            f"{m['id']}: the reference must be labelled as such, got: {note!r}"
        )


def test_non_usd_reference_needs_no_fx_rate_and_invents_none() -> None:
    """The stage thresholds are DIMENSIONLESS (2.0x, 0.70, a sum of percentages), so the comparison
    is exact rather than converted. Nothing may claim a conversion or a rate."""
    ils = _ue_metrics("ILS")
    for m in [m for m in ils if m.get("benchmark_reference_rating")]:
        blob = json.dumps(m).lower()
        assert "fx rate" not in blob and "converted at" not in blob, (
            f"{m['id']}: reference grade must not imply a currency conversion: {blob[:200]}"
        )


def test_usd_models_gain_no_reference_fields() -> None:
    """Pure addition: a USD review must be byte-for-byte unaffected by this."""
    for m in _ue_metrics(None):
        assert "benchmark_reference_rating" not in m, f"{m['id']} leaked a reference field into a USD review"


def test_ltv_cac_is_not_suppressed_for_non_usd() -> None:
    """Scope guard: the caveat is applied to burn_multiple and rule_of_40 only. LTV/CAC is a
    dimensionless ratio with no ARR floor attached and keeps its grade in any currency."""
    ils = _ue_metrics("ILS")
    ltv_cac = _by_id(ils, "ltv_cac_ratio") or _by_id(ils, "ltv_cac")
    if ltv_cac is not None:
        assert "benchmark_reference_rating" not in ltv_cac, (
            "LTV/CAC should never have been routed through the non-USD caveat"
        )


def _runway_payload(cash: float, burn: float, mrr: float, growth: float) -> str:
    """Runway input with an explicit opex line so net burn is derivable."""
    return json.dumps(
        {
            "company": {"name": "TestCo", "stage": "seed"},
            "revenue": {"mrr": {"value": mrr}, "growth_rate_monthly": growth},
            "cash": {"current_balance": cash, "monthly_net_burn": burn, "balance_date": "2026-06"},
            "expenses": {"opex_monthly": [{"category": "All", "amount": mrr + burn, "start_month": "2026-01"}]},
        }
    )


def _base_scenario(data: dict) -> dict:
    return next((s for s in data["scenarios"] if s["name"].lower().startswith("base")), data["scenarios"][0])


def test_runway_reports_static_runway_alongside_the_projection() -> None:
    """The projection holds burn FLAT while revenue compounds, so it can turn a short cash
    position into 'default alive'. The static number is the floor under that and must be present."""
    rc, data, _ = run_script("runway.py", ["--pretty"], stdin_data=_runway_payload(6_200_000, 900_000, 420_000, 0.12))
    assert rc == 0
    base = _base_scenario(data)
    assert base["static_runway_months"] == 6.9, base["static_runway_months"]


def test_runway_verdict_names_the_flat_burn_assumption_when_static_runway_is_short() -> None:
    """A ~7-month cash position surfacing as 'low risk' is the founder-trust bug: the verdict
    has to carry the assumption that produced it."""
    rc, data, _ = run_script("runway.py", ["--pretty"], stdin_data=_runway_payload(6_200_000, 900_000, 420_000, 0.12))
    assert rc == 0
    verdict = data["risk_assessment"]
    assert "6.9 months" in verdict
    assert "burn staying flat" in verdict or "holds burn flat" in verdict


def test_runway_default_alive_without_profitability_is_not_called_low_risk() -> None:
    """default_alive is also set when cash merely never depletes in the window (incl. grant-driven),
    which is not the same as profitable — so it must not read as 'Low risk' unqualified."""
    rc, data, _ = run_script("runway.py", ["--pretty"], stdin_data=_runway_payload(6_200_000, 900_000, 420_000, 0.12))
    assert rc == 0
    base = _base_scenario(data)
    if base["default_alive"] and not base.get("became_profitable"):
        assert not data["risk_assessment"].startswith("Low risk")


def test_runway_static_months_absent_when_not_burning() -> None:
    """Revenue already covering opex means there is no static runway to report (no division)."""
    rc, data, _ = run_script(
        "runway.py",
        ["--pretty"],
        stdin_data=json.dumps(
            {
                "company": {"name": "TestCo", "stage": "seed"},
                "revenue": {"mrr": {"value": 500000}, "growth_rate_monthly": 0.05},
                "cash": {"current_balance": 3_000_000, "monthly_net_burn": -50_000, "balance_date": "2026-06"},
                "expenses": {"opex_monthly": [{"category": "All", "amount": 450000, "start_month": "2026-01"}]},
            }
        ),
    )
    assert rc == 0
    assert _base_scenario(data)["static_runway_months"] is None


def _implausible_bm_payload(currency: str | None) -> str:
    """burn 900K/mo against net-new ARR of 50K*0.01*12 = 6K -> 150x, past the >50 cutoff.

    ARR is set material (600K) so the USD $500K floor does NOT gate it — this isolates the
    implausibly-high branch itself, which is the one the non-USD caveat cannot reach.
    """
    payload = copy.deepcopy(_VALID_INPUTS)
    if currency:
        payload["currency"] = currency
    payload["revenue"] = {
        "mrr": {"value": 50000},
        "growth_rate_monthly": 0.01,
        "arr": {"value": 600000},
    }
    payload["cash"]["monthly_net_burn"] = 900000
    return json.dumps(payload)


def _burn_multiple(data: dict) -> dict | None:
    hits = [m for m in data["metrics"] if m["id"] == "burn_multiple"]
    return hits[0] if hits else None


def test_implausible_burn_multiple_usd_blames_inputs() -> None:
    """USD keeps today's wording: with the floor active, an implausible ratio really is suspect data."""
    rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=_implausible_bm_payload(None))
    assert rc == 0
    bm = _burn_multiple(data)
    assert bm is not None and bm["rating"] == "not_rated"
    assert "check input consistency" in bm["evidence"]


def test_implausible_burn_multiple_non_usd_blames_the_ungated_arr_base() -> None:
    """Non-USD skips the USD-denominated ARR materiality floor by design, so an implausible
    ratio there is most likely an immaterial base that was never gated — NOT a data error.
    Sending the founder to hunt a nonexistent input bug is the defect. The uniform non-USD
    caveat cannot fix this one: it early-returns on `not_rated`."""
    rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=_implausible_bm_payload("ILS"))
    assert rc == 0
    bm = _burn_multiple(data)
    assert bm is not None and bm["rating"] == "not_rated"
    assert "check input consistency" not in bm["evidence"]
    assert "materiality floor" in bm["evidence"]
    assert "ILS" in bm["evidence"]


def test_non_usd_arr_floor_asymmetry_is_deliberate_and_explained() -> None:
    """An immaterial ARR is withheld in USD but computed in non-USD (the floor is USD-denominated).
    That asymmetry is intended — but the non-USD side must SAY why rather than assert a benchmark."""
    payload = copy.deepcopy(_VALID_INPUTS)
    # Keep the fixture's revenue block (it carries the growth data the ratio needs) and only make the
    # ARR base immaterial — that is the single variable this asymmetry turns on.
    payload["revenue"]["arr"] = {"value": 100000}
    payload["cash"]["monthly_net_burn"] = 900000
    rc_usd, usd, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(payload))
    payload["currency"] = "ILS"
    rc_ils, ils, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(payload))
    assert rc_usd == 0 and rc_ils == 0
    bm_usd, bm_ils = _burn_multiple(usd), _burn_multiple(ils)
    assert bm_usd is not None and bm_usd["rating"] == "not_applicable"
    assert bm_ils is not None and bm_ils["rating"] != "not_applicable"
    assert "materiality floor" in bm_ils["evidence"] or "USD-denominated" in bm_ils["evidence"]


# --- unit_economics.py tests ---


def test_unit_economics_basic() -> None:
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "metrics" in data
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    assert "cac" in metrics_by_name
    assert "ltv" in metrics_by_name
    assert "gross_margin" in metrics_by_name
    assert "ltv_cac_ratio" in metrics_by_name


def test_unit_economics_flags_insufficient_data_below_two_metrics() -> None:
    """When <2 metrics are computable, unit_economics self-declares insufficient_data
    (mirroring runway.py) so the downstream gate can accept-with-warning."""
    minimal = {
        "company": {
            "company_name": "MinCo",
            "slug": "minco",
            "stage": "pre-seed",
            "sector": "B2B SaaS",
            "geography": "US",
            "revenue_model_type": "saas-plg",
        },
        "cash": {"current_balance": 500000},
    }
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(minimal))
    assert rc == 0, stderr
    assert data is not None
    assert data["summary"]["computed"] < 2
    assert data.get("insufficient_data") is True


def test_unit_economics_no_insufficient_flag_when_metrics_computed() -> None:
    """A rich model computing >=2 metrics carries no insufficient_data flag."""
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc == 0, stderr
    assert data is not None
    assert data["summary"]["computed"] >= 2
    assert "insufficient_data" not in data


def test_unit_economics_burn_multiple() -> None:
    inputs = {**_VALID_INPUTS}
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    if "burn_multiple" in metrics_by_name:
        assert metrics_by_name["burn_multiple"]["value"] is not None


def test_unit_economics_missing_optional_fields() -> None:
    """Should handle missing optional fields gracefully."""
    minimal = {
        "company": {
            "company_name": "MinCo",
            "slug": "minco",
            "stage": "pre-seed",
            "sector": "B2B SaaS",
            "geography": "US",
            "revenue_model_type": "saas-plg",
        },
        "revenue": {"mrr": {"value": 5000, "as_of": "2025-12"}},
    }
    payload = json.dumps(minimal)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    assert "gross_margin" not in metrics_by_name or metrics_by_name["gross_margin"].get("value") is None


def test_unit_economics_ratings() -> None:
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    valid_ratings = {"strong", "acceptable", "warning", "fail", "not_rated", "contextual", "not_applicable"}
    for metric in data["metrics"]:
        if metric.get("value") is not None:
            assert metric["rating"] in valid_ratings


def test_unit_economics_burn_multiple_computed_wins() -> None:
    """When compute inputs and provided are close (<2x ratio), computed burn_multiple is used.

    Derivation (period-matched formula):
      mrr=50000, g=0.08, burn=80000
      net_new_arr = 50000*0.08*12 = 48000
      computed = 80000/48000 = 1.67x
      provided = 1.7; ratio = 1.7/1.67 = 1.02 < 2.0 → computed (1.67) wins
    """
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["unit_economics"]["burn_multiple"] = 1.7
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    assert "burn_multiple" in metrics_by_name
    # Computed value (1.67) should be used, not the reported 1.7
    assert metrics_by_name["burn_multiple"]["value"] == 1.67
    assert metrics_by_name["burn_multiple"]["value"] != 1.7
    # burn_multiple_lifetime should NOT exist
    assert "burn_multiple_lifetime" not in metrics_by_name


def test_unit_economics_burn_multiple_fallback() -> None:
    """When compute inputs are missing, reported burn_multiple is used as fallback."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # Remove compute inputs: monthly_burn, mrr, growth_rate
    inputs["cash"].pop("monthly_net_burn", None)
    inputs["revenue"].pop("mrr", None)
    inputs["revenue"].pop("growth_rate_monthly", None)
    # Provide reported burn_multiple
    inputs["unit_economics"]["burn_multiple"] = 0.66
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    assert "burn_multiple" in metrics_by_name
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] == 0.66
    assert bm["rating"] == "not_rated"
    assert "reported" in bm["evidence"].lower()
    # burn_multiple_lifetime should NOT exist
    assert "burn_multiple_lifetime" not in metrics_by_name


def test_unit_economics_rule_of_40_below_1m_arr() -> None:
    """Rule of 40 should be not_applicable when ARR < $1M."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 130000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "not_applicable"
    assert "not meaningful" in r40["evidence"].lower() or "$1M" in r40["evidence"]


def test_unit_economics_rule_of_40_above_1m_arr() -> None:
    """Rule of 40 should use operating margin (burn-derived) when ARR >= $5M."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # ARR intentionally inflated above MRR*12 to test R40 $5M+ benchmark path
    inputs["revenue"]["arr"]["value"] = 6000000
    inputs["cash"]["monthly_net_burn"] = 30000  # op_margin = -30K/50K = -60%
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] != "not_applicable"
    assert r40["rating"] != "contextual"  # operating margin → benchmark-rated
    assert r40["value"] is not None
    # growth ≈ 151.8%, op_margin = -60%, R40 ≈ 91.8
    assert 85 < r40["value"] < 100
    assert "operating margin" in r40["evidence"].lower()


def test_unit_economics_rule_of_40_negative() -> None:
    """R40 can be negative when burn far exceeds revenue."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # ARR intentionally inflated above MRR*12 to test R40 $5M+ benchmark path
    inputs["revenue"]["arr"]["value"] = 6000000
    inputs["cash"]["monthly_net_burn"] = 80000  # op_margin = -80K/50K = -160%
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["value"] is not None
    assert r40["value"] < 0  # growth ≈ 151.8% + (-160%) ≈ -8.2
    assert "operating margin" in r40["evidence"].lower()


def test_unit_economics_rule_of_40_gross_margin_fallback() -> None:
    """R40 should fall back to gross margin (contextual) when burn data missing."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 1200000
    del inputs["cash"]["monthly_net_burn"]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["value"] is not None
    assert r40["rating"] == "contextual"
    assert "gross margin" in r40["evidence"].lower()
    assert "overstates" in r40["evidence"].lower()


def test_unit_economics_rule_of_40_operating_margin_preferred() -> None:
    """When both burn+MRR and gross margin are available, operating margin wins."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # ARR intentionally inflated above MRR*12 to test R40 $5M+ benchmark path
    inputs["revenue"]["arr"]["value"] = 6000000
    inputs["cash"]["monthly_net_burn"] = 30000
    inputs["unit_economics"]["gross_margin"] = 0.75  # should be ignored for R40
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert "operating margin" in r40["evidence"].lower()
    assert "gross margin" not in r40["evidence"].lower()


def test_unit_economics_rule_of_40_sign_error_fallback() -> None:
    """Negative monthly_net_burn (wrong sign) should trigger gross margin fallback."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 1200000
    inputs["cash"]["monthly_net_burn"] = -80000  # wrong sign → op_margin > 100%
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "contextual"
    assert "gross margin" in r40["evidence"].lower()
    assert "sign error" in stderr.lower() or "exceeds 100%" in stderr.lower()


def test_unit_economics_rule_of_40_sign_error_no_gm() -> None:
    """Sign error + no gross margin → not_rated (can't compute R40 at all)."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 1200000
    inputs["cash"]["monthly_net_burn"] = -80000
    del inputs["unit_economics"]["gross_margin"]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["value"] is None
    assert r40["rating"] == "not_rated"
    assert "implausible" in r40["evidence"].lower()
    assert "exceeds 100%" in stderr.lower()


def test_unit_economics_ltv_zero_churn_capped() -> None:
    """LTV with 0% churn should be capped at 60-month horizon with a label."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["churn_monthly"] = 0.0
    inputs["unit_economics"]["ltv"] = {
        "value": 38235,
        "method": "formula",
        "inputs": {"arpu_monthly": 500, "gross_margin": 0.75, "churn_monthly": 0.0},
        "observed_vs_assumed": "assumed",
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    ltv = metrics_by_name["ltv"]
    assert ltv["value"] is not None
    assert "capped" in ltv["evidence"].lower() or "5-year" in ltv["evidence"].lower()


def test_unit_economics_ltv_zero_churn_missing_arpu() -> None:
    """LTV with 0% churn but missing arpu should be not_rated with warning evidence."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["unit_economics"]["ltv"] = {
        "value": 1840000,
        "method": "formula",
        "inputs": {"churn_monthly": 0.0, "gross_margin": 0.75},
        "observed_vs_assumed": "assumed",
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    ltv = metrics_by_name["ltv"]
    assert ltv["rating"] == "not_rated"
    assert "could not apply 5-year cap" in ltv["evidence"].lower()
    assert "missing arpu" in ltv["evidence"].lower()
    # Structured warning
    warnings = data.get("warnings", [])
    codes = [w["code"] for w in warnings]
    assert "LTV_CAP_MISSING_INPUTS" in codes


def test_unit_economics_ltv_zero_churn_missing_gm() -> None:
    """LTV with 0% churn but missing gross_margin should be not_rated with warning evidence."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["unit_economics"]["ltv"] = {
        "value": 1840000,
        "method": "formula",
        "inputs": {"arpu_monthly": 500, "churn_monthly": 0.0},
        "observed_vs_assumed": "assumed",
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    ltv = metrics_by_name["ltv"]
    assert ltv["rating"] == "not_rated"
    assert "could not apply 5-year cap" in ltv["evidence"].lower()
    # Structured warning
    warnings = data.get("warnings", [])
    codes = [w["code"] for w in warnings]
    assert "LTV_CAP_MISSING_INPUTS" in codes


def test_unit_economics_ltv_zero_churn_with_inputs_no_warning() -> None:
    """LTV with 0% churn and both arpu+gm present should NOT emit LTV_CAP_MISSING_INPUTS."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["unit_economics"]["ltv"]["inputs"]["churn_monthly"] = 0.0
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    warnings = data.get("warnings", [])
    codes = [w["code"] for w in warnings]
    assert "LTV_CAP_MISSING_INPUTS" not in codes


def test_unit_economics_ltv_cac_contextual_when_assumed() -> None:
    """LTV/CAC from assumed inputs should be rated 'contextual', not hard pass/fail."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    if "ltv_cac_ratio" in metrics_by_name:
        assert metrics_by_name["ltv_cac_ratio"]["rating"] == "contextual"


def test_unit_economics_output_flag() -> None:
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as out:
        out_path = out.name
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("unit_economics.py", ["-o", out_path], stdin_data=payload)
    assert rc == 0
    assert data is not None and data["ok"] is True
    with open(out_path) as f:
        written = json.load(f)
    os.unlink(out_path)
    assert "metrics" in written


def test_unit_economics_confidence_qualifier() -> None:
    """data_confidence: 'estimated' appends qualifier to rated metric evidence."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["company"]["data_confidence"] = "estimated"
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    rated_metrics = [m for m in data["metrics"] if m["rating"] not in ("not_rated", "not_applicable")]
    assert len(rated_metrics) > 0, "Expected some rated metrics"
    for m in rated_metrics:
        assert "estimated" in m["evidence"].lower(), (
            f"Metric '{m['name']}' evidence should contain estimated qualifier: {m['evidence']}"
        )
        assert m.get("confidence") == "estimated", f"Metric '{m['name']}' should have confidence='estimated'"


def test_unit_economics_confidence_no_rating_change() -> None:
    """Ratings are identical regardless of data_confidence."""
    payload_exact = json.dumps(_VALID_INPUTS)
    rc1, data_exact, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=payload_exact)
    assert rc1 == 0 and data_exact is not None

    inputs_est = json.loads(json.dumps(_VALID_INPUTS))
    inputs_est["company"]["data_confidence"] = "estimated"
    payload_est = json.dumps(inputs_est)
    rc2, data_est, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=payload_est)
    assert rc2 == 0 and data_est is not None

    ratings_exact = {m["name"]: m["rating"] for m in data_exact["metrics"]}
    ratings_est = {m["name"]: m["rating"] for m in data_est["metrics"]}
    for name in ratings_exact:
        assert ratings_exact[name] == ratings_est.get(name, ratings_exact[name]), (
            f"Rating for '{name}' changed: exact={ratings_exact[name]} vs estimated={ratings_est.get(name)}"
        )


def test_unit_economics_confidence_exact_no_qualifier() -> None:
    """data_confidence: 'exact' (default) adds no qualifier or confidence field."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    for m in data["metrics"]:
        assert "estimated" not in m["evidence"].lower()
        assert "confidence" not in m


# --- runway.py tests ---


def test_runway_basic() -> None:
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "scenarios" in data
    assert len(data["scenarios"]) >= 3  # base, slow, crisis


def test_runway_auto_generates_scenarios() -> None:
    """When inputs don't include scenarios, script generates slow and crisis."""
    inputs = {k: v for k, v in _VALID_INPUTS.items() if k != "scenarios"}
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    scenario_names = {s["name"] for s in data["scenarios"]}
    assert "base" in scenario_names
    assert "slow" in scenario_names
    assert "crisis" in scenario_names


def test_runway_decision_points() -> None:
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    for scenario in data["scenarios"]:
        assert "runway_months" in scenario
        assert "cash_out_date" in scenario or scenario.get("runway_months") is None
        assert "decision_point" in scenario


def test_runway_default_alive() -> None:
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    for scenario in data["scenarios"]:
        assert "default_alive" in scenario
        assert isinstance(scenario["default_alive"], bool)


def test_runway_custom_scenarios() -> None:
    inputs = {
        **_VALID_INPUTS,
        "scenarios": {
            "base": {"growth_rate": 0.08, "burn_change": 0},
            "optimistic": {"growth_rate": 0.12, "burn_change": -0.05},
        },
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    scenario_names = {s["name"] for s in data["scenarios"]}
    assert "optimistic" in scenario_names


def test_runway_iia_grant_disbursement() -> None:
    """IIA grants add cash to projections during disbursement period."""
    inputs_with_grant = {
        **_VALID_INPUTS,
        "cash": {
            **_VALID_INPUTS["cash"],
            "grants": {
                "iia_approved": 120000,
                "iia_disbursement_months": 12,
                "iia_start_month": 1,
            },
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs_with_grant))
    assert rc == 0
    assert data is not None
    base_with = next(s for s in data["scenarios"] if s["name"] == "base")
    # Run without grants to compare
    rc2, data2, _ = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc2 == 0
    assert data2 is not None
    base_without = next(s for s in data2["scenarios"] if s["name"] == "base")
    # Grant should extend runway or improve cash at same month
    if base_with["runway_months"] is not None and base_without["runway_months"] is not None:
        assert base_with["runway_months"] >= base_without["runway_months"]
    # Limitations should mention IIA
    assert any("IIA" in lim for lim in data["limitations"])


def test_runway_fx_adjustment() -> None:
    """FX adjustment affects ILS-denominated expenses in scenarios."""
    inputs_with_fx = {
        **_VALID_INPUTS,
        "israel_specific": {
            "fx_rate_ils_usd": 3.65,
            "ils_expense_fraction": 0.6,
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs_with_fx))
    assert rc == 0
    assert data is not None
    # Auto-generated crisis scenario should have fx_adjustment > 0
    crisis = next(s for s in data["scenarios"] if s["name"] == "crisis")
    assert crisis["fx_adjustment"] == 0.10
    # Base should have 0
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    assert base["fx_adjustment"] == 0.0
    # Limitations should mention FX
    assert any("FX" in lim for lim in data["limitations"])


def test_runway_non_usd_currency_not_formatted_with_bare_dollar_sign() -> None:
    """When currency is a non-USD code, cash-balance messaging must be
    currency-tagged (e.g., "1,200,000 INR") rather than a bare "$" sign that
    implies USD."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["cash"].pop("monthly_net_burn", None)
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "$" not in data["risk_assessment"]
    assert "INR" in data["risk_assessment"]


def test_runway_absent_currency_still_uses_dollar_sign() -> None:
    """Back-compat: currency absent must keep today's bare-$ formatting."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["cash"].pop("monthly_net_burn", None)
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "$" in data["risk_assessment"]


def test_runway_output_echoes_currency() -> None:
    """The output JSON must carry the resolved currency code so downstream
    consumers (compose_report.py, visualize.py) can format currency-aware
    without re-reading inputs.json themselves."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0, stderr
    assert data is not None
    assert data.get("currency") == "INR"


def test_runway_output_defaults_currency_to_usd_when_absent() -> None:
    """Back-compat: absent currency echoes as 'USD' in the output."""
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc == 0, stderr
    assert data is not None
    assert data.get("currency") == "USD"


def test_runway_output_echoes_currency_on_insufficient_data_paths() -> None:
    """Currency must be echoed even on the early-return insufficient-data
    branches (both cash and burn missing; burn known but cash missing; cash
    known but burn missing) — not only the full-computation return path."""
    both_missing = json.loads(json.dumps(_VALID_INPUTS))
    both_missing["currency"] = "INR"
    both_missing["cash"] = {}
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(both_missing))
    assert rc == 0, stderr
    assert data is not None and data.get("currency") == "INR"

    burn_known_cash_missing = json.loads(json.dumps(_VALID_INPUTS))
    burn_known_cash_missing["currency"] = "INR"
    burn_known_cash_missing["cash"] = {"monthly_net_burn": 80000}
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(burn_known_cash_missing))
    assert rc == 0, stderr
    assert data is not None and data.get("currency") == "INR"

    cash_known_burn_missing = json.loads(json.dumps(_VALID_INPUTS))
    cash_known_burn_missing["currency"] = "INR"
    cash_known_burn_missing["cash"] = {"current_balance": 2000000}
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(cash_known_burn_missing))
    assert rc == 0, stderr
    assert data is not None and data.get("currency") == "INR"


def test_runway_burn_sensitivity_grid_skipped_for_non_usd() -> None:
    """The 500K-5M cash-level sensitivity grid is a USD-hypothetical grid
    (fixed round-number USD cash levels) — it must not be presented as if
    those numbers were native non-USD currency amounts. Skip it for non-USD
    rather than mislabeling raw USD figures as native currency."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["cash"] = {"monthly_net_burn": 80000}
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0, stderr
    assert data is not None
    assert not data.get("burn_sensitivity")
    assert any("INR" in w for w in data.get("warnings", []))


def test_runway_post_raise() -> None:
    """Post-raise computation shows extended runway."""
    inputs_with_raise = {
        **_VALID_INPUTS,
        "cash": {
            **_VALID_INPUTS["cash"],
            "fundraising": {"target_raise": 5000000, "expected_close": "2026-06"},
        },
        "bridge": {"raise_amount": 5000000, "runway_target_months": 24},
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs_with_raise))
    assert rc == 0
    assert data is not None
    assert "post_raise" in data
    assert data["post_raise"] is not None
    assert data["post_raise"]["raise_amount"] == 5000000
    assert data["post_raise"]["new_cash"] > _VALID_INPUTS["cash"]["current_balance"]
    # Post-raise runway should be longer than pre-raise
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    if base["runway_months"] is not None and data["post_raise"]["new_runway_months"] is not None:
        assert data["post_raise"]["new_runway_months"] > base["runway_months"]


def test_runway_no_post_raise_without_fundraising() -> None:
    """post_raise is None when no fundraising data is provided."""
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc == 0
    assert data is not None
    assert data["post_raise"] is None


def test_runway_threshold_scenario() -> None:
    """Runway output includes a 'threshold' scenario with minimum viable growth rate."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    scenario_names = {s["name"] for s in data["scenarios"]}
    assert "threshold" in scenario_names
    threshold = next(s for s in data["scenarios"] if s["name"] == "threshold")
    assert "growth_rate" in threshold
    assert threshold["growth_rate"] is not None
    assert threshold["growth_rate"] >= 0
    assert threshold["growth_rate"] <= _VALID_INPUTS["revenue"]["growth_rate_monthly"]


def test_runway_threshold_narrative() -> None:
    """Risk assessment includes minimum viable growth language."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    risk = data["risk_assessment"].lower()
    assert "at least" in risk or "minimum" in risk or "need" in risk


def test_runway_threshold_already_dead() -> None:
    """When even base scenario is not default-alive, threshold still present."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["cash"]["monthly_net_burn"] = 500000
    inputs["cash"]["current_balance"] = 1000000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    threshold = next((s for s in data["scenarios"] if s["name"] == "threshold"), None)
    assert threshold is not None


def test_runway_missing_cash_data() -> None:
    """Should handle missing cash fields gracefully."""
    minimal = {
        "company": {
            "company_name": "MinCo",
            "slug": "minco",
            "stage": "pre-seed",
            "sector": "B2B SaaS",
            "geography": "US",
            "revenue_model_type": "saas-plg",
        },
    }
    payload = json.dumps(minimal)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None


def test_runway_burn_change_one_time_step_up() -> None:
    """burn_change should be a one-time step-up, not monthly compounding."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # Use the slow scenario which has burn_change: 0.10
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    slow = next(s for s in data["scenarios"] if s["name"] == "slow")
    projections = slow["monthly_projections"]
    assert len(projections) >= 3
    # After the one-time step-up, expenses should be flat across all months
    month1_expenses = projections[0]["expenses"]
    month2_expenses = projections[1]["expenses"]
    month3_expenses = projections[2]["expenses"]
    # With one-time step-up: month1 == month2 == month3 (no compounding)
    # Allow tiny FP tolerance
    assert abs(month2_expenses - month1_expenses) < 0.01, (
        f"Expenses should be flat after step-up: month1={month1_expenses}, month2={month2_expenses}"
    )
    assert abs(month3_expenses - month1_expenses) < 0.01, (
        f"Expenses should be flat after step-up: month1={month1_expenses}, month3={month3_expenses}"
    )


def test_runway_growth_deceleration() -> None:
    """Effective growth rate should decay over time."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    projections = base["monthly_projections"]
    assert len(projections) >= 6
    # Compute implied growth rates from revenue: g_t = (R_t / R_{t-1}) - 1
    # Month 1 revenue grows from revenue0; subsequent months from prior month
    growth_rates = []
    for i in range(1, min(len(projections), 12)):
        prev_rev = projections[i - 1]["revenue"]
        curr_rev = projections[i]["revenue"]
        if prev_rev > 0:
            growth_rates.append(curr_rev / prev_rev - 1)
    # Growth rates must strictly decrease (not just non-increasing).
    # Math: with MRR=50000, growth=8%, decay=3%, the implied rate drops ~0.24pp per month.
    # Revenue is rounded to 2 decimals, shifting implied rates by at most ~0.001pp.
    # The 0.1% relative tolerance has ~15x headroom over rounding noise.
    assert len(growth_rates) >= 2, "Need at least 2 implied growth rates"
    for i in range(1, len(growth_rates)):
        assert growth_rates[i] < growth_rates[i - 1] * 0.999, (
            f"Growth rate must strictly decay: month {i + 2} rate {growth_rates[i]:.6f} "
            f"not less than month {i + 1} rate {growth_rates[i - 1]:.6f}"
        )


def test_runway_decayed_trajectory_leq_constant() -> None:
    """Decayed revenue trajectory should be <= constant-rate trajectory after month 1."""
    # We compare the actual (decayed) revenue to what constant-rate would produce
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    projections = base["monthly_projections"]
    growth_rate = _VALID_INPUTS["revenue"]["growth_rate_monthly"]
    mrr = _VALID_INPUTS["revenue"]["mrr"]["value"]
    # Compute constant-rate trajectory
    constant_rev = mrr
    for i, p in enumerate(projections):
        constant_rev = constant_rev * (1 + growth_rate)
        # After month 2 (index 1+), decayed must be strictly less than constant.
        # Skip index 1 (month 2) where rounding may compress the small delta;
        # by month 3+ the cumulative gap is well above rounding noise.
        if i > 1:
            assert p["revenue"] < constant_rev - 1.0, (
                f"Month {i + 1}: decayed revenue {p['revenue']:.2f} should be strictly "
                f"less than constant-rate {constant_rev:.2f}"
            )


def test_runway_threshold_solver_with_decay() -> None:
    """Threshold solver should find a viable rate; with decay it must be higher than base rate
    would be without decay for a cash-tight scenario."""
    # Use tight cash so threshold rate is meaningfully above 0
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["cash"]["current_balance"] = 500000  # tight cash
    inputs["cash"]["monthly_net_burn"] = 80000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    threshold = next((s for s in data["scenarios"] if s["name"] == "threshold"), None)
    assert threshold is not None
    assert threshold["growth_rate"] is not None
    # With decay, the solver needs a higher initial rate to compensate.
    # The threshold rate should be > 0 for this cash-tight scenario.
    assert threshold["growth_rate"] > 0.001, (
        f"Threshold rate {threshold['growth_rate']:.4f} should be meaningfully positive "
        f"for a cash-tight scenario with growth decay"
    )


def test_runway_passes_confidence_through() -> None:
    """data_confidence from company appears in runway output."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["company"]["data_confidence"] = "estimated"
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data.get("data_confidence") == "estimated"


def test_runway_no_confidence_when_exact() -> None:
    """data_confidence defaults to 'exact' and is omitted from output."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    # 'exact' is the default; field should not be in output
    assert "data_confidence" not in data


# --- compose_report.py helpers ---

_VALID_CHECKLIST: dict[str, Any] = {
    "items": [
        {
            "id": item_id,
            "category": "Test",
            "label": f"Label for {item_id}",
            "status": "pass",
            "evidence": f"Evidence for {item_id}",
            "notes": None,
        }
        for item_id in _CHECKLIST_IDS
    ],
    "summary": {
        "total": 46,
        "pass": 46,
        "fail": 0,
        "warn": 0,
        "not_applicable": 0,
        "score_pct": 100.0,
        "overall_status": "strong",
        "by_category": {},
        "failed_items": [],
        "warned_items": [],
    },
}

_VALID_UNIT_ECONOMICS: dict[str, Any] = {
    "metrics": [
        {
            "name": "cac",
            "value": 1500,
            "rating": "acceptable",
            "evidence": "Fully loaded CAC",
            "benchmark_source": "test",
            "benchmark_as_of": "2024",
        },
        {
            "name": "ltv",
            "value": 6000,
            "rating": "strong",
            "evidence": "Formula-based",
            "benchmark_source": "test",
            "benchmark_as_of": "2024",
        },
        {
            "name": "gross_margin",
            "value": 0.75,
            "rating": "strong",
            "evidence": "75% GM",
            "benchmark_source": "test",
            "benchmark_as_of": "2024",
        },
    ],
    "summary": {"computed": 3, "strong": 2, "acceptable": 1, "warning": 0, "fail": 0},
}

_VALID_RUNWAY: dict[str, Any] = {
    "company": {"name": "TestCo", "slug": "testco", "stage": "seed"},
    "baseline": {"net_cash": 2000000, "monthly_burn": 80000, "monthly_revenue": 50000},
    "scenarios": [
        {
            "name": "base",
            "runway_months": 25,
            "cash_out_date": "2028-01",
            "decision_point": "2027-01",
            "default_alive": True,
            "monthly_projections": [],
        },
        {
            "name": "slow",
            "runway_months": 18,
            "cash_out_date": "2027-06",
            "decision_point": "2026-06",
            "default_alive": False,
            "monthly_projections": [],
        },
        {
            "name": "crisis",
            "runway_months": 12,
            "cash_out_date": "2026-12",
            "decision_point": "2025-12",
            "default_alive": False,
            "monthly_projections": [],
        },
    ],
    "risk_assessment": "Adequate runway under base case.",
    "limitations": [],
    "warnings": [],
}


def _make_fmr_artifact_dir(artifacts: dict[str, Any]) -> str:
    d = tempfile.mkdtemp(prefix="test-compose-fmr-")
    for name, data in artifacts.items():
        path = os.path.join(d, name)
        with open(path, "w") as f:
            if isinstance(data, str):
                f.write(data)
            else:
                json.dump(data, f)
    return d


def _run_compose(artifact_dir: str, extra_args: list[str] | None = None) -> tuple[int, dict | None, str]:
    args = ["--dir", artifact_dir, "--pretty"]
    if extra_args:
        args.extend(extra_args)
    return run_script("compose_report.py", args)


# --- compose_report.py tests ---


def test_compose_complete_set() -> None:
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    assert "report_markdown" in data
    assert "validation" in data
    assert data["validation"]["status"] in ("clean", "warnings")


def test_compose_missing_required_artifact() -> None:
    """Missing required artifacts should exit 1."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 1
    assert "required artifacts missing" in stderr


def test_compose_missing_only_optional_artifact() -> None:
    """Missing model_data.json (optional) should succeed without high-severity warnings."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    # No high-severity warnings for missing optional
    missing_artifact_warnings = [w for w in data["validation"].get("warnings", []) if w["code"] == "MISSING_ARTIFACT"]
    assert not missing_artifact_warnings, "model_data.json is optional - should not trigger MISSING_ARTIFACT"


def test_compose_corrupt_artifact() -> None:
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": "not valid json{{{",
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["validation"]["warnings"]]
    assert "CORRUPT_ARTIFACT" in codes


def test_compose_strict_mode() -> None:
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": "corrupt",
        }
    )
    rc, data, stderr = _run_compose(d, extra_args=["--strict"])
    assert rc == 1


# --- Pipeline integration: feed realistic data through all scripts ---


def test_pipeline_extract_to_compose() -> None:
    """End-to-end: extract_model → checklist + unit_economics + runway → compose_report.

    This verifies schema compatibility across ALL five data-producing scripts.
    Each script's output must be consumable by downstream scripts without
    transformation.
    """
    # Step 0: Run extract_model on CSV fixture
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
        f.write("Month,Revenue,Expenses,Net\n2025-01,50000,80000,-30000\n2025-02,55000,82000,-27000\n")
        f.flush()
        csv_path = f.name
    rc_ex, extract_data, ex_stderr = run_script("extract_model.py", ["--file", csv_path, "--pretty"])
    os.unlink(csv_path)
    assert rc_ex == 0, f"extract_model.py failed: {ex_stderr}"
    assert extract_data is not None
    assert "sheets" in extract_data

    # Step 1: Build checklist items and run checklist.py
    checklist_input = {"items": _make_checklist_items()}
    rc_ck, checklist_data, _ = run_script("checklist.py", ["--pretty"], stdin_data=json.dumps(checklist_input))
    assert rc_ck == 0 and checklist_data is not None, "checklist.py failed"

    # Step 2: Run unit_economics on inputs
    rc_ue, ue_data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc_ue == 0 and ue_data is not None, "unit_economics.py failed"

    # Step 3: Run runway on inputs
    rc_rw, runway_data, _ = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc_rw == 0 and runway_data is not None, "runway.py failed"

    # Step 4: Feed all outputs to compose_report
    d = tempfile.mkdtemp(prefix="test-pipeline-")
    for name, data in [
        ("inputs.json", _VALID_INPUTS),
        ("model_data.json", extract_data),
        ("checklist.json", checklist_data),
        ("unit_economics.json", ue_data),
        ("runway.json", runway_data),
    ]:
        with open(os.path.join(d, name), "w") as f:
            json.dump(data, f)

    rc_cr, report, stderr = _run_compose(d)
    assert rc_cr == 0, f"compose_report failed on pipeline output: {stderr}"
    assert report is not None
    assert "report_markdown" in report
    # No high-severity warnings = schemas are compatible
    if "warnings" in report.get("validation", {}):
        for w in report["validation"]["warnings"]:
            assert w.get("severity") != "high", f"Pipeline produced high-severity warning: {w}"


# --- Agent structural smoke test ---


def test_compose_deck_format_severity_downgrade() -> None:
    """CHECKLIST_FAILURES severity should be 'medium' when model_format is deck."""
    inputs_deck = json.loads(json.dumps(_VALID_INPUTS))
    inputs_deck["company"]["model_format"] = "deck"
    checklist_failing = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist_failing["summary"]["overall_status"] = "major_revision"
    checklist_failing["summary"]["fail"] = 23
    checklist_failing["summary"]["failed_items"] = [{"id": f"STRUCT_0{i}"} for i in range(1, 10)]
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_deck,
            "checklist.json": checklist_failing,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    checklist_warnings = [w for w in data["validation"]["warnings"] if w["code"] == "CHECKLIST_FAILURES"]
    for w in checklist_warnings:
        assert w["severity"] == "medium", "CHECKLIST_FAILURES should be medium for deck format"


def test_compose_model_completeness_section() -> None:
    """Deck format report includes Model Completeness section."""
    inputs_deck = json.loads(json.dumps(_VALID_INPUTS))
    inputs_deck["company"]["model_format"] = "deck"
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_deck,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    assert "Model Completeness" in data["report_markdown"]


def test_compose_no_model_completeness_for_spreadsheet() -> None:
    """Spreadsheet format report should NOT include Model Completeness section."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    assert "Model Completeness" not in data["report_markdown"]


def test_compose_infinite_runway_rendering() -> None:
    """When runway_months is None (default_alive), renders 'Infinite' not 'None months'."""
    runway_infinite = json.loads(json.dumps(_VALID_RUNWAY))
    # Set base scenario to infinite runway (default alive)
    runway_infinite["scenarios"][0]["runway_months"] = None
    runway_infinite["scenarios"][0]["cash_out_date"] = None
    runway_infinite["scenarios"][0]["decision_point"] = None
    runway_infinite["scenarios"][0]["default_alive"] = True
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_infinite,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    # Should NOT contain "None months" anywhere in the report
    assert "None months" not in md, "Report should not render 'None months'"
    # Should contain the formatted infinite runway text
    assert "Infinite" in md or "profitability" in md.lower(), "Report should indicate infinite runway / profitability"


def test_compose_default_alive_coaching_payload_note() -> None:
    """coaching_payload carries base_runway_note when base scenario is default-alive with null runway."""
    runway_da = json.loads(json.dumps(_VALID_RUNWAY))
    # Base scenario: default_alive=True, runway_months=None
    runway_da["scenarios"][0]["runway_months"] = None
    runway_da["scenarios"][0]["cash_out_date"] = None
    runway_da["scenarios"][0]["decision_point"] = None
    runway_da["scenarios"][0]["default_alive"] = True
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_da,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    payload = data.get("coaching_payload", {})
    # runway_months is null for default-alive
    assert payload.get("runway_months") is None
    # base_runway_note must be present and explain the null
    assert "base_runway_note" in payload, "coaching_payload must carry base_runway_note for default-alive base scenario"
    note = payload["base_runway_note"]
    assert "default-alive" in note.lower() or "default_alive" in note.lower(), (
        f"base_runway_note should explain default-alive semantics; got: {note!r}"
    )
    assert "null" in note.lower() or "by design" in note.lower(), (
        f"base_runway_note should state runway_months is null by design; got: {note!r}"
    )


def test_compose_coaching_payload_carries_static_runway() -> None:
    """The Main-Thread Return treats runway_months as a HEADLINE field, but it is legitimately null
    for a default-alive company. static_runway_months travels with it so the headline always has a
    concrete number — otherwise the step has a field it cannot render."""
    runway_da = json.loads(json.dumps(_VALID_RUNWAY))
    runway_da["scenarios"][0]["runway_months"] = None
    runway_da["scenarios"][0]["default_alive"] = True
    runway_da["scenarios"][0]["static_runway_months"] = 6.9
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_da,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0 and data is not None
    payload = data.get("coaching_payload", {})
    assert payload.get("runway_months") is None
    assert payload.get("static_runway_months") == 6.9, (
        "coaching_payload must carry static_runway_months so a null runway_months still has a "
        f"reportable headline number; got: {payload.get('static_runway_months')!r}"
    )


def test_compose_default_alive_note_absent_when_not_default_alive() -> None:
    """coaching_payload does NOT carry base_runway_note when base scenario is not default-alive."""
    # _VALID_RUNWAY base scenario has default_alive=True but runway_months=25 (non-null)
    runway_normal = json.loads(json.dumps(_VALID_RUNWAY))
    runway_normal["scenarios"][0]["default_alive"] = False
    runway_normal["scenarios"][0]["runway_months"] = 20
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_normal,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    payload = data.get("coaching_payload", {})
    assert "base_runway_note" not in payload, (
        "base_runway_note should only appear for default-alive null-runway base scenarios"
    )


def test_compose_post_raise_in_report() -> None:
    """Post-raise data appears in runway section when present."""
    runway_with_post = json.loads(json.dumps(_VALID_RUNWAY))
    runway_with_post["post_raise"] = {
        "raise_amount": 5000000,
        "new_cash": 7000000,
        "new_runway_months": 48,
        "new_cash_out_date": "2029-12",
        "meets_target": True,
    }
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_with_post,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "Post-Raise" in md or "post_raise" in md.lower() or "$5" in md


# --- compose_report.py data confidence rendering tests ---


def test_compose_report_data_quality_line() -> None:
    """'Data Quality: Estimated' in executive summary when data_confidence != exact."""
    inputs_est = json.loads(json.dumps(_VALID_INPUTS))
    inputs_est["company"]["data_confidence"] = "estimated"
    inputs_est["company"]["model_format"] = "deck"
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_est,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "Data Quality" in md
    assert "Estimated" in md


def test_compose_report_estimated_label() -> None:
    """Score label is 'Deck Financial Readiness' when estimated + model_maturity_pct is null."""
    inputs_est = json.loads(json.dumps(_VALID_INPUTS))
    inputs_est["company"]["data_confidence"] = "estimated"
    inputs_est["company"]["model_format"] = "deck"
    checklist_deck = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist_deck["summary"]["model_maturity_pct"] = None
    checklist_deck["summary"]["business_quality_pct"] = 100.0
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_est,
            "checklist.json": checklist_deck,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "Deck Financial Readiness" in md or "business quality only" in md.lower()


def test_compose_report_deck_biz_quality_none_uses_model_quality_label() -> None:
    """When business_quality_pct is None (all business items N/A) and there is no
    spreadsheet model, the report must NOT print the 'business quality only' label
    with the overall score — that would be a mislabeled number.  Instead it must
    fall back to the generic 'Model Quality' label.

    Regression: the old code substituted `score` (overall) for `bq_score` but kept
    the 'business quality only' label, misleading readers about what the number meant.
    """
    inputs_est = json.loads(json.dumps(_VALID_INPUTS))
    inputs_est["company"]["data_confidence"] = "estimated"
    inputs_est["company"]["model_format"] = "deck"
    checklist_biz_na = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist_biz_na["summary"]["model_maturity_pct"] = None
    checklist_biz_na["summary"]["business_quality_pct"] = None  # all biz items N/A
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_est,
            "checklist.json": checklist_biz_na,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    # MUST NOT print the mislabeled line
    assert "business quality only" not in md.lower(), (
        "Report should not print 'business quality only' when business_quality_pct is None"
    )
    # MUST fall back to generic label
    assert "Model Quality" in md, "Report should use 'Model Quality' label when business_quality_pct is None"


def test_compose_report_exact_label() -> None:
    """Score label is 'Model Quality' when data_confidence is exact."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "Model Quality" in md


def test_compose_report_unit_economics_estimated_header() -> None:
    """Unit economics section notes when metrics are based on estimated inputs."""
    inputs_est = json.loads(json.dumps(_VALID_INPUTS))
    inputs_est["company"]["data_confidence"] = "estimated"
    ue_est = json.loads(json.dumps(_VALID_UNIT_ECONOMICS))
    for m in ue_est["metrics"]:
        m["confidence"] = "estimated"
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_est,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": ue_est,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "estimated" in md.lower()


def test_compose_report_stale_artifact_mismatched_run_ids() -> None:
    """Mismatched run_id across artifacts triggers STALE_ARTIFACT warning."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {"run_id": "run-001"}
    checklist = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist["metadata"] = {"run_id": "run-001"}
    ue = json.loads(json.dumps(_VALID_UNIT_ECONOMICS))
    ue["metadata"] = {"run_id": "run-002"}  # stale!
    runway = json.loads(json.dumps(_VALID_RUNWAY))
    runway["metadata"] = {"run_id": "run-001"}
    d = _make_fmr_artifact_dir(
        {"inputs.json": inputs, "checklist.json": checklist, "unit_economics.json": ue, "runway.json": runway}
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["validation"]["warnings"]]
    assert "STALE_ARTIFACT" in codes


def test_compose_report_matching_run_ids_no_stale_warning() -> None:
    """Matching run_id across all artifacts produces no STALE_ARTIFACT warning."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {"run_id": "run-001"}
    checklist = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist["metadata"] = {"run_id": "run-001"}
    ue = json.loads(json.dumps(_VALID_UNIT_ECONOMICS))
    ue["metadata"] = {"run_id": "run-001"}
    runway = json.loads(json.dumps(_VALID_RUNWAY))
    runway["metadata"] = {"run_id": "run-001"}
    d = _make_fmr_artifact_dir(
        {"inputs.json": inputs, "checklist.json": checklist, "unit_economics.json": ue, "runway.json": runway}
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["validation"]["warnings"]]
    assert "STALE_ARTIFACT" not in codes


def test_compose_report_no_run_ids_graceful() -> None:
    """No run_id in any artifact → graceful degradation, no STALE_ARTIFACT."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["validation"]["warnings"]]
    assert "STALE_ARTIFACT" not in codes


def test_compose_report_surfaces_warning_overrides() -> None:
    """Warning overrides from inputs.json metadata appear in the report."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {
        "warning_overrides": [
            {
                "code": "BURN_MULTIPLE_SUSPECT",
                "reason": "Enterprise SaaS with lumpy deal flow; TTM burn multiple is 5.7x",
                "reviewed_by": "agent",
                "timestamp": "2026-03-05T17:30:00Z",
            }
        ]
    }
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "Acknowledged Warnings" in md
    assert "BURN_MULTIPLE_SUSPECT" in md
    assert "Enterprise SaaS with lumpy deal flow" in md
    # Agent overrides appear in "Acknowledged Warnings" without reviewer suffix
    assert "Burn Multiple Suspect" in md


def test_compose_report_no_overrides_no_section() -> None:
    """No warning overrides → no Acknowledged Warnings section."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    assert "Acknowledged Warnings" not in data["report_markdown"]


# --- B1: sector_type derivation from revenue_model_type ---


def test_checklist_sector_type_derived_from_revenue_model_type() -> None:
    """sector_type auto-derived from revenue_model_type when not provided."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "AI infrastructure",
        "revenue_model_type": "ai-native",
        "traits": [],
        # no sector_type — should derive "ai-native" from revenue_model_type
    }
    items = _make_checklist_items(overrides={"SECTOR_40": {"status": "pass", "evidence": "Inference costs modeled"}})
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    s40 = next(i for i in data["items"] if i["id"] == "SECTOR_40")
    assert s40["status"] == "pass", "SECTOR_40 should not be auto-gated when ai-native derived"


def test_checklist_sector_type_saas_no_sector_items() -> None:
    """SaaS revenue_model_type derives sector_type='saas', no sector items triggered."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "B2B SaaS",
        "revenue_model_type": "saas-sales-led",
        "traits": [],
        # no sector_type — should derive "saas"
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    for item_id in ("SECTOR_39", "SECTOR_40", "SECTOR_41", "SECTOR_42", "SECTOR_43", "SECTOR_44"):
        item = next(i for i in data["items"] if i["id"] == item_id)
        assert item["status"] == "not_applicable", f"{item_id} should be gated for saas sector_type"


def test_checklist_annual_contracts_sector_gate() -> None:
    """annual-contracts revenue_model_type triggers SECTOR_44 (deferred revenue)."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "Enterprise SaaS",
        "revenue_model_type": "annual-contracts",
        "traits": [],
        # no sector_type — should derive "annual-contracts"
    }
    items = _make_checklist_items(overrides={"SECTOR_44": {"status": "pass", "evidence": "Deferred revenue tracked"}})
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    s44 = next(i for i in data["items"] if i["id"] == "SECTOR_44")
    assert s44["status"] == "pass", "SECTOR_44 should not be auto-gated for annual-contracts"


# --- B2: --strict behavior for deck format ---


def test_compose_strict_mode_deck_format_checklist_failures_not_blocking() -> None:
    """--strict should not exit 1 for deck format CHECKLIST_FAILURES alone."""
    inputs_deck = json.loads(json.dumps(_VALID_INPUTS))
    inputs_deck["company"]["model_format"] = "deck"
    checklist_failing = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist_failing["summary"]["overall_status"] = "major_revision"
    checklist_failing["summary"]["fail"] = 23
    checklist_failing["summary"]["failed_items"] = [{"id": f"STRUCT_0{i}"} for i in range(1, 10)]
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_deck,
            "checklist.json": checklist_failing,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d, extra_args=["--strict"])
    assert rc == 0, f"--strict should not block on CHECKLIST_FAILURES for deck format: {stderr}"
    assert data is not None
    checklist_warnings = [w for w in data["validation"]["warnings"] if w["code"] == "CHECKLIST_FAILURES"]
    assert len(checklist_warnings) > 0, "CHECKLIST_FAILURES warning should still be present"
    assert checklist_warnings[0]["severity"] == "medium", "Severity should remain medium"


def test_compose_strict_mode_medium_warnings_do_not_block() -> None:
    """--strict should NOT exit 1 for medium-severity warnings (findings, not data errors)."""
    # Create runway with inconsistent cash to trigger RUNWAY_INCONSISTENCY (medium)
    runway_inconsistent = json.loads(json.dumps(_VALID_RUNWAY))
    runway_inconsistent["baseline"]["net_cash"] = 500000  # differs >10% from inputs cash 2M
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_inconsistent,
        }
    )
    rc, data, stderr = _run_compose(d, extra_args=["--strict"])
    assert rc == 0, "--strict should not block on medium-severity warnings like RUNWAY_INCONSISTENCY"
    assert data is not None
    # But the warning should still be present in the output
    warnings = data["validation"]["warnings"]
    codes = [w["code"] for w in warnings]
    assert "RUNWAY_INCONSISTENCY" in codes


def test_compose_validation_includes_model_format() -> None:
    """Validation result includes model_format for --strict context."""
    inputs_deck = json.loads(json.dumps(_VALID_INPUTS))
    inputs_deck["company"]["model_format"] = "deck"
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_deck,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    assert data["validation"]["model_format"] == "deck"


def test_compose_validation_model_format_default_spreadsheet() -> None:
    """Validation result defaults model_format to spreadsheet."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    assert data["validation"]["model_format"] == "spreadsheet"


# --- v0.4.2 Phase 3: coaching_payload + uuid marker + severity truncation ---


def _make_fmr_v042_artifact_dir(checklist_overrides: dict | None = None) -> str:
    """Build a complete FMR artifact dir for v0.4.2 coaching_payload tests."""
    checklist = {**_VALID_CHECKLIST, **(checklist_overrides or {})}
    return _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": checklist,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )


def test_compose_severity_map_complete() -> None:
    """WARNING_SEVERITY contains all expected codes including MARKER_COLLISION."""
    import subprocess as _subprocess

    snippet = (
        f"import sys, os; sys.path.insert(0, {FMR_SCRIPTS_DIR!r}); "
        "from compose_report import WARNING_SEVERITY; "
        "import json; print(json.dumps(WARNING_SEVERITY))"
    )
    result = _subprocess.run([sys.executable, "-c", snippet], capture_output=True, text=True)
    try:
        sev_map = json.loads(result.stdout)
    except (json.JSONDecodeError, TypeError) as exc:
        raise AssertionError(f"can't import WARNING_SEVERITY: stdout={result.stdout}, stderr={result.stderr}") from exc

    expected_codes = [
        "CORRUPT_ARTIFACT",
        "MISSING_ARTIFACT",
        "STALE_ARTIFACT",
        "CHECKLIST_FAILURES",
        "MISSING_OPTIONAL_ARTIFACT",
        "CHECKLIST_INCOMPLETE",
        "RUNWAY_INCONSISTENCY",
        "METRICS_GAPS",
        "MARKER_COLLISION",
    ]
    for code in expected_codes:
        assert code in sev_map, f"WARNING_SEVERITY missing code: {code}"
    assert sev_map["MARKER_COLLISION"] == "low", "MARKER_COLLISION should be low severity"


def test_compose_emits_coaching_payload() -> None:
    """compose emits a coaching_payload block with all v0.4.2 fields."""
    import re

    d = _make_fmr_v042_artifact_dir()
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    assert "coaching_payload" in data, "report.json missing coaching_payload block"

    payload = data["coaching_payload"]
    assert payload["schema_version"] == "v0.4.2-financial-model-review"

    # All expected top-level keys present
    for key in (
        "schema_version",
        "summary",
        "failed_items",
        "warned_items",
        "high_severity_warnings",
        "company_name",
        "review_dir",
        "report_path",
        "insertion_marker",
        "truncated",
        "truncated_count",
    ):
        assert key in payload, f"coaching_payload missing key: {key}"

    # Summary mirrors checklist counts
    s = payload["summary"]
    for sk in ("score_pct", "overall_status", "total", "pass", "fail", "warn", "not_applicable"):
        assert sk in s, f"coaching_payload.summary missing {sk}"

    # Company name sourced from inputs.json → company.company_name
    assert payload["company_name"] == "TestCo"

    # Insertion marker matches uuid format
    assert re.fullmatch(r"<!-- COACHING_INSERTION_POINT_[0-9a-f]{8} -->", payload["insertion_marker"]), (
        f"unexpected marker shape: {payload['insertion_marker']}"
    )

    # Backward-compat: existing top-level keys still present
    assert "report_markdown" in data
    assert "validation" in data


def test_compose_inserts_uuid_marker() -> None:
    """report_markdown contains exactly one uuid marker matching coaching_payload.insertion_marker."""
    import re

    d = _make_fmr_v042_artifact_dir()
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None

    md = data["report_markdown"]
    matches = re.findall(r"<!-- COACHING_INSERTION_POINT_[0-9a-f]{8} -->", md)
    assert len(matches) == 1, f"expected exactly one marker, found {len(matches)}: {matches}"
    assert matches[0] == data["coaching_payload"]["insertion_marker"], (
        "marker in report_markdown must equal coaching_payload.insertion_marker"
    )


def test_compose_warns_on_marker_collision() -> None:
    """Body content containing the marker substring triggers MARKER_COLLISION (non-fatal)."""
    # Adversarial: a failed_item evidence string containing the literal marker substring.
    adversarial_items = [
        {
            "id": item_id,
            "category": "Test",
            "label": f"Label for {item_id}",
            "status": "pass",
            "evidence": f"Evidence for {item_id}",
            "notes": None,
        }
        for item_id in _CHECKLIST_IDS
    ]
    adversarial_items[0] = {
        "id": _CHECKLIST_IDS[0],
        "category": "Overall",
        "label": "Test fail",
        "status": "fail",
        "evidence": "Sneaky body content with <!-- COACHING_INSERTION_POINT_aaaaaaaa --> embedded",
        "notes": "Watch out",
    }
    checklist_overrides: dict[str, Any] = {
        "items": adversarial_items,
        "summary": {
            "total": 46,
            "pass": 45,
            "fail": 1,
            "warn": 0,
            "not_applicable": 0,
            "score_pct": 97.8,
            "overall_status": "strong",
            "by_category": {},
            "failed_items": [
                {
                    "id": _CHECKLIST_IDS[0],
                    "category": "Overall",
                    "label": "Test fail",
                    "evidence": "Sneaky body content with <!-- COACHING_INSERTION_POINT_aaaaaaaa --> embedded",
                    "notes": "Watch out",
                    "severity": "medium",
                }
            ],
            "warned_items": [],
        },
    }
    d = _make_fmr_v042_artifact_dir(checklist_overrides=checklist_overrides)
    rc, data, stderr = _run_compose(d)
    # Compose still succeeds (warning, not error)
    assert rc == 0, stderr
    assert data is not None
    codes = [w["code"] for w in data["validation"]["warnings"]]
    assert "MARKER_COLLISION" in codes, f"expected MARKER_COLLISION in warnings, got: {codes}"


def test_payload_arrays_match_summary_counts() -> None:
    """coaching_payload.failed_items length matches summary.fail; warned_items matches summary.warn (no truncation)."""
    items = [
        {
            "id": item_id,
            "category": "Test",
            "label": f"Label for {item_id}",
            "status": "pass",
            "evidence": "ok",
            "notes": None,
        }
        for item_id in _CHECKLIST_IDS
    ]
    items[0] = {**items[0], "status": "fail", "severity": "medium"}
    items[1] = {**items[1], "status": "fail", "severity": "medium"}
    items[2] = {**items[2], "status": "warn", "severity": "low"}

    failed_items = [
        {"id": _CHECKLIST_IDS[0], "category": "Test", "label": "L0", "evidence": "e", "severity": "medium"},
        {"id": _CHECKLIST_IDS[1], "category": "Test", "label": "L1", "evidence": "e", "severity": "medium"},
    ]
    warned_items = [
        {"id": _CHECKLIST_IDS[2], "category": "Test", "label": "L2", "evidence": "e", "severity": "low"},
    ]
    checklist_overrides: dict[str, Any] = {
        "items": items,
        "summary": {
            "total": 46,
            "pass": 43,
            "fail": 2,
            "warn": 1,
            "not_applicable": 0,
            "score_pct": 93.5,
            "overall_status": "strong",
            "by_category": {},
            "failed_items": failed_items,
            "warned_items": warned_items,
        },
    }
    d = _make_fmr_v042_artifact_dir(checklist_overrides=checklist_overrides)
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    payload = data["coaching_payload"]
    assert len(payload["failed_items"]) == payload["summary"]["fail"] == 2
    assert len(payload["warned_items"]) == payload["summary"]["warn"] == 1


def test_payload_truncation_under_30() -> None:
    """When total actionable items <= 30: truncated=false, truncated_count=0, all items present."""
    # 10 fails + 10 warns = 20 total — under threshold
    failed_items = [
        {
            "id": f"CASH_{20 + i}",
            "category": "Expenses, Cash & Runway",
            "label": f"F{i}",
            "evidence": "e",
            "severity": "high",
        }
        for i in range(10)
    ]
    warned_items = [
        {
            "id": f"UNIT_{10 + i}",
            "category": "Revenue & Unit Economics",
            "label": f"W{i}",
            "evidence": "e",
            "severity": "medium",
        }
        for i in range(10)
    ]
    checklist_overrides: dict[str, Any] = {
        "summary": {
            "total": 46,
            "pass": 26,
            "fail": 10,
            "warn": 10,
            "not_applicable": 0,
            "score_pct": 60.0,
            "overall_status": "major_revision",
            "by_category": {},
            "failed_items": failed_items,
            "warned_items": warned_items,
        },
    }
    d = _make_fmr_v042_artifact_dir(checklist_overrides=checklist_overrides)
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    payload = data["coaching_payload"]
    assert payload["truncated"] is False, "should not truncate when total <= 30"
    assert payload["truncated_count"] == 0
    assert len(payload["failed_items"]) == 10
    assert len(payload["warned_items"]) == 10


def test_payload_truncation_over_30() -> None:
    """When total actionable items > 30: truncated=true, severity ordering preserved, dropped_count correct."""
    # 20 high-severity fails + 15 medium warns = 35 total → 5 dropped
    # Expected: all 20 high fails + 10 medium warns kept (30 total)
    failed_items = [
        {
            "id": f"CASH_{20 + i}",
            "category": "Expenses, Cash & Runway",
            "label": f"HighFail{i}",
            "evidence": f"e{i}",
            "severity": "high",
        }
        for i in range(20)
    ]
    warned_items = [
        {
            "id": f"UNIT_{10 + i}",
            "category": "Revenue & Unit Economics",
            "label": f"MedWarn{i}",
            "evidence": f"w{i}",
            "severity": "medium",
        }
        for i in range(15)
    ]
    checklist_overrides: dict[str, Any] = {
        "summary": {
            "total": 46,
            "pass": 11,
            "fail": 20,
            "warn": 15,
            "not_applicable": 0,
            "score_pct": 30.0,
            "overall_status": "major_revision",
            "by_category": {},
            "failed_items": failed_items,
            "warned_items": warned_items,
        },
    }
    d = _make_fmr_v042_artifact_dir(checklist_overrides=checklist_overrides)
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    payload = data["coaching_payload"]

    assert payload["truncated"] is True
    assert payload["truncated_count"] == 5  # 35 - 30

    # Total kept is exactly 30
    assert len(payload["failed_items"]) + len(payload["warned_items"]) == 30

    # All 20 high-severity fails are kept (high > medium in priority)
    assert len(payload["failed_items"]) == 20

    # Only 10 of the 15 medium warns are kept
    assert len(payload["warned_items"]) == 10

    # Severity ordering: all failed_items should be high, all warned_items medium
    for item in payload["failed_items"]:
        assert item["severity"] == "high", f"Expected high severity in failed_items: {item}"
    for item in payload["warned_items"]:
        assert item["severity"] == "medium", f"Expected medium severity in warned_items: {item}"

    # Original order preserved within each severity tier (first 10 warns kept)
    kept_warn_ids = [item["id"] for item in payload["warned_items"]]
    assert kept_warn_ids == [f"UNIT_{10 + i}" for i in range(10)], (
        f"Original order not preserved in warned_items: {kept_warn_ids}"
    )


def test_coaching_payload_includes_runway_months() -> None:
    """Context B's success payload reports runway_months 'from coaching_payload'
    — so compose must actually emit it (base scenario value, null for default-alive)."""
    d = _make_fmr_v042_artifact_dir()
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    payload = data["coaching_payload"]
    assert "runway_months" in payload, "coaching_payload missing runway_months key"
    # _VALID_RUNWAY base scenario has runway_months=25
    assert payload["runway_months"] == 25, (
        f"Expected runway_months=25 from base scenario, got {payload['runway_months']!r}"
    )


def test_coaching_payload_runway_months_null_for_default_alive() -> None:
    """When all scenarios are default-alive (runway_months: null), coaching_payload
    must still emit runway_months: null (not omit the key)."""
    import copy

    default_alive_runway = copy.deepcopy(_VALID_RUNWAY)
    for s in default_alive_runway["scenarios"]:
        s["runway_months"] = None
        s["default_alive"] = True
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": default_alive_runway,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    payload = data["coaching_payload"]
    assert "runway_months" in payload, "coaching_payload must contain runway_months key even for default-alive"
    assert payload["runway_months"] is None, (
        f"Expected runway_months=None for default-alive company, got {payload['runway_months']!r}"
    )


# --- B3: burn multiple ARR floor ---


def test_unit_economics_burn_multiple_below_500k_arr() -> None:
    """Burn multiple not applicable below $500K ARR."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 130000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] == "not_applicable"
    assert "$500K" in bm["evidence"] or "not meaningful" in bm["evidence"].lower()


def test_unit_economics_burn_multiple_above_500k_arr() -> None:
    """Burn multiple computed normally above $500K ARR."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 600000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] != "not_applicable"
    assert bm["value"] is not None


def test_unit_economics_burn_multiple_fallback_below_500k_arr() -> None:
    """Reported burn multiple also gated by ARR floor."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 130000
    # Remove compute inputs so fallback path is used
    inputs["cash"].pop("monthly_net_burn", None)
    inputs["revenue"].pop("mrr", None)
    inputs["revenue"].pop("growth_rate_monthly", None)
    # Provide reported burn_multiple
    inputs["unit_economics"]["burn_multiple"] = 2.5
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] == "not_applicable", "ARR floor should gate even the fallback path"
    assert "$500K" in bm["evidence"] or "not meaningful" in bm["evidence"].lower()


# --- Currency determinism: burn_multiple / rule_of_40 are currency-agnostic
# --- RATIOS; only the USD-denominated materiality floor and stage benchmark
# --- table are USD-bound. A non-USD model must still get the ratio computed,
# --- downgraded to `contextual` (benchmark/floor not verifiable) — not
# --- silently withheld (not_rated, value None) and not silently passed through
# --- the ordinary benchmark rating either. This must hold whether or not
# --- revenue.arr.value happens to be present. ---


def test_unit_economics_burn_multiple_non_usd_currency_computes_ratio_as_contextual() -> None:
    """A non-USD-currency model's burn multiple ratio is currency-agnostic and must
    still be computed — only the USD stage benchmark and $500K materiality floor
    are not verifiable, so the rating downgrades to `contextual` with the ratio
    shown, rather than withholding the value entirely."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["revenue"]["arr"]["value"] = 123_000_000  # native INR, not USD
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] == "contextual"
    assert bm["value"] is not None  # the ratio itself is currency-agnostic
    assert "INR" in bm["evidence"]
    assert "$" not in bm["evidence"]


def test_unit_economics_burn_multiple_non_usd_currency_same_regardless_of_arr_presence() -> None:
    """The non-USD treatment must not depend on whether revenue.arr.value is
    present — deleting it must not silently let the metric through the ordinary
    benchmark-rated path (more data must not mean less scrutiny, and vice versa)."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["revenue"].pop("arr", None)
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] == "contextual"
    assert bm["value"] is not None


def test_unit_economics_burn_multiple_non_usd_low_raw_arr_not_gated_not_applicable() -> None:
    """The $500K floor is USD-denominated; a small raw non-USD ARR number must not
    trigger the not_applicable low-ARR gate — we cannot tell whether it clears
    materiality without a currency conversion, so the gate must not fire."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["revenue"]["arr"]["value"] = 100_000  # raw number is below the USD floor,
    # but this is INR, not a USD reading — the floor must not gate on it.
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] != "not_applicable"
    assert bm["value"] is not None


def test_unit_economics_rule_of_40_non_usd_currency_computes_ratio_as_contextual() -> None:
    """Same currency-agnostic-ratio treatment for Rule of 40's $1M ARR gate."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["revenue"]["arr"]["value"] = 82_000_000  # native INR, not USD
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "contextual"
    assert r40["value"] is not None
    assert "INR" in r40["evidence"]
    assert "$" not in r40["evidence"]


def test_unit_economics_rule_of_40_non_usd_currency_same_regardless_of_arr_presence() -> None:
    """ARR-absent path must receive the identical contextual treatment as ARR-present."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["revenue"].pop("arr", None)
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "contextual"
    assert r40["value"] is not None


def test_unit_economics_rule_of_40_non_usd_low_raw_arr_not_gated_by_usd_5m_floor() -> None:
    """The '$5M ARR' not-benchmark-compared floor on the operating-margin path is
    USD-denominated; a small raw non-USD ARR must not hit it (it would leak a bare
    $5M into evidence and compare native units against a USD threshold). Non-USD
    falls through to the benchmark branch and gets the uniform contextual caveat."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    inputs["revenue"]["arr"]["value"] = 4_000_000  # native INR, below the USD $5M floor
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "contextual"
    assert r40["value"] is not None
    assert "5M ARR" not in r40["evidence"], f"USD $5M floor leaked for non-USD: {r40['evidence']}"
    assert "$" not in r40["evidence"]


def test_unit_economics_currency_usd_explicit_matches_absent() -> None:
    """currency: 'USD' must behave identically to currency being absent."""
    rc1, data1, stderr1 = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    inputs_usd = json.loads(json.dumps(_VALID_INPUTS))
    inputs_usd["currency"] = "USD"
    rc2, data2, stderr2 = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inputs_usd))
    assert rc1 == 0, stderr1
    assert rc2 == 0, stderr2
    assert data1 is not None and data2 is not None
    assert data1["metrics"] == data2["metrics"]


def test_unit_economics_output_echoes_currency() -> None:
    """The output JSON must carry the resolved currency code so downstream
    consumers (compose_report.py, visualize.py) can format currency-aware
    without re-reading inputs.json themselves."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0, stderr
    assert data is not None
    assert data.get("currency") == "INR"


def test_unit_economics_output_defaults_currency_to_usd_when_absent() -> None:
    """Back-compat: absent currency echoes as 'USD' in the output."""
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc == 0, stderr
    assert data is not None
    assert data.get("currency") == "USD"


def test_unit_economics_cac_ltv_arr_per_fte_evidence_currency_tagged() -> None:
    """CAC, LTV, and ARR/FTE evidence strings must be currency-tagged for a
    non-USD model, not left as a bare '$' figure — these metrics aren't gated
    by the USD ARR floor at all, so they were the one bare-$ site the earlier
    currency fix missed entirely."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["currency"] = "INR"
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0, stderr
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}

    cac = metrics_by_name["cac"]
    assert "$" not in cac["evidence"]
    assert "INR" in cac["evidence"]

    ltv = metrics_by_name["ltv"]
    assert "$" not in ltv["evidence"]
    assert "INR" in ltv["evidence"]

    arr_fte = metrics_by_name["arr_per_fte"]
    assert "$" not in arr_fte["evidence"]
    assert "INR" in arr_fte["evidence"]


def test_unit_economics_cac_ltv_arr_per_fte_evidence_usd_unchanged() -> None:
    """Back-compat: absent/USD currency keeps the ordinary bare-$ evidence text."""
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(_VALID_INPUTS))
    assert rc == 0, stderr
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    assert "$" in metrics_by_name["cac"]["evidence"]
    assert "$" in metrics_by_name["ltv"]["evidence"]
    assert "$" in metrics_by_name["arr_per_fte"]["evidence"]


# --- New tests: FMR postmortem fixes ---


def test_unit_economics_annual_contracts_saas() -> None:
    """annual-contracts model type should be treated as SaaS."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["company"]["revenue_model_type"] = "annual-contracts"
    inputs["revenue"]["arr"]["value"] = 1200000  # above R40 $1M floor
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    # SaaS-only metrics should be computed, not not_applicable
    assert metrics_by_name["magic_number"]["rating"] != "not_applicable"
    assert metrics_by_name["rule_of_40"]["rating"] != "not_applicable"


def test_unit_economics_rule_of_40_contextual_band_1m_5m() -> None:
    """R40 should be contextual between $1M and $5M ARR with operating margin."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 3000000
    inputs["cash"]["monthly_net_burn"] = 30000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "contextual"
    assert "not benchmark-compared below $5M ARR" in r40["evidence"]


def test_unit_economics_rule_of_40_above_5m_benchmarked() -> None:
    """R40 above $5M ARR with operating margin should be benchmark-rated."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 6000000
    inputs["cash"]["monthly_net_burn"] = 30000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] not in ("contextual", "not_applicable", "not_rated")
    assert "benchmark" in r40["evidence"].lower() or "strong" in r40["evidence"].lower()


def test_unit_economics_rule_of_40_hyper_growth_above_5m() -> None:
    """R40 hyper-growth should still be contextual even above $5M ARR."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["arr"]["value"] = 6000000
    inputs["revenue"]["growth_rate_monthly"] = 0.15  # annualized ~435%
    inputs["cash"]["monthly_net_burn"] = 30000
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    r40 = metrics_by_name["rule_of_40"]
    assert r40["rating"] == "contextual"
    assert "hyper" in r40["evidence"].lower()


def test_unit_economics_burn_multiple_hyper_growth_contextual() -> None:
    """Burn multiple should be contextual when annualized growth > 200%."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["growth_rate_monthly"] = 0.15  # annualized ~435%
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["rating"] == "contextual"
    assert "hyper-growth" in bm["evidence"].lower()


def test_unit_economics_burn_multiple_seed_vs_series_a_thresholds() -> None:
    """Seed burn_multiple thresholds (2.0/2.5/3.0) differ from series-a (1.5/2.0/2.5)."""
    # Target BM = 2.0 using period-matched formula: monthly_burn / (mrr * g * 12)
    # mrr=50000, g=0.08 → net_new_arr = 50000*0.08*12 = 48000
    # burn = 2.0 * 48000 = 96000 → BM = 96000/48000 = 2.0
    # Seed (strong<=2.0): 2.0 ≤ 2.0 → "strong"; Series-A (strong<=1.5): 2.0 > 1.5 → "acceptable"
    for stage, expected_rating in [("seed", "strong"), ("series-a", "acceptable")]:
        inputs = json.loads(json.dumps(_VALID_INPUTS))
        inputs["company"]["stage"] = stage
        inputs["cash"]["monthly_net_burn"] = 96000
        inputs["revenue"]["growth_rate_monthly"] = 0.08
        payload = json.dumps(inputs)
        rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
        assert rc == 0
        assert data is not None
        metrics_by_name = {m["name"]: m for m in data["metrics"]}
        bm = metrics_by_name["burn_multiple"]
        assert bm["value"] == 2.0
        assert bm["rating"] == expected_rating, (
            f"Stage {stage}: burn_mult 2.0 expected {expected_rating}, got {bm['rating']}"
        )


def test_unit_economics_burn_multiple_ttm_monthly_arr() -> None:
    """12+ monthly entries with arr field → TTM path used."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # Build 12 monthly entries with ARR growing from 400K to 1M
    inputs["revenue"]["monthly"] = [
        {"month": f"2025-{m:02d}", "total": 33333 + i * 5000, "arr": 400000 + i * 50000}
        for i, m in enumerate(range(1, 13))
    ]
    # net_new_arr = 950000 - 400000 = 550000
    # burn = 80000/mo → annual burn = 960000
    # burn_mult = 960000 / 550000 = 1.75
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] == 1.75, f"Expected TTM burn multiple 1.75, got {bm['value']}"
    assert "TTM actual" in bm["evidence"]


def test_unit_economics_burn_multiple_ttm_monthly_total_only() -> None:
    """12+ monthly entries with only total (no arr) → total*12 approximation."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # Build 12 monthly entries with total growing from 30K to 80K (no arr field)
    inputs["revenue"]["monthly"] = [
        {"month": f"2025-{m:02d}", "total": 30000 + i * (50000 / 11)} for i, m in enumerate(range(1, 13))
    ]
    # latest total ≈ 80000 → arr approx = 960000
    # earliest total = 30000 → arr approx = 360000
    # net_new_arr = 960000 - 360000 = 600000
    # burn = 80000/mo → annual burn = 960000
    # burn_mult = 960000 / 600000 = 1.6
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] is not None
    assert "TTM actual" in bm["evidence"]


def test_unit_economics_burn_multiple_quarterly_yoy() -> None:
    """4+ quarterly entries → YoY quarterly path used."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # No monthly entries — force quarterly path
    inputs["revenue"].pop("monthly", None)
    inputs["revenue"]["quarterly"] = [
        {"quarter": "2025-Q1", "total": 100000, "arr": 400000},
        {"quarter": "2025-Q2", "total": 125000, "arr": 500000},
        {"quarter": "2025-Q3", "total": 150000, "arr": 600000},
        {"quarter": "2025-Q4", "total": 200000, "arr": 800000},
    ]
    # net_new_arr = 800000 - 400000 = 400000
    # burn = 80000/mo → annual burn = 960000
    # burn_mult = 960000 / 400000 = 2.4
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] == 2.4, f"Expected quarterly burn multiple 2.4, got {bm['value']}"
    assert "YoY (quarterly) actual" in bm["evidence"]


def test_unit_economics_burn_multiple_growth_rate_fallback() -> None:
    """<12 monthly + <4 quarterly → growth-rate fallback."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # Only 3 monthly entries — not enough for TTM
    inputs["revenue"]["monthly"] = [{"month": f"2025-{m:02d}", "total": 50000} for m in range(10, 13)]
    # Only 2 quarterly entries — not enough for YoY
    inputs["revenue"]["quarterly"] = [
        {"quarter": "2025-Q3", "total": 150000, "arr": 600000},
        {"quarter": "2025-Q4", "total": 200000, "arr": 800000},
    ]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] is not None
    # Growth-rate fallback doesn't say "TTM" or "YoY"
    assert "TTM" not in bm["evidence"]
    assert "YoY" not in bm["evidence"]


def test_unit_economics_burn_multiple_ttm_13_months_full_window() -> None:
    """13 monthly entries → true 12-month lookback (index -13)."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # 13 entries: month 0 (arr=400K) through month 12 (arr=1M)
    # With 13 entries, lookback is -13 → index 0 → arr=400K
    # net_new_arr = 1000000 - 400000 = 600000
    # burn = 80K/mo → annual = 960K → burn_mult = 960K / 600K = 1.6
    inputs["revenue"]["monthly"] = [
        {"month": f"2025-{m:02d}" if m <= 12 else f"2026-{m - 12:02d}", "arr": 400000 + i * 50000}
        for i, m in enumerate(range(0, 13))
    ]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] == 1.6, f"Expected 1.6, got {bm['value']}"
    assert "TTM actual" in bm["evidence"]


def test_unit_economics_burn_multiple_quarterly_5_entries_full_yoy() -> None:
    """5 quarterly entries → true 4-quarter YoY lookback (index -5)."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"].pop("monthly", None)
    # 5 entries: Q1 (arr=300K) through Q5 (arr=900K)
    # With 5 entries, lookback is -5 → index 0 → arr=300K
    # net_new_arr = 900000 - 300000 = 600000
    # burn = 80K/mo → annual = 960K → burn_mult = 960K / 600K = 1.6
    inputs["revenue"]["quarterly"] = [
        {"quarter": f"2024-Q{q}", "arr": 300000 + i * 150000} for i, q in enumerate([1, 2, 3, 4])
    ] + [{"quarter": "2025-Q1", "arr": 900000}]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] == 1.6, f"Expected 1.6, got {bm['value']}"
    assert "YoY (quarterly) actual" in bm["evidence"]


def test_unit_economics_burn_multiple_divergence_warning() -> None:
    """When TTM and growth-rate burn multiples diverge >2x, emit BURN_MULTIPLE_DIVERGENCE warning."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # TTM: arr grows from 200K to 800K → net_new_arr = 600K, burn_mult = (80K*12)/600K = 1.6x
    # Growth-rate (period-matched): MRR=50K, growth=0.02 → net_new_arr_monthly = 50K*0.02*12 = 12K
    #   _gr_burn_mult = 80K / 12K = 6.67x
    # Ratio = max(1.6, 6.67) / min(1.6, 6.67) = 4.17 > 2.0 → divergence warning fires
    inputs["revenue"]["growth_rate_monthly"] = 0.02  # low stated growth
    inputs["revenue"]["monthly"] = [
        {"month": f"2025-{m:02d}", "arr": 200000 + i * (600000 / 11)} for i, m in enumerate(range(1, 13))
    ]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    # Should use TTM path
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert "TTM actual" in bm["evidence"]
    # Should have divergence warning
    warnings = data.get("warnings", [])
    codes = [w["code"] for w in warnings]
    assert "BURN_MULTIPLE_DIVERGENCE" in codes, f"Expected divergence warning, got: {warnings}"


def test_unit_economics_burn_multiple_no_divergence_warning() -> None:
    """When TTM and growth-rate burn multiples are close, no BURN_MULTIPLE_DIVERGENCE."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    # ARR grows at exactly ΔARR_per_month = mrr*g*12 = 50K*0.08*12 = 48K per month
    # 12 entries (i=0..11) → ts_net_new_arr = arr[11]-arr[0] = 11 * 48K = 528K
    # TTM BM = (80K*12) / 528K = 1.82x
    # GR BM (period-matched) = 80K / (50K*0.08*12) = 80K/48K = 1.67x
    # Ratio = max(1.82, 1.67) / min(1.82, 1.67) = 1.09 → below 2.0 threshold, no warning
    inputs["revenue"]["monthly"] = [
        {"month": f"2025-{m:02d}", "total": 50000 + i * 4000, "arr": 600000 + i * 48000}
        for i, m in enumerate(range(1, 13))
    ]
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    warnings = data.get("warnings", [])
    codes = [w["code"] for w in warnings]
    assert "BURN_MULTIPLE_DIVERGENCE" not in codes


def test_unit_economics_burn_multiple_reference_regression() -> None:
    """Reference regression: mrr=58500, g=0.08, burn=42000 → burn_multiple=0.75, rating='strong'.

    Derivation (period-matched Sacks convention):
      ΔMRR = 58500 * 0.08 = 4680
      net_new_arr_per_month = ΔMRR * 12 = 4680 * 12 = 56160
      burn_multiple = monthly_burn / net_new_arr_per_month = 42000 / 56160 = 0.7479... → round(2) = 0.75
      Seed benchmark: strong ≤ 2.0 → 0.75 ≤ 2.0 → 'strong'
    """
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["mrr"] = {"value": 58500, "as_of": "2025-12"}
    inputs["revenue"]["growth_rate_monthly"] = 0.08
    inputs["cash"]["monthly_net_burn"] = 42000
    # Ensure growth-rate fallback is used (no monthly/quarterly time-series)
    inputs["revenue"].pop("monthly", None)
    inputs["revenue"].pop("quarterly", None)
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    bm = metrics_by_name["burn_multiple"]
    assert bm["value"] == 0.75, f"Expected 0.75, got {bm['value']}"
    assert bm["rating"] == "strong", f"Expected 'strong' (≤ seed threshold 2.0), got {bm['rating']!r}"


def test_unit_economics_ai_gross_margin_seed_vs_series_a() -> None:
    """AI gross margin adjustment: -5pt for seed, -10pt for series-a."""
    for stage, expected_adj in [("seed", 0.05), ("series-a", 0.10)]:
        inputs = json.loads(json.dumps(_VALID_INPUTS))
        inputs["company"]["stage"] = stage
        inputs["company"]["sector"] = "ai-native"
        inputs["unit_economics"]["gross_margin"] = 0.70
        payload = json.dumps(inputs)
        rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
        assert rc == 0
        assert data is not None
        metrics_by_name = {m["name"]: m for m in data["metrics"]}
        gm = metrics_by_name["gross_margin"]
        assert f"{expected_adj:.0%} discount" in gm["evidence"]


def test_unit_economics_benchmark_nested_object() -> None:
    """Benchmark-rated metrics should include a nested benchmark object."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    metrics_by_name = {m["name"]: m for m in data["metrics"]}
    # Gross margin should have benchmark
    gm = metrics_by_name["gross_margin"]
    assert "benchmark" in gm, f"gross_margin should have benchmark nested object: {gm}"
    assert gm["benchmark"]["target"] is not None
    assert gm["benchmark"]["source"] != ""


def test_runway_arr_12_fallback() -> None:
    """When MRR is missing but ARR present, runway should use ARR/12."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    del inputs["revenue"]["mrr"]
    inputs["revenue"]["arr"]["value"] = 600000  # ARR/12 = 50K
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["baseline"]["monthly_revenue"] == 50000.0
    assert "ARR/12" in stderr or "ARR/12" in str(data.get("warnings", []))


def test_sector_alias_fintech_to_saas() -> None:
    """fintech should resolve to saas sector type."""
    shared_scripts = os.path.join(os.path.dirname(SCRIPT_DIR), "scripts")
    sys.path.insert(0, shared_scripts)
    try:
        from founder_context import _derive_sector_type  # type: ignore[import-not-found]

        assert _derive_sector_type("fintech") == "saas"
        assert _derive_sector_type("B2B fintech") == "saas"
        assert _derive_sector_type("cybersecurity") == "saas"
        assert _derive_sector_type("edtech") == "saas"
        assert _derive_sector_type("transactional fintech") == "transactional-fintech"
        assert _derive_sector_type("payment processing") == "transactional-fintech"
        assert _derive_sector_type("payments infrastructure") == "transactional-fintech"
    finally:
        sys.path.remove(shared_scripts)


# --- validate_inputs.py sanity check tests ---


def _make_inputs(
    stage: str = "series-a",
    mrr: float = 100_000,
    burn: float = 200_000,
    growth: float | None = 0.10,
) -> dict[str, Any]:
    """Build a minimal inputs.json for validate_inputs tests."""
    inputs: dict[str, Any] = {
        "company": {
            "company_name": "TestCo",
            "slug": "testco",
            "stage": stage,
            "sector": "SaaS",
            "geography": "US",
            "revenue_model_type": "saas-plg",
        },
        "revenue": {
            "mrr": {"value": mrr, "as_of": "2025-01"},
        },
        "cash": {
            "current_balance": 2_000_000,
            "balance_date": "2025-01",
            "monthly_net_burn": burn,
        },
    }
    if growth is not None:
        inputs["revenue"]["growth_rate_monthly"] = growth
    return inputs


def test_validate_burn_revenue_suspect_series_a() -> None:
    """Series A burn > 5x MRR triggers BURN_REVENUE_SUSPECT with critical flag."""
    inputs = _make_inputs(stage="series-a", mrr=100_000, burn=600_000)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    assert data["valid"] is True
    assert data["has_critical_warnings"] is True
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_REVENUE_SUSPECT" in codes
    w = next(w for w in data["warnings"] if w["code"] == "BURN_REVENUE_SUSPECT")
    assert w["critical"] is True


def test_validate_burn_revenue_normal_no_warning() -> None:
    """Series A burn < 5x MRR does not trigger warning."""
    # burn=80K < 5*100K=500K → no BURN_REVENUE_SUSPECT
    # burn_multiple = (80K*12)/(100K*0.1*12) = 8x → no BURN_MULTIPLE_SUSPECT
    inputs = _make_inputs(stage="series-a", mrr=100_000, burn=80_000)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_REVENUE_SUSPECT" not in codes
    assert data["has_critical_warnings"] is False


def test_validate_burn_revenue_seed_higher_threshold() -> None:
    """Seed stage uses 10x threshold — 8x burn should not trigger."""
    # burn=40K < 10*50K=500K → no BURN_REVENUE_SUSPECT
    # burn_multiple = (40K*12)/(50K*0.1*12) = 8x → no BURN_MULTIPLE_SUSPECT
    inputs = _make_inputs(stage="seed", mrr=50_000, burn=40_000)  # 0.8x
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_REVENUE_SUSPECT" not in codes


class TestValidateInputsLateStageCoverage:
    """A late-stage founder must clear the enum and be gated at their real stage.

    The shared founder context accepts seven stages; this validator's enum and
    its seed+/series-a+ membership sets were written with five. The gap is not
    symmetrical: the enum rejects loudly, while the membership sets fail open to
    the pre-seed branch, so a late-stage company is silently held to no threshold
    at all.
    """

    LATE_STAGES = ("series-c", "series-d")

    def test_late_stage_clears_the_enum(self) -> None:
        for stage in self.LATE_STAGES:
            inputs = _make_inputs(stage=stage)
            rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
            assert rc == 0
            assert data is not None
            enum_errors = [
                e for e in data.get("errors", []) if e.get("code") == "ENUM_ERROR" and "stage" in json.dumps(e)
            ]
            assert not enum_errors, f"{stage} rejected by the stage enum: {enum_errors}"

    def test_late_stage_uses_the_series_a_burn_threshold(self) -> None:
        """Not the pre-seed 'skip' branch — the defect the enum error masks."""
        for stage in self.LATE_STAGES:
            # 6x MRR: above the series-a+ threshold of 5x, below the seed+ 10x.
            inputs = _make_inputs(stage=stage, mrr=100_000, burn=600_000)
            rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
            assert rc == 0
            assert data is not None
            codes = [w["code"] for w in data["warnings"]]
            assert "BURN_REVENUE_SUSPECT" in codes, (
                f"{stage} burn at 6x MRR did not trip the series-a+ threshold — it "
                f"fell through to the pre-seed skip branch. warnings={codes}"
            )

    def test_pre_seed_still_skips_the_burn_threshold(self) -> None:
        """Guard the by-design case: pre-seed is excluded on purpose, not by omission."""
        inputs = _make_inputs(stage="pre-seed", mrr=100_000, burn=600_000)
        rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
        assert rc == 0
        assert data is not None
        assert "BURN_REVENUE_SUSPECT" not in [w["code"] for w in data["warnings"]]

    def test_later_stage_still_uses_the_series_a_threshold(self) -> None:
        """Regression guard for a stage that already worked."""
        inputs = _make_inputs(stage="later", mrr=100_000, burn=600_000)
        rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
        assert rc == 0
        assert data is not None
        assert "BURN_REVENUE_SUSPECT" in [w["code"] for w in data["warnings"]]


def _load_fmr_unit_economics_module() -> Any:
    import importlib.util

    path = os.path.join(FMR_SCRIPTS_DIR, "unit_economics.py")
    spec = importlib.util.spec_from_file_location("fmr_unit_economics_module", path)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    sys.modules["fmr_unit_economics_module"] = mod
    spec.loader.exec_module(mod)
    return mod


class TestImplausibleValuesAreNotStrengths:
    """A higher-is-better scale must be able to say "that is not a real number".

    Rating an absurd value `strong` means the check stops checking exactly where
    a mis-scaled input is most likely — the same reasoning behind the existing
    burn-multiple and operating-margin guards.
    """

    @staticmethod
    def _ue() -> Any:
        return _load_fmr_unit_economics_module()

    def test_structurally_impossible_values_are_flagged(self) -> None:
        m = self._ue()
        for metric, value in (("gross_margin", 7.8), ("grr", 1.4)):
            assert m._implausibility_note(metric, value, pct=True), (
                f"{metric}={value} is impossible and must not read as a strength"
            )

    def test_units_error_magnitudes_are_flagged(self) -> None:
        m = self._ue()
        assert m._implausibility_note("nrr", 5.0, pct=True)
        assert m._implausibility_note("ltv_cac_ratio", 120.0, pct=False)

    def test_real_values_are_left_alone(self) -> None:
        """Anti-vacuity: an elite-but-real figure must still rate normally."""
        m = self._ue()
        for metric, value, pct in (
            ("gross_margin", 0.92, True),
            ("grr", 0.97, True),
            ("nrr", 1.6, True),
            ("ltv_cac_ratio", 8.0, False),
        ):
            assert m._implausibility_note(metric, value, pct=pct) is None, (
                f"{metric}={value} is high but achievable — flagging it would be a false positive"
            )

    def test_a_metric_without_a_ceiling_is_never_flagged(self) -> None:
        m = self._ue()
        assert m._implausibility_note("rule_of_40", 250.0, pct=False) is None


class TestAgentSuppliedDisclosureReachesTheReport:
    """A defaulted value must be distinguishable from a founder-stated one in the
    artifact people keep, not only in the chat turn where it was confirmed."""

    @staticmethod
    def _compose() -> Any:
        return _load_fmr_compose_module()

    def test_declared_defaults_render(self) -> None:
        mod = self._compose()
        out = mod._section_agent_supplied({"agent_supplied": ["cash.monthly_net_burn", "growth.growth_rate_monthly"]})
        assert "Agent-Supplied Values" in out
        assert "cash.monthly_net_burn" in out
        assert "growth.growth_rate_monthly" in out

    def test_nothing_renders_when_nothing_was_defaulted(self) -> None:
        """Anti-vacuity: an empty or absent declaration must produce no section."""
        mod = self._compose()
        assert mod._section_agent_supplied({"agent_supplied": []}) == ""
        assert mod._section_agent_supplied({}) == ""
        assert mod._section_agent_supplied(None) == ""


class TestStructuralErrorEvidenceExists:
    """The structural-error criterion is scored entirely on broken cells, so the
    tally has to be produced and routed to whoever scores it."""

    def test_error_cells_are_tallied(self) -> None:
        import importlib.util

        path = os.path.join(FMR_SCRIPTS_DIR, "extract_model.py")
        spec = importlib.util.spec_from_file_location("fmr_extract_model_mod", path)
        assert spec is not None and spec.loader is not None
        mod = importlib.util.module_from_spec(spec)
        sys.modules["fmr_extract_model_mod"] = mod
        spec.loader.exec_module(mod)

        sheets = [{"rows": [["Revenue", 100, "#REF!"], ["Costs", "#DIV/0!", "#REF!"]]}]
        assert mod._count_structural_errors(sheets) == {"#REF!": 2, "#DIV/0!": 1}
        assert mod._count_structural_errors([{"rows": [["ok", 1]]}]) == {}

    def test_the_assessor_is_told_where_the_evidence_is(self) -> None:
        base = os.path.dirname(FMR_SCRIPTS_DIR)
        with open(os.path.join(base, "SKILL.md"), encoding="utf-8") as f:
            skill = f.read()
        agent_path = os.path.join(os.path.dirname(SCRIPT_DIR), "agents", "financial-model-review.md")
        with open(agent_path, encoding="utf-8") as f:
            agent = f.read()
        for doc, name in ((skill, "SKILL.md"), (agent, "agent body")):
            assert "structural_errors" in doc, f"{name} never names the tally"
            assert "not_applicable" in doc, f"{name} does not say what to do when it is absent"


class TestCurrencyRuleIsNotContradicted:
    """The agent body is resident in context on every dispatch, so a boilerplate
    default there outranks a skill rule in practice."""

    def test_no_usd_default_contradicts_the_native_currency_rule(self) -> None:
        agent_path = os.path.join(os.path.dirname(SCRIPT_DIR), "agents", "financial-model-review.md")
        with open(agent_path, encoding="utf-8") as f:
            agent = f.read()
        assert "Currency is USD unless" not in agent
        assert "native currency" in agent.lower()


class TestValidateExtractionStageRangeSubstitution:
    """A stage with no plausibility ranges must borrow DOWNWARD, and say so.

    The bounds are floors — only the low end is compared — so standing in a
    lower stage's minimum makes the check more permissive, not less. Defaulting
    a Series C to seed's $50K floor means almost any cash figure reads as
    plausible and the check quietly stops checking.
    """

    @staticmethod
    def _ve() -> Any:
        import importlib.util

        path = os.path.join(FMR_SCRIPTS_DIR, "validate_extraction.py")
        spec = importlib.util.spec_from_file_location("fmr_validate_extraction_stage_mod", path)
        assert spec is not None and spec.loader is not None
        mod = importlib.util.module_from_spec(spec)
        sys.modules["fmr_validate_extraction_stage_mod"] = mod
        spec.loader.exec_module(mod)
        return mod

    def test_stages_above_the_table_borrow_from_the_nearest_lower_stage(self) -> None:
        mod = self._ve()
        for stage in ("series-c", "series-d", "later"):
            ranges, basis = mod._resolve_stage_ranges(stage)
            assert basis is not None, f"{stage} substituted silently"
            assert basis["resolved_to"] == "series-b", (
                f"{stage} borrowed from {basis['resolved_to']}, not the nearest lower stage"
            )
            assert ranges["cash_balance"][0] == mod._STAGE_RANGES["series-b"]["cash_balance"][0]

    def test_substitution_is_stricter_than_the_old_seed_default(self) -> None:
        """The floor must rise, not fall — this is what makes it not a false negative."""
        mod = self._ve()
        seed_floor = mod._STAGE_RANGES["seed"]["cash_balance"][0]
        ranges, _ = mod._resolve_stage_ranges("series-c")
        assert ranges["cash_balance"][0] > seed_floor

    def test_a_stage_with_its_own_ranges_substitutes_nothing(self) -> None:
        """Anti-vacuity: no disclosure when nothing was borrowed."""
        mod = self._ve()
        for stage in ("pre-seed", "seed", "series-a", "series-b"):
            ranges, basis = mod._resolve_stage_ranges(stage)
            assert basis is None, f"{stage} has its own ranges; reporting a substitution is wrong"
            assert ranges is mod._STAGE_RANGES[stage]

    def test_an_unrecognized_token_keeps_the_seed_default(self) -> None:
        """Off-ladder tokens carry no ordering, so there is nothing to descend from."""
        mod = self._ve()
        ranges, basis = mod._resolve_stage_ranges("not-a-stage")
        assert ranges is mod._STAGE_RANGES["seed"]
        assert basis is not None and basis["resolved_to"] == "seed"

    def test_the_check_result_carries_the_substitution(self) -> None:
        mod = self._ve()
        inputs = {
            "company": {"company_name": "TestCo", "stage": "series-c"},
            "cash": {"current_balance": 40_000_000, "monthly_net_burn": 900_000},
            "expenses": {"headcount": []},
        }
        result = mod._check_scale_plausibility(inputs, {"sheets": []})
        assert result.get("stage_basis") is not None, "substitution not surfaced on the check"
        assert result["stage_basis"]["requested"] == "series-c"
        assert "series-b" in result["message"]

    def test_a_non_usd_model_reports_no_substitution(self) -> None:
        """The ranges are never consulted for a non-USD model, so nothing was borrowed."""
        mod = self._ve()
        inputs = {
            "company": {"company_name": "TestCo", "stage": "series-c"},
            "currency": "EUR",
            "cash": {"current_balance": 40_000_000, "monthly_net_burn": 900_000},
            "expenses": {"headcount": []},
        }
        result = mod._check_scale_plausibility(inputs, {"sheets": []})
        assert "stage_basis" not in result

    def test_an_unrecognized_off_ladder_token_lands_on_seed_not_the_ladder(self) -> None:
        """Confirms the two substitution paths land differently: an on-ladder
        stage with no published ranges (series-c/-d/later) descends to the
        nearest lower published stage (series-b); a token that isn't on the
        ladder at all (typo, unknown value) has no ordering to descend from
        and lands on seed instead. A prior audit got this backwards in both
        directions, so this is pinned directly against the code."""
        mod = self._ve()
        for on_ladder_unpublished in ("series-c", "series-d", "later"):
            _ranges, basis = mod._resolve_stage_ranges(on_ladder_unpublished)
            assert basis is not None
            assert basis["resolved_to"] == "series-b", (
                f"{on_ladder_unpublished} should descend the ladder to series-b, got {basis['resolved_to']}"
            )
        for off_ladder in ("growth", "not-a-real-stage", "bridge"):
            _ranges, basis = mod._resolve_stage_ranges(off_ladder)
            assert basis is not None
            assert basis["resolved_to"] == "seed", (
                f"{off_ladder} is not on the ladder; it should land on seed, not descend it, got {basis['resolved_to']}"
            )
            assert basis["reason"] == "stage not recognized"

    def test_a_missing_stage_is_disclosed_via_the_resolver_sentinel(self) -> None:
        """`_STAGE_UNSPECIFIED` is the sentinel callers must pass instead of
        defaulting to "seed" themselves before the resolver ever sees it."""
        mod = self._ve()
        ranges, basis = mod._resolve_stage_ranges(mod._STAGE_UNSPECIFIED)
        assert ranges is mod._STAGE_RANGES["seed"]
        assert basis is not None
        assert basis["resolved_to"] == "seed"
        assert basis["reason"] != "stage not recognized", (
            "a missing stage must carry its own reason, distinct from an "
            "unrecognized token, even though both resolve to seed"
        )

    def test_a_missing_stage_field_is_disclosed_not_silently_seeded(self) -> None:
        """ITEM 26b: `company.stage` absent entirely used to hit
        `.get("stage", "seed")`, which resolves straight through
        `stage in _STAGE_RANGES` with no substitution recorded — no
        disclosure at all, unlike the unrecognized-token path. The check
        result must now carry `stage_basis` in this case too."""
        mod = self._ve()
        inputs = {
            "company": {"company_name": "TestCo"},  # no "stage" key at all
            "cash": {"current_balance": 40_000_000, "monthly_net_burn": 900_000},
            "expenses": {"headcount": []},
        }
        result = mod._check_scale_plausibility(inputs, {"sheets": []})
        assert result.get("stage_basis") is not None, "missing stage substituted silently"
        assert result["stage_basis"]["resolved_to"] == "seed"
        assert result["stage_basis"]["reason"] != "stage not recognized"

    def test_a_missing_company_block_is_also_disclosed(self) -> None:
        """The same gap one level up: no "company" key at all."""
        mod = self._ve()
        inputs = {
            "cash": {"current_balance": 40_000_000, "monthly_net_burn": 900_000},
            "expenses": {"headcount": []},
        }
        result = mod._check_scale_plausibility(inputs, {"sheets": []})
        assert result.get("stage_basis") is not None, "missing company block substituted silently"
        assert result["stage_basis"]["resolved_to"] == "seed"

    def test_missing_stage_uses_the_same_seed_floor_as_before(self) -> None:
        """The fix must not change WHICH ranges are used for a missing stage —
        only whether the substitution is disclosed. Same seed floor, now
        surfaced instead of silent."""
        mod = self._ve()
        ranges, _basis = mod._resolve_stage_ranges(mod._STAGE_UNSPECIFIED)
        assert ranges["cash_balance"][0] == mod._STAGE_RANGES["seed"]["cash_balance"][0]
        assert ranges["monthly_burn"][0] == mod._STAGE_RANGES["seed"]["monthly_burn"][0]

    def test_plausibility_vote_also_discloses_a_missing_stage_via_resolver(self) -> None:
        """`_plausibility_vote` (the --fix gate) does the same `.get(..., "seed")`
        pattern as `_check_scale_plausibility` — confirm it now routes through
        the resolver too, rather than defaulting before the resolver runs."""
        mod = self._ve()
        inputs: dict[str, Any] = {
            "company": {"company_name": "TestCo"},
            "cash": {"current_balance": 40_000_000, "monthly_net_burn": 900_000},
            "expenses": {"headcount": []},
        }
        # Same ranges as an explicit "seed" — the vote's plausibility outcome
        # is unaffected by this fix, only the disclosure path upstream is.
        plausible_explicit, checked_explicit = mod._plausibility_vote(
            {**inputs, "company": {"company_name": "TestCo", "stage": "seed"}}
        )
        plausible_missing, checked_missing = mod._plausibility_vote(inputs)
        assert plausible_missing == plausible_explicit
        assert checked_missing == checked_explicit


class TestUnitEconomicsBenchmarkDisclosure:
    """A substituted stage benchmark must be visible in the artifact.

    Published medians do not exist for every stage. Substituting a neighbouring
    stage's is the honest option, but it changes every rating derived from it,
    so the substitution has to travel with the numbers rather than sit on stderr.
    """

    @staticmethod
    def _ue_inputs(stage: str) -> dict[str, Any]:
        return {
            "company": {
                "company_name": "TestCo",
                "slug": "testco",
                "stage": stage,
                "sector": "SaaS",
                "geography": "US",
            },
            "revenue": {"mrr": {"value": 500_000, "as_of": "2025-01"}, "growth_rate_monthly": 0.05},
            "unit_economics": {"gross_margin": 0.78, "cac": 12_000},
            "cash": {
                "current_balance": 20_000_000,
                "balance_date": "2025-01",
                "monthly_net_burn": 800_000,
            },
        }

    def test_substituted_stage_is_disclosed(self) -> None:
        for stage in ("series-c", "series-d"):
            rc, data, stderr = run_script(
                "unit_economics.py",
                ["--pretty"],
                stdin_data=json.dumps(self._ue_inputs(stage)),
            )
            assert rc == 0
            assert data is not None
            basis = data.get("benchmark_basis")
            assert basis is not None, f"{stage} rated against substituted benchmarks silently"
            assert basis["requested"] == stage
            assert basis["resolved_to"] in _load_fmr_unit_economics_module().STAGE_BENCHMARKS

    def test_stage_with_its_own_benchmarks_discloses_nothing(self) -> None:
        """Anti-vacuity: the key must be absent when no substitution happened."""
        for stage in ("seed", "series-a"):
            rc, data, stderr = run_script(
                "unit_economics.py",
                ["--pretty"],
                stdin_data=json.dumps(self._ue_inputs(stage)),
            )
            assert rc == 0
            assert data is not None
            assert "benchmark_basis" not in data, f"{stage} has its own benchmarks; disclosing a substitution is wrong"


def test_validate_burn_revenue_seed_above_threshold() -> None:
    """Seed stage burn > 10x MRR triggers warning."""
    inputs = _make_inputs(stage="seed", mrr=50_000, burn=600_000)  # 12x
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_REVENUE_SUSPECT" in codes


def test_validate_burn_revenue_pre_seed_skipped() -> None:
    """Pre-seed stage skips burn-to-revenue check entirely."""
    inputs = _make_inputs(stage="pre-seed", mrr=1_000, burn=200_000)  # 200x
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_REVENUE_SUSPECT" not in codes


def test_validate_burn_revenue_zero_mrr_no_trigger() -> None:
    """Zero MRR does not trigger burn-to-revenue check (pre-revenue guard)."""
    inputs = _make_inputs(stage="series-a", mrr=0, burn=500_000)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_REVENUE_SUSPECT" not in codes


def test_validate_burn_multiple_suspect() -> None:
    """Extreme burn multiple (> 10x) triggers BURN_MULTIPLE_SUSPECT."""
    # burn=1.44M/mo, MRR=170K, growth=1% → monthly net-new ARR = 170K * 0.01 * 12
    # = 20.4K; annual ≈ 244.8K; burn_multiple = (1.44M * 12) / 244.8K ≈ 71x
    inputs = _make_inputs(stage="series-a", mrr=170_000, burn=1_440_000, growth=0.01)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_MULTIPLE_SUSPECT" in codes
    assert data["has_critical_warnings"] is True


def test_validate_burn_multiple_suspect_ttm_overrides_growth_rate() -> None:
    """Time-series burn multiple should prevent false positive from growth-rate shortcut."""
    # Growth-rate shortcut: burn=150K, MRR=50K, growth=2% → monthly net-new ARR
    # = 50K*0.02*12 = 12K → burn_multiple = 150K/12K = 12.5x → would trigger
    # But TTM time-series: ARR grew from 400K to 1.4M → net_new_arr = 1M
    # burn_multiple = (150K*12)/1M = 1.8x → should NOT trigger
    inputs = _make_inputs(stage="series-a", mrr=50_000, burn=150_000, growth=0.02)
    inputs["revenue"]["monthly"] = [
        {"month": f"2025-{m:02d}", "arr": 400000 + i * (1000000 / 11)} for i, m in enumerate(range(1, 13))
    ]
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_MULTIPLE_SUSPECT" not in codes, (
        "TTM time-series should prevent false positive from growth-rate shortcut"
    )


def test_validate_burn_multiple_no_growth_no_trigger() -> None:
    """Missing growth data does not trigger burn multiple check."""
    inputs = _make_inputs(stage="series-a", mrr=170_000, burn=1_440_000, growth=None)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "BURN_MULTIPLE_SUSPECT" not in codes


def test_validate_has_critical_warnings_false_when_clean() -> None:
    """Clean inputs produce has_critical_warnings: false."""
    # burn=80K < 5*100K → no BURN_REVENUE_SUSPECT
    # burn_multiple = (80K*12)/(100K*0.1*12) = 8x → no BURN_MULTIPLE_SUSPECT
    inputs = _make_inputs(stage="series-a", mrr=100_000, burn=80_000, growth=0.10)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    assert data["has_critical_warnings"] is False


def test_validate_warning_overrides_valid() -> None:
    """Valid warning_overrides pass structural validation."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {
        "warning_overrides": [
            {
                "code": "BURN_MULTIPLE_SUSPECT",
                "reason": "Enterprise SaaS with lumpy deal flow",
                "reviewed_by": "agent",
                "timestamp": "2026-03-05T17:30:00Z",
            }
        ]
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    errors = data.get("errors", [])
    override_errors = [e for e in errors if e["code"].startswith("OVERRIDE_")]
    assert override_errors == [], f"Valid overrides should not produce errors: {override_errors}"


def test_validate_warning_overrides_missing_keys() -> None:
    """warning_overrides entry missing required keys produces OVERRIDE_MISSING_KEYS."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {
        "warning_overrides": [{"code": "BURN_MULTIPLE_SUSPECT"}]  # missing reason, reviewed_by, timestamp
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    errors = data.get("errors", [])
    codes = [e["code"] for e in errors]
    assert "OVERRIDE_MISSING_KEYS" in codes


def test_validate_warning_overrides_invalid_reviewer() -> None:
    """warning_overrides entry with bad reviewed_by produces OVERRIDE_INVALID_REVIEWER."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {
        "warning_overrides": [
            {
                "code": "BURN_MULTIPLE_SUSPECT",
                "reason": "test",
                "reviewed_by": "nobody",
                "timestamp": "2026-03-05T17:30:00Z",
            }
        ]
    }
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    errors = data.get("errors", [])
    codes = [e["code"] for e in errors]
    assert "OVERRIDE_INVALID_REVIEWER" in codes


# --- validate_inputs.py override filtering tests (direct import) ---


def _import_validate() -> Any:
    """Import validate() directly from validate_inputs.py."""
    sys.path.insert(0, FMR_SCRIPTS_DIR)
    try:
        import validate_inputs  # type: ignore[import-not-found]

        return validate_inputs.validate
    finally:
        sys.path.remove(FMR_SCRIPTS_DIR)


def test_validate_founder_override_does_not_clear_critical() -> None:
    """Founder override is informational — does NOT clear has_critical_warnings."""
    validate = _import_validate()
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["growth_rate_monthly"] = 0.0
    inputs["revenue"]["mrr"] = {"value": 50000, "as_of": "2026-01"}
    inputs["revenue"]["customers"] = 100

    result = validate(inputs)
    assert result["has_critical_warnings"] is True

    # Founder override: still has_critical_warnings
    warning_field = next(w["field"] for w in result["warnings"] if w["code"] == "GROWTH_RATE_ZERO_SUSPECT")
    inputs.setdefault("metadata", {})["warning_overrides"] = [
        {
            "code": "GROWTH_RATE_ZERO_SUSPECT",
            "field": warning_field,
            "reason": "pivot phase",
            "reviewed_by": "founder",
            "timestamp": "2026-03-09T14:00:00Z",
        }
    ]
    result2 = validate(inputs)
    assert result2["has_critical_warnings"] is True  # founder override does NOT clear


def test_validate_agent_override_clears_critical() -> None:
    """Agent override clears has_critical_warnings."""
    validate = _import_validate()
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["growth_rate_monthly"] = 0.0
    inputs["revenue"]["mrr"] = {"value": 50000, "as_of": "2026-01"}
    inputs["revenue"]["customers"] = 100

    result = validate(inputs)
    assert result["has_critical_warnings"] is True

    warning_field = next(w["field"] for w in result["warnings"] if w["code"] == "GROWTH_RATE_ZERO_SUSPECT")
    inputs.setdefault("metadata", {})["warning_overrides"] = [
        {
            "code": "GROWTH_RATE_ZERO_SUSPECT",
            "field": warning_field,
            "reason": "pivot phase — confirmed by founder",
            "reviewed_by": "agent",
            "timestamp": "2026-03-09T14:00:00Z",
        }
    ]
    result2 = validate(inputs)
    assert result2["has_critical_warnings"] is False
    # Warning still in list, just not blocking
    assert any(w["code"] == "GROWTH_RATE_ZERO_SUSPECT" for w in result2["warnings"])


def test_validate_honors_legacy_overrides_without_field() -> None:
    """Override without field (legacy agent format) suppresses by code only."""
    validate = _import_validate()
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["revenue"]["growth_rate_monthly"] = 0.0
    inputs["revenue"]["mrr"] = {"value": 50000, "as_of": "2026-01"}
    inputs["revenue"]["customers"] = 100

    result = validate(inputs)
    assert result["has_critical_warnings"] is True

    # Legacy override: code only, no field
    inputs.setdefault("metadata", {})["warning_overrides"] = [
        {
            "code": "GROWTH_RATE_ZERO_SUSPECT",
            "reason": "pivot phase",
            "reviewed_by": "agent",
            "timestamp": "2026-03-09T14:00:00Z",
        }
    ]
    result2 = validate(inputs)
    assert result2["has_critical_warnings"] is False


def test_unit_economics_propagates_run_id() -> None:
    """unit_economics.py propagates metadata.run_id from input to output."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {"run_id": "test-run-001"}
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data.get("metadata", {}).get("run_id") == "test-run-001"


def test_unit_economics_no_run_id_no_metadata() -> None:
    """unit_economics.py without run_id in input produces no metadata in output."""
    payload = json.dumps(_VALID_INPUTS)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "metadata" not in data


def test_runway_propagates_run_id() -> None:
    """runway.py propagates metadata.run_id from input to output."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["metadata"] = {"run_id": "test-run-002"}
    payload = json.dumps(inputs)
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data.get("metadata", {}).get("run_id") == "test-run-002"


def test_checklist_propagates_run_id() -> None:
    """checklist.py propagates metadata.run_id from input to output."""
    items = _make_checklist_items()
    payload = json.dumps(
        {
            "items": items,
            "company": {"stage": "seed", "geography": "us", "sector": "B2B SaaS"},
            "metadata": {"run_id": "test-run-003"},
        }
    )
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data.get("metadata", {}).get("run_id") == "test-run-003"


_CO = {
    "company_name": "TestCo",
    "slug": "testco",
    "stage": "seed",
    "sector": "SaaS",
    "geography": "US",
    "revenue_model_type": "saas-plg",
}


def test_validate_date_format_errors() -> None:
    """validate_inputs.py catches malformed YYYY-MM dates."""
    payload = json.dumps(
        {
            "company": _CO,
            "cash": {
                "current_balance": 100000,
                "monthly_net_burn": 10000,
                "balance_date": "not-a-month",
            },
            "revenue": {
                "mrr": {"value": 5000, "as_of": "2026-2"},
            },
        }
    )
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["valid"] is False
    error_fields = {e["field"] for e in data.get("errors", [])}
    assert "cash.balance_date" in error_fields
    assert "revenue.mrr.as_of" in error_fields


def test_validate_enum_errors() -> None:
    """validate_inputs.py catches invalid enum values."""
    payload = json.dumps(
        {
            "company": {**_CO, "stage": "unicorn"},
            "structure": {"formatting_quality": "fair"},
        }
    )
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["valid"] is False
    error_fields = {e["field"] for e in data.get("errors", [])}
    assert "company.stage" in error_fields
    assert "structure.formatting_quality" in error_fields


# --- Currency determinism: top-level `currency` field ---


def test_validate_accepts_currency_field() -> None:
    """A top-level currency ISO code is a recognized, valid field."""
    payload = json.dumps({"company": _CO, "currency": "INR"})
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["valid"] is True
    assert data["errors"] == []


def test_validate_rejects_non_string_currency() -> None:
    """A non-string currency value is a structural TYPE_ERROR, not silently ignored."""
    payload = json.dumps({"company": _CO, "currency": 123})
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["valid"] is False
    error_fields = {e["field"] for e in data.get("errors", [])}
    assert "currency" in error_fields


def test_validate_absent_currency_backward_compatible() -> None:
    """Omitting currency entirely must validate identically to today (no new
    errors/warnings introduced by the field's mere absence)."""
    payload = json.dumps({"company": _CO})
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["valid"] is True
    assert data["errors"] == []


def test_validate_time_series_date_format() -> None:
    """validate_inputs.py catches malformed dates in time series arrays."""
    payload = json.dumps(
        {
            "company": _CO,
            "revenue": {
                "monthly": [
                    {"month": "2025-01", "actual": True, "total": 1000},
                    {"month": "Jan-2025", "actual": False, "total": 2000},
                ],
            },
        }
    )
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert data["valid"] is False
    assert any("Jan-2025" in e.get("message", "") for e in data.get("errors", []))


def test_validate_inputs_referenced_by_agent() -> None:
    """validate_inputs.py should exist on disk."""
    scripts_dir = os.path.join(SCRIPT_DIR, "..", "skills", "financial-model-review", "scripts")
    assert os.path.isfile(os.path.join(scripts_dir, "validate_inputs.py"))


# --- Agent structural smoke test ---


def test_agent_definition_references_valid_scripts() -> None:
    """All scripts referenced in agent workflow exist on disk."""
    agent_path = os.path.join(SCRIPT_DIR, "..", "agents", "financial-model-review.md")
    assert os.path.isfile(agent_path), "Agent definition not found"
    scripts_dir = os.path.join(SCRIPT_DIR, "..", "skills", "financial-model-review", "scripts")
    expected_scripts = [
        "extract_model.py",
        "checklist.py",
        "unit_economics.py",
        "runway.py",
        "compose_report.py",
        "visualize.py",
    ]
    for script in expected_scripts:
        assert os.path.isfile(os.path.join(scripts_dir, script)), f"Agent references {script} but it doesn't exist"
    # Verify SKILL.md exists
    skill_md = os.path.join(SCRIPT_DIR, "..", "skills", "financial-model-review", "SKILL.md")
    assert os.path.isfile(skill_md), "SKILL.md not found"


# --- ARPU field-name fallback tests ---


class TestValidateInputsArpuFallback:
    """validate_inputs.py must detect ARPU issues regardless of field name."""

    def _make_inputs_with_arpu(self, field_name: str, arpu_val: float) -> dict:
        """Build minimal inputs with ARPU under the given field name."""
        return {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}, "customers": 10},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {
                "ltv": {
                    "value": 6000,
                    "inputs": {field_name: arpu_val, "churn_monthly": 0.03, "gross_margin": 0.75},
                },
                "gross_margin": 0.75,
            },
        }

    def test_arpu_monthly_triggers_suspect(self) -> None:
        """ARPU_SUSPECT fires with canonical field name arpu_monthly."""
        inp = self._make_inputs_with_arpu("arpu_monthly", 60000)  # >= MRR
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "ARPU_SUSPECT" in codes

    def test_arpu_old_name_triggers_suspect(self) -> None:
        """ARPU_SUSPECT fires with old schema field name arpu."""
        inp = self._make_inputs_with_arpu("arpu", 60000)  # >= MRR
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "ARPU_SUSPECT" in codes

    def test_arpu_old_name_consistency_check(self) -> None:
        """ARPU_INCONSISTENT fires with old field name when ARPU*customers != MRR."""
        inp = self._make_inputs_with_arpu("arpu", 3000)  # 3000*10=30000 vs MRR 50000 → >20% gap
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "ARPU_INCONSISTENT" in codes

    def test_arpu_monthly_preferred_over_arpu(self) -> None:
        """When both arpu_monthly and arpu exist, arpu_monthly wins."""
        inp = self._make_inputs_with_arpu("arpu_monthly", 5000)
        inp["unit_economics"]["ltv"]["inputs"]["arpu"] = 60000  # old name, wrong value
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "ARPU_SUSPECT" not in codes  # 5000 < 50000 MRR, so no suspect


class TestValidateInputsArpuCritical:
    """ARPU_SUSPECT should be critical; CUSTOMERS_MISSING when LTV present but no customers."""

    def test_arpu_suspect_is_critical(self) -> None:
        """ARPU_SUSPECT should block at the stop-gate."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}, "customers": 10},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {
                "ltv": {"value": 6000, "inputs": {"arpu_monthly": 60000, "churn_monthly": 0.03, "gross_margin": 0.75}},
                "gross_margin": 0.75,
            },
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        suspect = [w for w in data["warnings"] if w["code"] == "ARPU_SUSPECT"]
        assert len(suspect) == 1
        assert suspect[0].get("critical") is True
        assert data["has_critical_warnings"] is True

    def test_customers_missing_with_ltv_at_seed(self) -> None:
        """CUSTOMERS_MISSING fires when LTV inputs present but revenue.customers absent."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {
                "ltv": {"value": 6000, "inputs": {"arpu_monthly": 5000, "churn_monthly": 0.03, "gross_margin": 0.75}},
                "gross_margin": 0.75,
            },
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CUSTOMERS_MISSING" in codes

    def test_customers_missing_not_at_preseed(self) -> None:
        """CUSTOMERS_MISSING should NOT fire at pre-seed."""
        inp = {
            "company": {"stage": "pre-seed"},
            "revenue": {"mrr": {"value": 5000}},
            "cash": {"monthly_net_burn": 20000},
            "unit_economics": {
                "ltv": {"value": 6000, "inputs": {"arpu_monthly": 5000}},
                "gross_margin": 0.75,
            },
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CUSTOMERS_MISSING" not in codes

    def test_customers_present_no_warning(self) -> None:
        """No CUSTOMERS_MISSING when revenue.customers is populated."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}, "customers": 10},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {
                "ltv": {"value": 6000, "inputs": {"arpu_monthly": 5000, "churn_monthly": 0.03, "gross_margin": 0.75}},
                "gross_margin": 0.75,
            },
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CUSTOMERS_MISSING" not in codes


class TestValidateInputsCashZero:
    """validate_inputs.py must flag $0 cash balance at seed+ as suspicious."""

    def test_zero_cash_at_series_a(self) -> None:
        """CASH_ZERO_SUSPECT fires as critical at series-a."""
        inp = {
            "company": {"stage": "series-a"},
            "revenue": {"mrr": {"value": 150000}},
            "cash": {"current_balance": 0, "monthly_net_burn": 500000},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        suspect = [w for w in data["warnings"] if w["code"] == "CASH_ZERO_SUSPECT"]
        assert len(suspect) == 1
        assert suspect[0].get("critical") is True
        assert data["has_critical_warnings"] is True

    def test_zero_cash_at_seed(self) -> None:
        """CASH_ZERO_SUSPECT fires at seed."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}},
            "cash": {"current_balance": 0, "monthly_net_burn": 80000},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CASH_ZERO_SUSPECT" in codes

    def test_zero_cash_at_preseed_no_warning(self) -> None:
        """CASH_ZERO_SUSPECT does NOT fire at pre-seed."""
        inp = {
            "company": {"stage": "pre-seed"},
            "revenue": {"mrr": {"value": 5000}},
            "cash": {"current_balance": 0, "monthly_net_burn": 20000},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CASH_ZERO_SUSPECT" not in codes

    def test_nonzero_cash_no_warning(self) -> None:
        """Normal cash balance produces no CASH_ZERO_SUSPECT."""
        inp = {
            "company": {"stage": "series-a"},
            "revenue": {"mrr": {"value": 150000}},
            "cash": {"current_balance": 5000000, "monthly_net_burn": 500000},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CASH_ZERO_SUSPECT" not in codes

    def test_null_cash_no_zero_warning(self) -> None:
        """Null cash triggers MISSING_CASH_BALANCE, not CASH_ZERO_SUSPECT."""
        inp = {
            "company": {"stage": "series-a"},
            "revenue": {"mrr": {"value": 150000}},
            "cash": {"monthly_net_burn": 500000},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "CASH_ZERO_SUSPECT" not in codes
        assert "MISSING_CASH_BALANCE" in codes


# --- ARPU/churn field-name fallback in unit_economics.py ---


class TestUnitEconomicsArpuFallback:
    """unit_economics.py LTV cap must work with both arpu and arpu_monthly."""

    def _make_inputs_zero_churn(self, arpu_field: str, arpu_val: float) -> dict:
        return {
            "company": {"stage": "seed", "sector": "B2B SaaS", "revenue_model_type": "saas-sales-led"},
            "revenue": {"mrr": {"value": 50000}, "arr": {"value": 600000}, "growth_rate_monthly": 0.08},
            "cash": {"current_balance": 2000000, "monthly_net_burn": 80000},
            "unit_economics": {
                "cac": {"total": 1500},
                "ltv": {
                    "value": 999999,
                    "inputs": {arpu_field: arpu_val, "churn_monthly": 0, "gross_margin": 0.75},
                },
                "gross_margin": 0.75,
            },
        }

    def test_zero_churn_cap_with_arpu_monthly(self) -> None:
        """60-month cap applies with canonical arpu_monthly."""
        inp = self._make_inputs_zero_churn("arpu_monthly", 500)
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        assert data is not None
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] == 500 * 0.75 * 60  # 22500

    def test_zero_churn_cap_with_arpu_old_name(self) -> None:
        """60-month cap applies with old schema field name arpu."""
        inp = self._make_inputs_zero_churn("arpu", 500)
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        assert data is not None
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] == 500 * 0.75 * 60  # 22500


# --- DERIVED_METRIC_REDUNDANT informational warning ---


class TestValidateInputsDerivedMetric:
    """validate_inputs.py warns when burn_multiple is provided alongside compute inputs."""

    def test_redundant_with_growth_inputs(self) -> None:
        """Warning fires when burn, mrr, and growth are all present."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}, "growth_rate_monthly": 0.08},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {"burn_multiple": 3.4, "gross_margin": 0.75},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "DERIVED_METRIC_REDUNDANT" in codes
        redundant = [w for w in data["warnings"] if w["code"] == "DERIVED_METRIC_REDUNDANT"]
        assert redundant[0].get("critical") is not True  # informational only

    def test_redundant_with_time_series(self) -> None:
        """Warning fires when monthly time-series has >= 12 entries."""
        monthly = [{"month": f"2024-{m:02d}", "actual": True, "total": 10000 + m * 1000} for m in range(1, 13)]
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}, "monthly": monthly},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {"burn_multiple": 3.4, "gross_margin": 0.75},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "DERIVED_METRIC_REDUNDANT" in codes

    def test_no_warning_when_fallback_needed(self) -> None:
        """No warning when compute inputs are missing (fallback is legitimate)."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}},  # no growth, no monthly
            "cash": {},  # no burn
            "unit_economics": {"burn_multiple": 3.4, "gross_margin": 0.75},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "DERIVED_METRIC_REDUNDANT" not in codes

    def test_no_warning_when_no_provided_bm(self) -> None:
        """No warning when burn_multiple is not provided."""
        inp = {
            "company": {"stage": "seed"},
            "revenue": {"mrr": {"value": 50000}, "growth_rate_monthly": 0.08},
            "cash": {"monthly_net_burn": 80000},
            "unit_economics": {"gross_margin": 0.75},
        }
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        codes = [w["code"] for w in data["warnings"]]
        assert "DERIVED_METRIC_REDUNDANT" not in codes


class TestBurnMultipleProvidedPreference:
    """When growth-rate burn multiple diverges >2x from provided, prefer provided."""

    def _make_inputs(self, growth_rate: float, provided_bm: float | None = None) -> dict[str, Any]:
        inp: dict[str, Any] = {
            "company": {"stage": "series-a", "sector": "B2B SaaS", "revenue_model_type": "saas-sales-led"},
            "revenue": {"mrr": {"value": 153603}, "arr": {"value": 1843235}, "growth_rate_monthly": growth_rate},
            "cash": {"current_balance": 1500000, "monthly_net_burn": 561000},
            "unit_economics": {"gross_margin": 0.784},
        }
        if provided_bm is not None:
            inp["unit_economics"]["burn_multiple"] = provided_bm
        return inp

    def test_divergent_prefers_provided(self) -> None:
        """15x growth-rate vs 3.05x provided → use 3.05x, warn about divergence.

        Derivation (period-matched formula):
          mrr=153603, g=0.02, burn=561000
          net_new_arr = 153603*0.02*12 = 36864.72
          computed = 561000/36864.72 = 15.22x
          ratio = max(15.22, 3.05) / min(15.22, 3.05) = 4.99 > 2.0 → prefer provided
        """
        inp = self._make_inputs(0.02, provided_bm=3.05)
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        bm = next(m for m in data["metrics"] if m["name"] == "burn_multiple")
        assert bm["value"] == 3.05
        # Value goes through normal rating branches (sanity + benchmark)
        # Divergence detail is in the warning
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "BURN_MULTIPLE_REPORTED_DIVERGENCE" in codes

    def test_close_values_uses_computed(self) -> None:
        """When growth-rate and provided are close, use computed.

        Derivation (period-matched formula):
          mrr=100000, g=0.05, burn=50000
          net_new_arr = 100000*0.05*12 = 60000
          computed = 50000/60000 = 0.83x
          provided = 0.85; ratio = 0.85/0.83 = 1.02 < 2.0 → use computed (0.83)
        """
        inp = {
            "company": {"stage": "series-a", "sector": "B2B SaaS", "revenue_model_type": "saas-sales-led"},
            "revenue": {"mrr": {"value": 100000}, "arr": {"value": 1200000}, "growth_rate_monthly": 0.05},
            "cash": {"current_balance": 5000000, "monthly_net_burn": 50000},
            "unit_economics": {"gross_margin": 0.75, "burn_multiple": 0.85},
        }
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        bm = next(m for m in data["metrics"] if m["name"] == "burn_multiple")
        assert bm["value"] == 0.83

    def test_no_provided_uses_computed(self) -> None:
        """Without provided burn_multiple, always use computed (existing behavior).

        Derivation (period-matched formula):
          mrr=153603, g=0.118, burn=561000
          net_new_arr = 153603*0.118*12 = 217,571.9
          computed = 561000/217571.9 = 2.58x
        """
        inp = self._make_inputs(0.118, provided_bm=None)
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        bm = next(m for m in data["metrics"] if m["name"] == "burn_multiple")
        # period-matched: 561000 / (153603*0.118*12) = 2.58
        assert bm["value"] == 2.58

    def test_time_series_path_unaffected_by_divergence_check(self) -> None:
        """Time-series burn multiple path should not be affected by the growth-rate divergence check."""
        inp = self._make_inputs(0.118, provided_bm=3.05)
        # Add TTM revenue data to trigger time-series path
        inp["revenue"]["quarterly"] = [
            {"quarter": "Q1-2025", "arr": 400000},
            {"quarter": "Q2-2025", "arr": 420000},
            {"quarter": "Q3-2025", "arr": 450000},
            {"quarter": "Q4-2025", "arr": 480000},
        ]
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        bm = next(m for m in data["metrics"] if m["name"] == "burn_multiple")
        # Time-series path should be used, not the provided value
        assert bm["value"] != 3.05
        # Should NOT produce BURN_MULTIPLE_REPORTED_DIVERGENCE (that's growth-rate path only)
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "BURN_MULTIPLE_REPORTED_DIVERGENCE" not in codes


class TestUnitEconomicsLtvSynthesis:
    """unit_economics.py synthesizes LTV when ltv.inputs missing but revenue has the data."""

    def _base_inputs(self) -> dict:
        return {
            "company": {"stage": "series-a", "sector": "B2B SaaS", "revenue_model_type": "saas-sales-led"},
            "revenue": {
                "mrr": {"value": 153603},
                "arr": {"value": 1843235},
                "growth_rate_monthly": 0.118,
                "customers": 45,
                "churn_monthly": 0.0067,
            },
            "cash": {"current_balance": 1500000, "monthly_net_burn": 561000},
            "unit_economics": {"gross_margin": 0.784},
        }

    def test_synthesizes_ltv_from_revenue(self) -> None:
        """LTV computed from revenue.customers + revenue.churn_monthly + gross_margin."""
        inp = self._base_inputs()
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        # arpu = 153603/45 = 3413.40, ltv = 3413.40 * 0.784 / 0.0067 ≈ 399,277.01
        expected_ltv = round(153603 / 45 * 0.784 / 0.0067, 2)
        assert ltv["value"] == expected_ltv
        assert "synthesized" in ltv["evidence"].lower() or "computed" in ltv["evidence"].lower()

    def test_no_synthesis_when_ltv_inputs_present(self) -> None:
        """Don't synthesize when ltv.inputs already has data."""
        inp = self._base_inputs()
        inp["unit_economics"]["ltv"] = {
            "value": 50000,
            "inputs": {"arpu_monthly": 3413, "churn_monthly": 0.0067, "gross_margin": 0.784},
        }
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] == 50000  # uses provided, not synthesized

    def test_preserves_existing_ltv_value_when_inputs_missing(self) -> None:
        """When ltv.value is provided but ltv.inputs is missing, synthesis fills inputs but keeps the value."""
        inp = self._base_inputs()
        inp["unit_economics"]["ltv"] = {"value": 75000}  # value present, inputs absent
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] == 75000  # must NOT be overwritten by synthesis
        # Evidence should NOT claim value was synthesized — only inputs were filled
        assert "synthesized from revenue.customers" not in ltv["evidence"].lower()

    def test_no_synthesis_without_customers(self) -> None:
        """Can't compute ARPU without customer count."""
        inp = self._base_inputs()
        del inp["revenue"]["customers"]
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] is None  # no data

    def test_no_synthesis_without_churn(self) -> None:
        """Can't compute LTV without churn."""
        inp = self._base_inputs()
        del inp["revenue"]["churn_monthly"]
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] is None

    def test_no_synthesis_with_zero_customers(self) -> None:
        """Zero customers → can't compute ARPU, no synthesis."""
        inp = self._base_inputs()
        inp["revenue"]["customers"] = 0
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        assert ltv["value"] is None

    def test_zero_churn_applies_60mo_cap(self) -> None:
        """Zero churn → 60-month LTV cap even with synthesized inputs."""
        inp = self._base_inputs()
        inp["revenue"]["churn_monthly"] = 0
        rc, data, _ = run_script("unit_economics.py", ["--pretty"], stdin_data=json.dumps(inp))
        assert rc == 0
        ltv = next(m for m in data["metrics"] if m["name"] == "ltv")
        # arpu = 153603/45 = 3413.40, capped ltv = 3413.40 * 0.784 * 60 = 160,564
        expected = round(153603 / 45 * 0.784 * 60, 2)
        assert ltv["value"] == expected


# ---------------------------------------------------------------------------
# ARPU-vs-MRR derived divergence regression tests
# ---------------------------------------------------------------------------


def test_validate_arpu_derived_divergence() -> None:
    """When stated ARPU diverges >20% from MRR/customers, warn."""
    inputs = _make_inputs(stage="series-a", mrr=100_000, burn=80_000, growth=0.10)
    inputs["revenue"]["customers"] = 50
    inputs["unit_economics"] = {
        "ltv": {
            "inputs": {"arpu_monthly": 3400}  # 100K/50 = 2000, 3400 is 70% higher
        }
    }
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "ARPU_INCONSISTENT" in codes


def test_validate_arpu_derived_close_no_warning() -> None:
    """When stated ARPU is within 20% of MRR/customers, no warning."""
    inputs = _make_inputs(stage="series-a", mrr=100_000, burn=80_000, growth=0.10)
    inputs["revenue"]["customers"] = 50
    inputs["unit_economics"] = {
        "ltv": {
            "inputs": {"arpu_monthly": 2100}  # 100K/50 = 2000, 2100 is 5% higher
        }
    }
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["warnings"]]
    assert "ARPU_INCONSISTENT" not in codes


# ---------------------------------------------------------------------------
# Runway: cash direction warning + profitability regression tests
# ---------------------------------------------------------------------------


def test_runway_no_cash_direction_warning_when_profitable() -> None:
    """Growth-driven profitability should not trigger cash direction warning."""
    # High growth (20% MoM) with moderate burn → revenue overtakes expenses
    inputs = {
        "company": {
            "company_name": "GrowthCo",
            "slug": "growthco",
            "stage": "seed",
            "sector": "SaaS",
            "geography": "US",
        },
        "revenue": {
            "mrr": {"value": 50_000, "as_of": "2025-01"},
            "growth_rate_monthly": 0.20,
        },
        "cash": {
            "current_balance": 500_000,
            "balance_date": "2025-01",
            "monthly_net_burn": 100_000,
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    # With 20% MoM growth, revenue overtakes expenses → became profitable
    assert base["default_alive"] is True
    # Cash direction warning should NOT fire — growth explains cash increase
    assert base.get("cash_direction_warning") is None


def test_runway_cash_direction_warning_with_grant_no_profitability() -> None:
    """Grant-funded cash increase without profitability should still warn.

    IIA grants add cash monthly, making the company never run out of cash
    (default_alive = True), but the company never becomes cash-flow positive
    (revenue < expenses throughout). Cash increases due to grants, not
    operational profitability — the warning should fire.
    """
    inputs = {
        "company": {"company_name": "GrantCo", "slug": "grantco", "stage": "seed", "sector": "SaaS", "geography": "IL"},
        "revenue": {
            "mrr": {"value": 10_000, "as_of": "2025-01"},
            "growth_rate_monthly": 0.0,  # zero growth → never profitable
        },
        "cash": {
            "current_balance": 500_000,
            "balance_date": "2025-01",
            "monthly_net_burn": 20_000,
            "grants": {
                "iia_approved": 3_000_000,  # 50K/mo for 60 months
                "iia_disbursement_months": 60,
                "iia_start_month": 1,
            },
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    # Revenue (10K) < opex (30K) throughout, never becomes profitable
    # But default_alive is True (never runs out of cash due to grants)
    # Cash increases due to grant inflows — warning SHOULD fire
    assert base["default_alive"] is True
    final_cash = base["monthly_projections"][-1]["cash_balance"]
    assert final_cash > 500_000  # cash increased
    assert base.get("cash_direction_warning") is not None
    # Risk narrative should NOT say "reaches profitability"
    assert "reaches profitability" not in data.get("risk_assessment", "")


# === v0.4.1 Phase 3 Task 9: compose on-disk verification + tolerant JSON extraction ===


def test_compose_verifies_outputs_exist_after_write(tmp_path: Any) -> None:
    """After successful compose, both report.json and report.md must exist on disk."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    json_path = os.path.join(d, "report.json")
    md_path = os.path.join(d, "report.md")
    rc, _, err = run_script(
        "compose_report.py",
        ["-d", d, "-o", json_path, "--write-md", md_path],
    )
    assert rc == 0, err
    assert os.path.isfile(json_path)
    assert os.path.isfile(md_path)
    assert os.path.getsize(json_path) > 0
    assert os.path.getsize(md_path) > 0


def test_compose_exits_nonzero_if_write_md_path_unwritable(tmp_path: Any) -> None:
    """Compose must exit nonzero if --write-md target dir doesn't exist and can't be created."""
    import pathlib

    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    # Point --write-md at a path inside a read-only parent
    ro_parent = pathlib.Path(str(tmp_path)) / "readonly"
    ro_parent.mkdir(mode=0o555)
    bad_md_path = str(ro_parent / "no-write" / "report.md")
    json_path = os.path.join(d, "report.json")
    rc, _, err = run_script(
        "compose_report.py",
        ["-d", d, "-o", json_path, "--write-md", bad_md_path],
    )
    assert rc != 0, "compose should exit nonzero when --write-md target is unwritable"
    # Cleanup: restore writable mode so tmp_path can be deleted
    os.chmod(str(ro_parent), 0o755)


# === v0.4.1 Phase 3 Task 9: tolerant JSON extraction from sub-agent messages ===


def test_extract_dispatch_json_raw_object() -> None:
    import sys

    sys.path.insert(
        0,
        os.path.join(os.path.dirname(SCRIPT_DIR), "skills", "financial-model-review", "scripts"),
    )
    from _dispatch_json import extract_dispatch_json  # type: ignore[import-not-found]

    assert extract_dispatch_json('{"a": 1, "b": 2}') == {"a": 1, "b": 2}


def test_extract_dispatch_json_fenced() -> None:
    import sys

    sys.path.insert(
        0,
        os.path.join(os.path.dirname(SCRIPT_DIR), "skills", "financial-model-review", "scripts"),
    )
    from _dispatch_json import extract_dispatch_json  # type: ignore[import-not-found]

    assert extract_dispatch_json('```json\n{"a": 1}\n```') == {"a": 1}


def test_extract_dispatch_json_nested() -> None:
    """Critical regression test: must not truncate on inner }."""
    import sys

    sys.path.insert(
        0,
        os.path.join(os.path.dirname(SCRIPT_DIR), "skills", "financial-model-review", "scripts"),
    )
    from _dispatch_json import extract_dispatch_json  # type: ignore[import-not-found]

    text = '```json\n{"a": {"b": 1}, "c": 2}\n```'
    assert extract_dispatch_json(text) == {"a": {"b": 1}, "c": 2}


def test_extract_dispatch_json_embedded_in_prose() -> None:
    import sys

    sys.path.insert(
        0,
        os.path.join(os.path.dirname(SCRIPT_DIR), "skills", "financial-model-review", "scripts"),
    )
    from _dispatch_json import extract_dispatch_json  # type: ignore[import-not-found]

    text = 'Here is the result:\n{"a": 1, "b": 2}\nLet me know if anything is wrong.'
    assert extract_dispatch_json(text) == {"a": 1, "b": 2}


def test_extract_dispatch_json_raises_when_no_json() -> None:
    import sys

    import pytest

    sys.path.insert(
        0,
        os.path.join(os.path.dirname(SCRIPT_DIR), "skills", "financial-model-review", "scripts"),
    )
    from _dispatch_json import extract_dispatch_json  # type: ignore[import-not-found]

    with pytest.raises(ValueError):
        extract_dispatch_json("Just some prose with no JSON object anywhere.")


# --- validate_inputs.py expense-coverage monthly_total regression ---


def _make_inputs_with_expenses(
    monthly_total: float | None = None,
    mrr_value: float | None = None,
    burn: float = 50_000,
) -> dict[str, Any]:
    """Build a validate_inputs fixture with headcount + opex but no mrr.value.

    extracted = 4 engineers × 180K/12 + 40K opex = 60K + 40K = 100K.
    With burn=50K and revenue=200K: expected_expenses = 250K;
    threshold = 125K; 100K < 125K → EXPENSE_COVERAGE_SUSPECT must fire.
    """
    inputs: dict[str, Any] = {
        "company": {
            "company_name": "TestCo",
            "slug": "testco",
            "stage": "series-a",
            "sector": "SaaS",
            "geography": "US",
            "revenue_model_type": "saas-plg",
        },
        "revenue": {},
        "cash": {
            "current_balance": 2_000_000,
            "balance_date": "2025-01",
            "monthly_net_burn": burn,
        },
        "expenses": {
            "headcount": [
                {"role": "Engineering", "count": 4, "salary_annual": 180_000},
            ],
            "opex_monthly": [
                {"category": "cloud", "amount": 40_000},
            ],
        },
    }
    if mrr_value is not None:
        inputs["revenue"]["mrr"] = {"value": mrr_value, "as_of": "2025-01"}
    if monthly_total is not None:
        inputs["revenue"]["monthly_total"] = monthly_total
    return inputs


def test_expense_coverage_counts_monthly_total_revenue() -> None:
    """expected_expenses must include revenue.monthly_total when mrr is absent.

    Regression: the old code read only revenue.mrr.value; when mrr was absent
    rev fell back to 0, understating expected_expenses and suppressing the
    coverage check for monthly_total-only companies.

    Numbers: burn=50K, monthly_total=200K, extracted=100K.
    Fixed → expected=250K, threshold=125K, 100K < 125K → warning FIRES.
    Bug   → expected= 50K, threshold= 25K, 100K >= 25K → warning silent (RED).
    """
    inputs = _make_inputs_with_expenses(monthly_total=200_000)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data.get("warnings", [])]
    assert "EXPENSE_COVERAGE_SUSPECT" in codes, (
        f"Expected EXPENSE_COVERAGE_SUSPECT with monthly_total revenue, got: {codes}"
    )


def test_expense_coverage_mrr_and_monthly_total_parity() -> None:
    """Same numbers via revenue.mrr.value must also trigger EXPENSE_COVERAGE_SUSPECT.

    Parity check: the mrr path should already work; this ensures both paths
    produce identical outcomes.
    """
    inputs = _make_inputs_with_expenses(mrr_value=200_000)
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data.get("warnings", [])]
    assert "EXPENSE_COVERAGE_SUSPECT" in codes, f"Expected EXPENSE_COVERAGE_SUSPECT with mrr revenue, got: {codes}"


# --- Task 16a: checklist.py business_quality_pct None when biz_applicable == 0 ---

# Business item IDs (all non-structural categories)
_BIZ_ITEM_IDS: list[str] = [
    "UNIT_10",
    "UNIT_11",
    "UNIT_12",
    "UNIT_13",
    "UNIT_14",
    "UNIT_15",
    "UNIT_16",
    "UNIT_17",
    "UNIT_18",
    "UNIT_19",
    "METRIC_33",
    "METRIC_34",
    "METRIC_35",
    "BRIDGE_36",
    "BRIDGE_37",
    "BRIDGE_38",
    "SECTOR_39",
    "SECTOR_40",
    "SECTOR_41",
    "SECTOR_42",
    "SECTOR_43",
    "SECTOR_44",
    "OVERALL_45",
    "OVERALL_46",
]


def test_checklist_business_quality_pct_none_when_all_biz_na() -> None:
    """business_quality_pct must be None (not 0.0) when every business item is
    not_applicable — mirrors the existing model_maturity_pct None semantics.

    Regression: the old else-branch returned 0.0, misrepresenting 'no applicable
    items' as 'zero score'."""
    na_overrides = {item_id: {"status": "not_applicable", "evidence": "N/A"} for item_id in _BIZ_ITEM_IDS}
    items = _make_checklist_items(overrides=na_overrides)
    payload = json.dumps({"items": items})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    biz_pct = data["summary"]["business_quality_pct"]
    assert biz_pct is None, f"Expected business_quality_pct=None when all business items are N/A, got: {biz_pct}"


# --- Task 16b: compose_report.py RUNWAY_INCONSISTENCY near-zero cash guard ---


def test_compose_runway_inconsistency_suppressed_for_near_zero_cash() -> None:
    """RUNWAY_INCONSISTENCY must NOT fire when inputs_cash is near-zero (< 1000).

    Regression: the guard used `inputs_cash != 0` so a cash balance of $1 with
    debt $0 triggered a 49 900% delta against any non-trivial runway net_cash,
    producing a spurious warning.  Fixed to abs(inputs_cash) >= 1000."""
    # inputs_cash = 1 (current_balance=1, no debt)
    # runway net_cash = 500  →  delta_pct ≈ 49 900% > 10%
    # Bug  → inputs_cash(1) != 0 is True  → warning fires
    # Fix  → abs(1) >= 1000 is False       → warning suppressed
    inputs_tiny_cash = json.loads(json.dumps(_VALID_INPUTS))
    inputs_tiny_cash["cash"]["current_balance"] = 1
    inputs_tiny_cash["cash"].pop("debt", None)

    runway_different = json.loads(json.dumps(_VALID_RUNWAY))
    runway_different["baseline"]["net_cash"] = 500

    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_tiny_cash,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_different,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    warnings = data["validation"]["warnings"]
    codes = [w["code"] for w in warnings]
    assert "RUNWAY_INCONSISTENCY" not in codes, (
        f"RUNWAY_INCONSISTENCY should be suppressed for near-zero cash (inputs_cash=1), got: {codes}"
    )


# --- Task 16c: runway.py breakeven warning when burn == 0 and no cash balance ---


def test_runway_breakeven_warning_when_burn_zero_no_cash() -> None:
    """runway.py must emit a 'breakeven' warning when monthly_net_burn=0 and
    current_balance is absent — the sensitivity table is not meaningful in
    this case."""
    inputs: dict[str, Any] = {
        "company": {
            "company_name": "TestCo",
            "slug": "testco",
            "stage": "series-a",
            "sector": "SaaS",
            "geography": "US",
        },
        "cash": {
            "monthly_net_burn": 0,
            # deliberately no current_balance
        },
        "revenue": {
            "mrr": {"value": 100_000, "as_of": "2025-01"},
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0
    assert data is not None
    warnings_list = data.get("warnings", [])
    # Warnings may be strings or dicts
    warnings_text = " ".join(w if isinstance(w, str) else str(w.get("message", w)) for w in warnings_list)
    assert "breakeven" in warnings_text.lower(), (
        f"Expected a 'breakeven' warning when burn=0 and no cash balance, got: {warnings_list}"
    )


# ---------------------------------------------------------------------------
# Null-coercion regression: present-but-null numeric fields must not crash the
# math producers. The corrections layer writes None for blank/cleared cells.
# ---------------------------------------------------------------------------


def test_runway_null_debt_does_not_crash() -> None:
    """cash.debt: null must not raise TypeError (the .get default only applies
    to missing keys, not explicit JSON null)."""
    inputs = {
        "company": {"company_name": "TestCo", "stage": "seed"},
        "cash": {"current_balance": 1_000_000, "monthly_net_burn": 50_000, "debt": None},
    }
    rc, data, stderr = run_script("runway.py", stdin_data=json.dumps(inputs))
    assert rc == 0, f"runway.py crashed on null debt: {stderr}"
    assert data.get("baseline", {}).get("net_cash") == 1_000_000


def test_runway_null_grant_and_target_fields_do_not_crash() -> None:
    """IIA grant fields, ils_expense_fraction, and runway_target_months set to
    null must not raise TypeError."""
    inputs = {
        "company": {"company_name": "TestCo", "stage": "seed"},
        "cash": {
            "current_balance": 1_000_000,
            "monthly_net_burn": 50_000,
            "grants": {
                "iia_approved": 500_000,
                "iia_disbursement_months": None,
                "iia_start_month": None,
            },
            "fundraising": {"target_raise": 2_000_000},
        },
        "bridge": {"runway_target_months": None},
        "israel_specific": {"fx_rate_ils_usd": 3.5, "ils_expense_fraction": None},
    }
    rc, data, stderr = run_script("runway.py", stdin_data=json.dumps(inputs))
    assert rc == 0, f"runway.py crashed on null grant/target fields: {stderr}"
    assert data.get("post_raise") is not None


def test_unit_economics_null_headcount_count_does_not_crash() -> None:
    """A headcount entry with count/salary/burden null must not raise TypeError
    in either the ARR/FTE path or the magic-number S&M loop."""
    inputs = {
        "company": {"company_name": "TestCo", "stage": "seed"},
        "revenue_model_type": "saas-plg",
        "revenue": {
            "arr": {"value": 2_000_000},
            "mrr": {"value": 166_666},
            "growth_rate_monthly": 0.1,
        },
        "expenses": {
            "headcount": [
                {"role": "sales", "count": None, "salary_annual": None, "burden_pct": None},
                {"role": "eng", "count": 5, "salary_annual": 150_000, "burden_pct": 0.3},
            ]
        },
    }
    rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
    assert rc == 0, f"unit_economics.py crashed on null headcount count: {stderr}"
    assert data is not None


def test_compose_marker_collision_reflected_in_status_and_report() -> None:
    """When body content contains the marker prefix, MARKER_COLLISION must be
    in validation.warnings, status must be 'warnings' (not 'clean'), and the
    warning must render in the report's Validation Warnings section — i.e. the
    pre-scan happens before status + the Warnings section are finalized."""
    inputs_collide = json.loads(json.dumps(_VALID_INPUTS))
    inputs_collide["company"]["company_name"] = "TestCo <!-- COACHING_INSERTION_POINT_deadbeef -->"
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_collide,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    codes = [w["code"] for w in data["validation"]["warnings"]]
    assert "MARKER_COLLISION" in codes, "MARKER_COLLISION must be recorded in validation.warnings"
    assert data["validation"]["status"] == "warnings", "status must reflect MARKER_COLLISION"
    # The warning must actually render in the report body (not just JSON). The
    # Validation Warnings section humanizes the code to its label.
    md = data["report_markdown"]
    assert "## Validation Warnings" in md
    assert "Marker Collision" in md, "MARKER_COLLISION must render in the Warnings section"


def test_compose_fmt_usd_negative_net_cash() -> None:
    """Negative net_cash (debt > balance) must render as '-$..' not '$-..'."""
    inputs_debt = json.loads(json.dumps(_VALID_INPUTS))
    inputs_debt["cash"]["current_balance"] = 500_000
    inputs_debt["cash"]["debt"] = 2_000_000
    runway_neg = json.loads(json.dumps(_VALID_RUNWAY))
    runway_neg["baseline"]["net_cash"] = -1_500_000
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_debt,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_neg,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "$-1,500,000.00" not in md, "Negative must not fall through to $-.. form"
    assert "-$1.5M" in md, "Negative net cash should render as -$1.5M"


# === Currency-aware report formatting (non-USD models) ===


def test_compose_non_usd_report_has_no_bare_dollar_signs() -> None:
    """End-to-end: a non-USD-denominated model's report.md must not print a bare
    '$' on any monetary value — CAC/LTV in the Unit Economics table, Net
    Cash/Monthly Burn/Monthly Revenue and the raise amounts in the Runway
    section must all be tagged with the model's native ISO code instead.
    unit_economics.json / runway.json now echo `currency` in their own output,
    so compose_report.py reads it from those artifacts rather than inputs.json."""
    inputs_inr = json.loads(json.dumps(_VALID_INPUTS))
    inputs_inr["currency"] = "INR"
    ue_inr = json.loads(json.dumps(_VALID_UNIT_ECONOMICS))
    ue_inr["currency"] = "INR"
    runway_inr = json.loads(json.dumps(_VALID_RUNWAY))
    runway_inr["currency"] = "INR"
    runway_inr["post_raise"] = {
        "raise_amount": 5_000_000,
        "new_cash": 7_000_000,
        "new_runway_months": 30,
        "meets_target": True,
    }
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": inputs_inr,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": ue_inr,
            "runway.json": runway_inr,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    report = data["report_markdown"]
    assert not re.search(r"\$\d", report), f"Found a bare-$ monetary value in a non-USD report:\n{report}"
    assert "INR" in report


def test_compose_usd_report_unaffected_by_currency_threading() -> None:
    """Back-compat: an explicit currency: 'USD' (or absent) must still render
    with the ordinary bare-$ formatting — the currency-aware formatting only
    changes behavior for a genuinely non-USD model."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    report = data["report_markdown"]
    assert "$" in report
    assert "USD" not in report  # back-compat: no currency tag noise for USD


# === run_id CLI stamping (alignment with the cross-skill contract) ===


def test_checklist_cli_run_id_overrides_stdin() -> None:
    """checklist.py: --run-id stamps metadata.run_id, overriding stdin (CLI > stdin)."""
    payload = json.dumps({"items": _make_checklist_items(), "metadata": {"run_id": "STDIN"}})
    rc, data, stderr = run_script("checklist.py", ["--pretty", "--run-id", "CLI-WINS"], stdin_data=payload)
    assert rc == 0, stderr
    assert data is not None and data.get("metadata", {}).get("run_id") == "CLI-WINS"


def test_unit_economics_cli_run_id_overrides_stdin() -> None:
    """unit_economics.py: --run-id overrides stdin-passthrough run_id."""
    inp = dict(_VALID_INPUTS)
    inp["metadata"] = {"run_id": "STDIN"}
    rc, data, stderr = run_script("unit_economics.py", ["--pretty", "--run-id", "CLI-WINS"], stdin_data=json.dumps(inp))
    assert rc == 0, stderr
    assert data is not None and data.get("metadata", {}).get("run_id") == "CLI-WINS"


def test_runway_cli_run_id_overrides_stdin() -> None:
    """runway.py: --run-id overrides stdin-passthrough run_id."""
    inp = dict(_VALID_INPUTS)
    inp["metadata"] = {"run_id": "STDIN"}
    rc, data, stderr = run_script("runway.py", ["--pretty", "--run-id", "CLI-WINS"], stdin_data=json.dumps(inp))
    assert rc == 0, stderr
    assert data is not None and data.get("metadata", {}).get("run_id") == "CLI-WINS"


# === Fix C: default-alive note includes projected breakeven month ===


def test_runway_default_alive_note_includes_breakeven_month() -> None:
    """When a scenario is default-alive, its note must include the projected breakeven month.

    The note already explains default_alive semantics; this extends it with
    '; projected to reach cash-flow breakeven around month N of the projection'.
    """
    inputs = {
        "company": {
            "company_name": "GrowthCo",
            "slug": "growthco",
            "stage": "seed",
            "sector": "SaaS",
            "geography": "US",
        },
        "revenue": {
            "mrr": {"value": 50_000, "as_of": "2025-01"},
            "growth_rate_monthly": 0.20,  # 20% MoM → reaches breakeven within ~3 months
        },
        "cash": {
            "current_balance": 2_000_000,
            "balance_date": "2025-01",
            "monthly_net_burn": 100_000,
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0, stderr
    assert data is not None
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    assert base["default_alive"] is True
    note = base.get("note", "")
    assert "breakeven" in note.lower(), f"default-alive note must include breakeven month; got: {note!r}"
    # Note must name a specific month number
    import re as _re

    assert _re.search(r"month \d+", note), f"note must say 'month N' for a specific breakeven month; got: {note!r}"


def test_runway_default_alive_note_no_breakeven_when_not_profitable() -> None:
    """When default-alive via grant (not profitability), no breakeven month in note."""
    inputs = {
        "company": {"company_name": "GrantCo", "slug": "grantco", "stage": "seed", "sector": "SaaS", "geography": "IL"},
        "revenue": {
            "mrr": {"value": 10_000, "as_of": "2025-01"},
            "growth_rate_monthly": 0.0,  # zero growth — never reaches breakeven
        },
        "cash": {
            "current_balance": 500_000,
            "balance_date": "2025-01",
            "monthly_net_burn": 20_000,
            "grants": {
                "iia_approved": 3_000_000,
                "iia_disbursement_months": 60,
                "iia_start_month": 1,
            },
        },
    }
    rc, data, stderr = run_script("runway.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert rc == 0, stderr
    assert data is not None
    base = next(s for s in data["scenarios"] if s["name"] == "base")
    assert base["default_alive"] is True
    note = base.get("note", "")
    # Without profitability, the note should not claim a breakeven month
    assert "breakeven" not in note.lower() or "month" not in note.lower(), (
        f"Note should not claim a breakeven month when company never becomes profitable; got: {note!r}"
    )


def test_compose_default_alive_breakeven_month_in_runway_table() -> None:
    """Infinite runway row in report should include '~month N' when breakeven is derivable."""
    import copy

    runway_da = copy.deepcopy(_VALID_RUNWAY)
    # Make base scenario default-alive with null runway_months
    runway_da["scenarios"][0]["runway_months"] = None
    runway_da["scenarios"][0]["cash_out_date"] = None
    runway_da["scenarios"][0]["decision_point"] = None
    runway_da["scenarios"][0]["default_alive"] = True
    # Add synthetic projections where month 5 reaches breakeven
    runway_da["scenarios"][0]["monthly_projections"] = [
        {
            "month": i,
            "cash_balance": 2_000_000,
            "revenue": 50_000 + i * 20_000,
            "expenses": 150_000,
            "net_burn": 150_000 - (50_000 + i * 20_000),
        }
        for i in range(1, 10)
    ]
    # Month 5: revenue = 50K + 5*20K = 150K, net_burn = 150K - 150K = 0
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": runway_da,
        }
    )
    rc, data, stderr = _run_compose(d)
    assert rc == 0, stderr
    assert data is not None
    md = data["report_markdown"]
    assert "breakeven" in md.lower() or "~month" in md.lower(), (
        f"Compose report should include breakeven month for default-alive scenario; got runway section:\n{md}"
    )


# === Fix D: apply_corrections.py neutral informational stderr for corrected payload ===


def test_apply_corrections_corrected_payload_no_warning_on_stderr() -> None:
    """apply_corrections.py must NOT print 'Warning:' or 'legacy' to stderr for the
    corrected-object payload shape — that is the documented dispatch contract.
    Instead it emits a neutral informational line.
    """
    original: dict[str, Any] = {
        "company": {"company_name": "TestCo", "stage": "seed"},
        "cash": {"current_balance": 1_000_000, "monthly_net_burn": 50_000},
    }
    # corrected-object shape (the documented dispatch contract)
    payload = {
        "corrected": {
            "company": {"company_name": "TestCo", "stage": "seed"},
            "cash": {"current_balance": 1_200_000, "monthly_net_burn": 50_000},
        },
        "corrections": [{"path": "cash.current_balance", "was": 1_000_000, "now": 1_200_000}],
    }
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as orig_f:
        json.dump(original, orig_f)
        orig_path = orig_f.name
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as corr_f:
        json.dump(payload, corr_f)
        corr_path = corr_f.name
    with tempfile.TemporaryDirectory() as out_dir:
        rc, data, stderr = run_script(
            "apply_corrections.py",
            [corr_path, "--original", orig_path, "--output-dir", out_dir, "--pretty"],
        )
    os.unlink(orig_path)
    os.unlink(corr_path)
    assert rc == 0, f"apply_corrections failed: {stderr}"
    # Must NOT contain "Warning:" or "legacy" — that is alarming for a normal run
    assert "Warning:" not in stderr, f"stderr must not contain 'Warning:' for corrected-payload shape; got: {stderr!r}"
    assert "legacy" not in stderr.lower(), f"stderr must not say 'legacy' for corrected-payload shape; got: {stderr!r}"
    # Must contain a neutral informational line
    assert "Info:" in stderr or "info:" in stderr.lower() or "corrected-object" in stderr.lower(), (
        f"stderr should contain a neutral informational message; got: {stderr!r}"
    )


# ===========================================================================
# Key-coverage tests: compose_report.py summary key reads vs. producer output
# ===========================================================================
#
# Invariant: every key the producer writes to a summary block must either be
# read by compose_report's section renderer OR appear in an explicit
# exclusion list (with a documented reason for skipping).
#
# Direction tested: produced ⊆ read ∪ explicitly-excluded.
# A new key added to checklist.py or unit_economics.py summary that the
# renderer silently ignores will fail with the offending key listed.
# ===========================================================================


def _load_compose_report_module() -> Any:
    """Import compose_report.py as a module with a unique sys.modules key."""
    import importlib.util
    import types

    key = "_fmr_keycov_compose_report"
    if key in sys.modules:
        return sys.modules[key]
    script_path = os.path.join(FMR_SCRIPTS_DIR, "compose_report.py")
    spec = importlib.util.spec_from_file_location(key, script_path)
    assert spec is not None and spec.loader is not None
    mod = types.ModuleType(key)
    mod.__spec__ = spec  # type: ignore[assignment]
    mod.__file__ = script_path  # type: ignore[assignment]
    sys.modules[key] = mod
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


class TestChecklistSummaryKeysCoverage:
    """compose_report._section_checklist must read every key that checklist.py
    writes to the summary block, OR the key must be in the explicit exclusion set.

    Producer (checklist.py) summary block keys:
        total, pass, fail, warn, not_applicable, score_pct,
        business_quality_pct, model_maturity_pct, overall_status,
        by_category, failed_items, warned_items

    Renderer (_section_checklist) reads via summary.get(...):
        score_pct, total, pass, fail, warn, not_applicable,
        overall_status, failed_items, warned_items, by_category

    Additionally _section_executive_summary reads:
        overall_status, score_pct, model_maturity_pct, business_quality_pct

    Explicit exclusions (not rendered by _section_checklist but used
    elsewhere or intentionally omitted from the Markdown section):
        business_quality_pct — rendered in _section_executive_summary only
                               (deck-only score; not a checklist-section concern)
        model_maturity_pct  — rendered in _section_executive_summary only
                               (structure sub-score; not a checklist-section concern)
    """

    # Keys checklist.py writes to summary{}.
    PRODUCED_SUMMARY_KEYS: set[str] = {
        "total",
        "pass",
        "fail",
        "warn",
        "not_applicable",
        "score_pct",
        "business_quality_pct",
        "model_maturity_pct",
        "overall_status",
        "by_category",
        "failed_items",
        "warned_items",
    }

    # Keys intentionally NOT rendered in _section_checklist but consumed
    # by _section_executive_summary (documented design split).
    EXCLUDED_FROM_CHECKLIST_SECTION: set[str] = {
        "business_quality_pct",
        "model_maturity_pct",
    }

    def test_produced_keys_rendered_or_excluded(self) -> None:
        """Every checklist summary key must be read by _section_checklist or
        appear in EXCLUDED_FROM_CHECKLIST_SECTION."""
        import re

        compose_script = os.path.join(FMR_SCRIPTS_DIR, "compose_report.py")
        with open(compose_script, encoding="utf-8") as fh:
            src = fh.read()

        # Locate _section_checklist function body
        fn_match = re.search(r"def _section_checklist\(.*?\n(?=def |\Z)", src, re.DOTALL)
        assert fn_match, "Could not locate _section_checklist in compose_report.py"
        fn_body = fn_match.group(0)

        # Extract every summary.get("...") key read in the function
        read_keys = set(re.findall(r'summary\.get\("([^"]+)"', fn_body))
        assert len(read_keys) >= 5, (
            f"Expected at least 5 summary.get() calls in _section_checklist, got {len(read_keys)}: {sorted(read_keys)}"
        )

        unread_and_not_excluded = self.PRODUCED_SUMMARY_KEYS - read_keys - self.EXCLUDED_FROM_CHECKLIST_SECTION
        assert not unread_and_not_excluded, (
            f"compose_report._section_checklist silently ignores checklist summary key(s): "
            f"{sorted(unread_and_not_excluded)}. Either read and render each key or add it to "
            f"EXCLUDED_FROM_CHECKLIST_SECTION with a documented reason."
        )

    def test_excluded_keys_read_by_executive_summary(self) -> None:
        """Keys excluded from _section_checklist must be read by
        _section_executive_summary (confirming they are rendered somewhere
        and the exclusion is not a silent drop)."""
        import re

        compose_script = os.path.join(FMR_SCRIPTS_DIR, "compose_report.py")
        with open(compose_script, encoding="utf-8") as fh:
            src = fh.read()

        fn_match = re.search(r"def _section_executive_summary\(.*?\n(?=def |\Z)", src, re.DOTALL)
        assert fn_match, "Could not locate _section_executive_summary in compose_report.py"
        fn_body = fn_match.group(0)

        exec_read_keys = set(re.findall(r'summary\.get\("([^"]+)"', fn_body))

        not_rendered_anywhere = self.EXCLUDED_FROM_CHECKLIST_SECTION - exec_read_keys
        assert not not_rendered_anywhere, (
            f"Checklist summary key(s) excluded from _section_checklist are ALSO not rendered "
            f"by _section_executive_summary: {sorted(not_rendered_anywhere)}. "
            f"Either render them somewhere or remove from PRODUCED_SUMMARY_KEYS."
        )

    def test_produced_summary_keys_min_count(self) -> None:
        """Guard against vacuous tests: producer summary must have >= 12 keys."""
        assert len(self.PRODUCED_SUMMARY_KEYS) >= 12, (
            f"PRODUCED_SUMMARY_KEYS expected >= 12 entries, got {len(self.PRODUCED_SUMMARY_KEYS)}. "
            f"Update when checklist.py changes its summary schema."
        )

    def test_live_producer_summary_keys_all_rendered_or_excluded(self) -> None:
        """Live checklist.py output summary keys must all appear in the read set
        or exclusion list.  Runs checklist.py on the existing _VALID_CHECKLIST fixture."""
        import re

        checklist_input = {
            "items": _VALID_CHECKLIST["items"],
            "company": {
                "company_name": "TestCo",
                "stage": "seed",
                "geography": "US",
                "revenue_model_type": "saas-sales-led",
            },
        }
        rc, data, stderr = run_script("checklist.py", stdin_data=json.dumps(checklist_input))
        assert rc == 0, f"checklist.py failed: {stderr}"
        assert isinstance(data, dict) and "summary" in data, "checklist.py output missing 'summary'"

        live_summary_keys = set(k for k, v in data["summary"].items() if v is not None)
        assert len(live_summary_keys) >= 8, (
            f"Expected >= 8 non-None summary keys from live checklist producer, got {live_summary_keys}"
        )

        compose_script = os.path.join(FMR_SCRIPTS_DIR, "compose_report.py")
        with open(compose_script, encoding="utf-8") as fh:
            src = fh.read()
        fn_match = re.search(r"def _section_checklist\(.*?\n(?=def |\Z)", src, re.DOTALL)
        assert fn_match
        read_keys = set(re.findall(r'summary\.get\("([^"]+)"', fn_match.group(0)))

        fn_exec_match = re.search(r"def _section_executive_summary\(.*?\n(?=def |\Z)", src, re.DOTALL)
        assert fn_exec_match
        exec_read_keys = set(re.findall(r'summary\.get\("([^"]+)"', fn_exec_match.group(0)))

        all_rendered = read_keys | exec_read_keys
        unrendered = live_summary_keys - all_rendered - self.EXCLUDED_FROM_CHECKLIST_SECTION
        assert not unrendered, (
            f"Live checklist.py summary key(s) not rendered by any compose_report section "
            f"and not in exclusion list: {sorted(unrendered)}."
        )


class TestUnitEconSummaryKeysCoverage:
    """compose_report._section_unit_economics must read every key that
    unit_economics.py writes to the summary block, OR the key must be in
    the explicit exclusion set.

    Producer (unit_economics.py) summary block keys:
        computed, strong, acceptable, warning, fail,
        not_rated, contextual, not_applicable

    Renderer (_section_unit_economics) reads:
        strong, acceptable, warning, fail

    Explicit exclusions (informational / display only):
        computed      — number of metrics that returned a non-None value;
                        used in visualize.py executive summary, not in the
                        Markdown section text.
        not_rated     — count of metrics with no benchmark; informational.
        contextual    — count of metrics rated contextual; informational.
        not_applicable — count of metrics that don't apply (SaaS-only, etc.);
                         informational.
    """

    PRODUCED_SUMMARY_KEYS: set[str] = {
        "computed",
        "strong",
        "acceptable",
        "warning",
        "fail",
        "not_rated",
        "contextual",
        "not_applicable",
    }

    # Keys the Markdown section doesn't render because they are informational
    # counts that don't affect the coaching narrative.
    EXCLUDED_INFORMATIONAL: set[str] = {
        "computed",
        "not_rated",
        "contextual",
        "not_applicable",
    }

    def test_produced_keys_rendered_or_excluded(self) -> None:
        """Every unit-economics summary key must be read by _section_unit_economics
        or appear in EXCLUDED_INFORMATIONAL."""
        import re

        compose_script = os.path.join(FMR_SCRIPTS_DIR, "compose_report.py")
        with open(compose_script, encoding="utf-8") as fh:
            src = fh.read()

        fn_match = re.search(r"def _section_unit_economics\(.*?\n(?=def |\Z)", src, re.DOTALL)
        assert fn_match, "Could not locate _section_unit_economics in compose_report.py"
        fn_body = fn_match.group(0)

        read_keys = set(re.findall(r'ue_summary\.get\("([^"]+)"', fn_body))
        assert len(read_keys) >= 4, (
            f"Expected >= 4 ue_summary.get() calls in _section_unit_economics, "
            f"got {len(read_keys)}: {sorted(read_keys)}"
        )

        unread_and_not_excluded = self.PRODUCED_SUMMARY_KEYS - read_keys - self.EXCLUDED_INFORMATIONAL
        assert not unread_and_not_excluded, (
            f"compose_report._section_unit_economics silently ignores unit-economics "
            f"summary key(s): {sorted(unread_and_not_excluded)}. Either render each key "
            f"or add it to EXCLUDED_INFORMATIONAL with a documented reason."
        )

    def test_excluded_keys_are_genuinely_informational(self) -> None:
        """Excluded keys must actually exist in the producer's output (not phantom).

        Verifies each excluded key appears in a live unit_economics.py run so
        the exclusion list doesn't silently mask producer renames.
        """
        full_saas_inputs = {
            "company": {
                "company_name": "TestCo",
                "stage": "seed",
                "revenue_model_type": "saas-sales-led",
            },
            "revenue": {
                "arr": {"value": 600_000, "as_of": "2025-12"},
                "mrr": {"value": 50_000, "as_of": "2025-12"},
                "growth_rate_monthly": 0.08,
            },
            "cash": {"current_balance": 2_000_000, "monthly_net_burn": 80_000},
            "unit_economics": {
                "cac": {"total": 1_500, "fully_loaded": True},
                "gross_margin": 0.75,
            },
        }
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(full_saas_inputs))
        assert rc == 0, f"unit_economics.py failed: {stderr}"
        assert isinstance(data, dict) and "summary" in data

        live_summary_keys = set(data["summary"].keys())
        phantom_exclusions = self.EXCLUDED_INFORMATIONAL - live_summary_keys
        assert not phantom_exclusions, (
            f"EXCLUDED_INFORMATIONAL contains key(s) that unit_economics.py does NOT actually "
            f"emit: {sorted(phantom_exclusions)}. Remove phantom entries from the exclusion list."
        )

    def test_produced_summary_keys_min_count(self) -> None:
        """Guard against vacuous tests: producer summary must have >= 8 keys."""
        assert len(self.PRODUCED_SUMMARY_KEYS) >= 8, (
            f"PRODUCED_SUMMARY_KEYS expected >= 8 entries, got {len(self.PRODUCED_SUMMARY_KEYS)}."
        )


# ---------------------------------------------------------------------------
# Magic number: period-matched formula
# ---------------------------------------------------------------------------
#
# Source: Scale Venture Partners, "Magic Number Math"
#   "Take the change in subscription revenue between two quarters, annualize it
#    (multiply by four), and divide the result by the sales and marketing spend
#    for the earlier of the two quarters."
# Period-matched monthly equivalent: net-new ARR (ΔMRR × 12) ÷ monthly S&M.
#
# Worked example: MRR $100K, 5% MoM (ΔMRR $5K), S&M $600K/yr ($50K/mo)
#   correct = (5K × 12) / 50K = 60K / 50K = 1.2
#   old bug = 60K / 600K = 0.1  (divided by annual instead of monthly)
# ---------------------------------------------------------------------------


class TestMagicNumberFormula:
    """unit_economics.py magic number must divide net-new ARR by monthly S&M."""

    _SAAS_INPUTS_TEMPLATE: dict[str, Any] = {
        "company": {
            "company_name": "TestCo",
            "stage": "seed",
            "sector": "B2B SaaS",
            "geography": "US",
            "revenue_model_type": "saas-sales-led",
        },
        "revenue": {
            "arr": {"value": 1_200_000, "as_of": "2025-12"},
            "mrr": {"value": 100_000, "as_of": "2025-12"},
            "growth_rate_monthly": 0.05,
        },
        "cash": {"current_balance": 2_000_000, "monthly_net_burn": 80_000},
        "expenses": {
            "headcount": [
                {
                    "role": "sales",
                    "count": 1,
                    "salary_annual": 600_000,
                    "burden_pct": 0.0,
                }
            ]
        },
    }

    def test_magic_number_period_matched(self) -> None:
        """Magic number ≈ 1.2 for MRR=100K, growth=5%, S&M=600K/yr (50K/mo).

        Derivation:
          net_new_ARR = ΔMRR × 12 = (100K × 0.05) × 12 = 60K
          monthly_sm  = 600K / 12 = 50K
          magic       = 60K / 50K = 1.2

        Period-mismatch (old bug) gives 60K / 600K = 0.1 — 12x understated.
        """
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(self._SAAS_INPUTS_TEMPLATE))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        magic = metrics.get("magic_number")
        assert magic is not None, "magic_number metric missing from output"
        assert magic["value"] is not None, "magic_number value is null"

        # Period-matched result ≈ 1.2; old bug gives 0.1
        assert abs(magic["value"] - 1.2) < 0.05, (
            f"magic_number expected ≈ 1.2 (period-matched), got {magic['value']:.4f}. "
            f"Verify net-new ARR (ΔMRR×12) is divided by monthly S&M."
        )

    def test_magic_number_not_12x_understated(self) -> None:
        """Magic number must NOT be 12x smaller than the correct value."""
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(self._SAAS_INPUTS_TEMPLATE))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        magic = metrics.get("magic_number")
        assert magic is not None
        if magic["value"] is not None:
            assert magic["value"] > 0.5, (
                f"magic_number = {magic['value']:.4f} is suspiciously low. "
                f"Expected ≈ 1.2 for the test fixture. Old bug gives 0.1."
            )


# ---------------------------------------------------------------------------
# GRR sanity: GRR > 1.0 is impossible by definition
# ---------------------------------------------------------------------------
#
# Source: Bessemer, "Gross Dollar Retention"
#   "GDR nets out the revenue from customers who turned off or downgraded…
#    but does not account for any expansion."
# Therefore GDR (GRR) ≤ 100% by definition — a value > 1.0 is a data error.
# ---------------------------------------------------------------------------


class TestGRRSanity:
    """validate_inputs.py Layer 3 must flag GRR > 1.0 as impossible."""

    def _make_grr_inputs(self, grr: float) -> dict[str, Any]:
        return {
            "company": {
                "company_name": "TestCo",
                "stage": "seed",
                "sector": "B2B SaaS",
                "geography": "US",
                "revenue_model_type": "saas-sales-led",
            },
            "revenue": {
                "mrr": {"value": 50_000, "as_of": "2025-12"},
                "grr": grr,
            },
            "cash": {"current_balance": 1_000_000, "monthly_net_burn": 80_000},
        }

    def test_grr_above_one_flagged(self) -> None:
        """GRR = 1.05 (105%) must trigger GRR_ABOVE_ONE in Layer 3."""
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(self._make_grr_inputs(1.05)))
        assert rc == 0
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "GRR_ABOVE_ONE" in codes, f"Expected GRR_ABOVE_ONE warning for grr=1.05. Got: {codes}"

    def test_grr_exactly_one_not_flagged(self) -> None:
        """GRR = 1.0 (100%) is at the boundary; no impossible-value flag."""
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(self._make_grr_inputs(1.0)))
        assert rc == 0
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "GRR_ABOVE_ONE" not in codes, f"GRR = 1.0 (100%) should not trigger GRR_ABOVE_ONE. Got: {codes}"

    def test_grr_normal_not_flagged(self) -> None:
        """GRR = 0.90 (90%) is valid; no impossible-value flag."""
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(self._make_grr_inputs(0.90)))
        assert rc == 0
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "GRR_ABOVE_ONE" not in codes

    def test_grr_above_one_is_critical(self) -> None:
        """GRR_ABOVE_ONE must be marked critical (impossible value, not just unusual)."""
        rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(self._make_grr_inputs(1.1)))
        assert rc == 0
        suspects = [w for w in data.get("warnings", []) if w["code"] == "GRR_ABOVE_ONE"]
        assert suspects, "GRR_ABOVE_ONE warning not raised"
        assert suspects[0].get("critical") is True, (
            f"GRR_ABOVE_ONE must be marked critical (impossible value). Got: {suspects[0]}"
        )


class TestBurnMultipleSuspectPeriodMatch:
    """The BURN_MULTIPLE_SUSPECT growth-rate fallback divides ANNUAL burn, so
    its net-new-ARR estimate must also be annual (monthly net-new ARR x 12).
    A period mismatch overstates the multiple 12x and false-flags healthy
    companies as data errors."""

    def _make_inputs(self, mrr: float, growth: float, burn: float) -> dict[str, Any]:
        return {
            "company": {
                "company_name": "TestCo",
                "stage": "seed",
                "sector": "B2B SaaS",
                "geography": "US",
                "revenue_model_type": "saas-sales-led",
            },
            "revenue": {
                "mrr": {"value": mrr, "as_of": "2025-12"},
                "growth_rate_monthly": growth,
            },
            "cash": {"current_balance": 1_000_000, "monthly_net_burn": burn},
        }

    def test_healthy_burn_multiple_not_flagged(self) -> None:
        """Burn 80K / MRR 50K / 8% MoM is a 1.67x burn multiple — no flag."""
        rc, data, _ = run_script(
            "validate_inputs.py", ["--pretty"], stdin_data=json.dumps(self._make_inputs(50_000, 0.08, 80_000))
        )
        assert rc == 0
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "BURN_MULTIPLE_SUSPECT" not in codes, f"1.67x burn multiple must not be flagged as suspect. Got: {codes}"

    def test_extreme_burn_multiple_flagged(self) -> None:
        """Burn 200K / MRR 100K / 0.1% MoM is a ~167x burn multiple — flagged."""
        rc, data, _ = run_script(
            "validate_inputs.py", ["--pretty"], stdin_data=json.dumps(self._make_inputs(100_000, 0.001, 200_000))
        )
        assert rc == 0
        codes = [w["code"] for w in data.get("warnings", [])]
        assert "BURN_MULTIPLE_SUSPECT" in codes, f"~167x burn multiple must be flagged as suspect. Got: {codes}"


# ---------------------------------------------------------------------------
# Rule of 40 growth-basis disclosure
# ---------------------------------------------------------------------------
#
# unit_economics.py annualizes the current MoM growth rate using
# (1+g)^12 - 1 — a forward annualization, not realized YoY.
# The evidence string must disclose this to avoid misleading readers.
# ---------------------------------------------------------------------------


class TestRuleOf40GrowthDisclosure:
    """Rule of 40 evidence string must disclose that growth is annualized from MoM."""

    _SAAS_INPUTS: dict[str, Any] = {
        "company": {
            "company_name": "TestCo",
            "stage": "series-a",
            "sector": "B2B SaaS",
            "geography": "US",
            "revenue_model_type": "saas-sales-led",
        },
        "revenue": {
            "arr": {"value": 6_000_000, "as_of": "2025-12"},
            "mrr": {"value": 500_000, "as_of": "2025-12"},
            "growth_rate_monthly": 0.08,
        },
        "cash": {"current_balance": 5_000_000, "monthly_net_burn": 300_000},
        "unit_economics": {"gross_margin": 0.75},
    }

    def test_r40_evidence_discloses_annualized_growth(self) -> None:
        """Rule of 40 evidence must include 'annualized' to signal forward projection."""
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(self._SAAS_INPUTS))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        r40 = metrics.get("rule_of_40")
        assert r40 is not None, "rule_of_40 metric missing"
        assert r40.get("rating") not in (
            "not_applicable",
            None,
        ), f"rule_of_40 not computed: {r40}"

        evidence = r40.get("evidence", "") or ""
        assert "annualized" in evidence.lower(), (
            f"rule_of_40 evidence must disclose that growth is annualized from MoM rate. Got: {evidence!r}"
        )


# ---------------------------------------------------------------------------
# Item 5: Runway scenarios table — Assumptions column
# ---------------------------------------------------------------------------


class TestRunwayAssumptionsColumn:
    """compose_report._section_runway() must render an Assumptions column
    with growth_rate, burn_change, fx_adjustment from scenario fields."""

    _BASE_ARTIFACTS = {
        "inputs.json": _VALID_INPUTS,
        "checklist.json": _VALID_CHECKLIST,
        "unit_economics.json": _VALID_UNIT_ECONOMICS,
    }

    def _run_with_runway(self, runway: dict) -> str:
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "runway.json": runway})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        return data["report_markdown"]  # type: ignore[no-any-return]

    def test_assumptions_column_header_present(self) -> None:
        """Scenarios table header must include 'Assumptions'."""
        md = self._run_with_runway(_VALID_RUNWAY)
        assert "Assumptions" in md

    def test_growth_rate_rendered_in_assumptions(self) -> None:
        """growth_rate field appears as 'growth X%/mo' in Assumptions column."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][1]["growth_rate"] = 0.05  # slow scenario
        md = self._run_with_runway(runway)
        assert "growth 5%/mo" in md

    def test_burn_change_rendered_in_assumptions(self) -> None:
        """burn_change field appears as 'burn X%' in Assumptions column."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][2]["burn_change"] = 0.20  # crisis scenario
        md = self._run_with_runway(runway)
        assert "burn +20%" in md

    def test_burn_change_negative_rendered_with_sign(self) -> None:
        """Negative burn_change shows negative sign (cost cut)."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][1]["burn_change"] = -0.15
        md = self._run_with_runway(runway)
        assert "burn -15%" in md

    def test_fx_adjustment_skipped_when_zero(self) -> None:
        """fx_adjustment == 0 must not appear in the table (no noise)."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][0]["fx_adjustment"] = 0.0
        md = self._run_with_runway(runway)
        # "fx" should not appear when value is zero
        assert "fx" not in md

    def test_no_assumptions_fields_renders_dash_or_empty(self) -> None:
        """When no scenario has growth_rate/burn_change/fx, Assumptions column
        cells are empty or '—' — not a Python None literal."""
        md = self._run_with_runway(_VALID_RUNWAY)
        assert "None" not in md


# ---------------------------------------------------------------------------
# Item 6: Executive summary — breakeven month derivation for default-alive
# ---------------------------------------------------------------------------


class TestDefaultAliveBreakevenInExecSummary:
    """When the base scenario is default-alive (runway_months=None), the
    executive summary Base Runway line must show the breakeven month derived
    from monthly_projections, matching the scenarios table cell."""

    _BASE_ARTIFACTS = {
        "inputs.json": _VALID_INPUTS,
        "checklist.json": _VALID_CHECKLIST,
        "unit_economics.json": _VALID_UNIT_ECONOMICS,
    }

    def _run_with_runway(self, runway: dict) -> str:
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "runway.json": runway})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        return data["report_markdown"]  # type: ignore[no-any-return]

    def test_exec_summary_shows_infinite_not_null_for_default_alive(self) -> None:
        """Base Runway line must not say 'None months' when default-alive."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][0]["runway_months"] = None
        runway["scenarios"][0]["cash_out_date"] = None
        runway["scenarios"][0]["decision_point"] = None
        runway["scenarios"][0]["default_alive"] = True
        md = self._run_with_runway(runway)
        assert "None months" not in md

    def test_exec_summary_shows_breakeven_month_from_projections(self) -> None:
        """When projections contain a month where net_burn <= 0, exec summary
        must reference that month (profitability / month N)."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][0]["runway_months"] = None
        runway["scenarios"][0]["cash_out_date"] = None
        runway["scenarios"][0]["decision_point"] = None
        runway["scenarios"][0]["default_alive"] = True
        runway["scenarios"][0]["monthly_projections"] = [
            {"month": 1, "net_burn": 20000},
            {"month": 2, "net_burn": 5000},
            {"month": 3, "net_burn": -1000},  # breakeven at month 3
            {"month": 4, "net_burn": -5000},
        ]
        md = self._run_with_runway(runway)
        # Should reference month 3 (breakeven)
        assert "3" in md

    def test_exec_summary_consistent_with_scenarios_table(self) -> None:
        """Both exec summary and scenarios table must indicate default-alive /
        infinite — no conflicting finite number in one and None in the other."""
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][0]["runway_months"] = None
        runway["scenarios"][0]["cash_out_date"] = None
        runway["scenarios"][0]["decision_point"] = None
        runway["scenarios"][0]["default_alive"] = True
        md = self._run_with_runway(runway)
        # "None months" must not appear in either location
        assert "None months" not in md
        # Report must use a consistent signal — Infinite or profitability
        assert "Infinite" in md or "profitability" in md.lower() or "default-alive" in md.lower()


# ---------------------------------------------------------------------------
# Fix A/B: static runway paired with the default-alive "Infinite" headline
# ---------------------------------------------------------------------------


class TestStaticRunwayPairedWithInfiniteHeadline:
    """finding 18/22: a default-alive scenario's 'Infinite' runway headline must
    never travel without the static (today's-burn, no-growth) floor — the
    projection holds burn flat while revenue compounds, an assumption a growing
    company's hiring plan rarely survives."""

    _BASE_ARTIFACTS = {
        "inputs.json": _VALID_INPUTS,
        "checklist.json": _VALID_CHECKLIST,
        "unit_economics.json": _VALID_UNIT_ECONOMICS,
    }

    def _run_with_runway(self, runway: dict) -> str:
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "runway.json": runway})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        return data["report_markdown"]  # type: ignore[no-any-return]

    def _default_alive_runway(self, static_months: float | None) -> dict:
        runway = json.loads(json.dumps(_VALID_RUNWAY))
        runway["scenarios"][0]["runway_months"] = None
        runway["scenarios"][0]["cash_out_date"] = None
        runway["scenarios"][0]["decision_point"] = None
        runway["scenarios"][0]["default_alive"] = True
        if static_months is not None:
            runway["scenarios"][0]["static_runway_months"] = static_months
        return runway  # type: ignore[no-any-return]

    def test_exec_summary_shows_static_runway_and_flat_burn_caveat(self) -> None:
        """The exec summary's 'Infinite' headline must be paired with the
        concrete static number and a flat-burn caveat, not left bare."""
        md = self._run_with_runway(self._default_alive_runway(6.9))
        assert "Infinite" in md
        assert "6.9 months" in md
        assert "flat" in md.lower()

    def test_exec_summary_omits_static_line_when_static_absent(self) -> None:
        """No static number available (e.g. zero initial burn) — must not
        fabricate one, and must not crash."""
        md = self._run_with_runway(self._default_alive_runway(None))
        assert "At Today's Burn" not in md

    def test_scenarios_table_has_runway_at_todays_burn_column(self) -> None:
        """The Scenarios markdown table gains a 'Runway at Today's Burn'
        column, populated from static_runway_months for every scenario row —
        not just base."""
        runway = self._default_alive_runway(6.9)
        runway["scenarios"][1]["static_runway_months"] = 12.3
        runway["scenarios"][2]["static_runway_months"] = 4.1
        md = self._run_with_runway(runway)
        assert "Runway at Today's Burn" in md
        assert "6.9 months" in md
        assert "12.3 months" in md
        assert "4.1 months" in md

    def test_scenarios_table_dash_when_static_missing_for_a_row(self) -> None:
        """A scenario row with no static_runway_months renders the em-dash
        placeholder in the new column, not a Python None literal."""
        md = self._run_with_runway(self._default_alive_runway(6.9))
        assert "Runway at Today's Burn" in md
        # "slow" and "crisis" in _VALID_RUNWAY carry no static_runway_months —
        # their row's new column must be the placeholder dash, not "None".
        slow_row = next(line for line in md.splitlines() if line.startswith("| slow "))
        assert "None" not in slow_row
        assert "—" in slow_row


# ---------------------------------------------------------------------------
# Item 7: Corrections Applied section from extraction_corrections.json
# ---------------------------------------------------------------------------


class TestCorrectionsAppliedSection:
    """compose_report must render a 'Corrections Applied' section when
    extraction_corrections.json is present; must be absent without it."""

    _BASE_ARTIFACTS = {
        "inputs.json": _VALID_INPUTS,
        "checklist.json": _VALID_CHECKLIST,
        "unit_economics.json": _VALID_UNIT_ECONOMICS,
        "runway.json": _VALID_RUNWAY,
    }

    _CORRECTIONS = {
        "timestamp": "2026-06-12T10:00:00Z",
        "corrections": [
            {"path": "revenue.arr.value", "was": "1000000", "now": "12000000", "source": "founder"},
            {"path": "cash.current_balance", "was": "500000", "now": "5000000", "source": "founder"},
        ],
    }

    def test_corrections_section_present_when_file_exists(self) -> None:
        """'Corrections Applied' heading appears when extraction_corrections.json provided."""
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": self._CORRECTIONS})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        assert "Corrections Applied" in data["report_markdown"]

    def test_corrections_field_names_in_table(self) -> None:
        """Corrected field paths appear as rows in the corrections table."""
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": self._CORRECTIONS})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        assert "revenue.arr.value" in md
        assert "cash.current_balance" in md

    def test_corrections_was_now_values_in_table(self) -> None:
        """Original and corrected values appear in the table."""
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": self._CORRECTIONS})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        assert "1000000" in md or "12000000" in md  # was/now values visible

    _DISPATCH_CORRECTIONS = {
        "corrections": [
            {"path": "revenue.arr.value", "old": 1000000, "new": 12000000, "reason": "source tab B12"},
        ],
    }

    def test_corrections_old_new_values_in_table(self) -> None:
        """The dispatch-shape old/new keys render actual values, not '?' placeholders."""
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": self._DISPATCH_CORRECTIONS})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        assert "1000000" in md or "12000000" in md
        assert "| ? | ?" not in md

    def test_corrections_reason_rendered(self) -> None:
        """The correction reason is surfaced in the table."""
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": self._DISPATCH_CORRECTIONS})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        assert "source tab B12" in data["report_markdown"]

    def test_corrections_null_null_row_renders_not_found(self) -> None:
        """A field genuinely absent from source (old=null, new=null) renders a
        'not in source' cell, not an ambiguous '?'."""
        corr = {
            "corrections": [
                {"path": "cash.debt", "old": None, "new": None, "reason": "not visible in source"},
            ],
        }
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": corr})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        assert "not in source" in md
        assert "cash.debt |" in md and "| ? |" not in md.split("cash.debt")[1].split("\n")[0]

    def test_corrections_caption_provenance_neutral(self) -> None:
        """The caption must not claim corrections were founder-authored (they may be
        agent-authored during INPUTS_REVIEW)."""
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": self._DISPATCH_CORRECTIONS})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        assert "corrected by the founder" not in data["report_markdown"]

    def test_corrections_replace_array_lengths_rendered(self) -> None:
        """replace_array entries render was_length/now_length, not '?'."""
        corr = {
            "corrections": [
                {"path": "revenue.monthly", "type": "replace_array", "was_length": 3, "now_length": 12},
            ],
        }
        d = _make_fmr_artifact_dir({**self._BASE_ARTIFACTS, "extraction_corrections.json": corr})
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        row = md.split("revenue.monthly")[1].split("\n")[0]
        assert "3" in row and "12" in row and "?" not in row

    def test_corrections_section_absent_without_file(self) -> None:
        """'Corrections Applied' does not appear when file is absent."""
        d = _make_fmr_artifact_dir(self._BASE_ARTIFACTS)
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        assert "Corrections Applied" not in data["report_markdown"]

    def test_compose_succeeds_without_corrections_file(self) -> None:
        """Missing extraction_corrections.json must not fail or warn."""
        d = _make_fmr_artifact_dir(self._BASE_ARTIFACTS)
        rc, data, stderr = _run_compose(d)
        assert rc == 0
        assert data is not None
        # No MISSING_ARTIFACT warning for corrections file
        warnings = data["validation"].get("warnings", [])
        missing = [w for w in warnings if w["code"] == "MISSING_ARTIFACT" and "corrections" in w.get("detail", "")]
        assert not missing


# ---------------------------------------------------------------------------
# Item 8: Explorer footer line in report_markdown
# ---------------------------------------------------------------------------


class TestExplorerFooterLine:
    """compose_report must append a footer line pointing founders to explore.html
    for what-if scenarios."""

    _BASE_ARTIFACTS = {
        "inputs.json": _VALID_INPUTS,
        "checklist.json": _VALID_CHECKLIST,
        "unit_economics.json": _VALID_UNIT_ECONOMICS,
        "runway.json": _VALID_RUNWAY,
    }

    def test_explorer_footer_present_in_report_markdown(self) -> None:
        """report_markdown footer must reference the interactive explorer."""
        d = _make_fmr_artifact_dir(self._BASE_ARTIFACTS)
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        # Footer must mention explore.html or explore.py
        assert "explore" in md.lower()

    def test_explorer_footer_mentions_what_if(self) -> None:
        """Footer should indicate what-if / interactive use."""
        d = _make_fmr_artifact_dir(self._BASE_ARTIFACTS)
        rc, data, stderr = _run_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        assert data is not None
        md = data["report_markdown"]
        assert "what-if" in md.lower() or "scenarios" in md.lower()


# ---------------------------------------------------------------------------
# Change 1: Magic number includes non-headcount marketing opex
# ---------------------------------------------------------------------------
#
# Source: Scale Venture Partners, "Magic Number Math"
#   The denominator is "sales and marketing spend for the earlier of the two
#   quarters" — i.e. ALL S&M, not just headcount.
#
# Non-headcount marketing lives in expenses.opex_monthly[] entries whose
# category (case-insensitive) matches the SM_OPEX_CATEGORIES frozenset.
# opex_monthly.amount is MONTHLY, so the annual base adds amount * 12.
#
# Worked examples:
#   A) Only marketing opex, no sales HC:
#      MRR=100K, growth=5%, marketing_opex=$50K/mo → $600K/yr
#      net_new_ARR = 100K * 0.05 * 12 = 60K
#      monthly_sm  = 600K / 12 = 50K
#      magic       = 60K / 50K = 1.2
#      (old code: no HC → not_rated; new code: should compute)
#
#   B) Sales HC + marketing opex:
#      MRR=100K, growth=5%
#      sales_HC_annual = 600K, marketing_opex = $20K/mo = $240K/yr
#      combined_annual = 600K + 240K = 840K
#      monthly_sm  = 840K / 12 = 70K
#      net_new_ARR = 60K
#      magic       = 60K / 70K ≈ 0.857  (old: 60K/50K = 1.2 — too optimistic)
# ---------------------------------------------------------------------------


class TestMagicNumberWithMarketingOpex:
    """Magic number denominator must include non-headcount marketing opex."""

    _BASE_COMPANY: dict[str, Any] = {
        "company_name": "TestCo",
        "stage": "seed",
        "sector": "B2B SaaS",
        "geography": "US",
        "revenue_model_type": "saas-sales-led",
    }
    _BASE_REVENUE: dict[str, Any] = {
        "arr": {"value": 1_200_000, "as_of": "2025-12"},
        "mrr": {"value": 100_000, "as_of": "2025-12"},
        "growth_rate_monthly": 0.05,
    }
    _BASE_CASH: dict[str, Any] = {
        "current_balance": 2_000_000,
        "monthly_net_burn": 80_000,
    }

    def test_magic_number_opex_only_marketing_computes(self) -> None:
        """Marketing opex alone (no sales HC) should now produce a magic number.

        Derivation (example A):
          net_new_ARR  = 100K * 0.05 * 12 = 60K
          opex_monthly = $50K/mo → annual = $600K → monthly = $50K
          magic        = 60K / 50K = 1.2
        Old behaviour: no headcount → sm_spend_annual = 0 → not_rated
        New behaviour: marketing opex included → magic = 1.2
        """
        inputs: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": self._BASE_REVENUE,
            "cash": self._BASE_CASH,
            "expenses": {
                "opex_monthly": [
                    {"category": "marketing", "amount": 50_000, "start_month": "2025-01"},
                ]
            },
        }
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        magic = metrics.get("magic_number")
        assert magic is not None, "magic_number metric missing"
        assert magic["value"] is not None, (
            "magic_number value is null — marketing opex alone must be sufficient to compute"
        )
        assert magic["rating"] != "not_rated", (
            f"magic_number should not be not_rated when marketing opex is present. Got: {magic}"
        )
        # net_new_ARR=60K, monthly_sm=50K → magic=1.2
        assert abs(magic["value"] - 1.2) < 0.05, (
            f"Expected magic_number ≈ 1.2 (marketing opex $50K/mo), got {magic['value']:.4f}"
        )

    def test_magic_number_combined_base_lower_than_headcount_only(self) -> None:
        """Adding marketing opex to existing sales HC lowers the magic number.

        Derivation (example B):
          net_new_ARR       = 100K * 0.05 * 12 = 60K
          sales_HC_annual   = $600K  → monthly = $50K
          marketing_opex    = $20K/mo → annual = $240K
          combined_annual   = $840K  → combined_monthly = $70K
          magic_combined    = 60K / 70K ≈ 0.857
          magic_HC_only     = 60K / 50K = 1.2
        Adding opex must lower the magic number (more conservative, matches definition).
        """
        inputs_hc_only: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": self._BASE_REVENUE,
            "cash": self._BASE_CASH,
            "expenses": {
                "headcount": [
                    {
                        "role": "sales",
                        "count": 1,
                        "salary_annual": 600_000,
                        "burden_pct": 0.0,
                    }
                ]
            },
        }
        inputs_combined: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": self._BASE_REVENUE,
            "cash": self._BASE_CASH,
            "expenses": {
                "headcount": [
                    {
                        "role": "sales",
                        "count": 1,
                        "salary_annual": 600_000,
                        "burden_pct": 0.0,
                    }
                ],
                "opex_monthly": [
                    {"category": "marketing", "amount": 20_000, "start_month": "2025-01"},
                ],
            },
        }
        rc1, data1, _ = run_script("unit_economics.py", stdin_data=json.dumps(inputs_hc_only))
        rc2, data2, _ = run_script("unit_economics.py", stdin_data=json.dumps(inputs_combined))
        assert rc1 == 0 and rc2 == 0

        magic_hc = {m["name"]: m for m in data1.get("metrics", [])}["magic_number"]
        magic_combined = {m["name"]: m for m in data2.get("metrics", [])}["magic_number"]

        assert magic_hc["value"] is not None and magic_combined["value"] is not None
        assert magic_combined["value"] < magic_hc["value"], (
            f"Combined base must lower magic number: "
            f"hc-only={magic_hc['value']:.4f}, combined={magic_combined['value']:.4f}"
        )
        # Spot-check combined value ≈ 0.857 (60K / 70K)
        assert abs(magic_combined["value"] - round(60_000 / 70_000, 2)) < 0.05, (
            f"Expected magic_number ≈ {round(60_000 / 70_000, 2):.4f} for combined base, "
            f"got {magic_combined['value']:.4f}"
        )

    def test_magic_number_evidence_names_headcount_and_opex(self) -> None:
        """Evidence string must honestly name the combined base when both are present."""
        inputs: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": self._BASE_REVENUE,
            "cash": self._BASE_CASH,
            "expenses": {
                "headcount": [{"role": "marketing", "count": 1, "salary_annual": 120_000, "burden_pct": 0.0}],
                "opex_monthly": [
                    {"category": "ads", "amount": 10_000, "start_month": "2025-01"},
                ],
            },
        }
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
        assert rc == 0, f"failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        magic = metrics.get("magic_number")
        assert magic is not None and magic["value"] is not None
        evidence = magic.get("evidence", "")
        # Evidence must mention the opex component in the base description
        assert "opex" in evidence.lower() or "marketing" in evidence.lower(), (
            f"Evidence should name the opex base component. Got: {evidence!r}"
        )

    def test_magic_number_category_case_insensitive(self) -> None:
        """Category matching must be case-insensitive ('Advertising', 'DEMAND GEN', etc.)."""
        for cat in ("Advertising", "DEMAND GEN", "Demand Generation", "S&M", "Sales & Marketing"):
            inputs: dict[str, Any] = {
                "company": self._BASE_COMPANY,
                "revenue": self._BASE_REVENUE,
                "cash": self._BASE_CASH,
                "expenses": {
                    "opex_monthly": [
                        {"category": cat, "amount": 50_000, "start_month": "2025-01"},
                    ]
                },
            }
            rc, data, _ = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
            assert rc == 0
            metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
            magic = metrics["magic_number"]
            assert magic["value"] is not None, f"Category '{cat}' not matched — magic_number should compute"

    def test_magic_number_non_sm_opex_excluded(self) -> None:
        """Cloud/engineering opex must NOT be counted in the S&M base."""
        inputs_no_sm: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": self._BASE_REVENUE,
            "cash": self._BASE_CASH,
            "expenses": {
                "opex_monthly": [
                    {"category": "cloud", "amount": 50_000, "start_month": "2025-01"},
                    {"category": "engineering tools", "amount": 20_000, "start_month": "2025-01"},
                ]
            },
        }
        rc, data, _ = run_script("unit_economics.py", stdin_data=json.dumps(inputs_no_sm))
        assert rc == 0
        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        magic = metrics["magic_number"]
        # No S&M in opex → should still be not_rated (cloud/tools are not S&M)
        assert magic["value"] is None, f"Non-S&M opex categories must not trigger magic_number. Got: {magic}"


# ---------------------------------------------------------------------------
# Change 2: Rule of 40 uses realized YoY growth when revenue history exists
# ---------------------------------------------------------------------------
#
# Source: Brad Feld (canonical R40): growth = "year-over-year growth rate"
# of revenue.  When ≥12 monthly entries (or ≥5 quarterly) exist, use the
# realized YoY rate; otherwise fall back to annualized-MoM with disclosure.
#
# Realized YoY derivation using monthly time-series:
#   latest_arr    = ARR at final month entry
#   year_ago_arr  = ARR 12 months earlier (index -13 if 13 entries, else index 0)
#   yoy_growth_%  = (latest_arr - year_ago_arr) / year_ago_arr * 100
#
# Worked example — 13 monthly entries, ARR grows from 400K to 1M:
#   yoy_growth = (1000K - 400K) / 400K * 100 = 150.0%
#   op_margin  = -monthly_burn / mrr = -30K / 83.3K ≈ -36% → let's use gross for simplicity
#   gross_margin = 0.75 → R40 = 150 + 75 = 225
#   (annualized-MoM from growth_rate_monthly=0.08: (1.08^12-1)*100 ≈ 151.8% — very close
#    but would differ for non-constant-growth companies)
#
# Worked example (fallback, no monthly history):
#   growth_rate_monthly=0.08 → annualized = (1.08^12-1)*100 ≈ 151.8%
#   evidence must include "annualized from current MoM rate"
# ---------------------------------------------------------------------------


class TestRuleOf40RealizedYoY:
    """Rule of 40 uses realized YoY when ≥12 months of history exist."""

    _BASE_COMPANY: dict[str, Any] = {
        "company_name": "TestCo",
        "stage": "series-a",
        "sector": "B2B SaaS",
        "geography": "US",
        "revenue_model_type": "saas-sales-led",
    }

    def _make_inputs_with_monthly(
        self,
        n_months: int,
        arr_start: float,
        arr_end: float,
        mrr_current: float,
        burn: float,
    ) -> dict[str, Any]:
        """Build inputs with n_months of monthly ARR history, linear progression."""
        step = (arr_end - arr_start) / max(n_months - 1, 1)
        monthly = [
            {
                "month": f"2025-{i + 1:02d}" if i < 12 else f"2026-{i - 11:02d}",
                "arr": round(arr_start + step * i),
            }
            for i in range(n_months)
        ]
        return {
            "company": self._BASE_COMPANY,
            "revenue": {
                "arr": {"value": arr_end, "as_of": "2025-12"},
                "mrr": {"value": mrr_current, "as_of": "2025-12"},
                "growth_rate_monthly": 0.08,  # also present — should NOT be used for growth
                "monthly": monthly,
            },
            "cash": {"current_balance": 5_000_000, "monthly_net_burn": burn},
            "unit_economics": {"gross_margin": 0.75},
        }

    def test_r40_realized_yoy_with_12_monthly_entries(self) -> None:
        """With 12 monthly entries, R40 uses realized YoY growth (not annualized MoM).

        Derivation (12 entries, arr 400K→1M, index-lookback is index 0 = 400K):
          year_ago_arr = 400K (entry[0])
          latest_arr   = 1M  (entry[-1])
          yoy_growth   = (1M - 400K) / 400K * 100 = 150.0%
          gross_margin = 0.75 → margin_pct = 75
          r40          = 150.0 + 75 = 225.0
          (annualized-MoM at 8% would give ≈151.8 — deliberately close but not identical
           for clean testing; see fallback test for the MoM path)
        The evidence must say "realized YoY" (not "annualized from current MoM rate").
        """
        inputs = self._make_inputs_with_monthly(
            n_months=12, arr_start=400_000, arr_end=1_000_000, mrr_current=83_333, burn=100_000
        )
        # Force gross-margin path: remove monthly_net_burn  → op_margin unavailable
        del inputs["cash"]["monthly_net_burn"]

        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        r40 = metrics.get("rule_of_40")
        assert r40 is not None, "rule_of_40 missing"
        assert r40["value"] is not None, f"rule_of_40 not computed: {r40}"

        evidence = r40.get("evidence", "")
        assert "realized yoy" in evidence.lower(), (
            f"Evidence must say 'realized YoY' when monthly history is present. Got: {evidence!r}"
        )
        assert "annualized from current mom" not in evidence.lower(), (
            f"Evidence must NOT say annualized-from-MoM when history is present. Got: {evidence!r}"
        )

    def test_r40_no_history_falls_back_to_annualized_mom(self) -> None:
        """Without monthly history, R40 falls back to annualized-MoM with disclosure.

        This test ensures the fallback path still carries the 'annualized from current MoM rate'
        disclosure required by TestRuleOf40GrowthDisclosure.
        """
        inputs: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": {
                "arr": {"value": 6_000_000, "as_of": "2025-12"},
                "mrr": {"value": 500_000, "as_of": "2025-12"},
                "growth_rate_monthly": 0.08,
                # No "monthly" or "quarterly" entries
            },
            "cash": {"current_balance": 5_000_000, "monthly_net_burn": 300_000},
            "unit_economics": {"gross_margin": 0.75},
        }
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        r40 = metrics.get("rule_of_40")
        assert r40 is not None and r40["value"] is not None, f"r40 not computed: {r40}"

        evidence = r40.get("evidence", "")
        assert "annualized" in evidence.lower(), (
            f"Fallback evidence must say 'annualized' (from MoM rate). Got: {evidence!r}"
        )

    def test_r40_realized_yoy_with_5_quarterly_entries(self) -> None:
        """With 5 quarterly entries (full YoY window), R40 uses realized YoY.

        Derivation:
          entry 0: arr=400K (year_ago)
          entry 4: arr=1M  (latest)
          yoy_growth = (1M - 400K) / 400K * 100 = 150.0%
          gross_margin = 0.75 → r40 = 225.0
        Evidence must say 'realized YoY'.
        """
        inputs: dict[str, Any] = {
            "company": self._BASE_COMPANY,
            "revenue": {
                "arr": {"value": 1_000_000, "as_of": "2025-Q4"},
                "mrr": {"value": 83_333, "as_of": "2025-Q4"},
                "growth_rate_monthly": 0.08,
                "quarterly": [
                    {"quarter": "2024-Q4", "arr": 400_000},
                    {"quarter": "2025-Q1", "arr": 600_000},
                    {"quarter": "2025-Q2", "arr": 700_000},
                    {"quarter": "2025-Q3", "arr": 850_000},
                    {"quarter": "2025-Q4", "arr": 1_000_000},
                ],
            },
            "cash": {"current_balance": 5_000_000},  # no monthly_net_burn → gross path
            "unit_economics": {"gross_margin": 0.75},
        }
        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        r40 = metrics.get("rule_of_40")
        assert r40 is not None and r40["value"] is not None, f"r40 not computed: {r40}"

        evidence = r40.get("evidence", "")
        assert "realized yoy" in evidence.lower(), (
            f"Evidence must say 'realized YoY' for 5-quarter window. Got: {evidence!r}"
        )

    def test_r40_fewer_than_12_monthly_falls_back_to_mom(self) -> None:
        """With only 11 monthly entries (< 12), fall back to annualized-MoM."""
        inputs = self._make_inputs_with_monthly(
            n_months=11, arr_start=400_000, arr_end=1_000_000, mrr_current=83_333, burn=100_000
        )
        del inputs["cash"]["monthly_net_burn"]

        rc, data, stderr = run_script("unit_economics.py", stdin_data=json.dumps(inputs))
        assert rc == 0, f"unit_economics.py failed: {stderr}"

        metrics = {m["name"]: m for m in data.get("metrics", []) if isinstance(m, dict)}
        r40 = metrics.get("rule_of_40")
        assert r40 is not None and r40["value"] is not None

        evidence = r40.get("evidence", "")
        assert "annualized" in evidence.lower(), (
            f"With < 12 monthly entries, evidence must say 'annualized'. Got: {evidence!r}"
        )


# --- Sector-aware gross margin benchmarks ---


def _gm_payload(
    model_type: str,
    gm: float,
    stage: str = "seed",
    traits: list[str] | None = None,
    basis: str | None = None,
    ai_cogs: bool = False,
) -> str:
    """Build a unit_economics payload with a specific revenue model type and gross margin."""
    inputs = json.loads(json.dumps(_VALID_INPUTS))
    inputs["company"]["stage"] = stage
    inputs["company"]["revenue_model_type"] = model_type
    if traits is not None:
        inputs["company"]["traits"] = traits
    inputs["unit_economics"]["gross_margin"] = gm
    if basis is not None:
        inputs["unit_economics"]["gross_margin_basis"] = basis
    if ai_cogs:
        inputs["expenses"]["cogs"]["inference_costs"] = 4000
    return json.dumps(inputs)


def _gm_metric(payload: str) -> dict[str, Any]:
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0, f"unit_economics.py failed: {stderr}"
    assert data is not None
    metric: dict[str, Any] = {m["name"]: m for m in data["metrics"]}["gross_margin"]
    return metric


def test_unit_economics_gm_retail_sector_benchmark() -> None:
    """A retail company's mid-50s gross margin rates strong against the retail table, not SaaS."""
    gm = _gm_metric(_gm_payload("retail", 0.55))
    assert gm["rating"] == "strong", f"0.55 GM should be strong for retail, got {gm['rating']!r}"
    assert "Damodaran" in gm["benchmark_source"], f"retail bar must cite its source: {gm['benchmark_source']!r}"
    assert gm["benchmark_as_of"] == "2026-01"
    assert "sector benchmark" in gm["evidence"], f"evidence should name the sector bar: {gm['evidence']!r}"
    # No basis provided: the product-GM assumption must be disclosed, not silent.
    assert "product" in gm["evidence"].lower(), f"evidence must disclose the basis assumption: {gm['evidence']!r}"


def test_unit_economics_gm_hardware_sector_benchmark() -> None:
    """A hardware company's 45% gross margin rates acceptable against the hardware table."""
    gm = _gm_metric(_gm_payload("hardware", 0.45))
    assert gm["rating"] == "acceptable", f"0.45 GM should be acceptable for hardware, got {gm['rating']!r}"
    assert "Damodaran" in gm["benchmark_source"]


def test_unit_economics_gm_hardware_subscription_contextual() -> None:
    """hardware-subscription is contextual: the device-only 50% rule excludes service blends,
    and a single GM number cannot be decomposed into hardware vs service margin."""
    gm = _gm_metric(_gm_payload("hardware-subscription", 0.45))
    assert gm["rating"] == "contextual", f"got {gm['rating']!r}"
    assert "split" in gm["evidence"].lower(), f"evidence should ask for the margin split: {gm['evidence']!r}"
    assert gm["benchmark_source"] != ""


def test_unit_economics_gm_consumer_subscription_benchmark() -> None:
    """A consumer-subscription 50% gross margin rates acceptable against its own table."""
    gm = _gm_metric(_gm_payload("consumer-subscription", 0.50))
    assert gm["rating"] == "acceptable", f"0.50 GM should be acceptable for consumer sub, got {gm['rating']!r}"
    assert "Damodaran" in gm["benchmark_source"]


def test_unit_economics_gm_marketplace_contextual() -> None:
    """Marketplace gross margin is contextual: basis-dependent, never pass/fail."""
    gm = _gm_metric(_gm_payload("marketplace", 0.60))
    assert gm["rating"] == "contextual", f"got {gm['rating']!r}"
    assert "basis" in gm["evidence"].lower(), f"evidence must explain the revenue-basis caveat: {gm['evidence']!r}"
    assert gm["benchmark_source"] != ""


def test_unit_economics_gm_transactional_fintech_contextual() -> None:
    """Transactional fintech gross margin is contextual (take-rate/net-revenue basis)."""
    gm = _gm_metric(_gm_payload("transactional-fintech", 0.60))
    assert gm["rating"] == "contextual", f"got {gm['rating']!r}"
    assert "basis" in gm["evidence"].lower()


def test_unit_economics_gm_ai_discount_composes_with_sector_table() -> None:
    """With material AI COGS present, the AI adjustment applies to the selected sector table."""
    gm = _gm_metric(_gm_payload("retail", 0.46, traits=["ai-powered"], ai_cogs=True))
    # retail strong bar 0.50 - 0.05 seed AI adjustment = 0.45
    assert gm["rating"] == "strong", f"0.46 GM should clear the AI-adjusted retail bar, got {gm['rating']!r}"
    assert "AI-adjusted" in gm["evidence"]


def test_unit_economics_gm_ai_trait_alone_no_discount() -> None:
    """The ai-powered trait alone earns no margin concession without AI costs in COGS."""
    gm = _gm_metric(_gm_payload("retail", 0.46, traits=["ai-powered"]))
    assert gm["rating"] == "acceptable", f"trait without AI COGS must not discount the bar: {gm['rating']!r}"
    assert "AI-adjusted" not in gm["evidence"]


def test_unit_economics_gm_sector_tables_stage_invariant() -> None:
    """Non-SaaS gross margin tables do not vary by stage (sources are not stage-segmented)."""
    for stage in ("pre-seed", "seed", "series-a"):
        gm = _gm_metric(_gm_payload("hardware", 0.45, stage=stage))
        assert gm["rating"] == "acceptable", f"stage {stage}: got {gm['rating']!r}"


def test_unit_economics_gm_saas_table_unchanged() -> None:
    """SaaS models keep the existing stage-keyed KeyBanc benchmark."""
    gm = _gm_metric(_gm_payload("saas-sales-led", 0.72))
    assert gm["rating"] == "acceptable", f"0.72 GM at seed should stay acceptable for SaaS, got {gm['rating']!r}"
    assert "KeyBanc" in gm["benchmark_source"]


def test_unit_economics_gm_unknown_model_type_falls_back_to_saas() -> None:
    """An empty/unknown revenue model type keeps the SaaS benchmark, with the assumption disclosed."""
    gm = _gm_metric(_gm_payload("", 0.72))
    assert gm["rating"] == "acceptable"
    assert "KeyBanc" in gm["benchmark_source"]
    assert "assumed" in gm["evidence"].lower(), f"fallback must disclose the SaaS assumption: {gm['evidence']!r}"


def test_unit_economics_gm_non_product_basis_contextual() -> None:
    """A non-product gross_margin_basis makes the metric contextual — the threshold tables all
    assume product/service gross margin, and store contribution is a different metric."""
    gm = _gm_metric(_gm_payload("retail", 0.20, basis="store_contribution"))
    assert gm["rating"] == "contextual", f"store-contribution margin must not rate fail: {gm['rating']!r}"
    assert "store" in gm["evidence"].lower() or "basis" in gm["evidence"].lower()


def test_unit_economics_gm_product_basis_rated() -> None:
    """An explicit product basis rates normally against the sector table."""
    gm = _gm_metric(_gm_payload("retail", 0.55, basis="product"))
    assert gm["rating"] == "strong", f"got {gm['rating']!r}"


def test_unit_economics_gm_threshold_boundaries() -> None:
    """Exact-threshold values rate at the tier (>= semantics) for every sector table."""
    cases = [
        ("retail", 0.50, "strong"),
        ("retail", 0.35, "acceptable"),
        ("retail", 0.25, "warning"),
        ("retail", 0.249, "fail"),
        ("hardware", 0.50, "strong"),
        ("hardware", 0.40, "acceptable"),
        ("hardware", 0.25, "warning"),
        ("consumer-subscription", 0.65, "strong"),
        ("consumer-subscription", 0.45, "acceptable"),
        ("consumer-subscription", 0.30, "warning"),
    ]
    for model_type, gm_value, expected in cases:
        gm = _gm_metric(_gm_payload(model_type, gm_value))
        assert gm["rating"] == expected, f"{model_type} @ {gm_value}: expected {expected}, got {gm['rating']!r}"


def test_validate_inputs_gross_margin_basis_enum() -> None:
    """validate_inputs accepts the documented gross_margin_basis values and rejects others."""
    ok = json.dumps(
        {"company": _CO, "unit_economics": {"gross_margin": 0.5, "gross_margin_basis": "store_contribution"}}
    )
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=ok)
    assert rc == 0
    assert "unit_economics.gross_margin_basis" not in {e["field"] for e in data.get("errors", [])}

    bad = json.dumps({"company": _CO, "unit_economics": {"gross_margin": 0.5, "gross_margin_basis": "ebitda"}})
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=bad)
    assert rc == 0
    assert "unit_economics.gross_margin_basis" in {e["field"] for e in data.get("errors", [])}


def test_unit_economics_gm_usage_based_contextual() -> None:
    """usage-based is contextual: healthy consumption models span passthrough-heavy CPaaS
    (~51%) to software-margin platforms — a single bar mis-rates one end."""
    gm = _gm_metric(_gm_payload("usage-based", 0.51))
    assert gm["rating"] == "contextual", f"a ~51% CPaaS-style GM must not be pass/fail: {gm['rating']!r}"
    assert "passthrough" in gm["evidence"].lower(), f"evidence should explain the spread: {gm['evidence']!r}"
    assert gm["benchmark_source"] != ""


def test_unit_economics_cac_contextual_for_retail() -> None:
    """Retail CAC is contextual (store-driven vs paid acquisition varies too widely)."""
    payload = _gm_payload("retail", 0.55)
    rc, data, stderr = run_script("unit_economics.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    cac = {m["name"]: m for m in data["metrics"]}["cac"]
    assert cac["rating"] == "contextual", f"got {cac['rating']!r}"


def test_validate_inputs_accepts_retail_revenue_model() -> None:
    """The revenue_model_type enum accepts 'retail'."""
    payload = json.dumps({"company": {**_CO, "revenue_model_type": "retail"}})
    rc, data, stderr = run_script("validate_inputs.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    error_fields = {e["field"] for e in data.get("errors", [])}
    assert "company.revenue_model_type" not in error_fields, f"retail must be a valid enum value: {data.get('errors')}"


def test_checklist_retail_derives_sector_type() -> None:
    """retail revenue_model_type derives sector_type without warning; no sector item fires."""
    company = {
        "stage": "seed",
        "geography": "us",
        "sector": "Retail",
        "revenue_model_type": "retail",
        "traits": [],
        # no sector_type — should derive "retail"
    }
    items = _make_checklist_items()
    payload = json.dumps({"items": items, "company": company})
    rc, data, stderr = run_script("checklist.py", ["--pretty"], stdin_data=payload)
    assert rc == 0
    assert data is not None
    assert "could not derive" not in stderr, f"retail should derive a sector_type: {stderr!r}"
    for item_id in ("SECTOR_39", "SECTOR_40", "SECTOR_41", "SECTOR_42", "SECTOR_43", "SECTOR_44"):
        item = next(i for i in data["items"] if i["id"] == item_id)
        assert item["status"] == "not_applicable", f"{item_id} should be gated for retail sector_type"


# --- extract_model.py used-range hardening ---


def _write_degenerate_xlsx(path: str) -> None:
    """Small real table + one styled-but-empty cell far away, ballooning the
    declared used-range (~3M cells) the way stray formatting does."""
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Revenue"
    ws.append(["Metric", "2025-01", "2025-02"])
    ws.append(["MRR", 50000, 55000])
    ws.append(["Customers", 100, 110])
    ws.cell(row=100000, column=30).fill = PatternFill(start_color="FFFF0000", fill_type="solid")
    wb.save(path)


def _write_unknown_extension_xlsx(path: str) -> None:
    """Normal small table whose sheet XML carries an unknown extLst extension,
    which makes openpyxl emit a UserWarning at load time."""
    import zipfile

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Metric", "2025-01"])
    ws.append(["MRR", 50000])
    wb.save(path)

    with zipfile.ZipFile(path) as zin:
        data = {n: zin.read(n) for n in zin.namelist()}
    xml = data["xl/worksheets/sheet1.xml"].decode()
    inj = '<extLst><ext uri="{DEADBEEF-0000-0000-0000-000000000000}"/></extLst></worksheet>'
    data["xl/worksheets/sheet1.xml"] = xml.replace("</worksheet>", inj).encode()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, b in data.items():
            zout.writestr(n, b)


def test_extract_model_degenerate_used_range_bounded() -> None:
    """A formatting-bloated used-range collapses to the populated region with a
    warning — not a multi-megabyte blob of null cells."""
    import pytest

    pytest.importorskip("openpyxl")
    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "patho.xlsx")
        _write_degenerate_xlsx(path)
        rc, raw, stderr = run_script_raw("extract_model.py", ["--file", path])
        assert rc == 0, f"extraction failed: {stderr}"
        assert len(raw) < 1_000_000, f"degenerate range must not balloon the output: {len(raw)} bytes"
        data = json.loads(raw)
        sheet = data["sheets"][0]
        assert sheet["row_count"] <= 10, f"rows must trim to the populated region: {sheet['row_count']}"
        assert sheet["col_count"] <= 5, f"cols must trim to the populated region: {sheet['col_count']}"
        labels = [str(r[0]) for r in sheet["rows"] if r]
        assert "MRR" in labels, "real data must survive the trim"
        warnings_list = data.get("extraction_warnings", [])
        assert warnings_list, "a truncated extraction must carry a warning"
        assert any("Revenue" in w for w in warnings_list), f"warning must name the sheet: {warnings_list}"


def test_extract_model_degenerate_receipt_carries_warning() -> None:
    """With -o, the truncation warning is surfaced in the write receipt."""
    import pytest

    pytest.importorskip("openpyxl")
    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "patho.xlsx")
        out_path = os.path.join(d, "model_data.json")
        _write_degenerate_xlsx(path)
        rc, receipt, stderr = run_script("extract_model.py", ["--file", path, "-o", out_path])
        assert rc == 0
        assert receipt.get("ok") is True
        assert receipt.get("extraction_warnings"), f"receipt must surface the warning: {receipt}"


def test_extract_model_openpyxl_warning_captured_not_leaked() -> None:
    """openpyxl load warnings are captured into extraction_warnings, not stderr."""
    import pytest

    pytest.importorskip("openpyxl")
    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "warned.xlsx")
        _write_unknown_extension_xlsx(path)
        rc, data, stderr = run_script("extract_model.py", ["--file", path, "--pretty"])
        assert rc == 0
        assert "Unknown extension" not in stderr, "openpyxl warning must not leak to stderr"
        warnings_list = data.get("extraction_warnings", [])
        assert any("Unknown extension" in w for w in warnings_list), (
            f"openpyxl warning must be captured in the output: {warnings_list}"
        )
        labels = [str(r[0]) for r in data["sheets"][0]["rows"] if r]
        assert "MRR" in labels, "data must still extract despite the warning"


def test_extract_model_normal_xlsx_no_extraction_warnings() -> None:
    """A normal model extracts exactly as before: no extraction_warnings key."""
    import pytest

    pytest.importorskip("openpyxl")
    import openpyxl

    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "normal.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Revenue"
        ws.append(["Metric", "2025-01", "2025-02"])
        ws.append(["MRR", 50000, 55000])
        wb.save(path)
        rc, data, stderr = run_script("extract_model.py", ["--file", path, "--pretty"])
        assert rc == 0
        assert "extraction_warnings" not in data, "normal models must not grow new keys"
        assert data["sheets"][0]["row_count"] == 2
        assert data["sheets"][0]["col_count"] == 3


def test_model_data_output_is_line_navigable() -> None:
    """Regression lock: run the extractor exactly as SKILL.md Step 2 does
    (--pretty, -o) and confirm model_data.json has many more lines than the
    sheet has rows — so downstream Grep/paged-Read can target regions. Guards
    against an edit that drops --pretty and re-ships a single multi-MB line."""
    import pytest

    pytest.importorskip("openpyxl")
    from openpyxl import Workbook  # type: ignore[import-untyped]

    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "m.xlsx")
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Metric", "M1", "M2"])
        for i in range(20):
            ws.append([f"Row {i}", i, i + 1])
        wb.save(path)
        wb.close()
        out = os.path.join(d, "model_data.json")
        rc, receipt, stderr = run_script("extract_model.py", ["--file", path, "--pretty", "-o", out])
        assert rc == 0, stderr
        with open(out, encoding="utf-8") as fh:
            text = fh.read()
        line_count = text.count("\n")
        sheet = json.loads(text)["sheets"][0]
        assert line_count > sheet["row_count"], (
            f"pretty output must stay line-navigable (lines {line_count} > rows {sheet['row_count']})"
        )


def _write_midsize_sparse_xlsx(path: str) -> None:
    """A tight table plus one styled-but-empty cell ~5,200 rows down, inflating
    the declared used-range to ~26k cells — below the old 100k sparse-guard
    floor AND the 2M degenerate threshold, the sub-degenerate window the old
    guard skipped entirely. Plus a dense control sheet with no trailing nulls."""
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    sparse = wb.active
    assert sparse is not None
    sparse.title = "Sparse"
    sparse.append(["Metric", "M1", "M2", "M3", "M4"])
    for i in range(29):
        sparse.append([f"Item {i}", i, i + 1, i + 2, i + 3])
    sparse.cell(row=5200, column=5).fill = PatternFill(start_color="FFFF0000", fill_type="solid")

    dense = wb.create_sheet("Dense")
    dense.append(["Metric", "M1", "M2", "M3", "M4"])
    for i in range(29):
        dense.append([f"Row {i}", i, i + 1, i + 2, i + 3])
    wb.save(path)


def test_extract_trims_trailing_null_region_on_midsize_sparse_sheet() -> None:
    """The trailing-null bounding-box trim is unconditional on the sub-degenerate
    path: a ~26k-cell stray-formatting balloon (below the old 100k sparse floor)
    must collapse to its populated region, losslessly, while a dense control
    sheet with no trailing nulls stays unchanged (byte-identical no-op)."""
    import pytest

    pytest.importorskip("openpyxl")
    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "midsize.xlsx")
        _write_midsize_sparse_xlsx(path)
        rc, raw, stderr = run_script_raw("extract_model.py", ["--file", path])
        assert rc == 0, f"extraction failed: {stderr}"
        data = json.loads(raw)
        sheets = {s["name"]: s for s in data["sheets"]}

        sparse = sheets["Sparse"]
        assert sparse["row_count"] <= 35, (
            f"sub-degenerate stray-formatting balloon must trim trailing nulls: {sparse['row_count']}"
        )
        assert sparse["col_count"] == 5, f"cols must trim to the populated region: {sparse['col_count']}"
        assert "Item 0" in [str(r[0]) for r in sparse["rows"] if r], "real data must survive the trim"
        warnings_list = data.get("extraction_warnings", [])
        assert any("Sparse" in w for w in warnings_list), (
            f">25%-trimmed sheet must carry a warning naming it: {warnings_list}"
        )

        dense = sheets["Dense"]
        assert dense["row_count"] == 29 and dense["col_count"] == 5, (
            f"dense control must be unchanged: {dense['row_count']}x{dense['col_count']}"
        )
        assert not any("Dense" in w for w in warnings_list), "a no-op trim on a dense sheet must not warn"
        assert dense["rows"][0][0] == "Row 0", "dense data must be intact (lossless no-op)"


# --- extract_model.py used-range hardening: ratio trigger, caps, filters ---


def _load_extract_model_module() -> Any:
    import importlib.util

    path = os.path.join(FMR_SCRIPTS_DIR, "extract_model.py")
    spec = importlib.util.spec_from_file_location("fmr_extract_model_module", path)
    assert spec is not None and spec.loader is not None
    mod = importlib.util.module_from_spec(spec)
    sys.modules["fmr_extract_model_module"] = mod
    spec.loader.exec_module(mod)
    return mod


class _FakeCell:
    def __init__(self, value: Any, coordinate: str) -> None:
        self.value = value
        self.coordinate = coordinate


class _FakeWS:
    """Duck-typed worksheet: declares a huge range, yields dense rows."""

    def __init__(self, title: str, declared_rows: int, declared_cols: int, real_rows: int, real_cols: int) -> None:
        self.title = title
        self.max_row = declared_rows
        self.max_column = declared_cols
        self._real_rows = real_rows
        self._real_cols = real_cols

    def iter_rows(self, min_col: int = 1, max_col: int | None = None, values_only: bool = False) -> Any:
        width = min(max_col or self.max_column, self.max_column)
        for i in range(min(self._real_rows, self.max_row)):
            yield tuple(_FakeCell("x" if j < self._real_cols else None, f"R{i + 1}C{j + 1}") for j in range(width))


def test_extract_model_sub_threshold_sparse_bloat_trimmed() -> None:
    """A sparse formatting-bloated sheet BELOW the absolute degenerate threshold
    still collapses to its populated region (populated-ratio trigger)."""
    import pytest

    pytest.importorskip("openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill

    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "subthreshold.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Revenue"
        ws.append(["Metric", "2025-01", "2025-02"])
        ws.append(["MRR", 50000, 55000])
        # ~1.04M declared cells: under the absolute threshold, sparse in the extreme
        ws.cell(row=40000, column=26).fill = PatternFill(start_color="FF00FF00", fill_type="solid")
        wb.save(path)
        rc, raw, stderr = run_script_raw("extract_model.py", ["--file", path])
        assert rc == 0, f"extraction failed: {stderr}"
        assert len(raw) < 1_000_000, f"sparse bloat below the threshold still ballooned: {len(raw)} bytes"
        data = json.loads(raw)
        sheet = data["sheets"][0]
        assert sheet["row_count"] <= 10, f"rows must trim to the populated region: {sheet['row_count']}"
        labels = [str(r[0]) for r in sheet["rows"] if r]
        assert "MRR" in labels
        assert any("Revenue" in w for w in data.get("extraction_warnings", []))


def test_extract_model_degenerate_column_cap_disclosed() -> None:
    """When the declared range is wider than the column cap, the warning says
    the far columns were never scanned — no silent column drop."""
    import pytest

    pytest.importorskip("openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill

    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "widepatho.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Revenue"
        ws.append(["Metric", "2025-01"])
        ws.append(["MRR", 50000])
        ws.cell(row=100000, column=2000).fill = PatternFill(start_color="FFFF0000", fill_type="solid")
        wb.save(path)
        rc, data, stderr = run_script("extract_model.py", ["--file", path, "--pretty"])
        assert rc == 0
        warnings_list = data.get("extraction_warnings", [])
        assert any("1,024" in w and "not scanned" in w for w in warnings_list), (
            f"column cap must be disclosed: {warnings_list}"
        )


def test_extract_model_read_sheet_rows_respects_cell_budget() -> None:
    """A dense degenerate sheet is truncated to the cell budget with a warning,
    and a smaller remaining workbook budget truncates further."""
    import pytest

    pytest.importorskip("openpyxl")
    mod = _load_extract_model_module()

    ws = _FakeWS("Dense", declared_rows=1_000_000, declared_cols=2_000, real_rows=2_000, real_cols=600)
    rows, coords, notes, kept = mod._read_sheet_rows(ws, cell_budget=1_000_000)
    assert kept == len(rows) * len(rows[0])
    assert kept <= 1_000_000, f"kept cells must respect the budget: {kept}"
    assert notes and any("dropped" in n for n in notes)

    ws2 = _FakeWS("Dense2", declared_rows=1_000_000, declared_cols=2_000, real_rows=2_000, real_cols=600)
    rows2, _coords2, notes2, kept2 = mod._read_sheet_rows(ws2, cell_budget=1_200)
    assert kept2 <= 1_200, f"a depleted workbook budget must bind: {kept2}"
    assert notes2


def test_extract_model_foreign_warnings_not_captured() -> None:
    """Only openpyxl-origin warnings land in extraction_warnings; anything else
    is re-emitted to stderr untouched."""
    import warnings as warnings_mod

    import pytest

    pytest.importorskip("openpyxl")
    import openpyxl

    mod = _load_extract_model_module()
    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "normal.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Metric", "2025-01"])
        ws.append(["MRR", 50000])
        wb.save(path)

        original = mod._read_sheet_rows

        def _warning_reader(ws: Any, cell_budget: int) -> Any:
            warnings_mod.warn("synthetic non-openpyxl deprecation", DeprecationWarning, stacklevel=2)
            return original(ws, cell_budget)

        mod._read_sheet_rows = _warning_reader
        try:
            data = mod.extract_xlsx(path)
        finally:
            mod._read_sheet_rows = original
        captured = data.get("extraction_warnings", [])
        assert not any("synthetic non-openpyxl" in w for w in captured), (
            f"foreign warnings must not enter extraction output: {captured}"
        )


def test_extract_model_receipt_warning_list_capped() -> None:
    """A many-sheet pathological workbook does not flood the -o receipt: at most
    10 warnings plus a summary line (the JSON payload keeps the full list)."""
    import pytest

    pytest.importorskip("openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill

    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "manysheets.xlsx")
        out_path = os.path.join(d, "model_data.json")
        wb = openpyxl.Workbook()
        first = wb.active
        assert first is not None
        wb.remove(first)
        for i in range(12):
            ws = wb.create_sheet(f"Tab{i}")
            ws.append(["Metric", "2025-01"])
            ws.append(["MRR", 50000])
            ws.cell(row=100000, column=30).fill = PatternFill(start_color="FFFF0000", fill_type="solid")
        wb.save(path)
        rc, receipt, stderr = run_script("extract_model.py", ["--file", path, "-o", out_path])
        assert rc == 0
        rw = receipt.get("extraction_warnings", [])
        assert len(rw) == 11, f"receipt must cap at 10 warnings + summary line: {len(rw)}"
        assert "more" in rw[-1], f"summary line must count the rest: {rw[-1]!r}"
        with open(out_path, encoding="utf-8") as f:
            payload = json.load(f)
        assert len(payload["extraction_warnings"]) == 12, "the JSON payload keeps the full list"


def test_extract_model_output_size_warning() -> None:
    """An unusually large serialized output adds a receipt warning (dense real
    data is kept, never silently truncated — but the size is called out)."""
    import pytest

    pytest.importorskip("openpyxl")
    import openpyxl

    mod = _load_extract_model_module()
    with tempfile.TemporaryDirectory(prefix="test-extract-") as d:
        path = os.path.join(d, "normal.xlsx")
        out_path = os.path.join(d, "model_data.json")
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Metric", "2025-01"])
        ws.append(["MRR", 50000])
        wb.save(path)

        mod._OUTPUT_BYTES_WARN = 100  # force the threshold for the test
        old_argv = sys.argv
        sys.argv = ["extract_model.py", "--file", path, "-o", out_path]
        import io
        from contextlib import redirect_stdout

        buf = io.StringIO()
        try:
            with redirect_stdout(buf):
                mod.main()
        finally:
            sys.argv = old_argv
        receipt = json.loads(buf.getvalue())
        rw = receipt.get("extraction_warnings", [])
        assert any("bytes" in w for w in rw), f"oversized output must be called out in the receipt: {rw}"


# ---------------------------------------------------------------------------
# METRIC_SELF_CONTRADICTION — one figure, one source
#
# The observed failure: unit_economics.py computed a burn multiple of 4.5x while
# the checklist sub-agent independently derived ~7x in its evidence prose. The
# founder is handed two numbers and cannot tell which to take to an investor.
#
# The upstream ambiguity that caused it is fixed at the source; these tests lock
# the compose-time BACKSTOP, which must catch the class without crying wolf on
# the many legitimate reasons a second number appears near a metric name.
# ---------------------------------------------------------------------------


def _ue_with_ratios(burn_multiple: float = 4.5, ltv_cac: float = 3.2) -> dict[str, Any]:
    return {
        "metrics": [
            {
                "id": "burn_multiple",
                "name": "burn_multiple",
                "value": burn_multiple,
                "rating": "warning",
                "evidence": "Net burn over net-new ARR",
                "benchmark": {"target": 2.0, "source": "test", "as_of": "2025-Q1"},
                "benchmark_source": "test",
                "benchmark_as_of": "2025-Q1",
            },
            {
                "id": "ltv_cac_ratio",
                "name": "ltv_cac_ratio",
                "value": ltv_cac,
                "rating": "acceptable",
                "evidence": "LTV over CAC",
                "benchmark": {"target": 3.0, "source": "test", "as_of": "2025-Q1"},
                "benchmark_source": "test",
                "benchmark_as_of": "2025-Q1",
            },
        ],
    }


def _checklist_with_evidence(evidence: str) -> dict[str, Any]:
    checklist = json.loads(json.dumps(_VALID_CHECKLIST))
    checklist["items"][0]["evidence"] = evidence
    checklist["items"][0]["id"] = "METRIC_34"
    typed_checklist: dict[str, Any] = checklist
    return typed_checklist


def _compose_codes(unit_economics: dict[str, Any], checklist: dict[str, Any]) -> list[str]:
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": checklist,
            "unit_economics.json": unit_economics,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0
    assert data is not None
    return [w["code"] for w in data["validation"]["warnings"]]


def test_metric_self_contradiction_catches_the_observed_burn_multiple_conflict() -> None:
    """The live failure: computed 4.5x, checklist evidence says ~7x."""
    codes = _compose_codes(
        _ue_with_ratios(burn_multiple=4.5),
        _checklist_with_evidence("Burn multiple of approximately 7x is far above the 2.0x seed benchmark."),
    )
    assert "METRIC_SELF_CONTRADICTION" in codes


def test_metric_self_contradiction_names_both_figures() -> None:
    """The founder must be able to adjudicate, which needs both numbers."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _checklist_with_evidence("Burn multiple of 7x observed."),
            "unit_economics.json": _ue_with_ratios(burn_multiple=4.5),
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0
    assert data is not None
    msg = next(w["message"] for w in data["validation"]["warnings"] if w["code"] == "METRIC_SELF_CONTRADICTION")
    assert "4.5" in msg and "7" in msg
    assert "unit_economics.json" in msg, "must name which artifact is authoritative"


def test_metric_self_contradiction_founder_message_reaches_report_md_not_json_message() -> None:
    """report.json keeps the artifact-naming `message`; report.md renders the
    plain-language `founder_message` instead -- no artifact filename, no
    third-person reference to the founder.
    """
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _checklist_with_evidence("Burn multiple of 7x observed."),
            "unit_economics.json": _ue_with_ratios(burn_multiple=4.5),
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0
    assert data is not None

    warning = next(w for w in data["validation"]["warnings"] if w["code"] == "METRIC_SELF_CONTRADICTION")
    assert "founder_message" in warning
    founder_msg = warning["founder_message"]
    assert "unit_economics.json" not in founder_msg
    assert "the founder" not in founder_msg
    assert "4.5" in founder_msg and "7" in founder_msg

    # report.json's `message` is untouched -- still names the authoritative artifact.
    assert "unit_economics.json" in warning["message"]

    # report.md renders the founder-facing text, not the raw artifact-naming message.
    report_md = data["report_markdown"]
    assert founder_msg in report_md
    contradiction_lines = [line for line in report_md.splitlines() if "different numbers for the" in line]
    assert contradiction_lines, "expected a rendered Warnings line for METRIC_SELF_CONTRADICTION"
    for line in contradiction_lines:
        assert "unit_economics.json" not in line
        assert "the founder" not in line


def test_metric_self_contradiction_allows_quoting_the_benchmark() -> None:
    """ "4.5x vs the 2.0x benchmark" is the NORMAL phrasing — never flag it.

    Without this the check fires on almost every well-written evidence line and
    the reader learns to ignore it, which is worse than not having the check.
    """
    codes = _compose_codes(
        _ue_with_ratios(burn_multiple=4.5),
        _checklist_with_evidence("Burn multiple of 4.5x exceeds the 2.0x strong benchmark for seed stage."),
    )
    assert "METRIC_SELF_CONTRADICTION" not in codes


def test_metric_self_contradiction_tolerates_prose_rounding() -> None:
    """A computed 4.53 written as "4.5x" is rounding, not a contradiction."""
    codes = _compose_codes(
        _ue_with_ratios(burn_multiple=4.53),
        _checklist_with_evidence("Burn multiple of 4.5x is the headline efficiency figure."),
    )
    assert "METRIC_SELF_CONTRADICTION" not in codes


def test_metric_self_contradiction_ignores_numbers_not_about_the_metric() -> None:
    """A stray figure elsewhere in the sentence must not attach to the metric."""
    codes = _compose_codes(
        _ue_with_ratios(burn_multiple=4.5),
        _checklist_with_evidence("Net burn grew and headcount is up 7x since the last round."),
    )
    assert "METRIC_SELF_CONTRADICTION" not in codes


def test_metric_self_contradiction_skips_percent_and_currency_metrics() -> None:
    """Scope is ratio metrics only — 0.75 / "75%" / "$75K" are the same value.

    Locking the exclusion so a later change does not widen the check into the
    false-positive territory it was deliberately kept out of.
    """
    ue = {
        "metrics": [
            {
                "id": "gross_margin",
                "name": "gross_margin",
                "value": 0.75,
                "rating": "strong",
                "evidence": "75% GM",
                "benchmark_source": "test",
                "benchmark_as_of": "2024",
            },
        ],
    }
    codes = _compose_codes(ue, _checklist_with_evidence("Gross margin of 75% is strong."))
    assert "METRIC_SELF_CONTRADICTION" not in codes


def test_metric_self_contradiction_absent_when_evidence_agrees() -> None:
    """No false positive on the happy path."""
    codes = _compose_codes(
        _ue_with_ratios(burn_multiple=4.5, ltv_cac=3.2),
        _checklist_with_evidence("Burn multiple of 4.5x and an LTV/CAC of 3.2 are both in range."),
    )
    assert "METRIC_SELF_CONTRADICTION" not in codes


# ---------------------------------------------------------------------------
# Non-USD reference grades must reach the founder
#
# With the USD-denominated absolute floors correctly suppressed, LTV/CAC, burn
# multiple and Rule of 40 all land `contextual` at once — a page of numbers with
# no assessment, for exactly the audience most likely to file in a local
# currency. The dimensionless comparison DID run (a 1.5x ratio is 1.5x in any
# currency, no FX involved); these lock that its grade is surfaced as a clearly
# marked reference rather than discarded.
# ---------------------------------------------------------------------------


def _ue_with_reference_grade() -> dict[str, Any]:
    return {
        "currency": "ILS",
        "metrics": [
            {
                "id": "burn_multiple",
                "name": "burn_multiple",
                "value": 2.1,
                "rating": "contextual",
                "evidence": "Net burn over net-new ARR; ratio shown but not benchmark-compared",
                "benchmark_reference_rating": "acceptable",
                "benchmark_reference_source": "CFO Advisors 2025",
                "benchmark_reference_as_of": "2025-Q1",
                "benchmark_source": "",
                "benchmark_as_of": "",
            },
            {
                "id": "cac",
                "name": "cac",
                "value": 1500,
                "rating": "acceptable",
                "evidence": "Fully loaded CAC",
                "benchmark_source": "test",
                "benchmark_as_of": "2024",
            },
        ],
        "summary": {},
    }


def test_non_usd_reference_grade_is_shown_not_swallowed() -> None:
    """The report must show the graded reference, not a bare "Contextual"."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _ue_with_reference_grade(),
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "Acceptable (reference)" in md, "the preserved grade must reach the founder"
    # And it must be explained, since an unexplained "(reference)" is its own puzzle.
    assert "reference grade" in md.lower()
    assert "no FX conversion" in md, "must state that no rate was invented"
    # A metric with a real rating is untouched.
    assert "| Acceptable |" in md


def test_reference_grade_note_precedes_the_metrics_table() -> None:
    """The note must not land between the header row and the separator row.

    Appending it after the two header lines silently breaks the markdown table —
    caught only by reading the rendered output, so pin the ordering.
    """
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _ue_with_reference_grade(),
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    note_at = md.lower().find("reference grade")
    header_at = md.find("| Metric | Value | Rating | Evidence |")
    separator_at = md.find("|--------|-------|--------|----------|")
    assert -1 < note_at < header_at < separator_at, "note must sit above the intact table header"
    assert md[header_at:separator_at].strip() == "| Metric | Value | Rating | Evidence |"


def test_usd_review_shows_no_reference_grade_machinery() -> None:
    """Back-compat: a USD model must not gain the note or the "(reference)" suffix."""
    d = _make_fmr_artifact_dir(
        {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
    )
    rc, data, _ = _run_compose(d)
    assert rc == 0
    assert data is not None
    md = data["report_markdown"]
    assert "(reference)" not in md
    assert "reference grade" not in md.lower()


# ---------------------------------------------------------------------------
# UNDECLARED_AGENT_VALUE — provenance for values the founder never stated
#
# Live: a conversational run wrote `bridge.runway_target_months: 24` for a founder
# who never mentioned a runway target. The value was harmless (runway.py defaults
# to 24 anyway) but inputs.json recorded it indistinguishably from a stated input,
# so nothing downstream could tell the difference. Same defect market-sizing fixed
# with `founder_stated_inputs`, approached from the other side: there we assert the
# founder's numbers survive, here we assert the agent's are labelled as its own.
#
# The fix is provenance, not prohibition — defaulting is allowed, hiding it is not.
# ---------------------------------------------------------------------------


def _conversational_inputs(**over: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "company": {
            "name": "TestCo",
            "stage": "seed",
            "model_format": "conversational",
            "currency": "ILS",
        },
        "revenue": {"mrr": 90000},
        "burn": {"net_monthly": -380000},
        "cash": {"current_balance": 4200000},
        "bridge": {"raise_amount": 12000000, "runway_target_months": 24},
        "metadata": {"run_id": "T"},
    }
    base.update(over)
    return base


def _validate_codes(inputs: dict[str, Any]) -> list[str]:
    rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(inputs))
    assert data is not None, "validate_inputs.py produced no JSON"
    return [w["code"] for w in data.get("warnings", [])]


def test_undeclared_agent_value_flags_the_live_case() -> None:
    """Conversational run carrying a computation-feeding field with no declaration."""
    assert "UNDECLARED_AGENT_VALUE" in _validate_codes(_conversational_inputs())


def test_undeclared_agent_value_names_the_offending_paths() -> None:
    """The founder must be told WHICH fields are unattributed, not just that some are."""
    rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(_conversational_inputs()))
    assert data is not None
    msg = next(w["message"] for w in data["warnings"] if w["code"] == "UNDECLARED_AGENT_VALUE")
    assert "bridge.runway_target_months" in msg
    assert "agent_supplied" in msg, "must name the field that resolves it"


def test_declaring_the_field_clears_the_warning() -> None:
    """Defaulting is permitted — declaring it is the requirement."""
    inputs = _conversational_inputs(agent_supplied=["bridge.runway_target_months"])
    assert "UNDECLARED_AGENT_VALUE" not in _validate_codes(inputs)


def test_empty_declaration_is_a_declaration() -> None:
    """`[]` means "the founder stated all of these" — absent means unanswered.

    The distinction is the whole mechanism: an optional field that can be silently
    omitted cannot force the question, which is why absence is what we flag.
    """
    assert "UNDECLARED_AGENT_VALUE" not in _validate_codes(_conversational_inputs(agent_supplied=[]))


def test_spreadsheet_runs_are_exempt() -> None:
    """Extraction has its own provenance chain (validate_extraction.py)."""
    inputs = _conversational_inputs()
    inputs["company"]["model_format"] = "spreadsheet"
    assert "UNDECLARED_AGENT_VALUE" not in _validate_codes(inputs)


def test_no_computation_feeding_field_means_nothing_to_declare() -> None:
    """No false positive when the agent supplied nothing that feeds a computation."""
    inputs = _conversational_inputs()
    del inputs["bridge"]
    assert "UNDECLARED_AGENT_VALUE" not in _validate_codes(inputs)


def test_undeclared_agent_value_is_not_critical() -> None:
    """It is a disclosure gap, not a data error — it must not block the review."""
    rc, data, _ = run_script("validate_inputs.py", ["--pretty"], stdin_data=json.dumps(_conversational_inputs()))
    assert data is not None
    w = next(x for x in data["warnings"] if x["code"] == "UNDECLARED_AGENT_VALUE")
    assert w["critical"] is False


# --- producer refusal + downstream detection (fleet-wide loud-failure pass) ----


def test_fmr_producers_refuse_loudly_instead_of_clobbering() -> None:
    """`checklist.py`, `unit_economics.py` and `runway.py` all shared one defect.

    Each exited 0 on a validation error, printed an `{"ok":true}` receipt, and wrote an
    analysis-free stub over its own canonical artifact. SKILL.md's producer-error branch is
    written as "the pipe fails next", so with exit 0 it could never fire, and the prior good
    artifact was gone. Exit 1 + not writing `-o` is the fix; both halves are asserted.
    """
    cases = [
        ("checklist.py", "checklist.json", json.dumps({"notitems": 1})),
        ("unit_economics.py", "unit_economics.json", json.dumps({"nocompany": 1})),
        ("runway.py", "runway.json", json.dumps({"nocompany": 1})),
    ]
    for script, artifact, payload in cases:
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, artifact)
            with open(out, "w") as f:
                f.write('{"sentinel": true}')
            rc, data, stderr = run_script(script, ["-o", out], stdin_data=payload)
            assert rc == 1, f"{script}: a rejected input must exit non-zero"
            assert stderr.strip(), f"{script}: a rejected run must say so on stderr"
            assert data is not None and data["validation"]["status"] == "invalid", script
            with open(out) as f:
                assert json.load(f) == {"sentinel": True}, f"{script}: canonical artifact clobbered"


def test_compose_flags_an_invalid_fmr_artifact_at_high_severity() -> None:
    """A rejected checklist/unit-economics/runway step must be loud downstream too.

    Before this the only signals were medium — hence suppressible via accepted_warnings — and
    each named a symptom rather than the cause.
    """
    for name in ("checklist.json", "unit_economics.json", "runway.json"):
        arts: dict[str, Any] = {
            "inputs.json": _VALID_INPUTS,
            "checklist.json": _VALID_CHECKLIST,
            "unit_economics.json": _VALID_UNIT_ECONOMICS,
            "runway.json": _VALID_RUNWAY,
        }
        arts[name] = {"validation": {"status": "invalid", "errors": ["bad input"]}}
        d = _make_fmr_artifact_dir(arts)
        rc, data, err = _run_compose(d)
        assert data is not None, err
        hits = [w for w in data["validation"]["warnings"] if w["code"] == "ARTIFACT_INVALID"]
        assert hits, f"{name}: a rejected producer artifact must raise ARTIFACT_INVALID"
        assert hits[0]["severity"] == "high", f"{name}: must not be acceptable-away"
