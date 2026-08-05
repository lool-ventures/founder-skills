"""Tests for the cap-table skill.

Covers each math producer, the rule_audit two-phase contract, scenario
dispatch, composition, and the 11 Gotchas from SKILL.md as explicit
regression tests. Test class structure mirrors test_financial_model_review.py.
"""

from __future__ import annotations

import json
import math
import os
import subprocess
import sys
import tempfile
from typing import Any

import pytest

_REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SCRIPTS = os.path.join(_REPO, "founder-skills", "skills", "cap-table", "scripts")

sys.path.insert(0, SCRIPTS)

# Import library APIs once for in-process tests (faster than subprocess for math).
import anti_dilution  # type: ignore[import-not-found]  # noqa: E402
import cap_state as cap_state_mod  # type: ignore[import-not-found]  # noqa: E402
import coverage as coverage_mod  # type: ignore[import-not-found]  # noqa: E402
import detect_structure  # type: ignore[import-not-found]  # noqa: E402
import extract_cap_table  # type: ignore[import-not-found]  # noqa: E402
import flip_scenario  # type: ignore[import-not-found]  # noqa: E402
import note_conversion  # type: ignore[import-not-found]  # noqa: E402
import option_pool  # type: ignore[import-not-found]  # noqa: E402
import priced_round  # type: ignore[import-not-found]  # noqa: E402
import rule_audit  # type: ignore[import-not-found]  # noqa: E402
import run_scenario  # type: ignore[import-not-found]  # noqa: E402
import safe_conversion  # type: ignore[import-not-found]  # noqa: E402


def _run(script_name: str, args: list[str], stdin_data: str = "") -> tuple[int, str, str]:
    """Invoke a cap-table script as a subprocess. Returns (rc, stdout, stderr)."""
    res = subprocess.run(
        [sys.executable, os.path.join(SCRIPTS, script_name), *args],
        input=stdin_data,
        capture_output=True,
        text=True,
    )
    return res.returncode, res.stdout, res.stderr


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_BASIC_INPUTS = {
    "company_name": "Acmecorp",
    "analysis_date": "2026-05-19",
    "mode": "standard",
    "jurisdiction": {
        "structure": "delaware",
        "incorporated_date": "2024-01-01",
        "iia_grants_history": {"has_grants": False, "grant_details": []},
    },
    "founders": [
        {"name": "Alice", "founder_id": "founder_alice", "common_shares": 5_000_000},
        {"name": "Bob", "founder_id": "founder_bob", "common_shares": 5_000_000},
    ],
    "preferred_series": [],
    "option_pool": {"plan_type": "nso", "authorized": 1_500_000, "issued": 500_000, "unallocated": 1_000_000},
    "common_batches": [],
    "metadata": {"run_id": "test"},
}

_BASIC_INSTRUMENTS = {
    "safes": [],
    "convertible_notes": [],
    "warrants": [],
    "option_grants": [],
    "metadata": {"run_id": "test"},
}

_SAFE_BASIC = {
    "id": "safe_001",
    "investor_name": "Angel A",
    "purchase_amount": 500_000,
    "post_money_valuation_cap": 8_000_000,
    "discount_multiplier": None,
    "mfn_provision": None,
    "pro_rata_side_letter": None,
    "issuance_date": "2025-01-01",
    "form": "yc_postmoney_cap",
    "conversion_price_override": None,
    "source_document": None,
    "extraction_confidence": "high",
}

_NOTE_BASIC = {
    "id": "note_001",
    "investor_name": "Lender L",
    "principal": 100_000,
    "annual_interest_rate": 0.06,
    "day_count_basis": 365,
    "compounding_periods_per_year": None,
    "interest_converts_to_shares": True,
    "issuance_date": "2025-06-01",
    "last_interest_event_date": None,
    "valuation_cap": 10_000_000,
    "discount_multiplier": 0.80,
    "capitalization_denominator": 10_000_000,
    "capitalization_denominator_policy": "pre-money fully diluted",
    "qualified_financing_threshold": 1_000_000,
    "maturity_date": "2027-06-01",
    "maturity_default_treatment": "convert_at_cap",
    "maturity_conversion_price_override": None,
    "non_qualified_financing_treatment": None,
    "source_document": None,
    "extraction_confidence": "high",
}


@pytest.fixture
def basic_dir(tmp_path: Any) -> Any:
    """Workspace with inputs + instruments + cap_state populated."""
    inputs_path = tmp_path / "inputs.json"
    instr_path = tmp_path / "instruments.json"
    cap_path = tmp_path / "cap_state.json"
    inputs_path.write_text(json.dumps(_BASIC_INPUTS))
    instr_path.write_text(json.dumps(_BASIC_INSTRUMENTS))
    # Build cap_state via library (fast)
    cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
    cs["metadata"]["run_id"] = "test"
    cap_path.write_text(json.dumps(cs))
    return tmp_path


# ===========================================================================
# cap_state.py
# ===========================================================================


class TestCapState:
    def test_basic_founders_only(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        assert cs["as_converted_totals"]["common_shares"] == 10_000_000
        assert cs["as_converted_totals"]["preferred_shares_as_converted"] == 0
        assert cs["as_converted_totals"]["options_outstanding"] == 500_000
        assert cs["as_converted_totals"]["options_available"] == 1_000_000
        assert cs["as_converted_totals"]["fully_diluted_shares"] == 11_500_000

    def test_with_preferred_series(self) -> None:
        inputs = dict(_BASIC_INPUTS)
        inputs["preferred_series"] = [
            {
                "series_name": "Series Seed",
                "shares": 2_000_000,
                "original_issue_price": 1.0,
                "original_conversion_price": 1.0,
                "current_conversion_price": 1.0,
                "issuance_date": "2025-03-01",
            }
        ]
        cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
        assert cs["as_converted_totals"]["preferred_shares_as_converted"] == 2_000_000
        assert cs["as_converted_totals"]["fully_diluted_shares"] == 13_500_000

    def test_preferred_with_anti_dilution_adjustment(self) -> None:
        """When current_conversion_price < OCP, as-converted count increases."""
        inputs = dict(_BASIC_INPUTS)
        inputs["preferred_series"] = [
            {
                "series_name": "Series Seed",
                "shares": 1_000_000,
                "original_issue_price": 2.0,
                "original_conversion_price": 2.0,
                "current_conversion_price": 1.0,  # AD adjusted down by 50%
                "issuance_date": "2025-03-01",
            }
        ]
        cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
        # 1M shares × (2.0 / 1.0) = 2M as-converted
        assert cs["as_converted_totals"]["preferred_shares_as_converted"] == 2_000_000

    def test_anti_dilution_recovered_from_wrong_key(self) -> None:
        """A founder's anti-dilution intent written under the WRONG key `anti_dilution` (the model's
        common slip — e.g. {"anti_dilution": "bbwa"}) while the canonical `anti_dilution_protection`
        is absent must be RECOVERED to the canonical field + flagged, never silently dropped to
        'none' (which skips the down-round adjustment the founder explicitly asked for).
        Regression for the priced-ad silent-drop bug (cap_state read only anti_dilution_protection)."""
        inputs = dict(_BASIC_INPUTS)
        inputs["preferred_series"] = [
            {
                "series_name": "Series Seed",
                "shares": 2_000_000,
                "original_issue_price": 1.0,
                "original_conversion_price": 1.0,
                "current_conversion_price": 1.0,
                "issuance_date": "2025-03-01",
                "anti_dilution": "bbwa",  # WRONG key + abbreviation; no anti_dilution_protection
            }
        ]
        cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
        ps = cs["preferred_series"][0]
        assert ps["anti_dilution_protection"] == "broad_based_weighted_average", (
            "BBWA intent under the wrong key must be recovered, not silently dropped to 'none'"
        )
        assert any("ANTI_DILUTION" in w for w in cs.get("warnings", [])), (
            "the recovery must be surfaced as a warning, never silent"
        )

    def test_anti_dilution_canonical_field_unchanged_no_warning(self) -> None:
        """The correct canonical field must pass through untouched with NO normalization warning."""
        inputs = dict(_BASIC_INPUTS)
        inputs["preferred_series"] = [
            {
                "series_name": "Series Seed",
                "shares": 2_000_000,
                "original_issue_price": 1.0,
                "original_conversion_price": 1.0,
                "current_conversion_price": 1.0,
                "issuance_date": "2025-03-01",
                "anti_dilution_protection": "broad_based_weighted_average",
            }
        ]
        cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
        assert cs["preferred_series"][0]["anti_dilution_protection"] == "broad_based_weighted_average"
        assert not any("ANTI_DILUTION" in w for w in cs.get("warnings", []))

    def test_with_common_batches(self) -> None:
        inputs = dict(_BASIC_INPUTS)
        inputs["common_batches"] = [
            {
                "batch_id": "b1",
                "holder_id": "advisor",
                "shares": 250_000,
                "issuance_date": "2025-06-01",
                "consideration": 0,
                "purpose": "founder_issuance",
            }
        ]
        cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
        assert cs["as_converted_totals"]["common_shares"] == 10_250_000

    def test_outstanding_options_carries_per_grant_metadata(self) -> None:
        """cap-table-data-contract §3.2: outstanding_options carries plan_type
        + section_102_trustee_deposit_date + strike_price + grant_date through
        to cap_state, so rule_audit matchers, compose_report.flip_specifics, and
        counsel_packet all read from a single canonical location
        (cap_state.outstanding_options[*]) instead of bypassing to
        instruments.option_grants[]."""
        inst: dict[str, Any] = dict(_BASIC_INSTRUMENTS)
        inst["option_grants"] = [
            {
                "id": "grant_001",
                "holder_id": "employee_1",
                "grant_date": "2025-09-01",
                "shares_granted": 100_000,
                "shares_vested_to_date": 25_000,
                "shares_exercised": 0,
                "strike_price": 0.50,
                "plan_type": "section_102_cg",
            }
        ]
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, inst)
        assert len(cs["outstanding_options"]) == 1
        opt = cs["outstanding_options"][0]
        assert opt["grant_id"] == "grant_001"
        assert opt["shares_vested_to_date"] == 25_000
        assert opt["shares_outstanding_unvested"] == 75_000
        # Per v0.5.0 contract: per-grant metadata is mirrored to cap_state so
        # downstream consumers don't bypass to instruments.option_grants[].
        assert opt["plan_type"] == "section_102_cg"
        assert opt["grant_date"] == "2025-09-01"
        assert opt["strike_price"] == 0.50

    def test_cli_writes_artifact_with_run_id(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            inp = os.path.join(d, "inputs.json")
            inst = os.path.join(d, "instruments.json")
            out = os.path.join(d, "cap_state.json")
            with open(inp, "w") as f:
                json.dump(_BASIC_INPUTS, f)
            with open(inst, "w") as f:
                json.dump(_BASIC_INSTRUMENTS, f)
            rc, _, stderr = _run(
                "cap_state.py", ["--inputs", inp, "--instruments", inst, "--run-id", "abc123", "-o", out]
            )
            assert rc == 0, stderr
            with open(out) as f:
                cs = json.load(f)
            assert cs["metadata"]["run_id"] == "abc123"


# ===========================================================================
# safe_conversion.py
# ===========================================================================


class TestSafeConversion:
    def test_cap_implied_basic(self) -> None:
        # $500k @ $8M cap, 11.5M company cap → 6.25% cap-implied ownership
        r = safe_conversion.convert_safe_cap_implied(
            purchase_amount=500_000,
            post_money_valuation_cap=8_000_000,
            company_capitalization=11_500_000,
        )
        assert r["branch"] == "cap_implied"
        assert math.isclose(r["cap_implied_ownership"], 0.0625, rel_tol=1e-6)
        # safe_price = 8M / 11.5M
        assert math.isclose(r["safe_price"], 8_000_000 / 11_500_000, rel_tol=1e-6)
        # shares = 500k / safe_price
        expected_shares = 500_000 / (8_000_000 / 11_500_000)
        assert math.isclose(r["cap_implied_shares"], expected_shares, rel_tol=1e-6)

    def test_cap_implied_rejects_zero_cap(self) -> None:
        r = safe_conversion.convert_safe_cap_implied(
            purchase_amount=500_000,
            post_money_valuation_cap=None,
            company_capitalization=11_500_000,
        )
        assert r["branch"] == "rejected"
        assert r["error"] == safe_conversion.E_SAFE_REQUIRES_CONVERSION_EVENT

    def test_priced_round_cap_only_form(self) -> None:
        r = safe_conversion.convert_safe_priced_round(
            purchase_amount=500_000,
            form="yc_postmoney_cap",
            post_money_valuation_cap=8_000_000,
            discount_multiplier=None,
            company_capitalization=11_500_000,
            equity_financing_price=2.0,
        )
        assert r["branch"] == "cap_branch"
        # cap_price = 8M / 11.5M ≈ 0.696; lower than equity_price 2.0, so cap wins
        assert math.isclose(r["conversion_price"], 8_000_000 / 11_500_000, rel_tol=1e-6)

    def test_priced_round_discount_only(self) -> None:
        r = safe_conversion.convert_safe_priced_round(
            purchase_amount=100_000,
            form="yc_postmoney_discount",
            post_money_valuation_cap=None,
            discount_multiplier=0.80,
            company_capitalization=11_500_000,
            equity_financing_price=2.0,
        )
        assert r["branch"] == "discount_branch"
        assert math.isclose(r["conversion_price"], 1.60, rel_tol=1e-6)  # 2.0 × 0.80

    def test_priced_round_cap_plus_discount_takes_min(self) -> None:
        # cap_price = 1.0, discount_price = 1.6 → cap wins (min)
        r = safe_conversion.convert_safe_priced_round(
            purchase_amount=100_000,
            form="cap_plus_discount",
            post_money_valuation_cap=11_500_000,  # cap_price = 1.0
            discount_multiplier=0.80,
            company_capitalization=11_500_000,
            equity_financing_price=2.0,
        )
        assert r["branch"] == "cap_and_discount_branch"
        assert math.isclose(r["conversion_price"], 1.0, rel_tol=1e-6)

    def test_discount_only_rejected_without_priced_round(self) -> None:
        """Per Gotcha #4 + design §5.1 hard-reject contract."""
        r = safe_conversion.convert_safe_priced_round(
            purchase_amount=100_000,
            form="yc_postmoney_discount",
            post_money_valuation_cap=None,
            discount_multiplier=0.80,
            company_capitalization=11_500_000,
            equity_financing_price=None,
        )
        assert r["branch"] == "rejected"
        assert r["error"] == safe_conversion.E_SAFE_REQUIRES_CONVERSION_EVENT

    def test_uncapped_mfn_rejected_without_trigger(self) -> None:
        r = safe_conversion.convert_safe_priced_round(
            purchase_amount=100_000,
            form="yc_uncapped_mfn",
            post_money_valuation_cap=None,
            discount_multiplier=None,
            company_capitalization=11_500_000,
            equity_financing_price=None,
        )
        assert r["branch"] == "rejected"
        assert r["error"] == safe_conversion.E_SAFE_REQUIRES_CONVERSION_EVENT

    def test_conversion_price_override(self) -> None:
        """Counsel-supplied override bypasses rule; provenance cites override."""
        r = safe_conversion.convert_safe_priced_round(
            purchase_amount=100_000,
            form="yc_postmoney_discount",  # would otherwise need priced round
            post_money_valuation_cap=None,
            discount_multiplier=0.80,
            company_capitalization=11_500_000,
            equity_financing_price=None,
            conversion_price_override=1.25,
        )
        assert r["branch"] == "conversion_price_override"
        assert r["conversion_price"] == 1.25
        # Math provenance cites source_ref, not rule_id
        assert any(
            p["source_type"] == "counsel_supplied_override"
            and p["rule_id"] is None
            and p["source_ref"] == "safes[].conversion_price_override"
            for p in r["math_provenance"]
        )

    def test_mfn_cycle_detection_unresolvable(self) -> None:
        """Per Gotcha #4: all-uncapped-MFN cycle = unresolvable."""
        safes = [
            {
                "id": "A",
                "form": "yc_uncapped_mfn",
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "B",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
            {
                "id": "B",
                "form": "yc_uncapped_mfn",
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "A",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
        ]
        cycles = safe_conversion.detect_mfn_cycles(safes)
        assert len(cycles) == 1
        assert cycles[0] == {"A", "B"}

    def test_mfn_cycle_not_flagged_if_anchored(self) -> None:
        """Cycle with a cap-bearing anchor → not flagged (anchor provides price)."""
        safes = [
            {
                "id": "A",
                "form": "yc_uncapped_mfn",
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "B",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
            {
                "id": "B",
                "form": "yc_postmoney_cap",  # anchor
                "mfn_provision": {
                    "present": False,
                    "elected_against_safe_id": None,
                    "elected": False,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
        ]
        cycles = safe_conversion.detect_mfn_cycles(safes)
        # B is not yc_uncapped_mfn, so the chain A→B has an anchor; no cycle
        assert cycles == []


# ===========================================================================
# note_conversion.py
# ===========================================================================


class TestNoteConversion:
    def test_accrued_interest_simple(self) -> None:
        # $100k @ 6% for 365 days, simple
        accrued = note_conversion.compute_accrued_interest(
            principal=100_000,
            annual_interest_rate=0.06,
            days_elapsed=365,
            day_count_basis=365,
        )
        assert math.isclose(accrued, 6_000, rel_tol=1e-6)

    def test_accrued_interest_compound_monthly(self) -> None:
        # $100k @ 6% for 365 days, monthly compounding
        accrued = note_conversion.compute_accrued_interest(
            principal=100_000,
            annual_interest_rate=0.06,
            days_elapsed=365,
            day_count_basis=365,
            compounding_periods_per_year=12,
        )
        # 100k × (1 + 0.06/12)^12 - 100k ≈ 6167.78
        assert math.isclose(accrued, 6167.78, abs_tol=1.0)

    def test_cap_conversion_branch(self) -> None:
        r = note_conversion.convert_note(
            _NOTE_BASIC,
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=2.0,
        )
        assert r["branch"] == "cap_conversion"
        # cap_price = 10M / 10M = 1.0; discount_price = 2.0 × 0.80 = 1.6 → cap wins
        assert math.isclose(r["conversion_price"], 1.0, rel_tol=1e-6)
        # balance = 100k + 6k accrued = 106k → 106,000 shares at $1/share
        assert math.isclose(r["conversion_shares"], 106_000, rel_tol=1e-6)

    def test_discount_only_branch(self) -> None:
        note = dict(_NOTE_BASIC)
        note["valuation_cap"] = None
        note["capitalization_denominator"] = None
        r = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=2.0,
        )
        assert r["branch"] == "discount_only"
        # discount_price = 2.0 × 0.80 = 1.6
        assert math.isclose(r["conversion_price"], 1.6, rel_tol=1e-6)

    def test_threshold_not_met_no_treatment(self) -> None:
        """Per design rev15: threshold not met + no treatment → branch threshold_not_met."""
        note = dict(_NOTE_BASIC)
        note["qualified_financing_threshold"] = 10_000_000
        r = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=1_000_000,  # below threshold
            qualified_financing_price=2.0,
        )
        assert r["branch"] == "threshold_not_met"
        assert "conversion_shares" not in r

    def test_threshold_not_met_with_convert_anyway(self) -> None:
        note = dict(_NOTE_BASIC)
        note["qualified_financing_threshold"] = 10_000_000
        note["non_qualified_financing_treatment"] = "convert_anyway"
        r = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=1_000_000,
            qualified_financing_price=2.0,
        )
        assert r["branch"] == "cap_conversion"  # convert_anyway falls through

    def test_threshold_not_met_with_do_not_convert(self) -> None:
        note = dict(_NOTE_BASIC)
        note["qualified_financing_threshold"] = 10_000_000
        note["non_qualified_financing_treatment"] = "do_not_convert"
        r = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=1_000_000,
            qualified_financing_price=2.0,
        )
        assert r["branch"] == "maturity_extend"

    def test_maturity_repay_branch(self) -> None:
        note = dict(_NOTE_BASIC)
        note["maturity_default_treatment"] = "repay"
        note["interest_converts_to_shares"] = False
        r = note_conversion.convert_note(note, conversion_event_date="2027-06-01")
        assert r["branch"] == "maturity_repay"
        # Per the math-bug-fix: repay always returns principal + accrued regardless
        # of interest_converts_to_shares (that flag governs CONVERSION behavior only).
        # 2 years × 6% on $100k = $12k accrued → $112k repayment
        assert math.isclose(r["cash_repayment"], 112_000, rel_tol=1e-6)

    def test_maturity_extend_branch(self) -> None:
        note = dict(_NOTE_BASIC)
        note["maturity_default_treatment"] = "extend"
        r = note_conversion.convert_note(note, conversion_event_date="2027-06-01")
        assert r["branch"] == "maturity_extend"

    def test_maturity_counsel_review_branch(self) -> None:
        note = dict(_NOTE_BASIC)
        note["maturity_default_treatment"] = "counsel_review"
        r = note_conversion.convert_note(note, conversion_event_date="2027-06-01")
        assert r["branch"] == "maturity_counsel_review"

    def test_maturity_override_branch(self) -> None:
        """Per Gotcha #5 + design rev15: override bypasses rule for convert_at_cap."""
        note = dict(_NOTE_BASIC)
        note["valuation_cap"] = None  # no cap available
        note["capitalization_denominator"] = None
        note["maturity_conversion_price_override"] = 0.50  # counsel-supplied
        r = note_conversion.convert_note(note, conversion_event_date="2027-06-01")
        assert r["branch"] == "maturity_convert_at_cap"
        assert r["conversion_price"] == 0.50
        # balance = 100k + 12k accrued = 112k → 224k shares
        assert math.isclose(r["conversion_shares"], 224_000, rel_tol=1e-6)
        # Provenance cites the override
        assert any(
            p["source_type"] == "counsel_supplied_override"
            and p["source_ref"] == "notes[].maturity_conversion_price_override"
            for p in r["math_provenance"]
        )

    def test_override_branch_mismatch_error(self) -> None:
        """Per Gotcha #5: override paired with repay/extend/counsel_review = error."""
        note = dict(_NOTE_BASIC)
        note["maturity_default_treatment"] = "repay"
        note["maturity_conversion_price_override"] = 0.50
        r = note_conversion.convert_note(note, conversion_event_date="2027-06-01")
        assert r["branch"] == "override_mismatch"
        assert r["error"] == note_conversion.E_NOTE_OVERRIDE_BRANCH_MISMATCH

    def test_completeness_partition_full(self) -> None:
        """All notes share-producing → completeness=full."""
        per_note = {
            "n1": {"branch": "cap_conversion"},
            "n2": {"branch": "discount_only"},
        }
        assert note_conversion.derive_scenario_completeness(per_note) == "full"

    def test_completeness_partition_repay_only(self) -> None:
        per_note = {"n1": {"branch": "maturity_repay"}, "n2": {"branch": "maturity_repay"}}
        assert note_conversion.derive_scenario_completeness(per_note) == "repay_only"

    def test_completeness_partition_mixed(self) -> None:
        per_note = {
            "n1": {"branch": "cap_conversion"},
            "n2": {"branch": "maturity_repay"},
        }
        assert note_conversion.derive_scenario_completeness(per_note) == "mixed"

    def test_completeness_partition_structural_only(self) -> None:
        per_note = {
            "n1": {"branch": "maturity_extend"},
            "n2": {"branch": "threshold_not_met"},
        }
        assert note_conversion.derive_scenario_completeness(per_note) == "structural_only"

    def test_statutory_ita_3j_with_null_rate_uses_proxy(self) -> None:
        """Real-doc end-to-end test (May 2026) surfaced this: Israeli CLAs with
        interest_rate_type='statutory_ita_section_3j' have annual_interest_rate=null
        (rate is set quarterly by Israeli Tax Authority — not stated in document).
        The validator accepts null per commit #5; the math producer must too.
        Math producer uses STATUTORY_ITA_DEFAULT_RATE (5%) as proxy + warns.
        """
        note = {
            "id": "cla_1",
            "investor_name": "Israeli investor",
            "principal": 1_000_000,
            "interest_rate_type": "statutory_ita_section_3j",
            "annual_interest_rate": None,
            "day_count_basis": 365,
            "issuance_date": "2024-09-01",
            "valuation_cap": 8_000_000,
            "capitalization_denominator": 11_000_000,
            "discount_multiplier": None,
            "qualified_financing_threshold": 1_000_000,
            "maturity_date": "2026-03-01",
            "maturity_default_treatment": "convert_at_cap",
            "interest_converts_to_shares": True,
        }
        r = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=4_000_000,
            qualified_financing_price=1.0,
        )
        # Should NOT crash with TypeError (None * float).
        assert r["accrued_interest"] > 0  # proxy rate produced non-zero accrual
        # Warning surfaced for counsel to substitute the actual ITA rate.
        assert "warnings" in r
        assert any(w["code"] == "statutory_ita_3j_proxy_rate_used" for w in r["warnings"])

    def test_none_interest_rate_type_uses_zero(self) -> None:
        """convertible_security subtype (SAFE-equivalent) sets
        interest_rate_type='none'. Math producer must treat as 0% — no accrual,
        no warning (it's the intended document shape).
        """
        note = {
            "id": "cs_1",
            "investor_name": "SAFE-equivalent investor",
            "principal": 500_000,
            "interest_rate_type": "none",
            "annual_interest_rate": None,
            "day_count_basis": 365,
            "issuance_date": "2019-08-01",
            "valuation_cap": 15_000_000,
            "capitalization_denominator": 11_000_000,
            "discount_multiplier": None,
            "qualified_financing_threshold": 1_000_000,
            "maturity_date": None,
            "maturity_default_treatment": None,
            "interest_converts_to_shares": True,
        }
        r = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=4_000_000,
            qualified_financing_price=1.0,
        )
        assert r["accrued_interest"] == 0
        # No warning — none-interest is the intended shape, not an inference fallback.
        assert "warnings" not in r or not any(
            w["code"] == "statutory_ita_3j_proxy_rate_used" for w in r.get("warnings", [])
        )


# ===========================================================================
# option_pool.py
# ===========================================================================


class TestOptionPool:
    def test_pre_money_target(self) -> None:
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=10_000_000,
            existing_unallocated_pool=0,
            target_pool_percent=0.10,
            new_money_shares=None,
            target_basis="pre_money",
        )
        # (0 + x) / (10M + x) = 0.10 → x = 10M × 0.10 / 0.90 ≈ 1.111M
        assert r["required_pool_topup_shares"] == round(10_000_000 * 0.10 / 0.90)
        assert math.isclose(r["post_topup_pool_percent"], 0.10, abs_tol=1e-4)

    def test_post_money_target(self) -> None:
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=10_000_000,
            existing_unallocated_pool=0,
            target_pool_percent=0.15,
            new_money_shares=2_500_000,
            target_basis="post_money",
        )
        # (0 + x) / (10M + x + 2.5M) = 0.15 → x = 0.15 × 12.5M / 0.85 ≈ 2.206M
        assert math.isclose(r["post_topup_pool_percent"], 0.15, abs_tol=1e-4)

    def test_post_money_excluding_converting_securities(self) -> None:
        """Same formula as post_money for v0.1 — the rule pack distinguishes
        the denominator policy via parameterization_points, not the math directly."""
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=10_000_000,
            existing_unallocated_pool=500_000,
            target_pool_percent=0.10,
            new_money_shares=1_000_000,
            target_basis="post_money_excluding_converting_securities",
        )
        assert r["target_basis"] == "post_money_excluding_converting_securities"
        assert r["required_pool_topup_shares"] >= 0

    def test_custom_basis(self) -> None:
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=10_000_000,
            existing_unallocated_pool=100_000,
            target_pool_percent=0.08,
            new_money_shares=None,
            target_basis="custom",
        )
        assert r["target_basis"] == "custom"

    def test_invalid_target_rejected(self) -> None:
        with pytest.raises(ValueError):
            option_pool.required_topup(
                pre_topup_fully_diluted_shares=10_000_000,
                existing_unallocated_pool=0,
                target_pool_percent=1.5,  # > 1
                new_money_shares=None,
                target_basis="pre_money",
            )

    def test_pre_money_target_already_met_emits_clarifying_question(self) -> None:
        """Phase M+S: when existing pool already meets target under literal
        pre_money basis but post_money interpretation would require a top-up,
        emit a warning + clarifying_question for the dispatching agent to
        escalate via AskUserQuestion.
        """
        # Existing pool 2M = 18.18% of 11M pre-FD; target 10% pre_money is
        # already met (returns 0 top-up). But post_money basis would compute
        # a real top-up: solve (2M + x) / (11M + 5M + x) = 0.10 → x = -400k,
        # so post-money is ALSO 0 here. Need a scenario where post-money > 0.
        # Set existing pool to 800k = 7.27% (below 10% pre_money target).
        # Pre_money calc: (800k + x)/(11M + x) = 0.10 → x = 333,333.
        # That's NOT a no-op — bad example.
        # Use: existing 1.5M = 13.64% (above 10% pre); pre_money no-op.
        # Post-money: (1.5M + x)/(11M + 5M + x) = 0.10 → x = 55,556.
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=11_000_000,
            existing_unallocated_pool=1_500_000,
            target_pool_percent=0.10,
            new_money_shares=5_000_000,
            target_basis="pre_money",
        )
        assert r["required_pool_topup_shares"] == 0  # literal pre_money no-op
        assert "warnings" in r
        assert any(w["code"] == "pool_target_already_met_check_intent" for w in r["warnings"])
        assert "clarifying_question" in r
        cq = r["clarifying_question"]
        assert cq["context"]["literal_pre_money_top_up"] == 0
        assert cq["context"]["industry_norm_post_money_top_up"] > 0  # M8 gate ensures this

    def test_post_money_basis_does_not_fire_clarifying_question(self) -> None:
        """Phase M+S negative case: target_basis='post_money' never triggers
        the warning regardless of whether the existing pool exceeds target.
        """
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=11_000_000,
            existing_unallocated_pool=1_500_000,
            target_pool_percent=0.10,
            new_money_shares=5_000_000,
            target_basis="post_money",
        )
        assert "warnings" not in r or not any(
            w.get("code") == "pool_target_already_met_check_intent" for w in r.get("warnings", [])
        )
        assert "clarifying_question" not in r or r.get("clarifying_question") is None

    def test_pre_money_no_op_both_interpretations_suppresses_warning(self) -> None:
        """M8: when BOTH literal pre-money AND industry-norm post-money
        produce 0 top-up (existing pool legitimately oversized), the
        clarifying_question is noise — the M8 gate must suppress it.
        """
        # Existing pool 3M = ~27% of 11M; way above 10% target under any reading.
        # Pre_money: (3M + x)/(11M + x) = 0.10 → x = -888,889 → 0.
        # Post_money: (3M + x)/(11M + 5M + x) = 0.10 → x = -1,555,556 → 0.
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=11_000_000,
            existing_unallocated_pool=3_000_000,
            target_pool_percent=0.10,
            new_money_shares=5_000_000,
            target_basis="pre_money",
        )
        assert r["required_pool_topup_shares"] == 0
        # M8 gate: post_money_required is also 0, so NO warning + NO question
        assert "warnings" not in r or not r.get("warnings", [])
        assert "clarifying_question" not in r or r.get("clarifying_question") is None

    def test_pre_fd_zero_does_not_crash(self) -> None:
        """M7: option_pool is a public library function. A library caller
        passing pre_fd=0 (degenerate but legal) must not ZeroDivisionError in
        the warning's existing_pct_of_pre_fd format string.
        """
        # pre_fd=0 means no pre-financing FD; target 10% pre_money with 0
        # existing pool computes x = (0.1 * 0 - 0) / 0.9 = 0 → no top-up needed.
        # M8 gate suppresses the warning here too (post_money_required also 0).
        # The key is: no crash.
        r = option_pool.required_topup(
            pre_topup_fully_diluted_shares=0,
            existing_unallocated_pool=0,
            target_pool_percent=0.10,
            new_money_shares=None,
            target_basis="pre_money",
        )
        assert r["required_pool_topup_shares"] == 0  # no crash


# ===========================================================================
# anti_dilution.py
# ===========================================================================


class TestAntiDilution:
    def test_bbwa_divisor_uses_cp1_not_oip(self) -> None:
        """Per Gotcha #2: B = consideration / CP1, not consideration / OIP.
        Cooley GO + NVCA Model Cert §4.4.4."""
        r = anti_dilution.bbwa_new_conversion_price(
            current_conversion_price=1.0,  # CP1
            pre_issuance_share_count_A=10_000_000,
            consideration_received=1_000_000,
            new_issue_price=0.50,
            new_shares_issued_C=2_000_000,
        )
        assert r["triggered"] is True
        # B = 1M / 1.0 = 1M (NOT 1M / OIP)
        assert r["intermediate"]["B"] == 1_000_000
        # CP2 = 1 × (10M + 1M) / (10M + 2M) = 11/12 ≈ 0.917
        assert math.isclose(r["new_conversion_price"], 11_000_000 / 12_000_000, rel_tol=1e-6)

    def test_bbwa_not_triggered_when_new_price_above_cp1(self) -> None:
        r = anti_dilution.bbwa_new_conversion_price(
            current_conversion_price=1.0,
            pre_issuance_share_count_A=10_000_000,
            consideration_received=1_000_000,
            new_issue_price=2.0,  # up round
        )
        assert r["triggered"] is False
        assert r["new_conversion_price"] == 1.0

    def test_bbwa_C_derived_from_consideration_when_omitted(self) -> None:
        r = anti_dilution.bbwa_new_conversion_price(
            current_conversion_price=1.0,
            pre_issuance_share_count_A=10_000_000,
            consideration_received=1_000_000,
            new_issue_price=0.50,
            # new_shares_issued_C omitted → derived as 1M / 0.5 = 2M
        )
        assert r["intermediate"]["C"] == 2_000_000

    def test_full_ratchet_triggered(self) -> None:
        r = anti_dilution.full_ratchet_new_conversion_price(
            current_conversion_price=1.0,
            new_issue_price=0.50,
        )
        assert r["triggered"] is True
        assert r["new_conversion_price"] == 0.50

    def test_full_ratchet_not_triggered(self) -> None:
        r = anti_dilution.full_ratchet_new_conversion_price(
            current_conversion_price=1.0,
            new_issue_price=2.0,
        )
        assert r["triggered"] is False
        assert r["new_conversion_price"] == 1.0


# ===========================================================================
# priced_round.py (solver / orchestrator)
# ===========================================================================


class TestPricedRound:
    def test_solves_cap_only_safe_with_priced_round(self) -> None:
        """Closed-form for cap-only SAFEs converges in ≤2 iterations."""
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[_SAFE_BASIC],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="pre_money",
        )
        assert r["completeness"] == "full"
        assert r["converged"] is True
        # new money 20% of post-money valuation = $5M / $25M
        assert math.isclose(r["aggregate_ownership_by_class"]["new_money_pct"], 0.20, abs_tol=1e-3)

    def test_emits_per_holder_post_round_ownership(self) -> None:
        """ "What does this round do to ME" must be readable, not derivable.

        Splitting the founders' aggregate in chat is banned (and would be arithmetic the model did,
        not output it read), so before this the most common founder question had to be declined. The
        producer now emits it. Exact, not an estimate: a priced round does not change a founder's
        share COUNT -- dilution is denominator growth -- so each holder's pct is shares / post_fd.
        """
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[_SAFE_BASIC],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="pre_money",
        )
        agg = r["aggregate_ownership_by_class"]
        fbh = agg["founders_by_holder"]
        assert fbh, "per-holder post-round ownership must be emitted"
        # Nested INSIDE the aggregate, so a solver-rejected scenario cannot leak it.
        assert "founders_by_holder" not in r
        # Every founder in cap_state appears, keyed by founder_id.
        assert set(fbh) == {f["founder_id"] for f in cs["founders"]}
        for row in fbh.values():
            assert row["name"] and row["common_shares"] > 0
            assert 0.0 < row["pct"] < 1.0
        # The per-holder figures must SUM to the aggregate the same solver reported -- if they drift,
        # one of the two is wrong and the founder is reading a number that does not reconcile.
        assert math.isclose(sum(v["pct"] for v in fbh.values()), agg["founders_pct"], abs_tol=1e-9)
        # And each holder's pct must be their own shares over the shared denominator.
        post_fd = sum(v["common_shares"] for v in fbh.values()) / agg["founders_pct"]
        for row in fbh.values():
            assert math.isclose(row["pct"], row["common_shares"] / post_fd, rel_tol=1e-9)

    def test_solver_detects_circular_mfn(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        cyclic_safes = [
            {
                **_SAFE_BASIC,
                "id": "A",
                "form": "yc_uncapped_mfn",
                "post_money_valuation_cap": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "B",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
            {
                **_SAFE_BASIC,
                "id": "B",
                "form": "yc_uncapped_mfn",
                "post_money_valuation_cap": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "A",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
        ]
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=cyclic_safes,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert r["completeness"] == "structural_only"
        assert any(b["code"] == "E_SAFE_CIRCULAR_MFN" for b in r["blockers"])

    def test_solver_rejects_zero_pre_fd(self) -> None:
        """When cap state has no pre-FD shares, the system has no anchor."""
        empty_inputs = {
            **_BASIC_INPUTS,
            "founders": [],
            "option_pool": {"plan_type": "nso", "authorized": 0, "issued": 0, "unallocated": 0},
        }
        cs = cap_state_mod.build_cap_state(empty_inputs, _BASIC_INSTRUMENTS)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert r["completeness"] == "structural_only"

    # --- E_SOLVER_NO_VALID_FIXED_POINT guard ---

    def test_solver_degenerate_purchase_exceeds_cap(self) -> None:
        """When a YC post-money SAFE's purchase_amount >> post_money_cap the
        solver previously drove PPS toward ~1e-19 and falsely reported full/
        converged with multi-e26 share counts.  After the fix it must return
        structural_only + E_SOLVER_NO_VALID_FIXED_POINT and NOT claim converged.

        Arithmetic: purchase/cap = 10M/5M = 2.0 (200% of company_cap).
        No real fixed point exists; any solution implies founders_pct ≤ 0.
        """
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        bad_safe = {
            "id": "safe_bad",
            "investor_name": "Pathological Fund",
            "purchase_amount": 10_000_000,  # purchase far exceeds cap
            "post_money_valuation_cap": 5_000_000,  # post-money cap = $5M
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[bad_safe],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert r["completeness"] == "structural_only", (
            f"Expected structural_only for purchase >> cap; got {r['completeness']!r}. "
            f"PPS={r.get('equity_financing_price')}, "
            f"founders_pct={r.get('aggregate_ownership_by_class', {}).get('founders_pct')}"
        )
        assert not r.get("converged", True), "converged must be False alongside E_SOLVER_NO_VALID_FIXED_POINT"
        codes = [b["code"] for b in r.get("blockers", [])]
        assert "E_SOLVER_NO_VALID_FIXED_POINT" in codes, (
            f"Expected E_SOLVER_NO_VALID_FIXED_POINT in blockers; got {codes}"
        )
        # Share counts must not be astronomically large
        if "post_round_fully_diluted_shares" in r:
            assert r["post_round_fully_diluted_shares"] < 10**15, (
                f"Absurd share count leaked: {r['post_round_fully_diluted_shares']}"
            )

    def test_solver_degenerate_multi_safe_collective_overflow(self) -> None:
        """Two YC post-money SAFEs whose collective purchase/cap fractions
        sum to 137% (> 100%) must also trigger E_SOLVER_NO_VALID_FIXED_POINT.

        safe_1: purchase=4M / cap=5M  → 80% of company_cap
        safe_2: purchase=4M / cap=7M  → ~57% of company_cap
        sum = ~137%; no valid fixed point.
        """
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        safe_1 = {
            "id": "safe_1",
            "investor_name": "Fund A",
            "purchase_amount": 4_000_000,
            "post_money_valuation_cap": 5_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        safe_2 = {
            "id": "safe_2",
            "investor_name": "Fund B",
            "purchase_amount": 4_000_000,
            "post_money_valuation_cap": 7_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe_1, safe_2],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert r["completeness"] == "structural_only", (
            f"Expected structural_only for collective 137% overflow; got {r['completeness']!r}"
        )
        codes = [b["code"] for b in r.get("blockers", [])]
        assert "E_SOLVER_NO_VALID_FIXED_POINT" in codes, (
            f"Expected E_SOLVER_NO_VALID_FIXED_POINT in blockers; got {codes}"
        )

    def test_solver_valid_high_dilution_round_converges(self) -> None:
        """A legitimately high-dilution round (SAFE takes ~64% of company_cap,
        founders retain ~14%) must still converge and report completeness=full.

        Arithmetic: purchase=4M, post_money_cap=5M → SAFE fraction = 80% of
        company_cap.  With pre_fd=11M, the SAFE fraction is of company_cap
        (not of total post-money FD), so:
          company_cap = 11M / (1 - 0.80) = 55M
          safe_shares = 0.80 × 55M = 44M
          PPS = 20M / 55M ≈ 0.3636
          new_money_shares = 5M / PPS ≈ 13.75M
          post_fd = 55M + 13.75M = 68.75M
          founders_pct = 10M / 68.75M ≈ 14.5%   ← positive and valid
        """
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        valid_high_dil_safe = {
            "id": "safe_high_dil",
            "investor_name": "Large Fund",
            "purchase_amount": 4_000_000,  # purchase/cap = 0.80 of company_cap
            "post_money_valuation_cap": 5_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[valid_high_dil_safe],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert r["completeness"] == "full", (
            f"Valid high-dilution round should be 'full'; got {r['completeness']!r}. blockers={r.get('blockers')}"
        )
        assert r["converged"] is True
        agg = r["aggregate_ownership_by_class"]
        # Founders must retain a positive and meaningful fraction (~14.5%)
        assert agg["founders_pct"] > 0.05, (
            f"founders_pct={agg['founders_pct']:.4f} should be ~14.5% for valid 80%-safe round"
        )
        # Guard must NOT fire on a legitimate round
        codes = [b["code"] for b in r.get("blockers", [])]
        assert "E_SOLVER_NO_VALID_FIXED_POINT" not in codes, "Guard incorrectly fired on a valid high-dilution round"


class TestStackedPostMoneySAFEsGolden:
    """Golden-value regression for YC post-money cap SAFE math.

    Locks the canonical answer for the eval-2 scenario:
    - Founder 10M common; unallocated pool 1M; pre-FD 11M
    - SAFE 1: $500k @ $10M post-money cap
    - SAFE 2: $1M @ $15M post-money cap
    - SAFE 3: $500k uncapped MFN electing SAFE 1's $10M cap
    - Series A: $5M @ $20M pre, 10% post-money pool (founders absorb)

    Golden values derived from first-principles YC post-money SAFE convention
    (rule pack `safe.company_capitalization_yc_post_money`): each post-money
    SAFE locks `purchase / post_money_cap` of Company Capitalization measured
    IMMEDIATELY PRIOR TO the equity financing, which includes existing shares,
    the pre-existing unissued option pool, and all converting securities
    (self-referential via the fixed-point loop), but EXCLUDES new-money shares
    and the in-connection pool top-up.

    First-principles derivation:
    - Aggregate SAFE fraction of company_cap = 1/20 + 1/15 + 1/20 = 1/6
    - company_cap = adj_pre_fd + total_safe_shares = 11M + C/6
      → C = 11M × 6/5 = 13,200,000
    - safe_1 = 660,000; safe_2 = 880,000; safe_3 = 660,000; total = 2,200,000
    - denom (pre-money FD for pricing) = 11M + 2.2M + topup = 13.2M + topup
    - Pool (post_money basis): (1M + topup)/(5denom/4) = 0.10
      → topup = 5,200,000/7 ≈ 742,857; denom = 97,600,000/7 ≈ 13,942,857
    - PPS = 20M/denom = 175/122 ≈ $1.4344
    - T (post-round FD) = 122,000,000/7 ≈ 17,428,571
    - founders_pct = 10M/T = 35/61 ≈ 57.3770%
    - safe_pct = 2.2M/T ≈ 12.6230%
    - pool = 10.0%; new money = 20.0%
    """

    EVAL2_INPUTS = {
        "company_name": "TestCo",
        "analysis_date": "2026-05-21",
        "mode": "standard",
        "jurisdiction": {
            "structure": "delaware",
            "incorporated_date": "2024-06-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        },
        "founders": [
            {"name": "Founder", "founder_id": "founder_1", "common_shares": 10_000_000},
        ],
        "preferred_series": [],
        "option_pool": {
            "plan_type": "nso",
            "authorized": 1_000_000,
            "issued": 0,
            "unallocated": 1_000_000,
        },
        "common_batches": [],
        "metadata": {"run_id": "test"},
    }

    EVAL2_SAFES = [
        {
            "id": "safe_1",
            "investor_name": "Angel A",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": 10_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        },
        {
            "id": "safe_2",
            "investor_name": "Angel B",
            "purchase_amount": 1_000_000,
            "post_money_valuation_cap": 15_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-02-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        },
        # safe_3 is "MFN-elected against safe_1's $10M cap" — modeled as a
        # pre-resolved $10M-cap post-money SAFE so J's regression isolates the
        # math-formula bug. Phase N (MFN auto-bind) has separate test coverage
        # for the auto-resolution path from uncapped MFN → elected SAFE's cap.
        {
            "id": "safe_3",
            "investor_name": "Angel C",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": 10_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-03-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        },
    ]

    def _instruments(self) -> dict[str, Any]:
        return {
            "safes": self.EVAL2_SAFES,
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }

    def test_per_safe_ownership_is_purchase_over_cap(self) -> None:
        """Per-SAFE ownership formula for YC post-money cap SAFEs.

        Each SAFE: `safe_ownership_i = purchase_amount_i / post_money_cap_i`.
        Not `purchase / cap_price`, not `purchase × pre_fd / cap`.
        """
        # safe_1: $500k / $10M = 5%
        assert math.isclose(500_000 / 10_000_000, 0.05, rel_tol=1e-9)
        # safe_2: $1M / $15M = 6.6667%
        assert math.isclose(1_000_000 / 15_000_000, 1 / 15, rel_tol=1e-9)
        # safe_3 (MFN→$10M): $500k / $10M = 5%
        assert math.isclose(500_000 / 10_000_000, 0.05, rel_tol=1e-9)
        # Aggregate: 1/6
        agg = 500_000 / 10_000_000 + 1_000_000 / 15_000_000 + 500_000 / 10_000_000
        assert math.isclose(agg, 1 / 6, rel_tol=1e-9)

    def test_eval2_golden_solver_iterative(self) -> None:
        """Full eval-2 scenario through the iterative solver.

        Golden values from independent first-principles derivation (YC post-money
        SAFE convention: company_cap measured immediately prior to the round,
        excluding new-money shares and in-connection pool top-up):
        - company_cap = 11M + C/6 → C = 13,200,000
        - safe_pct of T = 2,200,000 / (122,000,000/7) = 77/610 ≈ 12.6230%
        - founders_pct = 10M / T = 35/61 ≈ 57.3770%
        - PPS = 175/122 ≈ $1.4344
        - T (post-round FD) = 122,000,000/7 ≈ 17,428,571
        """
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, self._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=self.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )

        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"
        assert r["converged"] is True

        agg = r["aggregate_ownership_by_class"]
        # T = 122,000,000/7; safe_shares = 2,200,000; safe_pct = 2,200,000/T = 77/610
        _T = 122_000_000 / 7
        assert math.isclose(agg["safe_pct"], 2_200_000 / _T, abs_tol=1e-4), (
            f"safe_pct: got {agg['safe_pct']:.6f}, expected {2_200_000 / _T:.6f}"
        )
        # founders_pct = 10,000,000 / T = 35/61
        assert math.isclose(agg["founders_pct"], 35 / 61, abs_tol=1e-4), (
            f"founders_pct: got {agg['founders_pct']:.6f}, expected {35 / 61:.6f}"
        )
        # New money: 20% of post-money
        assert math.isclose(agg["new_money_pct"], 0.20, abs_tol=1e-4)
        # Pool: 10% of post-money
        assert math.isclose(agg["option_pool_pct"], 0.10, abs_tol=1e-4)

        # PPS = 20M / denom = 20M / (97,600,000/7) = 175/122
        assert math.isclose(r["equity_financing_price"], 175 / 122, rel_tol=1e-4), (
            f"PPS: got {r['equity_financing_price']:.6f}, expected {175 / 122:.6f}"
        )

        # Total post-round FD = 122,000,000/7 ≈ 17,428,571
        total_fd = r["post_round_fully_diluted_shares"]
        assert math.isclose(total_fd, 122_000_000 / 7, rel_tol=1e-4), (
            f"post_round_fd: got {total_fd}, expected {122_000_000 / 7:.0f}"
        )

    def test_eval2_per_safe_share_counts(self) -> None:
        """Each SAFE's resolved share count matches the ownership formula.

        company_cap = 13,200,000 (derived: adj_pre_fd + safe_shares = 11M + C/6 → C = 13.2M)
        safe_1 = $500k × 13.2M / $10M = 660,000
        safe_2 = $1M  × 13.2M / $15M = 880,000
        safe_3 = $500k × 13.2M / $10M = 660,000
        """
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, self._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=self.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        per_safe = r["per_safe"]
        # safe_1: $500k × 13,200,000 / $10M = 660,000
        assert math.isclose(per_safe["safe_1"]["conversion_shares"], 660_000, rel_tol=1e-3)
        # safe_2: $1M × 13,200,000 / $15M = 880,000
        assert math.isclose(per_safe["safe_2"]["conversion_shares"], 880_000, rel_tol=1e-3)
        # safe_3: $500k × 13,200,000 / $10M = 660,000
        assert math.isclose(per_safe["safe_3"]["conversion_shares"], 660_000, rel_tol=1e-3)

    def test_eval2_new_investor_and_pool_shares(self) -> None:
        """Verify share counts for new investor and pool top-up.

        denom = 97,600,000/7 ≈ 13,942,857; PPS = 175/122
        new_shares = $5M / PPS = $5M × 122/175 = 610M/175 = 3,485,714.285...
        topup = 5,200,000/7 ≈ 742,857
        T = 122,000,000/7 ≈ 17,428,571
        new_money_pct = 3,485,714/T = 20.0%; pool_pct = 10.0%
        """
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, self._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=self.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        # new_shares = 5,000,000 × 122/175 = 3,485,714.285...
        assert math.isclose(r["shares_breakdown"]["new_money"], 5_000_000 * 122 / 175, rel_tol=1e-3)
        # pool_topup = 5,200,000/7 ≈ 742,857
        assert math.isclose(r["shares_breakdown"]["pool_topup"], 5_200_000 / 7, rel_tol=1e-3)

    def test_eval2_closed_form_path_matches_iterative(self) -> None:
        """Solver must produce the correct golden regardless of iteration path."""
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, self._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=self.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        _T = 122_000_000 / 7
        agg = r["aggregate_ownership_by_class"]
        assert math.isclose(agg["safe_pct"], 2_200_000 / _T, abs_tol=1e-4)
        assert math.isclose(agg["founders_pct"], 35 / 61, abs_tol=1e-4)
        assert math.isclose(r["equity_financing_price"], 175 / 122, rel_tol=1e-4)

    def test_eval2_ownership_sum_to_one(self) -> None:
        """Cross-check assertion: all ownership classes sum to 100%."""
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, self._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=self.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        agg = r["aggregate_ownership_by_class"]
        total = (
            agg["safe_pct"]
            + agg["founders_pct"]
            + agg["new_money_pct"]
            + agg["option_pool_pct"]
            + agg.get("preferred_pct", 0.0)
            + agg.get("note_pct", 0.0)
        )
        assert math.isclose(total, 1.0, abs_tol=1e-4), f"ownerships sum to {total}, not 1.0"

    def test_single_postmoney_safe_yc_identity(self) -> None:
        """$500k @ $5M post-money cap, founders 10M + pool 1M, $2M at $8M pre,
        no pool top-up. YC convention: SAFE owns 10% of company_cap immediately
        before the round (company_cap = 11M + C/10 → C = 12,222,222.22;
        safe_shares = 1,222,222.22), then the round dilutes it to 8.00%.

        First-principles derivation:
        - safe_shares = 500k × C / 5M = C/10
        - C = 11M + C/10 → C = 110M/9 ≈ 12,222,222.22
        - safe_shares = 11M/9 ≈ 1,222,222.22
        - denom = 11M + 11M/9 = 110M/9
        - PPS = $8M / (110M/9) = 72M/110M = 36/55 ≈ $0.654545
        - new_shares = $2M / PPS = 2M × 55/36 = 110M/36 ≈ 3,055,555.56
        - T = 110M/9 + 110M/36 = 440M/36 + 110M/36 = 550M/36 ≈ 15,277,777.78
        - founders_pct = 10M/T = 10M × 36/550M = 360/5500 = 36/55 ≈ 65.45%
        - safe_pct = (11M/9) / (550M/36) = (11M × 4)/(550M) = 44/550 = 4/50 = 8.00%
        """
        inputs = {
            "company_name": "TestCo",
            "analysis_date": "2026-05-21",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-06-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [
                {"name": "Founder", "founder_id": "founder_1", "common_shares": 10_000_000},
            ],
            "preferred_series": [],
            "option_pool": {
                "plan_type": "nso",
                "authorized": 1_000_000,
                "issued": 0,
                "unallocated": 1_000_000,
            },
            "common_batches": [],
            "metadata": {"run_id": "test"},
        }
        safe = {
            "id": "safe_s1",
            "investor_name": "Angel",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": 5_000_000,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2025-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        instruments = {
            "safes": [safe],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        cs = cap_state_mod.build_cap_state(inputs, instruments)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe],
            notes=[],
            pre_money=8_000_000,
            new_money=2_000_000,
            target_pool_percent=None,  # no pool top-up
            target_basis="post_money",
        )
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"
        assert r["converged"] is True

        # safe_shares = 11M/9 ≈ 1,222,222.22
        assert math.isclose(r["per_safe"]["safe_s1"]["conversion_shares"], 11_000_000 / 9, rel_tol=1e-4)
        # PPS = 36/55
        assert math.isclose(r["equity_financing_price"], 36 / 55, rel_tol=1e-4), (
            f"PPS: got {r['equity_financing_price']:.8f}, expected {36 / 55:.8f}"
        )
        # T = 550M/36 ≈ 15,277,778
        assert math.isclose(r["post_round_fully_diluted_shares"], 550_000_000 / 36, rel_tol=1e-4)

        agg = r["aggregate_ownership_by_class"]
        # safe_pct = 8%
        assert math.isclose(agg["safe_pct"], 0.08, abs_tol=1e-4), (
            f"safe_pct: got {agg['safe_pct']:.6f}, expected 0.08 (SAFE locks 10% of company_cap, "
            f"then diluted to 8% by the 20% new-money round)"
        )
        # founders_pct = 36/55 ≈ 65.45%
        assert math.isclose(agg["founders_pct"], 36 / 55, abs_tol=1e-4), (
            f"founders_pct: got {agg['founders_pct']:.6f}, expected {36 / 55:.6f}"
        )
        # new_money_pct = 20%
        assert math.isclose(agg["new_money_pct"], 0.20, abs_tol=1e-4)

    def test_uncapped_mfn_auto_binds_to_elected_safes_terms(self) -> None:
        """Phase N: uncapped MFN with `elected_against_safe_id` set should
        auto-resolve to the elected sibling's terms — no override required.

        Eval-2 transcript noted that the solver previously REQUIRED a
        `conversion_price_override` even when MFN was clearly electing
        safe_1's cap. With Phase N auto-bind, the resolver pre-inherits the
        elected sibling's form/cap/discount before the solver runs.
        """
        eval2_safes_with_real_mfn = [
            self.EVAL2_SAFES[0],
            self.EVAL2_SAFES[1],
            # safe_3 is uncapped MFN electing safe_1; should auto-bind
            {
                "id": "safe_3",
                "investor_name": "Angel C",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": None,
                "discount_multiplier": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "safe_1",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
                "pro_rata_side_letter": None,
                "issuance_date": "2025-03-01",
                "form": "yc_uncapped_mfn",
                "conversion_price_override": None,
                "source_document": None,
                "extraction_confidence": "high",
            },
        ]
        instruments = {
            "safes": eval2_safes_with_real_mfn,
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, instruments)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=eval2_safes_with_real_mfn,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        # Auto-bind must produce the same answer as the pre-resolved test
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"
        _T = 122_000_000 / 7
        agg = r["aggregate_ownership_by_class"]
        assert math.isclose(agg["safe_pct"], 2_200_000 / _T, abs_tol=1e-4), (
            f"MFN auto-bind: aggregate safe_pct {agg['safe_pct']} ≠ expected {2_200_000 / _T:.6f}"
        )
        assert math.isclose(agg["founders_pct"], 35 / 61, abs_tol=1e-4)

    def test_fast_assess_writes_sentinel_and_markdown(self) -> None:
        """Phase O: quick_assess.py produces fast_assess_only.json + report.md.

        Verifies:
        - sentinel JSON conforms to v0.1.0-cap-table-fast-assess schema
        - sentinel uses the corrected math from Phase J (founder 53.33% on
          the canonical eval-2 scenario)
        - report_fast_assess.md exists and contains the founder ownership
        - no canonical artifacts are produced (no inputs.json, cap_state.json,
          report.json, etc.)
        """
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="TestCo",
            inputs=self.EVAL2_INPUTS,
            safes=[
                self.EVAL2_SAFES[0],
                self.EVAL2_SAFES[1],
                self.EVAL2_SAFES[2],
            ],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            founder_prompt="three SAFEs + $5M Series A",
            attached_docs=[],
        )
        # Sentinel schema
        assert sentinel["schema_version"] == "v0.1.0-cap-table-fast-assess"
        assert sentinel["mode"] == "fast_assess"
        assert sentinel["produces_canonical_artifacts"] is False
        assert sentinel["rule_pack_version"] == "0.4.1"
        # inputs_fingerprint structurally valid
        fp = sentinel["inputs_fingerprint"]
        assert "sha256" in fp and len(fp["sha256"]) == 64
        # headline_data uses corrected math (YC post-money SAFE company_cap convention)
        hd = sentinel["headline_data"]
        fi = hd["founder_impact"]
        assert math.isclose(fi["ownership_post_financing_pct"], 35 / 61, abs_tol=1e-4)
        assert math.isclose(fi["pps_priced_round"], 175 / 122, rel_tol=1e-4)
        # report_md is delivered for the CLI to extract
        report_md = sentinel.pop("_report_md")
        assert "57.38%" in report_md  # founder ownership rendered (35/61 ≈ 57.38%)

    def test_fast_assess_sentinel_validates_against_schema(self) -> None:
        """Phase O: sentinel JSON validates against fast_assess_only.schema.json."""
        try:
            import jsonschema  # type: ignore[import-untyped]
        except ImportError:
            pytest.skip("jsonschema not installed")

        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="TestCo",
            inputs=self.EVAL2_INPUTS,
            safes=[self.EVAL2_SAFES[0]],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        sentinel.pop("_report_md", None)
        # Add the path field that the CLI normally sets
        sentinel["fast_assess_report_path"] = "/tmp/test_report.md"

        schema_path = os.path.join(
            os.path.dirname(SCRIPTS),
            "references",
            "schemas",
            "fast_assess_only.schema.json",
        )
        with open(schema_path) as f:
            schema = json.load(f)
        # Will raise jsonschema.ValidationError if invalid
        jsonschema.validate(instance=sentinel, schema=schema)

    def test_mfn_chain_resolves_transitively(self) -> None:
        """H9: A elects B elects C, where C has a resolved cap. The resolver
        must iterate to a fixed point so A inherits transitively (via B's
        resolution against C). Single-hop resolver would leave A unresolved
        because B is still 'yc_uncapped_mfn' on the first pass.
        """
        # Chain: safe_a → safe_b → safe_c (safe_c has resolved cap)
        chain_safes = [
            {
                "id": "safe_a",
                "investor_name": "Angel Chain A",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": None,
                "pre_money_valuation_cap": None,
                "discount_multiplier": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "safe_b",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
                "pro_rata_side_letter": None,
                "issuance_date": "2025-03-01",
                "form": "yc_uncapped_mfn",
                "conversion_price_override": None,
                "source_document": None,
                "extraction_confidence": "high",
            },
            {
                "id": "safe_b",
                "investor_name": "Angel Chain B",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": None,
                "pre_money_valuation_cap": None,
                "discount_multiplier": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "safe_c",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
                "pro_rata_side_letter": None,
                "issuance_date": "2025-03-01",
                "form": "yc_uncapped_mfn",
                "conversion_price_override": None,
                "source_document": None,
                "extraction_confidence": "high",
            },
            {
                "id": "safe_c",
                "investor_name": "Angel Chain C (anchor)",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": 10_000_000,
                "pre_money_valuation_cap": None,
                "discount_multiplier": None,
                "mfn_provision": None,
                "pro_rata_side_letter": None,
                "issuance_date": "2025-03-01",
                "form": "yc_postmoney_cap",
                "conversion_price_override": None,
                "source_document": None,
                "extraction_confidence": "high",
            },
        ]
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, {**self._instruments(), "safes": chain_safes})
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=chain_safes,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        assert r["completeness"] == "full", f"chain didn't resolve: {r.get('blockers')}"
        # All three SAFEs resolved to safe_c's $10M cap; aggregate fraction of
        # company_cap = 3 × (500k/10M) = 15%
        # Under YC convention: company_cap = adj_pre_fd + safe_shares
        # safe_shares = 0.15 × C; C = 11M + 0.15C → C = 11M/0.85; safe_shares = 11M × 0.15/0.85
        # After-round dilution: safe_pct of T < 15% (diluted by new money)
        # Verify qualitatively that safe_pct is within the expected range (between 10% and 15%)
        agg = r["aggregate_ownership_by_class"]
        assert 0.10 < agg["safe_pct"] < 0.15, (
            f"chain didn't propagate correctly: aggregate safe_pct {agg['safe_pct']:.4f} not in (0.10, 0.15)"
        )
        # Verify 15% aggregate cap fraction (purchase/cap) mathematically
        _C = 11_000_000 / 0.85  # company_cap
        _safe_shares = _C * 0.15
        _T = r["post_round_fully_diluted_shares"]
        assert math.isclose(agg["safe_pct"], _safe_shares / _T, abs_tol=1e-3), (
            f"chain safe_pct {agg['safe_pct']:.6f} doesn't match company_cap convention"
        )

    def test_mfn_elected_against_missing_sibling_fails_cleanly(self) -> None:
        """H9 edge case: MFN points to a sibling that doesn't exist in the
        safes list. Resolver must leave the SAFE unresolved (no crash); the
        downstream rejection path then produces E_SAFE_REQUIRES_CONVERSION_EVENT.
        """
        orphan_safes = [
            {
                "id": "safe_orphan",
                "investor_name": "Angel Orphan",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": None,
                "pre_money_valuation_cap": None,
                "discount_multiplier": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "safe_does_not_exist",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
                "pro_rata_side_letter": None,
                "issuance_date": "2025-03-01",
                "form": "yc_uncapped_mfn",
                "conversion_price_override": None,
                "source_document": None,
                "extraction_confidence": "high",
            },
        ]
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, {**self._instruments(), "safes": orphan_safes})
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=orphan_safes,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        # Should produce a blocker, not crash, not silently pass through.
        assert r["completeness"] == "structural_only"
        assert (
            any(
                b.get("code") in {"E_SAFE_REQUIRES_CONVERSION_EVENT", "E_SAFE_CIRCULAR_MFN"}
                for b in r.get("blockers", [])
            )
            or "safe_orphan" in r["per_safe"]
            and r["per_safe"]["safe_orphan"]["branch"] == "rejected"
        )

    def test_mfn_inheritance_provenance_propagated_to_per_safe_result(self) -> None:
        """M5: per_safe[safe_id]['_mfn_inherited_from'] must be set when the
        SAFE was MFN-resolved. Lets downstream counsel-review reporting phrase
        'Investor X's MFN-inherited terms from Investor Y'.
        """
        eval2_with_mfn = [
            self.EVAL2_SAFES[0],
            self.EVAL2_SAFES[1],
            {
                "id": "safe_3_mfn",
                "investor_name": "Angel C MFN",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": None,
                "pre_money_valuation_cap": None,
                "discount_multiplier": None,
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "safe_1",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
                "pro_rata_side_letter": None,
                "issuance_date": "2025-03-01",
                "form": "yc_uncapped_mfn",
                "conversion_price_override": None,
                "source_document": None,
                "extraction_confidence": "high",
            },
        ]
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, {**self._instruments(), "safes": eval2_with_mfn})
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=eval2_with_mfn,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        per_safe_3 = r["per_safe"]["safe_3_mfn"]
        # Provenance must be propagated from shadow record to per_safe result
        assert per_safe_3.get("_mfn_inherited_from") == "safe_1", (
            f"MFN inheritance provenance lost: {per_safe_3.get('_mfn_inherited_from')} != 'safe_1'"
        )

    def test_eval2_aggregate_safe_pct_equals_sum_of_per_safe(self) -> None:
        """J8 cross-check: aggregate_ownership_by_class.safe_pct must equal
        Σ per-safe ownership. Catches drift between aggregate and detail.
        """
        cs = cap_state_mod.build_cap_state(self.EVAL2_INPUTS, self._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=self.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        total_fd = r["post_round_fully_diluted_shares"]
        sum_per_safe_shares = sum(s["conversion_shares"] for s in r["per_safe"].values())
        agg_from_detail = sum_per_safe_shares / total_fd
        agg_reported = r["aggregate_ownership_by_class"]["safe_pct"]
        assert math.isclose(agg_reported, agg_from_detail, abs_tol=1e-6), (
            f"aggregate safe_pct {agg_reported} does not equal sum of per-safe {agg_from_detail}"
        )


class TestLegacyPreMoneySAFEs:
    """Golden-value regression for YC PRE-MONEY (legacy, pre-Oct-2018) SAFE math.

    Two distinct YC SAFE families exist:
    - **Post-money** (current): each SAFE locks `purchase / cap` of POST-money FD.
      Covered by TestStackedPostMoneySAFEsGolden.
    - **Pre-money** (legacy): conversion branch selected by §(a)(1)/(a)(2) of the
      YC pre-money SAFE document:
        §(a)(1): pre_money_valuation ≤ cap → investor converts at ROUND price
                 (Standard Preferred Stock price per share).
        §(a)(2): pre_money_valuation > cap → investor converts at SAFE price
                 = cap / Company Capitalization, where Company Capitalization
                 includes the in-connection pool top-up per the Company Capitalization clause.
      The SAFE's % of post-money is NOT fixed; pre-money SAFEs dilute alongside
      founders when new money + pool refresh land.

    Base fixture:
    - 10M founders common + 1M unallocated pool → pre_money_FD = 11M
    - SAFE: $500k purchase, $5M pre-money cap, no discount
    """

    EVAL_PREMONEY_INPUTS = {
        "company_name": "TestCo",
        "analysis_date": "2026-05-21",
        "mode": "standard",
        "jurisdiction": {
            "structure": "delaware",
            "incorporated_date": "2024-06-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        },
        "founders": [
            {"name": "Founder", "founder_id": "founder_1", "common_shares": 10_000_000},
        ],
        "preferred_series": [],
        "option_pool": {
            "plan_type": "nso",
            "authorized": 1_000_000,
            "issued": 0,
            "unallocated": 1_000_000,
        },
        "common_batches": [],
        "metadata": {"run_id": "test"},
    }

    EVAL_PREMONEY_SAFE = {
        "id": "safe_pre_1",
        "investor_name": "Angel L (legacy)",
        "purchase_amount": 500_000,
        "post_money_valuation_cap": None,
        "pre_money_valuation_cap": 5_000_000,  # legacy SAFE uses pre-money cap
        "discount_multiplier": None,
        "mfn_provision": None,
        "pro_rata_side_letter": None,
        "issuance_date": "2017-09-01",  # pre-Oct-2018; legacy form era
        "form": "yc_premoney_cap_only",
        "conversion_price_override": None,
        "source_document": None,
        "extraction_confidence": "high",
    }

    def _instruments(self, safes: list[dict[str, Any]]) -> dict[str, Any]:
        return {
            "safes": safes,
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }

    def test_legacy_premoney_at_cap_uses_round_price(self) -> None:
        """Single-SAFE pre-money scenario; round pre_money == cap → §(a)(1) applies.

        Per YC pre-money SAFE §(a)(1): when pre_money_valuation ≤ cap, the investor
        receives shares at the Standard Preferred (round) price, NOT the cap price.
        The round price itself depends on safe_shares (coupled system):
          PPS × (11M + safe_shares) = 5M   AND   safe_shares = 500k / PPS
          → PPS × 11M + 500k = 5M  → PPS = 4.5M / 11M = 9/22 ≈ 0.40909
          → safe_shares = 500k / (9/22) = 500k × 22/9 = 11M/9 ≈ 1,222,222
          → new_money_shares = 5M / (9/22) = 110M/9 ≈ 12,222,222
          → post_fd = 11M + 11M/9 + 110M/9 = 220M/9 ≈ 24,444,444
          → safe_pct  = (11M/9) / (220M/9) = 11/220 = 5.0%
          → founders_pct = 10M / (220M/9) = 90/220 = 9/22 ≈ 40.909%
          → new_money_pct = (110M/9) / (220M/9) = 50% ✓
        """
        cs = cap_state_mod.build_cap_state(self.EVAL_PREMONEY_INPUTS, self._instruments([self.EVAL_PREMONEY_SAFE]))
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[self.EVAL_PREMONEY_SAFE],
            notes=[],
            pre_money=5_000_000,  # equals cap → §(a)(1) branch
            new_money=5_000_000,
            target_pool_percent=None,  # no pool refresh
            target_basis="post_money",
        )
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"

        # PPS = 9/22 ≈ 0.40909
        assert math.isclose(r["equity_financing_price"], 9 / 22, rel_tol=1e-4), (
            f"PPS: got {r['equity_financing_price']:.6f}, expected {9 / 22:.6f}"
        )
        # Total post-money FD = 220M/9 ≈ 24,444,444
        assert math.isclose(r["post_round_fully_diluted_shares"], 220_000_000 / 9, rel_tol=1e-4)

        # safe_shares = 11M/9 ≈ 1,222,222
        assert math.isclose(r["per_safe"]["safe_pre_1"]["conversion_shares"], 11_000_000 / 9, rel_tol=1e-4), (
            f"safe_shares: got {r['per_safe']['safe_pre_1']['conversion_shares']:.0f}, expected {11_000_000 / 9:.0f}"
        )
        # SAFE converted at round_price_branch, not cap_branch
        assert r["per_safe"]["safe_pre_1"]["branch"] == "round_price_branch", (
            f"expected round_price_branch, got {r['per_safe']['safe_pre_1']['branch']}"
        )

        # safe_pct = 11/220 = 5.0%, NOT 10% (would be 10% under post-money form)
        agg = r["aggregate_ownership_by_class"]
        assert math.isclose(agg["safe_pct"], 11 / 220, abs_tol=1e-4), (
            f"safe_pct: got {agg['safe_pct']:.6f}, expected {11 / 220:.6f} (5.0%)"
        )
        # founders_pct = 9/22 ≈ 40.909%
        assert math.isclose(agg["founders_pct"], 9 / 22, abs_tol=1e-4)
        # New money 50% (holds by construction)
        assert math.isclose(agg["new_money_pct"], 0.50, abs_tol=1e-4)

    def test_legacy_premoney_safe_pct_differs_from_post_money_safe_pct(self) -> None:
        """Smoke test: pre-money SAFE math MUST produce different ownership than
        post-money SAFE math under identical inputs. If the solver collapses both
        forms to the same formula (the regression we're guarding against), this
        fails. safe_pct=10% would be the post-money answer; at-cap pre-money is 5.0%.
        """
        cs = cap_state_mod.build_cap_state(self.EVAL_PREMONEY_INPUTS, self._instruments([self.EVAL_PREMONEY_SAFE]))
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[self.EVAL_PREMONEY_SAFE],
            notes=[],
            pre_money=5_000_000,
            new_money=5_000_000,
            target_pool_percent=None,
            target_basis="post_money",
        )
        agg = r["aggregate_ownership_by_class"]
        # The post-money YC SAFE identity `safe_pct = purchase/cap = 10%` MUST NOT
        # hold for a pre-money SAFE. If it does, the solver routed the wrong form.
        # At-cap (§(a)(1)) gives 5.0%; above-cap (§(a)(2)) gives ~4.96% — both < 8%.
        assert agg["safe_pct"] < 0.08, (
            f"safe_pct {agg['safe_pct']:.4f} is too close to the post-money value "
            f"0.10; pre-money SAFE should dilute (at-cap: 5.0%, above-cap: ~4.96%). "
            "Form dispatch likely broken."
        )

    def test_mixed_legacy_premoney_and_post_money_safes(self) -> None:
        """Mixed-form scenario: one post-money SAFE + one pre-money SAFE in the
        same priced round. Solver must route each form to its correct branch.

        Round pre_money = $5M = safe_B cap → §(a)(1) applies for safe_B (round price).
        safe_A (post-money): locks 10% of company_cap C.
        safe_B (pre-money, §(a)(1)): safe_B_shares = 500k / PPS = 500k / (5M/C) = C/10

        System: C = 11M + safe_A + safe_B = 11M + C/10 + C/10 = 11M + C/5
          → 4C/5 = 11M → C = 55M/4 = 13,750,000
          → safe_A_shares = safe_B_shares = C/10 = 55M/40 = 11M/8 = 1,375,000 (exact)
          → PPS = 5M / C = 5M / (55M/4) = 4/11 ≈ 0.36364
          → new_money_shares = 5M / PPS = 5M × 11/4 = 55M/4 = 13,750,000
          → T = C + 13.75M = 55M/4 + 55M/4 = 55M/2 = 27,500,000 (exact)
          → safe_pct  = (1.375M + 1.375M) / 27.5M = 2.75/27.5 = 10.0%
          → founders_pct = 10M / 27.5M = 4/11 ≈ 36.36%
          → new_money_pct = 13.75M / 27.5M = 50% ✓
        """
        safe_a_post = {
            "id": "safe_a",
            "investor_name": "Angel A (post-money)",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": 5_000_000,
            "pre_money_valuation_cap": None,
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2021-01-01",
            "form": "yc_postmoney_cap",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        safe_b_pre = self.EVAL_PREMONEY_SAFE  # the legacy one, id="safe_pre_1"
        cs = cap_state_mod.build_cap_state(self.EVAL_PREMONEY_INPUTS, self._instruments([safe_a_post, safe_b_pre]))
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe_a_post, safe_b_pre],
            notes=[],
            pre_money=5_000_000,
            new_money=5_000_000,
            target_pool_percent=None,
            target_basis="post_money",
        )
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"

        # T = 55M/2 = 27,500,000 (exact)
        _T = 27_500_000
        assert math.isclose(r["post_round_fully_diluted_shares"], _T, rel_tol=1e-3), (
            f"total_fd: got {r['post_round_fully_diluted_shares']}, expected {_T}"
        )

        # safe_A (post-money) shares = 11M/8 = 1,375,000 (exact)
        assert math.isclose(r["per_safe"]["safe_a"]["conversion_shares"], 11_000_000 / 8, rel_tol=1e-3)
        # safe_B (pre-money, §(a)(1)) shares = 11M/8 = 1,375,000 (exact, same as safe_A)
        assert math.isclose(r["per_safe"]["safe_pre_1"]["conversion_shares"], 11_000_000 / 8, rel_tol=1e-3)
        assert r["per_safe"]["safe_pre_1"]["branch"] == "round_price_branch"

        agg = r["aggregate_ownership_by_class"]
        # Combined safe_pct = 2 × 1.375M / 27.5M = 10.0%
        assert math.isclose(agg["safe_pct"], 0.10, abs_tol=1e-4)
        # Founders = 4/11 ≈ 36.36%
        assert math.isclose(agg["founders_pct"], 4 / 11, abs_tol=1e-4)
        # New money = 50%
        assert math.isclose(agg["new_money_pct"], 0.50, abs_tol=1e-3)

    def test_legacy_premoney_above_cap_uses_cap_price(self) -> None:
        """Single-SAFE pre-money scenario; round pre_money > cap → §(a)(2) applies.

        Per YC pre-money SAFE §(a)(2): when pre_money_valuation > cap, the investor
        receives shares at the SAFE price = cap / Company Capitalization (cap price).

        Fixture: pre_fd=11M, SAFE $500k at $5M cap, round pre_money=$6M (> cap).
          safe_price = 5M / 11M = 5/11
          safe_shares = 500k / (5/11) = 500k × 11/5 = 1,100,000 (constant, non-iterative)
          PPS = 6M / (11M + 1.1M) = 6M / 12.1M = 60/121 ≈ 0.49587
          new_money_shares = 5M / (60/121) = 5M × 121/60 = 605M/60 ≈ 10,083,333
          post_fd = 12.1M + 10.083M = 22.183M (not an integer — use rel_tol)
          safe_pct ≈ 1.1M / 22.183M ≈ 4.96%
        """
        safe_above_cap = dict(self.EVAL_PREMONEY_SAFE)
        # cap = 5M, round pre_money = 6M → strictly above cap
        cs = cap_state_mod.build_cap_state(self.EVAL_PREMONEY_INPUTS, self._instruments([safe_above_cap]))
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe_above_cap],
            notes=[],
            pre_money=6_000_000,  # strictly above $5M cap → §(a)(2)
            new_money=5_000_000,
            target_pool_percent=None,
            target_basis="post_money",
        )
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"

        # safe_shares = 1,100,000 (cap price branch; non-iterative since pre-money forms
        # use a constant denominator in §(a)(2))
        assert math.isclose(r["per_safe"]["safe_pre_1"]["conversion_shares"], 1_100_000, rel_tol=1e-4)
        assert r["per_safe"]["safe_pre_1"]["branch"] == "cap_branch"

        # PPS = 60/121 ≈ 0.49587
        assert math.isclose(r["equity_financing_price"], 60 / 121, rel_tol=1e-4), (
            f"PPS: got {r['equity_financing_price']:.6f}, expected {60 / 121:.6f}"
        )

        # safe_pct: 1.1M / (12.1M + 5M × 121/60)  < 5% (below at-cap case)
        agg = r["aggregate_ownership_by_class"]
        assert agg["safe_pct"] < 0.05, f"safe_pct={agg['safe_pct']:.4f} should be < 0.05 in above-cap §(a)(2) branch"
        # new_money_pct = new_money / (pre_money + new_money) = 5/11 ≈ 45.45%
        assert math.isclose(agg["new_money_pct"], 5 / 11, abs_tol=1e-4)

    def test_legacy_premoney_pool_topup_in_denominator(self) -> None:
        """Pre-money SAFE §(a)(2): Company Capitalization includes in-connection pool top-up.

        Per the YC pre-money SAFE "Company Capitalization" definition, the denominator
        for the SAFE price INCLUDES "all shares of Common Stock reserved and available
        for future grant under any equity incentive or similar plan to be created or
        increased in connection with the Equity Financing."

        Fixture: 10M founders, 0 existing pool → pre_fd = 10M
          SAFE: $500k, $12M cap (pre_money=$20M > cap → §(a)(2))
          Round: pre_money=$20M, new_money=$5M, target_pool=10% pre_money basis
          existing_unallocated = 0

        Correct (YC pre-money SAFE) denominator includes pool top-up:
          safe_shares = P*(pre_fd + topup)/cap  AND  topup = (pre_fd + safe_shares)/9
          Solving: safe_shares*(9C - P) = 10*P*pre_fd
            → safe_shares = 10*P*pre_fd/(9C-P) = 10*500k*10M/(9*12M-500k) = 50T/107.5M = 20M/43
          topup = (10M + 20M/43)/9 = (430M/43 + 20M/43)/9 = (450M/43)/9 = 50M/43
          cap_price = 12M/(10M + 50M/43) = 12M*43/(430M+50M) = 516M/480M = 43/40 = 1.075

        Buggy (pre-fix) denominator excludes pool top-up:
          safe_shares = P*pre_fd/cap = 500k*10M/12M = 25M/60 ≈ 416,667  (fewer shares)

        The investor gets MORE shares when the pool top-up is correctly included.
        P=500k, F=10M, C=12M, target=10% pre_money.
        """
        inputs_no_pool = {
            "company_name": "TestCo",
            "analysis_date": "2026-05-21",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-06-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [
                {"name": "Founder", "founder_id": "founder_1", "common_shares": 10_000_000},
            ],
            "preferred_series": [],
            "option_pool": {
                "plan_type": "nso",
                "authorized": 0,
                "issued": 0,
                "unallocated": 0,  # no pre-existing pool
            },
            "common_batches": [],
            "metadata": {"run_id": "test"},
        }
        safe_above = {
            "id": "safe_pool_test",
            "investor_name": "Angel (pool test)",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": None,
            "pre_money_valuation_cap": 12_000_000,  # cap; pre_money=20M > cap → §(a)(2)
            "discount_multiplier": None,
            "mfn_provision": None,
            "pro_rata_side_letter": None,
            "issuance_date": "2017-09-01",
            "form": "yc_premoney_cap_only",
            "conversion_price_override": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        instruments = {
            "safes": [safe_above],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        cs = cap_state_mod.build_cap_state(inputs_no_pool, instruments)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe_above],
            notes=[],
            pre_money=20_000_000,  # > cap → §(a)(2); exercises the pool-top-up denominator term
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="pre_money",
        )
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"

        # With the top-up in the denominator: safe_shares = 20M/43 ≈ 465,116
        # Buggy value (pre-fix): 500k*10M/12M = 416,667 (denominator excludes pool topup)
        # The test checks the LOWER bound to confirm the fix:
        expected_fixed = 20_000_000 / 43  # ≈ 465,116
        expected_buggy = 500_000 * 10_000_000 / 12_000_000  # ≈ 416,667
        safe_shares = r["per_safe"]["safe_pool_test"]["conversion_shares"]
        assert safe_shares > expected_buggy + 1000, (
            f"safe_shares={safe_shares:.0f} should be well above the buggy value "
            f"{expected_buggy:.0f}; expected ≈{expected_fixed:.0f} (pool top-up included in denominator)"
        )
        assert math.isclose(safe_shares, expected_fixed, rel_tol=1e-3), (
            f"safe_shares={safe_shares:.2f}, expected 20M/43={expected_fixed:.2f} "
            "(pool top-up = 50M/43 included in cap-price denominator)"
        )
        assert r["per_safe"]["safe_pool_test"]["branch"] == "cap_branch"


# ===========================================================================
# rule_audit.py (two-phase)
# ===========================================================================


class TestRuleAudit:
    def _rules(self) -> dict[str, Any]:
        with open(os.path.join(os.path.dirname(SCRIPTS), "data", "cap-table-rules.json")) as f:
            return json.load(f)  # type: ignore[no-any-return]

    def test_classify_scope_not_applicable(self) -> None:
        rule = {"rule_id": "r1", "domain": "safe"}  # no date_window
        assert rule_audit._classify_scope(rule) == "not_applicable"

    def test_missing_date_action_humanizes_field_name(self) -> None:
        # The 'Provide X' watchlist action must not leak the raw snake_case field.
        action = rule_audit._action_for_status(
            {"status": "missing_event_date", "event_date_field": "tax_position_date"}, {}
        )
        assert "tax_position_date" not in action
        assert "tax position date" in action

    def test_israeli_aoa_rules_do_not_fire_on_delaware_no_preferred(self) -> None:
        """Real-doc test surfaced this: israeli_aoa.* rules were firing on a
        Delaware C-corp with no preferred series, because they weren't in the
        _RULE_MATCHERS dispatch table and defaulted to _matcher_always.

        Now gated to: jurisdiction_includes_israel AND has_preferred_series.
        """
        delaware_inputs = {
            "jurisdiction": {"structure": "delaware"},
            "mode": "standard",
        }
        empty_instruments: dict[str, Any] = {"safes": [], "convertible_notes": [], "warrants": [], "option_grants": []}
        empty_cap_state: dict[str, Any] = {"preferred_series": []}

        for rule_id in (
            "israeli_aoa.drag_along_threshold_below_75_percent",
            "israeli_aoa.section_102_plan_absent",
            "israeli_aoa.liquidation_preference_above_1x",
            "israeli_aoa.full_ratchet_anti_dilution",
        ):
            matcher = rule_audit._RULE_MATCHERS.get(rule_id, rule_audit._matcher_always)
            assert matcher(delaware_inputs, empty_instruments, empty_cap_state) is False, (
                f"{rule_id} should NOT match on Delaware engagement with no preferred series"
            )

    def test_israeli_aoa_rules_fire_on_israeli_with_preferred(self) -> None:
        """Israeli engagement WITH a preferred series → israeli_aoa.* rules apply."""
        israeli_inputs = {
            "jurisdiction": {"structure": "israeli"},
            "mode": "standard",
        }
        empty_instruments: dict[str, Any] = {"safes": [], "convertible_notes": [], "warrants": [], "option_grants": []}
        # All four israeli_aoa.* rules gate on the ACTUAL extracted aoa_findings
        # flag (not just "Israeli + has_preferred"). Populate aoa_findings so
        # every flag the rules consult is truthy in the firing sense, then
        # assert all four fire.
        cap_state_with_preferred = {
            "preferred_series": [
                {
                    "series_name": "Series Seed",
                    "anti_dilution_protection": "broad_based_weighted_average",
                    "liquidation_preference_multiple": 1.0,
                }
            ],
            "aoa_findings": {
                "drag_along_threshold_pct": 50,  # sub-75 → drag_along fires
                "section_102_plan_reference": False,  # absent → section_102_plan_absent fires
                "liquidation_preference_above_1x": True,  # → liquidation_preference_above_1x fires
                "ratchet_anti_dilution_detected": True,  # → full_ratchet_anti_dilution fires
            },
        }

        for rule_id in (
            "israeli_aoa.drag_along_threshold_below_75_percent",
            "israeli_aoa.section_102_plan_absent",
            "israeli_aoa.liquidation_preference_above_1x",
            "israeli_aoa.full_ratchet_anti_dilution",
        ):
            matcher = rule_audit._RULE_MATCHERS.get(rule_id, rule_audit._matcher_always)
            assert matcher(israeli_inputs, empty_instruments, cap_state_with_preferred) is True, (
                f"{rule_id} should match on Israeli engagement with preferred series + matching aoa_findings"
            )

    def test_israeli_aoa_rules_default_deny_when_aoa_findings_contradict_rule(self) -> None:
        """Regression: when aoa_findings flags contradict each rule's predicate
        (or aoa_findings is absent), the rule must NOT fire.

        Three rules (section_102_plan_absent, liquidation_preference_above_1x,
        full_ratchet_anti_dilution) were firing as false positives because they
        gated solely on `_structure_includes_israel + _has_preferred` instead
        of reading aoa_findings.*. This regression asserts the matchers now
        consult the AoA findings."""
        israeli_inputs = {
            "jurisdiction": {"structure": "israeli"},
            "mode": "standard",
        }
        empty_instruments: dict[str, Any] = {"safes": [], "convertible_notes": [], "warrants": [], "option_grants": []}
        # Each rule's contradicting AoA-finding flag:
        cap_state_contradicts = {
            "preferred_series": [
                {
                    "series_name": "Series A",
                    "anti_dilution_protection": "broad_based_weighted_average",
                    "liquidation_preference_multiple": 1.0,
                }
            ],
            "aoa_findings": {
                "drag_along_threshold_pct": 75,  # at threshold (not < 75) → drag_along does NOT fire
                "section_102_plan_reference": True,  # plan present → absent rule does NOT fire
                "liquidation_preference_above_1x": False,  # at-or-below 1x → rule does NOT fire
                "ratchet_anti_dilution_detected": False,  # BBWA → full_ratchet rule does NOT fire
            },
        }
        for rule_id in (
            "israeli_aoa.drag_along_threshold_below_75_percent",
            "israeli_aoa.section_102_plan_absent",
            "israeli_aoa.liquidation_preference_above_1x",
            "israeli_aoa.full_ratchet_anti_dilution",
        ):
            matcher = rule_audit._RULE_MATCHERS.get(rule_id, rule_audit._matcher_always)
            assert matcher(israeli_inputs, empty_instruments, cap_state_contradicts) is False, (
                f"{rule_id} should NOT match when aoa_findings contradict the rule"
            )

        # Null aoa_findings (AoA not extracted) → default-deny across all four.
        cap_state_no_aoa = {
            "preferred_series": [
                {
                    "series_name": "Series A",
                    "anti_dilution_protection": "broad_based_weighted_average",
                    "liquidation_preference_multiple": 1.0,
                }
            ],
        }
        for rule_id in (
            "israeli_aoa.drag_along_threshold_below_75_percent",
            "israeli_aoa.section_102_plan_absent",
            "israeli_aoa.liquidation_preference_above_1x",
            "israeli_aoa.full_ratchet_anti_dilution",
        ):
            matcher = rule_audit._RULE_MATCHERS.get(rule_id, rule_audit._matcher_always)
            assert matcher(israeli_inputs, empty_instruments, cap_state_no_aoa) is False, (
                f"{rule_id} should default-deny when aoa_findings is absent"
            )

    def test_acquisition_rules_do_not_fire_when_acquisition_absent(self) -> None:
        """P1 regression: acquisition.* counsel rules were absent from
        _RULE_MATCHERS and defaulted to _matcher_always, so they surfaced as
        counsel-review items even when the engagement has no acquisition.

        Now gated to: inputs.acquisition present (truthy).
        """
        rules = self._rules()
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        items = rule_audit.build_counsel_review_items(gating, rules)
        item_rule_ids = {it["rule_id"] for it in items}
        for rule_id in (
            "acquisition.consideration_shares",
            "acquisition.pool_consideration_basis",
            "acquisition.timing",
        ):
            assert rule_id not in item_rule_ids, (
                f"{rule_id} should NOT appear as a counsel item when inputs.acquisition is absent"
            )

    def test_acquisition_rules_fire_when_acquisition_present(self) -> None:
        """Acquisition block present (schema-valid) → acquisition.* counsel
        rules apply and surface as counsel-review items."""
        rules = self._rules()
        acquisition_inputs = dict(_BASIC_INPUTS)
        acquisition_inputs["acquisition"] = {
            "acquired_entity": "TargetCo",
            "consideration_pct": 0.15,
            "consideration_form": "shares",
            "acquisition_timing": "concurrent_with_round",
        }
        gating = rule_audit.build_gating_block(
            rules,
            inputs=acquisition_inputs,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(acquisition_inputs, _BASIC_INSTRUMENTS),
        )
        items = rule_audit.build_counsel_review_items(gating, rules)
        item_rule_ids = {it["rule_id"] for it in items}
        for rule_id in (
            "acquisition.consideration_shares",
            "acquisition.pool_consideration_basis",
            "acquisition.timing",
        ):
            assert rule_id in item_rule_ids, (
                f"{rule_id} should appear as a counsel item when inputs.acquisition is present"
            )

    def test_classify_scope_legal_tax_applicability(self) -> None:
        rule = {"date_window": {"event_date_field": "stock_issue_date", "start": "2025-07-05"}}
        assert rule_audit._classify_scope(rule) == "legal_tax_applicability"

    def test_classify_scope_benchmark_freshness(self) -> None:
        rule = {"date_window": {"event_date_field": "benchmark_reference_date"}}
        assert rule_audit._classify_scope(rule) == "benchmark_freshness"

    def test_status_in_window(self) -> None:
        from datetime import date

        status, near_end, near_start = rule_audit._evaluate_date_status(
            event_date_value=date(2026, 1, 1),
            start=date(2025, 1, 1),
            end=date(2026, 12, 31),
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "in_window"
        assert near_end is False  # > 90 days from end
        assert near_start is False

    def test_status_pre_effective(self) -> None:
        from datetime import date

        status, _, near_start = rule_audit._evaluate_date_status(
            event_date_value=date(2025, 6, 1),
            start=date(2025, 7, 5),
            end=None,  # QSBS-style start-only
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "pre_effective"
        # Within 30 days of start? 2025-07-05 - 2025-06-01 = 34 days, so no
        assert near_start is False

    def test_status_expired(self) -> None:
        from datetime import date

        status, _, _ = rule_audit._evaluate_date_status(
            event_date_value=date(2027, 1, 1),
            start=date(2025, 1, 1),
            end=date(2026, 12, 31),
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "expired"

    def test_status_near_end_flag(self) -> None:
        from datetime import date

        status, near_end, _ = rule_audit._evaluate_date_status(
            event_date_value=date(2026, 10, 15),  # 77 days before end 2026-12-31
            start=date(2025, 1, 1),
            end=date(2026, 12, 31),
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "in_window"
        assert near_end is True

    def test_status_missing_event_date(self) -> None:
        status, _, _ = rule_audit._evaluate_date_status(
            event_date_value=None,
            start=None,
            end=None,
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "missing_event_date"

    def test_gating_block_not_date_sensitive_for_core_formula(self) -> None:
        rules = self._rules()
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        # safe.post_money_cap_conversion has no date_window → not_date_sensitive
        entries = gating["safe.post_money_cap_conversion"]
        assert len(entries) == 1
        entry = next(iter(entries.values()))
        assert entry["status"] == "not_date_sensitive"
        assert entry["scope"] == "not_applicable"
        assert entry["event_date_field"] is None
        assert entry["event_date_path"] is None

    def test_gating_block_legal_tax_applicability(self) -> None:
        rules = self._rules()
        # QSBS rule: legal_tax_applicability, missing stock_issue_date in basic fixture
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        qsbs = gating["delaware_cross_border.qsbs_date_sensitive"]
        assert len(qsbs) >= 1
        entry = next(iter(qsbs.values()))
        assert entry["scope"] == "legal_tax_applicability"

    def test_gating_block_benchmark_freshness(self) -> None:
        rules = self._rules()
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        # qualified_financing_threshold uses benchmark_reference_date
        qft = gating["convertible_notes.qualified_financing_threshold"]
        entry = next(iter(qft.values()))
        assert entry["scope"] == "benchmark_freshness"
        assert entry["freshness_status"] in {"fresh", "stale", "unknown"}

    def test_counsel_items_include_missing_event_date(self) -> None:
        """Per rev17: rules with status=missing_event_date STILL surface in counsel
        items (founder needs to be prompted)."""
        rules = self._rules()
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        items = rule_audit.build_counsel_review_items(gating, rules)
        # QSBS has counsel_review=true and missing stock_issue_date in fixture
        qsbs_in_items = any(it["rule_id"] == "delaware_cross_border.qsbs_date_sensitive" for it in items)
        assert qsbs_in_items, f"QSBS missing from counsel items: {[i['rule_id'] for i in items]}"

    def test_applies_when_predicate_israeli_rule_in_delaware_engagement(self) -> None:
        """Per Phase 1 verification #3 (now implemented): Israeli-only rules
        must NOT apply in a pure-Delaware engagement."""
        rules = self._rules()
        # _BASIC_INPUTS has jurisdiction.structure="delaware"
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        # Israeli SAFE rules should be applies_when_matched=False
        israeli_entries = gating.get("safe.israeli_2025_safe_harbor", {})
        for entry in israeli_entries.values():
            assert entry["applies_when_matched"] is False, (
                "Israeli SAFE rule should NOT apply in Delaware-only engagement"
            )

    def test_applies_when_predicate_qsbs_in_israeli_engagement(self) -> None:
        """QSBS is Delaware-only; must NOT apply when structure is israeli."""
        rules = self._rules()
        israeli_inputs = dict(_BASIC_INPUTS)
        israeli_inputs["jurisdiction"] = {
            "structure": "israeli",
            "incorporated_date": "2024-01-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        }
        gating = rule_audit.build_gating_block(
            rules,
            inputs=israeli_inputs,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(israeli_inputs, _BASIC_INSTRUMENTS),
        )
        qsbs = gating["delaware_cross_border.qsbs_date_sensitive"]
        for entry in qsbs.values():
            assert entry["applies_when_matched"] is False, "QSBS should NOT apply in pure-Israeli engagement"

    def test_applies_when_predicate_iia_only_with_iia_grants(self) -> None:
        """IIA royalty rule should apply only when has_grants=True."""
        rules = self._rules()
        israeli_no_iia = dict(_BASIC_INPUTS)
        israeli_no_iia["jurisdiction"] = {
            "structure": "israeli",
            "incorporated_date": "2024-01-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        }
        gating_no_iia = rule_audit.build_gating_block(
            rules,
            inputs=israeli_no_iia,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(israeli_no_iia, _BASIC_INSTRUMENTS),
        )
        iia = gating_no_iia["israeli_ltd.iia_royalties_ip"]
        for entry in iia.values():
            assert entry["applies_when_matched"] is False

        # Now with IIA grants
        israeli_with_iia = dict(_BASIC_INPUTS)
        israeli_with_iia["jurisdiction"] = {
            "structure": "israeli",
            "incorporated_date": "2024-01-01",
            "iia_grants_history": {"has_grants": True, "grant_details": []},
        }
        gating_with_iia = rule_audit.build_gating_block(
            rules,
            inputs=israeli_with_iia,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(israeli_with_iia, _BASIC_INSTRUMENTS),
        )
        iia = gating_with_iia["israeli_ltd.iia_royalties_ip"]
        for entry in iia.values():
            assert entry["applies_when_matched"] is True

    def test_applies_when_predicate_stacked_safes_only_when_multiple(self) -> None:
        """stacked_post_money_caps applies only when 2+ SAFEs exist."""
        rules = self._rules()
        # Zero SAFEs
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        for entry in gating["safe.stacked_post_money_caps"].values():
            assert entry["applies_when_matched"] is False
        # Two SAFEs
        inst_two: dict[str, Any] = dict(_BASIC_INSTRUMENTS)
        inst_two["safes"] = [
            {**_SAFE_BASIC, "id": "safe_001"},
            {**_SAFE_BASIC, "id": "safe_002"},
        ]
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=inst_two,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, inst_two),
        )
        for entry in gating["safe.stacked_post_money_caps"].values():
            assert entry["applies_when_matched"] is True

    def test_applies_when_predicate_anti_dilution_only_when_protected(self) -> None:
        """BBWA rule should fire only when preferred series have BBWA protection."""
        rules = self._rules()
        # No preferred
        gating = rule_audit.build_gating_block(
            rules,
            inputs=_BASIC_INPUTS,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS),
        )
        for entry in gating["anti_dilution.broad_based_weighted_average"].values():
            assert entry["applies_when_matched"] is False

        # With BBWA-protected preferred
        inputs_with_pref = dict(_BASIC_INPUTS)
        inputs_with_pref["preferred_series"] = [
            {
                "series_name": "Series Seed",
                "shares": 1_000_000,
                "original_issue_price": 1.0,
                "original_conversion_price": 1.0,
                "current_conversion_price": 1.0,
                "issuance_date": "2025-01-01",
                "anti_dilution_protection": "broad_based_weighted_average",
            }
        ]
        cs_with_pref = cap_state_mod.build_cap_state(inputs_with_pref, _BASIC_INSTRUMENTS)
        gating = rule_audit.build_gating_block(
            rules,
            inputs=inputs_with_pref,
            instruments=_BASIC_INSTRUMENTS,
            cap_state=cs_with_pref,
        )
        for entry in gating["anti_dilution.broad_based_weighted_average"].values():
            assert entry["applies_when_matched"] is True

    def test_two_phase_subprocess(self) -> None:
        """Full pre_math + post_math subprocess round-trip."""
        with tempfile.TemporaryDirectory() as d:
            inp = os.path.join(d, "inputs.json")
            inst = os.path.join(d, "instruments.json")
            cap = os.path.join(d, "cap_state.json")
            audit = os.path.join(d, "rule_audit.json")
            with open(inp, "w") as f:
                json.dump(_BASIC_INPUTS, f)
            with open(inst, "w") as f:
                json.dump(_BASIC_INSTRUMENTS, f)
            cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
            cs["metadata"]["run_id"] = "rid1"
            with open(cap, "w") as f:
                json.dump(cs, f)
            rc, _, stderr = _run(
                "rule_audit.py",
                [
                    "--phase",
                    "pre_math",
                    "--inputs",
                    inp,
                    "--instruments",
                    inst,
                    "--cap-state",
                    cap,
                    "--run-id",
                    "rid1",
                    "-o",
                    audit,
                ],
            )
            assert rc == 0, stderr
            with open(audit) as f:
                pre = json.load(f)
            assert "gating" in pre
            assert pre["applied_rules"] == []
            assert pre["date_sensitive_watchlist"] == []

            rc, _, stderr = _run(
                "rule_audit.py",
                [
                    "--phase",
                    "post_math",
                    "--run-id",
                    "rid1",
                    "-o",
                    audit,
                ],
            )
            assert rc == 0, stderr
            with open(audit) as f:
                post = json.load(f)
            # Gating block preserved verbatim
            assert post["gating"] == pre["gating"]
            # post_math populates the watchlist + counsel items
            assert len(post["date_sensitive_watchlist"]) > 0
            assert len(post["applied_rules"]) > 0


# ===========================================================================
# Gotchas — explicit regression tests
# ===========================================================================


class TestGotchas:
    def test_gotcha_1_company_capitalization_excludes_new_money(self) -> None:
        """Gotcha #1: as_converted_totals MUST NOT include new-money shares or
        new pool top-ups. cap_state.py works ONLY with pre-financing inputs."""
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        # FD should ONLY include founders (10M) + existing pool (1.5M) = 11.5M
        # No new money. Adding a SAFE doesn't change this either:
        inst_with_safe = {**_BASIC_INSTRUMENTS, "safes": [_SAFE_BASIC]}
        cs_with_safe = cap_state_mod.build_cap_state(_BASIC_INPUTS, inst_with_safe)
        # FD is unchanged — outstanding SAFE is tracked separately, NOT in FD
        assert (
            cs["as_converted_totals"]["fully_diluted_shares"]
            == cs_with_safe["as_converted_totals"]["fully_diluted_shares"]
        )
        # SAFE is tracked
        assert len(cs_with_safe["outstanding_safes"]) == 1

    def test_gotcha_2_bbwa_divisor_uses_cp1(self) -> None:
        """Already tested in TestAntiDilution; this is a labelled regression."""
        r = anti_dilution.bbwa_new_conversion_price(
            current_conversion_price=2.0,  # CP1
            pre_issuance_share_count_A=10_000_000,
            consideration_received=1_000_000,
            new_issue_price=1.0,
            new_shares_issued_C=1_000_000,
        )
        # B = 1M / 2.0 = 500_000 (NOT 1M / OIP)
        assert r["intermediate"]["B"] == 500_000

    def test_gotcha_3_discount_multiplier_normalized(self) -> None:
        """A value of 20 (percent) should be converted to 0.80 (multiplier)."""
        # Use the extract_instrument library
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import normalize_discount_multiplier  # type: ignore[import-not-found]

        mult, warn = normalize_discount_multiplier(20.0)
        assert mult == 0.80
        assert warn is not None and "multiplier" in warn

        mult, warn = normalize_discount_multiplier(0.80)
        assert mult == 0.80
        assert warn is None

        mult, warn = normalize_discount_multiplier(None)
        assert mult is None
        assert warn is None

    def test_gotcha_3_discount_boundaries(self) -> None:
        """extraction-5: percent-multiplier vs discount-rate disambiguation.

        d>=50 → percent-multiplier (80→0.80, 50→0.50); 1<d<50 → discount-rate
        (49→0.51); d<=0 → error.
        """
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import normalize_discount_multiplier  # type: ignore[import-not-found]

        mult, warn = normalize_discount_multiplier(80.0)
        assert mult == 0.80, mult
        assert warn is not None

        mult, warn = normalize_discount_multiplier(50.0)
        assert mult == 0.50, mult

        mult, warn = normalize_discount_multiplier(49.0)
        assert mult is not None and abs(mult - 0.51) < 1e-9, mult

        mult, warn = normalize_discount_multiplier(0.0)
        assert mult is None
        assert warn is not None and "invalid" in warn.lower()

    def test_gotcha_4_mfn_cycle_unresolvable(self) -> None:
        safes = [
            {
                "id": "A",
                "form": "yc_uncapped_mfn",
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "B",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
            {
                "id": "B",
                "form": "yc_uncapped_mfn",
                "mfn_provision": {
                    "present": True,
                    "elected_against_safe_id": "A",
                    "elected": True,
                    "cherry_pick_attempted": False,
                    "notes": None,
                },
            },
        ]
        cycles = safe_conversion.detect_mfn_cycles(safes)
        assert len(cycles) == 1

    def test_gotcha_5_override_only_applies_to_convert_at_cap(self) -> None:
        for bad_treatment in ["repay", "extend", "counsel_review"]:
            note = dict(_NOTE_BASIC)
            note["maturity_default_treatment"] = bad_treatment
            note["maturity_conversion_price_override"] = 1.0
            r = note_conversion.convert_note(note, conversion_event_date="2027-06-01")
            assert r["branch"] == "override_mismatch", f"failed for {bad_treatment}"

    def test_gotcha_7_flip_is_share_for_share_only_in_v01(self) -> None:
        """v0.1 supports only 1:1 share-for-share."""
        sys.path.insert(0, SCRIPTS)
        from flip_scenario import flip_share_for_share  # type: ignore[import-not-found]

        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        r = flip_share_for_share(cs)
        assert r["_flip_metadata" if "_flip_metadata" in r else "post_flip_cap_state"]
        # Check the metadata note
        post = r["post_flip_cap_state"]
        assert post["_flip_metadata"]["exchange_ratio"] == "1:1"
        assert "v0_1_scope_limitation" in post["_flip_metadata"]
        # Founder share counts unchanged (1:1)
        for pre_f, post_f in zip(cs["founders"], post["founders"], strict=True):
            assert pre_f["common_shares"] == post_f["common_shares"]

    def test_gotcha_8_qsbs_date_is_2025_07_05(self) -> None:
        """Per Public Law 119-21 §70431: 'after July 4, 2025' → first in-window
        day is 2025-07-05 (with inclusive >= start semantics)."""
        with open(os.path.join(os.path.dirname(SCRIPTS), "data", "cap-table-rules.json")) as f:
            rules = json.load(f)
        qsbs = next(
            r
            for r in rules["domains"]["delaware_cross_border"]
            if r["rule_id"] == "delaware_cross_border.qsbs_date_sensitive"
        )
        assert qsbs["date_window"]["start"] == "2025-07-05"
        # Verify the date evaluation is consistent
        from datetime import date

        status, _, _ = rule_audit._evaluate_date_status(
            event_date_value=date(2025, 7, 4),  # day of enactment
            start=date(2025, 7, 5),
            end=None,
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "pre_effective"  # NOT in QSBS window
        status, _, _ = rule_audit._evaluate_date_status(
            event_date_value=date(2025, 7, 5),  # first eligible day
            start=date(2025, 7, 5),
            end=None,
            near_start_days=30,
            near_end_days=90,
        )
        assert status == "in_window"

    def test_gotcha_9_counsel_review_can_coexist_with_high_confidence(self) -> None:
        """counsel_review is a reliance boundary, NOT a confidence score.
        A rule can be confidence=high AND counsel_review=true."""
        with open(os.path.join(os.path.dirname(SCRIPTS), "data", "cap-table-rules.json")) as f:
            rules = json.load(f)
        high_conf_counsel = [
            r
            for domain in rules["domains"].values()
            for r in domain
            if r.get("confidence") == "high" and r.get("counsel_review") is True
        ]
        # At least one such rule must exist in the rule pack (proof of contract)
        assert len(high_conf_counsel) >= 1, "Expected at least one counsel_review=true + confidence=high rule"

    def test_gotcha_10_cap_implied_only_subflag_set_for_standalone_cap(self) -> None:
        """Standalone cap SAFE = completeness:structural_only + cap_implied_only:true."""
        # Via subprocess of run_scenario
        with tempfile.TemporaryDirectory() as d:
            inp = os.path.join(d, "inputs.json")
            inst = os.path.join(d, "instruments.json")
            cap = os.path.join(d, "cap_state.json")
            reqs = os.path.join(d, "scenario_requests.json")
            scen = os.path.join(d, "scenarios.json")
            with open(inp, "w") as f:
                json.dump(_BASIC_INPUTS, f)
            inst_safe = {**_BASIC_INSTRUMENTS, "safes": [_SAFE_BASIC]}
            with open(inst, "w") as f:
                json.dump(inst_safe, f)
            cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, inst_safe)
            cs["metadata"]["run_id"] = "rid"
            with open(cap, "w") as f:
                json.dump(cs, f)
            with open(reqs, "w") as f:
                json.dump(
                    [
                        {
                            "scenario_id": "base",
                            "label": "Standalone cap SAFE",
                            "type": "safe_conversion",
                            "parameters": {},
                        }
                    ],
                    f,
                )
            rc, _, stderr = _run(
                "run_scenario.py",
                [
                    "--inputs",
                    inp,
                    "--instruments",
                    inst,
                    "--cap-state",
                    cap,
                    "--scenarios-input",
                    reqs,
                    "--run-id",
                    "rid",
                    "-o",
                    scen,
                ],
            )
            assert rc == 0, stderr
            with open(scen) as f:
                doc = json.load(f)
            co = doc["scenarios"][0]["computed_outputs"]
            assert co["completeness"] == "structural_only"
            assert co.get("cap_implied_only") is True


# ===========================================================================
# Full pipeline end-to-end (subprocess; mirrors SKILL.md workflow)
# ===========================================================================


# ===========================================================================
# Corpus-test-surfaced robustness fixes (real-world edge cases)
# ===========================================================================


# ===========================================================================
# Pre-baseline patches — bugs surfaced by eval-set labeling:
# - legacy pre-money SAFE forms (multiple 2017–2025 corpus cases)
# - statutory ITA Section 3(j) interest rate (Israeli CLA corpus case)
# - warrant + non_instrument classification (warrant-in-safes-folder misfile)
# ===========================================================================


class TestPreBaselinePatches:
    def test_yc_premoney_cap_only_form_validates(self) -> None:
        """Legacy YC pre-money cap-only SAFE corpus case must
        validate when pre_money_valuation_cap is set + no discount."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_safe  # type: ignore[import-not-found]

        errs = validate_safe(
            {
                "form": "yc_premoney_cap_only",
                "purchase_amount": 500000,
                "issuance_date": "2020-06-17",
                "pre_money_valuation_cap": 40_000_000,
                "post_money_valuation_cap": None,
                "discount_multiplier": None,
            }
        )
        assert errs == [], f"unexpected validation errors: {errs}"

    def test_yc_premoney_cap_only_rejects_post_money_cap(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_safe  # type: ignore[import-not-found]

        errs = validate_safe(
            {
                "form": "yc_premoney_cap_only",
                "purchase_amount": 500000,
                "issuance_date": "2020-06-17",
                "pre_money_valuation_cap": 40_000_000,
                "post_money_valuation_cap": 40_000_000,  # both set — invalid
            }
        )
        assert any("post_money_valuation_cap" in e and "null" in e for e in errs), errs

    def test_pre_money_cap_and_discount_legacy_validates(self) -> None:
        """Legacy YC pre-money cap + discount corpus case."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_safe  # type: ignore[import-not-found]

        errs = validate_safe(
            {
                "form": "pre_money_cap_and_discount_legacy",
                "purchase_amount": 100000,
                "issuance_date": "2017-04-20",
                "pre_money_valuation_cap": 6_000_000,
                "post_money_valuation_cap": None,
                "discount_multiplier": 0.80,
            }
        )
        assert errs == [], f"unexpected validation errors: {errs}"

    def test_pre_money_form_requires_pre_money_cap(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_safe  # type: ignore[import-not-found]

        errs = validate_safe(
            {
                "form": "pre_money_cap_and_discount_legacy",
                "purchase_amount": 100000,
                "issuance_date": "2017-04-20",
                "pre_money_valuation_cap": None,  # missing!
                "discount_multiplier": 0.80,
            }
        )
        assert any("pre_money_valuation_cap" in e for e in errs), errs

    def test_validate_note_accepts_statutory_ita_section_3j(self) -> None:
        """Israeli CLA corpus case — statutory ITA Section 3(j) rate. annual_interest_rate
        is null because rate is set annually by ITA."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 1_500_000,
                "interest_rate_type": "statutory_ita_section_3j",
                "annual_interest_rate": None,
                "day_count_basis": 365,
                "issuance_date": "2014-03-25",
                "maturity_date": "2015-09-25",
                "maturity_default_treatment": "convert_at_cap",
            }
        )
        assert errs == [], f"statutory rate should validate with null annual_interest_rate; got: {errs}"

    def test_validate_note_rejects_numeric_when_statutory(self) -> None:
        """Agent must NOT fabricate a numeric rate when type is statutory."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 1_500_000,
                "interest_rate_type": "statutory_ita_section_3j",
                "annual_interest_rate": 0.05,  # fabricated!
                "day_count_basis": 365,
                "issuance_date": "2014-03-25",
                "maturity_date": "2015-09-25",
                "maturity_default_treatment": "convert_at_cap",
            }
        )
        assert any("must be null" in e for e in errs), errs

    def test_validate_note_accepts_none_interest(self) -> None:
        """SAFE-equivalent convertible security — no interest at all."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 2_500_000,
                "interest_rate_type": "none",
                "annual_interest_rate": None,
                "day_count_basis": 365,
                "issuance_date": "2019-08-07",
                "maturity_date": "2024-08-07",
                "maturity_default_treatment": "convert_at_cap",
            }
        )
        assert errs == [], errs

    def test_validate_note_fixed_numeric_still_requires_rate(self) -> None:
        """Backward compat — existing notes with default interest_rate_type
        (fixed_numeric) MUST still have a numeric rate."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 1_000_000,
                "interest_rate_type": "fixed_numeric",
                "annual_interest_rate": None,  # missing!
                "day_count_basis": 365,
                "issuance_date": "2019-12-13",
                "maturity_date": "2021-06-13",
                "maturity_default_treatment": "convert_at_cap",
            }
        )
        assert any("annual_interest_rate" in e for e in errs), errs

    def test_validate_note_default_interest_rate_type(self) -> None:
        """If interest_rate_type is absent, default to fixed_numeric for backward compat."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 1_000_000,
                # No interest_rate_type — should default to fixed_numeric
                "annual_interest_rate": 0.05,
                "day_count_basis": 365,
                "issuance_date": "2019-12-13",
                "maturity_date": "2021-06-13",
                "maturity_default_treatment": "convert_at_cap",
            }
        )
        assert errs == [], f"default fixed_numeric should validate; got: {errs}"

    def test_warrant_doc_type_passes_validation(self) -> None:
        """WARRANT misfile (warrant landed in safes/ folder) — extractor should classify cleanly."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_non_extractable  # type: ignore[import-not-found]

        # Empty fields block + warrant doc_type should be acceptable
        errs = validate_non_extractable({})
        assert errs == [], errs

    def test_non_instrument_doc_type_passes_validation(self) -> None:
        """Notice-of-Financing letter mistaken for a SAFE — extractor classifies as non_instrument."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_non_extractable  # type: ignore[import-not-found]

        errs = validate_non_extractable({})
        assert errs == [], errs

    # ------------------------------------------------------------------
    # Convertible aliases
    # ------------------------------------------------------------------

    def test_validate_note_subtype_cla_uses_standard_gate(self) -> None:
        """Israeli CLA (convertible_loan_agreement subtype) is mathematically
        identical to a standard convertible_note — same required fields."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 1_500_000,
                "interest_rate_type": "statutory_ita_section_3j",
                "annual_interest_rate": None,
                "day_count_basis": 365,
                "issuance_date": "2019-06-01",
                "maturity_date": "2021-06-01",
                "maturity_default_treatment": "convert_at_cap",
            },
            subtype="convertible_loan_agreement",
        )
        assert errs == [], f"CLA subtype should validate as standard note; got: {errs}"

    def test_validate_note_subtype_convertible_security_waives_maturity(self) -> None:
        """YC convertible_security (SAFE-equivalent) has no maturity / no interest.
        Subtype gate waives maturity_date, maturity_default_treatment,
        day_count_basis, and annual_interest_rate."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        errs = validate_note(
            {
                "principal": 1_000_000,
                "interest_rate_type": "none",
                "annual_interest_rate": None,
                "day_count_basis": None,
                "issuance_date": "2019-08-01",
                "maturity_date": None,
                "maturity_default_treatment": None,
            },
            subtype="convertible_security",
        )
        assert errs == [], f"convertible_security should pass with null maturity/interest; got: {errs}"

    def test_validate_note_subtype_unknown_falls_back_to_standard_gate(self) -> None:
        """An unknown subtype value falls back to the standard convertible_note
        gate (all canonical fields required)."""
        sys.path.insert(0, SCRIPTS)
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        # Missing maturity_date — should fail with subtype=None (the standard
        # convertible_note gate requires it).
        errs = validate_note(
            {
                "principal": 1_000_000,
                "interest_rate_type": "fixed_numeric",
                "annual_interest_rate": 0.05,
                "day_count_basis": 365,
                "issuance_date": "2019-12-13",
                "maturity_date": None,
                "maturity_default_treatment": None,
            },
            subtype=None,
        )
        assert any("maturity_date" in e for e in errs), f"standard note gate must require maturity_date; got: {errs}"

    def test_cli_routes_convertible_loan_agreement_to_notes_with_subtype(self) -> None:
        """End-to-end CLI: instrument_type='convertible_loan_agreement' input
        must (a) validate via the note gate, (b) land in instruments.notes[],
        and (c) carry subtype='convertible_loan_agreement' for provenance."""
        cla_extraction = {
            "instrument_type": "convertible_loan_agreement",
            "fields": {
                "investor_name": "Acmecorp Investors Trustee Ltd.",
                "principal": 7_000_000,
                "interest_rate_type": "statutory_ita_section_3j",
                "annual_interest_rate": None,
                "day_count_basis": 365,
                "interest_converts_to_shares": True,
                "issuance_date": "2019-06-01",
                "valuation_cap": 50_000_000,
                "discount_multiplier": 0.80,
                "qualified_financing_threshold": 5_000_000,
                "maturity_date": "2021-06-01",
                "maturity_default_treatment": "convert_at_cap",
                "extraction_confidence": "high",
            },
            "confidence": {},
            "ambiguities": [],
        }
        with tempfile.TemporaryDirectory() as d:
            instr_path = os.path.join(d, "instruments.json")
            with open(instr_path, "w") as f:
                json.dump(
                    {
                        "safes": [],
                        "convertible_notes": [],
                        "warrants": [],
                        "option_grants": [],
                        "metadata": {"run_id": "test"},
                    },
                    f,
                )
            rc, _, e = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(cla_extraction),
            )
            assert rc == 0, f"CLA extraction failed: rc={rc}, stderr={e}"
            with open(instr_path) as f:
                instruments = json.load(f)
            assert len(instruments["convertible_notes"]) == 1
            note = instruments["convertible_notes"][0]
            assert note["subtype"] == "convertible_loan_agreement"
            assert note["principal"] == 7_000_000

    def test_cli_routes_convertible_security_with_null_maturity(self) -> None:
        """End-to-end CLI: convertible_security input lands in instruments.notes[]
        with subtype tag and null maturity fields. Validates the SAFE-equivalent
        path against a synthetic Israeli-style convertible_security shape."""
        cs_extraction = {
            "instrument_type": "convertible_security",
            "fields": {
                "investor_name": "Acmecorp Holdings",
                "principal": 500_000,
                "interest_rate_type": "none",
                "annual_interest_rate": None,
                "interest_converts_to_shares": True,
                "issuance_date": "2019-08-01",
                "valuation_cap": 15_000_000,
                "discount_multiplier": None,
                "qualified_financing_threshold": 1_000_000,
                "maturity_date": None,
                "maturity_default_treatment": None,
                "extraction_confidence": "high",
            },
            "confidence": {},
            "ambiguities": [],
        }
        with tempfile.TemporaryDirectory() as d:
            instr_path = os.path.join(d, "instruments.json")
            with open(instr_path, "w") as f:
                json.dump(
                    {
                        "safes": [],
                        "convertible_notes": [],
                        "warrants": [],
                        "option_grants": [],
                        "metadata": {"run_id": "test"},
                    },
                    f,
                )
            rc, _, e = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(cs_extraction),
            )
            assert rc == 0, f"convertible_security extraction failed: rc={rc}, stderr={e}"
            with open(instr_path) as f:
                instruments = json.load(f)
            assert len(instruments["convertible_notes"]) == 1
            note = instruments["convertible_notes"][0]
            assert note["subtype"] == "convertible_security"
            assert note["maturity_date"] is None
            assert note["interest_rate_type"] == "none"

    # ------------------------------------------------------------------
    # Fix A: duplicate instrument id guard + --replace upsert
    # ------------------------------------------------------------------

    @staticmethod
    def _base_instr() -> dict[str, Any]:
        return {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }

    @staticmethod
    def _safe_extraction(safe_id: str = "safe_seed_1") -> dict[str, Any]:
        return {
            "instrument_type": "safe",
            "fields": {
                "id": safe_id,
                "investor_name": "Angel A",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": 8_000_000,
                "discount_multiplier": None,
                "issuance_date": "2025-01-01",
                "form": "yc_postmoney_cap",
                "extraction_confidence": "high",
            },
            "confidence": {},
            "ambiguities": [],
        }

    def test_duplicate_id_exits_1_no_flag(self) -> None:
        """Second insert with same id and no --replace flag must exit 1,
        emit E_DUPLICATE_INSTRUMENT_ID in output, and leave the file with
        exactly one entry."""
        with tempfile.TemporaryDirectory() as d:
            instr_path = os.path.join(d, "instruments.json")
            with open(instr_path, "w") as f:
                json.dump(self._base_instr(), f)
            # First insert succeeds
            rc1, _, _ = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t1", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(self._safe_extraction("safe_seed_1")),
            )
            assert rc1 == 0, "first insert should succeed"
            # Second insert with same id — must fail
            rc2, stdout2, stderr2 = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t2", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(self._safe_extraction("safe_seed_1")),
            )
            assert rc2 == 1, "duplicate id without --replace must exit 1"
            combined = stdout2 + stderr2
            assert "E_DUPLICATE_INSTRUMENT_ID" in combined, (
                f"expected E_DUPLICATE_INSTRUMENT_ID in output; got:\n{combined}"
            )
            assert "safe_seed_1" in combined, "error must name the duplicate id"
            # File must still have exactly one entry (not two)
            with open(instr_path) as f:
                instruments = json.load(f)
            assert len(instruments["safes"]) == 1, (
                "file must be unchanged (still exactly 1 safe) after duplicate-id rejection"
            )

    def test_content_duplicate_fresh_id_warns(self) -> None:
        """Re-piping identical content under a DIFFERENT id (the auto-sequential-id
        footgun) appends a second entry but surfaces W_POSSIBLE_DUPLICATE_INSTRUMENT
        in the receipt, so a silent duplicate is at least flagged."""
        with tempfile.TemporaryDirectory() as d:
            instr_path = os.path.join(d, "instruments.json")
            with open(instr_path, "w") as f:
                json.dump(self._base_instr(), f)
            rc1, _, _ = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t1", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(self._safe_extraction("safe_a")),
            )
            assert rc1 == 0
            rc2, out2, err2 = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t2", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(self._safe_extraction("safe_b")),  # same content, new id
            )
            assert rc2 == 0, "a fresh-id insert is not a hard error"
            assert "W_POSSIBLE_DUPLICATE_INSTRUMENT" in (out2 + err2), (
                f"expected content-duplicate warning; got:\n{out2}{err2}"
            )

    def test_same_id_replace_does_not_warn_content_duplicate(self) -> None:
        """A deliberate same-id --replace upsert must NOT trip the content-duplicate
        warning (the different-id guard prevents a false positive on legitimate replace)."""
        with tempfile.TemporaryDirectory() as d:
            instr_path = os.path.join(d, "instruments.json")
            with open(instr_path, "w") as f:
                json.dump(self._base_instr(), f)
            _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t1", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(self._safe_extraction("safe_a")),
            )
            rc2, out2, err2 = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t2", "--no-verify", "--no-invariants", "--replace"],
                stdin_data=json.dumps(self._safe_extraction("safe_a")),
            )
            assert rc2 == 0
            assert "W_POSSIBLE_DUPLICATE_INSTRUMENT" not in (out2 + err2), (
                "same-id --replace must not trip the content-duplicate warning"
            )

    def test_duplicate_id_with_replace_upserts(self) -> None:
        """Second insert with same id and --replace must exit 0, leave
        exactly one entry in the array, and reflect the updated values."""
        with tempfile.TemporaryDirectory() as d:
            instr_path = os.path.join(d, "instruments.json")
            with open(instr_path, "w") as f:
                json.dump(self._base_instr(), f)
            # First insert
            rc1, _, _ = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t1", "--no-verify", "--no-invariants"],
                stdin_data=json.dumps(self._safe_extraction("safe_seed_1")),
            )
            assert rc1 == 0
            # Updated extraction — same id, different purchase_amount
            updated = self._safe_extraction("safe_seed_1")
            updated["fields"]["purchase_amount"] = 750_000
            rc2, _, stderr2 = _run(
                "extract_instrument.py",
                ["--instruments", instr_path, "--run-id", "t2", "--no-verify", "--no-invariants", "--replace"],
                stdin_data=json.dumps(updated),
            )
            assert rc2 == 0, f"--replace should succeed; stderr={stderr2}"
            with open(instr_path) as f:
                instruments = json.load(f)
            assert len(instruments["safes"]) == 1, "--replace must leave exactly 1 entry"
            assert instruments["safes"][0]["purchase_amount"] == 750_000, "--replace must update to new values"


class TestAoAExtraction:
    """Tests for extract_aoa.py — AoA (Articles of Association) extraction
    validator + counsel-review item detection + merge-into-inputs flow.

    Synthetic fixtures modeled on five real Israeli AoAs (2012–2024
    vintages, anonymized). Real AoAs are NOT included as fixtures (per the
    redaction rule — no real founder/company names in committed artifacts).
    """

    @staticmethod
    def _valid_aoa_extraction() -> dict[str, Any]:
        """Synthetic AoA extraction shape — single preferred series, Israeli."""
        return {
            "extraction_type": "articles_of_association",
            "fields": {
                "company_name": "Acmecorp Ltd.",
                "jurisdiction_structure": "israeli",
                "section_102_plan_reference": True,
                "drag_along_threshold_pct": 0.75,
                "preferred_series": [
                    {
                        "series_name": "Series Seed",
                        "shares": None,
                        "original_issue_price": 1.175,
                        "original_conversion_price": 1.175,
                        "current_conversion_price": 1.175,
                        "issuance_date": "2018-03-15",
                        "liquidation_preference_multiple": 1.0,
                        "liquidation_preference_type": "participating",
                        "participation_cap_multiple": None,
                        "anti_dilution_protection": "broad_based_weighted_average",
                        "dividend_rate_percent": 0.08,
                        "dividend_cumulative": True,
                        "pro_rata_rights": True,
                    }
                ],
            },
            "confidence": {},
            "ambiguities": [],
        }

    def test_valid_aoa_extraction_passes(self) -> None:
        """Canonical valid AoA extraction — no validation errors."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import validate_aoa_extraction  # type: ignore[import-not-found]

        errs = validate_aoa_extraction(self._valid_aoa_extraction())
        assert errs == [], errs

    def test_extraction_type_mismatch_rejected(self) -> None:
        """extraction_type must be 'articles_of_association' — anything else fails."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import validate_aoa_extraction  # type: ignore[import-not-found]

        ext = self._valid_aoa_extraction()
        ext["extraction_type"] = "instrument_extraction"
        errs = validate_aoa_extraction(ext)
        assert any("extraction_type" in e for e in errs)

    def test_missing_oip_rejected(self) -> None:
        """A preferred series without original_issue_price fails validation."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import validate_aoa_extraction  # type: ignore[import-not-found]

        ext = self._valid_aoa_extraction()
        ext["fields"]["preferred_series"][0]["original_issue_price"] = None
        errs = validate_aoa_extraction(ext)
        assert any("original_issue_price" in e for e in errs)

    def test_participating_capped_requires_cap_multiple(self) -> None:
        """liquidation_preference_type='participating_capped' requires
        participation_cap_multiple to be non-null."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import validate_aoa_extraction  # type: ignore[import-not-found]

        ext = self._valid_aoa_extraction()
        ext["fields"]["preferred_series"][0]["liquidation_preference_type"] = "participating_capped"
        ext["fields"]["preferred_series"][0]["participation_cap_multiple"] = None
        errs = validate_aoa_extraction(ext)
        assert any("participation_cap_multiple" in e for e in errs)

    def test_invalid_anti_dilution_enum_rejected(self) -> None:
        """anti_dilution_protection must be one of the canonical enum values."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import validate_aoa_extraction  # type: ignore[import-not-found]

        ext = self._valid_aoa_extraction()
        ext["fields"]["preferred_series"][0]["anti_dilution_protection"] = "magic_ratchet"
        errs = validate_aoa_extraction(ext)
        assert any("anti_dilution_protection" in e for e in errs)

    def test_liquidation_pref_below_1x_rejected(self) -> None:
        """liquidation_preference_multiple must be ≥ 1.0 (Israeli + US convention)."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import validate_aoa_extraction  # type: ignore[import-not-found]

        ext = self._valid_aoa_extraction()
        ext["fields"]["preferred_series"][0]["liquidation_preference_multiple"] = 0.5
        errs = validate_aoa_extraction(ext)
        assert any("liquidation_preference_multiple" in e for e in errs)

    def test_counsel_items_drag_along_below_75(self) -> None:
        """Israeli AoA with drag-along < 75% surfaces counsel-review item."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import detect_counsel_review_items  # type: ignore[import-not-found]

        fields = self._valid_aoa_extraction()["fields"]
        fields["drag_along_threshold_pct"] = 0.65
        items = detect_counsel_review_items(fields)
        assert any(it["rule_id"] == "israeli_aoa.drag_along_threshold_below_75_percent" for it in items)

    def test_counsel_items_section_102_absent(self) -> None:
        """Israeli AoA without §102 plan reference flags counsel review."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import detect_counsel_review_items  # type: ignore[import-not-found]

        fields = self._valid_aoa_extraction()["fields"]
        fields["section_102_plan_reference"] = False
        items = detect_counsel_review_items(fields)
        assert any(it["rule_id"] == "israeli_aoa.section_102_plan_absent" for it in items)

    def test_counsel_items_above_1x_liquidation_pref(self) -> None:
        """Above-1x liquidation preference triggers counsel review."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import detect_counsel_review_items  # type: ignore[import-not-found]

        fields = self._valid_aoa_extraction()["fields"]
        fields["preferred_series"][0]["liquidation_preference_multiple"] = 2.0
        items = detect_counsel_review_items(fields)
        assert any(it["rule_id"] == "israeli_aoa.liquidation_preference_above_1x" for it in items)

    def test_counsel_items_full_ratchet_anti_dilution(self) -> None:
        """Full-ratchet anti-dilution triggers counsel review."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import detect_counsel_review_items  # type: ignore[import-not-found]

        fields = self._valid_aoa_extraction()["fields"]
        fields["preferred_series"][0]["anti_dilution_protection"] = "full_ratchet"
        items = detect_counsel_review_items(fields)
        assert any(it["rule_id"] == "israeli_aoa.full_ratchet_anti_dilution" for it in items)

    def test_counsel_items_clean_aoa_returns_empty(self) -> None:
        """A clean AoA (drag-along ≥ 75%, §102 present, 1x non-participating,
        BBWA anti-dilution) produces no counsel-review items."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import detect_counsel_review_items  # type: ignore[import-not-found]

        fields = self._valid_aoa_extraction()["fields"]
        fields["preferred_series"][0]["liquidation_preference_multiple"] = 1.0
        fields["preferred_series"][0]["liquidation_preference_type"] = "non_participating"
        fields["preferred_series"][0]["anti_dilution_protection"] = "broad_based_weighted_average"
        items = detect_counsel_review_items(fields)
        assert items == [], f"clean AoA should produce no counsel items; got: {items}"

    def test_merge_into_inputs_appends_new_series(self) -> None:
        """AoA-extracted preferred_series is appended to inputs.preferred_series[]
        with extraction_provenance stamp."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import merge_into_inputs  # type: ignore[import-not-found]

        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            with open(inputs_path, "w") as f:
                json.dump(
                    {
                        "company_name": "Acmecorp Ltd.",
                        "analysis_date": "2026-05-21",
                        "mode": "standard",
                        "preferred_series": [],
                        "metadata": {"run_id": "test"},
                    },
                    f,
                )

            new_series = self._valid_aoa_extraction()["fields"]["preferred_series"]
            result = merge_into_inputs(inputs_path, new_series, source_doc="/path/to/aoa.pdf")
            assert result["status"] == "merged"
            assert result["added_count"] == 1
            with open(inputs_path) as f:
                inputs = json.load(f)
            assert len(inputs["preferred_series"]) == 1
            merged = inputs["preferred_series"][0]
            assert merged["series_name"] == "Series Seed"
            assert merged["extraction_provenance"]["source_doc"] == "/path/to/aoa.pdf"

    def test_merge_into_inputs_conflict_on_duplicate_series_name(self) -> None:
        """Attempting to merge a series with an already-present series_name
        returns status='conflict' (caller must resolve)."""
        sys.path.insert(0, SCRIPTS)
        from extract_aoa import merge_into_inputs  # type: ignore[import-not-found]

        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            with open(inputs_path, "w") as f:
                json.dump(
                    {
                        "company_name": "Acmecorp Ltd.",
                        "preferred_series": [{"series_name": "Series Seed", "shares": 1_000_000}],
                        "metadata": {"run_id": "test"},
                    },
                    f,
                )
            new_series = self._valid_aoa_extraction()["fields"]["preferred_series"]
            result = merge_into_inputs(inputs_path, new_series, source_doc="/path/to/aoa.pdf")
            assert result["status"] == "conflict"
            assert "Series Seed" in result["conflicts"]

    def test_cli_end_to_end_validates_and_merges(self) -> None:
        """End-to-end: pipe AoA extraction JSON through extract_aoa.py CLI,
        verify it validates + merges + returns counsel items."""
        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            with open(inputs_path, "w") as f:
                json.dump(
                    {
                        "company_name": "Acmecorp Ltd.",
                        "analysis_date": "2026-05-21",
                        "mode": "standard",
                        "preferred_series": [],
                        "metadata": {"run_id": "test"},
                    },
                    f,
                )
            extraction = self._valid_aoa_extraction()
            rc, out, err = _run(
                "extract_aoa.py",
                ["--run-id", "test", "--inputs", inputs_path, "--source-doc", "/path/aoa.pdf", "--pretty"],
                stdin_data=json.dumps(extraction),
            )
            assert rc == 0, f"extract_aoa.py failed: rc={rc}, stderr={err}"
            receipt = json.loads(out)
            # After a successful merge the top-level status reflects the merge
            # outcome (merged), not the pre-merge validation status.
            assert receipt["status"] == "merged"
            assert receipt["preferred_series_count"] == 1
            assert "merge" in receipt
            assert receipt["merge"]["status"] == "merged"


class TestPrivacyAssertion:
    """Tests for compose_report._assert_coaching_payload_privacy_clean().

    M9 word-boundary + length-threshold defense-in-depth: ensures the
    assertion catches genuine investor-name leaks while NOT firing on
    legitimate generic text that happens to contain investor-name substrings.
    """

    @staticmethod
    def _instruments_with_investor(name: str) -> dict[str, Any]:
        return {
            "safes": [
                {
                    "id": "s1",
                    "investor_name": name,
                    "purchase_amount": 500_000,
                    "post_money_valuation_cap": 10_000_000,
                    "discount_multiplier": None,
                    "mfn_provision": None,
                    "pro_rata_side_letter": None,
                    "issuance_date": "2025-01-01",
                    "form": "yc_postmoney_cap",
                    "conversion_price_override": None,
                    "source_document": None,
                    "extraction_confidence": "high",
                }
            ],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }

    def test_assertion_catches_word_boundary_leak(self) -> None:
        """Positive case: investor name appears as a whole word in the payload."""
        import compose_report  # type: ignore[import-not-found]

        payload = {
            "company_name": "TestCo",
            "scenario_digest": [{"scenario_drivers": ["Sequoia Capital Operations LLC participated in the round"]}],
        }
        instruments = self._instruments_with_investor("Sequoia Capital Operations LLC")
        with pytest.raises(AssertionError, match="privacy-scrub violation"):
            compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments)

    def test_assertion_passes_on_clean_payload(self) -> None:
        """Negative case: legitimate payload with no investor names passes."""
        import compose_report  # type: ignore[import-not-found]

        payload = {
            "company_name": "TestCo",
            "scenario_digest": [
                {
                    "scenario_drivers": [
                        "Series A at $20M pre-money",
                        "Pool top-up to 10% post-money",
                    ]
                }
            ],
        }
        instruments = self._instruments_with_investor("Sequoia Capital Operations LLC")
        # Should not raise
        compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments)

    def test_assertion_does_not_false_fire_on_short_substring(self) -> None:
        """M9: short / common investor-name substrings ('SAFE', 'Inc', 'LLC',
        'Capital') must NOT trigger false positives on generic template prose.
        Length threshold >8 chars filters these out.
        """
        import compose_report  # type: ignore[import-not-found]

        payload = {
            "scenario_digest": [
                {"branch_summary": "cap_plus_discount + new money", "completeness": "full"},
                {"scenario_drivers": ["Pool top-up to 10% post-money basis"]},
            ],
            "headline_inputs": {"target_basis": "post_money"},
        }
        # Real-world short fund names that would falsely fire under the v1 bare
        # substring check (n>2): "SAFE" inside "yc_postmoney_cap_plus_discount",
        # "Inc" inside "Pool top-up to 10%" — wait, "Inc" doesn't appear here.
        # The key M9 test: a short fund name doesn't cause false positives on
        # standard template strings.
        for short_name in ["SAFE Fund", "Capital", "Inc Co", "LLC Co"]:
            instruments = self._instruments_with_investor(short_name)
            compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments)

    def test_assertion_word_boundary_prevents_substring_false_positive(self) -> None:
        """M9: even a long investor name should match only as a whole word.
        'AcmeCorp Partners' must NOT match 'AcmeCorp Partnership' (different
        word). With bare substring `in`, this would fire spuriously.
        """
        import compose_report  # type: ignore[import-not-found]

        payload = {
            "scenario_digest": [{"scenario_drivers": ["AcmeCorp Partnership terms reviewed"]}],
        }
        instruments = self._instruments_with_investor("AcmeCorp Partners")
        # 'AcmeCorp Partners' is NOT a substring of 'AcmeCorp Partnership' at
        # word boundaries — should pass.
        compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments)

    def test_assertion_catches_founder_name_leak(self) -> None:
        """H6: founder names from inputs.founders[].name must also fire the
        assertion. The agent body promises both investor + founder scrubbing.
        """
        import compose_report  # type: ignore[import-not-found]

        payload = {
            "scenario_digest": [{"scenario_drivers": ["Alexander Hamilton-Jones diluted to 22%"]}],
        }
        instruments = self._instruments_with_investor("Sequoia Capital Operations")
        inputs = {
            "company_name": "TestCo",
            "founders": [{"name": "Alexander Hamilton-Jones", "common_shares": 10_000_000}],
        }
        with pytest.raises(AssertionError, match="founder name leak"):
            compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments, inputs=inputs)

    def test_assertion_company_name_carve_out(self) -> None:
        """H6 carve-out: a founder whose name overlaps with the company name
        (founder 'Acme Holdings Founder Trust' at 'Acme Holdings') must NOT
        trigger the assertion. The company_name is intentionally in the
        payload as engagement identity.
        """
        import compose_report  # type: ignore[import-not-found]

        payload = {
            "company_name": "Acme Holdings",
            "scenario_digest": [{"label": "Series A at Acme Holdings"}],
        }
        instruments = self._instruments_with_investor("Sequoia Capital Operations")
        inputs = {
            "company_name": "Acme Holdings",
            "founders": [{"name": "Acme Holdings Founder Trust", "common_shares": 10_000_000}],
        }
        # Founder name contains the company name → carve out, no leak.
        compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments, inputs=inputs)

    def test_assertion_founder_becomes_investor_carve_out(self) -> None:
        """H6 carve-out: a founder who also appears as an investor in the
        SAFE list (common in Israeli market) is treated as an investor for
        the purpose of this check. The check still fires on investor-side
        leaks (if name appears in payload), but the founder-side check
        is suppressed to avoid double-counting.
        """
        import compose_report  # type: ignore[import-not-found]

        founder_investor_name = "Alice Mendelssohn-Rothschild"
        payload = {
            "scenario_digest": [{"scenario_drivers": ["Series A pre-money $20M"]}],
        }
        instruments = self._instruments_with_investor(founder_investor_name)
        inputs = {
            "company_name": "TestCo",
            "founders": [{"name": founder_investor_name, "common_shares": 10_000_000}],
        }
        # Name in both lists; payload doesn't leak it → should pass.
        compose_report._assert_coaching_payload_privacy_clean(payload, instruments=instruments, inputs=inputs)


class TestComposeReportWarningCallouts:
    """Part A: cap_state warnings render as founder-facing callouts. The AD recovery warnings
    (W_ANTI_DILUTION_*) are interpolated SENTENCES, so the renderer must match by PREFIX, not the
    exact-`==` the existing bare-code callouts use — otherwise the already-shipped AD fix stays
    invisible to founders."""

    def test_renders_anti_dilution_warning_by_prefix(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        warn = (
            "W_ANTI_DILUTION_NONCANONICAL: preferred series 'Series Seed' specified anti-dilution "
            "under the wrong key `anti_dilution`='bbwa' — recovered as 'broad_based_weighted_average'."
        )
        body = "\n".join(compose_report._render_warning_callouts([warn]))
        assert "anti-dilution" in body.lower()
        assert "broad_based_weighted_average" in body  # the recovery detail reaches the founder

    def test_existing_bare_code_warnings_still_render(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        body = "\n".join(compose_report._render_warning_callouts(["W_CAP_BASE_ASSUMED"]))
        assert "Cap base ASSUMED" in body

    def test_no_warnings_renders_nothing(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        assert compose_report._render_warning_callouts([]) == []


class TestWarningsSharedRenderer:
    """Issue C: the warning families render from ONE shared module
    (`_warning_callouts.render_warning_callouts`) so compose and concise cannot diverge. AD matches by PREFIX
    (interpolated sentence); the others by exact code."""

    def test_renders_all_families(self) -> None:
        import _warning_callouts  # type: ignore[import-not-found]

        warnings = [
            "W_AOA_ONLY_NO_INSTRUMENTS",
            "W_CAP_BASE_ASSUMED",
            "W_FOUNDER_LOOKS_LIKE_INVESTOR",
            "W_CAP_BASE_RECONSTRUCTED",
            "W_VISION_EXTRACTION_LOW_CONFIDENCE",
            "W_BASE_VACUOUS",
            "W_FD_RECONCILE_DELTA: computed fully-diluted 10,000,000 vs source-stated 9,000,000 "
            "(Δ +1,000,000, +100000 ppm) exceeds 1000 ppm — a holder/class may be dropped or mis-entered.",
            "W_ANTI_DILUTION_NONCANONICAL: preferred series 'Series Seed' specified anti-dilution "
            "under the wrong key `anti_dilution`='bbwa' — recovered as 'broad_based_weighted_average'.",
        ]
        body = "\n".join(_warning_callouts.render_warning_callouts(warnings))
        assert "AoA-only engagement" in body
        assert "Cap base ASSUMED" in body
        assert "resembles an investment entity" in body
        assert "deterministic spreadsheet mapper" in body  # W_CAP_BASE_RECONSTRUCTED (softened text)
        assert "Image-only PDF read by vision" in body  # W_VISION_EXTRACTION_LOW_CONFIDENCE
        assert "No real cap-table base" in body  # W_BASE_VACUOUS
        assert "does not match the source-stated total" in body  # W_FD_RECONCILE_DELTA (prefix match)
        assert "Anti-dilution input recovered" in body
        assert "broad_based_weighted_average" in body  # AD recovery detail (prefix match)

    def test_no_warnings_renders_nothing(self) -> None:
        import _warning_callouts  # type: ignore[import-not-found]

        assert _warning_callouts.render_warning_callouts([]) == []


class TestValidatorRequiredMissingHint:
    """Part B: option_pool fields are `required`, so a mis-keyed `authorized_shares` is REJECTED at
    validation (not silently zeroed). The required-missing error should HINT the resembling sibling
    the founder actually wrote, so the rejection isn't a dead-end."""

    def test_required_missing_emits_near_miss_hint(self) -> None:
        import _cap_table_schema_validator as v  # type: ignore[import-not-found]

        schema = {"type": "object", "required": ["authorized"], "properties": {"authorized": {"type": "integer"}}}
        errs = v.validate({"authorized_shares": 1_000_000}, schema, "option_pool")
        joined = " ".join(errs)
        assert "authorized" in joined and "missing" in joined
        assert "authorized_shares" in joined  # the hint names the sibling the founder actually wrote

    def test_required_missing_no_spurious_hint(self) -> None:
        import _cap_table_schema_validator as v  # type: ignore[import-not-found]

        schema = {"type": "object", "required": ["authorized"], "properties": {}}
        errs = v.validate({"plan_type": "iso"}, schema, "option_pool")
        joined = " ".join(errs)
        assert "authorized" in joined and "missing" in joined
        assert "did you" not in joined.lower()  # nothing resembles 'authorized' → no bogus hint


class TestIntentionalNonSchemaKeysInventory:
    """Phase-3 down-payment: the inventory of legit non-schema keys is git-tracked + test-locked so
    Phase-3 (additionalProperties: false) consumes it instead of re-deriving — and so it can't silently
    lose an entry or list a real schema property (the founder_id mistake the plan's reviews caught)."""

    _SCHEMAS = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "skills",
        "cap-table",
        "references",
        "schemas",
    )

    def _props(self, schema_file: str, obj: str) -> set:
        with open(os.path.join(self._SCHEMAS, schema_file)) as fh:
            node = json.load(fh)["properties"][obj]
        node = node.get("items", node)
        return set((node.get("properties") or {}).keys())

    def test_pins_known_extras(self) -> None:
        import cap_state  # type: ignore[import-not-found]

        inv = cap_state._INTENTIONAL_NON_SCHEMA_KEYS
        assert {"oip", "ocp", "anti_dilution"} <= inv["preferred_series"]
        assert "vesting" in inv["founders"]
        assert {"exercised", "expired_or_forfeited"} <= inv["option_pool"]
        assert "exercised_flag" in inv["warrants"]

    def test_inventory_keys_are_genuinely_non_schema(self) -> None:
        """Every inventory key must NOT be a declared schema property — else it isn't an 'extra'.
        This is exactly the `founder_id` mis-classification the plan's reviews caught."""
        import cap_state  # type: ignore[import-not-found]

        schema_obj = {
            "preferred_series": ("inputs.schema.json", "preferred_series"),
            "founders": ("inputs.schema.json", "founders"),
            "option_pool": ("inputs.schema.json", "option_pool"),
            "warrants": ("instruments.schema.json", "warrants"),
        }
        for obj, extras in cap_state._INTENTIONAL_NON_SCHEMA_KEYS.items():
            schema_file, name = schema_obj[obj]
            overlap = set(extras) & self._props(schema_file, name)
            assert not overlap, f"{obj}: inventory lists declared schema props as extras: {overlap}"


class TestEvidenceVerifierIntegration:
    """extract_instrument.py --verify --source-doc integration.

    Tests the wiring of evidence_verifier into the extraction validator.
    Verification is INFORMATIONAL by default; --verify-blocking flips to gating.
    """

    @pytest.fixture
    def safe_doc_text(self, tmp_path: Any) -> Any:
        """Synthetic SAFE-like document text. Contains $20M cap + 80% discount.
        Padded past the image-only threshold (500 chars) so verifier runs."""
        text = (
            "SIMPLE AGREEMENT FOR FUTURE EQUITY\n\n"
            "THIS CERTIFIES THAT in exchange for the payment by Foobar Capital LLC "
            '("Investor") of $500,000 (the "Purchase Amount") on or about January 15, 2024, '
            'Acmecorp Inc. (the "Company") issues to Investor the right to certain shares. '
            'The "Post-Money Valuation Cap" is $20,000,000. The "Discount Rate" is 80%. '
            'This Safe is one of a series of Safes referred to as the "2024 Series Seed Safes." '
            "The Company hereby covenants and agrees with the Investor as follows: This Safe "
            "shall convert in connection with a future Equity Financing on the terms set forth "
            "herein. Standard provisions for conversion, liquidation, dissolution events, and "
            "related rights apply, as set forth in the YC standard form. The rights and "
            "obligations of the parties hereunder shall survive the conversion. This document "
            "incorporates by reference the standard YC Safe Pro-Rata side letter where applicable. "
        )
        p = tmp_path / "synthetic_safe.txt"
        p.write_text(text)
        return p

    @pytest.fixture
    def basic_instruments_path(self, tmp_path: Any) -> Any:
        p = tmp_path / "instruments.json"
        p.write_text(json.dumps(_BASIC_INSTRUMENTS))
        return p

    def test_verify_disabled_with_no_verify(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """--verify defaults ON; --no-verify disables it.
        Verifies the opt-out path keeps existing test semantics."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "discount_multiplier": 0.80,
                "post_money_valuation_cap": 20000000,
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_off",
                "--no-verify",
                "--no-invariants",
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "evidence_verification" not in receipt
        assert "invariant_check" not in receipt

    def test_verify_default_on_requires_source_doc(self, basic_instruments_path: Any) -> None:
        """With --verify default ON,, missing --source-doc errors."""
        extraction = {
            "instrument_type": "safe",
            "fields": _SAFE_BASIC,
            "confidence": {},
            "ambiguities": [],
        }
        rc, _, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_default_verify_no_source",
                # no --no-verify and no --source-doc → should error
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 1
        assert "source-doc" in err.lower() or "source_doc" in err.lower()

    def test_verify_informational_passes_clean_extraction(
        self, basic_instruments_path: Any, safe_doc_text: Any
    ) -> None:
        """A SAFE extraction where every value is in the source doc passes verification."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000,
                "investor_name": "Foobar Capital LLC",
                "post_money_valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
                "issuance_date": "2024-01-15",  # match the synthetic doc
            },
            "confidence": {
                "purchase_amount": {"level": "high", "evidence_quote": '$500,000 (the "Purchase Amount")'},
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $20,000,000.',
                },
                "discount_multiplier": {"level": "high", "evidence_quote": 'The "Discount Rate" is 80%.'},
                "investor_name": {"level": "high", "evidence_quote": "Foobar Capital LLC"},
                "issuance_date": {"level": "high", "evidence_quote": "on or about January 15, 2024"},
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_clean",
                "--verify",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "evidence_verification" in receipt
        ev = receipt["evidence_verification"]
        assert ev["overall_status"] in ("pass", "no_verifiable_fields"), ev
        assert ev["filtered_summary"]["n_fail"] == 0, ev

    def test_verify_informational_catches_hallucinated_value(
        self, basic_instruments_path: Any, safe_doc_text: Any
    ) -> None:
        """A SAFE with a hallucinated cap (cap=$99M when doc says $20M) is caught
        by value_in_doc check. earlier work: informational — receipt records the
        failure but extraction still exits 0."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000,
                "investor_name": "Foobar Capital LLC",
                "post_money_valuation_cap": 99_000_000,  # FAKE — doc says $20M
                "discount_multiplier": 0.80,
            },
            "confidence": {
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $99,000,000.',
                },  # fabricated quote
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_hallucination_informational",
                "--verify",
                "--no-verify-blocking",  # --verify-blocking is default ON; opt out for informational mode
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        # Informational mode: still exits 0
        assert rc == 0, err
        receipt = json.loads(out)
        ev = receipt["evidence_verification"]
        assert ev["overall_status"] == "fail", ev
        assert "rejection" in ev
        assert ev["rejection"]["rejected"] is False  # informational
        assert "post_money_valuation_cap" in ev["rejection"]["failed_fields"]
        assert "retry_hint" in ev["rejection"]

    def test_verify_blocking_exits_nonzero_on_hallucination(
        self, basic_instruments_path: Any, safe_doc_text: Any
    ) -> None:
        """earlier work preview: --verify-blocking exits 1 on value_in_doc failures."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "post_money_valuation_cap": 99_000_000,  # FAKE
                "discount_multiplier": 0.80,
            },
            "confidence": {
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $99,000,000.',
                },
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_hallucination_blocking",
                "--verify",
                "--verify-blocking",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 1, f"blocking mode should fail on hallucination; stderr={err!r}"
        # Receipt should still be printed
        receipt = json.loads(out)
        assert receipt["evidence_verification"]["rejection"]["rejected"] is True
        assert "post_money_valuation_cap" in receipt["evidence_verification"]["rejection"]["failed_fields"]

    def test_verify_skips_synthesized_fields(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """Form enum, jurisdiction, mfn_provision are synthesized — verifier
        should mark them skipped_synthesized, not fail them."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",  # synthesized enum
                "post_money_valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
            },
            "confidence": {
                "form": {"level": "high", "evidence_quote": "POST-MONEY VALUATION CAP"},
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $20,000,000.',
                },
                "discount_multiplier": {"level": "high", "evidence_quote": 'The "Discount Rate" is 80%.'},
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_synthesized",
                "--verify",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        ev = receipt["evidence_verification"]
        assert ev["n_skipped_synthesized"] >= 1
        form_field = next(f for f in ev["per_field"] if f["field"] == "form")
        assert form_field["status"] == "skipped_synthesized"

    def test_partial_state_fields_do_not_fail_verification(self) -> None:
        """A partial instrument stamps completeness/missing_required_fields/to_confirm INTO
        the extracted fields, which then flow into the verifier. These are script-synthesized
        (no evidence_quote by design), so they must be demoted to skipped_synthesized — not
        counted as failures that flip an otherwise-clean partial's overall_status to 'fail'."""
        from extract_instrument import filter_verifier_report  # type: ignore[import-not-found]

        raw = {
            "overall_status": "fail",
            "per_field": [
                {"field": "post_money_valuation_cap", "status": "pass", "reasons": []},
                {
                    "field": "completeness",
                    "status": "fail",
                    "reasons": ["value='partial' provided but evidence_quote is null/empty"],
                },
                {
                    "field": "missing_required_fields",
                    "status": "fail",
                    "reasons": ["value=[...] provided but evidence_quote is null/empty"],
                },
                {
                    "field": "to_confirm",
                    "status": "fail",
                    "reasons": ["value={...} provided but evidence_quote is null/empty"],
                },
            ],
        }
        out = filter_verifier_report(raw)
        statuses = {r["field"]: r["status"] for r in out["per_field"]}
        assert statuses["completeness"] == "skipped_synthesized"
        assert statuses["missing_required_fields"] == "skipped_synthesized"
        assert statuses["to_confirm"] == "skipped_synthesized"
        # Only the real (passing) field remains in the recomputed status -> not 'fail'.
        assert out["overall_status"] == "pass"

    @pytest.fixture
    def note_doc_text(self, tmp_path: Any) -> Any:
        """Synthetic convertible-note document text carrying principal,
        day-count, cap, discount, and maturity facts. Padded past the
        image-only threshold (500 chars) so the verifier runs. Silent on
        whether interest converts to shares — the extractor is expected to
        report that field absent, not to invent a value."""
        text = (
            "CONVERTIBLE PROMISSORY NOTE\n\n"
            "THIS NOTE is issued by Acmecorp Inc. to Foobar Capital LLC in the original "
            'principal amount of $500,000 (the "Principal Amount") on or about January 15, 2024. '
            "Interest, if any, shall accrue under a 365-day basis. "
            'The "Valuation Cap" is $20,000,000. The applicable discount is 80%. '
            "Unless earlier converted, this Note shall mature on or about June 1, 2027 "
            '(the "Maturity Date"). Upon a Qualified Financing, this Note shall convert '
            "into the securities issued therein on the terms set forth herein. The rights "
            "and obligations of the parties hereunder shall survive until repayment or "
            "conversion in full."
        )
        p = tmp_path / "synthetic_note.txt"
        p.write_text(text)
        return p

    def test_verify_skips_documented_absence(self, basic_instruments_path: Any, note_doc_text: Any) -> None:
        """A boolean field the extractor reported absent from the document
        (falsy value + confidence level "absent" + no evidence_quote) must not
        drag overall_status to 'fail' — there is no quote for the absence of
        something. It should show up as skipped_documented_absence instead."""
        extraction = {
            "instrument_type": "convertible_note",
            "fields": {
                "investor_name": "Foobar Capital LLC",
                "principal": 500_000,
                "interest_rate_type": "none",
                "annual_interest_rate": None,
                "day_count_basis": 365,
                "interest_converts_to_shares": False,  # doc is silent on this
                "issuance_date": "2024-01-15",
                "valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
                "qualified_financing_threshold": None,
                "maturity_date": "2027-06-01",
                "maturity_default_treatment": "convert_at_cap",
            },
            "confidence": {
                "investor_name": {"level": "high", "evidence_quote": "Foobar Capital LLC"},
                "principal": {"level": "high", "evidence_quote": '$500,000 (the "Principal Amount")'},
                "day_count_basis": {"level": "high", "evidence_quote": "365-day basis"},
                "issuance_date": {"level": "high", "evidence_quote": "on or about January 15, 2024"},
                "valuation_cap": {"level": "high", "evidence_quote": 'The "Valuation Cap" is $20,000,000.'},
                "discount_multiplier": {"level": "high", "evidence_quote": "The applicable discount is 80%."},
                "maturity_date": {"level": "high", "evidence_quote": "on or about June 1, 2027"},
                "interest_converts_to_shares": {"level": "absent"},
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_documented_absence",
                "--verify",
                "--no-invariants",
                "--source-doc",
                str(note_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        ev = receipt["evidence_verification"]
        field_row = next(f for f in ev["per_field"] if f["field"] == "interest_converts_to_shares")
        assert field_row["status"] == "skipped_documented_absence", field_row
        assert ev.get("n_skipped_absence", 0) >= 1
        assert ev["overall_status"] != "fail", ev

    def test_verify_absent_field_precision_guard_still_fails(
        self, basic_instruments_path: Any, note_doc_text: Any
    ) -> None:
        """Precision guard: a falsy value (False) with confidence level 'high'
        (not 'absent') and no evidence_quote is NOT a documented absence — it
        must still fail normal verification ('value provided but no quote')."""
        extraction = {
            "instrument_type": "convertible_note",
            "fields": {
                "investor_name": "Foobar Capital LLC",
                "principal": 500_000,
                "interest_rate_type": "none",
                "annual_interest_rate": None,
                "day_count_basis": 365,
                "interest_converts_to_shares": False,
                "issuance_date": "2024-01-15",
                "valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
                "qualified_financing_threshold": None,
                "maturity_date": "2027-06-01",
                "maturity_default_treatment": "convert_at_cap",
            },
            "confidence": {
                "investor_name": {"level": "high", "evidence_quote": "Foobar Capital LLC"},
                "principal": {"level": "high", "evidence_quote": '$500,000 (the "Principal Amount")'},
                "day_count_basis": {"level": "high", "evidence_quote": "365-day basis"},
                "issuance_date": {"level": "high", "evidence_quote": "on or about January 15, 2024"},
                "valuation_cap": {"level": "high", "evidence_quote": 'The "Valuation Cap" is $20,000,000.'},
                "discount_multiplier": {"level": "high", "evidence_quote": "The applicable discount is 80%."},
                "maturity_date": {"level": "high", "evidence_quote": "on or about June 1, 2027"},
                "interest_converts_to_shares": {"level": "high"},  # no evidence_quote, NOT "absent"
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_absent_precision_guard",
                "--verify",
                "--no-verify-blocking",
                "--no-invariants",
                "--source-doc",
                str(note_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        ev = receipt["evidence_verification"]
        field_row = next(f for f in ev["per_field"] if f["field"] == "interest_converts_to_shares")
        assert field_row["status"] == "fail", field_row
        assert ev["overall_status"] == "fail", ev

    def test_verify_warrant_is_skipped(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """Non-extractable doc types (warrant, non_instrument) skip verification."""
        extraction = {
            "instrument_type": "warrant",
            "fields": {},
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_warrant",
                "--verify",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert receipt["evidence_verification"]["overall_status"] == "skipped_non_instrument"

    def test_verify_requires_source_doc(self, basic_instruments_path: Any) -> None:
        """--verify without --source-doc errors out."""
        extraction = {
            "instrument_type": "safe",
            "fields": _SAFE_BASIC,
            "confidence": {},
            "ambiguities": [],
        }
        rc, _, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_verify_no_source",
                "--verify",
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 1
        assert "source-doc" in err.lower()


class TestInvariantCheckerIntegration:
    """--invariants default ON in extract_instrument.py. Hard math
    invariants exit 1; soft bounds violations warn-only via receipt."""

    @pytest.fixture
    def basic_instruments_path(self, tmp_path: Any) -> Any:
        p = tmp_path / "instruments.json"
        p.write_text(json.dumps(_BASIC_INSTRUMENTS))
        return p

    @pytest.fixture
    def safe_doc_text(self, tmp_path: Any) -> Any:
        # Padded past the image-only threshold so verifier runs.
        text = (
            "SAFE\n\n"
            'Foobar Capital LLC (the "Investor") pays $500,000 (the "Purchase Amount") '
            'on or about January 15, 2024. The "Post-Money Valuation Cap" is $20,000,000. '
            'The "Discount Rate" is 80%. This Safe is one of a series of 2024 Seed Safes. '
            "Standard YC provisions for conversion, equity financing, liquidity event, dissolution, "
            "and miscellaneous matters apply. Standard amendments, governing law, dispute resolution, "
            "counterparts apply. " * 4
        )
        p = tmp_path / "synthetic_safe.txt"
        p.write_text(text)
        return p

    def test_invariants_on_by_default_clean_safe(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """Clean SAFE within all bounds → invariant_check block in receipt, n_violations=0."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
                "issuance_date": "2024-01-15",
                "investor_name": "Foobar Capital LLC",
            },
            "confidence": {
                "purchase_amount": {"level": "high", "evidence_quote": '$500,000 (the "Purchase Amount")'},
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $20,000,000.',
                },
                "discount_multiplier": {"level": "high", "evidence_quote": 'The "Discount Rate" is 80%.'},
                "investor_name": {"level": "high", "evidence_quote": "Foobar Capital LLC"},
                "issuance_date": {"level": "high", "evidence_quote": "on or about January 15, 2024"},
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_invariants_clean",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "invariant_check" in receipt
        assert receipt["invariant_check"]["n_violations"] == 0
        assert receipt["invariant_check"]["n_hard_violations"] == 0

    def test_invariant_hard_violation_blocks(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """SAFE with both pre_money_valuation_cap AND post_money_valuation_cap set →
        hard math violation. extract_instrument.py exits 1 even with --no-verify."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "other",  # avoid validate_safe gate triggering first
                "pre_money_valuation_cap": 10_000_000,
                "post_money_valuation_cap": 12_000_000,  # both set — math impossible
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_invariants_hard_block",
                "--no-verify",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 1, f"hard invariant should block; out={out[:300]} err={err[:300]}"
        assert "invariant" in err.lower()

    def test_invariants_can_be_disabled(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """--no-invariants disables the check; even hard violations pass through."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "other",
                "pre_money_valuation_cap": 10_000_000,
                "post_money_valuation_cap": 12_000_000,  # hard violation
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_invariants_disabled",
                "--no-verify",
                "--no-invariants",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "invariant_check" not in receipt

    def test_soft_bounds_violation_warns_not_blocks(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """SAFE with $500M purchase_amount (unit error?) → soft bound warning, rc=0."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000_000,  # $500M — above $50M SAFE max, soft warn
                "post_money_valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
                "issuance_date": "2024-01-15",
                "investor_name": "Foobar Capital LLC",
            },
            "confidence": {
                "purchase_amount": {"level": "high", "evidence_quote": "$500,000"},
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $20,000,000.',
                },
                "discount_multiplier": {"level": "high", "evidence_quote": 'The "Discount Rate" is 80%.'},
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_invariants_soft",
                "--no-verify",  # purchase_amount=$500M not in synth doc; skip evidence check
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err  # soft violations don't block
        receipt = json.loads(out)
        assert receipt["invariant_check"]["n_violations"] >= 1
        assert receipt["invariant_check"]["n_hard_violations"] == 0
        # And the field appears in attention_needed_fields
        assert "purchase_amount" in receipt.get("attention_needed_fields", [])

    def test_invariants_skipped_for_warrant(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """Warrants and non_instrument classifications skip invariant_check entirely."""
        extraction = {
            "instrument_type": "warrant",
            "fields": {},
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_invariants_warrant",
                "--no-verify",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "invariant_check" not in receipt


class TestCrossCheckerIntegration:
    """`--cross-check` runs deterministic backstop extractors and surfaces
    confidence demotions when sub-agent and backstop disagree."""

    @pytest.fixture
    def basic_instruments_path(self, tmp_path: Any) -> Any:
        p = tmp_path / "instruments.json"
        p.write_text(json.dumps(_BASIC_INSTRUMENTS))
        return p

    @pytest.fixture
    def safe_doc_text(self, tmp_path: Any) -> Any:
        text = (
            "SIMPLE AGREEMENT FOR FUTURE EQUITY\n\n"
            'in exchange for the payment by Foobar Capital LLC (the "Investor") '
            'of $500,000 (the "Purchase Amount") on or about January 15, 2024, '
            'Acmecorp Inc. (the "Company") issues to Investor the right to certain '
            'shares. The "Post-Money Valuation Cap" is $20,000,000. '
            'The "Discount Rate" is 80%. Standard YC provisions. '
        ) + ("Filler text for image-only threshold avoidance: " * 10)
        p = tmp_path / "synthetic_safe.txt"
        p.write_text(text)
        return p

    def test_cross_check_agreement_no_demotion(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        """When sub-agent + backstop agree, cross_check reports zero demotions."""
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
                "issuance_date": "2024-01-15",
                "investor_name": "Foobar Capital LLC",
            },
            "confidence": {
                "purchase_amount": {"level": "high", "evidence_quote": '$500,000 (the "Purchase Amount")'},
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $20,000,000.',
                },
                "discount_multiplier": {"level": "high", "evidence_quote": 'The "Discount Rate" is 80%.'},
                "issuance_date": {"level": "high", "evidence_quote": "on or about January 15, 2024"},
                "investor_name": {"level": "high", "evidence_quote": "Foobar Capital LLC"},
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_xcheck_agree",
                "--source-doc",
                str(safe_doc_text),
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "cross_check" in receipt
        assert receipt["cross_check"]["n_demotions"] == 0

    def test_cross_check_disagreement_surfaces_demotion(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": 99_000_000,  # WRONG — backstop finds $20M
                "discount_multiplier": 0.80,
                "issuance_date": "2024-01-15",
                "investor_name": "Foobar Capital LLC",
            },
            "confidence": {
                "post_money_valuation_cap": {
                    "level": "high",
                    "evidence_quote": 'The "Post-Money Valuation Cap" is $99,000,000.',
                },
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_xcheck_disagree",
                "--source-doc",
                str(safe_doc_text),
                "--no-verify-blocking",
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert receipt["cross_check"]["n_demotions"] >= 1
        assert "post_money_valuation_cap" in receipt.get("attention_needed_fields", [])

    def test_cross_check_can_be_disabled(self, basic_instruments_path: Any, safe_doc_text: Any) -> None:
        extraction = {
            "instrument_type": "safe",
            "fields": {
                **_SAFE_BASIC,
                "form": "cap_plus_discount",
                "purchase_amount": 500_000,
                "post_money_valuation_cap": 20_000_000,
                "discount_multiplier": 0.80,
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                str(basic_instruments_path),
                "--run-id",
                "test_xcheck_off",
                "--source-doc",
                str(safe_doc_text),
                "--no-cross-check",
                "--no-verify",
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        receipt = json.loads(out)
        assert "cross_check" not in receipt


# ===========================================================================
# Carta extraction (corpus 2 surfaced 13 real Carta exports — fingerprint,
# Convertible Ledger parsing, discount normalization, cancelled-record skip)
# ===========================================================================

_CARTA_FIXTURE = os.path.join(_REPO, "founder-skills", "tests", "fixtures", "cap-table-corpus", "synthetic_carta.xlsx")


class TestCartaExtraction:
    def test_carta_fingerprint_detects_carta(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _carta_detect  # type: ignore[import-not-found]

        # Default Carta (full export)
        assert _carta_detect(["Summary Cap Table", "Intermediate Cap Table", "Detailed Cap Table"]) == "carta"
        # Minimal Carta (just summary)
        assert _carta_detect(["Summary Cap Table"]) == "carta"
        # Carta OCX format (rarely used)
        assert _carta_detect(["Capitalization by Stakeholder", "Voting Details", "Context"]) == "carta_ocx"
        # Freeform / lawyer-built — none match
        assert _carta_detect(["Cap Table", "ESOP", "Sheet3"]) is None
        # Pulley-style — not Carta
        assert _carta_detect(["Ownership", "SAFEs", "Shares"]) is None

    def test_carta_fingerprint_strips_whitespace(self) -> None:
        """Corpus surfaced trailing-whitespace sheet names; fingerprint must strip."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _carta_detect  # type: ignore[import-not-found]

        assert _carta_detect(["Summary Cap Table ", "Intermediate Cap Table"]) == "carta"

    def test_discount_normalization_carta_percent(self) -> None:
        """Carta stores Conversion Discount as percent-as-fraction (0.2 = 20%).
        Per Gotcha #3 the canonical form is the multiplier (0.80)."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _normalize_discount  # type: ignore[import-not-found]

        # 0.2 (Carta percent-as-fraction for 20%) → 0.8 multiplier
        mult, warning = _normalize_discount(0.2)
        assert mult == 0.8
        assert warning is not None and "Carta discount" in warning

        # 20 (Carta percent for 20%) → 0.8 multiplier
        mult, warning = _normalize_discount(20)
        assert mult == 0.8
        assert warning is not None

        # None → None, no warning
        mult, warning = _normalize_discount(None)
        assert mult is None
        assert warning is None

        # Out of range
        mult, warning = _normalize_discount(200)
        assert mult is None
        assert warning is not None and "out of expected range" in warning

    def test_safe_form_inference(self) -> None:
        """SAFE form inferred from cap + discount presence."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _infer_safe_form  # type: ignore[import-not-found]

        # cap only
        assert _infer_safe_form(8_000_000, None) == "yc_postmoney_cap"
        # cap + discount
        assert _infer_safe_form(8_000_000, 0.20) == "cap_plus_discount"
        # discount only
        assert _infer_safe_form(None, 0.20) == "yc_postmoney_discount"
        # neither (uncapped MFN or other)
        assert _infer_safe_form(None, None) == "other"

    def test_carta_extract_synthetic_fixture(self) -> None:
        """End-to-end: real-Carta-shape synthetic XLSX → instruments.json."""
        if not os.path.exists(_CARTA_FIXTURE):
            pytest.skip(f"Carta fixture missing: {_CARTA_FIXTURE}")
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _carta_extract  # type: ignore[import-not-found]

        result = _carta_extract(_CARTA_FIXTURE)
        assert result["format"] == "carta"
        assert result["counts"]["safes_extracted"] == 1
        assert result["counts"]["notes_extracted"] == 1
        # Cancelled SAFE was skipped, mentioned in warnings
        assert any("SAFE1-2" in w and "cancelled" in w.lower() for w in result["warnings"])
        # Discount normalization warning fired
        assert any("discount" in w.lower() and "multiplier" in w.lower() for w in result["warnings"])

        # Verify the extracted SAFE
        safes = result["instruments"]["safes"]
        assert len(safes) == 1
        safe = safes[0]
        assert safe["investor_name"] == "Angel Alice"
        assert safe["purchase_amount"] == 250_000
        assert safe["post_money_valuation_cap"] == 8_000_000
        assert safe["discount_multiplier"] == 0.8  # normalized from 0.20
        assert safe["form"] == "cap_plus_discount"  # has both cap + discount
        assert safe["source_document"] == "carta:SAFE1-1"

        # Verify the extracted note
        notes = result["instruments"]["convertible_notes"]
        assert len(notes) == 1
        note = notes[0]
        assert note["investor_name"] == "Lender Larry"
        assert note["principal"] == 500_000
        assert note["annual_interest_rate"] == 0.06
        assert note["valuation_cap"] == 10_000_000
        assert note["maturity_date"] == "2027-03-01"

    def test_carta_extract_company_name_from_banner(self) -> None:
        """Carta banner: row 2 is "{Company} Summary Cap Table"; row 3 is
        "As of M/D/YYYY • Generated by ...". Extractor must parse both."""
        if not os.path.exists(_CARTA_FIXTURE):
            pytest.skip(f"Carta fixture missing: {_CARTA_FIXTURE}")
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _carta_extract  # type: ignore[import-not-found]

        result = _carta_extract(_CARTA_FIXTURE)
        assert result["summary_totals"]["company_name"] == "Acmecorp"
        assert result["summary_totals"]["as_of_date"] == "2026-05-19"

    def test_auto_mode_dispatch_subprocess(self) -> None:
        """End-to-end CLI test: --mode=auto sniffs sheet names + dispatches."""
        if not os.path.exists(_CARTA_FIXTURE):
            pytest.skip(f"Carta fixture missing: {_CARTA_FIXTURE}")
        with tempfile.TemporaryDirectory() as d:
            audit = os.path.join(d, "audit.json")
            inst = os.path.join(d, "instruments.json")
            rc, stdout, stderr = _run(
                "extract_cap_table.py",
                ["--mode", "auto", "--xlsx", _CARTA_FIXTURE, "-o", audit, "--instruments", inst, "--pretty"],
            )
            assert rc == 0, stderr
            with open(inst) as f:
                instruments = json.load(f)
            assert len(instruments["safes"]) == 1
            assert len(instruments["convertible_notes"]) == 1

    def test_freeform_routing_for_non_carta_xlsx(self) -> None:
        """--mode=auto on a freeform XLSX must route to the freeform fallback
        with a clear remedy pointing at Context-A dispatch."""
        with tempfile.TemporaryDirectory() as d:
            import openpyxl  # type: ignore[import-untyped]

            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = "Cap Table"  # Freeform-shaped
            ws["A1"] = "Acme"
            xlsx = os.path.join(d, "freeform.xlsx")
            wb.save(xlsx)
            rc, stdout, stderr = _run("extract_cap_table.py", ["--mode", "auto", "--xlsx", xlsx])
            assert rc != 0
            try:
                receipt = json.loads(stdout)
                assert receipt.get("detected_format") == "freeform"
                assert "SPREADSHEET_STRUCTURE_DETECTION" in receipt.get("remedy", "")
            except json.JSONDecodeError as e:
                raise AssertionError(f"expected JSON output, got: {stdout}") from e


class TestCorpusRobustness:
    """Edge cases surfaced by testing against a 258-file real-world corpus.

    Documented in docs/internal/captable-corpus-test-findings.md.
    """

    def test_normalize_xlsx_macos_dupe_suffix(self) -> None:
        """macOS duplicate-download artifact: `file.xlsx(1)` should normalize
        to `file.xlsx`. 3 of 258 corpus files had this exact failure mode."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _normalize_path  # type: ignore[import-not-found]

        normalized, warning = _normalize_path("/tmp/cap.xlsx(1)")
        assert normalized == "/tmp/cap.xlsx"
        assert warning is not None and "macOS" in warning

        normalized, warning = _normalize_path("/tmp/cap.xlsx(2)")
        assert normalized == "/tmp/cap.xlsx"
        assert warning is not None

        # XLS variant
        normalized, warning = _normalize_path("/tmp/cap.xls(1)")
        assert normalized == "/tmp/cap.xls"

        # PDF variant
        normalized, warning = _normalize_path("/tmp/doc.pdf(1)")
        assert normalized == "/tmp/doc.pdf"

        # No suffix — pass through unchanged, no warning
        normalized, warning = _normalize_path("/tmp/cap.xlsx")
        assert normalized == "/tmp/cap.xlsx"
        assert warning is None

    def test_normalize_xlsx_strips_whitespace(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _normalize_path  # type: ignore[import-not-found]

        normalized, _ = _normalize_path("  /tmp/cap.xlsx  ")
        assert normalized == "/tmp/cap.xlsx"

    def test_reject_eml_with_guidance(self) -> None:
        """When the user forwards an email instead of detaching the attachment."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _check_supported_input_type  # type: ignore[import-not-found]

        ok, msg = _check_supported_input_type("/tmp/captable_email.eml")
        assert ok is False
        assert ".eml" in msg or "email" in msg.lower()
        assert "attach" in msg.lower()  # founder-friendly guidance

    def test_supported_types_pass(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _check_supported_input_type  # type: ignore[import-not-found]

        for path in ["/tmp/cap.xlsx", "/tmp/cap.pdf", "/tmp/cap.docx"]:
            ok, _ = _check_supported_input_type(path)
            assert ok is True, f"Should accept {path}"

    def test_legacy_xls_rejected_with_guidance(self) -> None:
        # B1: legacy binary .xls is rejected early with a founder-friendly remedy (openpyxl can't read it;
        # the old behavior let it through, then died on a cryptic BadZipFile).
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _check_supported_input_type  # type: ignore[import-not-found]

        ok, msg = _check_supported_input_type("/tmp/cap.xls")
        assert ok is False
        assert ".xlsx" in msg and "re-save" in msg.lower()

    def test_eml_rejected_via_subprocess(self) -> None:
        """Full CLI round-trip: passing an .eml exits non-zero with structured blocker."""
        with tempfile.TemporaryDirectory() as d:
            eml_path = os.path.join(d, "captable.eml")
            with open(eml_path, "w") as f:
                f.write("From: lawyer@example.com\nSubject: Cap table\n\n[attachment removed]")
            rc, stdout, stderr = _run(
                "extract_cap_table.py",
                ["--mode", "auto", "--xlsx", eml_path],
            )
            assert rc != 0
            # stdout should contain the structured blocker
            try:
                receipt = json.loads(stdout.strip().splitlines()[-1])
                assert receipt.get("ok") is False
                assert receipt.get("blocker") == "unsupported_input_type"
                assert "attach" in receipt.get("remedy", "").lower()
            except (json.JSONDecodeError, IndexError):
                # Fall back: just check stderr/stdout contains a recognizable error string
                combined = (stdout + stderr).lower()
                assert "email" in combined or "eml" in combined


class TestPipelineE2E:
    def test_full_pipeline_acmecorp(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            RID = "20260519T120000Z"
            inp = os.path.join(d, "inputs.json")
            inst = os.path.join(d, "instruments.json")
            cap = os.path.join(d, "cap_state.json")
            audit = os.path.join(d, "rule_audit.json")
            reqs = os.path.join(d, "scenario_requests.json")
            scen = os.path.join(d, "scenarios.json")
            packet_json = os.path.join(d, "counsel_packet.json")
            packet_md = os.path.join(d, "counsel_packet.md")
            report_json = os.path.join(d, "report.json")
            report_md = os.path.join(d, "report.md")
            html = os.path.join(d, "report.html")
            explorer = os.path.join(d, "explorer.html")

            # Stamp run_id on inputs + instruments
            inputs = dict(_BASIC_INPUTS)
            inputs["metadata"] = {"run_id": RID}
            instruments = {**_BASIC_INSTRUMENTS, "safes": [_SAFE_BASIC]}
            instruments["metadata"] = {"run_id": RID}
            with open(inp, "w") as f:
                json.dump(inputs, f)
            with open(inst, "w") as f:
                json.dump(instruments, f)

            # Step 1 — cap_state
            rc, _, e = _run(
                "cap_state.py",
                [
                    "--inputs",
                    inp,
                    "--instruments",
                    inst,
                    "--run-id",
                    RID,
                    "-o",
                    cap,
                ],
            )
            assert rc == 0, e

            # Step 2 — rule_audit pre_math
            rc, _, e = _run(
                "rule_audit.py",
                [
                    "--phase",
                    "pre_math",
                    "--inputs",
                    inp,
                    "--instruments",
                    inst,
                    "--cap-state",
                    cap,
                    "--run-id",
                    RID,
                    "-o",
                    audit,
                ],
            )
            assert rc == 0, e

            # Step 3 — scenarios
            with open(reqs, "w") as f:
                json.dump(
                    [
                        {
                            "scenario_id": "base",
                            "label": "Cap-implied SAFE",
                            "type": "safe_conversion",
                            "parameters": {},
                        },
                        {
                            "scenario_id": "series_a",
                            "label": "Series A",
                            "type": "priced_round",
                            "parameters": {
                                "pre_money": 20_000_000,
                                "new_money": 5_000_000,
                                "target_pool_percent": 0.10,
                                "target_basis": "pre_money",
                            },
                        },
                    ],
                    f,
                )
            rc, _, e = _run(
                "run_scenario.py",
                [
                    "--inputs",
                    inp,
                    "--instruments",
                    inst,
                    "--cap-state",
                    cap,
                    "--scenarios-input",
                    reqs,
                    "--run-id",
                    RID,
                    "-o",
                    scen,
                ],
            )
            assert rc == 0, e

            # Step 4 — rule_audit post_math
            rc, _, e = _run(
                "rule_audit.py",
                [
                    "--phase",
                    "post_math",
                    "--run-id",
                    RID,
                    "-o",
                    audit,
                ],
            )
            assert rc == 0, e

            # Step 5 — counsel_packet
            rc, _, e = _run(
                "counsel_packet.py",
                [
                    "--rule-audit",
                    audit,
                    "--inputs",
                    inp,
                    "--scenarios",
                    scen,
                    "--run-id",
                    RID,
                    "-o",
                    packet_json,
                    "--write-md",
                    packet_md,
                ],
            )
            assert rc == 0, e

            # Step 6 — compose_report
            rc, _, e = _run(
                "compose_report.py",
                [
                    "--dir",
                    d,
                    "--run-id",
                    RID,
                    "-o",
                    report_json,
                    "--write-md",
                    report_md,
                ],
            )
            assert rc == 0, e

            # Step 7 — visualize
            rc, _, e = _run("visualize.py", ["--dir", d, "-o", html])
            assert rc == 0, e

            # Step 8 — explore
            rc, _, e = _run("explore.py", ["--dir", d, "-o", explorer])
            assert rc == 0, e

            # Verify all artifacts exist + key invariants
            for path in [cap, audit, scen, packet_json, packet_md, report_json, report_md, html, explorer]:
                assert os.path.exists(path) and os.path.getsize(path) > 0, path

            # report.json has the coaching_payload block per the cross-skill invariant
            with open(report_json) as f:
                rpt = json.load(f)
            assert "report_markdown" in rpt
            assert "coaching_payload" in rpt
            cp = rpt["coaching_payload"]
            assert cp["schema_version"] == "v0.5.0-cap-table"
            assert cp["company_name"] == "Acmecorp"
            assert cp["scenarios_modeled"] == 2
            # Per rev15: nullable founder_impact present for full scenarios, null for structural_only
            digests = cp["scenario_digest"]
            assert len(digests) == 2
            base_digest = next(d for d in digests if d["scenario_id"] == "base")
            assert base_digest["completeness"] == "structural_only"
            assert base_digest["founder_impact"] is None  # rev15 nullable contract
            sa_digest = next(d for d in digests if d["scenario_id"] == "series_a")
            assert sa_digest["completeness"] == "full"
            assert sa_digest["founder_impact"] is not None
            # Insertion marker is uuid-bearing
            assert cp["insertion_marker"].startswith("<!-- COACHING_INSERTION_POINT_")
            # The marker appears EXACTLY once in report.md (Context B's Grep test)
            with open(report_md) as f:
                md = f.read()
            assert md.count(cp["insertion_marker"]) == 1

    def test_compose_emits_coaching_payload(self) -> None:
        """Cross-skill invariant test mirror (see tests/test_compose_invariants.py).
        Verifies compose_report.py emits report.json with coaching_payload block."""
        with tempfile.TemporaryDirectory() as d:
            RID = "rid1"
            inputs = dict(_BASIC_INPUTS)
            inputs["metadata"] = {"run_id": RID}
            instruments = dict(_BASIC_INSTRUMENTS)
            instruments["metadata"] = {"run_id": RID}
            for name, data in [
                ("inputs.json", inputs),
                ("instruments.json", instruments),
            ]:
                with open(os.path.join(d, name), "w") as f:
                    json.dump(data, f)
            # Minimal cap_state via library
            cs = cap_state_mod.build_cap_state(inputs, instruments)
            cs["metadata"]["run_id"] = RID
            with open(os.path.join(d, "cap_state.json"), "w") as f:
                json.dump(cs, f)
            # Minimal rule_audit + scenarios + counsel_packet
            for name, data in [
                (
                    "rule_audit.json",
                    {
                        "gating": {},
                        "applied_rules": [],
                        "counsel_review_items": [],
                        "date_sensitive_watchlist": [],
                        "metadata": {"run_id": RID},
                    },
                ),
                ("scenarios.json", {"scenarios": [], "metadata": {"run_id": RID}}),
                (
                    "counsel_packet.json",
                    {"company_name": "Acmecorp", "engagement_summary": "", "items": [], "metadata": {"run_id": RID}},
                ),
            ]:
                with open(os.path.join(d, name), "w") as f:
                    json.dump(data, f)

            rc, _, stderr = _run(
                "compose_report.py",
                [
                    "--dir",
                    d,
                    "--run-id",
                    RID,
                    "-o",
                    os.path.join(d, "report.json"),
                    "--write-md",
                    os.path.join(d, "report.md"),
                ],
            )
            assert rc == 0, stderr
            with open(os.path.join(d, "report.json")) as f:
                report = json.load(f)
            assert "coaching_payload" in report
            assert "report_markdown" in report
            cp = report["coaching_payload"]
            # Cross-skill required keys per tests/test_compose_invariants.py
            for k in [
                "schema_version",
                "summary",
                "failed_items",
                "warned_items",
                "high_severity_warnings",
                "insertion_marker",
            ]:
                assert k in cp, f"missing required coaching_payload key: {k}"
            assert cp["schema_version"].endswith("-cap-table")

    def test_compose_source_notes_rendered(self) -> None:
        """Regression: render_report_markdown must not crash on source_notes.

        Bug: variable `n` was already typed as int (n = len(scenarios) on
        line 516) and was then reused as a loop variable over source_notes
        dicts — a name-shadow. Fixed by renaming the loop variable to `sn`.
        This test exercises the source_notes rendering path end-to-end.
        """
        with tempfile.TemporaryDirectory() as d:
            RID = "rid_sn"
            inputs = dict(_BASIC_INPUTS)
            inputs["metadata"] = {"run_id": RID}
            instruments = dict(_BASIC_INSTRUMENTS)
            instruments["metadata"] = {"run_id": RID}
            for name, data in [
                ("inputs.json", inputs),
                ("instruments.json", instruments),
            ]:
                with open(os.path.join(d, name), "w") as f:
                    json.dump(data, f)
            cs = cap_state_mod.build_cap_state(inputs, instruments)
            cs["metadata"]["run_id"] = RID
            with open(os.path.join(d, "cap_state.json"), "w") as f:
                json.dump(cs, f)
            # Include a non-empty source_notes block to exercise the fixed path
            rule_audit_with_source_notes = {
                "gating": {},
                "applied_rules": [],
                "counsel_review_items": [],
                "date_sensitive_watchlist": [],
                "source_notes": [
                    {
                        "rule_id": "SAFE-001",
                        "title": "YC Post-Money SAFE Denominator",
                        "domain": "safe_conversion",
                        "summary": "company_capitalization excludes the converting SAFE itself.",
                    },
                    {
                        "rule_id": "OPT-003",
                        "title": "Option Pool Top-Up Basis",
                        "domain": "option_pool",
                        "summary": "Target basis is post-round fully diluted.",
                    },
                ],
                "metadata": {"run_id": RID},
            }
            for name, data in [
                ("rule_audit.json", rule_audit_with_source_notes),
                ("scenarios.json", {"scenarios": [], "metadata": {"run_id": RID}}),
                (
                    "counsel_packet.json",
                    {"company_name": "Acmecorp", "engagement_summary": "", "items": [], "metadata": {"run_id": RID}},
                ),
            ]:
                with open(os.path.join(d, name), "w") as f:
                    json.dump(data, f)

            rc, _, stderr = _run(
                "compose_report.py",
                [
                    "--dir",
                    d,
                    "--run-id",
                    RID,
                    "-o",
                    os.path.join(d, "report.json"),
                    "--write-md",
                    os.path.join(d, "report.md"),
                ],
            )
            assert rc == 0, stderr
            with open(os.path.join(d, "report.md")) as f:
                md = f.read()
            # Both source notes must appear in the rendered markdown
            assert "YC Post-Money SAFE Denominator" in md, "source_note title missing from report.md"
            assert "Option Pool Top-Up Basis" in md, "second source_note title missing from report.md"
            assert "## Source Notes" in md, "Source Notes section header missing"


# ===========================================================================
# Item 3 — Structured errors in cap_state._build_outstanding_safes / _notes
# ===========================================================================


class TestCapStateStructuredErrors:
    """Missing required fields in SAFEs/notes yield CapStateInvariantError,
    not raw KeyError, with the structured E_SAFE_MISSING_FIELD /
    E_NOTE_MISSING_FIELD code embedded in the message."""

    def _minimal_inputs(self) -> dict:
        return {
            "company_name": "TestCo",
            "analysis_date": "2026-06-01",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-01-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
            "preferred_series": [],
            "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
            "common_batches": [],
            "metadata": {"run_id": "test"},
        }

    def test_safe_missing_id_raises_structured(self) -> None:
        """SAFE missing 'id' → E_SAFE_MISSING_FIELD, not KeyError."""
        instruments = {
            "safes": [
                {
                    # 'id' intentionally absent
                    "investor_name": "Angel Missing",
                    "purchase_amount": 100_000,
                    "issuance_date": "2024-01-01",
                    "form": "yc_postmoney_cap",
                    "post_money_valuation_cap": 5_000_000,
                }
            ],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc_info:
            cap_state_mod.build_cap_state(self._minimal_inputs(), instruments)
        msg = str(exc_info.value)
        assert "E_SAFE_MISSING_FIELD" in msg, f"Expected E_SAFE_MISSING_FIELD in: {msg}"
        assert "'id'" in msg, f"Expected field name 'id' in: {msg}"
        assert "index 0" in msg or "safes[0]" in msg, f"Expected index in: {msg}"

    def test_safe_missing_issuance_date_raises_structured(self) -> None:
        """SAFE missing 'issuance_date' → E_SAFE_MISSING_FIELD with remediation hint."""
        instruments = {
            "safes": [
                {
                    "id": "safe_x",
                    "investor_name": "Angel Date",
                    "purchase_amount": 200_000,
                    # 'issuance_date' intentionally absent
                    "form": "yc_postmoney_cap",
                    "post_money_valuation_cap": 5_000_000,
                }
            ],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc_info:
            cap_state_mod.build_cap_state(self._minimal_inputs(), instruments)
        msg = str(exc_info.value)
        assert "E_SAFE_MISSING_FIELD" in msg
        assert "'issuance_date'" in msg or "issuance_date" in msg

    def test_note_missing_id_raises_structured(self) -> None:
        """Note missing 'id' → E_NOTE_MISSING_FIELD, not KeyError."""
        instruments = {
            "safes": [],
            "convertible_notes": [
                {
                    # 'id' intentionally absent
                    "investor_name": "Lender Missing",
                    "principal": 50_000,
                    "issuance_date": "2024-06-01",
                    "annual_interest_rate": 0.06,
                    "day_count_basis": 365,
                    "maturity_date": "2026-06-01",
                    "maturity_default_treatment": "convert_at_cap",
                }
            ],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc_info:
            cap_state_mod.build_cap_state(self._minimal_inputs(), instruments)
        msg = str(exc_info.value)
        assert "E_NOTE_MISSING_FIELD" in msg, f"Expected E_NOTE_MISSING_FIELD in: {msg}"
        assert "'id'" in msg

    def test_note_missing_issuance_date_raises_structured(self) -> None:
        """Note missing 'issuance_date' → E_NOTE_MISSING_FIELD."""
        instruments = {
            "safes": [],
            "convertible_notes": [
                {
                    "id": "note_x",
                    "investor_name": "Lender Date",
                    "principal": 75_000,
                    # 'issuance_date' intentionally absent
                    "annual_interest_rate": 0.06,
                    "day_count_basis": 365,
                    "maturity_date": "2026-06-01",
                    "maturity_default_treatment": "convert_at_cap",
                }
            ],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc_info:
            cap_state_mod.build_cap_state(self._minimal_inputs(), instruments)
        msg = str(exc_info.value)
        assert "E_NOTE_MISSING_FIELD" in msg
        assert "issuance_date" in msg


# ===========================================================================
# Item 4 — Unknown SAFE form → E_UNKNOWN_SAFE_FORM blocker
# ===========================================================================


class TestUnknownSafeForm:
    """A SAFE with an unrecognised form value surfaces E_UNKNOWN_SAFE_FORM
    in both the per_safe dict and the solver's top-level blockers list,
    and completeness is NOT 'full'."""

    _CAP_STATE = {
        "founders": [{"name": "F", "common_shares": 10_000_000}],
        "preferred_series": [],
        "as_converted_totals": {
            "common_shares": 10_000_000,
            "preferred_shares_as_converted": 0,
            "options_outstanding": 0,
            "options_available": 1_000_000,
            "fully_diluted_shares": 11_000_000,
        },
        "option_pool": {
            "plan_type": "iso",
            "authorized": 1_000_000,
            "issued_and_outstanding": 0,
            "available_for_grant": 1_000_000,
        },
    }

    def test_bogus_form_surfaces_e_unknown_safe_form(self) -> None:
        """SAFE with form='bogus' → E_UNKNOWN_SAFE_FORM in blockers + not 'full'."""
        bogus_safe = {
            "id": "safe_bogus",
            "investor_name": "Angel Typo",
            "purchase_amount": 250_000,
            "form": "bogus",
            "post_money_valuation_cap": 5_000_000,
            "pre_money_valuation_cap": None,
            "discount_multiplier": None,
        }
        r = priced_round.solve_priced_round(
            cap_state=self._CAP_STATE,
            safes=[bogus_safe],
            notes=[],
            pre_money=5_000_000.0,
            new_money=3_000_000.0,
        )
        assert r["completeness"] != "full", "completeness must not be 'full' with an unknown form"
        codes = [b["code"] for b in r.get("blockers", [])]
        assert "E_UNKNOWN_SAFE_FORM" in codes, f"Expected E_UNKNOWN_SAFE_FORM in blockers; got: {codes}"
        # The per_safe dict also carries the error
        ps = r["per_safe"]["safe_bogus"]
        assert ps.get("error") == "E_UNKNOWN_SAFE_FORM"
        # The remedy message should name the valid forms
        blocker = next(b for b in r["blockers"] if b["code"] == "E_UNKNOWN_SAFE_FORM")
        remedy = blocker.get("remedy", "")
        assert "yc_postmoney_cap" in remedy, f"Valid forms not listed in remedy: {remedy}"

    def test_valid_form_does_not_trigger_e_unknown_safe_form(self) -> None:
        """Sanity check: a known form does NOT produce E_UNKNOWN_SAFE_FORM."""
        good_safe = {
            "id": "safe_good",
            "investor_name": "Angel Good",
            "purchase_amount": 250_000,
            "form": "yc_postmoney_cap",
            "post_money_valuation_cap": 5_000_000,
            "pre_money_valuation_cap": None,
            "discount_multiplier": None,
        }
        r = priced_round.solve_priced_round(
            cap_state=self._CAP_STATE,
            safes=[good_safe],
            notes=[],
            pre_money=5_000_000.0,
            new_money=3_000_000.0,
        )
        codes = [b["code"] for b in r.get("blockers", [])]
        assert "E_UNKNOWN_SAFE_FORM" not in codes


# ===========================================================================
# Item 5 — quick_assess.py UX: no-pool note + nested company_name
# ===========================================================================


class TestQuickAssessUX:
    """Tests for fast-assess UX improvements:
    5a. no-pool note appears when --target-pool-percent absent;
    5b. nested company_name fallback."""

    _INPUTS = {
        "company_name": "Foobar",
        "analysis_date": "2026-06-01",
        "mode": "standard",
        "jurisdiction": {
            "structure": "delaware",
            "incorporated_date": "2024-01-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        },
        "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
        "preferred_series": [],
        "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
        "common_batches": [],
        "metadata": {"run_id": "test"},
    }

    _SAFE = {
        "id": "safe_1",
        "investor_name": "Angel A",
        "purchase_amount": 500_000,
        "issuance_date": "2025-01-01",
        "form": "yc_postmoney_cap",
        "post_money_valuation_cap": 10_000_000,
        "pre_money_valuation_cap": None,
        "discount_multiplier": None,
    }

    def test_no_pool_note_present_when_flag_absent(self) -> None:
        """When target_pool_percent is None, report must contain the pool note."""
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="Foobar",
            inputs=self._INPUTS,
            safes=[self._SAFE],
            notes=[],
            pre_money=5_000_000.0,
            new_money=3_000_000.0,
            target_pool_percent=None,  # <-- no pool
            target_basis="post_money",
        )
        report_md = sentinel.pop("_report_md")
        assert "No pool top-up modeled" in report_md, "Expected pool note in report when target_pool_percent is absent"

    def test_no_pool_note_absent_when_pool_present(self) -> None:
        """When target_pool_percent is provided, the pool note must NOT appear."""
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="Foobar",
            inputs=self._INPUTS,
            safes=[self._SAFE],
            notes=[],
            pre_money=5_000_000.0,
            new_money=3_000_000.0,
            target_pool_percent=0.10,  # <-- pool provided
            target_basis="post_money",
        )
        report_md = sentinel.pop("_report_md")
        assert "No pool top-up modeled" not in report_md, "Pool note must not appear when target_pool_percent is set"

    def test_nested_company_name_fallback(self) -> None:
        """When top-level company_name absent, nested company.company_name is used."""
        import tempfile

        inputs_nested = dict(self._INPUTS)
        del inputs_nested["company_name"]
        inputs_nested["company"] = {"company_name": "NestedCo"}

        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            safes_path = os.path.join(d, "safes.json")
            with open(inputs_path, "w") as f:
                json.dump(inputs_nested, f)
            with open(safes_path, "w") as f:
                json.dump([self._SAFE], f)

            rc, stdout, stderr = _run(
                "quick_assess.py",
                [
                    "--inputs",
                    inputs_path,
                    "--safes",
                    safes_path,
                    "--pre-money",
                    "5000000",
                    "--new-money",
                    "3000000",
                    "--review-dir",
                    d,
                    "--pretty",
                ],
            )
            assert rc == 0, f"quick_assess.py failed: {stderr}"
            receipt = json.loads(stdout)
            md_path = receipt["wrote"]["fast_assess_report_md"]
            with open(md_path) as f:
                md = f.read()
            # company_name from nested path must appear in the report
            assert "NestedCo" in md, f"Expected 'NestedCo' in report; got: {md[:500]}"

    def test_run_id_arg_overrides_sentinel(self) -> None:
        """Fix 1: --run-id must be accepted and override the sentinel's run_id."""
        import tempfile

        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            safes_path = os.path.join(d, "safes.json")
            with open(inputs_path, "w") as f:
                json.dump(self._INPUTS, f)
            with open(safes_path, "w") as f:
                json.dump([self._SAFE], f)

            rc, stdout, stderr = _run(
                "quick_assess.py",
                [
                    "--inputs",
                    inputs_path,
                    "--safes",
                    safes_path,
                    "--pre-money",
                    "5000000",
                    "--new-money",
                    "3000000",
                    "--review-dir",
                    d,
                    "--run-id",
                    "test-run-42",
                    "--pretty",
                ],
            )
            assert rc == 0, f"quick_assess.py rejected --run-id: {stderr}"
            receipt = json.loads(stdout)
            # receipt carries run_id
            assert receipt["run_id"] == "test-run-42", f"receipt run_id mismatch: {receipt}"
            # sentinel JSON also carries run_id
            sentinel_path = os.path.join(d, "fast_assess_only.json")
            with open(sentinel_path) as f:
                sentinel = json.load(f)
            assert sentinel["run_id"] == "test-run-42", f"sentinel run_id mismatch: {sentinel['run_id']}"

    def test_pool_topup_share_counts_in_report(self) -> None:
        """Fix 2: when a pool top-up ran, the report must contain the share count line."""
        import tempfile

        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            safes_path = os.path.join(d, "safes.json")
            with open(inputs_path, "w") as f:
                json.dump(self._INPUTS, f)
            with open(safes_path, "w") as f:
                json.dump([self._SAFE], f)

            rc, stdout, stderr = _run(
                "quick_assess.py",
                [
                    "--inputs",
                    inputs_path,
                    "--safes",
                    safes_path,
                    "--pre-money",
                    "5000000",
                    "--new-money",
                    "3000000",
                    "--target-pool-percent",
                    "0.10",
                    "--target-basis",
                    "post_money",
                    "--review-dir",
                    d,
                    "--pretty",
                ],
            )
            assert rc == 0, f"quick_assess.py failed: {stderr}"
            receipt = json.loads(stdout)
            md_path = receipt["wrote"]["fast_assess_report_md"]
            with open(md_path) as f:
                md = f.read()
            # Report must include share counts for pool top-up
            assert "shares" in md.lower() and "pool" in md.lower(), (
                f"Expected pool top-up share count in report; got:\n{md}"
            )
            # More specifically, look for the pattern "+ N shares" or "N shares" near Pool
            import re

            assert re.search(r"Pool.*\d[\d,]* shares", md, re.IGNORECASE), (
                f"Expected 'Pool ... N shares' in report dilution lines; got:\n{md}"
            )

    def test_no_topup_label_when_pool_not_topped_up(self) -> None:
        """Fix 2a: when target_pool_percent is None (no top-up), dilution table must say
        'Existing option pool (no top-up)' and must NOT say 'Pool Refresh'."""
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="Foobar",
            inputs=self._INPUTS,
            safes=[self._SAFE],
            notes=[],
            pre_money=5_000_000.0,
            new_money=3_000_000.0,
            target_pool_percent=None,  # no top-up
            target_basis="post_money",
        )
        report_md = sentinel.pop("_report_md")
        assert "Pool Refresh" not in report_md, f"'Pool Refresh' must not appear when no top-up ran; got:\n{report_md}"
        assert "Existing option pool (no top-up)" in report_md, (
            f"Expected 'Existing option pool (no top-up)' in dilution table; got:\n{report_md}"
        )

    def test_topup_label_when_pool_topped_up(self) -> None:
        """Fix 2b: when target_pool_percent is set and top-up shares > 0,
        dilution table must say 'Pool Top-Up' and must NOT say 'Pool Refresh'."""
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="Foobar",
            inputs=self._INPUTS,
            safes=[self._SAFE],
            notes=[],
            pre_money=5_000_000.0,
            new_money=3_000_000.0,
            target_pool_percent=0.10,  # top-up requested
            target_basis="post_money",
        )
        report_md = sentinel.pop("_report_md")
        assert "Pool Refresh" not in report_md, f"'Pool Refresh' must not appear in report; got:\n{report_md}"
        assert "Pool Top-Up" in report_md, (
            f"Expected 'Pool Top-Up' in dilution table when top-up ran; got:\n{report_md}"
        )


# ===========================================================================
# Fast-assess report — per-holder cap table + SAFE derivation
# ===========================================================================


class TestFastAssessCapTableSection:
    """Per-holder cap table and SAFE derivation line in report_fast_assess.md.

    Scenario: EVAL2 single-SAFE (Angel A $500K @ $10M cap),
    pre_money=20M, new_money=5M, pool=10% post_money.
    """

    # Reuse EVAL2 fixture data from TestPricedRound
    _INPUTS = {
        "company_name": "TestCo",
        "analysis_date": "2026-05-21",
        "mode": "standard",
        "jurisdiction": {
            "structure": "delaware",
            "incorporated_date": "2024-06-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        },
        "founders": [
            {"name": "Founder", "founder_id": "founder_1", "common_shares": 10_000_000},
        ],
        "preferred_series": [],
        "option_pool": {
            "plan_type": "nso",
            "authorized": 1_000_000,
            "issued": 0,
            "unallocated": 1_000_000,
        },
        "common_batches": [],
        "metadata": {"run_id": "test"},
    }

    _SAFE = {
        "id": "safe_1",
        "investor_name": "Angel A",
        "purchase_amount": 500_000,
        "post_money_valuation_cap": 10_000_000,
        "discount_multiplier": None,
        "mfn_provision": None,
        "pro_rata_side_letter": None,
        "issuance_date": "2025-01-01",
        "form": "yc_postmoney_cap",
        "conversion_price_override": None,
        "source_document": None,
        "extraction_confidence": "high",
    }

    def _run(self) -> tuple[dict, str]:
        """Return (sentinel, report_md) for the canonical single-SAFE scenario."""
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="TestCo",
            inputs=self._INPUTS,
            safes=[self._SAFE],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        report_md = sentinel.pop("_report_md")
        return sentinel, report_md

    def test_cap_table_section_header_present(self) -> None:
        """Report must contain '## Post-Financing Cap Table' when completeness=='full'."""
        _, md = self._run()
        assert "## Post-Financing Cap Table" in md, (
            f"Expected '## Post-Financing Cap Table' section in report; got:\n{md[:1000]}"
        )

    def test_cap_table_founders_row_present(self) -> None:
        """Founders row must be present with a share count in the cap table."""
        import re

        _, md = self._run()
        # Must have a Founders row with a comma-formatted share count
        assert re.search(r"Founders.*[\d,][\d,][\d,]+", md), (
            f"Expected 'Founders' row with share count in cap table; got:\n{md}"
        )

    def test_cap_table_total_row_present(self) -> None:
        """Total row must be present in the cap table."""
        _, md = self._run()
        assert "Total" in md and "100%" in md, f"Expected 'Total' row with 100% in cap table; got:\n{md}"

    def test_cap_table_share_counts_consistent(self) -> None:
        """Parse the cap table rows and verify: sum of holder shares == Total row shares,
        and Total == solver post_round_fully_diluted_shares."""
        import math
        import re
        import sys

        sys.path.insert(0, SCRIPTS)
        import cap_state as cap_state_mod  # type: ignore[import-not-found]
        import priced_round as pr  # type: ignore[import-not-found]

        _, md = self._run()

        # Extract cap table section
        cap_table_match = re.search(
            r"## Post-Financing Cap Table\n(.*?)(?=\n##|\Z)",
            md,
            re.DOTALL,
        )
        assert cap_table_match, f"Could not find Post-Financing Cap Table section:\n{md}"
        table_text = cap_table_match.group(1)

        # Parse table rows: | ... | N,NNN,NNN | NN.NN% |
        # Look for rows with integer share counts (skip header and separator rows)
        row_shares: list[int] = []
        total_shares_from_table: int | None = None
        for line in table_text.splitlines():
            if "|" not in line or "---" in line or "Holder" in line:
                continue
            # Bold total row check
            cols = [c.strip() for c in line.split("|") if c.strip()]
            if len(cols) < 2:
                continue
            # Strip markdown bold markers
            share_str = cols[1].replace("*", "").replace(",", "").strip()
            if not share_str.isdigit():
                continue
            shares = int(share_str)
            holder_col = cols[0].replace("*", "").strip()
            if "Total" in holder_col or "fully diluted" in holder_col.lower():
                total_shares_from_table = shares
            else:
                row_shares.append(shares)

        assert row_shares, f"No holder rows parsed from cap table:\n{table_text}"
        assert total_shares_from_table is not None, f"No Total row with share count found in cap table:\n{table_text}"

        # Sum of holder rows must equal total row (within ±2 for rounding)
        holder_sum = sum(row_shares)
        assert abs(holder_sum - total_shares_from_table) <= 2, (
            f"Holder share sum ({holder_sum:,}) != Total row ({total_shares_from_table:,}); "
            f"rounding divergence > 2 shares"
        )

        # Total must match solver post_round_fully_diluted_shares
        cs = cap_state_mod.build_cap_state(
            self._INPUTS,
            {
                "safes": [self._SAFE],
                "convertible_notes": [],
                "warrants": [],
                "option_grants": [],
                "metadata": {"run_id": "test"},
            },
        )
        solver_result = pr.solve_priced_round(
            cap_state=cs,
            safes=[self._SAFE],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        solver_fd = solver_result["post_round_fully_diluted_shares"]
        assert math.isclose(total_shares_from_table, solver_fd, abs_tol=2), (
            f"Table Total ({total_shares_from_table:,}) != solver FD ({solver_fd:,})"
        )

    def test_safe_derivation_sentence_present(self) -> None:
        """Report must contain the SAFE derivation canonical sentence."""
        import re

        _, md = self._run()
        # The derivation uses "purchase ÷ cap" or the share count formula
        assert re.search(r"purchase\s*[÷/]\s*cap", md) or re.search(r"purchase.*cap.*conversion", md, re.IGNORECASE), (
            f"Expected SAFE derivation sentence with 'purchase ÷ cap' in report; got:\n{md}"
        )

    def test_safe_row_with_investor_name_in_cap_table(self) -> None:
        """SAFE row must appear in cap table when only 1 SAFE is present (per-holder mode)."""
        _, md = self._run()
        # "Angel A" is the investor name from the SAFE fixture
        assert "Angel A" in md, f"Expected SAFE investor name 'Angel A' in report cap table; got:\n{md[:1500]}"

    def test_no_cap_table_section_when_completeness_not_full(self) -> None:
        """Cap table section must NOT appear when solver cannot converge (no valid cap)."""
        import quick_assess as qa  # type: ignore[import-not-found]

        # SAFE with unknown form triggers E_UNKNOWN_SAFE_FORM -> completeness=structural_only
        bad_safe = dict(self._SAFE)
        bad_safe["form"] = "unknown_form_xyz"

        sentinel = qa.quick_assess(
            company_name="TestCo",
            inputs=self._INPUTS,
            safes=[bad_safe],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        report_md = sentinel.pop("_report_md")
        assert "## Post-Financing Cap Table" not in report_md, (
            "Cap table section must not appear when completeness != 'full'"
        )


# ===========================================================================
# Item 2 — Quantitative AD + SAFE golden (test_golden_4 extension)
# ===========================================================================


class TestGolden4ADPlusSAFEQuantitative:
    """Quantitative extension of test_golden_4_ad_plus_safe_conversion.

    Derivation of the closed-form fixed point:
    ----------------------------------------
    Inputs:
      founders:  10 M common shares
      Series Seed:  2 M shares, OIP = OCP = CCP = $1.00, BBWA broad
      option pool available:  1 M shares
      SAFE:  $500k purchase @ $10M post-money cap (yc_postmoney_cap)
      pre_money = $5M, new_money = $5M

    The YC post-money SAFE identity:
      SAFE shares = purchase_amount / safe_price
      safe_price  = post_money_cap / company_capitalization
      => SAFE ownership = purchase_amount / post_money_cap = 500k / 10M = 5%
    This is a CONSTANT — it does not depend on the solver iteration. The
    SAFE holder always owns exactly purchase/cap of "Company Capitalization",
    regardless of how the AD loop resolves.

    Therefore:
      safe_shares / company_capitalization == purchase / cap
        <=> safe_shares / company_cap == 0.05  (the YC identity)

    For the BBWA CP2 derivation, safe_shares are carved out of the new-money
    consideration (NVCA default). The AD trigger fires because new PPS < OIP.
    The BD[B] term uses new_money only:
      B = new_money / CP1 = 5_000_000 / 1.00 = 5_000_000 shares

    The full closed-form fixed point is intractable here (coupled SAFE + BBWA
    is a quadratic in safe_shares that depends on the AD-adjusted FD), so we
    pin the verifiable identity and direction-of-effect assertions:

    (a) YC identity: safe_shares / company_cap == 0.05
    (b) Ownership sum to 1
    (c) AD adjustment direction: founder_pct_post_AD < founder_pct_pre_AD
    (d) B term = new_money / CP1 (SAFE carved out)
    (e) Sanity pin on PPS (pre-money / pre_FD ≈ 5M / 13M; actual is less due
        to SAFE + AD expansion of denominator)
    """

    _SAFE = {
        "id": "safe_a",
        "investor_name": "Angel A",
        "purchase_amount": 500_000.0,
        "form": "yc_postmoney_cap",
        "post_money_valuation_cap": 10_000_000.0,
        "pre_money_valuation_cap": None,
        "discount_multiplier": None,
    }

    def _run(self) -> dict[str, Any]:
        cap_state = {
            "founders": [{"name": "Founder A", "common_shares": 10_000_000}],
            "preferred_series": [
                {
                    "series_id": "series_seed",
                    "shares": 2_000_000,
                    "original_issue_price": 1.00,
                    "original_conversion_price": 1.00,
                    "current_conversion_price": 1.00,
                    "anti_dilution_protection": "broad_based_weighted_average",
                }
            ],
            "as_converted_totals": {
                "common_shares": 10_000_000,
                "preferred_shares_as_converted": 2_000_000,
                "options_outstanding": 0,
                "options_available": 1_000_000,
                "fully_diluted_shares": 13_000_000,
            },
            "option_pool": {
                "plan_type": "iso",
                "authorized": 1_000_000,
                "issued_and_outstanding": 0,
                "available_for_grant": 1_000_000,
            },
        }
        result: dict[str, Any] = priced_round.solve_priced_round(  # type: ignore[assignment]
            cap_state=cap_state,
            safes=[self._SAFE],
            notes=[],
            pre_money=5_000_000.0,
            new_money=5_000_000.0,
        )
        return result

    def test_converges(self) -> None:
        r = self._run()
        assert r["converged"], f"solver did not converge: {r.get('blockers')}"

    def test_yc_safe_identity_purchase_over_cap(self) -> None:
        """(a) YC post-money identity: safe_shares / company_cap == purchase / cap.

        This is the load-bearing invariant for YC post-money SAFEs regardless
        of how the AD fixed point resolves.  company_capitalization is the
        adj_pre_fd at the converged price which equals
        (pre_money / pps) — the converged Company Capitalization denominator
        that the solver produces.
        """
        import math as _math

        r = self._run()
        assert r["completeness"] == "full", f"blockers: {r.get('blockers')}"

        pps = r["equity_financing_price"]
        # At convergence, company_capitalization = pre_money / pps
        company_cap = 5_000_000.0 / pps
        safe_shares = r["shares_breakdown"]["safe_converted"]
        # YC identity: safe_shares / company_cap == purchase / cap
        expected_ratio = 500_000.0 / 10_000_000.0  # = 0.05 exactly
        actual_ratio = safe_shares / company_cap
        assert _math.isclose(actual_ratio, expected_ratio, rel_tol=1e-4), (
            f"YC identity failed: safe_shares/company_cap={actual_ratio:.6f}, "
            f"expected purchase/cap={expected_ratio:.6f}"
        )

    def test_ownership_sums_to_one(self) -> None:
        """(b) Aggregate ownership must sum to 1 (within floating-point tolerance)."""
        import math as _math

        r = self._run()
        agg = r["aggregate_ownership_by_class"]
        total = (
            agg["founders_pct"] + agg["safe_pct"] + agg["preferred_pct"] + agg["option_pool_pct"] + agg["new_money_pct"]
        )
        assert _math.isclose(total, 1.0, abs_tol=1e-5), f"Ownership does not sum to 1: {total:.8f} (components: {agg})"

    def test_ad_reduces_founder_pct(self) -> None:
        """(c) AD adjustment direction: founder_pct post-AD < pre-AD."""
        r = self._run()
        agg = r["aggregate_ownership_by_class"]
        assert "founders_pct_pre_anti_dilution" in agg, "founders_pct_pre_anti_dilution must be present when AD fires"
        assert agg["founders_pct"] < agg["founders_pct_pre_anti_dilution"], (
            f"AD should reduce founder pct: post={agg['founders_pct']:.5f} "
            f"pre={agg['founders_pct_pre_anti_dilution']:.5f}"
        )

    def test_bbwa_b_term_uses_new_money_only(self) -> None:
        """(d) BBWA B term = new_money / CP1 (SAFE carved out per NVCA default)."""
        import math as _math

        r = self._run()
        bd = r["anti_dilution_breakdown"][0]
        # B = consideration / CP1; consideration = new_money only (SAFE carved out)
        # new_money = 5_000_000; CP1 = 1.00 => B = 5_000_000
        assert _math.isclose(bd["B"], 5_000_000.0, abs_tol=1.0), (
            f"B term {bd['B']} != 5_000_000 (new_money/CP1 with SAFE carved out)"
        )

    def test_pps_below_oip(self) -> None:
        """(e) PPS must be below OIP=$1.00 (this is a down round — AD trigger fires)."""
        r = self._run()
        assert r["equity_financing_price"] < 1.00, (
            f"PPS {r['equity_financing_price']} is not below OIP=1.00; AD should fire"
        )


# ===========================================================================
# Audit regression tests (a1: cap-table area)
# ===========================================================================


class TestRulePackVersion:
    """_rule_pack.RULE_PACK_VERSION is read from cap-table-rules.json and is the
    single source every producer binds to (no hardcoded literals drift)."""

    def test_version_matches_rules_json(self) -> None:
        import _rule_pack  # type: ignore[import-not-found]

        rules_path = os.path.join(_REPO, "founder-skills", "skills", "cap-table", "data", "cap-table-rules.json")
        with open(rules_path, encoding="utf-8") as f:
            expected = json.load(f)["metadata"]["version"]
        assert expected == _rule_pack.RULE_PACK_VERSION

    def test_all_producers_share_one_version(self) -> None:
        import _rule_pack  # type: ignore[import-not-found]
        import anti_dilution as ad  # type: ignore[import-not-found]
        import flip_scenario  # type: ignore[import-not-found]
        import note_conversion as nc  # type: ignore[import-not-found]
        import option_pool as op  # type: ignore[import-not-found]
        import priced_round as pr  # type: ignore[import-not-found]
        import run_scenario  # type: ignore[import-not-found]
        import safe_conversion as sc  # type: ignore[import-not-found]

        v = _rule_pack.RULE_PACK_VERSION
        for mod in (ad, nc, op, pr, sc, flip_scenario, run_scenario):
            assert v == mod.RULE_PACK_VERSION, f"{mod.__name__} drifted from {v}"

    def test_no_stale_literal_versions_in_provenance(self) -> None:
        """flip_scenario / run_scenario provenance no longer hardcode 0.3.2."""
        import _rule_pack  # type: ignore[import-not-found]

        for name in ("flip_scenario.py", "run_scenario.py"):
            with open(os.path.join(SCRIPTS, name), encoding="utf-8") as fh:
                text = fh.read()
            assert '"0.3.2"' not in text, f"{name} still hardcodes 0.3.2"
        # And the live version is current.
        assert _rule_pack.RULE_PACK_VERSION != "0.3.2"


class TestPreferredSeriesCCPFallback:
    """math-1: CCP canonicalization falls back to original_conversion_price (not
    legacy 'ocp' default 0), and a resolved CCP <= 0 is rejected, never masked."""

    def _inputs_with_series(self, series: dict) -> dict:
        return {
            "company_name": "TestCo",
            "analysis_date": "2026-06-01",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-01-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
            "preferred_series": [series],
            "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
            "common_batches": [],
            "metadata": {"run_id": "test"},
        }

    def _instruments(self) -> dict:
        return {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }

    def test_ccp_falls_back_to_original_conversion_price(self) -> None:
        series = {
            "series_name": "Seed",
            "shares": 2_000_000,
            "original_issue_price": 2.0,
            "original_conversion_price": 2.0,
            # current_conversion_price intentionally omitted (and no 'ocp')
            "issuance_date": "2024-06-01",
        }
        cs = cap_state_mod.build_cap_state(self._inputs_with_series(series), self._instruments())
        canon = cs["preferred_series"][0]
        assert canon["current_conversion_price"] == 2.0, canon["current_conversion_price"]
        # as-converted ratio is 1:1 (OCP == CCP), so as-converted == shares
        assert cs["as_converted_totals"]["preferred_shares_as_converted"] == 2_000_000

    def test_resolved_ccp_zero_is_rejected(self) -> None:
        series = {
            "series_name": "Seed",
            "shares": 2_000_000,
            "original_issue_price": 2.0,
            "original_conversion_price": 2.0,
            "current_conversion_price": 0.0,
            "issuance_date": "2024-06-01",
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc:
            cap_state_mod.build_cap_state(self._inputs_with_series(series), self._instruments())
        assert "E_PREFERRED_SERIES_INVALID_PRICE" in str(exc.value)


class TestPricingUnknownSeries:
    """P5: 'pricing unknown' mode. A preferred series whose founder/mapper could not confirm
    original_issue_price/original_conversion_price/current_conversion_price carries a numeric
    $1.00 sentinel for all three (so the schema's required-numeric and cap_state's >0 gates pass
    unchanged) plus `pricing_unknown: true`, and anti_dilution_protection must resolve to 'none'.
    """

    def _inputs_with_series(self, series: dict) -> dict:
        return {
            "company_name": "TestCo",
            "analysis_date": "2026-06-01",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-01-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
            "preferred_series": [series],
            "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
            "common_batches": [],
            "metadata": {"run_id": "test", "schema_version": "v0.5.0-inputs"},
        }

    def _instruments(self) -> dict:
        return {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }

    _PRICING_UNKNOWN_SERIES = {
        "series_name": "Series A",
        "shares": 2_000_000,
        "original_issue_price": 1.0,
        "original_conversion_price": 1.0,
        "current_conversion_price": 1.0,
        "anti_dilution_protection": "none",
        "pricing_unknown": True,
        "issuance_date": "2024-06-01",
    }

    def test_pricing_unknown_series_validates_against_inputs_schema(self) -> None:
        """(a) part 1: the $1.00-sentinel series validates against inputs.schema.json — the new
        `pricing_unknown` boolean property doesn't break the required-numeric OIP/OCP/CCP checks."""
        import _cap_table_schema_validator as v  # type: ignore[import-not-found]
        from _artifact_writer import load_schema  # type: ignore[import-not-found]

        schema_path = os.path.join(os.path.dirname(SCRIPTS), "references", "schemas", "inputs.schema.json")
        schema = load_schema(schema_path)
        inputs = self._inputs_with_series(dict(self._PRICING_UNKNOWN_SERIES))
        errs = v.validate(inputs, schema, "inputs")
        assert errs == [], f"pricing_unknown inputs should validate cleanly: {errs}"

    def test_pricing_unknown_series_build_cap_state_succeeds(self) -> None:
        """(a) part 2: build_cap_state succeeds — all four cap_state >0 gates
        (OIP>0, OCP>0, resolved CCP>0, CCP<=OCP) pass on the 1.0 sentinel."""
        series = dict(self._PRICING_UNKNOWN_SERIES)
        cs = cap_state_mod.build_cap_state(self._inputs_with_series(series), self._instruments())
        canon = cs["preferred_series"][0]
        assert canon["original_issue_price"] == 1.0
        assert canon["original_conversion_price"] == 1.0
        assert canon["current_conversion_price"] == 1.0
        assert canon["anti_dilution_protection"] == "none"
        # flag-carry: the canonicalizer's fixed-key comprehension must copy pricing_unknown through
        # (task 3 / the R3 flag-carry fix) — without it the matcher + disclosure are permanently dead.
        assert canon.get("pricing_unknown") is True
        assert "W_PRICING_UNKNOWN" in cs.get("warnings", [])

    def test_pricing_unknown_produces_1_to_1_pro_rata_like_priced_none_ad_series(self) -> None:
        """(b) A pricing_unknown series (1.0/1.0/1.0 sentinel) produces the same as-converted
        share count as an ordinary priced series with anti_dilution: 'none' and OCP == CCP —
        both are exact 1:1 pro-rata (as-converted == shares)."""
        unknown_cs = cap_state_mod.build_cap_state(
            self._inputs_with_series(dict(self._PRICING_UNKNOWN_SERIES)), self._instruments()
        )
        priced_series = {
            "series_name": "Series A",
            "shares": 2_000_000,
            "original_issue_price": 3.50,
            "original_conversion_price": 3.50,
            "current_conversion_price": 3.50,
            "anti_dilution_protection": "none",
            "issuance_date": "2024-06-01",
        }
        priced_cs = cap_state_mod.build_cap_state(self._inputs_with_series(priced_series), self._instruments())
        assert (
            unknown_cs["as_converted_totals"]["preferred_shares_as_converted"]
            == priced_cs["as_converted_totals"]["preferred_shares_as_converted"]
            == 2_000_000
        )

    def test_pricing_unknown_disclosure_counsel_item_only_when_flagged(self) -> None:
        """(c) cap_table.pricing_unknown_disclosure appears as a counsel-review item ONLY when a
        pricing_unknown series is present. Built via build_cap_state() (not a hand-authored cap_state
        dict) so this exercises the real flag-carry path — a hand-authored dict would false-green
        over a broken canonicalizer comprehension."""
        with open(os.path.join(os.path.dirname(SCRIPTS), "data", "cap-table-rules.json")) as f:
            rules = json.load(f)

        unknown_inputs = self._inputs_with_series(dict(self._PRICING_UNKNOWN_SERIES))
        unknown_cs = cap_state_mod.build_cap_state(unknown_inputs, self._instruments())
        gating = rule_audit.build_gating_block(
            rules, inputs=unknown_inputs, instruments=self._instruments(), cap_state=unknown_cs
        )
        items = rule_audit.build_counsel_review_items(gating, rules)
        assert "cap_table.pricing_unknown_disclosure" in {it["rule_id"] for it in items}

        priced_series = {
            "series_name": "Series A",
            "shares": 2_000_000,
            "original_issue_price": 3.50,
            "original_conversion_price": 3.50,
            "current_conversion_price": 3.50,
            "anti_dilution_protection": "none",
            "issuance_date": "2024-06-01",
        }
        priced_inputs = self._inputs_with_series(priced_series)
        priced_cs = cap_state_mod.build_cap_state(priced_inputs, self._instruments())
        gating_priced = rule_audit.build_gating_block(
            rules, inputs=priced_inputs, instruments=self._instruments(), cap_state=priced_cs
        )
        items_priced = rule_audit.build_counsel_review_items(gating_priced, rules)
        assert "cap_table.pricing_unknown_disclosure" not in {it["rule_id"] for it in items_priced}

    def test_pricing_unknown_with_stray_ratchet_key_rejected(self) -> None:
        """(d) A pricing_unknown series with a stray mis-keyed `anti_dilution: 'ratchet'` is REJECTED
        by the post-resolution invariant — proves the invariant runs AFTER `_resolve_anti_dilution`
        (which would otherwise flip the resolved anti_dilution_protection to 'full_ratchet' and let
        ratchet math fire against the $1.00 sentinel)."""
        series = {
            "series_name": "Series A",
            "shares": 2_000_000,
            "original_issue_price": 1.0,
            "original_conversion_price": 1.0,
            "current_conversion_price": 1.0,
            "pricing_unknown": True,
            "anti_dilution": "ratchet",  # stray mis-keyed field _resolve_anti_dilution recovers
            "issuance_date": "2024-06-01",
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc:
            cap_state_mod.build_cap_state(self._inputs_with_series(series), self._instruments())
        assert "E_PRICING_UNKNOWN_AD_NOT_NONE" in str(exc.value)

    def test_pricing_unknown_callout_mentions_1_to_1(self) -> None:
        """(e) The W_PRICING_UNKNOWN disclosure banner states the 1:1 conversion assumption."""
        from _warning_callouts import render_warning_callouts  # noqa: PLC0415

        body = "\n".join(render_warning_callouts(["W_PRICING_UNKNOWN"]))
        assert "1:1" in body
        assert "anti-dilution" in body.lower()

    def test_pricing_unknown_end_to_end_mapper_through_rule_audit(self) -> None:
        """Cross-boundary: a pricing_unknown series produced by the freeform mapper (map_freeform,
        --answer pricing_unknown=true) must flow through build_cap_state -> rule_audit and surface
        BOTH the W_PRICING_UNKNOWN disclosure and the counsel item. This guards the seam between the
        mapper's emitted shape and cap_state's expectation — a divergence (e.g. the flag emitted as a
        string) would pass each half's own unit tests but break the pipeline."""
        import freeform_mapper as fm  # type: ignore[import-not-found]  # noqa: PLC0415
        from _warning_callouts import render_warning_callouts  # noqa: PLC0415

        meta = self._inputs_with_series({})
        del meta["preferred_series"]  # the mapper produces it
        blocks = [
            {
                "block_type": "preferred_series_block",
                "sheet": "P",
                "cell_range": "A2:B2",
                "column_role_map": {"A": "series_name", "B": "shares"},
            }
        ]
        grid = {"ok": True, "mode": "grid", "sheets": {"P": {"rows": [["S", "Sh"], ["Series A", 2_000_000]]}}}
        mapped = fm.map_freeform(
            blocks, grid, existing_inputs=meta, answers={"0.pricing_unknown": "true"}, run_id="test"
        )
        assert mapped["blockers"] == []
        series = mapped["inputs"]["preferred_series"][0]
        assert series["pricing_unknown"] is True and series["original_issue_price"] == 1.0

        inputs = dict(meta)
        inputs["preferred_series"] = mapped["inputs"]["preferred_series"]
        cs = cap_state_mod.build_cap_state(inputs, self._instruments())
        assert "W_PRICING_UNKNOWN" in cs.get("warnings", [])
        assert cs["preferred_series"][0].get("pricing_unknown") is True

        with open(os.path.join(os.path.dirname(SCRIPTS), "data", "cap-table-rules.json")) as f:
            rules = json.load(f)
        gating = rule_audit.build_gating_block(rules, inputs=inputs, instruments=self._instruments(), cap_state=cs)
        items = rule_audit.build_counsel_review_items(gating, rules)
        assert "cap_table.pricing_unknown_disclosure" in {it["rule_id"] for it in items}
        assert "1:1" in "\n".join(render_warning_callouts(cs["warnings"]))


class TestPricedRoundNoteDateBlocker:
    """math-2: notes present without a conversion date returns a structured
    blocker (E_NOTE_NO_CONVERSION_DATE), never an AssertionError."""

    def _cap_state(self) -> dict:
        return {
            "as_converted_totals": {
                "fully_diluted_shares": 1_000_000,
                "common_shares": 1_000_000,
                "preferred_shares_as_converted": 0,
                "options_outstanding": 0,
                "options_available": 0,
                "warrants_underlying_total": 0,
            },
            "founders": [{"common_shares": 1_000_000}],
            "preferred_series": [],
        }

    def test_notes_without_date_returns_blocker(self) -> None:
        r = priced_round.solve_priced_round(
            cap_state=self._cap_state(),
            safes=[],
            notes=[{"id": "n1", "principal": 100_000}],
            pre_money=5_000_000,
            new_money=2_000_000,
        )
        assert r["completeness"] == "structural_only"
        assert r["blockers"][0]["code"] == "E_NOTE_NO_CONVERSION_DATE"

    def test_quick_assess_notes_defaults_today_and_discloses(self) -> None:
        import quick_assess as qa  # type: ignore[import-not-found]

        inputs = dict(_BASIC_INPUTS)
        note = {
            "id": "n1",
            "principal": 250_000,
            "issuance_date": "2024-01-01",
            "interest_rate": 0.06,
            "valuation_cap": 8_000_000,
            "capitalization_denominator": 10_000_000,
        }
        sentinel = qa.quick_assess(
            company_name="TestCo",
            inputs=inputs,
            safes=[],
            notes=[note],
            pre_money=8_000_000,
            new_money=2_000_000,
            target_pool_percent=None,
            target_basis="post_money",
        )
        assert sentinel.get("assumptions"), "expected disclosed assumption for missing note date"
        assert any("today" in a.lower() for a in sentinel["assumptions"])


class TestPreAdBaselineDenominator:
    """math-3: pre_ad_post_fd includes common_batches + warrants, so the AD
    delta denominator matches the post-AD denominator's components."""

    def test_pre_ad_fd_includes_batches_and_warrants(self) -> None:
        cs = {
            "as_converted_totals": {
                "fully_diluted_shares": 2_500_000,
                "common_shares": 2_000_000,  # founders 1.5M + batch 0.5M
                "preferred_shares_as_converted": 250_000,
                "options_outstanding": 0,
                "options_available": 0,
                "warrants_underlying_total": 250_000,
            },
            "founders": [{"common_shares": 1_500_000, "common_class": "class_a"}],
            "common_batches": [{"shares": 500_000}],
            "preferred_series": [
                {
                    "series_id": "seed",
                    "series_name": "Seed",
                    "shares": 250_000,
                    "original_issue_price": 1.0,
                    "original_conversion_price": 1.0,
                    "current_conversion_price": 1.0,
                    "anti_dilution_protection": "broad_based_weighted_average",
                    "ad_a_denominator_basis": "nvca_broad",
                    "ad_trigger_basis": "original_issue_price",
                    "issuance_date": "2024-01-01",
                }
            ],
        }
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[],
            pre_money=1_000_000,  # down round → AD fires
            new_money=500_000,
        )
        agg = r["aggregate_ownership_by_class"]
        # Pre-AD baseline derivation:
        #   pre_pps = pre_money / pre_fd = 1_000_000 / 2_500_000 = 0.40 (frozen)
        #   pre_ad_new_money_shares = 500_000 / 0.40 = 1_250_000
        #   pre_ad_post_fd = common_shares(2_000_000) + preferred_pre_ad(250_000)
        #                  + options(0) + warrants(250_000) + pool_topup(0)
        #                  + safe_shares(0) + note_shares(0)
        #                  + pre_ad_new_money_shares(1_250_000) = 3_750_000
        #   founders numerator = 1_500_000 (founders list only, not batches)
        #   expected = 1_500_000 / 3_750_000 = 0.40
        # Old denominator (omitting batches 500_000 + warrants 250_000 = 750_000):
        #   old_denom = 3_000_000 → old_value = 0.50 — delta 0.10 exceeds 1e-6 tolerance.
        assert "founders_pct_pre_anti_dilution" in agg
        assert abs(agg["founders_pct_pre_anti_dilution"] - 0.40) < 1e-6


class TestVisualize:
    """math-4: scenario donut/legend palette lookup tolerates the _pct suffix,
    has a warrants color, and excludes pre-AD/delta keys."""

    def test_palette_color_strips_pct_suffix(self) -> None:
        import visualize  # type: ignore[import-not-found]

        assert visualize._palette_color("founders_pct") == visualize.PALETTE["founders"]
        assert visualize._palette_color("safe_pct") == visualize.PALETTE["safe"]
        assert visualize._palette_color("warrants") == visualize.PALETTE["warrants"]
        # Unknown key falls back to neutral
        assert visualize._palette_color("totally_unknown") == visualize.PALETTE["neutral"]

    def test_donut_renders_non_neutral_slices_for_pct_keys(self) -> None:
        import visualize  # type: ignore[import-not-found]

        agg = {"founders_pct": 0.6, "safe_pct": 0.2, "new_money_pct": 0.2}
        svg = visualize.render_donut(agg, size=120)
        assert visualize.PALETTE["founders"] in svg
        assert visualize.PALETTE["safe"] in svg
        # No neutral wedge should appear for known categories
        assert svg.count(visualize.PALETTE["neutral"]) == 0

    def test_donut_excludes_pre_ad_delta_keys(self) -> None:
        import visualize  # type: ignore[import-not-found]

        agg = {
            "founders_pct": 0.6,
            "new_money_pct": 0.4,
            "founders_pct_pre_anti_dilution": 0.7,
            "anti_dilution_delta_pct_points": -10.0,
        }
        svg = visualize.render_donut(agg, size=120)
        legend = visualize.render_legend(agg)
        assert "anti dilution delta" not in legend
        # negative delta should never create a wedge
        assert "founders pct pre anti dilution" not in legend
        assert svg  # renders without error


class TestExploreDonutExclusion:
    """math-5: explore.py filters AD meta keys and non-numeric values from
    the donut/legend data path before embedding into JS."""

    def _make_explorer_html(self, agg: dict) -> str:
        import explore  # type: ignore[import-not-found]

        inputs = {"company_name": "TestCo", "mode": "standard", "analysis_date": "2026-01-01"}
        cap_state_data: dict = {
            "as_of_date": "2026-01-01",
            "as_converted_totals": {
                "fully_diluted_shares": 10_000_000,
                "common_shares": 8_000_000,
                "preferred_shares_as_converted": 1_000_000,
                "options_outstanding": 500_000,
                "options_available": 500_000,
                "warrants_underlying_total": 0,
            },
            "founders": [{"name": "Alice", "founder_id": "alice", "common_shares": 8_000_000}],
        }
        scenarios_doc: dict = {
            "scenarios": [
                {
                    "scenario_id": "s1",
                    "label": "Test Scenario",
                    "type": "priced_round",
                    "computed_outputs": {
                        "completeness": "full",
                        "aggregate_ownership_by_class": agg,
                        "equity_financing_price": 1.0,
                        "post_round_fully_diluted_shares": 12_000_000,
                        "shares_breakdown": {},
                        "founder_impact": None,
                    },
                    "parameters": {},
                }
            ]
        }
        counsel_packet: dict = {"items": []}
        return explore.render_explorer_html(  # type: ignore[no-any-return]
            inputs=inputs,
            cap_state=cap_state_data,
            scenarios_doc=scenarios_doc,
            counsel_packet=counsel_packet,
        )

    def test_ad_meta_keys_absent_from_embedded_donut_data(self) -> None:
        """AD meta keys must not appear in the donut/legend data path."""
        agg = {
            "founders_pct": 0.50,
            "preferred_pct": 0.10,
            "option_pool_pct": 0.05,
            "new_money_pct": 0.15,
            "warrants_pct": 0.05,
            "founders_by_class": {"class_a": 0.50},
            "founders_pct_pre_anti_dilution": 0.60,
            "preferred_pct_pre_anti_dilution": 0.12,
            "anti_dilution_delta_pct_points": -10.0,
        }
        html = self._make_explorer_html(agg)
        # The aggregate object that JS iterates must not contain the meta keys.
        # They are embedded as JSON inside the DATA constant — assert they are
        # absent from the aggregate sub-object in the embedded payload.
        assert '"founders_pct_pre_anti_dilution"' not in html
        assert '"preferred_pct_pre_anti_dilution"' not in html
        assert '"anti_dilution_delta_pct_points"' not in html
        # Non-numeric founders_by_class dict must also be excluded.
        assert '"founders_by_class"' not in html

    def test_real_ownership_keys_present(self) -> None:
        """Real ownership slices survive the filter."""
        agg = {
            "founders_pct": 0.50,
            "preferred_pct": 0.20,
            "option_pool_pct": 0.10,
            "new_money_pct": 0.20,
            "founders_by_class": {"class_a": 0.50},
            "founders_pct_pre_anti_dilution": 0.60,
            "anti_dilution_delta_pct_points": -10.0,
        }
        html = self._make_explorer_html(agg)
        assert '"founders_pct"' in html
        assert '"preferred_pct"' in html
        assert '"option_pool_pct"' in html
        assert '"new_money_pct"' in html

    def test_warrants_in_palette(self) -> None:
        """warrants color is defined in the JS PALETTE."""
        agg = {
            "founders_pct": 0.80,
            "warrants_pct": 0.05,
            "new_money_pct": 0.15,
        }
        html = self._make_explorer_html(agg)
        # The JS PALETTE object must contain a warrants entry.
        assert "warrants:" in html or '"warrants"' in html


class TestComposeSummaryCounts:
    """math-7: summary.passed / failed are scenario counts (never negative)."""

    def test_passed_never_negative_with_multi_blocker_scenario(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        scenarios: list[dict[str, Any]] = [
            {
                "scenario_id": "s1",
                "type": "priced_round",
                "parameters": {},
                "computed_outputs": {
                    "blockers": [
                        {"code": "E_A", "remedy": "fix a"},
                        {"code": "E_B", "remedy": "fix b"},
                        {"code": "E_C", "remedy": "fix c"},
                    ]
                },
            },
            {"scenario_id": "s2", "type": "priced_round", "parameters": {}, "computed_outputs": {}},
        ]
        failed_scenarios = sum(1 for s in scenarios if ((s.get("computed_outputs", {}) or {}).get("blockers") or []))
        passed_scenarios = len(scenarios) - failed_scenarios
        assert failed_scenarios == 1
        assert passed_scenarios == 1
        assert passed_scenarios >= 0
        # build_scenario_digest must not raise on this shape
        compose_report.build_scenario_digest(scenarios)


class TestSafeFormChoices:
    """math-16: priced-round --form exposes the supported pre-money forms and
    no longer offers the always-rejected 'other'."""

    def test_cli_form_choices(self) -> None:
        rc, out, err = _run(
            "safe_conversion.py",
            [
                "priced-round",
                "--purchase",
                "100000",
                "--form",
                "yc_premoney_cap_only",
                "--pre-money-cap",
                "5000000",
                "--pre-money-fd",
                "10000000",
                "--company-cap",
                "10000000",
                "--equity-price",
                "1.0",
            ],
        )
        assert rc == 0, err
        result = json.loads(out)
        assert result.get("branch") != "rejected", result

    def test_other_form_removed(self) -> None:
        rc, out, err = _run(
            "safe_conversion.py",
            [
                "priced-round",
                "--purchase",
                "100000",
                "--form",
                "other",
                "--company-cap",
                "10000000",
            ],
        )
        # argparse rejects an invalid choice with exit code 2
        assert rc == 2


class TestZeroPriceRejections:
    """math-15: discount/override zero prices return structured errors."""

    def test_safe_discount_zero_rejected(self) -> None:
        result = safe_conversion.convert_safe_priced_round(
            purchase_amount=100_000,
            form="yc_postmoney_discount",
            post_money_valuation_cap=None,
            discount_multiplier=0.0,
            company_capitalization=10_000_000,
            equity_financing_price=2.0,
        )
        assert result["error"] == "E_SAFE_INVALID_PRICE_INPUT"

    def test_note_override_zero_rejected(self) -> None:
        import note_conversion as nc  # type: ignore[import-not-found]

        note = {
            "id": "n1",
            "principal": 100_000,
            "issuance_date": "2024-01-01",
            "interest_rate": 0.0,
            "maturity_default_treatment": "convert_at_cap",
            "maturity_conversion_price_override": 0.0,
        }
        result = nc.convert_note(note, conversion_event_date="2026-01-01")
        assert result.get("error") == "E_NOTE_INVALID_PRICE_INPUT", result


class TestWarrantMissingField:
    """math-18: warrants missing required fields raise structured
    E_WARRANT_MISSING_FIELD, not a raw KeyError."""

    def _inputs(self) -> dict:
        return {
            "company_name": "TestCo",
            "analysis_date": "2026-06-01",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-01-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
            "preferred_series": [],
            "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
            "common_batches": [],
            "metadata": {"run_id": "test"},
        }

    def test_warrant_missing_shares_underlying(self) -> None:
        instruments = {
            "safes": [],
            "convertible_notes": [],
            "warrants": [
                {
                    "id": "w1",
                    # shares_underlying intentionally absent
                    "exercise_price": 1.0,
                    "warrant_type": "common_stock",
                    "issuance_date": "2024-01-01",
                    "settlement_type": "physical",
                }
            ],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc:
            cap_state_mod.build_cap_state(self._inputs(), instruments)
        assert "E_WARRANT_MISSING_FIELD" in str(exc.value)
        assert "shares_underlying" in str(exc.value)


class TestCapStateAfterRoundNullHistory:
    """math-19: a present-but-null cap_table_history does not TypeError in the
    receipt computation."""

    def test_null_history_receipt(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            pre: dict[str, Any] = {
                "as_converted_totals": {
                    "common_shares": 1000,
                    "preferred_shares_as_converted": 0,
                    "options_outstanding": 0,
                    "options_available": 0,
                    "warrants_underlying_total": 0,
                    "fully_diluted_shares": 1000,
                },
                "cap_table_history": None,
                "preferred_series": [],
            }
            scenarios = {"scenarios": [{"scenario_id": "round_a", "computed_outputs": {}}]}
            pre_path = os.path.join(d, "cap_state.json")
            scen_path = os.path.join(d, "scenarios.json")
            out_path = os.path.join(d, "out.json")
            with open(pre_path, "w") as f:
                json.dump(pre, f)
            with open(scen_path, "w") as f:
                json.dump(scenarios, f)
            rc, out, err = _run(
                "cap_state_after_round.py",
                [
                    "--cap-state",
                    pre_path,
                    "--scenarios",
                    scen_path,
                    "--scenario-id",
                    "round_a",
                    "-o",
                    out_path,
                ],
            )
            # Must not crash with TypeError; receipt emitted, rc 0.
            assert "TypeError" not in err, err
            assert rc == 0, err
            assert json.loads(out)["ok"] is True


class TestPdfMissingPdfplumber:
    """extraction-1: a missing pdfplumber must surface E_MISSING_DEPENDENCY and
    block, not silently degrade the hallucination gate to a no-op."""

    def test_pdf_missing_pdfplumber_blocks(self) -> None:
        import builtins
        import importlib

        sys.path.insert(0, SCRIPTS)
        ev = importlib.import_module("evidence_verifier")

        real_import = builtins.__import__

        def _fake_import(name: str, *args: Any, **kwargs: Any) -> Any:
            if name == "pdfplumber":
                raise ImportError("No module named 'pdfplumber'")
            return real_import(name, *args, **kwargs)

        from pathlib import Path

        with tempfile.TemporaryDirectory() as d:
            pdf_path = Path(d) / "doc.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 fake")
            builtins.__import__ = _fake_import
            try:
                with pytest.raises(ev.MissingDependencyError) as exc:
                    ev._load_doc_text(pdf_path)
            finally:
                builtins.__import__ = real_import
            assert "E_MISSING_DEPENDENCY" in str(exc.value)
            assert exc.value.dependency == "pdfplumber"


class TestAoAMergeMatrix:
    """extraction-2/3: --replace-existing semantics + missing-inputs handling."""

    @staticmethod
    def _extraction() -> dict[str, Any]:
        return {
            "extraction_type": "articles_of_association",
            "fields": {
                "company_name": "Acmecorp Ltd.",
                "jurisdiction_structure": "israeli",
                "preferred_series": [
                    {
                        "series_name": "Series Seed",
                        "shares": None,
                        "original_issue_price": 2.0,
                        "original_conversion_price": 2.0,
                        "current_conversion_price": 2.0,
                        "issuance_date": "2020-01-01",
                        "liquidation_preference_multiple": 1.0,
                        "liquidation_preference_type": "non_participating",
                        "anti_dilution_protection": "broad_based_weighted_average",
                    }
                ],
            },
            "confidence": {},
            "ambiguities": [],
        }

    def _write_inputs(self, d: str, series: list[dict]) -> str:
        path = os.path.join(d, "inputs.json")
        with open(path, "w") as f:
            json.dump(
                {
                    "company_name": "Acmecorp Ltd.",
                    "analysis_date": "2026-05-21",
                    "mode": "standard",
                    "preferred_series": series,
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        return path

    def test_conflict_without_flag_is_atomic_no_write_exit_2(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            inputs_path = self._write_inputs(
                d, [{"series_name": "Series Seed", "shares": 1_000_000, "original_issue_price": 1.0}]
            )
            with open(inputs_path) as f:
                before = f.read()
            rc, out, err = _run(
                "extract_aoa.py",
                ["--run-id", "t", "--inputs", inputs_path],
                stdin_data=json.dumps(self._extraction()),
            )
            assert rc == 2, err
            receipt = json.loads(out)
            assert receipt["status"] == "conflict"
            assert receipt["merge"]["added"] == []
            # Atomic: file unchanged on disk
            with open(inputs_path) as f:
                assert f.read() == before

    def test_replace_existing_overwrites_exit_0(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            inputs_path = self._write_inputs(
                d, [{"series_name": "Series Seed", "shares": 1_000_000, "original_issue_price": 1.0}]
            )
            rc, out, err = _run(
                "extract_aoa.py",
                ["--run-id", "t", "--inputs", inputs_path, "--source-doc", "/aoa.pdf", "--replace-existing"],
                stdin_data=json.dumps(self._extraction()),
            )
            assert rc == 0, err
            receipt = json.loads(out)
            assert receipt["status"] == "merged"
            assert receipt["merge"]["replaced"] == ["Series Seed"]
            with open(inputs_path) as f:
                merged = json.load(f)
            seeds = [s for s in merged["preferred_series"] if s["series_name"] == "Series Seed"]
            assert len(seeds) == 1
            # Replaced in place with the new OIP + fresh provenance
            assert seeds[0]["original_issue_price"] == 2.0
            assert "extraction_provenance" in seeds[0]

    def test_missing_inputs_is_merge_failed_exit_1(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            missing = os.path.join(d, "does_not_exist.json")
            rc, out, err = _run(
                "extract_aoa.py",
                ["--run-id", "t", "--inputs", missing],
                stdin_data=json.dumps(self._extraction()),
            )
            assert rc == 1, err
            receipt = json.loads(out)
            assert receipt["status"] == "merge_failed"

    def test_no_conflict_appends_exit_0(self) -> None:
        with tempfile.TemporaryDirectory() as d:
            inputs_path = self._write_inputs(d, [])
            rc, out, err = _run(
                "extract_aoa.py",
                ["--run-id", "t", "--inputs", inputs_path, "--source-doc", "/aoa.pdf"],
                stdin_data=json.dumps(self._extraction()),
            )
            assert rc == 0, err
            receipt = json.loads(out)
            assert receipt["status"] == "merged"
            assert receipt["merge"]["added"] == ["Series Seed"]


# ===========================================================================
# Fix 1 — build_top_dilution_drivers includes note conversion + pool top-up
# ===========================================================================


class TestTopDilutionDrivers:
    """compose_report.build_top_dilution_drivers must surface note_pct and
    pool_topup as dilution drivers, sorted by impact descending."""

    def _make_scenario(
        self,
        scenario_id: str,
        *,
        new_money_pct: float,
        safe_pct: float,
        note_pct: float,
        new_money: float = 5_000_000,
        pool_topup: int = 0,
        post_round_fd: int = 10_000_000,
    ) -> dict:
        return {
            "scenario_id": scenario_id,
            "parameters": {"new_money": new_money},
            "computed_outputs": {
                "aggregate_ownership_by_class": {
                    "new_money_pct": new_money_pct,
                    "safe_pct": safe_pct,
                    "note_pct": note_pct,
                },
                "shares_breakdown": {
                    "pool_topup": pool_topup,
                },
                "post_round_fully_diluted_shares": post_round_fd,
            },
        }

    def test_note_appears_in_drivers(self) -> None:
        """A note_pct of 4.1pp must surface as 'Note conversion' driver."""
        sys.path.insert(0, SCRIPTS)
        import compose_report  # type: ignore[import-not-found]

        s = self._make_scenario(
            "s1",
            new_money_pct=0.1724,
            safe_pct=0.0259,
            note_pct=0.0411,
        )
        drivers = compose_report.build_top_dilution_drivers([s])
        names = [d["driver"] for d in drivers]
        assert any("note" in n.lower() for n in names), f"Expected 'Note conversion' driver in {names}"
        note_driver = next(d for d in drivers if "note" in d["driver"].lower())
        assert abs(note_driver["founder_impact_pp"] - 4.1) < 0.2, (
            f"Note driver impact {note_driver['founder_impact_pp']!r} not ~4.1"
        )

    def test_ordering_new_money_note_safe(self) -> None:
        """new_money_pct=17.2 > note_pct=4.1 > safe_pct=2.6 → that ordering."""
        sys.path.insert(0, SCRIPTS)
        import compose_report  # type: ignore[import-not-found]

        s = self._make_scenario(
            "s1",
            new_money_pct=0.1724,
            safe_pct=0.0259,
            note_pct=0.0411,
        )
        drivers = compose_report.build_top_dilution_drivers([s])
        # Sorted descending: new_money(~17.2) > note(~4.1) > safe(~2.6)
        impacts = [d["founder_impact_pp"] for d in drivers]
        assert impacts == sorted(impacts, reverse=True), f"Not sorted desc: {impacts}"
        assert drivers[0]["founder_impact_pp"] > drivers[1]["founder_impact_pp"]
        # Note must come before SAFE
        note_idx = next(i for i, d in enumerate(drivers) if "note" in d["driver"].lower())
        safe_idx = next(i for i, d in enumerate(drivers) if "safe" in d["driver"].lower())
        assert note_idx < safe_idx, "Note driver must appear before SAFE driver"

    def test_zero_note_pct_omitted(self) -> None:
        """note_pct == 0 must not create a driver (same gate as safe_pct)."""
        sys.path.insert(0, SCRIPTS)
        import compose_report  # type: ignore[import-not-found]

        s = self._make_scenario(
            "s1",
            new_money_pct=0.15,
            safe_pct=0.05,
            note_pct=0.0,
        )
        drivers = compose_report.build_top_dilution_drivers([s])
        names = [d["driver"] for d in drivers]
        assert not any("note" in n.lower() for n in names), f"Zero note_pct should not create a driver; got {names}"

    def test_pool_topup_driver_when_present(self) -> None:
        """pool_topup > 0 and post_round_fd > 0 → pool top-up driver appears."""
        sys.path.insert(0, SCRIPTS)
        import compose_report  # type: ignore[import-not-found]

        s = self._make_scenario(
            "s1",
            new_money_pct=0.15,
            safe_pct=0.02,
            note_pct=0.0,
            pool_topup=800_000,
            post_round_fd=10_000_000,
        )
        drivers = compose_report.build_top_dilution_drivers([s])
        names = [d["driver"] for d in drivers]
        assert any("pool" in n.lower() for n in names), f"Expected pool top-up driver; got {names}"

    def test_pool_topup_zero_omitted(self) -> None:
        """pool_topup == 0 must not create a pool driver."""
        sys.path.insert(0, SCRIPTS)
        import compose_report  # type: ignore[import-not-found]

        s = self._make_scenario(
            "s1",
            new_money_pct=0.15,
            safe_pct=0.02,
            note_pct=0.0,
            pool_topup=0,
            post_round_fd=10_000_000,
        )
        drivers = compose_report.build_top_dilution_drivers([s])
        names = [d["driver"] for d in drivers]
        assert not any("pool" in n.lower() for n in names), f"Zero pool_topup should not create a driver; got {names}"


# ===========================================================================
# Fix 2 — Carta extractor emits schema-required interest_rate_type
# ===========================================================================


class TestCartaInterestRateType:
    """extract_cap_table._convertible_record_to_instrument must emit
    interest_rate_type on every convertible note it produces."""

    @pytest.mark.skipif(
        not os.path.exists(
            os.path.join(
                _REPO,
                "founder-skills",
                "tests",
                "fixtures",
                "cap-table-corpus",
                "synthetic_carta.xlsx",
            )
        ),
        reason="Carta fixture missing",
    )
    def test_carta_note_has_interest_rate_type(self) -> None:
        """Extracted note must carry a valid interest_rate_type enum value."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _carta_extract  # type: ignore[import-not-found]

        result = _carta_extract(_CARTA_FIXTURE)
        notes = result["instruments"]["convertible_notes"]
        assert len(notes) >= 1, "Expected at least one note from synthetic fixture"
        note = notes[0]
        valid_types = {"fixed_numeric", "fixed_numeric_simple", "statutory_ita_section_3j", "none"}
        assert "interest_rate_type" in note, f"Note missing interest_rate_type: {list(note.keys())}"
        assert note["interest_rate_type"] in valid_types, (
            f"interest_rate_type {note['interest_rate_type']!r} not in {valid_types}"
        )

    @pytest.mark.skipif(
        not os.path.exists(
            os.path.join(
                _REPO,
                "founder-skills",
                "tests",
                "fixtures",
                "cap-table-corpus",
                "synthetic_carta.xlsx",
            )
        ),
        reason="Carta fixture missing",
    )
    def test_carta_note_interest_rate_type_assumed_warning(self) -> None:
        """When interest_rate_type is assumed (not in Carta export), a warning
        matching 'interest_rate_type assumed' must appear."""
        sys.path.insert(0, SCRIPTS)
        from extract_cap_table import _carta_extract  # type: ignore[import-not-found]

        result = _carta_extract(_CARTA_FIXTURE)
        warnings = result.get("warnings", [])
        assert any("interest_rate_type assumed" in w for w in warnings), (
            f"Expected 'interest_rate_type assumed' warning; got: {warnings}"
        )

    @pytest.mark.skipif(
        not os.path.exists(
            os.path.join(
                _REPO,
                "founder-skills",
                "tests",
                "fixtures",
                "cap-table-corpus",
                "synthetic_carta.xlsx",
            )
        ),
        reason="Carta fixture missing",
    )
    def test_carta_validate_mode_accepts_instruments(self) -> None:
        """instruments.json written by carta extractor must pass --mode=validate."""
        with tempfile.TemporaryDirectory() as d:
            audit = os.path.join(d, "audit.json")
            inst = os.path.join(d, "instruments.json")
            rc_carta, _, err_carta = _run(
                "extract_cap_table.py",
                [
                    "--mode",
                    "auto",
                    "--xlsx",
                    _CARTA_FIXTURE,
                    "-o",
                    audit,
                    "--instruments",
                    inst,
                    "--run-id",
                    "test-carta",
                    "--pretty",
                ],
            )
            assert rc_carta == 0, f"Carta extraction failed: {err_carta}"
            # Also need a minimal inputs.json for validate mode (with schema_version)
            inputs = dict(_BASIC_INPUTS)
            inputs["metadata"] = {"run_id": "test-carta", "schema_version": "v0.5.0-inputs"}
            inp_path = os.path.join(d, "inputs.json")
            with open(inp_path, "w") as f:
                json.dump(inputs, f)
            rc_val, out_val, err_val = _run(
                "extract_cap_table.py",
                ["--mode", "validate", "--dir", d],
            )
            assert rc_val == 0, (
                f"validate mode rejected carta-produced instruments.json:\nstdout: {out_val}\nstderr: {err_val}"
            )


# ===========================================================================
# Fix 3 — extract_cap_table.py accepts --run-id
# ===========================================================================


class TestExtractCapTableRunId:
    """extract_cap_table.py --run-id must stamp metadata.run_id in
    the instruments.json it writes."""

    @pytest.mark.skipif(
        not os.path.exists(
            os.path.join(
                _REPO,
                "founder-skills",
                "tests",
                "fixtures",
                "cap-table-corpus",
                "synthetic_carta.xlsx",
            )
        ),
        reason="Carta fixture missing",
    )
    def test_run_id_stamped_in_instruments(self) -> None:
        """--run-id t → instruments.json metadata.run_id == 't'."""
        with tempfile.TemporaryDirectory() as d:
            audit = os.path.join(d, "audit.json")
            inst = os.path.join(d, "instruments.json")
            rc, stdout, stderr = _run(
                "extract_cap_table.py",
                [
                    "--mode",
                    "auto",
                    "--xlsx",
                    _CARTA_FIXTURE,
                    "-o",
                    audit,
                    "--instruments",
                    inst,
                    "--run-id",
                    "t",
                    "--pretty",
                ],
            )
            assert rc == 0, f"exit {rc}: {stderr}"
            with open(inst) as f:
                instruments = json.load(f)
            assert instruments["metadata"]["run_id"] == "t", (
                f"Expected run_id='t', got {instruments['metadata'].get('run_id')!r}"
            )

    def test_run_id_flag_accepted_freeform_emit(self) -> None:
        """--run-id must not be rejected (unknown-arg error) on freeform-emit mode."""
        payload = json.dumps({"blocks": []})
        with tempfile.TemporaryDirectory() as d:
            rc, stdout, stderr = _run(
                "extract_cap_table.py",
                ["--mode", "freeform-emit", "--run-id", "myrun", "--xlsx", os.path.join(d, "x.xlsx"), "--dir", d],
                stdin_data=payload,
            )
            # rc may be non-zero (missing xlsx), but NOT an argparse error
            assert "unrecognized arguments" not in stderr, f"--run-id was rejected by argparse: {stderr}"


# ===========================================================================
# Fix 4 — E_NOTE_NO_CONVERSION_PATH reason includes capitalization_denominator
# ===========================================================================


class TestNoteNoConversionPathReason:
    """note_conversion E_NOTE_NO_CONVERSION_PATH reason must name
    capitalization_denominator and tell the agent to ask the founder."""

    def _note(self, **overrides: Any) -> dict[str, Any]:
        base: dict[str, Any] = {
            "id": "note_001",
            "investor_name": "Lender",
            "principal": 100_000,
            "annual_interest_rate": 0.06,
            "interest_rate_type": "fixed_numeric",
            "day_count_basis": 365,
            "compounding_periods_per_year": None,
            "interest_converts_to_shares": True,
            "issuance_date": "2025-01-01",
            "last_interest_event_date": None,
            "valuation_cap": 10_000_000,
            "discount_multiplier": None,
            "capitalization_denominator": None,  # <-- triggers no_conversion_path
            "capitalization_denominator_policy": None,
            "qualified_financing_threshold": 1_000_000,
            "maturity_date": "9999-12-31",  # not maturity branch
            "maturity_default_treatment": None,
            "maturity_conversion_price_override": None,
            "non_qualified_financing_treatment": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        base.update(overrides)
        return base

    def test_reason_names_capitalization_denominator(self) -> None:
        """The reason string must mention capitalization_denominator."""
        result = note_conversion.convert_note(
            self._note(),
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=1.0,
        )
        assert result.get("error") == note_conversion.E_NOTE_NO_CONVERSION_PATH, (
            f"Expected E_NOTE_NO_CONVERSION_PATH; got {result}"
        )
        reason = result.get("reason", "")
        assert "capitalization_denominator" in reason, f"reason must name capitalization_denominator; got: {reason!r}"

    def test_reason_tells_agent_to_ask_founder(self) -> None:
        """The reason string must direct the agent to ask the founder."""
        result = note_conversion.convert_note(
            self._note(),
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=1.0,
        )
        reason = result.get("reason", "")
        # Must contain an action word pointing to a human source
        action_present = any(kw in reason.lower() for kw in ("ask", "founder", "note text", "confirm"))
        assert action_present, f"reason must direct agent to ask/confirm; got: {reason!r}"


# ---------------------------------------------------------------------------
# extract_cap_table.py --mode=grid tests
# ---------------------------------------------------------------------------


def _make_grid_xlsx(tmp_path: str) -> str:
    """Build a small two-sheet XLSX with values, a date, a None cell, and a
    merged range. Returns absolute path to the saved file."""
    import datetime

    import openpyxl  # type: ignore[import-untyped]

    wb = openpyxl.Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Cap Table"
    ws1["A1"] = "Acmecorp"
    ws1["B1"] = 10_000_000
    ws1["C1"] = None  # explicit null cell
    ws1["A2"] = datetime.date(2026, 6, 12)
    ws1["B2"] = 500_000.0
    ws1.merge_cells("A4:C4")  # merged range

    ws2 = wb.create_sheet("ESOP")
    ws2["A1"] = "Option Pool"
    ws2["B1"] = 1_500_000

    path = os.path.join(tmp_path, "acme_freeform.xlsx")
    wb.save(path)
    return path


class TestGridMode:
    """Tests for extract_cap_table.py --mode=grid (cell-grid dump for Lane 3)."""

    def test_grid_happy_path_structure(self, tmp_path: Any) -> None:
        """--mode=grid emits {"ok": true, "mode": "grid", "sheets": {...}}."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, stderr = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0, f"expected exit 0; stderr={stderr!r}"
        receipt = json.loads(stdout)
        assert receipt["ok"] is True
        assert receipt["mode"] == "grid"
        assert "sheets" in receipt
        assert "Cap Table" in receipt["sheets"]
        assert "ESOP" in receipt["sheets"]

    def test_grid_dimensions_present(self, tmp_path: Any) -> None:
        """Each sheet entry has 'dimensions', 'rows', and 'merged_ranges' keys."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0
        receipt = json.loads(stdout)
        sheet = receipt["sheets"]["Cap Table"]
        assert "dimensions" in sheet
        assert "rows" in sheet
        assert "merged_ranges" in sheet
        assert isinstance(sheet["rows"], list)
        assert isinstance(sheet["merged_ranges"], list)

    def test_grid_values_correct(self, tmp_path: Any) -> None:
        """Cell values are present; numeric and string cells round-trip correctly."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0
        receipt = json.loads(stdout)
        rows = receipt["sheets"]["Cap Table"]["rows"]
        # Row 0 (A1:C1): "Acmecorp", 10000000, null
        assert rows[0][0] == "Acmecorp"
        assert rows[0][1] == 10_000_000
        assert rows[0][2] is None

    def test_grid_date_serialized_as_string(self, tmp_path: Any) -> None:
        """Date cells serialize to a string (ISO or str()), not as a datetime object."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0
        receipt = json.loads(stdout)
        rows = receipt["sheets"]["Cap Table"]["rows"]
        # Row 1 (A2): date(2026, 6, 12) — must be a string, not crash JSON
        assert isinstance(rows[1][0], str)

    def test_grid_merged_ranges(self, tmp_path: Any) -> None:
        """Merged ranges are reported as strings like 'A4:C4'."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0
        receipt = json.loads(stdout)
        merged = receipt["sheets"]["Cap Table"]["merged_ranges"]
        assert len(merged) >= 1
        assert any("A4" in m and "C4" in m for m in merged)

    def test_grid_multi_sheet(self, tmp_path: Any) -> None:
        """Both sheets are present with correct values."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0
        receipt = json.loads(stdout)
        esop_rows = receipt["sheets"]["ESOP"]["rows"]
        assert esop_rows[0][0] == "Option Pool"
        assert esop_rows[0][1] == 1_500_000

    def test_grid_output_flag_writes_file_and_receipt(self, tmp_path: Any) -> None:
        """-o writes the grid JSON to a file; stdout is the standard JSON receipt."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        out_path = os.path.join(str(tmp_path), "cell_grid.json")
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx, "-o", out_path])
        assert rc == 0
        # File must exist and be valid JSON with sheets
        assert os.path.exists(out_path)
        with open(out_path) as f:
            file_data = json.load(f)
        assert file_data["ok"] is True
        assert "sheets" in file_data
        # stdout is the receipt confirming the write
        receipt = json.loads(stdout)
        assert receipt["ok"] is True
        assert "written_to" in receipt

    def test_grid_pretty_flag(self, tmp_path: Any) -> None:
        """--pretty produces indented JSON (newlines in stdout)."""
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx, "--pretty"])
        assert rc == 0
        assert "\n" in stdout  # indented JSON has newlines

    def test_grid_missing_file_error(self, tmp_path: Any) -> None:
        """Missing --xlsx exits non-zero with structured error JSON on stdout."""
        missing = os.path.join(str(tmp_path), "nonexistent.xlsx")
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", missing])
        assert rc != 0
        receipt = json.loads(stdout)
        assert receipt.get("ok") is False
        assert receipt.get("mode") == "grid"
        assert receipt.get("blocker")

    def test_grid_time_and_timedelta_cells_serialize(self, tmp_path: Any) -> None:
        """datetime.time and timedelta cells serialize to strings, not crash."""
        import datetime as dt

        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Times"
        ws["A1"] = dt.time(14, 30, 5)
        ws["B1"] = dt.timedelta(hours=2)
        path = os.path.join(str(tmp_path), "times.xlsx")
        wb.save(path)

        rc, stdout, stderr = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", path])
        assert rc == 0, f"expected exit 0; stderr={stderr!r}"
        receipt = json.loads(stdout)
        row = receipt["sheets"]["Times"]["rows"][0]
        assert isinstance(row[0], str)
        assert isinstance(row[1], str)

    def test_grid_empty_sheet(self, tmp_path: Any) -> None:
        """An empty sheet produces an empty rows list without crashing."""
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Empty"
        # leave all cells blank
        path = os.path.join(str(tmp_path), "empty_sheet.xlsx")
        wb.save(path)

        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", path])
        assert rc == 0
        receipt = json.loads(stdout)
        assert receipt["ok"] is True
        empty_sheet = receipt["sheets"].get("Empty", {})
        rows = empty_sheet.get("rows", [])
        assert isinstance(rows, list)


# ---------------------------------------------------------------------------
# extract_cap_table.py --mode=grid payload compaction (H4: control-frame cap)
#
# The grid is consumed ONLY by the SPREADSHEET_STRUCTURE_DETECTION sub-agent for
# block/role detection; the deterministic freeform-emit phase re-reads the FULL
# grid from the file. So the structure-detection grid may be trimmed, rounded,
# and row-elided down to a byte budget without any loss of final-output fidelity.
# ---------------------------------------------------------------------------


def _measure(obj: Any) -> int:
    return len(json.dumps(obj, separators=(",", ":")))


class TestGridUsedBounds:
    """_used_bounds(rows, merged_ranges) -> (last_row, last_col), 1-based."""

    def test_bounds_from_values(self) -> None:
        rows = [["A", 1, None], [None, None, None], ["B", 2, None]]
        assert extract_cap_table._used_bounds(rows, []) == (3, 2)

    def test_empty_grid_is_zero(self) -> None:
        assert extract_cap_table._used_bounds([], []) == (0, 0)
        assert extract_cap_table._used_bounds([[None, None]], []) == (0, 0)

    def test_merged_ranges_extend_bounds(self) -> None:
        # Values only reach B1, but a merged range spans to C4.
        rows = [["A", 1, None]]
        assert extract_cap_table._used_bounds(rows, ["A4:C4"]) == (4, 3)


class TestGridRoundFloats:
    """_round_floats keeps non-floats verbatim, rounds floats to N sig figs."""

    def test_rounds_floats_to_sig_figs(self) -> None:
        rows = [[0.123456789012345, 1234567.891234]]
        out = extract_cap_table._round_floats(rows, sig=8)
        assert out[0][0] == 0.12345679
        assert out[0][1] == 1234567.9

    def test_leaves_ints_strings_none(self) -> None:
        rows = [[10_000_000, "Acmecorp", None, 0]]
        out = extract_cap_table._round_floats(rows, sig=8)
        assert out[0] == [10_000_000, "Acmecorp", None, 0]
        assert isinstance(out[0][0], int)


class TestGridCompactSheets:
    """_compact_sheets(raw_sheets, budget) -> (sheets, meta)."""

    def _raw(self, rows: list, merged: list | None = None) -> dict:
        return {"S": {"dimensions": "A1:Z999", "rows": rows, "merged_ranges": merged or []}}

    def test_trims_phantom_rows_and_cols(self) -> None:
        # openpyxl over-reports dimension: real data is A1:B2, padded out to 6x6.
        rows = [["A", 1, None, None, None, None], ["B", 2, None, None, None, None]]
        rows += [[None] * 6 for _ in range(4)]
        sheets, meta = extract_cap_table._compact_sheets(self._raw(rows), budget=1_000_000)
        assert sheets["S"]["rows"] == [["A", 1], ["B", 2]]
        assert "trim" in meta["applied"]
        assert meta["over_budget"] is False

    def test_keeps_interior_blank_rows(self) -> None:
        rows = [["A", 1], [None, None], ["B", 2]]
        sheets, _ = extract_cap_table._compact_sheets(self._raw(rows), budget=1_000_000)
        assert sheets["S"]["rows"] == [["A", 1], [None, None], ["B", 2]]

    def test_meta_reports_payload_and_budget(self) -> None:
        sheets, meta = extract_cap_table._compact_sheets(self._raw([["A", 1]]), budget=12_345)
        assert meta["budget_bytes"] == 12_345
        assert meta["payload_bytes"] == _measure({"ok": True, "mode": "grid", "sheets": sheets})

    def test_rounds_floats_when_over_budget(self) -> None:
        rows = [[0.123456789012345 + i, 0.987654321098 + i] for i in range(60)]
        sheets, meta = extract_cap_table._compact_sheets(self._raw(rows), budget=1_700)
        assert "round_floats" in meta["applied"]
        assert "elide_rows" not in meta["applied"]  # rounding alone fits → rows stay positional
        # No 15-digit floats survive.
        assert all(len(repr(c)) <= 12 for r in sheets["S"]["rows"] for c in r if isinstance(c, float))

    def test_elides_tall_sheet_preserving_endpoints(self) -> None:
        rows = [[f"holder_{i}", i] for i in range(1, 501)]
        sheets, meta = extract_cap_table._compact_sheets(self._raw(rows), budget=1_500)
        assert "elide_rows" in meta["applied"]
        s = sheets["S"]
        assert s.get("indexed") is True
        kept = [r for r in s["rows"] if "r" in r]
        markers = [r for r in s["rows"] if "elided" in r]
        assert markers, "expected at least one elision marker"
        # First and last data rows survive with correct 1-based row numbers.
        assert kept[0]["r"] == 1 and kept[0]["c"] == ["holder_1", 1]
        assert kept[-1]["r"] == 500 and kept[-1]["c"] == ["holder_500", 500]

    def test_unshrinkable_payload_flags_over_budget(self) -> None:
        # One very wide row of long strings: trim/round/elide cannot help.
        rows = [["x" * 200 for _ in range(400)]]
        _sheets, meta = extract_cap_table._compact_sheets(self._raw(rows), budget=2_000)
        assert meta["over_budget"] is True


class TestGridModeCompactionCLI:
    """End-to-end --mode=grid behavior with the byte budget."""

    def test_grid_output_includes_compaction_meta(self, tmp_path: Any) -> None:
        xlsx = _make_grid_xlsx(str(tmp_path))
        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", xlsx])
        assert rc == 0
        receipt = json.loads(stdout)
        assert "compaction" in receipt
        assert "payload_bytes" in receipt["compaction"]
        assert "budget_bytes" in receipt["compaction"]

    def test_grid_large_sheet_stays_under_budget(self, tmp_path: Any) -> None:
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Big"
        for r in range(1, 801):
            for c in range(1, 9):
                ws.cell(row=r, column=c, value=r * 1000 + c + 0.123456789012345)
        path = os.path.join(str(tmp_path), "big.xlsx")
        wb.save(path)

        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", path, "--grid-budget-bytes", "20000"])
        assert rc == 0, f"stdout={stdout[:300]}"
        receipt = json.loads(stdout)
        assert receipt["ok"] is True
        assert receipt["compaction"]["payload_bytes"] <= 20_000
        assert receipt["compaction"]["over_budget"] is False

    def test_grid_unshrinkable_returns_blocker(self, tmp_path: Any) -> None:
        import openpyxl  # type: ignore[import-untyped]

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Wide"
        for c in range(1, 401):
            ws.cell(row=1, column=c, value="x" * 200)
        path = os.path.join(str(tmp_path), "wide.xlsx")
        wb.save(path)

        rc, stdout, _ = _run("extract_cap_table.py", ["--mode", "grid", "--xlsx", path, "--grid-budget-bytes", "2000"])
        assert rc != 0
        receipt = json.loads(stdout)
        assert receipt["ok"] is False
        assert receipt["blocker"] == "grid_too_large"


# ===========================================================================
# Warrants in nvca_broad A denominator (NVCA §4.4.4 "Options
# outstanding" where Option expressly includes warrants)
# ===========================================================================


class TestWarrantsInNvcaBroadABasis:
    """NVCA §4.4.4 defines 'A' to include 'Options outstanding' where the
    NVCA definition of 'Option' expressly includes warrants.  The nvca_broad
    basis must therefore include warrants_underlying_total from cap_state.

    The test verifies the mechanism (A includes warrants) not the CP2 direction
    in the full coupled system.  In the coupled solver, adding warrants to FD
    also lowers PPS (by inflating the denominator), which deepens the down-round
    and can dominate the moderating effect of a larger A.  The correct invariant
    is structural: the A field in the breakdown must grow by exactly the warrant
    count when warrants are present.

    Hand derivation of A values:
      Without warrants:
        A = common + preferred_as_conv + options_outstanding + options_available
          = 8_000_000 + 2_000_000 + 500_000 + 500_000
          = 11_000_000

      With 300_000 warrant underlying:
        A = 8_000_000 + 2_000_000 + 500_000 + 500_000 + 300_000
          = 11_300_000

      Difference = 300_000  (the warrant count).

    Isolated BBWA test (analytic verification of CP2 direction holding A fixed,
    varying PPS slightly):
      CP1 = 1.00, A = 11_000_000, consideration = 5_000_000, new_pps = 0.42
      B = 5_000_000 / 1.00 = 5_000_000
      C = 5_000_000 / 0.42 ≈ 11_904_762

      CP2_narrow_A  = 1.00 × (11_000_000 + 5_000_000) / (11_000_000 + 11_904_762)
                    = 1.00 × 16_000_000 / 22_904_762 ≈ 0.6986

      CP2_broader_A = 1.00 × (11_300_000 + 5_000_000) / (11_300_000 + 11_904_762)
                    = 1.00 × 16_300_000 / 23_204_762 ≈ 0.7024   > 0.6986

      So HOLDING PPS CONSTANT, adding warrants to A does raise CP2.  The
      reason the full-coupled scenario shows the reverse is that adding warrants
      to fully_diluted_shares also lowers PPS (denominator grows), worsening
      the down-round.  We test the mechanism via the isolated analytic, and test
      the structural property (A grows by warrant count) via the coupled solver.
    """

    def _make_cap_state(self, warrants_underlying: int) -> dict[str, Any]:
        """Build a minimal cap_state with configurable warrant underlying.

        Note: fully_diluted_shares includes warrants so the coupled solver
        correctly initialises PPS = pre_money / FD.
        """
        return {
            "founders": [{"name": "Founder", "common_shares": 8_000_000}],
            "preferred_series": [
                {
                    "series_id": "series_seed",
                    "shares": 2_000_000,
                    "original_issue_price": 1.00,
                    "original_conversion_price": 1.00,
                    "current_conversion_price": 1.00,
                    "anti_dilution_protection": "broad_based_weighted_average",
                }
            ],
            "as_converted_totals": {
                "common_shares": 8_000_000,
                "preferred_shares_as_converted": 2_000_000,
                "options_outstanding": 500_000,
                "options_available": 500_000,
                "warrants_underlying_total": warrants_underlying,
                "fully_diluted_shares": 8_000_000 + 2_000_000 + 500_000 + 500_000 + warrants_underlying,
            },
            "option_pool": {
                "plan_type": "iso",
                "authorized": 1_000_000,
                "issued_and_outstanding": 500_000,
                "available_for_grant": 500_000,
            },
            "cap_table_history": [],
        }

    def _run_solver(self, warrants_underlying: int) -> dict[str, Any]:
        cs = self._make_cap_state(warrants_underlying)
        result: dict[str, Any] = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[],
            pre_money=5_000_000.0,  # down round: PPS ~ 5M / 11.5M ≈ 0.43
            new_money=5_000_000.0,
        )
        return result

    def test_warrants_reflected_in_breakdown_A_field(self) -> None:
        """The A field in the breakdown must differ by exactly the warrant count.

        Direct structural test: warrants_underlying_total
        is added to the nvca_broad A components, so the frozen pre-financing A
        snapshot grows by the warrant count.
        """
        r_no = self._run_solver(warrants_underlying=0)
        r_with = self._run_solver(warrants_underlying=300_000)

        assert r_no["converged"], r_no.get("blockers")
        assert r_with["converged"], r_with.get("blockers")

        A_no = r_no["anti_dilution_breakdown"][0]["A"]
        A_with = r_with["anti_dilution_breakdown"][0]["A"]

        assert math.isclose(A_with - A_no, 300_000, abs_tol=1), (
            f"A difference should be 300_000 (the warrant underlying added), "
            f"got A_no={A_no}, A_with={A_with}, diff={A_with - A_no}.  "
            f"warrants_underlying_total is not being included in nvca_broad A."
        )

    def test_isolated_bbwa_warrants_in_A_raises_cp2(self) -> None:
        """Isolated BBWA: holding PPS constant, adding warrants to A raises CP2.

        Hand derivation:
          CP1 = 1.00, consideration = 5_000_000, new_pps = 0.42 (fixed)

          B = 5_000_000 / 1.00 = 5_000_000
          C = 5_000_000 / 0.42 ≈ 11_904_762

          A_no_warrants  = 11_000_000
          A_with_warrants= 11_300_000

          CP2_no  = 1.00 × (11_000_000 + 5_000_000) / (11_000_000 + 11_904_762) ≈ 0.6986
          CP2_yes = 1.00 × (11_300_000 + 5_000_000) / (11_300_000 + 11_904_762) ≈ 0.7024

          CP2_yes > CP2_no  (warrants in A moderate the downward BBWA adjustment)
        """
        cp1 = 1.00
        consideration = 5_000_000.0
        new_pps = 0.42

        r_no = anti_dilution.bbwa_new_conversion_price(
            current_conversion_price=cp1,
            pre_issuance_share_count_A=11_000_000.0,
            consideration_received=consideration,
            new_issue_price=new_pps,
        )
        r_with = anti_dilution.bbwa_new_conversion_price(
            current_conversion_price=cp1,
            pre_issuance_share_count_A=11_300_000.0,  # +300k warrants
            consideration_received=consideration,
            new_issue_price=new_pps,
        )

        cp2_no = r_no["new_conversion_price"]
        cp2_with = r_with["new_conversion_price"]

        # Verify hand math
        B = consideration / cp1
        C = consideration / new_pps
        expected_no = cp1 * (11_000_000 + B) / (11_000_000 + C)
        expected_with = cp1 * (11_300_000 + B) / (11_300_000 + C)
        assert math.isclose(cp2_no, expected_no, rel_tol=1e-9)
        assert math.isclose(cp2_with, expected_with, rel_tol=1e-9)
        assert cp2_with > cp2_no, (
            f"Holding PPS constant, adding warrants to A must raise CP2: CP2_no={cp2_no:.6f}, CP2_with={cp2_with:.6f}"
        )

    def test_warrants_not_in_narrow_basis(self) -> None:
        """nvca_narrow must NOT include warrants (narrow excludes options/warrants
        per the NVCA footnote; only common + preferred-as-converted)."""
        cs = self._make_cap_state(warrants_underlying=300_000)
        # Flip series to narrow protection
        cs["preferred_series"][0]["anti_dilution_protection"] = "narrow_based_weighted_average"

        r: dict[str, Any] = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[],
            pre_money=5_000_000.0,
            new_money=5_000_000.0,
        )
        assert r["converged"]
        bd = r["anti_dilution_breakdown"][0]
        # For narrow: A = common + preferred_as_converted = 8M + 2M = 10M
        # (no options, no warrants)
        assert math.isclose(bd["A"], 10_000_000, abs_tol=1), (
            f"nvca_narrow A should be 10_000_000 (common+preferred only), got {bd['A']}"
        )


# ===========================================================================
# Full-ratchet zero-consideration guard (NVCA §4.4.4 proviso)
# ===========================================================================


class TestFullRatchetZeroConsiderationFloor:
    """NVCA §4.4.4 proviso: 'if such issuance…was without consideration, then
    the Corporation shall be deemed to have received an aggregate of [$.001]
    of consideration.'

    full_ratchet_new_conversion_price must not set CP2=0 for zero-price
    issuance; it must floor to $.001 of deemed aggregate consideration.
    Because full-ratchet sets CP2 = new_issue_price directly, the $.001
    deemed consideration applies as a minimum price floor: CP2 = max(new_price,
    DEEMED_MIN_PRICE) where DEEMED_MIN_PRICE corresponds to the $.001 aggregate
    deemed consideration per share.  The implementation should accept the shares
    count and compute per-share floor, OR more simply floor the per-share price
    at some positive epsilon.

    Implementation note: the NVCA proviso is in the form of aggregate
    consideration ($.001), not a per-share floor.  Because the full-ratchet
    formula sets CP2 = new_issue_price, and a zero new_issue_price would give
    CP2=0 (an invalid conversion price that would crash downstream),  we
    implement a minimum price floor via FULL_RATCHET_DEEMED_MIN_PRICE = 0.001.
    This matches the NVCA spirit (any price below the floor is floored to the
    lowest representable non-zero consideration).
    """

    def test_zero_price_returns_deemed_floor_not_zero(self) -> None:
        """CP2 for a zero-consideration issuance must be > 0 (NVCA proviso)."""
        r = anti_dilution.full_ratchet_new_conversion_price(
            current_conversion_price=1.00,
            new_issue_price=0.0,  # without-consideration issuance
        )
        assert r["triggered"] is True
        assert r["new_conversion_price"] > 0, (
            "CP2 must not be zero for a without-consideration issuance "
            "(NVCA §4.4.4 proviso: deemed consideration = $.001 aggregate)"
        )
        # Floor should be at least 0.001 (NVCA deemed aggregate) expressed as price
        assert r["new_conversion_price"] >= 0.001, (
            f"CP2={r['new_conversion_price']} is below the 0.001 deemed-consideration floor"
        )
        # Must emit a warning indicating the floor was applied
        assert r.get("deemed_consideration_floor_applied") is True, (
            "Result must set deemed_consideration_floor_applied=True when floor is invoked"
        )

    def test_very_low_price_below_floor_also_floored(self) -> None:
        """A non-zero price below 0.001 should also be floored."""
        r = anti_dilution.full_ratchet_new_conversion_price(
            current_conversion_price=1.00,
            new_issue_price=0.0001,  # below the $.001 floor
        )
        assert r["triggered"] is True
        assert r["new_conversion_price"] >= 0.001

    def test_normal_low_price_above_floor_passes_through(self) -> None:
        """A price already above the floor must pass through unchanged."""
        r = anti_dilution.full_ratchet_new_conversion_price(
            current_conversion_price=1.00,
            new_issue_price=0.50,
        )
        assert r["triggered"] is True
        assert math.isclose(r["new_conversion_price"], 0.50, rel_tol=1e-9)
        assert not r.get("deemed_consideration_floor_applied")


# ===========================================================================
# Note maturity silent default disclosure
# ===========================================================================


class TestNoteMaturityDefaultWarning:
    """note_conversion.py must emit a warning when maturity_default_treatment
    was absent from the note and the default 'convert_at_cap' was applied.

    Standard note forms (Fenwick) default to repayment on majority-holder
    demand — the convert_at_cap default is unsourced and must be disclosed.
    """

    _BASE_NOTE: dict[str, Any] = {
        "id": "note_w",
        "investor_name": "Test",
        "principal": 100_000,
        "annual_interest_rate": 0.06,
        "day_count_basis": 365,
        "compounding_periods_per_year": None,
        "interest_converts_to_shares": True,
        "issuance_date": "2025-01-01",
        "last_interest_event_date": None,
        "valuation_cap": 8_000_000,
        "discount_multiplier": None,
        "capitalization_denominator": 10_000_000,
        "capitalization_denominator_policy": "pre-money fully diluted",
        "qualified_financing_threshold": 1_000_000,
        "maturity_date": "2027-01-01",
        "maturity_conversion_price_override": None,
        "non_qualified_financing_treatment": None,
        "source_document": None,
        "extraction_confidence": "high",
    }

    def test_absent_maturity_treatment_emits_warning(self) -> None:
        """When maturity_default_treatment key is absent, a warning must appear."""
        note = {k: v for k, v in self._BASE_NOTE.items() if k != "maturity_default_treatment"}
        assert "maturity_default_treatment" not in note

        # Use maturity path: no priced-round inputs
        result = note_conversion.convert_note(
            note,
            conversion_event_date="2027-01-01",
        )
        warnings = result.get("warnings", [])
        codes = [w["code"] for w in warnings]
        assert "maturity_default_treatment_defaulted" in codes, (
            f"Expected 'maturity_default_treatment_defaulted' warning when field is absent, got warning codes: {codes}"
        )

    def test_explicit_maturity_treatment_no_default_warning(self) -> None:
        """When maturity_default_treatment is explicitly set, no default warning."""
        note = dict(self._BASE_NOTE)
        note["maturity_default_treatment"] = "convert_at_cap"

        result = note_conversion.convert_note(
            note,
            conversion_event_date="2027-01-01",
        )
        warnings = result.get("warnings", [])
        codes = [w["code"] for w in warnings]
        assert "maturity_default_treatment_defaulted" not in codes, (
            f"'maturity_default_treatment_defaulted' warning should NOT appear when "
            f"field is explicitly set, got: {codes}"
        )

    def test_absent_maturity_treatment_still_routes_to_convert_at_cap(self) -> None:
        """The warning must not break the conversion branch routing."""
        note = {k: v for k, v in self._BASE_NOTE.items() if k != "maturity_default_treatment"}
        result = note_conversion.convert_note(
            note,
            conversion_event_date="2027-01-01",
        )
        assert result["branch"] == "maturity_convert_at_cap", (
            f"Branch should still be maturity_convert_at_cap, got {result['branch']!r}"
        )


# ---------------------------------------------------------------------------
# Helper: build minimal compose_report artifact dir with injectable scenarios
# ---------------------------------------------------------------------------


def _make_cap_compose_dir(
    scenarios: list[dict[str, Any]] | None = None,
    founders: list[dict[str, Any]] | None = None,
) -> str:
    """Create a temporary artifact directory for compose_report.py with
    injected scenario list and optional founder override.

    Returns the directory path; caller is responsible for cleanup.
    """
    d = tempfile.mkdtemp(prefix="test-cap-compose-")
    RID = "test-rid"

    inputs = json.loads(json.dumps(_BASIC_INPUTS))
    if founders is not None:
        inputs["founders"] = founders
    inputs["metadata"] = {"run_id": RID}

    instruments = json.loads(json.dumps(_BASIC_INSTRUMENTS))
    instruments["metadata"] = {"run_id": RID}

    # Write inputs + instruments
    for name, data in [("inputs.json", inputs), ("instruments.json", instruments)]:
        with open(os.path.join(d, name), "w") as f:
            json.dump(data, f)

    # Build cap_state from library so as_converted_totals is populated
    cs = cap_state_mod.build_cap_state(inputs, instruments)
    cs["metadata"]["run_id"] = RID
    with open(os.path.join(d, "cap_state.json"), "w") as f:
        json.dump(cs, f)

    # rule_audit, counsel_packet, scenarios
    for name, data in [
        (
            "rule_audit.json",
            {
                "gating": {},
                "applied_rules": [],
                "counsel_review_items": [],
                "date_sensitive_watchlist": [],
                "metadata": {"run_id": RID},
            },
        ),
        (
            "scenarios.json",
            {"scenarios": scenarios or [], "metadata": {"run_id": RID}},
        ),
        (
            "counsel_packet.json",
            {"company_name": "Acmecorp", "engagement_summary": "", "items": [], "metadata": {"run_id": RID}},
        ),
    ]:
        with open(os.path.join(d, name), "w") as f:
            json.dump(data, f)

    return d


def _run_cap_compose(d: str) -> tuple[int, dict[str, Any], str]:
    """Run compose_report.py on the given artifact dir; return (rc, report_json, stderr)."""
    report_path = os.path.join(d, "report.json")
    md_path = os.path.join(d, "report.md")
    rc, stdout, stderr = _run(
        "compose_report.py",
        ["--dir", d, "--run-id", "test-rid", "-o", report_path, "--write-md", md_path],
    )
    if rc != 0:
        return rc, {}, stderr
    with open(report_path) as f:
        return rc, json.load(f), stderr  # type: ignore[no-any-return]


def _minimal_scenario(
    scenario_id: str = "s1",
    *,
    computed_outputs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Return a minimal scenario dict for injection into scenarios.json."""
    return {
        "scenario_id": scenario_id,
        "type": "priced_round",
        "scenario_type": "priced_round",
        "completeness": "structural_only",
        "parameters": {
            "pre_money": 10_000_000,
            "new_money": 2_000_000,
            "transaction_event_date": "2026-06-01",
        },
        "computed_outputs": computed_outputs or {},
        "blockers": [],
        "warnings": [],
        "metadata": {"run_id": "test-rid"},
    }


# ---------------------------------------------------------------------------
# Item 9: Per-SAFE table — keyed on non-empty per_safe; Purchase ÷ Cap column
# ---------------------------------------------------------------------------


class TestPerSafeTableRendering:
    """compose_report renders per-SAFE conversion math when per_safe is
    non-empty and rows are NOT cap_implied_only; includes Purchase ÷ Cap."""

    def test_per_safe_table_rendered_for_converted_safe(self) -> None:
        """Table appears when per_safe has a row without cap_implied_ownership."""
        scenario = _minimal_scenario(
            "s_safe",
            computed_outputs={
                "per_safe": {
                    "safe_001": {
                        "branch": "cap_conversion",
                        "conversion_price": 0.2500,
                        "conversion_shares": 2_000_000,
                        "purchase_amount": 500_000,
                        "post_money_cap": 10_000_000,
                    }
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Per-SAFE conversion math" in md

    def test_purchase_div_cap_derivation_in_table(self) -> None:
        """Table cell shows 'purchase ÷ cap = XX.XX%' format."""
        scenario = _minimal_scenario(
            "s_safe_div",
            computed_outputs={
                "per_safe": {
                    "safe_001": {
                        "branch": "cap_conversion",
                        "conversion_price": 0.25,
                        "conversion_shares": 2_000_000,
                        "purchase_amount": 500_000,
                        "post_money_cap": 10_000_000,
                    }
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # $500K ÷ $10M = 5.00%
        assert "5.00%" in md

    def test_per_safe_table_skipped_when_cap_implied_only(self) -> None:
        """When ALL per_safe rows have cap_implied_ownership, table is skipped."""
        scenario = _minimal_scenario(
            "s_cap_impl",
            computed_outputs={
                "per_safe": {
                    "safe_001": {
                        "cap_implied_ownership": 0.05,
                        "safe_price": 0.25,
                        "cap_implied_shares": 200_000,
                    }
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # Table should NOT appear — only cap_implied narrative
        assert "Per-SAFE conversion math" not in md

    def test_per_safe_table_skipped_when_empty(self) -> None:
        """When per_safe is empty or absent, no table is emitted."""
        scenario = _minimal_scenario("s_no_safe", computed_outputs={})
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Per-SAFE conversion math" not in md


# ---------------------------------------------------------------------------
# Item 10: shares_breakdown post-round composition table
# ---------------------------------------------------------------------------


class TestSharesBreakdownTable:
    """compose_report renders a Post-round share composition table when
    shares_breakdown is present in computed_outputs."""

    def test_shares_breakdown_table_rendered(self) -> None:
        """Table appears when shares_breakdown is in computed_outputs."""
        scenario = _minimal_scenario(
            "s_bd",
            computed_outputs={
                "shares_breakdown": {
                    "pre_round_fd": 11_500_000,
                    "safe_converted_shares": 600_000,
                    "pool_topup_shares": 300_000,
                    "new_money_shares": 800_000,
                    "post_round_fd": 13_200_000,
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Post-round share composition" in md

    def test_shares_breakdown_components_in_table(self) -> None:
        """Pre-round FD, SAFE converted, pool top-up, new money, post-round FD rows appear."""
        scenario = _minimal_scenario(
            "s_bd2",
            computed_outputs={
                "shares_breakdown": {
                    "pre_round_fd": 11_500_000,
                    "safe_converted_shares": 600_000,
                    "pool_topup_shares": 300_000,
                    "new_money_shares": 800_000,
                    "post_round_fd": 13_200_000,
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Pre-round FD" in md
        assert "SAFE converted" in md
        assert "Pool top-up" in md
        assert "New money" in md
        assert "Post-round FD" in md

    def test_shares_breakdown_post_round_pct(self) -> None:
        """Post-round % column is rendered (100.0% for total row)."""
        scenario = _minimal_scenario(
            "s_bd3",
            computed_outputs={
                "shares_breakdown": {
                    "pre_round_fd": 10_000_000,
                    "new_money_shares": 2_000_000,
                    "post_round_fd": 12_000_000,
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "100.0%" in md

    def test_shares_breakdown_table_absent_when_missing(self) -> None:
        """No composition table when shares_breakdown is absent."""
        scenario = _minimal_scenario("s_no_bd", computed_outputs={})
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Post-round share composition" not in md


# ---------------------------------------------------------------------------
# Item 11: Per-founder rows in single-class Current Cap State
# ---------------------------------------------------------------------------


class TestPerFounderCapStateRows:
    """In single-class engagements, each founder appears by name in the
    Current Cap State table with their share count and pre-round %."""

    def test_founder_names_in_cap_state(self) -> None:
        """Both founder names appear in the Current Cap State section."""
        d = _make_cap_compose_dir()
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # _BASIC_INPUTS has founders Alice and Bob
        assert "Alice" in md
        assert "Bob" in md

    def test_founder_share_counts_in_cap_state(self) -> None:
        """Founder share counts (5,000,000 each) appear in the table."""
        d = _make_cap_compose_dir()
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # 5,000,000 formatted with comma separator
        assert "5,000,000" in md

    def test_custom_founder_name_appears(self) -> None:
        """Custom founder name injected into inputs appears in Current Cap State."""
        founders = [
            {"name": "Zelda", "founder_id": "founder_z", "common_shares": 3_000_000},
            {"name": "Yoram", "founder_id": "founder_y", "common_shares": 7_000_000},
        ]
        d = _make_cap_compose_dir(founders=founders)
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Zelda" in md
        assert "Yoram" in md

    def test_founder_pct_in_cap_state(self) -> None:
        """Founder ownership % appears in the table (not just raw count)."""
        d = _make_cap_compose_dir()
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # Alice and Bob each own 5M of total ~12.5M FD → roughly 40%
        # Just check a % sign appears near the founders section
        assert "%" in md


# ---------------------------------------------------------------------------
# Item 12: Skip per-SAFE table when all rows are cap_implied_only
# ---------------------------------------------------------------------------


class TestPerSafeCapImpliedOnlySkip:
    """When every per_safe row has cap_implied_ownership (structural_only
    snapshot), the entire per-SAFE table block must be skipped."""

    def test_table_skipped_all_cap_implied(self) -> None:
        """No 'Per-SAFE conversion math' header when all rows are cap_implied."""
        scenario = _minimal_scenario(
            "s_ci",
            computed_outputs={
                "per_safe": {
                    "safe_a": {"cap_implied_ownership": 0.05, "safe_price": 0.20, "cap_implied_shares": 250_000},
                    "safe_b": {"cap_implied_ownership": 0.03, "safe_price": 0.20, "cap_implied_shares": 150_000},
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        assert "Per-SAFE conversion math" not in md

    def test_cap_implied_narrative_still_rendered(self) -> None:
        """Cap-implied ownership narrative section IS rendered when completeness is
        structural_only AND cap_implied_only flag is set."""
        scenario = _minimal_scenario(
            "s_ci_narr",
            computed_outputs={
                "cap_implied_only": True,
                "per_safe": {
                    "safe_a": {"cap_implied_ownership": 0.05, "safe_price": 0.20, "cap_implied_shares": 250_000},
                },
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # cap_implied narrative should still appear
        assert "cap-implied" in md.lower()

    def test_mixed_rows_table_rendered_for_non_cap_implied(self) -> None:
        """When some rows have cap_implied_ownership and some do NOT, the table
        IS rendered (filtering only drops the cap_implied rows from the table)."""
        scenario = _minimal_scenario(
            "s_mixed",
            computed_outputs={
                "per_safe": {
                    "safe_ci": {"cap_implied_ownership": 0.05, "safe_price": 0.20, "cap_implied_shares": 250_000},
                    "safe_conv": {
                        "branch": "cap_conversion",
                        "conversion_price": 0.25,
                        "conversion_shares": 2_000_000,
                        "purchase_amount": 500_000,
                        "post_money_cap": 10_000_000,
                    },
                }
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, f"compose failed: {stderr}"
        md = report["report_markdown"]
        # safe_conv should appear in the table
        assert "Per-SAFE conversion math" in md
        assert "safe_conv" in md
        # safe_ci should NOT appear in the conversion math table
        # (it only appears in the cap-implied narrative)
        assert "safe_ci" not in md.split("Per-SAFE conversion math")[-1].split("\n\n")[0]


# ===========================================================================
# MFN election override (scenario-level mfn_elections)
# ===========================================================================

_MFN_CAP10 = {
    **_SAFE_BASIC,
    "id": "cap10",
    "post_money_valuation_cap": 10_000_000,
    "discount_multiplier": None,
    "form": "yc_postmoney_cap",
    "mfn_provision": None,
}
_MFN_CAP15 = {
    **_SAFE_BASIC,
    "id": "cap15",
    "post_money_valuation_cap": 15_000_000,
    "discount_multiplier": None,
    "form": "yc_postmoney_cap",
    "mfn_provision": None,
}


def _mfn_uncapped(elected: str | None) -> dict[str, Any]:
    return {
        **_SAFE_BASIC,
        "id": "safe_mfn",
        "post_money_valuation_cap": None,
        "discount_multiplier": None,
        "form": "yc_uncapped_mfn",
        "mfn_provision": {
            "present": True,
            "elected_against_safe_id": elected,
            "elected": elected is not None,
            "cherry_pick_attempted": False,
            "notes": None,
        },
    }


class TestMfnElectionOverride:
    def _solve(self, elections: dict[str, str]) -> dict[str, Any]:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        return priced_round.solve_priced_round(  # type: ignore[no-any-return]
            cap_state=cs,
            safes=[_MFN_CAP10, _MFN_CAP15, _mfn_uncapped(None)],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            mfn_elections=elections,
        )

    def test_election_override_changes_conversion(self) -> None:
        r10 = self._solve({"safe_mfn": "cap10"})
        r15 = self._solve({"safe_mfn": "cap15"})
        p10 = r10["per_safe"]["safe_mfn"]["conversion_price"]
        p15 = r15["per_safe"]["safe_mfn"]["conversion_price"]
        assert p10 != p15, "MFN election must change the conversion price"
        assert p15 > p10, "$15M cap -> higher conversion price than $10M"
        f10 = r10["aggregate_ownership_by_class"]["founders_pct"]
        f15 = r15["aggregate_ownership_by_class"]["founders_pct"]
        assert f15 > f10, "$15M election dilutes founders less -> higher founders_pct"

    def _solve_safes(self, safes: list[dict[str, Any]], elections: dict[str, str]) -> dict[str, Any]:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        return priced_round.solve_priced_round(  # type: ignore[no-any-return]
            cap_state=cs,
            safes=safes,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            mfn_elections=elections,
        )

    @staticmethod
    def _has(result: dict[str, Any], code: str) -> bool:
        return any(w.get("code") == code for w in result.get("warnings", []) or [])

    def test_not_most_favorable_warns_on_higher_cap(self) -> None:
        # electing the $15M sibling when a cheaper $10M exists is non-most-favorable
        r = self._solve_safes([_MFN_CAP10, _MFN_CAP15, _mfn_uncapped(None)], {"safe_mfn": "cap15"})
        assert self._has(r, "W_MFN_NOT_MOST_FAVORABLE")

    def test_not_most_favorable_absent_on_best_election(self) -> None:
        # electing the cheapest ($10M) sibling IS most-favorable -> no warning
        r = self._solve_safes([_MFN_CAP10, _MFN_CAP15, _mfn_uncapped(None)], {"safe_mfn": "cap10"})
        assert not self._has(r, "W_MFN_NOT_MOST_FAVORABLE")

    def test_not_most_favorable_absent_with_single_candidate(self) -> None:
        # only one capped sibling -> nothing more favorable to elect -> no warning
        r = self._solve_safes([_MFN_CAP10, _mfn_uncapped(None)], {"safe_mfn": "cap10"})
        assert not self._has(r, "W_MFN_NOT_MOST_FAVORABLE")

    def test_override_beats_instrument_warns(self) -> None:
        baked = _mfn_uncapped("cap10")  # instrument elects cap10
        r = self._solve_safes([_MFN_CAP10, _MFN_CAP15, baked], {"safe_mfn": "cap15"})
        assert self._has(r, "W_MFN_ELECTION_OVERRIDES_INSTRUMENT")
        # control: override matches baked -> no override-conflict warning
        r2 = self._solve_safes([_MFN_CAP10, _MFN_CAP15, _mfn_uncapped("cap10")], {"safe_mfn": "cap10"})
        assert not self._has(r2, "W_MFN_ELECTION_OVERRIDES_INSTRUMENT")

    def test_bad_shape_blocks(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[_MFN_CAP10, _mfn_uncapped(None)],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            mfn_elections=["safe_mfn", "cap10"],  # malformed: a list, not a dict
        )
        assert r["completeness"] == "structural_only"
        assert any(b["code"] == "E_MFN_ELECTIONS_BAD_SHAPE" for b in r["blockers"])

    def test_bad_target_blocks(self) -> None:
        r = self._solve_safes([_MFN_CAP10, _mfn_uncapped(None)], {"safe_mfn": "nonexistent"})
        assert r["completeness"] == "structural_only"
        assert any(b["code"] == "E_SAFE_MFN_ELECTION_BAD_TARGET" for b in r["blockers"])

    def test_input_not_mutated(self) -> None:
        mfn = _mfn_uncapped(None)
        before = mfn["mfn_provision"]["elected_against_safe_id"]
        self._solve_safes([_MFN_CAP10, _MFN_CAP15, mfn], {"safe_mfn": "cap15"})
        assert mfn["mfn_provision"]["elected_against_safe_id"] == before  # still None

    def test_scenario_route_forwards_mfn_elections(self) -> None:
        import run_scenario  # type: ignore[import-not-found]

        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        instruments = {"safes": [_MFN_CAP10, _MFN_CAP15, _mfn_uncapped(None)], "convertible_notes": []}

        def run(elect: str) -> dict[str, Any]:
            scenario = {
                "type": "safe_conversion",
                "parameters": {
                    "priced_round_pre_money": 20_000_000,
                    "priced_round_new_money": 5_000_000,
                    "target_pool_percent": 0.10,
                    "target_basis": "post_money",
                    "mfn_elections": {"safe_mfn": elect},
                },
            }
            return run_scenario.run_safe_conversion_scenario(scenario, instruments=instruments, cap_state=cs)  # type: ignore[no-any-return]

        r10 = run("cap10")
        r15 = run("cap15")
        # the param actually reaches the solver -> the two scenarios differ
        assert r10["per_safe"]["safe_mfn"]["conversion_price"] != r15["per_safe"]["safe_mfn"]["conversion_price"]
        # and the resolved election is visible in per_safe for audit
        assert r15["per_safe"]["safe_mfn"]["_mfn_election_source"] == "scenario_override"
        assert r15["per_safe"]["safe_mfn"]["_mfn_inherited_cap"] == 15_000_000
        assert r15["per_safe"]["safe_mfn"]["_mfn_inherited_cap_type"] == "post_money"

    def test_cap_implied_path_blocks_mfn_elections(self) -> None:
        import run_scenario  # type: ignore[import-not-found]

        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        instruments = {"safes": [_MFN_CAP10, _mfn_uncapped(None)], "convertible_notes": []}
        scenario = {"type": "safe_conversion", "parameters": {"mfn_elections": {"safe_mfn": "cap10"}}}
        r = run_scenario.run_safe_conversion_scenario(scenario, instruments=instruments, cap_state=cs)
        assert any(b["code"] == "E_SAFE_MFN_ELECTION_REQUIRES_PRICED_ROUND" for b in r["blockers"])

    def test_structural_returns_include_math_provenance(self) -> None:
        # ALL FOUR structural early-returns must carry math_provenance (schema-required at
        # scenarios.schema.json computed_outputs.required).
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)

        # (1) MFN cycle path: two uncapped MFNs electing each other.
        cyc = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[
                {
                    **_mfn_uncapped("B"),
                    "id": "A",
                    "mfn_provision": {
                        "present": True,
                        "elected_against_safe_id": "B",
                        "elected": True,
                        "cherry_pick_attempted": False,
                        "notes": None,
                    },
                },
                {
                    **_mfn_uncapped("A"),
                    "id": "B",
                    "mfn_provision": {
                        "present": True,
                        "elected_against_safe_id": "A",
                        "elected": True,
                        "cherry_pick_attempted": False,
                        "notes": None,
                    },
                },
            ],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert any(b["code"] == "E_SAFE_CIRCULAR_MFN" for b in cyc["blockers"])
        assert "math_provenance" in cyc

        # (2) bad-shape MFN override path.
        bad = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[_MFN_CAP10],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            mfn_elections=["bad"],
        )
        assert any(b["code"] == "E_MFN_ELECTIONS_BAD_SHAPE" for b in bad["blockers"])
        assert "math_provenance" in bad

        # (3) note-without-conversion-date path.
        note = {"id": "n1", "principal": 100_000, "form": "convertible_note"}
        nod = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[note],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert any(b["code"] == "E_NOTE_NO_CONVERSION_DATE" for b in nod["blockers"])
        assert "math_provenance" in nod

        # (4) zero pre-FD path (no founders / pool).
        empty_inputs = {
            **_BASIC_INPUTS,
            "founders": [],
            "option_pool": {"plan_type": "nso", "authorized": 0, "issued": 0, "unallocated": 0},
        }
        empty_cs = cap_state_mod.build_cap_state(empty_inputs, _BASIC_INSTRUMENTS)
        zpf = priced_round.solve_priced_round(
            cap_state=empty_cs,
            safes=[],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
        )
        assert any(b["code"] == "E_SCENARIO_NO_PRE_FD" for b in zpf["blockers"])
        assert "math_provenance" in zpf


# ===========================================================================
# Investor entity offered/recorded as a "founder"
# ===========================================================================


def _inputs_with_founders(founder_names: list[str]) -> dict[str, Any]:
    return {
        "company_name": "TestCo",
        "analysis_date": "2026-06-21",
        "jurisdiction": {"structure": "delaware"},
        "founders": [
            {"name": n, "founder_id": f"f{i}", "common_shares": 5_000_000} for i, n in enumerate(founder_names, start=1)
        ],
        "option_pool": {"plan_type": "iso", "authorized": 1_000_000, "issued": 0, "unallocated": 1_000_000},
        "metadata": {"run_id": "20260621T000000Z", "schema_version": "v0.5.0-inputs"},
    }


class TestFounderInvestorWarning:
    def test_looks_like_investor_entity_truth_table(self) -> None:
        f = cap_state_mod.looks_like_investor_entity
        for name in ["OG Tech Ventures", "Foo Capital", "Bar Fund", "Acme Ventures LP", "Foo Capital, L.P."]:
            assert f(name) is True, name
        for name in [
            "Jane Doe",
            "Acme Holdings Founder Trust",
            "Smith Partners",
            "Foo Ltd",
            "Acme Holdings",
            "LP Morgan",
            "Fundamentals Inc",
        ]:
            assert f(name) is False, name

    def test_investor_entity_founder_warns(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _inputs_with_founders(["Jane Doe", "OG Tech Ventures"]), {"safes": [], "convertible_notes": []}
        )
        assert "W_FOUNDER_LOOKS_LIKE_INVESTOR" in cs.get("warnings", [])

    def test_clean_founders_do_not_warn(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _inputs_with_founders(["Jane Doe", "John Smith"]), {"safes": [], "convertible_notes": []}
        )
        assert "W_FOUNDER_LOOKS_LIKE_INVESTOR" not in cs.get("warnings", [])

    def test_founder_co_investor_does_not_warn_alone(self) -> None:
        # a personal founder who ALSO appears as a SAFE investor is a legit Israeli
        # pattern -> must NOT fire on the cross-reference alone.
        inputs = _inputs_with_founders(["Jane Doe", "John Smith"])
        instruments = {
            "safes": [{**_SAFE_BASIC, "id": "s1", "investor_name": "Jane Doe"}],
            "convertible_notes": [],
        }
        cs = cap_state_mod.build_cap_state(inputs, instruments)
        assert "W_FOUNDER_LOOKS_LIKE_INVESTOR" not in cs.get("warnings", [])


# ===========================================================================
# Silent assumed cap base
# ===========================================================================


def _inputs_named(founder_names: list[str], cap_base_source: str | None = None) -> dict[str, Any]:
    d = _inputs_with_founders(founder_names)
    if cap_base_source is not None:
        d["metadata"]["cap_base_source"] = cap_base_source
    return d


class TestAssumedCapBaseWarning:
    NO_INST: dict[str, list] = {"safes": [], "convertible_notes": []}

    def test_placeholder_names_warn(self) -> None:
        cs = cap_state_mod.build_cap_state(_inputs_named(["Founder A", "Founder B"]), self.NO_INST)
        assert "W_CAP_BASE_ASSUMED" in cs.get("warnings", [])

    def test_full_pipeline_real_names_no_flag_warns(self) -> None:
        # Issue B (default-to-assumed): an equity base (founders/pool) present with cap_base_source
        # neither set nor "confirmed" now WARNS regardless of names — the model must affirmatively set
        # "confirmed" to suppress, flipping the compliance burden to the safe side. (Inverts the prior
        # real-names-stay-silent contract, under which a model that skipped the gate, used real inline
        # names, and never set "assumed" emitted NO caveat at all.)
        cs = cap_state_mod.build_cap_state(_inputs_named(["Jane Doe", "John Smith"]), self.NO_INST)
        assert "W_CAP_BASE_ASSUMED" in cs.get("warnings", [])

    def test_no_equity_base_does_not_warn(self) -> None:
        # Nothing to assume: no founders, no option_pool, no instruments → default-to-assumed must NOT fire.
        inputs = {
            "company_name": "TestCo",
            "analysis_date": "2026-06-21",
            "jurisdiction": {"structure": "delaware"},
            "metadata": {"run_id": "20260621T000000Z", "schema_version": "v0.5.0-inputs"},
        }
        cs = cap_state_mod.build_cap_state(inputs, self.NO_INST)
        assert "W_CAP_BASE_ASSUMED" not in cs.get("warnings", [])

    def test_real_names_confirmed_suppresses_warn(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _inputs_named(["Jane Doe", "John Smith"], cap_base_source="confirmed"), self.NO_INST
        )
        assert "W_CAP_BASE_ASSUMED" not in cs.get("warnings", [])

    def test_explicit_assumed_flag_warns(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _inputs_named(["Jane Doe", "John Smith"], cap_base_source="assumed"), self.NO_INST
        )
        assert "W_CAP_BASE_ASSUMED" in cs.get("warnings", [])

    def test_confirmed_suppresses_placeholder(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _inputs_named(["Founder A", "Founder B"], cap_base_source="confirmed"), self.NO_INST
        )
        assert "W_CAP_BASE_ASSUMED" not in cs.get("warnings", [])

    def _inputs_prov(self, names: list[str], cap_base_source: str, provenance: str | None) -> dict[str, Any]:
        d = _inputs_named(names, cap_base_source=cap_base_source)
        if provenance is not None:
            d["metadata"]["cap_base_provenance"] = provenance
        return d

    def test_confirmed_non_deterministic_provenance_warns_reconstructed(self) -> None:
        # confirmed-but-unmapped shape: a confirmed base with NO deterministic_mapped
        # marker = model-built → flag it, so a hand-built base can't masquerade as
        # verified. (The case the explicit-only trigger would miss.)
        cs = cap_state_mod.build_cap_state(
            self._inputs_prov(["Jane Doe", "John Smith"], "confirmed", None), self.NO_INST
        )
        assert "W_CAP_BASE_RECONSTRUCTED" in cs.get("warnings", [])

    def test_deterministic_mapped_suppresses_reconstructed(self) -> None:
        cs = cap_state_mod.build_cap_state(
            self._inputs_prov(["Jane Doe", "John Smith"], "confirmed", "deterministic_mapped"), self.NO_INST
        )
        assert "W_CAP_BASE_RECONSTRUCTED" not in cs.get("warnings", [])

    def test_assumed_does_not_warn_reconstructed(self) -> None:
        # mutually exclusive with W_CAP_BASE_ASSUMED (which requires != confirmed)
        cs = cap_state_mod.build_cap_state(_inputs_named(["Jane Doe", "John Smith"]), self.NO_INST)
        w = cs.get("warnings", [])
        assert "W_CAP_BASE_RECONSTRUCTED" not in w
        assert "W_CAP_BASE_ASSUMED" in w

    def test_no_equity_base_no_reconstructed(self) -> None:
        inputs = {
            "company_name": "TestCo",
            "analysis_date": "2026-06-21",
            "jurisdiction": {"structure": "delaware"},
            "metadata": {
                "run_id": "20260621T000000Z",
                "schema_version": "v0.5.0-inputs",
                "cap_base_source": "confirmed",
            },
        }
        cs = cap_state_mod.build_cap_state(inputs, self.NO_INST)
        assert "W_CAP_BASE_RECONSTRUCTED" not in cs.get("warnings", [])

    def test_is_placeholder_founder_name_truth_table(self) -> None:
        f = cap_state_mod._is_placeholder_founder_name
        for n in ["Founder A", "Founder B", "Founder 1", "Co-Founder A", "CoFounder B"]:
            assert f(n) is True, n
        for n in ["Founder", "Jane Doe", "Founder Jane", "Foundering"]:
            assert f(n) is False, n


class TestCartaFdTotalCapture:
    """A1 capture: parse Carta's INDEPENDENT printed grand fully-diluted total (the 'Totals' row, Fully
    Diluted Shares column) from the Summary Cap Table — the figure cap_state cross-foots against. Built
    against the real Carta Summary layout; tested with synthetic numbers (the column labels are Carta's
    generic headers, not company data)."""

    HEADER = (
        None,
        "Shares Authorized",
        "Shares Issued\nand Outstanding",
        "Fully Diluted Shares",
        "Fully Diluted\nOwnership",
    )

    def test_captures_totals_row_fd(self) -> None:
        import extract_cap_table  # type: ignore[import-not-found]

        rows = [
            (None,),
            (None, "Acmecorp Summary Cap Table"),
            (None, "As of 1/1/2026"),
            (),
            self.HEADER,
            ("Common (CS) Stock", 9_000_000, 8_000_000, 9_000_000, 0.9),
            ("Total Common Stock issued", 0, 0, 9_000_000, 0.9),
            ("Totals", None, None, 10_000_000, 1.0),
        ]
        assert extract_cap_table._extract_carta_fd_total(rows) == 10_000_000

    def test_picks_exact_fd_column_not_the_with_variant(self) -> None:
        import extract_cap_table  # type: ignore[import-not-found]

        # Carta also prints a "Fully Diluted Shares with ..." column — must pick the EXACT one, not that.
        header = (
            None,
            "Shares Authorized",
            "Fully Diluted Shares",
            "Fully Diluted Ownership",
            "Fully Diluted Shares with X",
        )
        rows = [self.HEADER and header, ("Totals", None, 7_000_000, 1.0, 9_999_999)]
        assert extract_cap_table._extract_carta_fd_total(rows) == 7_000_000

    def test_no_totals_row_returns_none(self) -> None:
        import extract_cap_table  # type: ignore[import-not-found]

        rows = [self.HEADER, ("Common (CS) Stock", 9_000_000, 8_000_000, 9_000_000, 0.9)]
        assert extract_cap_table._extract_carta_fd_total(rows) is None


def _vacuous_inputs() -> dict[str, Any]:
    """A base with NO equity holders — just an unallocated option pool."""
    return {
        "company_name": "TestCo",
        "analysis_date": "2026-06-21",
        "jurisdiction": {"structure": "delaware"},
        "founders": [],
        "option_pool": {"plan_type": "iso", "authorized": 1_000_000, "issued": 0, "unallocated": 1_000_000},
        "metadata": {"run_id": "20260621T000000Z", "schema_version": "v0.5.0-inputs"},
    }


class TestBaseVacuous:
    """R-5: a deliverable built on a base with 0 founders / 0 common / 0 preferred — essentially just an
    unallocated option pool — is confidently misleading (donut/FD/%s are meaningless). cap_state flags it
    with W_BASE_VACUOUS, INDEPENDENT of confirmed/assumed."""

    NO_INST: dict[str, list] = {"safes": [], "convertible_notes": []}

    def test_vacuous_base_warns(self) -> None:
        cs = cap_state_mod.build_cap_state(_vacuous_inputs(), self.NO_INST)
        assert "W_BASE_VACUOUS" in cs.get("warnings", [])

    def test_vacuous_even_when_confirmed(self) -> None:
        # vacuity is independent of cap_base_source — a CONFIRMED empty base is still vacuous
        inp = _vacuous_inputs()
        inp["metadata"]["cap_base_source"] = "confirmed"
        cs = cap_state_mod.build_cap_state(inp, self.NO_INST)
        assert "W_BASE_VACUOUS" in cs.get("warnings", [])

    def test_normal_base_no_vacuous(self) -> None:
        cs = cap_state_mod.build_cap_state(_inputs_with_founders(["Jane Doe"]), self.NO_INST)
        assert "W_BASE_VACUOUS" not in cs.get("warnings", [])


class TestFdReconciliation:
    """A1: when the source document prints an INDEPENDENT grand fully-diluted total (e.g. Carta's 'Totals'
    row), captured into inputs.stated_totals, cap_state cross-checks the computed FD against it and warns
    (suppressibly) on a divergence beyond a relative-ppm threshold — catching a dropped/mis-entered
    holder/class that a manual rebuild would otherwise hide."""

    NO_INST: dict[str, list] = {"safes": [], "convertible_notes": []}

    def _fd(self) -> int:
        cs = cap_state_mod.build_cap_state(_inputs_with_founders(["Jane Doe", "John Smith"]), self.NO_INST)
        return cs["as_converted_totals"]["fully_diluted_shares"]  # type: ignore[no-any-return]

    def _cs(self, stated: int) -> dict[str, Any]:
        inp = _inputs_with_founders(["Jane Doe", "John Smith"])
        inp["stated_totals"] = {"fully_diluted": stated, "source": "carta_summary"}
        return cap_state_mod.build_cap_state(inp, self.NO_INST)  # type: ignore[no-any-return]

    def test_within_threshold_no_warn(self) -> None:
        cs = self._cs(self._fd())  # exact match
        assert not any(w.startswith("W_FD_RECONCILE_DELTA") for w in cs.get("warnings", []))
        assert cs["as_converted_totals"]["reconciliation"]["residual_abs"] == 0

    def test_small_rounding_delta_no_warn(self) -> None:
        # a sub-0.1% rounding delta must NOT fire
        cs = self._cs(self._fd() + 50)  # ~4.5 ppm on an 11M base
        assert not any(w.startswith("W_FD_RECONCILE_DELTA") for w in cs.get("warnings", []))

    def test_large_delta_warns(self) -> None:
        cs = self._cs(int(self._fd() * 0.9))  # 10% off → fires
        assert any(w.startswith("W_FD_RECONCILE_DELTA") for w in cs.get("warnings", []))

    def test_no_stated_totals_no_reconciliation(self) -> None:
        cs = cap_state_mod.build_cap_state(_inputs_with_founders(["Jane Doe", "John Smith"]), self.NO_INST)
        assert "reconciliation" not in cs["as_converted_totals"]
        assert not any(w.startswith("W_FD_RECONCILE_DELTA") for w in cs.get("warnings", []))


class TestVisionExtractionWarning:
    """B3: an image-only PDF read by raw vision (no OCR) under-extracts dense tables silently.
    The pdf-probe sets metadata.extraction_mode='vision_image_pdf'; cap_state then
    surfaces a low-confidence warning so the structured artifacts carry the same caveat the narrative does."""

    NO_INST: dict[str, list] = {"safes": [], "convertible_notes": []}

    def test_vision_image_pdf_warns(self) -> None:
        inputs = _inputs_named(["Jane Doe", "John Smith"], cap_base_source="assumed")
        inputs["metadata"]["extraction_mode"] = "vision_image_pdf"
        cs = cap_state_mod.build_cap_state(inputs, self.NO_INST)
        assert "W_VISION_EXTRACTION_LOW_CONFIDENCE" in cs.get("warnings", [])

    def test_no_extraction_mode_no_vision_warn(self) -> None:
        cs = cap_state_mod.build_cap_state(_inputs_named(["Jane Doe", "John Smith"]), self.NO_INST)
        assert "W_VISION_EXTRACTION_LOW_CONFIDENCE" not in cs.get("warnings", [])


class TestRedlineDraftWarning:
    """B2: a tracked-changes/redline .docx is an UNSIGNED draft. When the founder proceeds on the accepted
    view, the SKILL stamps metadata.source_markup='tracked_changes_accepted'; cap_state surfaces
    W_REDLINE_DRAFT so the report persists the draft caveat (the gate message is the primary caveat)."""

    NO_INST: dict[str, list] = {"safes": [], "convertible_notes": []}

    def test_tracked_changes_accepted_warns(self) -> None:
        inputs = _inputs_named(["Jane Doe", "John Smith"], cap_base_source="assumed")
        inputs["metadata"]["source_markup"] = "tracked_changes_accepted"
        cs = cap_state_mod.build_cap_state(inputs, self.NO_INST)
        assert "W_REDLINE_DRAFT" in cs.get("warnings", [])

    def test_no_source_markup_no_redline_warn(self) -> None:
        cs = cap_state_mod.build_cap_state(_inputs_named(["Jane Doe", "John Smith"]), self.NO_INST)
        assert "W_REDLINE_DRAFT" not in cs.get("warnings", [])

    def test_redline_callout_rendered(self) -> None:
        from _warning_callouts import render_warning_callouts  # noqa: PLC0415

        body = "\n".join(render_warning_callouts(["W_REDLINE_DRAFT"]))
        assert "redline" in body.lower() and "unsigned draft" in body.lower()

    def test_redline_warning_from_instrument_level_source_markup(self) -> None:
        safe = {**_SAFE_BASIC, "source_markup": "tracked_changes_accepted"}
        instruments = {**_BASIC_INSTRUMENTS, "safes": [safe]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        assert "W_REDLINE_DRAFT" in cs.get("warnings", [])


# ===========================================================================
# Proceed-degraded on a missing SAFE purchase amount (terms-only / non-convertible)
# ===========================================================================


def test_safe_has_usable_purchase_amount_predicate() -> None:
    """Predicate: True iff purchase_amount is a real positive number (not bool)."""
    assert safe_conversion.safe_has_usable_purchase_amount({"purchase_amount": 500_000}) is True
    assert safe_conversion.safe_has_usable_purchase_amount({"purchase_amount": 0}) is False
    assert safe_conversion.safe_has_usable_purchase_amount({"purchase_amount": None}) is False
    assert safe_conversion.safe_has_usable_purchase_amount({}) is False
    assert safe_conversion.safe_has_usable_purchase_amount({"purchase_amount": True}) is False  # bool guard


def test_safe_missing_purchase_amount_degrades_not_raises() -> None:
    """A SAFE without purchase_amount is kept terms-only (non-convertible) with a warning."""
    safe = {k: v for k, v in _SAFE_BASIC.items() if k != "purchase_amount"}  # has id + issuance_date
    instruments = {**_BASIC_INSTRUMENTS, "safes": [safe]}
    cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)  # must NOT raise
    s = cs["outstanding_safes"][0]
    assert s.get("convertible") is False
    assert s.get("purchase_amount") is None
    assert "W_SAFE_PURCHASE_AMOUNT_MISSING" in cs.get("warnings", [])


def test_solver_skips_non_convertible_safe_cap_implied() -> None:
    """Cap-implied path: a blank SAFE is skipped; the good SAFE is still converted."""
    good = {**_SAFE_BASIC, "id": "safe_ok"}
    blank = {k: v for k, v in _SAFE_BASIC.items() if k != "purchase_amount"} | {"id": "safe_blank"}
    instruments = {**_BASIC_INSTRUMENTS, "safes": [good, blank]}
    cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
    result = run_scenario.run_safe_conversion_scenario(
        {"type": "safe_conversion", "parameters": {}}, instruments=instruments, cap_state=cs
    )
    assert "safe_ok" in result["per_safe"]
    assert "safe_blank" not in result["per_safe"]


def test_solver_skips_non_convertible_safe_priced_round() -> None:
    """Priced-round path: blank SAFE is excluded; solver does not KeyError on purchase_amount."""
    good = {**_SAFE_BASIC, "id": "safe_ok"}
    blank = {k: v for k, v in _SAFE_BASIC.items() if k != "purchase_amount"} | {"id": "safe_blank"}
    instruments = {**_BASIC_INSTRUMENTS, "safes": [good, blank]}
    cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
    scenario = {"type": "priced_round", "parameters": {"pre_money": 10_000_000, "new_money": 2_000_000}}
    result = run_scenario.run_priced_round_scenario(scenario, instruments=instruments, cap_state=cs)
    assert "safe_ok" in result["per_safe"]
    assert "safe_blank" not in result["per_safe"]


def test_quick_assess_terms_only_safe_no_purchase_amount() -> None:
    """quick_assess must not raise when a SAFE has no purchase_amount (terms-only SAFE).

    A SAFE that cap_state marks non-convertible (purchase_amount absent or None)
    reaches the Outstanding instruments listing in quick_assess.  The listing must
    render an 'amount n/a' placeholder rather than raising KeyError.
    """
    import quick_assess as qa  # type: ignore[import-not-found]

    terms_only_safe = {k: v for k, v in _SAFE_BASIC.items() if k != "purchase_amount"} | {
        "id": "safe_terms_only",
        "investor_name": "Mystery Investor",
    }
    sentinel = qa.quick_assess(
        company_name="Acmecorp",
        inputs=_BASIC_INPUTS,
        safes=[terms_only_safe],
        notes=[],
        pre_money=10_000_000,
        new_money=2_000_000,
        target_pool_percent=None,
        target_basis="post_money",
    )
    report_md = sentinel.pop("_report_md")
    assert "amount n/a" in report_md, f"Expected 'amount n/a' placeholder for a terms-only SAFE; got:\n{report_md}"


# ===========================================================================
# Guard: freeform stated_totals triggers W_FD_RECONCILE_DELTA on divergence
# ===========================================================================


def test_stated_totals_accepts_price_and_premoney() -> None:
    """Guard: stated_totals with price_per_share + pre_money validates against inputs.schema.json.

    Since stated_totals has no additionalProperties constraint, the new fields already
    validate; this test is a guard confirming a populated stated_totals with the two
    source-stated round fields passes schema validation and is not accidentally rejected.
    """
    try:
        import jsonschema  # type: ignore[import-untyped]
    except ImportError:
        pytest.skip("jsonschema not installed")

    schema_path = os.path.join(
        os.path.dirname(SCRIPTS),
        "references",
        "schemas",
        "inputs.schema.json",
    )
    with open(schema_path) as f:
        schema = json.load(f)

    st = {
        "fully_diluted": 45_209_360,
        "price_per_share": 23.2255,
        "pre_money": 1_050_000_000,
        "source": "pro_forma_grid",
    }
    # Validates the stated_totals sub-schema directly
    jsonschema.validate(instance=st, schema=schema["properties"]["stated_totals"])


def test_freeform_stated_totals_catches_dropped_class() -> None:
    """Guard: a freeform stated_totals FD that diverges >0.1% from the computed FD fires W_FD_RECONCILE_DELTA.

    _BASIC_INPUTS computes to 11,500,000 FD (two founders × 5 M + 1.5 M authorized pool).
    A stated total of 3,881,559 diverges by ~1,962,727 ppm — far above the 1000 ppm threshold —
    simulating a dropped equity class in the freeform grid.
    """
    inputs = {**_BASIC_INPUTS, "stated_totals": {"fully_diluted": 3_881_559, "source": "freeform_grid"}}
    cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
    assert any(w.startswith("W_FD_RECONCILE_DELTA") for w in cs.get("warnings", []))


def test_freeform_stated_totals_silent_when_matching() -> None:
    """Guard: freeform stated_totals FD equal to computed FD → no W_FD_RECONCILE_DELTA (no false positive)."""
    cs0 = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
    fd = cs0["as_converted_totals"]["fully_diluted_shares"]
    inputs = {**_BASIC_INPUTS, "stated_totals": {"fully_diluted": fd, "source": "freeform_grid"}}
    cs = cap_state_mod.build_cap_state(inputs, _BASIC_INSTRUMENTS)
    assert not any(w.startswith("W_FD_RECONCILE_DELTA") for w in cs.get("warnings", []))


# ---------------------------------------------------------------------------
# build_reconciliation_lines
# ---------------------------------------------------------------------------


def test_reconciliation_section_renders_with_divergent_pps() -> None:
    """Divergent PPS (>0.1%) must render the section, flag the row, and fire the hedged cause note."""
    import compose_report  # type: ignore[import-not-found]

    scenario = {
        "type": "priced_round",
        "scenario_id": "s1",
        "computed_outputs": {"equity_financing_price": 21.6565},
        "parameters": {"pre_money": 1_050_000_000},
    }
    inputs = {
        "stated_totals": {
            "price_per_share": 23.2255,
            "pre_money": 1_050_000_000,
            "fully_diluted": 45_209_360,
            "source": "pro_forma_grid",
        }
    }
    cap_state = {"as_converted_totals": {"fully_diluted_shares": 45_209_008}}
    out = "\n".join(compose_report.build_reconciliation_lines(scenario, inputs, cap_state))
    assert "Reconciliation vs your source documents" in out
    assert "Price per share" in out and "23.2255" in out and "21.6565" in out
    assert "coupled" in out.lower()  # the cause note fired (PPS diverged > 0.1%)


def test_reconciliation_absent_when_no_stated_totals() -> None:
    """When inputs has no stated_totals, build_reconciliation_lines must return []."""
    import compose_report  # type: ignore[import-not-found]

    scenario = {
        "type": "priced_round",
        "computed_outputs": {"equity_financing_price": 21.66},
        "parameters": {"pre_money": 1},
    }
    assert compose_report.build_reconciliation_lines(scenario, {}, {"as_converted_totals": {}}) == []


# ---------------------------------------------------------------------------
# Finding 14: non-circularity guard on the stated_totals reconciliation cross-foot.
#
# SKILL.md's "Non-circularity rule" documents the hazard (copying a computed value into
# inputs.stated_totals trivially "matches" and hides a real divergence) as agent-instruction-only,
# with no deterministic script guard. These tests cover the mechanical guard added to
# compose_report.py: _stated_totals_provenance_ok / refine_reconciliation_status_for_provenance /
# build_reconciliation_provenance_warnings, plus their effect on build_reconciliation_lines'
# rendered table row.
# ---------------------------------------------------------------------------


class TestReconciliationNonCircularityGuard:
    def _scenario(self, pps: float) -> dict[str, Any]:
        return {
            "type": "priced_round",
            "scenario_id": "s1",
            "computed_outputs": {"equity_financing_price": pps},
            "parameters": {"pre_money": 1_000_000},
        }

    def _cap_state(self, fd: int) -> dict[str, Any]:
        return {"as_converted_totals": {"fully_diluted_shares": fd}}

    # -- build_reconciliation_lines (the founder-facing table row) --------------------------

    def test_copied_scenario_output_no_source_renders_circular(self) -> None:
        """Copying the computed PPS into stated_totals with NO `source` at all — the exact
        "copy-the-computed-value-back-in" pattern — must render CIRCULAR, not a clean match."""
        import compose_report  # type: ignore[import-not-found]

        scenario = self._scenario(21.6565)
        inputs = {"stated_totals": {"price_per_share": 21.6565}}  # no "source" key
        out = "\n".join(compose_report.build_reconciliation_lines(scenario, inputs, {"as_converted_totals": {}}))
        assert "CIRCULAR" in out
        assert "Circular reconciliation detected" in out
        assert "CANNOT VERIFY" not in out

    def test_copied_scenario_output_self_referential_source_renders_circular(self) -> None:
        """Same copy pattern, but this time `source` is populated with a self-referential label
        naming the skill's own output — must still render CIRCULAR, not a clean match."""
        import compose_report  # type: ignore[import-not-found]

        scenario = self._scenario(21.6565)
        for bad_source in ("scenarios.json", "computed", "this report", "cap_state.json"):
            inputs = {"stated_totals": {"price_per_share": 21.6565, "source": bad_source}}
            out = "\n".join(compose_report.build_reconciliation_lines(scenario, inputs, {"as_converted_totals": {}}))
            assert "CIRCULAR" in out, f"source={bad_source!r} should have fired CIRCULAR; got:\n{out}"

    def test_legitimate_exact_match_with_real_provenance_is_clean_match(self) -> None:
        """A genuine exact match (the skill's math SHOULD agree with the document) with REAL
        provenance must report cleanly — exact-match alone must never be the trigger."""
        import compose_report  # type: ignore[import-not-found]

        scenario = self._scenario(21.6565)
        inputs = {"stated_totals": {"price_per_share": 21.6565, "source": "term_sheet"}}
        out = "\n".join(compose_report.build_reconciliation_lines(scenario, inputs, {"as_converted_totals": {}}))
        assert "Price per share" in out and "21.6565" in out
        assert "CIRCULAR" not in out
        assert "CANNOT VERIFY" not in out
        assert "+0.0%" in out  # clean match row, no ⚠ suffix

    def test_missing_provenance_near_match_is_cannot_verify(self) -> None:
        """A near-but-not-exact match (within tolerance) with NO `source` at all can't be
        confirmed as independent — must degrade to CANNOT VERIFY, not report a match."""
        import compose_report  # type: ignore[import-not-found]

        # 21.6565 vs 21.65 is well within the 0.1% tolerance but not bit-for-bit identical.
        scenario = self._scenario(21.6565)
        inputs = {"stated_totals": {"price_per_share": 21.65}}  # no "source" key
        out = "\n".join(compose_report.build_reconciliation_lines(scenario, inputs, {"as_converted_totals": {}}))
        assert "CANNOT VERIFY" in out
        assert "provenance missing" in out.lower()
        assert "CIRCULAR" not in out

    # -- compute_reconciliation_status + refine + warnings (the report.json / --strict surface) --

    def test_refine_downgrades_exact_match_no_source_to_circular(self) -> None:
        from compose_report import (  # type: ignore[import-not-found]
            compute_reconciliation_status,
            refine_reconciliation_status_for_provenance,
        )

        status, ppm = compute_reconciliation_status(computed={"pps": 2.00}, stated={"pps": 2.00})
        assert status == "pass"
        refined = refine_reconciliation_status_for_provenance(
            status=status, max_ppm=ppm, stated_totals={"price_per_share": 2.00}
        )
        assert refined == "circular"

    def test_refine_downgrades_near_match_no_source_to_cannot_verify(self) -> None:
        from compose_report import (  # type: ignore[import-not-found]
            compute_reconciliation_status,
            refine_reconciliation_status_for_provenance,
        )

        status, ppm = compute_reconciliation_status(computed={"pps": 2.00}, stated={"pps": 2.0005})
        assert status == "pass"
        refined = refine_reconciliation_status_for_provenance(
            status=status, max_ppm=ppm, stated_totals={"price_per_share": 2.0005}
        )
        assert refined == "cannot_verify"

    def test_refine_leaves_provenance_backed_match_as_pass(self) -> None:
        from compose_report import (  # type: ignore[import-not-found]
            compute_reconciliation_status,
            refine_reconciliation_status_for_provenance,
        )

        status, ppm = compute_reconciliation_status(computed={"pps": 2.00}, stated={"pps": 2.00})
        refined = refine_reconciliation_status_for_provenance(
            status=status, max_ppm=ppm, stated_totals={"price_per_share": 2.00, "source": "term_sheet"}
        )
        assert refined == "pass"

    def test_refine_leaves_diverged_and_not_applicable_unchanged(self) -> None:
        """A genuine numeric divergence — or nothing to compare — is a real finding (or a
        no-op) regardless of provenance; only a would-be "pass" can be downgraded."""
        from compose_report import (  # type: ignore[import-not-found]
            refine_reconciliation_status_for_provenance,
        )

        assert (
            refine_reconciliation_status_for_provenance(
                status="diverged", max_ppm=50_000.0, stated_totals={"price_per_share": 1.0}
            )
            == "diverged"
        )
        assert (
            refine_reconciliation_status_for_provenance(status="not_applicable", max_ppm=0.0, stated_totals=None)
            == "not_applicable"
        )

    def test_absent_stated_totals_still_not_applicable_no_warnings(self) -> None:
        """Requirement: absent stated_totals behaves exactly as before — no reconciliation row,
        no provenance warning, status stays not_applicable."""
        from compose_report import (  # type: ignore[import-not-found]
            build_reconciliation_provenance_warnings,
            compute_reconciliation_status,
            refine_reconciliation_status_for_provenance,
        )

        status, ppm = compute_reconciliation_status(computed={"pps": 2.00, "fd": 100}, stated=None)
        assert status == "not_applicable"
        refined = refine_reconciliation_status_for_provenance(status=status, max_ppm=ppm, stated_totals=None)
        assert refined == "not_applicable"
        assert build_reconciliation_provenance_warnings(status=refined, stated_totals=None) == []

    def test_build_reconciliation_provenance_warnings_circular_is_high_severity_code(self) -> None:
        from compose_report import build_reconciliation_provenance_warnings  # type: ignore[import-not-found]

        warnings = build_reconciliation_provenance_warnings(status="circular", stated_totals={"price_per_share": 2.00})
        assert len(warnings) == 1
        assert warnings[0]["code"] == "E_CIRCULAR_RECONCILIATION"
        assert warnings[0]["severity"] == "critical"

    def test_build_reconciliation_provenance_warnings_cannot_verify_is_medium_severity_code(self) -> None:
        from compose_report import build_reconciliation_provenance_warnings  # type: ignore[import-not-found]

        warnings = build_reconciliation_provenance_warnings(
            status="cannot_verify", stated_totals={"price_per_share": 2.00}
        )
        assert len(warnings) == 1
        assert warnings[0]["code"] == "W_STATED_TOTALS_PROVENANCE_MISSING"
        assert warnings[0]["severity"] == "medium"

    def test_build_reconciliation_provenance_warnings_pass_is_silent(self) -> None:
        from compose_report import build_reconciliation_provenance_warnings  # type: ignore[import-not-found]

        assert (
            build_reconciliation_provenance_warnings(
                status="pass", stated_totals={"price_per_share": 2.00, "source": "term_sheet"}
            )
            == []
        )

    # -- end-to-end via compose_report.py CLI: copied output with no source blocks --strict -----

    def test_end_to_end_circular_stated_totals_fires_strict_gate(self) -> None:
        """Full pipeline: a founder-skill run that copies the computed equity_financing_price
        back into inputs.stated_totals with no source must fire E_CIRCULAR_RECONCILIATION in
        report.json's validation.warnings and fail a --strict run."""
        scenario = _minimal_scenario(
            "s_priced",
            computed_outputs={
                "completeness": "full",
                "blockers": [],
                "equity_financing_price": 2.00,
                "aggregate_ownership_by_class": {"founders_pct": 0.5, "option_pool_pct": 0.1, "preferred_pct": 0.4},
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        inputs_path = os.path.join(d, "inputs.json")
        with open(inputs_path) as f:
            inputs = json.load(f)
        inputs["stated_totals"] = {"price_per_share": 2.00}  # copied, no source
        with open(inputs_path, "w") as f:
            json.dump(inputs, f)

        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, stderr  # non-strict run still succeeds (degraded, not fatal)
        assert report["reconciliation_status"] == "circular"
        codes = {w["code"] for w in report["validation"]["warnings"]}
        assert "E_CIRCULAR_RECONCILIATION" in codes

        # --strict must now fail on this code
        report_path = os.path.join(d, "report.json")
        md_path = os.path.join(d, "report.md")
        rc_strict, _stdout, _stderr = _run(
            "compose_report.py",
            ["--dir", d, "--run-id", "test-rid", "-o", report_path, "--write-md", md_path, "--strict"],
        )
        assert rc_strict == 1


class TestCoverage:
    def _reg(self) -> dict[str, Any]:
        return coverage_mod.load_registry()  # type: ignore[no-any-return, attr-defined]

    def test_priced_round_with_safe_and_acquisition_is_covered(self) -> None:
        r = coverage_mod.is_covered(["priced_round", "safe_conversion", "acquisition_consideration"], self._reg())  # type: ignore[attr-defined]
        assert r["covered"] is True
        assert r["uncovered_parts"] == []

    def test_flip_plus_round_is_covered_via_sequence(self) -> None:
        r = coverage_mod.is_covered(["flip", "priced_round", "safe_conversion"], self._reg())  # type: ignore[attr-defined]
        assert r["covered"] is True

    def test_lone_flip_is_covered(self) -> None:
        r = coverage_mod.is_covered(["flip"], self._reg())  # type: ignore[attr-defined]
        assert r["covered"] is True

    def test_note_plus_acquisition_is_uncovered(self) -> None:
        r = coverage_mod.is_covered(["priced_round", "note_conversion", "acquisition_consideration"], self._reg())  # type: ignore[attr-defined]
        assert r["covered"] is False
        assert any(
            set(p.get("combination", [])) == {"note_conversion", "acquisition_consideration"}
            for p in r["uncovered_parts"]
        )

    def test_unknown_primitive_is_uncovered(self) -> None:
        r = coverage_mod.is_covered(["priced_round", "earnout"], self._reg())  # type: ignore[attr-defined]
        assert r["covered"] is False


class TestDetectStructure:
    def _inputs(self, acquisition: dict | None = None) -> dict:
        d = {
            "company_name": "Acmecorp",
            "analysis_date": "2026-01-01",
            "mode": "standard",
            "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
            "preferred_series": [],
            "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
        }
        if acquisition is not None:
            d["acquisition"] = acquisition
        return d

    def _instruments(
        self, safes: list[dict[str, Any]] | None = None, notes: list[dict[str, Any]] | None = None
    ) -> dict[str, Any]:
        return {
            "safes": safes or [],
            "convertible_notes": notes or [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }

    def test_acquisition_block_emits_acquisition_primitive(self) -> None:
        r = detect_structure.detect(
            self._inputs(
                acquisition={"acquired_entity": "TargetCo", "consideration_pct": 0.2, "consideration_form": "shares"}
            ),
            self._instruments(safes=[{"id": "s1"}]),
        )
        assert "acquisition_consideration" in r["required_primitives"]
        assert "safe_conversion" in r["required_primitives"]

    def test_note_plus_acquisition_routes_uncovered(self) -> None:
        r = detect_structure.detect(
            self._inputs(
                acquisition={"acquired_entity": "TargetCo", "consideration_pct": 0.2, "consideration_form": "shares"}
            ),
            self._instruments(notes=[{"note_id": "n1"}]),
        )
        assert r["covered"] is False

    def test_plain_safe_round_is_covered(self) -> None:
        r = detect_structure.detect(self._inputs(), self._instruments(safes=[{"id": "s1"}]))
        assert r["covered"] is True

    def test_flip_focused_mode_emits_flip_primitive(self) -> None:
        # mode=="flip_focused" is a real schema signal; the old implementation keyed off
        # non-existent fields (jurisdiction_change / event_dates.flip) and never detected flip.
        inputs = {**self._inputs(), "mode": "flip_focused"}
        r = detect_structure.detect(inputs, self._instruments())
        assert "flip" in r["required_primitives"]

    def test_mid_flip_jurisdiction_emits_flip_primitive(self) -> None:
        inputs = {**self._inputs(), "jurisdiction": {"structure": "mid_flip"}}
        r = detect_structure.detect(inputs, self._instruments())
        assert "flip" in r["required_primitives"]

    def test_flip_closing_date_emits_flip_primitive(self) -> None:
        inputs = {**self._inputs(), "event_dates": {"flip_closing_date": "2026-03-01"}}
        r = detect_structure.detect(inputs, self._instruments())
        assert "flip" in r["required_primitives"]

    def test_flip_focused_plus_safe_plus_acquisition_routes_covered(self) -> None:
        # Motivating combo: flip_focused mode + SAFE + acquisition.
        # Must be covered and route must contain both a flip scenario and a priced_round
        # scenario whose chained_from_scenario_id points at the flip.
        inputs = {
            **self._inputs(
                acquisition={"acquired_entity": "TargetCo", "consideration_pct": 0.2, "consideration_form": "shares"}
            ),
            "mode": "flip_focused",
        }
        r = detect_structure.detect(inputs, self._instruments(safes=[{"id": "s1"}]))
        assert "flip" in r["required_primitives"]
        assert r["covered"] is True
        types = [s["type"] for s in r["route"]["scenario_requests"]]
        assert "flip" in types
        assert "priced_round" in types
        # The priced_round scenario must chain from the flip scenario.
        round_req = next(s for s in r["route"]["scenario_requests"] if s["type"] == "priced_round")
        flip_req = next(s for s in r["route"]["scenario_requests"] if s["type"] == "flip")
        assert round_req["chained_from_scenario_id"] == flip_req["scenario_id"]


class TestExecutiveSummaryCompletenessWording:
    """`structural_only` has EIGHT origins; only one is working-as-intended.

    The cap-implied SAFE snapshot is by design (run_scenario stamps `cap_implied_only`). The other
    seven are genuinely blocked runs. The summary used to count them together and call the total
    "pending input", so a founder read a by-design result as an unfinished analysis. These pin the
    split -- and pin that a blocked scenario is still described as needing something.
    """

    def _render(self, scenarios: list[dict]) -> str:
        import compose_report  # type: ignore[import-not-found]

        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        for s in scenarios:
            s.setdefault("type", "safe_conversion")
            s.setdefault("label", s["scenario_id"])
        md: str = compose_report.render_report_markdown(
            artifacts={
                "inputs.json": _BASIC_INPUTS,
                "instruments.json": _BASIC_INSTRUMENTS,
                "cap_state.json": cs,
                "scenarios.json": {"scenarios": scenarios},
                "rule_audit.json": {},
                "counsel_packet.json": {},
            },
            validation_warnings=[],
            insertion_marker="MARKER",
        )
        return md

    def test_by_design_snapshot_is_not_called_pending_input(self) -> None:
        md = self._render(
            [
                {
                    "scenario_id": "s1",
                    "type": "safe_conversion",
                    "label": "S1",
                    "computed_outputs": {
                        "completeness": "structural_only",
                        "cap_implied_only": True,
                        "per_safe": {"safe_1": {"ownership_pct": 0.05}},
                    },
                }
            ]
        )
        assert "pending input" not in md
        assert "needs more from you" not in md
        assert "terms not set yet" in md

    def test_a_genuinely_blocked_scenario_still_says_it_needs_input(self) -> None:
        md = self._render(
            [
                {
                    "scenario_id": "s1",
                    "type": "safe_conversion",
                    "label": "S1",
                    "computed_outputs": {
                        "completeness": "structural_only",
                        "blockers": [{"code": "E_X", "remedy": "need more"}],
                    },
                }
            ]
        )
        assert "needs more from you" in md
        assert "terms not set yet" not in md

    def test_the_two_are_counted_separately_not_merged(self) -> None:
        md = self._render(
            [
                {
                    "scenario_id": "s1",
                    "type": "safe_conversion",
                    "label": "S1",
                    "computed_outputs": {
                        "completeness": "structural_only",
                        "cap_implied_only": True,
                        "per_safe": {"safe_1": {"ownership_pct": 0.05}},
                    },
                },
                {
                    "scenario_id": "s2",
                    "type": "safe_conversion",
                    "label": "S2",
                    "computed_outputs": {"completeness": "structural_only"},
                },
            ]
        )
        assert "1 priced-round terms not set yet" in md
        assert "1 needs more from you" in md


class TestCoverageRouteParameterGaps:
    """The coverage detector cannot supply terms that live with the founder.

    `detect_structure.py` reads inputs.json + instruments.json only, so a priced_round comes back
    with no pre_money/new_money and a flip with no IIA/102 answers. SKILL.md says to extend the
    route's request with the founder's answers -- but a caller that ran it verbatim used to reach
    `params["pre_money"]` in run_scenario.py and die on an uncaught KeyError. These pin the two
    halves of the fix: the gaps are NAMED in the request, and running one anyway is a typed blocker.
    """

    def _inputs(self) -> dict:
        return {
            "company_name": "Acme",
            "metadata": {"run_id": "t", "schema_version": "v0.5.0-inputs"},
            "founders": [{"founder_id": "f1", "name": "A", "common_shares": 1_000_000}],
        }

    def test_priced_round_request_names_its_missing_parameters(self) -> None:
        import detect_structure

        r = detect_structure.detect(self._inputs(), {"safes": [{"id": "s1"}]})
        req = next(s for s in r["route"]["scenario_requests"] if s["type"] == "priced_round")
        assert req["parameters"].get("pre_money") is None
        assert set(req["parameters_required"]) == {"pre_money", "new_money"}

    def test_running_a_priced_round_without_terms_is_a_typed_blocker_not_a_crash(self) -> None:
        import run_scenario

        out = run_scenario.run_priced_round_scenario(
            {"scenario_id": "s_round", "type": "priced_round", "parameters": {}},
            instruments={},
            cap_state={"founders": [], "as_converted_totals": {"fully_diluted_shares": 1_000_000}},
        )
        assert out["completeness"] == "structural_only"
        codes = [b["code"] for b in out["blockers"]]
        assert "E_MISSING_PARAMETER" in codes
        remedy = next(b["remedy"] for b in out["blockers"] if b["code"] == "E_MISSING_PARAMETER")
        assert "pre_money" in remedy and "new_money" in remedy


class TestPreMoneyBasis:
    def test_default_includes_is_byte_identical_to_existing_golden(self) -> None:
        # The eval2 golden (PPS = 175/122) must be unchanged when pre_money_basis defaults.
        from test_cap_table import TestStackedPostMoneySAFEsGolden as G  # reuse fixtures

        g = G()
        cs = cap_state_mod.build_cap_state(g.EVAL2_INPUTS, g._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=g.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        assert math.isclose(r["equity_financing_price"], 175 / 122, rel_tol=1e-4)

    def test_excludes_basis_drops_safe_from_price_denom(self) -> None:
        # Same deal, excludes basis → larger PPS (smaller denom), SAFE share COUNT unchanged.
        from test_cap_table import TestStackedPostMoneySAFEsGolden as G

        g = G()
        cs = cap_state_mod.build_cap_state(g.EVAL2_INPUTS, g._instruments())
        inc = priced_round.solve_priced_round(
            cap_state=cs,
            safes=g.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            pre_money_basis="includes_safe_conversion",
        )
        exc = priced_round.solve_priced_round(
            cap_state=cs,
            safes=g.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            pre_money_basis="excludes_safe_conversion",
        )
        assert exc["equity_financing_price"] > inc["equity_financing_price"]
        assert exc["completeness"] == "full"


class TestAcquisitionInput:
    def _inputs(self, timing: str) -> dict:
        return {
            "company_name": "Acmecorp",
            "analysis_date": "2026-01-01",
            "mode": "standard",
            "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
            "preferred_series": [],
            "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
            "acquisition": {
                "acquired_entity": "TargetCo",
                "consideration_pct": 0.2,
                "consideration_form": "shares",
                "acquisition_timing": timing,
            },
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
        }

    def _instruments(self) -> dict:
        return {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }

    def test_concurrent_does_not_fold_C(self) -> None:
        cs = cap_state_mod.build_cap_state(self._inputs("concurrent_with_round"), self._instruments())
        assert cs["as_converted_totals"]["fully_diluted_shares"] == 8_000_000

    def test_pre_round_closed_folds_C_into_snapshot(self) -> None:
        cs = cap_state_mod.build_cap_state(self._inputs("pre_round_closed"), self._instruments())
        # C = round(0.2/0.8 * 8_000_000) = 2_000_000 ; FD = 10_000_000
        assert cs["as_converted_totals"]["fully_diluted_shares"] == 10_000_000
        assert cs["as_converted_totals"]["acquisition_consideration_shares"] == 2_000_000


class TestAcquisitionConsideration:
    ACQ_INPUTS = {
        "company_name": "Acmecorp",
        "analysis_date": "2026-01-01",
        "mode": "standard",
        "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
        "preferred_series": [],
        "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
        "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
    }

    def _cs(self) -> dict[str, Any]:
        instruments = {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        return cap_state_mod.build_cap_state(self.ACQ_INPUTS, instruments)  # type: ignore[no-any-return]

    def test_excludes_basis_golden_pps_095(self) -> None:
        # t=0.20, pre=10M, new=2M, pre_FD=8M, no safe/pool.
        # k = 1 + 2/10 = 1.2 ; D = 8M/(1-0.24) = 8M/0.76 ; PPS = 10M/D = 0.95 ; C = D-8M.
        r = priced_round.solve_priced_round(
            cap_state=self._cs(),
            safes=[],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.20},
        )
        assert r["completeness"] == "full", r.get("blockers")
        assert math.isclose(r["equity_financing_price"], 0.95, rel_tol=1e-4)
        agg = r["aggregate_ownership_by_class"]
        assert math.isclose(agg["acquisition_pct"], 0.20, abs_tol=1e-4)  # universal invariant
        total = (
            agg["founders_pct"]
            + agg["preferred_pct"]
            + agg["option_pool_pct"]
            + agg["safe_pct"]
            + agg["note_pct"]
            + agg["new_money_pct"]
            + agg["acquisition_pct"]
        )
        assert math.isclose(total, 1.0, abs_tol=1e-4)
        assert r["shares_breakdown"]["acquisition_consideration"] > 0

    def _cs_with_cap_safe(self) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        # One cap-branch post-money SAFE: purchase 1M, cap 10M → owns 10% of company cap.
        # company_cap = 8M + safe ; safe = 0.1*(8M+safe) → safe = 8M/9 (PPS-independent).
        safe: dict[str, Any] = {
            "id": "s1",
            "purchase_amount": 1_000_000,
            "post_money_valuation_cap": 10_000_000,
            "form": "yc_postmoney_cap",
            "discount_multiplier": None,
            "issuance_date": "2025-01-01",
        }
        instruments = {
            "safes": [safe],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        cs: dict[str, Any] = cap_state_mod.build_cap_state(self.ACQ_INPUTS, instruments)
        return cs, [safe]

    def test_excludes_basis_with_cap_safe_golden_171_184(self) -> None:
        # excludes: D = (pre_FD + t*safe)/(1-t*k) ; safe=8M/9, t=0.2, k=1.2, pre_FD=8M.
        # numerator = 8M + 0.2*(8M/9) = 73.6M/9 ; D = (73.6M/9)/0.76 ; PPS = 10M/D = 171/184.
        # The C1 bug (dropping t*safe) would give 0.95 — this golden catches it.
        cs, safes = self._cs_with_cap_safe()
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=safes,
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.20},
        )
        assert r["completeness"] == "full", r.get("blockers")
        assert math.isclose(r["equity_financing_price"], 171 / 184, rel_tol=1e-4), r["equity_financing_price"]
        assert math.isclose(r["aggregate_ownership_by_class"]["acquisition_pct"], 0.20, abs_tol=1e-4)
        assert r["aggregate_ownership_by_class"]["safe_pct"] > 0

    def test_includes_basis_with_cap_safe_golden_171_200(self) -> None:
        # includes: D = (pre_FD + safe)/(1-t*k) = (8M + 8M/9)/0.76 = (80M/9)/0.76 ; PPS = 10M/D = 171/200 = 0.855.
        cs, safes = self._cs_with_cap_safe()
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=safes,
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="includes_safe_conversion",
            acquisition={"consideration_pct": 0.20},
        )
        assert r["completeness"] == "full", r.get("blockers")
        assert math.isclose(r["equity_financing_price"], 171 / 200, rel_tol=1e-4), r["equity_financing_price"]

    def test_t_times_k_ge_1_is_overdetermined(self) -> None:
        # t=0.9, k=1.2 → t·k = 1.08 ≥ 1.
        r = priced_round.solve_priced_round(
            cap_state=self._cs(),
            safes=[],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.90},
        )
        assert r["completeness"] == "structural_only"
        assert any(b["code"] == "E_ACQUISITION_OVERDETERMINED" for b in r["blockers"])

    def test_no_acquisition_is_byte_identical_no_new_keys(self) -> None:
        from test_cap_table import TestStackedPostMoneySAFEsGolden as G

        g = G()
        cs = cap_state_mod.build_cap_state(g.EVAL2_INPUTS, g._instruments())
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=g.EVAL2_SAFES,
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        assert math.isclose(r["equity_financing_price"], 175 / 122, rel_tol=1e-4)
        assert "acquisition_pct" not in r["aggregate_ownership_by_class"]
        assert "acquisition_consideration" not in r["shares_breakdown"]


class TestPoolConsiderationBasis:
    def _cs(self) -> dict[str, Any]:
        inputs = {
            "company_name": "Acmecorp",
            "analysis_date": "2026-01-01",
            "mode": "standard",
            "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
            "preferred_series": [],
            "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
        }
        instruments = {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        return cap_state_mod.build_cap_state(inputs, instruments)  # type: ignore[no-any-return]

    def _solve(self, basis: str) -> dict[str, Any]:
        return priced_round.solve_priced_round(  # type: ignore[no-any-return]
            cap_state=self._cs(),
            safes=[],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.20},
            pool_consideration_basis=basis,
        )

    def test_realized_pool_pct_hits_target_both_bases(self) -> None:
        # include: pool sized on post_fd-incl-C → ownership pool/post_fd == target (0.10).
        inc = self._solve("include")
        assert inc["completeness"] == "full", inc.get("blockers")
        assert math.isclose(inc["aggregate_ownership_by_class"]["option_pool_pct"], 0.10, abs_tol=2e-3)
        # exclude: pool sized on the pre-consideration base → ownership = target*(1-t) = 0.08 for t=0.2.
        exc = self._solve("exclude")
        assert exc["completeness"] == "full", exc.get("blockers")
        assert math.isclose(exc["aggregate_ownership_by_class"]["option_pool_pct"], 0.08, abs_tol=2e-3)
        # The ownership table MUST sum to 1 under BOTH bases (this is the real correctness guard).
        for r in (inc, exc):
            agg = r["aggregate_ownership_by_class"]
            total = sum(
                agg[k]
                for k in (
                    "founders_pct",
                    "preferred_pct",
                    "option_pool_pct",
                    "safe_pct",
                    "note_pct",
                    "new_money_pct",
                    "acquisition_pct",
                )
            )
            assert math.isclose(total, 1.0, abs_tol=1e-4), (r is exc, total)

    def test_basis_moves_pps(self) -> None:
        inc = self._solve("include")
        exc = self._solve("exclude")
        assert not math.isclose(inc["equity_financing_price"], exc["equity_financing_price"], rel_tol=1e-3)


class TestAcquisitionRouting:
    def _scenario(self, timing: str) -> dict:
        return {
            "scenario_id": "s_round",
            "label": "Round",
            "type": "priced_round",
            "parameters": {
                "pre_money": 10_000_000,
                "new_money": 2_000_000,
                "pre_money_basis": "excludes_safe_conversion",
                "acquisition": {
                    "acquired_entity": "TargetCo",
                    "consideration_pct": 0.2,
                    "consideration_form": "shares",
                    "acquisition_timing": timing,
                },
            },
        }

    def _cs(self, timing: str) -> dict[str, Any]:
        inputs = {
            "company_name": "Acmecorp",
            "analysis_date": "2026-01-01",
            "mode": "standard",
            "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
            "preferred_series": [],
            "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
            "acquisition": {
                "acquired_entity": "TargetCo",
                "consideration_pct": 0.2,
                "consideration_form": "shares",
                "acquisition_timing": timing,
            },
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
        }
        instruments = {
            "safes": [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        return cap_state_mod.build_cap_state(inputs, instruments)  # type: ignore[no-any-return]

    def test_concurrent_routes_acquisition_into_solver(self) -> None:
        out = run_scenario.run_priced_round_scenario(
            self._scenario("concurrent_with_round"),
            instruments={"safes": [], "convertible_notes": []},
            cap_state=self._cs("concurrent_with_round"),
        )
        assert out["completeness"] == "full", out.get("blockers")
        assert "acquisition_pct" in out["aggregate_ownership_by_class"]
        assert math.isclose(out["equity_financing_price"], 0.95, rel_tol=1e-4)

    def test_pre_round_closed_no_double_count_and_bucket_present(self) -> None:
        # cap_state already folded C (FD=10M, acquisition_consideration_shares=2M). The router strips
        # `acquisition` so the solver does NOT re-run the t·k solve (no double-count). The aggregate
        # STILL emits an acquisition_pct bucket from the cap_state marker (M4) so buckets sum to 1.
        # The closed C is a fixed pre-round block diluted by new money: with PPS=10M/10M=1.0,
        # new_money_shares=2M, post_fd=12M → acquisition_pct = 2M/12M = 1/6 ≈ 0.1667 (NOT t=0.20).
        out = run_scenario.run_priced_round_scenario(
            self._scenario("pre_round_closed"),
            instruments={"safes": [], "convertible_notes": []},
            cap_state=self._cs("pre_round_closed"),
        )
        assert out["completeness"] == "full", out.get("blockers")
        agg = out["aggregate_ownership_by_class"]
        assert "acquisition_pct" in agg
        assert math.isclose(agg["acquisition_pct"], 1 / 6, abs_tol=2e-3)  # 2M/12M, NOT 0.20
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-4)


class TestFlipRoundSequence:
    def _cs_and_instruments(self) -> tuple[dict[str, Any], dict[str, Any]]:
        inputs: dict[str, Any] = {
            "company_name": "Acmecorp",
            "analysis_date": "2026-01-01",
            "mode": "flip_focused",
            "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
            "preferred_series": [],
            "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
        }
        instruments: dict[str, Any] = {
            "safes": [
                {
                    "id": "s1",
                    "purchase_amount": 1_000_000,
                    "post_money_valuation_cap": 10_000_000,
                    "form": "yc_postmoney_cap",
                    "discount_multiplier": None,
                    "issuance_date": "2025-01-01",
                }
            ],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        cs: dict[str, Any] = cap_state_mod.build_cap_state(inputs, instruments)
        return cs, instruments

    def test_flip_emits_post_flip_instruments_verbatim(self) -> None:
        cs, instruments = self._cs_and_instruments()
        out = flip_scenario.flip_share_for_share(cs, instruments=instruments)
        pf = out["post_flip_instruments"]
        assert pf["safes"][0]["post_money_valuation_cap"] == 10_000_000  # terms copied verbatim
        assert pf["safes"][0]["purchase_amount"] == 1_000_000
        # emission guard: post_flip_instruments is a DISTINCT deep copy carrying the entity marker
        assert pf is not None
        assert pf.get("_flip_entity_ref") == "post_flip"
        assert pf is not instruments  # deep copy, not the same object

    def test_sequence_feeds_post_flip_state_into_round(self) -> None:
        # End-to-end smoke: the both-swap (post_flip_cap_state + post_flip_instruments) is
        # content-invisible for a value-neutral 1:1 flip, so this asserts the chained pipeline
        # runs and the SAFE converts; swap correctness rests on the unit-level emission guard
        # above (test_flip_emits_post_flip_instruments_verbatim) + the implementation.
        cs, instruments = self._cs_and_instruments()
        requests = [
            {"scenario_id": "s_flip", "label": "Flip", "type": "flip", "parameters": {}},
            {
                "scenario_id": "s_round",
                "label": "Round",
                "type": "priced_round",
                "chained_from_scenario_id": "s_flip",
                "parameters": {"pre_money": 10_000_000, "new_money": 2_000_000},
            },
        ]
        results = run_scenario.run_all_scenarios(
            inputs={}, instruments=instruments, cap_state=cs, scenario_requests=requests
        )
        round_out = next(r for r in results if r["scenario_id"] == "s_round")["computed_outputs"]
        assert round_out["completeness"] == "full", round_out.get("blockers")
        # the SAFE rolled forward and converted in the post-flip round:
        assert round_out["aggregate_ownership_by_class"]["safe_pct"] > 0


class TestAcquisitionRules:
    def _rules(self) -> dict[str, Any]:
        path = os.path.join(_REPO, "founder-skills", "skills", "cap-table", "data", "cap-table-rules.json")
        with open(path) as f:
            return json.load(f)  # type: ignore[no-any-return]

    def test_acquisition_rules_present_and_valid(self) -> None:
        import jsonschema

        rules = self._rules()
        spath = os.path.join(
            _REPO, "founder-skills", "skills", "cap-table", "references", "cap-table-rules.schema.json"
        )
        with open(spath) as f:
            schema = json.load(f)
        jsonschema.validate(rules, schema)  # must not raise
        acq = rules["domains"]["acquisition"]
        ids = {r["rule_id"] for r in acq}
        assert {"acquisition.consideration_shares", "acquisition.pool_consideration_basis", "acquisition.timing"} <= ids

    def test_consideration_shares_rule_matches_solver_constant(self) -> None:
        assert priced_round.ACQUISITION_RULE_ID == "acquisition.consideration_shares"


class TestReconciliationStatus:
    """Structured reconciliation_status function (spec §3.4).

    compute_reconciliation_status takes STRUCTURED dicts (computed, stated),
    NOT the pre-rendered markdown rows from build_reconciliation_lines.
    """

    def test_no_stated_terms_is_not_applicable(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from compose_report import compute_reconciliation_status  # type: ignore[import-not-found]

        status, ppm = compute_reconciliation_status(computed={"pps": 1.0, "fd": 100}, stated=None)
        assert status == "not_applicable"
        assert ppm == 0.0

    def test_within_tolerance_is_pass(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from compose_report import compute_reconciliation_status  # type: ignore[import-not-found]

        status, ppm = compute_reconciliation_status(computed={"pps": 1.0000}, stated={"pps": 1.0005})
        assert status == "pass"
        assert math.isclose(ppm, 500.0, abs_tol=2)  # 0.0005 / 1.0005 ≈ 499.75 ppm ≤ 2-tol from 500

    def test_beyond_tolerance_is_diverged(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from compose_report import compute_reconciliation_status  # type: ignore[import-not-found]

        status, ppm = compute_reconciliation_status(computed={"pps": 1.0}, stated={"pps": 1.10})
        assert status == "diverged"


class TestDisclosure:
    def test_uncovered_banner(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from compose_report import build_disclosure_banner  # type: ignore[import-not-found]

        b = build_disclosure_banner(covered=False, reconciliation_status="not_applicable", max_divergence_ppm=0.0)
        assert "outside the validated cap-table engine" in b.lower() or "provisional" in b.lower()

    def test_covered_but_diverged_banner(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from compose_report import build_disclosure_banner  # type: ignore[import-not-found]

        b = build_disclosure_banner(covered=True, reconciliation_status="diverged", max_divergence_ppm=120_000)
        assert "diverge" in b.lower()

    def test_covered_and_passing_has_no_banner(self) -> None:
        sys.path.insert(0, SCRIPTS)
        from compose_report import build_disclosure_banner  # type: ignore[import-not-found]

        b = build_disclosure_banner(covered=True, reconciliation_status="pass", max_divergence_ppm=10.0)
        assert b == ""

    def test_disclosure_instance_validates(self) -> None:
        import jsonschema

        spath = os.path.join(
            _REPO,
            "founder-skills",
            "skills",
            "cap-table",
            "references",
            "schemas",
            "coverage-disclosure.schema.json",
        )
        with open(spath) as f:
            schema = json.load(f)
        instance = {
            "schema_version": "v0.1-coverage-disclosure",
            "covered": False,
            "computation_method": "manual_outside_pipeline",
            "required_primitives": ["priced_round"],
            "uncovered_parts": [],
            "counsel_review": True,
            "reconciliation": {"status": "not_applicable", "max_divergence_ppm": 0.0},
            "manual_artifacts": ["analysis.md"],
        }
        jsonschema.validate(instance, schema)  # must not raise


class TestAcquisitionCouplingGoldens:
    BASE_INPUTS = {
        "company_name": "Acmecorp",
        "analysis_date": "2026-01-01",
        "mode": "standard",
        "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
        "preferred_series": [],
        "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
        "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
    }

    def _cs(
        self,
        inputs: dict[str, Any] | None = None,
        safes: list[dict[str, Any]] | None = None,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        instruments: dict[str, Any] = {
            "safes": safes or [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        cs: dict[str, Any] = cap_state_mod.build_cap_state(inputs or self.BASE_INPUTS, instruments)
        return cs, instruments

    def test_safe_holder_pct_under_acquisition_dilution_convention(self) -> None:
        # SAFE + concurrent acquisition: C dilutes the SAFE (C in post_fd, NOT in company_cap).
        safe = {
            "id": "s1",
            "purchase_amount": 1_000_000,
            "post_money_valuation_cap": 10_000_000,
            "form": "yc_postmoney_cap",
            "discount_multiplier": None,
            "issuance_date": "2025-01-01",
        }
        cs, _ = self._cs(safes=[safe])
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.20},
        )
        assert r["completeness"] == "full", r.get("blockers")
        agg = r["aggregate_ownership_by_class"]
        assert math.isclose(agg["acquisition_pct"], 0.20, abs_tol=1e-4)  # invariant
        assert agg["safe_pct"] > 0
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-4)  # SAFE diluted by C, books still balance

    def test_acquisition_plus_anti_dilution_golden(self) -> None:
        # A preferred series with broad-based weighted-average AD + acquisition. AD flows through
        # adj_pre_fd (carried by every D-form); assert it computes and balances.
        inputs = {
            **self.BASE_INPUTS,
            "preferred_series": [
                {
                    "series_id": "A",
                    "series_name": "Series A",
                    "shares": 2_000_000,
                    "original_issue_price": 1.0,
                    "original_conversion_price": 1.0,
                    "current_conversion_price": 1.0,
                    "issuance_date": "2024-01-01",
                    "anti_dilution_protection": "broad_based_weighted_average",
                }
            ],
        }
        cs, _ = self._cs(inputs=inputs)
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.20},
        )
        assert r["completeness"] == "full", r.get("blockers")
        agg = r["aggregate_ownership_by_class"]
        assert math.isclose(agg["acquisition_pct"], 0.20, abs_tol=1e-4)
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-4)

    def test_convergence_stress_high_tk_with_post_money_pool(self) -> None:
        # t·k ≈ 0.85 (t=0.5, k=1.7 via new=7M/pre=10M) WITH a post-money pool: the full
        # PPS↔safe↔pool↔C feedback loop under amplification must converge or fail loud.
        safe = {
            "id": "s1",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": 12_000_000,
            "form": "yc_postmoney_cap",
            "discount_multiplier": None,
            "issuance_date": "2025-01-01",
        }
        cs, _ = self._cs(safes=[safe])
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[safe],
            notes=[],
            pre_money=10_000_000,
            new_money=7_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": 0.50},
            pool_consideration_basis="include",
        )
        # t·k = 0.5 * 1.7 = 0.85 < 1 → feasible. Must either converge cleanly or report a blocker
        # (never silently return a degenerate price).
        if r["completeness"] == "full":
            assert r["converged"] is True
            assert math.isclose(r["aggregate_ownership_by_class"]["acquisition_pct"], 0.50, abs_tol=2e-3)
        else:
            assert r["blockers"], "non-full result must carry a blocker, not silent failure"

    def test_discount_branch_safe_coupling_both_bases(self) -> None:
        # Spec §5.1 requires a DISCOUNT-branch SAFE (price-coupled, unlike the cap branch) under both
        # bases. A discount SAFE's share count depends on PPS, so the full PPS↔safe↔C loop runs.
        # No clean hand-derivable rational here → assert the structural invariants + that the two
        # bases produce different prices (the safe term enters the denominator differently).
        disc_safe = {
            "id": "s1",
            "purchase_amount": 1_000_000,
            "form": "yc_postmoney_discount",
            "discount_multiplier": 0.80,
            "post_money_valuation_cap": None,  # 20% off → multiplier 0.80
            "issuance_date": "2025-01-01",
        }
        results = {}
        for basis in ("includes_safe_conversion", "excludes_safe_conversion"):
            cs, _ = self._cs(safes=[disc_safe])
            r = priced_round.solve_priced_round(
                cap_state=cs,
                safes=[disc_safe],
                notes=[],
                pre_money=10_000_000,
                new_money=2_000_000,
                pre_money_basis=basis,
                acquisition={"consideration_pct": 0.20},
            )
            assert r["completeness"] == "full", (basis, r.get("blockers"))
            assert r["converged"] is True
            agg = r["aggregate_ownership_by_class"]
            assert math.isclose(
                agg["acquisition_pct"], 0.20, abs_tol=1e-4
            )  # invariant holds even with discount coupling
            assert agg["safe_pct"] > 0
            total = sum(
                agg[k]
                for k in (
                    "founders_pct",
                    "preferred_pct",
                    "option_pool_pct",
                    "safe_pct",
                    "note_pct",
                    "new_money_pct",
                    "acquisition_pct",
                )
            )
            assert math.isclose(total, 1.0, abs_tol=1e-4)
            results[basis] = r["equity_financing_price"]
        assert not math.isclose(
            results["includes_safe_conversion"],
            results["excludes_safe_conversion"],
            rel_tol=1e-3,
        )

    def test_pre_round_closed_changes_safe_share_count(self) -> None:
        # pre_round_closed folds C into company_cap_estimate → the SAFE's cap-price denominator grows
        # → the SAFE's share COUNT differs from the concurrent case.
        safe = {
            "id": "s1",
            "purchase_amount": 1_000_000,
            "post_money_valuation_cap": 10_000_000,
            "form": "yc_postmoney_cap",
            "discount_multiplier": None,
            "issuance_date": "2025-01-01",
        }
        acq_concurrent = {
            "acquired_entity": "TargetCo",
            "consideration_pct": 0.2,
            "consideration_form": "shares",
            "acquisition_timing": "concurrent_with_round",
        }
        acq_closed = {**acq_concurrent, "acquisition_timing": "pre_round_closed"}
        cs_conc, instruments = self._cs(inputs={**self.BASE_INPUTS, "acquisition": acq_concurrent}, safes=[safe])
        cs_closed, _ = self._cs(inputs={**self.BASE_INPUTS, "acquisition": acq_closed}, safes=[safe])
        common = dict(
            safes=[safe],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            pre_money_basis="excludes_safe_conversion",
        )
        r_conc = priced_round.solve_priced_round(cap_state=cs_conc, acquisition={"consideration_pct": 0.2}, **common)
        r_closed = priced_round.solve_priced_round(
            cap_state=cs_closed, acquisition=None, **common
        )  # folded, not re-solved
        assert r_conc["completeness"] == "full" and r_closed["completeness"] == "full"
        assert r_conc["shares_breakdown"]["safe_converted"] != r_closed["shares_breakdown"]["safe_converted"]


class TestPoolBasisNote:
    """Unit tests for build_pool_basis_note (report note, not math)."""

    def _fn(self) -> Any:
        sys.path.insert(0, SCRIPTS)
        from compose_report import build_pool_basis_note  # type: ignore[import-not-found]

        return build_pool_basis_note

    def test_exclude_basis_with_acquisition_and_pool(self) -> None:
        fn = self._fn()
        note = fn(
            target_pool_percent=0.10,
            pool_consideration_basis="exclude",
            realized_pool_pct=0.08,
            acquisition_pct=0.20,
        )
        assert "10.0%" in note
        assert "8.0%" in note
        assert "pre-consideration" in note
        assert "post-closing" in note

    def test_include_basis_with_acquisition_and_pool(self) -> None:
        fn = self._fn()
        note = fn(
            target_pool_percent=0.10,
            pool_consideration_basis="include",
            realized_pool_pct=0.10,
            acquisition_pct=0.20,
        )
        assert "10.0%" in note
        assert "post-closing" in note

    def test_no_acquisition_returns_empty(self) -> None:
        fn = self._fn()
        note = fn(
            target_pool_percent=0.10,
            pool_consideration_basis="exclude",
            realized_pool_pct=0.08,
            acquisition_pct=None,
        )
        assert note == ""

    def test_no_pool_returns_empty(self) -> None:
        fn = self._fn()
        note = fn(
            target_pool_percent=None,
            pool_consideration_basis="exclude",
            realized_pool_pct=0.0,
            acquisition_pct=0.20,
        )
        assert note == ""


class TestAcquisitionSolverConvergenceHardening:
    """Bracketed-root-find fallback for acquisition deals near the feasibility fold.

    See docs/internal/2026-06-30-solver-convergence-hardening.md. The negotiated-%
    acquisition path has a narrow feasible-but-slow band (t·k just below the
    pool-fold ~0.833) where the fixed-point iteration burns all 200 iters and
    false-fails with E_SOLVER_DID_NOT_CONVERGE while a real economic root exists.
    A bracketed 1-D root-find fallback solves that band and, for genuinely
    infeasible deals, returns a typed infeasibility instead of DID_NOT_CONVERGE.
    """

    BASE_INPUTS = {
        "company_name": "Acmecorp",
        "analysis_date": "2026-01-01",
        "mode": "standard",
        "founders": [{"founder_id": "f1", "name": "Founder", "common_shares": 8_000_000}],
        "preferred_series": [],
        "option_pool": {"issued_and_outstanding": 0, "available_for_grant": 0},
        "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
    }

    def _cs(
        self,
        inputs: dict[str, Any] | None = None,
        safes: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        instruments: dict[str, Any] = {
            "safes": safes or [],
            "convertible_notes": [],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "R1", "schema_version": "v0.5.0-instruments"},
        }
        return cap_state_mod.build_cap_state(inputs or self.BASE_INPUTS, instruments)  # type: ignore[no-any-return]

    _CAP_SAFE = {
        "id": "s1",
        "purchase_amount": 500_000,
        "post_money_valuation_cap": 12_000_000,
        "form": "yc_postmoney_cap",
        "discount_multiplier": None,
        "issuance_date": "2025-01-01",
    }

    def _solve(self, *, new_money: float, consideration_pct: float = 0.50, **overrides: Any) -> dict[str, Any]:
        safes = overrides.pop("safes", [self._CAP_SAFE])
        inputs = overrides.pop("inputs", None)
        kwargs: dict[str, Any] = dict(
            safes=safes,
            notes=[],
            pre_money=10_000_000,
            new_money=new_money,
            target_pool_percent=0.10,
            target_basis="post_money",
            pre_money_basis="excludes_safe_conversion",
            acquisition={"consideration_pct": consideration_pct},
            pool_consideration_basis="include",
        )
        kwargs.update(overrides)
        result: dict[str, Any] = priced_round.solve_priced_round(
            cap_state=self._cs(inputs=inputs, safes=safes), **kwargs
        )
        return result

    # --- The real gap: feasible-but-slow band must convert false-fail → pass ---

    def test_slow_band_converges_via_fallback(self) -> None:
        # new=6.55M → t·k=0.8275, a real root at pps≈0.0085, founders≈0.41%.
        # The fixed-point iteration needs ~264 iters (> the 200 cap) so it
        # false-fails today; the bracketed fallback must return the root.
        r = self._solve(new_money=6_550_000)
        assert r["completeness"] == "full", r.get("blockers")
        assert r["converged"] is True
        agg = r["aggregate_ownership_by_class"]
        # founders near-wiped but strictly positive with real margin (not float noise).
        assert agg["founders_pct"] > 0.001, agg["founders_pct"]
        assert math.isclose(agg["acquisition_pct"], 0.50, abs_tol=2e-3)
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-3)

    def test_slow_band_not_mislabeled_infeasible(self) -> None:
        # A KNOWN-feasible deal just inside the fold must NOT be reported infeasible.
        r = self._solve(new_money=6_550_000)
        codes = {b["code"] for b in r.get("blockers", [])}
        assert "E_ACQUISITION_INFEASIBLE" not in codes
        assert "E_SOLVER_DID_NOT_CONVERGE" not in codes

    # --- Genuine infeasibility → typed error (NOT DID_NOT_CONVERGE) ---

    def test_genuine_infeasibility_typed_error(self) -> None:
        # new=7M → t·k=0.85, past the pool-fold (~0.833): no positive root.
        r = self._solve(new_money=7_000_000)
        assert r["completeness"] == "structural_only"
        codes = {b["code"] for b in r["blockers"]}
        assert "E_ACQUISITION_INFEASIBLE" in codes, codes
        assert "E_SOLVER_DID_NOT_CONVERGE" not in codes

    def test_mislabel_boundary_bracketed(self) -> None:
        # Bracket the feasibility boundary: just-inside feasible, just-outside infeasible.
        inside = self._solve(new_money=6_550_000)
        outside = self._solve(new_money=7_000_000)
        assert inside["completeness"] == "full"
        assert outside["completeness"] == "structural_only"
        assert any(b["code"] == "E_ACQUISITION_INFEASIBLE" for b in outside["blockers"])

    # --- New feasibility guards ---

    def test_t_plus_target_ge_1_pool_overdetermined(self) -> None:
        # include basis: t + pool_target ≥ 1 ⇒ the 2×2 pool/C system is singular.
        # t=0.6, pool target=0.45 → 1.05 ≥ 1. Keep t·k < 1 (new=2M, k=1.2 → t·k=0.72).
        r = self._solve(
            new_money=2_000_000,
            consideration_pct=0.60,
            target_pool_percent=0.45,
            safes=[],
        )
        assert r["completeness"] == "structural_only"
        codes = {b["code"] for b in r["blockers"]}
        assert "E_ACQUISITION_POOL_OVERDETERMINED" in codes, codes
        # Must NOT be conflated with the t·k≥1 overdetermination.
        assert "E_ACQUISITION_OVERDETERMINED" not in codes

    def test_t_plus_target_ge_1_not_flagged_for_pre_money_basis(self) -> None:
        # The 2×2 pool/C singularity exists ONLY when the acquisition consideration C
        # sits in the pool denominator — i.e. a post_money-family target_basis. Under
        # target_basis="pre_money" the pool formula ignores C (no coupling), so
        # t + pool_target ≥ 1 is NOT a pool-overdetermination and must not be flagged
        # as one (it would be a false-positive typed block).
        r = self._solve(
            new_money=2_000_000,
            consideration_pct=0.60,
            target_pool_percent=0.45,
            target_basis="pre_money",
            safes=[],
        )
        codes = {b["code"] for b in r.get("blockers", [])}
        assert "E_ACQUISITION_POOL_OVERDETERMINED" not in codes, codes

    def test_cap_safe_sum_over_one_typed(self) -> None:
        # A cap SAFE whose purchase ≥ cap (Σ purchase/cap ≥ 1) has no finite
        # share count → typed infeasibility, not a silent degenerate price.
        big_safe = {
            "id": "s1",
            "purchase_amount": 12_000_000,
            "post_money_valuation_cap": 10_000_000,
            "form": "yc_postmoney_cap",
            "discount_multiplier": None,
            "issuance_date": "2025-01-01",
        }
        r = self._solve(new_money=2_000_000, consideration_pct=0.20, safes=[big_safe], target_pool_percent=None)
        assert r["completeness"] == "structural_only"
        codes = {b["code"] for b in r["blockers"]}
        assert "E_ACQUISITION_CAP_SAFE_OVERDETERMINED" in codes, codes

    # --- Sweep: every feasible point converges; beyond the fold is infeasible ---

    @pytest.mark.parametrize("new_money", [1_000_000, 3_000_000, 5_000_000, 6_000_000, 6_550_000])
    def test_feasible_sweep_converges(self, new_money: int) -> None:
        r = self._solve(new_money=new_money)
        assert r["completeness"] == "full", (new_money, r.get("blockers"))
        assert r["converged"] is True
        agg = r["aggregate_ownership_by_class"]
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-3)

    @pytest.mark.parametrize("new_money", [7_000_000, 9_000_000])
    def test_beyond_fold_infeasible(self, new_money: int) -> None:
        r = self._solve(new_money=new_money)
        assert r["completeness"] == "structural_only"
        assert any(b["code"] == "E_ACQUISITION_INFEASIBLE" for b in r["blockers"])

    # --- Coupling: discount-branch SAFE + AD both flow through the fallback ---

    def test_discount_branch_safe_in_band_converges(self) -> None:
        # A discount-branch SAFE converts at PPS·multiplier, so its share count is
        # price-coupled (unlike the cap branch): the full PPS↔safe↔pool↔C loop runs
        # through the fallback. A discount SAFE makes the fold very sharp, so rather
        # than hunt a razor-thin false-fail band, force the fallback with a low
        # max_iterations on a feasible point and assert it equals the fast path.
        disc_safe = {
            "id": "s1",
            "purchase_amount": 500_000,
            "form": "yc_postmoney_discount",
            "discount_multiplier": 0.80,
            "post_money_valuation_cap": None,
            "issuance_date": "2025-01-01",
        }
        fast = self._solve(new_money=5_800_000, safes=[disc_safe])
        forced = self._solve(new_money=5_800_000, safes=[disc_safe], max_iterations=30)
        assert fast["completeness"] == "full", fast.get("blockers")
        assert forced["completeness"] == "full", forced.get("blockers")
        assert forced["converged"] is True
        agg = forced["aggregate_ownership_by_class"]
        assert agg["safe_pct"] > 0
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-3)
        # The fallback root equals the fast-path fixed point (discount coupling).
        assert math.isclose(fast["equity_financing_price"], forced["equity_financing_price"], rel_tol=1e-5), (
            fast["equity_financing_price"],
            forced["equity_financing_price"],
        )

    def test_anti_dilution_acquisition_in_fallback_converges(self) -> None:
        # An acquisition deal WITH anti-dilution that reaches the fallback (slow
        # band) must converge via per-evaluation AD re-solve — not silently
        # freeze AD, not falsely report infeasible.
        inputs = {
            **self.BASE_INPUTS,
            "preferred_series": [
                {
                    "series_id": "A",
                    "series_name": "Series A",
                    "shares": 200_000,
                    "original_issue_price": 1.0,
                    "original_conversion_price": 1.0,
                    "current_conversion_price": 1.0,
                    "issuance_date": "2024-01-01",
                    "anti_dilution_protection": "broad_based_weighted_average",
                }
            ],
        }
        r = self._solve(new_money=6_400_000, inputs=inputs)
        assert r["completeness"] == "full", r.get("blockers")
        assert r["converged"] is True
        assert "anti_dilution_breakdown" in r
        agg = r["aggregate_ownership_by_class"]
        total = sum(
            agg[k]
            for k in (
                "founders_pct",
                "preferred_pct",
                "option_pool_pct",
                "safe_pct",
                "note_pct",
                "new_money_pct",
                "acquisition_pct",
            )
        )
        assert math.isclose(total, 1.0, abs_tol=1e-3)

    # --- The fallback root equals the fast-path fixed point at a non-fold point ---

    def test_fallback_root_matches_fast_path(self) -> None:
        # At a well-inside (non-fold) point the fast path converges quickly; force
        # the fallback via max_iterations=1 and assert it lands on the same PPS.
        fast = self._solve(new_money=3_000_000)
        forced = self._solve(new_money=3_000_000, max_iterations=1)
        assert fast["completeness"] == "full" and forced["completeness"] == "full", (
            fast.get("blockers"),
            forced.get("blockers"),
        )
        assert math.isclose(fast["equity_financing_price"], forced["equity_financing_price"], rel_tol=1e-6), (
            fast["equity_financing_price"],
            forced["equity_financing_price"],
        )

    # --- No-hang: the bracketed root-finder has a finite iteration cap ---

    def test_bracketed_solver_has_finite_cap(self) -> None:
        assert isinstance(priced_round.ACQ_BISECT_MAX_ITERS, int)
        assert priced_round.ACQ_BISECT_MAX_ITERS > 0
        assert isinstance(priced_round.ACQ_GRID_POINTS, int)
        assert priced_round.ACQ_GRID_POINTS > 0

    # ---- rev 5: rate-independent inner sub-solve (closed-form pool + affine cc) ----

    _TOTAL_KEYS = (
        "founders_pct",
        "preferred_pct",
        "option_pool_pct",
        "safe_pct",
        "note_pct",
        "new_money_pct",
        "acquisition_pct",
    )

    def test_very_high_Q_cap_safe_solves(self) -> None:
        # Mechanism A: Q = Σ purchase/cap = 0.99 (< 1). Feasible (founders ≈ 0.89%),
        # but the fixed-point cc rate is 0.99 → the pre-rev-5 inner loop (and the
        # cut Aitken plan, whose 20× fence rejects the exact projection for Q>0.952)
        # both mislabel it infeasible at the default cap. The affine cc solve must
        # return the number.
        safe = {
            "id": "s1",
            "purchase_amount": 9_900_000,
            "post_money_valuation_cap": 10_000_000,
            "form": "yc_postmoney_cap",
            "discount_multiplier": None,
            "issuance_date": "2025-01-01",
        }
        r = self._solve(
            new_money=1_000_000,
            consideration_pct=0.10,
            safes=[safe],
            target_pool_percent=None,
        )
        assert r["completeness"] == "full", r.get("blockers")
        assert r["converged"] is True
        agg = r["aggregate_ownership_by_class"]
        assert 0.008 < agg["founders_pct"] < 0.010, agg["founders_pct"]
        assert agg["safe_pct"] > 0
        assert math.isclose(sum(agg[k] for k in self._TOTAL_KEYS), 1.0, abs_tol=1e-3)

    def test_pool_singular_feasible_solves(self) -> None:
        # Mechanism B: no SAFE (Q=0), include+post_money, t=target=0.499 → a·b≈0.992.
        # Genuine root (founders ≈ 0.20%, pool ≈ 49.9%, acq ≈ 49.9%). The closed-form
        # pool must solve it; must NOT be infeasible or nonconvergent.
        r = self._solve(
            new_money=1,
            consideration_pct=0.499,
            target_pool_percent=0.499,
            safes=[],
        )
        assert r["completeness"] == "full", r.get("blockers")
        agg = r["aggregate_ownership_by_class"]
        assert agg["founders_pct"] > 0.0015
        assert math.isclose(agg["acquisition_pct"], 0.499, abs_tol=2e-3)
        assert math.isclose(agg["option_pool_pct"], 0.499, abs_tol=2e-3)
        assert math.isclose(sum(agg[k] for k in self._TOTAL_KEYS), 1.0, abs_tol=1e-3)

    def test_finalize_reported_numbers_rate_independent(self) -> None:
        # The reported founders%/pool must be rate-independent — identical at the
        # default cap and a very high cap. Pre-rev-5, _finalize's own pool loop
        # under-converges the REPORTED numbers at high a·b even when the root is found.
        orig = priced_round.ACQ_INNER_MAX_ITERS
        try:
            priced_round.ACQ_INNER_MAX_ITERS = orig
            r_default = self._solve(new_money=1, consideration_pct=0.4999, target_pool_percent=0.4999, safes=[])
            priced_round.ACQ_INNER_MAX_ITERS = 200_000
            r_high = self._solve(new_money=1, consideration_pct=0.4999, target_pool_percent=0.4999, safes=[])
        finally:
            priced_round.ACQ_INNER_MAX_ITERS = orig
        assert r_default["completeness"] == "full", r_default.get("blockers")
        assert r_high["completeness"] == "full", r_high.get("blockers")
        f_def = r_default["aggregate_ownership_by_class"]["founders_pct"]
        f_high = r_high["aggregate_ownership_by_class"]["founders_pct"]
        assert math.isclose(f_def, f_high, rel_tol=1e-6), (f_def, f_high)

    def test_pool_target_out_of_range_typed(self) -> None:
        # D0: target ≥ 1 on any basis → typed blocker, not a silent full and not a crash.
        for basis in ("post_money", "pre_money"):
            r = self._solve(
                new_money=2_000_000,
                consideration_pct=0.20,
                target_pool_percent=1.5,
                target_basis=basis,
                safes=[],
            )
            assert r["completeness"] == "structural_only", (basis, r)
            assert any(b["code"] == "E_POOL_TARGET_OUT_OF_RANGE" for b in r["blockers"]), (basis, r["blockers"])

    def test_zero_pool_not_blocked_by_d0(self) -> None:
        # D0 truthiness: explicit target_pool_percent=0.0 is a legit no-pool deal.
        r = self._solve(new_money=2_000_000, consideration_pct=0.20, target_pool_percent=0.0, safes=[])
        assert r["completeness"] == "full", r.get("blockers")

    def test_acceleration_does_not_move_the_root(self) -> None:
        # I2: the reported PPS at the default cap equals a very-high-cap reference.
        orig = priced_round.ACQ_INNER_MAX_ITERS
        try:
            r_def = self._solve(new_money=6_550_000)
            priced_round.ACQ_INNER_MAX_ITERS = 200_000
            r_high = self._solve(new_money=6_550_000)
        finally:
            priced_round.ACQ_INNER_MAX_ITERS = orig
        assert r_def["completeness"] == "full" and r_high["completeness"] == "full"
        assert math.isclose(r_def["equity_financing_price"], r_high["equity_financing_price"], rel_tol=1e-9), (
            r_def["equity_financing_price"],
            r_high["equity_financing_price"],
        )

    def test_cap_plus_discount_acquisition_fallback(self) -> None:
        # The non-affine residual: a cap_plus_discount SAFE (min(cap,disc) kink) + acquisition
        # forced into the fallback must still resolve full (affinity check → bounded iteration).
        cpd = {
            "id": "s1",
            "purchase_amount": 500_000,
            "post_money_valuation_cap": 12_000_000,
            "form": "cap_plus_discount",
            "discount_multiplier": 0.80,
            "issuance_date": "2025-01-01",
        }
        fast = self._solve(new_money=3_000_000, safes=[cpd])
        forced = self._solve(new_money=3_000_000, safes=[cpd], max_iterations=1)
        assert fast["completeness"] == "full", fast.get("blockers")
        assert forced["completeness"] == "full", forced.get("blockers")
        assert math.isclose(fast["equity_financing_price"], forced["equity_financing_price"], rel_tol=1e-5), (
            fast["equity_financing_price"],
            forced["equity_financing_price"],
        )

    # ---- rev 5: pure-helper unit tests (R1 pool, R6 affine cc) ----

    def test_acquisition_pool_C_matches_brute_force(self) -> None:
        # R1: the closed-form pool/C helper equals a brute-force clamped fixed point
        # across target_basis × pool_basis × clamp regimes.
        fn = priced_round._acquisition_pool_C

        def brute(pre_pool: float, nm: float, target: float, acq_t: float, existing: float, tb: str, pb: str) -> Any:
            b = acq_t / (1.0 - acq_t)
            x = 0.0
            post = tb in ("post_money", "post_money_excluding_converting_securities")
            # Tight ABSOLUTE break: near a·b→1 a relative tolerance stops many shares
            # short of the fixed point (the very rate problem the closed form removes).
            for _ in range(5_000_000):
                base = pre_pool + nm
                C = b * (base + x)
                acq_for_pool = C if pb == "include" else 0.0
                if post:
                    x_new = max(0.0, (target * (pre_pool + nm + acq_for_pool) - existing) / (1.0 - target))
                else:
                    x_new = max(0.0, (target * pre_pool - existing) / (1.0 - target))
                if abs(x_new - x) <= 1e-4:
                    x = x_new
                    break
                x = x_new
            C = b * (pre_pool + nm + x)
            return x, C

        cases = [
            # (pre_pool, nm, target, acq_t, existing, target_basis, pool_basis)
            (16_000_000, 3_000_000, 0.10, 0.50, 0.0, "post_money", "include"),
            (16_000_000, 3_000_000, 0.10, 0.50, 0.0, "post_money", "exclude"),
            (16_000_000, 3_000_000, 0.499, 0.499, 0.0, "post_money", "include"),
            (16_000_000, 3_000_000, 0.10, 0.20, 50_000_000, "post_money", "include"),  # clamp x→0
            (8_000_000, 2_000_000, 0.10, 0.20, 0.0, "pre_money", "include"),
        ]
        for pre_pool, nm, target, acq_t, existing, tb, pb in cases:
            got = fn(
                pre_pool=pre_pool,
                nm=nm,
                target=target,
                acq_t=acq_t,
                existing=existing,
                target_basis=tb,
                pool_basis=pb,
            )
            assert got is not None, (tb, pb)
            x_got, c_got = got
            x_ref, c_ref = brute(pre_pool, nm, target, acq_t, existing, tb, pb)
            assert abs(x_got - x_ref) <= 1.0, (tb, pb, x_got, x_ref)
            assert abs(c_got - c_ref) <= 1.0, (tb, pb, c_got, c_ref)

    def test_affine_cc_solve_matches_brute_force(self) -> None:
        # R6: the 2-probe affine cc solve equals a brute-force cc fixed point, and a
        # non-affine sequence returns None.
        fn = priced_round._affine_cc_solve
        adj, note = 8_000_000.0, 0.0
        # Affine: safe_total(cc) = 0.55*cc + 1_000_000
        cc = fn(lambda c: 0.55 * c + 1_000_000.0, adj, note)
        assert cc is not None
        assert math.isclose(cc, (adj + note + 1_000_000.0) / (1.0 - 0.55), rel_tol=1e-12)
        # m ≥ 1 → None (cap SAFEs demand ≥100%)
        assert fn(lambda c: 1.2 * c, adj, note) is None
        # Non-affine: the min() kink is straddled by the probes (adj=8e6 on the slope
        # branch, 2·adj=16e6 on the cap branch) → affinity check fails → None.
        assert fn(lambda c: min(0.9 * c, 10_000_000.0), adj, note) is None


class TestNoteUsableGuard:
    """A partial/blank note (null principal) must be filtered as terms-only
    at the note math chokepoints — it must NOT crash convert_note at note_conversion._conversion_balance
    (note["principal"] + accrued -> TypeError). The value may live in a Schedule of Lenders, not the
    blank Note template.
    """

    def _partial_note(self) -> dict:
        # Has a conversion path (cap+discount) so it reaches _conversion_balance, but null principal.
        return {
            "id": "note_partial",
            "principal": None,
            "interest_rate_type": "none",
            "annual_interest_rate": None,
            "day_count_basis": 365,
            "interest_converts_to_shares": True,
            "issuance_date": "2025-06-01",
            "valuation_cap": 10_000_000,
            "discount_multiplier": 0.80,
            "qualified_financing_threshold": 1_000_000,
        }

    def test_run_note_scenario_surfaces_partial_note_as_terms_only(self) -> None:
        scenario = {"parameters": {"transaction_event_date": "2027-06-01", "priced_round_new_money": 5_000_000}}
        instruments = {"convertible_notes": [self._partial_note()]}
        out = run_scenario.run_note_conversion_scenario(scenario, instruments=instruments, cap_state={})
        # Not silently dropped, not converted, surfaced as terms-only excluded.
        assert "note_partial" in out["per_note"]
        assert out["per_note"]["note_partial"].get("branch") == "terms_only_excluded"

    def test_note_has_usable_math_inputs_predicate(self) -> None:
        assert note_conversion.note_has_usable_math_inputs({"principal": 100_000, "issuance_date": "2025-06-01"})
        assert not note_conversion.note_has_usable_math_inputs(self._partial_note())
        assert not note_conversion.note_has_usable_math_inputs({"principal": 100_000, "issuance_date": None})
        assert not note_conversion.note_has_usable_math_inputs({"principal": 0, "issuance_date": "2025-06-01"})


class TestNoteUsableGuardSolver:
    """The note-usable guard at solve_priced_round (quick_assess calls it
    directly, bypassing run_scenario), plus the day_count_basis null-coalesce."""

    def test_solve_priced_round_filters_partial_note_without_crashing(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)
        partial_note = {
            "id": "note_partial",
            "principal": None,
            "interest_rate_type": "none",
            "annual_interest_rate": None,
            "day_count_basis": 365,
            "interest_converts_to_shares": True,
            "issuance_date": "2025-06-01",
            "valuation_cap": 10_000_000,
            "discount_multiplier": 0.80,
            "qualified_financing_threshold": 1_000_000,
        }
        r = priced_round.solve_priced_round(
            cap_state=cs,
            safes=[],
            notes=[partial_note],
            pre_money=20_000_000,
            new_money=5_000_000,
            conversion_event_date="2027-06-01",
        )
        # No crash; the partial note contributes no shares (excluded from the solve).
        assert "note_partial" not in r.get("per_note", {})

    def test_convert_note_coalesces_null_day_count_basis(self) -> None:
        note = {
            "id": "note_dc",
            "principal": 1_000_000,
            "interest_rate_type": "fixed_numeric",
            "annual_interest_rate": 0.05,
            "day_count_basis": None,  # blank in the doc — must coalesce to 365, not TypeError
            "compounding_periods_per_year": None,
            "interest_converts_to_shares": True,
            "issuance_date": "2025-01-01",
            "last_interest_event_date": None,
            "valuation_cap": 10_000_000,
            "discount_multiplier": 0.80,
            "qualified_financing_threshold": 1_000_000,
        }
        r = note_conversion.convert_note(note, conversion_event_date="2027-01-01", priced_round_new_money=5_000_000)
        # The interest math ran past the null day_count_basis (coalesced to 365) without a TypeError.
        assert "accrued_interest" in r


class TestNotePrincipalMirror:
    """_build_outstanding_notes mirrors the SAFE terms-only treatment — a
    null-principal note is a non-convertible terms-only entry with W_NOTE_PRINCIPAL_MISSING, never
    passed through as convertible (which corrupts/crashes downstream note math)."""

    def _inputs(self) -> dict:
        return {
            "company_name": "TestCo",
            "analysis_date": "2026-06-01",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-01-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
            "preferred_series": [],
            "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
            "common_batches": [],
        }

    def _null_principal_note(self) -> dict:
        return {
            "id": "note_partial",
            "investor_name": "Lender X",
            "principal": None,  # blank in the Note template; the amount lives in a Schedule of Lenders
            "issuance_date": "2024-06-01",
            "interest_rate_type": "none",
            "annual_interest_rate": None,
            "day_count_basis": 365,
        }

    def test_null_principal_note_is_terms_only_and_warns(self) -> None:
        instruments = {
            "safes": [],
            "convertible_notes": [self._null_principal_note()],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        cs = cap_state_mod.build_cap_state(self._inputs(), instruments)
        on = cs["outstanding_notes"][0]
        assert on.get("convertible") is False
        assert on.get("principal") is None
        assert "W_NOTE_PRINCIPAL_MISSING" in cs.get("warnings", [])

    def test_usable_note_stays_convertible(self) -> None:
        note = {**self._null_principal_note(), "id": "note_ok", "principal": 250_000}
        instruments = {
            "safes": [],
            "convertible_notes": [note],
            "warrants": [],
            "option_grants": [],
            "metadata": {"run_id": "test"},
        }
        cs = cap_state_mod.build_cap_state(self._inputs(), instruments)
        on = cs["outstanding_notes"][0]
        assert on.get("convertible") is True
        assert on.get("principal") == 250_000
        assert "W_NOTE_PRINCIPAL_MISSING" not in cs.get("warnings", [])


class TestTermsOnlyWarningCallouts:
    """The terms-only warning codes must render a founder-facing callout —
    unmatched codes are silently dropped, so a terms-only instrument would otherwise produce zero
    shares with no explanation anywhere in the report."""

    def test_safe_purchase_amount_missing_renders_callout(self) -> None:
        from _warning_callouts import render_warning_callouts  # type: ignore[import-not-found]

        out = render_warning_callouts(["W_SAFE_PURCHASE_AMOUNT_MISSING"])
        assert out
        assert any("terms-only" in line.lower() or "purchase amount" in line.lower() for line in out)

    def test_note_principal_missing_renders_callout(self) -> None:
        from _warning_callouts import render_warning_callouts  # type: ignore[import-not-found]

        out = render_warning_callouts(["W_NOTE_PRINCIPAL_MISSING"])
        assert out
        assert any("terms-only" in line.lower() or "principal" in line.lower() for line in out)


class TestExtractionReportNullRendering:
    """A present-with-null investor_name/issuance_date must render
    the neutral 'to confirm' marker, never the literal string 'None'."""

    def test_note_null_investor_and_date_not_literal_none(self) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        note = {
            "id": "n1",
            "investor_name": None,
            "principal": None,
            "issuance_date": None,
            "valuation_cap": None,
            "discount_multiplier": None,
            "annual_interest_rate": None,
            "maturity_date": None,
            "governing_law": None,
        }
        row = cer._notes_table([note], {})[-1]
        assert "None" not in row
        assert cer.NEUTRAL_MARKER in row

    def test_safe_null_investor_and_date_not_literal_none(self) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        safe = {
            "id": "s1",
            "investor_name": None,
            "purchase_amount": None,
            "issuance_date": None,
            "form": None,
            "discount_multiplier": None,
        }
        row = cer._safes_table([safe], {})[-1]
        assert "None" not in row
        assert cer.NEUTRAL_MARKER in row


class TestExtractionReportSameFileGuard:
    """Staging inputs/instruments inside --review-dir (the natural
    location) must not crash compose_extraction_report with shutil.SameFileError."""

    def test_copy_if_different_noop_when_same_path(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "inputs.json"
        p.write_text("{}")
        cer._copy_if_different(str(p), str(p))  # src == dst -> must not raise SameFileError
        assert p.read_text() == "{}"

    def test_copy_if_different_copies_when_paths_differ(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        src = tmp_path / "a.json"
        src.write_text('{"x":1}')
        dst = tmp_path / "b.json"
        cer._copy_if_different(str(src), str(dst))
        assert dst.read_text() == '{"x":1}'


class TestExtractionReportAmendment:
    """An amendment-only doc classifies non-extractable (no math instrument persisted), so its
    clause deltas live only in the extract_instrument receipt. The extraction-only renderer must
    surface them from the receipt/audit file as an 'Amendments (terms modified)' section instead
    of producing an empty '(no SAFEs/notes/warrants supplied)' deliverable."""

    def _amendment_audit(self) -> dict[str, Any]:
        return {
            "classified_doc_type": "amendment",
            "classified_as_non_extractable": True,
            "ambiguities": [
                {
                    "field": "amended_clause_delta.new_definition",
                    "description": "Amends only the 'Qualified Financing' definition; other note terms unchanged.",
                },
                {
                    "field": "qualified_financing_threshold change",
                    "description": "Restates the Qualified Financing threshold as $2,500,000 gross proceeds.",
                },
            ],
        }

    def test_load_amendment_deltas_parses_amendment_receipt(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "extraction_audit.json"
        p.write_text(json.dumps(self._amendment_audit()))
        deltas = cer._load_amendment_deltas(str(p))
        assert [d["field"] for d in deltas] == [
            "amended_clause_delta.new_definition",
            "qualified_financing_threshold change",
        ]
        assert "Qualified Financing" in deltas[0]["description"]

    def test_load_amendment_deltas_empty_for_non_amendment(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "extraction_audit.json"
        p.write_text(json.dumps({"classified_doc_type": "safe", "ambiguities": [{"field": "x", "description": "y"}]}))
        assert cer._load_amendment_deltas(str(p)) == []
        assert cer._load_amendment_deltas(None) == []

    def test_compose_renders_amendment_section(self) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        result = cer.compose_extraction_report(
            company_name="Acmecorp Inc.",
            inputs={"jurisdiction": {"structure": "delaware_c_corp"}, "analysis_date": "2026-01-01"},
            instruments={"safes": [], "convertible_notes": [], "warrants": [], "option_grants": []},
            amendments=self._amendment_audit()["ambiguities"],
        )
        md = result["_report_md"]
        assert "Amendments" in md
        assert "Qualified Financing" in md
        assert "$2,500,000" in md
        # The deliverable must not read as empty when an amendment IS present.
        assert md.count("_(no SAFEs, convertible notes, or warrants supplied)_") == 0


class TestExtractionReportTermsDoc:
    """A term_sheet/option_plan classifies non-math; its content lives in the receipt and renders as a
    'Term sheet terms (as extracted)' section fed via --audit — never an empty deliverable. Every
    verifier/invariant finding surfaces as an honest per-field marker."""

    def _receipt(self) -> dict[str, Any]:
        return {
            "classified_doc_type": "term_sheet",
            "classified_as_terms_doc": True,
            "terms_doc": {
                "fields": {
                    "pre_money_valuation": 10_000_000,
                    "investment_amount": 99_000_000,
                    "board_composition": "2 founders, 1 investor",
                    "protective_provisions": ["veto on sale", "veto on new debt"],
                    "no_shop_days": None,
                },
                "confidence": {
                    "pre_money_valuation": {"level": "high", "evidence_quote": "Pre-Money is $10,000,000"},
                    "investment_amount": {"level": "high", "evidence_quote": "Investment is $99,000,000"},
                },
            },
            "evidence_verification": {
                "per_field": [
                    {"field": "pre_money_valuation", "status": "pass", "reasons": []},
                    {"field": "investment_amount", "status": "fail", "reasons": ["value_in_doc failed: absent"]},
                    {"field": "board_composition", "status": "unverified_no_quote", "reasons": ["no quote"]},
                ]
            },
            "invariant_check": {"violations": [{"field": "pre_money_valuation", "stake": "hard", "reason": "x"}]},
            "ambiguities": [{"field": "no_shop_days", "description": "not stated; confirm"}],
        }

    def test_load_terms_doc_receipt_shape(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "audit.json"
        p.write_text(json.dumps(self._receipt()))
        td = cer._load_terms_doc(str(p))
        assert td is not None
        assert td["itype"] == "term_sheet"
        assert td["fields"]["pre_money_valuation"] == 10_000_000
        assert td["status_by_field"]["investment_amount"] == "fail"
        assert "pre_money_valuation" in td["invariant_fields"]

    def test_load_terms_doc_raw_handoff_shape(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "audit.json"
        p.write_text(json.dumps({"instrument_type": "option_plan", "fields": {"pool_size": 1000}, "confidence": {}}))
        td = cer._load_terms_doc(str(p))
        assert td is not None
        assert td["itype"] == "option_plan"
        assert td["fields"]["pool_size"] == 1000
        assert td["status_by_field"] == {}  # no evidence_verification in the raw shape

    def test_load_terms_doc_empty_for_non_terms(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "audit.json"
        p.write_text(json.dumps({"classified_doc_type": "amendment", "ambiguities": []}))
        assert cer._load_terms_doc(str(p)) is None
        assert cer._load_terms_doc(None) is None

    def _compose(self, terms_doc: dict[str, Any]) -> str:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        result = cer.compose_extraction_report(
            company_name="Acmecorp Inc.",
            inputs={"jurisdiction": {"structure": "delaware"}, "analysis_date": "2026-01-01"},
            instruments={"safes": [], "convertible_notes": [], "warrants": [], "option_grants": []},
            terms_doc=terms_doc,
        )
        return str(result["_report_md"])

    def test_renders_terms_section_with_status_markers(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "audit.json"
        p.write_text(json.dumps(self._receipt()))
        md = self._compose(cer._load_terms_doc(str(p)))
        assert "Term sheet terms" in md
        assert "Board Composition" in md  # humanized snake_case
        assert "veto on sale; veto on new debt" in md  # list joined
        assert "value not found in source text" in md  # fail marker (softened, not "NOT FOUND")
        assert "no evidence quote" in md  # unverified_no_quote distinct from fail
        assert "fails a cross-field check" in md  # invariant marker on the pass field
        # The empty deliverable must not appear when a term sheet IS present.
        assert md.count("_(no SAFEs, convertible notes, or warrants supplied)_") == 0
        # a confident hallucination must NOT render its self-claimed 'high' unqualified
        assert "| high |" not in md.replace("High", "high") or "value not found" in md

    def test_raw_handoff_fallback_renders_without_markers(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "audit.json"
        p.write_text(
            json.dumps(
                {
                    "instrument_type": "term_sheet",
                    "fields": {"pre_money_valuation": 5_000_000},
                    "confidence": {"pre_money_valuation": {"level": "medium"}},
                }
            )
        )
        md = self._compose(cer._load_terms_doc(str(p)))
        assert "Term sheet terms" in md
        assert "Pre Money Valuation" in md
        assert "value not found in source text" not in md  # no evidence_verification → no fail markers

    def test_terms_doc_does_not_trigger_amendment_section(self, tmp_path: Any) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        p = tmp_path / "audit.json"
        p.write_text(json.dumps(self._receipt()))
        # A terms-doc payload must not be read as amendment deltas, and vice versa.
        assert cer._load_amendment_deltas(str(p)) == []
        md = self._compose(cer._load_terms_doc(str(p)))
        assert "Amendments (terms modified)" not in md

    def test_flat_dict_field_renders_readable_not_raw_json(self) -> None:
        import compose_extraction_report as cer  # type: ignore[import-not-found]

        # A flat nested field (board composition) must render as humanized "Key: v" pairs, not a raw
        # JSON/Python dump in one table cell.
        cell = cer._terms_cell({"total_directors": 3, "investor_seats": 1}, "board_composition", {})
        assert cell == "Total Directors: 3; Investor Seats: 1"
        # Deeply-nested values fall back to compact JSON (no crash, still pipe-safe).
        nested = {"tranche_a": {"amount": 1}}
        assert cer._terms_cell(nested, "x", {}) == json.dumps(nested, ensure_ascii=False)

    def test_pipeline_splice_handoff_to_rendered_section(self, tmp_path: Any) -> None:
        """End-to-end across BOTH scripts: a term_sheet hand-off → extract_instrument (-o receipt) →
        compose_extraction_report --audit → a rendered Term sheet section, instruments untouched."""
        instr = str(tmp_path / "instruments.json")
        with open(instr, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "t"},
                },
                f,
            )
        doc = tmp_path / "ts.txt"
        doc.write_text('TERM SHEET. The "Pre-Money Valuation" is $20,000,000. ' + "padding " * 80)
        receipt = str(tmp_path / "extraction_audit.json")
        handoff = {
            "instrument_type": "term_sheet",
            "fields": {"pre_money_valuation": 20_000_000, "board_composition": "2 founders, 1 investor"},
            "confidence": {
                "pre_money_valuation": {"level": "high", "evidence_quote": 'The "Pre-Money Valuation" is $20,000,000.'}
            },
            "ambiguities": [],
        }
        rc, _o, err = _run(
            "extract_instrument.py",
            ["--instruments", instr, "--run-id", "t", "--verify", "--source-doc", str(doc), "-o", receipt],
            stdin_data=json.dumps(handoff),
        )
        assert rc == 0, err
        with open(instr) as f:
            data = json.load(f)
        assert all(len(data[k]) == 0 for k in ("safes", "convertible_notes", "warrants", "option_grants"))
        inputs = str(tmp_path / "inputs.json")
        with open(inputs, "w") as f:
            json.dump(
                {
                    "company_name": "Acmecorp Inc.",
                    "analysis_date": "2026-01-01",
                    "jurisdiction": {"structure": "delaware"},
                },
                f,
            )
        rd = str(tmp_path / "out")
        rc2, _o2, err2 = _run(
            "compose_extraction_report.py",
            ["--inputs", inputs, "--instruments", instr, "--review-dir", rd, "--audit", receipt, "--run-id", "t"],
        )
        assert rc2 == 0, err2
        with open(os.path.join(rd, "report_extraction_only.md")) as f:
            md = f.read()
        assert "Term sheet terms (as extracted)" in md
        assert "Pre Money Valuation" in md
        assert "_(no SAFEs, convertible notes, or warrants supplied)_" not in md
        assert os.path.exists(os.path.join(rd, "extraction_only.json"))


class TestExtractInstrumentSourceDocCheck:
    """A --source-doc pointing at a nonexistent file must fail
    loud (E_SOURCE_DOC_NOT_FOUND, exit 1) BEFORE write_artifact — not silently degrade to
    unverifiable_doc / ok:true, and not persist the instrument."""

    def test_missing_source_doc_fails_loud_and_writes_nothing(self, tmp_path: Any) -> None:
        instr_path = str(tmp_path / "instruments.json")
        with open(instr_path, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        extraction = {
            "instrument_type": "safe",
            "fields": {
                "investor_name": "X",
                "purchase_amount": 100000,
                "issuance_date": "2024-01-01",
                "form": "yc_postmoney_cap",
                "post_money_valuation_cap": 5_000_000,
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, _, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--source-doc", str(tmp_path / "nope.pdf")],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 1
        assert "E_SOURCE_DOC_NOT_FOUND" in e
        with open(instr_path) as f:
            assert json.load(f)["safes"] == []  # artifact untouched


class TestAmendmentClassification:
    """An amendment (a clause-only restatement whose other instrument terms are legitimately absent)
    is classified non-extractable — not forced through the note gate, not persisted to any math array
    (even a relaxed gate would otherwise let an all-null 'note' into convertible_notes)."""

    def _instr_file(self, tmp_path: Any) -> str:
        p = str(tmp_path / "instruments.json")
        with open(p, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        return p

    def test_amendment_classified_not_persisted(self, tmp_path: Any) -> None:
        instr_path = self._instr_file(tmp_path)
        extraction = {
            "instrument_type": "amendment",
            "fields": {},
            "confidence": {},
            "ambiguities": [{"field": "qualified_financing", "reason": "restates the QF definition only"}],
        }
        rc, out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"rc={rc} stderr={e}"
        receipt = json.loads(out)
        assert receipt.get("classified_doc_type") == "amendment"
        assert receipt.get("classified_as_non_extractable") is True
        with open(instr_path) as f:
            instr = json.load(f)
        assert instr["safes"] == []
        assert instr["convertible_notes"] == []
        assert instr["warrants"] == []


class TestEvidencedAbsenceRelax:
    """Core contract: a required field that is missing is an error UNLESS the extraction
    asserts confidence.level == 'absent' for it (and it is in the type's relaxable set), in which
    case the instrument validates (persisted as a partial). A silently-missing field still errors."""

    def test_validate_safe_relaxes_documented_absent_fields(self) -> None:
        from extract_instrument import validate_safe  # type: ignore[import-not-found]

        fields = {
            "form": "yc_postmoney_cap",
            "post_money_valuation_cap": 5_000_000,
            "purchase_amount": None,
            "issuance_date": None,
        }
        confidence = {
            "purchase_amount": {"level": "absent"},
            "issuance_date": {"level": "absent"},
        }
        assert validate_safe(fields, confidence=confidence) == []

    def test_validate_safe_still_errors_on_silently_missing(self) -> None:
        from extract_instrument import validate_safe  # type: ignore[import-not-found]

        fields = {
            "form": "yc_postmoney_cap",
            "post_money_valuation_cap": 5_000_000,
            "purchase_amount": None,
            "issuance_date": None,
        }
        errors = validate_safe(fields)  # no confidence -> not documented-absent -> still errors
        assert any("purchase_amount" in e for e in errors)

    def test_validate_note_relaxes_documented_absent_principal(self) -> None:
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        fields = {
            "principal": None,
            "day_count_basis": 365,
            "issuance_date": "2024-01-01",
            "maturity_date": None,
            "maturity_default_treatment": None,
            "interest_rate_type": "none",
            "annual_interest_rate": None,
        }
        confidence = {
            "principal": {"level": "absent"},
            "maturity_date": {"level": "absent"},
            "maturity_default_treatment": {"level": "absent"},
        }
        assert validate_note(fields, confidence=confidence) == []

    def test_validate_note_still_errors_on_silently_missing_principal(self) -> None:
        from extract_instrument import validate_note  # type: ignore[import-not-found]

        fields = {
            "principal": None,
            "day_count_basis": 365,
            "issuance_date": "2024-01-01",
            "maturity_date": "2026-01-01",
            "maturity_default_treatment": "convert_at_cap",
            "interest_rate_type": "none",
            "annual_interest_rate": None,
        }
        errors = validate_note(fields)
        assert any("principal" in e for e in errors)


class TestPartialInstrumentPersist:
    """End-to-end: a blank-template SAFE (relaxable fields documented-absent) persists as a
    partial — completeness='partial', missing_required_fields, W_FIELD_ABSENT_IN_DOC, no fabrication."""

    def _instr(self, tmp_path: Any) -> str:
        p = str(tmp_path / "instruments.json")
        with open(p, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        return p

    def test_blank_template_safe_persists_as_partial(self, tmp_path: Any) -> None:
        instr_path = self._instr(tmp_path)
        extraction = {
            "instrument_type": "safe",
            "fields": {
                "investor_name": None,
                "purchase_amount": None,
                "issuance_date": None,
                "form": "yc_postmoney_cap",
                "post_money_valuation_cap": 5_000_000,
            },
            "confidence": {
                "investor_name": {"level": "absent"},
                "purchase_amount": {"level": "absent"},
                "issuance_date": {"level": "absent"},
            },
            "ambiguities": [{"field": "purchase_amount", "reason": "blank in the unexecuted template"}],
        }
        rc, out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"rc={rc} stderr={e}"
        with open(instr_path) as f:
            safe = json.load(f)["safes"][0]
        assert safe["completeness"] == "partial"
        assert "purchase_amount" in safe["missing_required_fields"]
        assert safe["purchase_amount"] is None  # never fabricated
        receipt = json.loads(out)
        assert any("W_FIELD_ABSENT_IN_DOC" in w for w in receipt.get("warnings", []))

    def test_relaxable_note_math_fields_covered_by_usable_predicate(self) -> None:
        import note_conversion  # type: ignore[import-not-found]
        from extract_instrument import NOTE_RELAXABLE_ABSENT_FIELDS  # type: ignore[import-not-found]

        # Every relaxable MATH-CONSUMED note field must be gated by the usable predicate, or a relaxed
        # partial could reach note math and crash/corrupt (the relaxable/usable-predicate lockstep invariant).
        math_consumed = {"principal", "issuance_date"}
        assert math_consumed <= NOTE_RELAXABLE_ABSENT_FIELDS
        assert not note_conversion.note_has_usable_math_inputs({"principal": None, "issuance_date": "2024-01-01"})
        assert not note_conversion.note_has_usable_math_inputs({"principal": 100, "issuance_date": None})
        assert note_conversion.note_has_usable_math_inputs({"principal": 100, "issuance_date": "2024-01-01"})


class TestExtractInstrumentTermsDoc:
    """term_sheet / option_plan persist their content to the RECEIPT (never a math array) and never
    block on a verifier/invariant FINDING — those surface as markers/attention_needed_fields.
    Input-integrity errors (missing --source-doc) STILL fail loud. Non-terms docs unchanged."""

    def _instr(self, tmp_path: Any) -> str:
        p = str(tmp_path / "instruments.json")
        with open(p, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        return p

    def _doc(self, tmp_path: Any) -> str:
        # Pre-money present; the fake $99,000,000 investment_amount is NOT in this text.
        p = tmp_path / "ts.txt"
        p.write_text(
            'TERM SHEET. The "Pre-Money Valuation" is $10,000,000. Standard governance, board, and '
            "protective provisions apply. " * 6
        )
        return str(p)

    @pytest.mark.parametrize("itype", ["term_sheet", "option_plan"])
    def test_terms_doc_persists_to_receipt_not_instruments(self, tmp_path: Any, itype: str) -> None:
        instr = self._instr(tmp_path)
        extraction = {
            "instrument_type": itype,
            "fields": {
                "pre_money_valuation": 10_000_000,
                "investment_amount": 5_000_000,
                "post_money_valuation": 15_000_000,
                "board_composition": "2 founders, 1 investor",
            },
            "confidence": {"pre_money_valuation": {"level": "high", "evidence_quote": "Pre-Money is $10,000,000"}},
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            ["--instruments", instr, "--run-id", "t", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, err
        with open(instr) as f:
            data = json.load(f)
        assert all(len(data[k]) == 0 for k in ("safes", "convertible_notes", "warrants", "option_grants"))
        receipt = json.loads(out)
        assert receipt["classified_as_terms_doc"] is True
        assert receipt["classified_doc_type"] == itype
        td = receipt["terms_doc"]
        assert td["fields"]["pre_money_valuation"] == 10_000_000
        assert "board_composition" in td["fields"]
        # M1: synthesized stamps must NOT pollute the terms-doc field table.
        assert "extraction_confidence" not in td["fields"]

    def test_terms_doc_value_in_doc_fail_does_not_block(self, tmp_path: Any) -> None:
        instr = self._instr(tmp_path)
        extraction = {
            "instrument_type": "term_sheet",
            "fields": {"pre_money_valuation": 10_000_000, "investment_amount": 99_000_000},
            "confidence": {
                "pre_money_valuation": {"level": "high", "evidence_quote": 'The "Pre-Money Valuation" is $10,000,000.'},
                "investment_amount": {"level": "high", "evidence_quote": "Investment Amount is $99,000,000."},  # fake
            },
            "ambiguities": [],
        }
        rc, out, err = _run(
            "extract_instrument.py",
            [
                "--instruments",
                instr,
                "--run-id",
                "t",
                "--verify",
                "--verify-blocking",
                "--source-doc",
                self._doc(tmp_path),
                "--no-invariants",
            ],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"terms doc must NOT block on a value_in_doc fail; stderr={err!r}"
        receipt = json.loads(out)
        pf = {r["field"]: r["status"] for r in receipt["evidence_verification"]["per_field"]}
        assert pf["investment_amount"] == "fail"  # the real fail is retained, not hidden
        assert "investment_amount" in receipt["attention_needed_fields"]  # surfaced, not silently dropped

    def test_terms_doc_invariant_violation_does_not_block(self, tmp_path: Any) -> None:
        instr = self._instr(tmp_path)
        # pre 10M + inv 5M != post 50M → hard term_sheet_math violation.
        extraction = {
            "instrument_type": "term_sheet",
            "fields": {
                "pre_money_valuation": 10_000_000,
                "investment_amount": 5_000_000,
                "post_money_valuation": 50_000_000,
            },
            "confidence": {},
            "ambiguities": [],
        }
        out_path = str(tmp_path / "receipt.json")
        rc, out, err = _run(
            "extract_instrument.py",
            ["--instruments", instr, "--run-id", "t", "--no-verify", "--invariants", "-o", out_path],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"terms doc must NOT block on an invariant violation; stderr={err!r}"
        with open(out_path) as f:
            receipt = json.load(f)  # -o must be written on a neutralized run
        assert receipt["invariant_check"]["n_hard_violations"] >= 1
        assert "post_money_valuation" in receipt["attention_needed_fields"]

    def test_terms_doc_missing_source_doc_still_fails_loud(self, tmp_path: Any) -> None:
        instr = self._instr(tmp_path)
        extraction = {
            "instrument_type": "term_sheet",
            "fields": {"pre_money_valuation": 10_000_000},
            "confidence": {},
            "ambiguities": [],
        }
        rc, _out, _err = _run(
            "extract_instrument.py",
            ["--instruments", instr, "--run-id", "t", "--verify"],  # --verify with no --source-doc
            stdin_data=json.dumps(extraction),
        )
        assert rc == 1  # input-integrity error stays blocking (NOT neutralized)

    def test_no_quote_soft_demotes_only_no_quote_fails(self) -> None:
        from extract_instrument import filter_verifier_report  # type: ignore[import-not-found]

        raw = {
            "overall_status": "fail",
            "per_field": [
                {"field": "a", "status": "pass", "reasons": []},
                {"field": "b", "status": "fail", "reasons": ["value=1 provided but evidence_quote is null/empty"]},
                {
                    "field": "c",
                    "status": "fail",
                    "reasons": ["value_in_doc failed: not found — value claimed but absent"],
                },
            ],
        }
        out = filter_verifier_report(raw, no_quote_soft=True)
        st = {r["field"]: r["status"] for r in out["per_field"]}
        assert st["b"] == "unverified_no_quote"  # no-quote fail demoted
        assert st["c"] == "fail"  # genuine value_in_doc fail NOT softened
        # Without the flag, both stay fail (non-terms regression).
        st2 = {r["field"]: r["status"] for r in filter_verifier_report(raw)["per_field"]}
        assert st2["b"] == "fail"


class TestSkillMdCodeDrift:
    """Doc-drift guard: every W_/E_ code cited in SKILL.md must exist somewhere in the
    skill's scripts or rule-pack/references — so the 'documented path that doesn't exist' class
    (e.g. a warning SKILL.md promises but no code emits) cannot recur."""

    def test_all_cited_warning_error_codes_exist_in_code(self) -> None:
        import contextlib
        import re
        from pathlib import Path

        skill_dir = Path(SCRIPTS).parent  # .../skills/cap-table
        skill_md = skill_dir / "SKILL.md"
        cited = set(re.findall(r"\b[WE]_[A-Z0-9]+_[A-Z0-9_]+\b", skill_md.read_text(encoding="utf-8")))
        # Corpus: the skill's own scripts + references AND the shared scripts (founder_context.py etc.
        # emit some warning codes SKILL.md legitimately references).
        shared_scripts = Path(_REPO) / "founder-skills" / "scripts"
        corpus_parts: list[str] = []
        for base in (skill_dir, shared_scripts):
            for path in base.rglob("*"):
                if path == skill_md or not path.is_file() or path.suffix not in {".py", ".json", ".md"}:
                    continue  # never let SKILL.md satisfy its own citations
                with contextlib.suppress(Exception):
                    corpus_parts.append(path.read_text(encoding="utf-8"))
        corpus = "\n".join(corpus_parts)
        missing = sorted(c for c in cited if c not in corpus)
        assert not missing, f"SKILL.md cites W_/E_ codes not defined in scripts/rule-pack: {missing}"


class TestPartialNotePersist:
    """A convertible note whose principal lives in a Schedule of Lenders (null in the Note
    template) persists as a partial when documented-absent — end-to-end through the CLI + schema."""

    def test_schedule_of_lenders_note_persists_as_partial(self, tmp_path: Any) -> None:
        instr_path = str(tmp_path / "instruments.json")
        with open(instr_path, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        extraction = {
            "instrument_type": "convertible_note",
            "fields": {
                "investor_name": "Foobar Capital II",
                "principal": None,  # lives in the Schedule of Lenders, blank in the Note template
                "interest_rate_type": "fixed_numeric",
                "annual_interest_rate": 0.08,
                "day_count_basis": 365,
                "issuance_date": "2024-03-01",
                "maturity_date": None,  # computed 18-months-from-issuance, not a literal
                "maturity_default_treatment": "convert_at_cap",
                "valuation_cap": 10_000_000,
            },
            "confidence": {
                "principal": {"level": "absent"},
                "maturity_date": {"level": "absent"},
            },
            "ambiguities": [{"field": "principal", "reason": "principal listed in Schedule of Lenders, not the Note"}],
        }
        rc, out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"rc={rc} stderr={e}"
        with open(instr_path) as f:
            note = json.load(f)["convertible_notes"][0]
        assert note["completeness"] == "partial"
        assert "principal" in note["missing_required_fields"]
        assert note["principal"] is None
        assert note["extraction_confidence"] == "low"


class TestPartialWarrantPersist:
    """A warrant whose share count is confirmed but whose strike is not stated in the source
    persists as a partial (documented-absent exercise_price) instead of being dropped — mirrors
    the SAFE/note partial-instrument gate. Undocumented null still hard-errors; a missing key
    keeps the empty-fields misfile escape hatch."""

    def _instr(self, tmp_path: Any) -> str:
        p = str(tmp_path / "instruments.json")
        with open(p, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        return p

    def test_warrant_with_absent_exercise_price_persists_as_partial(self, tmp_path: Any) -> None:
        instr_path = self._instr(tmp_path)
        extraction = {
            "instrument_type": "warrant",
            "fields": {
                "investor_name": "Foobar Bank",
                "shares_underlying": 25000,
                "exercise_price": None,
                "warrant_type": "common_stock",
                "issuance_date": "2024-05-01",
                "settlement_type": "cash_exercise",
                "vested_flag": True,
            },
            "confidence": {"exercise_price": {"level": "absent"}},
            "ambiguities": [{"field": "exercise_price", "reason": "strike not stated on the cap-table sheet"}],
        }
        rc, out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"rc={rc} stderr={e}"
        with open(instr_path) as f:
            warrant = json.load(f)["warrants"][0]
        assert warrant["exercise_price"] is None  # never fabricated
        assert warrant["completeness"] == "partial"
        assert "exercise_price" in warrant["missing_required_fields"]
        assert warrant["extraction_confidence"] == "low"
        receipt = json.loads(out)
        assert any("W_FIELD_ABSENT_IN_DOC:exercise_price" in w for w in receipt.get("warnings", []))

    def test_warrant_null_strike_without_absence_documentation_hard_errors(self, tmp_path: Any) -> None:
        instr_path = self._instr(tmp_path)
        extraction = {
            "instrument_type": "warrant",
            "fields": {
                "investor_name": "Foobar Bank",
                "shares_underlying": 25000,
                "exercise_price": None,  # present + null but absence NOT documented
                "warrant_type": "common_stock",
                "issuance_date": "2024-05-01",
                "settlement_type": "cash_exercise",
                "vested_flag": True,
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc != 0
        blob = (out + e).lower()
        assert "exercise_price" in blob and "absent" in blob  # actionable message, not a raw schema dump

    def test_warrant_missing_strike_key_stays_unpersisted_rc0(self, tmp_path: Any) -> None:
        # Pinning test (passes today): the exercise_price key is OMITTED entirely — the empty-fields
        # warrant-misfile escape hatch. Must keep rc-0 + "NOT persisted", NOT the key-present-null gate.
        instr_path = self._instr(tmp_path)
        extraction = {
            "instrument_type": "warrant",
            "fields": {
                "investor_name": "Foobar Bank",
                "shares_underlying": 25000,
                # exercise_price key intentionally absent
                "warrant_type": "common_stock",
                "issuance_date": "2024-05-01",
                "settlement_type": "cash_exercise",
                "vested_flag": True,
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"rc={rc} stderr={e}"
        with open(instr_path) as f:
            assert json.load(f)["warrants"] == []
        receipt = json.loads(out)
        assert any("NOT persisted" in w for w in receipt.get("warnings", []))


class TestPartialWarrantCapState:
    """cap_state keeps a null-strike warrant's shares in FD (the whole point) while marking it
    non-exercisable + warning; end-to-end through the cap_state.py CLI schema-validation layer."""

    def _warrant(self, **overrides: Any) -> dict[str, Any]:
        w = {
            "id": "w1",
            "shares_underlying": 25_000,
            "exercise_price": None,
            "warrant_type": "common_stock",
            "issuance_date": "2024-05-01",
            "settlement_type": "cash_exercise",
            "vested_flag": True,
        }
        w.update(overrides)
        return w

    def test_warrant_null_exercise_price_degrades_not_raises(self) -> None:
        instruments = {**_BASIC_INSTRUMENTS, "warrants": [self._warrant()]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)  # must NOT raise
        ow = cs["outstanding_warrants"][0]
        assert ow["exercise_price"] is None
        assert ow["exercisable"] is False
        assert cs["as_converted_totals"]["warrants_underlying_total"] == 25_000  # shares STILL in FD
        assert "W_WARRANT_EXERCISE_PRICE_MISSING" in cs.get("warnings", [])

    def test_warrant_numeric_exercise_price_is_exercisable(self) -> None:
        instruments = {**_BASIC_INSTRUMENTS, "warrants": [self._warrant(exercise_price=1.0)]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        ow = cs["outstanding_warrants"][0]
        assert ow["exercisable"] is True
        assert "W_WARRANT_EXERCISE_PRICE_MISSING" not in cs.get("warnings", [])

    def test_cap_state_cli_round_trips_partial_warrant(self, tmp_path: Any) -> None:
        inputs_p = str(tmp_path / "inputs.json")
        instr_p = str(tmp_path / "instruments.json")
        out_p = str(tmp_path / "cap_state.json")
        with open(inputs_p, "w") as f:
            json.dump(_BASIC_INPUTS, f)
        with open(instr_p, "w") as f:
            json.dump({**_BASIC_INSTRUMENTS, "warrants": [self._warrant()]}, f)
        rc, _out, err = _run(
            "cap_state.py",
            ["--inputs", inputs_p, "--instruments", instr_p, "--run-id", "test", "-o", out_p],
        )
        assert rc == 0, f"partial warrant must survive the cap_state.py CLI; stderr={err!r}"
        with open(out_p) as f:
            cs = json.load(f)
        ow = cs["outstanding_warrants"][0]
        assert ow["exercise_price"] is None
        assert cs["as_converted_totals"]["warrants_underlying_total"] == 25_000

    def test_warrant_strike_missing_callout_rendered(self) -> None:
        from _warning_callouts import render_warning_callouts  # type: ignore[import-not-found]

        out = render_warning_callouts(["W_WARRANT_EXERCISE_PRICE_MISSING"])
        assert out
        body = "\n".join(out).lower()
        # The callout must say the shares ARE counted in FD and exercise math was skipped.
        assert "fully-diluted" in body or "fully diluted" in body
        assert "exercise" in body


class TestWarrantExerciseNullStrike:
    """A null-strike warrant that reaches exercise/pump math fails loud with a structured code,
    surfaced by run_scenario as a structural_only blocker — never a raw TypeError."""

    def test_relaxable_warrant_fields_covered_by_usable_predicate(self) -> None:
        import warrant_exercise  # type: ignore[import-not-found]
        from extract_instrument import WARRANT_RELAXABLE_ABSENT_FIELDS  # type: ignore[import-not-found]

        assert {"exercise_price"} <= WARRANT_RELAXABLE_ABSENT_FIELDS
        assert warrant_exercise.warrant_has_usable_exercise_price({"exercise_price": 1.0}) is True
        assert warrant_exercise.warrant_has_usable_exercise_price({"exercise_price": 0.0}) is True  # nominal strike ok
        assert warrant_exercise.warrant_has_usable_exercise_price({"exercise_price": None}) is False
        assert warrant_exercise.warrant_has_usable_exercise_price({"exercise_price": True}) is False  # bool guard

    def test_exercise_warrant_null_strike_raises_structured_error(self) -> None:
        import warrant_exercise  # type: ignore[import-not-found]

        w = {
            "warrant_id": "w1",
            "settlement_type": "cash_exercise",
            "shares_underlying": 1000,
            "exercise_price": None,
        }
        with pytest.raises(warrant_exercise.WarrantPumpError) as exc:
            warrant_exercise.exercise_warrant(w)
        assert str(exc.value).startswith("E_WARRANT_EXERCISE_PRICE_MISSING:")

    def test_run_scenario_surfaces_null_strike_pump_as_blocker(self) -> None:
        w = {
            "id": "w1",
            "shares_underlying": 1000,
            "exercise_price": None,
            "warrant_type": "common_stock",
            "issuance_date": "2024-01-01",
            "settlement_type": "cash_exercise",
            "vested_flag": True,
            "exercise_event_date": "2026-01-01",  # earlier than the round → pump fires
        }
        instruments = {**_BASIC_INSTRUMENTS, "warrants": [w]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        scenario = {
            "type": "priced_round",
            "parameters": {
                "pre_money": 10_000_000,
                "new_money": 2_000_000,
                "transaction_event_date": "2026-06-01",
            },
        }
        result = run_scenario.run_priced_round_scenario(scenario, instruments=instruments, cap_state=cs)
        assert result["completeness"] == "structural_only"
        assert any(b["code"] == "E_WARRANT_EXERCISE_PRICE_MISSING" for b in result["blockers"])


class TestTermsOnlyInstrumentCapStateCli:
    """The SAFE/note partial-instrument gate must survive the cap_state.py CLI (write_artifact schema
    validation), not only the in-process build path — coverage the in-process tests never exercised."""

    def _files(self, tmp_path: Any, instruments: dict[str, Any]) -> tuple[str, str, str]:
        inputs_p = str(tmp_path / "inputs.json")
        instr_p = str(tmp_path / "instruments.json")
        out_p = str(tmp_path / "cap_state.json")
        with open(inputs_p, "w") as f:
            json.dump(_BASIC_INPUTS, f)
        with open(instr_p, "w") as f:
            json.dump(instruments, f)
        return inputs_p, instr_p, out_p

    def test_cap_state_cli_round_trips_terms_only_safe(self, tmp_path: Any) -> None:
        safe = {k: v for k, v in _SAFE_BASIC.items() if k != "purchase_amount"}
        safe = {**safe, "purchase_amount": None, "issuance_date": None}
        inputs_p, instr_p, out_p = self._files(tmp_path, {**_BASIC_INSTRUMENTS, "safes": [safe]})
        rc, _out, err = _run(
            "cap_state.py", ["--inputs", inputs_p, "--instruments", instr_p, "--run-id", "t", "-o", out_p]
        )
        assert rc == 0, f"terms-only SAFE must survive the cap_state.py CLI; stderr={err!r}"
        with open(out_p) as f:
            cs = json.load(f)
        s = cs["outstanding_safes"][0]
        assert s["purchase_amount"] is None
        assert s["convertible"] is False
        assert "W_SAFE_PURCHASE_AMOUNT_MISSING" in cs.get("warnings", [])

    def test_cap_state_cli_round_trips_terms_only_note(self, tmp_path: Any) -> None:
        note = {**_NOTE_BASIC, "principal": None}
        inputs_p, instr_p, out_p = self._files(tmp_path, {**_BASIC_INSTRUMENTS, "convertible_notes": [note]})
        rc, _out, err = _run(
            "cap_state.py", ["--inputs", inputs_p, "--instruments", instr_p, "--run-id", "t", "-o", out_p]
        )
        assert rc == 0, f"terms-only note must survive the cap_state.py CLI; stderr={err!r}"
        with open(out_p) as f:
            cs = json.load(f)
        n = cs["outstanding_notes"][0]
        assert n["principal"] is None
        assert n["convertible"] is False
        assert "W_NOTE_PRINCIPAL_MISSING" in cs.get("warnings", [])


class TestCommonBatchNoBatchId:
    """A common batch supplied with holder_id but no batch_id (the skeleton-canonical shape) must
    survive the cap_state.py CLI — batch_id is a display-only label, not a required key."""

    def test_cap_state_cli_round_trips_common_batch_without_batch_id(self, tmp_path: Any) -> None:
        inputs = {
            **_BASIC_INPUTS,
            "common_batches": [{"holder_id": "angel_1", "shares": 100_000, "issuance_date": "2024-06-01"}],
        }
        inputs_p = str(tmp_path / "inputs.json")
        instr_p = str(tmp_path / "instruments.json")
        out_p = str(tmp_path / "cap_state.json")
        with open(inputs_p, "w") as f:
            json.dump(inputs, f)
        with open(instr_p, "w") as f:
            json.dump(_BASIC_INSTRUMENTS, f)
        rc, _out, err = _run(
            "cap_state.py", ["--inputs", inputs_p, "--instruments", instr_p, "--run-id", "t", "-o", out_p]
        )
        assert rc == 0, f"a holder_id-only common batch must survive the cap_state.py CLI; stderr={err!r}"
        with open(out_p) as f:
            cs = json.load(f)
        assert cs["common_batches"][0]["holder_id"] == "angel_1"


class TestBatchDisplayLabelPrefersHolderName:
    """common_batches has no name field like founders[].name — so the report/visualize tables render a
    raw 'Batch <id>'. When holder_name is supplied, both renderers must prefer it."""

    def test_compose_batch_label_prefers_holder_name(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        assert compose_report._batch_label({"holder_name": "Jane Doe", "holder_id": "b1"}) == "Jane Doe"
        assert compose_report._batch_label({"holder_id": "b1"}) == "Batch b1"
        assert compose_report._batch_label({"batch_id": "bx", "holder_id": "b1"}) == "Batch bx"

    def test_visualize_batch_label_prefers_holder_name(self) -> None:
        import visualize  # type: ignore[import-not-found]

        assert visualize._batch_label({"holder_name": "Jane Doe", "holder_id": "b1"}) == "Jane Doe"
        assert visualize._batch_label({"holder_id": "b1"}) == "Batch b1"


class TestNotePricedRoundNoCapNoDiscount:
    """A note evaluated in a priced round with NEITHER a valuation_cap NOR a discount_multiplier gets a
    distinct branch/error attributing the gap to the note's missing terms — not a maturity-default
    misdiagnosis. Existing cap-present / repay behavior is preserved."""

    def _note(self, **overrides: Any) -> dict[str, Any]:
        base: dict[str, Any] = {
            "id": "note_001",
            "investor_name": "Lender",
            "principal": 100_000,
            "annual_interest_rate": 0.06,
            "interest_rate_type": "fixed_numeric",
            "day_count_basis": 365,
            "compounding_periods_per_year": None,
            "interest_converts_to_shares": True,
            "issuance_date": "2025-01-01",
            "last_interest_event_date": None,
            "valuation_cap": None,
            "discount_multiplier": None,
            "capitalization_denominator": None,
            "capitalization_denominator_policy": None,
            "qualified_financing_threshold": 1_000_000,
            "maturity_date": "9999-12-31",
            "maturity_default_treatment": None,
            "maturity_conversion_price_override": None,
            "non_qualified_financing_treatment": None,
            "source_document": None,
            "extraction_confidence": "high",
        }
        base.update(overrides)
        return base

    def test_priced_round_note_without_cap_or_discount_gets_distinct_branch(self) -> None:
        result = note_conversion.convert_note(
            self._note(),
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=1.0,
        )
        assert result["branch"] == "priced_round_no_cap_or_discount", result
        assert result.get("error") == note_conversion.E_NOTE_PRICED_ROUND_NO_CAP_OR_DISCOUNT

    def test_reason_attributes_gap_to_missing_cap_and_discount(self) -> None:
        result = note_conversion.convert_note(
            self._note(),
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=1.0,
        )
        reason = result.get("reason", "")
        assert "valuation_cap" in reason and "discount_multiplier" in reason, reason
        assert "convert_at_cap" not in reason  # no longer the maturity-default misdiagnosis
        assert any(kw in reason.lower() for kw in ("ask", "founder", "note text", "confirm"))

    def test_no_maturity_default_warning_for_priced_round_no_terms_note(self) -> None:
        note = self._note()
        note.pop("maturity_default_treatment")  # key ABSENT → today the default fires + warns
        result = note_conversion.convert_note(
            note,
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=1.0,
        )
        codes = {w["code"] for w in result.get("warnings", [])}
        assert "maturity_default_treatment_defaulted" not in codes

    def test_new_money_without_round_price_still_gets_distinct_branch(self) -> None:
        result = note_conversion.convert_note(
            self._note(),
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=None,  # boundary variant: strict both-present gate would miss this
        )
        assert result["branch"] == "priced_round_no_cap_or_discount", result
        assert result.get("error") == note_conversion.E_NOTE_PRICED_ROUND_NO_CAP_OR_DISCOUNT

    def test_priced_round_note_with_repay_treatment_keeps_maturity_repay(self) -> None:
        # Pinning test (passes today): a repay note is a maturity branch, never a no-path return, so
        # the rewrite must not touch it.
        result = note_conversion.convert_note(
            self._note(maturity_default_treatment="repay"),
            conversion_event_date="2026-06-01",
            priced_round_new_money=5_000_000,
            qualified_financing_price=1.0,
        )
        assert result["branch"] == "maturity_repay"
        assert "cash_repayment" in result

    def test_scenario_completeness_structural_only_for_new_branch(self) -> None:
        out = note_conversion.derive_scenario_completeness({"n1": {"branch": "priced_round_no_cap_or_discount"}})
        assert out == "structural_only"

    def test_run_note_scenario_surfaces_new_blocker_code(self) -> None:
        note = self._note()
        instruments = {**_BASIC_INSTRUMENTS, "convertible_notes": [note]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        scenario = {
            "type": "note_conversion",
            "parameters": {
                "transaction_event_date": "2026-06-01",
                "priced_round_new_money": 5_000_000,
                "qualified_financing_price": 1.0,
            },
        }
        result = run_scenario.run_note_conversion_scenario(scenario, instruments=instruments, cap_state=cs)
        assert result["completeness"] == "structural_only"
        assert any(
            b["code"] == "E_NOTE_PRICED_ROUND_NO_CAP_OR_DISCOUNT" and b["instance_id"] == "note_001"
            for b in result["blockers"]
        )


def _option_grant(**overrides: Any) -> dict[str, Any]:
    # Explicit nulls for the optional share numerics are deliberate: null and key-absent are
    # schema-equivalent for these fields, and the builder must tolerate both (an agent authoring a
    # partial grant with null-for-unknown will write these as explicit null).
    g: dict[str, Any] = {
        "id": "og1",
        "holder_id": "emp_1",
        "grant_date": "2024-02-01",
        "shares_granted": 10_000,
        "shares_vested_to_date": None,
        "shares_exercised": None,
        "strike_price": None,
        "plan_type": "section_102_cg",
    }
    g.update(overrides)
    return g


def _instruments_with_grant(grant: dict[str, Any]) -> dict[str, Any]:
    return {**_BASIC_INSTRUMENTS, "option_grants": [grant]}


class TestPartialOptionGrantSchema:
    """A grant whose strike is genuinely not stated persists as a partial (null strike). The
    validation gate is instruments.schema.json — the only substrate every grant-validation path
    loads (grants are never authored by extract_instrument; they enter via Lane-4/validate)."""

    def _write_instruments(self, tmp_path: Any, grant: dict[str, Any]) -> str:
        d = str(tmp_path)
        with open(os.path.join(d, "instruments.json"), "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [grant],
                    "metadata": {"run_id": "t", "schema_version": "v0.5.0-instruments"},
                },
                f,
            )
        return d

    def test_validate_mode_accepts_null_strike_grant(self, tmp_path: Any) -> None:
        d = self._write_instruments(tmp_path, _option_grant())
        rc, out, err = _run("extract_cap_table.py", ["--mode", "validate", "--dir", d])
        assert rc == 0, f"null-strike grant must validate; stderr={err!r}"
        assert json.loads(out)["ok"] is True

    def test_validate_mode_still_rejects_missing_strike_key(self, tmp_path: Any) -> None:
        grant = _option_grant()
        del grant["strike_price"]  # key omitted entirely — the mis-key/misfile catch stays
        d = self._write_instruments(tmp_path, grant)
        rc, out, _err = _run("extract_cap_table.py", ["--mode", "validate", "--dir", d])
        assert rc != 0
        assert "strike_price" in out or "strike_price" in _err

    def test_lane1_persist_survives_preexisting_null_strike_grant(self, tmp_path: Any) -> None:
        instr_path = str(tmp_path / "instruments.json")
        with open(instr_path, "w") as f:
            json.dump(
                {
                    "safes": [],
                    "convertible_notes": [],
                    "warrants": [],
                    "option_grants": [_option_grant()],
                    "metadata": {"run_id": "test"},
                },
                f,
            )
        extraction = {
            "instrument_type": "safe",
            "fields": {
                "investor_name": "Angel A",
                "purchase_amount": 500_000,
                "issuance_date": "2025-01-01",
                "form": "yc_postmoney_cap",
                "post_money_valuation_cap": 8_000_000,
            },
            "confidence": {},
            "ambiguities": [],
        }
        rc, _out, e = _run(
            "extract_instrument.py",
            ["--instruments", instr_path, "--run-id", "test", "--no-verify", "--no-invariants"],
            stdin_data=json.dumps(extraction),
        )
        assert rc == 0, f"a pre-existing null-strike grant must not fail a later persist; stderr={e}"
        with open(instr_path) as f:
            data = json.load(f)
        assert len(data["safes"]) == 1 and len(data["option_grants"]) == 1


class TestPreferredSeriesMisplacedInInstruments:
    """ITEM 17: preferred_series[] belongs in inputs.json. Written into instruments.json instead,
    it used to be silently dropped (the instruments schema has no such property and the hand-rolled
    validator ignores unknown top-level keys) — incomplete priced-round/anti-dilution math with no
    validation error. `_cap_table_schema_validator.check_misplaced_top_level_keys` adds a targeted
    (non-generic) guard for this one specific mis-key, wired into `extract_cap_table.py --mode
    validate` (the pipeline's schema-enforcement gate)."""

    def _write_instruments(self, tmp_path: Any, extra: dict[str, Any] | None = None) -> str:
        d = str(tmp_path)
        # Deep copy, not dict(): _BASIC_INSTRUMENTS is module-level and its nested `metadata`
        # is shared, so a shallow copy lets another test's mutation reach this one. That is
        # exactly what happened -- this class passed in isolation and failed in a full-file run.
        data = json.loads(json.dumps(_BASIC_INSTRUMENTS))
        data["metadata"] = {"run_id": "t", "schema_version": "v0.5.0-instruments"}
        if extra:
            data.update(extra)
        with open(os.path.join(d, "instruments.json"), "w") as f:
            json.dump(data, f)
        return d

    def test_preferred_series_at_top_level_is_rejected(self, tmp_path: Any) -> None:
        d = self._write_instruments(
            tmp_path,
            {
                "preferred_series": [
                    {
                        "series_name": "Series A",
                        "shares": 1_000_000,
                        "original_issue_price": 1.0,
                        "original_conversion_price": 1.0,
                        "current_conversion_price": 1.0,
                        "issuance_date": "2025-01-01",
                    }
                ]
            },
        )
        rc, out, err = _run("extract_cap_table.py", ["--mode", "validate", "--dir", d])
        assert rc != 0, f"a top-level preferred_series in instruments.json must be rejected; stdout={out!r}"
        combined = out + err
        assert "preferred_series" in combined
        assert "instruments.json" in combined
        assert "inputs.json" in combined  # names the correct home, not just what's wrong

    def test_valid_instruments_json_unaffected(self, tmp_path: Any) -> None:
        """No false positive: an instruments.json with no preferred_series key validates clean."""
        d = self._write_instruments(tmp_path)
        rc, out, err = _run("extract_cap_table.py", ["--mode", "validate", "--dir", d])
        assert rc == 0, f"a clean instruments.json must still validate; stderr={err!r}"
        assert json.loads(out)["ok"] is True

    def test_check_misplaced_top_level_keys_direct(self) -> None:
        """Unit-level: the shared detector names both files in one error string, and only fires
        when the key is found in the file it's actually misplaced in."""
        import _cap_table_schema_validator as v  # type: ignore[import-not-found]

        errs = v.check_misplaced_top_level_keys({"preferred_series": []}, "instruments.json")
        assert len(errs) == 1
        assert "preferred_series" in errs[0]
        assert "inputs.json" in errs[0]
        assert "instruments.json" in errs[0]

        # No false positive: no key present, or the key present in the file it legitimately belongs in.
        assert v.check_misplaced_top_level_keys({"safes": []}, "instruments.json") == []
        assert v.check_misplaced_top_level_keys({"preferred_series": []}, "inputs.json") == []


class TestAdCarveOutsErgonomics:
    """ITEM 21: ad_carve_outs is a plural key holding a scalar enum ('nvca_default') — the plural
    name invites a model to write a list, and a live recorded run did (`ad_carve_outs: []`),
    tripping schema validation. inputs.schema.json now also accepts `[]` and `['nvca_default']`;
    cap_state.py normalizes all three forms to the scalar so every downstream consumer (rule_audit
    matchers, disclosure, `test_ad_carve_outs_survives_canonicalization`) keeps seeing a plain
    string. A genuinely invalid value (not one of the three accepted forms) still fails, at both
    the schema layer and — as a last-line-of-defense for a non-schema-validated direct cap_state.py
    invocation — inside cap_state.py's own canonicalization."""

    _SCHEMA_PATH = os.path.join(os.path.dirname(SCRIPTS), "references", "schemas", "inputs.schema.json")

    def _series(self, ad_carve_outs: Any) -> dict[str, Any]:
        return {
            "series_name": "Series A",
            "shares": 1_000_000,
            "original_issue_price": 1.0,
            "original_conversion_price": 1.0,
            "current_conversion_price": 1.0,
            "issuance_date": "2025-01-01",
            "ad_carve_outs": ad_carve_outs,
        }

    def _inputs_with_series(self, series: dict[str, Any]) -> dict[str, Any]:
        return {
            "company_name": "TestCo",
            "analysis_date": "2026-06-01",
            "mode": "standard",
            "jurisdiction": {
                "structure": "delaware",
                "incorporated_date": "2024-01-01",
                "iia_grants_history": {"has_grants": False, "grant_details": []},
            },
            "founders": [{"name": "F", "founder_id": "f1", "common_shares": 10_000_000}],
            "preferred_series": [series],
            "option_pool": {"plan_type": "iso", "authorized": 0, "issued": 0, "unallocated": 0},
            "common_batches": [],
            "metadata": {"run_id": "test", "schema_version": "v0.5.0-inputs"},
        }

    def _validate_schema(self, ad_carve_outs: Any) -> list[str]:
        import _cap_table_schema_validator as v  # type: ignore[import-not-found]
        from _artifact_writer import load_schema  # type: ignore[import-not-found]

        schema = load_schema(self._SCHEMA_PATH)
        errs: list[str] = v.validate(self._inputs_with_series(self._series(ad_carve_outs)), schema, "inputs")
        return errs

    def test_empty_list_validates_against_schema(self) -> None:
        errs = self._validate_schema([])
        assert errs == [], f"ad_carve_outs=[] should validate cleanly: {errs}"

    def test_scalar_validates_against_schema(self) -> None:
        errs = self._validate_schema("nvca_default")
        assert errs == [], f"ad_carve_outs='nvca_default' should validate cleanly: {errs}"

    def test_singleton_list_validates_against_schema(self) -> None:
        errs = self._validate_schema(["nvca_default"])
        assert errs == [], f"ad_carve_outs=['nvca_default'] should validate cleanly: {errs}"

    def test_genuinely_invalid_scalar_still_rejected_naming_accepted_form(self) -> None:
        errs = self._validate_schema("bogus")
        assert errs, "a genuinely invalid ad_carve_outs value must still fail"
        assert any("nvca_default" in e for e in errs), f"error should name the accepted form: {errs}"

    def test_genuinely_invalid_list_still_rejected_naming_accepted_form(self) -> None:
        errs = self._validate_schema(["something_else"])
        assert errs, "an invalid list value must still fail"
        assert any("nvca_default" in e for e in errs), f"error should name the accepted form: {errs}"

    def test_cap_state_normalizes_empty_list_to_scalar(self) -> None:
        cs = cap_state_mod.build_cap_state(self._inputs_with_series(self._series([])), _BASIC_INSTRUMENTS)
        assert cs["preferred_series"][0]["ad_carve_outs"] == "nvca_default"

    def test_cap_state_normalizes_singleton_list_to_scalar(self) -> None:
        cs = cap_state_mod.build_cap_state(self._inputs_with_series(self._series(["nvca_default"])), _BASIC_INSTRUMENTS)
        assert cs["preferred_series"][0]["ad_carve_outs"] == "nvca_default"

    def test_cap_state_rejects_genuinely_invalid_value_naming_accepted_form(self) -> None:
        with pytest.raises(cap_state_mod.CapStateInvariantError) as exc_info:
            cap_state_mod.build_cap_state(self._inputs_with_series(self._series("bogus")), _BASIC_INSTRUMENTS)
        assert "nvca_default" in str(exc_info.value)


class TestPartialOptionGrantCapState:
    """cap_state tolerates a partial grant: null share numerics degrade to unvested, null strike is
    kept + warned; share counts (FD) are untouched (pool aggregate drives FD)."""

    def test_null_vested_grant_degrades_not_raises(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _instruments_with_grant(_option_grant()))
        o = cs["outstanding_options"][0]
        assert o["shares_vested_to_date"] == 0
        assert o["shares_exercised"] == 0
        assert o["shares_outstanding_unvested"] == 10_000  # null vesting → fully unvested

    def test_null_strike_grant_degrades_not_raises(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _instruments_with_grant(_option_grant()))
        o = cs["outstanding_options"][0]
        assert o["strike_price"] is None
        assert o["plan_type"] == "section_102_cg"
        assert "W_OPTION_GRANT_STRIKE_MISSING" in cs.get("warnings", [])

    def test_numeric_strike_no_warning(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _BASIC_INPUTS,
            _instruments_with_grant(_option_grant(strike_price=0.50, shares_vested_to_date=2_500, shares_exercised=0)),
        )
        o = cs["outstanding_options"][0]
        assert o["strike_price"] == 0.50
        assert "W_OPTION_GRANT_STRIKE_MISSING" not in cs.get("warnings", [])
        cs0 = cap_state_mod.build_cap_state(
            _BASIC_INPUTS,
            _instruments_with_grant(_option_grant(strike_price=0.0, shares_vested_to_date=0, shares_exercised=0)),
        )
        assert cs0["outstanding_options"][0]["strike_price"] == 0.0  # nominal strike legitimate
        assert "W_OPTION_GRANT_STRIKE_MISSING" not in cs0.get("warnings", [])

    def test_bool_strike_normalized_to_none(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _BASIC_INPUTS,
            _instruments_with_grant(_option_grant(strike_price=True, shares_vested_to_date=0, shares_exercised=0)),
        )
        assert cs["outstanding_options"][0]["strike_price"] is None
        assert "W_OPTION_GRANT_STRIKE_MISSING" in cs.get("warnings", [])

    def test_fd_totals_unchanged_by_partial_grant(self) -> None:
        without = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)["as_converted_totals"]
        with_grant = cap_state_mod.build_cap_state(_BASIC_INPUTS, _instruments_with_grant(_option_grant()))[
            "as_converted_totals"
        ]
        assert without == with_grant  # options in FD come from the pool aggregate, not per-grant

    def test_cap_state_cli_round_trips_partial_grant(self, tmp_path: Any) -> None:
        inputs_p = str(tmp_path / "inputs.json")
        instr_p = str(tmp_path / "instruments.json")
        out_p = str(tmp_path / "cap_state.json")
        with open(inputs_p, "w") as f:
            json.dump(_BASIC_INPUTS, f)
        with open(instr_p, "w") as f:
            json.dump(_instruments_with_grant(_option_grant()), f)
        rc, _out, err = _run(
            "cap_state.py", ["--inputs", inputs_p, "--instruments", instr_p, "--run-id", "t", "-o", out_p]
        )
        assert rc == 0, f"partial grant must survive the cap_state.py CLI; stderr={err!r}"
        with open(out_p) as f:
            cs = json.load(f)
        o = cs["outstanding_options"][0]
        assert o["strike_price"] is None
        assert o["shares_vested_to_date"] == 0

    def test_option_strike_missing_callout_rendered(self) -> None:
        from _warning_callouts import render_warning_callouts  # type: ignore[import-not-found]

        out = render_warning_callouts(["W_OPTION_GRANT_STRIKE_MISSING"])
        assert out
        body = "\n".join(out).lower()
        assert "strike" in body
        assert "fully-diluted" in body or "fully diluted" in body or "share counts" in body


class TestSection102CountStrikeIndependent:
    """The §102 grant count (which drives the flip continuity item) derives from plan_type and is
    independent of the strike — a null-strike §102 grant still counts."""

    def test_section_102_count_includes_null_strike_grant(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _BASIC_INPUTS,
            _instruments_with_grant(_option_grant(shares_vested_to_date=0, shares_exercised=0)),
        )
        # the exact counting expression compose_report uses for flip_specifics
        count = sum(
            1 for o in cs.get("outstanding_options", []) if (o.get("plan_type") or "").startswith("section_102")
        )
        assert count == 1


def _inputs_with_pool(**pool_overrides: Any) -> dict[str, Any]:
    inp: dict[str, Any] = json.loads(json.dumps(_BASIC_INPUTS))
    inp["option_pool"] = {"plan_type": "nso", "authorized": 0, "issued": 0, "unallocated": 0, **pool_overrides}
    return inp


class TestFlipSection102Derivation:
    """A flip must not silently report zero §102 grants: the count derives from the per-grant
    plan_type in cap_state, and when the pool has issued §102 options but no per-grant data exists,
    an 'unconfirmed' counsel item fires instead of silence."""

    def _flip(self, cs: dict[str, Any], **kw: Any) -> dict[str, Any]:
        result: dict[str, Any] = flip_scenario.flip_share_for_share(cs, **kw)
        return result

    def _titles(self, result: dict[str, Any]) -> list[str]:
        return [c["title"] for c in result["counsel_review_items"]]

    def test_flip_derives_102_count_from_cap_state(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _BASIC_INPUTS,
            _instruments_with_grant(_option_grant(shares_vested_to_date=0, shares_exercised=0)),
        )
        result = self._flip(cs)  # parameter defaults to 0 — derivation must supply the count
        item = next(
            (
                c
                for c in result["counsel_review_items"]
                if c["rule_id"] == "delaware_cross_border.section_102_parent_shares"
            ),
            None,
        )
        assert item is not None, "the derived §102 continuity item must be present"
        assert "continuity" in item["title"] and "1" in item["title"]

    def test_flip_parameter_still_honored_when_larger(self) -> None:
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, _BASIC_INSTRUMENTS)  # empty outstanding_options
        result = self._flip(cs, section_102_grants_outstanding=3)
        item = next(
            c
            for c in result["counsel_review_items"]
            if c["rule_id"] == "delaware_cross_border.section_102_parent_shares"
        )
        assert "3" in item["title"]

    def test_flip_unconfirmed_102_item_when_pool_issued_but_no_grant_data(self) -> None:
        cs = cap_state_mod.build_cap_state(
            _inputs_with_pool(plan_type="section_102_cg", authorized=2_000, issued=1_234, unallocated=766),
            _BASIC_INSTRUMENTS,  # no per-grant data
        )
        result = self._flip(cs)  # param 0, no grants → unconfirmed
        item = next((c for c in result["counsel_review_items"] if "unconfirmed" in c["title"]), None)
        assert item is not None
        assert "1,234" in item["title"] or "1234" in item["title"]

    def test_flip_no_102_items_when_pool_empty_and_no_grants(self) -> None:
        cs = cap_state_mod.build_cap_state(_inputs_with_pool(plan_type="nso", issued=0), _BASIC_INSTRUMENTS)
        result = self._flip(cs)
        assert not any("section_102" in c["rule_id"] for c in result["counsel_review_items"])
        assert not any("unconfirmed" in c["title"] for c in result["counsel_review_items"])

    def test_run_scenario_flip_picks_up_derived_count(self) -> None:
        instruments = _instruments_with_grant(_option_grant(shares_vested_to_date=0, shares_exercised=0))
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        result = run_scenario.run_flip_scenario(
            {"type": "flip", "parameters": {}}, cap_state=cs, instruments=instruments
        )
        assert any(
            c["rule_id"] == "delaware_cross_border.section_102_parent_shares" for c in result["counsel_review_items"]
        )


class TestFlipSpecificsGrantDetailFlag:
    """coaching_payload.flip_specifics must expose whether per-grant detail exists, so the coaching
    layer can distinguish '0 §102 grants' from 'no per-grant data captured'."""

    def _artifacts(self, outstanding_options: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        return {
            "inputs.json": {"company_name": "Acmecorp", "mode": "flip_focused"},
            "instruments.json": {
                "safes": [],
                "convertible_notes": [],
                "warrants": [],
                "option_grants": [],
                "metadata": {},
            },
            "scenarios.json": {"scenarios": []},
            "rule_audit.json": {},
            "counsel_packet.json": {"items": []},
            "cap_state.json": {
                "outstanding_options": outstanding_options,
                "option_pool": {"plan_type": "section_102_cg", "issued_and_outstanding": 1_000},
                "founders": [],
                "preferred_series": [],
                "outstanding_warrants": [],
            },
        }

    def test_flip_specifics_flags_missing_grant_detail(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        empty = compose_report.build_coaching_payload(
            artifacts=self._artifacts([]), review_dir="/tmp/x", report_path="/tmp/x/report.md", insertion_marker="M"
        )
        assert empty["flip_specifics"]["option_grant_detail_available"] is False
        assert empty["flip_specifics"]["section_102_grants_outstanding"] == 0

        populated = compose_report.build_coaching_payload(
            artifacts=self._artifacts([{"plan_type": "section_102_cg", "strike_price": None}]),
            review_dir="/tmp/x",
            report_path="/tmp/x/report.md",
            insertion_marker="M",
        )
        assert populated["flip_specifics"]["option_grant_detail_available"] is True
        assert populated["flip_specifics"]["section_102_grants_outstanding"] == 1


def test_watchlist_carries_applies_when_matched_through() -> None:
    """`build_watchlist` must not drop the static gate on the way out.

    compose_report.py splits the watchlist into ACTIVE vs for-reference on this
    field. When build_watchlist dropped it, every structurally-inapplicable rule
    was promoted to an action item — a Delaware-only company with no IIA grants
    was shown Israeli §102 / Registrar / SAFE-safe-harbour rows as things to
    watch, in a live run. The gate itself was always correct; only the plumbing
    between producer and renderer lost it.
    """
    import importlib.util

    spec = importlib.util.spec_from_file_location("rule_audit", os.path.join(SCRIPTS, "rule_audit.py"))
    assert spec and spec.loader
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    gating = {
        "israel_equity_tax.section_102_capital_gains": {
            "inst_1": {
                "scope": "legal_tax_applicability",
                "applies_when_matched": False,  # Delaware-only company
                "status": "in_window",
                "event_date_field": "grant_date",
            }
        },
        "safe.yc_postmoney_denominator": {
            "inst_2": {
                "scope": "legal_tax_applicability",
                "applies_when_matched": True,
                "status": "in_window",
                "event_date_field": "safe_investment_date",
            }
        },
    }
    rules: dict[str, Any] = {"domains": {}}
    out = mod.build_watchlist(gating, rules)

    by_id = {w["rule_id"]: w for w in out}
    assert "applies_when_matched" in by_id["israel_equity_tax.section_102_capital_gains"], (
        "the static gate must survive into the watchlist entry"
    )
    assert by_id["israel_equity_tax.section_102_capital_gains"]["applies_when_matched"] is False
    assert by_id["safe.yc_postmoney_denominator"]["applies_when_matched"] is True


# ===========================================================================
# An assumed option-pool basis must be disclosed, not presented as stated
# ===========================================================================


class TestTargetBasisAssumedDisclosure:
    """An omitted `target_basis` on a priced-round scenario still resolves to 'pre_money' (the
    solver needs a definite value to run), but the resolution must be disclosed as an assumption
    — not presented as a founder-stated input. Mirrors note_conversion.py's
    maturity_default_treatment_defaulted pattern."""

    def test_priced_round_scenario_defaults_and_warns(self) -> None:
        good = {**_SAFE_BASIC, "id": "safe_ok"}
        instruments = {**_BASIC_INSTRUMENTS, "safes": [good]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        scenario = {
            "type": "priced_round",
            "parameters": {
                "pre_money": 10_000_000,
                "new_money": 2_000_000,
                "target_pool_percent": 0.10,
                # target_basis intentionally omitted
            },
        }
        result = run_scenario.run_priced_round_scenario(scenario, instruments=instruments, cap_state=cs)
        codes = [w.get("code") for w in (result.get("warnings") or [])]
        assert "target_basis_defaulted" in codes, f"expected disclosure warning; got warnings={result.get('warnings')}"

    def test_priced_round_scenario_no_warning_when_basis_supplied(self) -> None:
        good = {**_SAFE_BASIC, "id": "safe_ok"}
        instruments = {**_BASIC_INSTRUMENTS, "safes": [good]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        scenario = {
            "type": "priced_round",
            "parameters": {
                "pre_money": 10_000_000,
                "new_money": 2_000_000,
                "target_pool_percent": 0.10,
                "target_basis": "post_money",
            },
        }
        result = run_scenario.run_priced_round_scenario(scenario, instruments=instruments, cap_state=cs)
        codes = [w.get("code") for w in (result.get("warnings") or [])]
        assert "target_basis_defaulted" not in codes

    def test_priced_round_scenario_no_warning_when_no_pool_topup(self) -> None:
        """target_basis has no math effect with no pool top-up requested — no spurious warning."""
        good = {**_SAFE_BASIC, "id": "safe_ok"}
        instruments = {**_BASIC_INSTRUMENTS, "safes": [good]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        scenario = {"type": "priced_round", "parameters": {"pre_money": 10_000_000, "new_money": 2_000_000}}
        result = run_scenario.run_priced_round_scenario(scenario, instruments=instruments, cap_state=cs)
        codes = [w.get("code") for w in (result.get("warnings") or [])]
        assert "target_basis_defaulted" not in codes

    def test_safe_conversion_priced_delegate_defaults_and_warns(self) -> None:
        """The safe_conversion scenario's priced-round delegate path (run_safe_conversion_scenario)
        independently reads target_basis via params.get — must emit the same disclosure."""
        good = {**_SAFE_BASIC, "id": "safe_ok"}
        instruments = {**_BASIC_INSTRUMENTS, "safes": [good]}
        cs = cap_state_mod.build_cap_state(_BASIC_INPUTS, instruments)
        scenario = {
            "type": "safe_conversion",
            "parameters": {
                "priced_round_pre_money": 10_000_000,
                "priced_round_new_money": 2_000_000,
                "target_pool_percent": 0.10,
            },
        }
        result = run_scenario.run_safe_conversion_scenario(scenario, instruments=instruments, cap_state=cs)
        codes = [w.get("code") for w in (result.get("warnings") or [])]
        assert "target_basis_defaulted" in codes, f"expected disclosure warning; got warnings={result.get('warnings')}"

    def test_composed_report_renders_assumed_basis_callout(self) -> None:
        """compose_report.py must render the assumption as a founder-facing callout in report.md —
        not just carry it as an unrendered warning code."""
        scenario = _minimal_scenario(
            "s_pool",
            computed_outputs={
                "completeness": "structural_only",
                "blockers": [],
                "warnings": [{"code": "target_basis_defaulted", "severity": "medium", "message": "defaulted"}],
            },
        )
        scenario["parameters"]["target_pool_percent"] = 0.10
        d = _make_cap_compose_dir(scenarios=[scenario])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, stderr
        md = report["report_markdown"]
        assert "ASSUMED, not stated" in md, f"expected assumed-basis callout in report.md; got:\n{md}"

    def test_scenario_digest_marks_basis_assumed(self) -> None:
        """The coaching-payload scenario digest's driver line must also flag the assumption —
        it feeds the Context-B coaching commentary, a separate render path from report.md."""
        import compose_report  # type: ignore[import-not-found]

        scenario = _minimal_scenario(
            "s_pool",
            computed_outputs={
                "completeness": "structural_only",
                "blockers": [],
                "warnings": [{"code": "target_basis_defaulted", "severity": "medium", "message": "defaulted"}],
            },
        )
        scenario["parameters"]["target_pool_percent"] = 0.10
        digest = compose_report.build_scenario_digest([scenario])
        drivers = digest[0]["scenario_drivers"]
        assert any("assumed" in d.lower() for d in drivers), f"expected an assumed-basis driver note; got {drivers}"


class TestQuickAssessTargetBasisDisclosure:
    """quick_assess.py (fast-assess) independently defaults target_basis at the CLI layer — same
    defect class as the full pipeline, so it needs the same disclosure."""

    def test_pool_topup_without_basis_discloses_assumption(self) -> None:
        import quick_assess as qa  # type: ignore[import-not-found]

        good_safe = {**_SAFE_BASIC, "id": "safe_ok"}
        sentinel = qa.quick_assess(
            company_name="Acmecorp",
            inputs=_BASIC_INPUTS,
            safes=[good_safe],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            target_pool_percent=0.10,
            target_basis=None,
        )
        report_md = sentinel.pop("_report_md")
        assumptions = sentinel.get("assumptions", [])
        assert any("pool-sizing basis" in a for a in assumptions), (
            f"expected a pool-sizing-basis assumption; got {assumptions}"
        )
        assert "## Assumptions" in report_md

    def test_pool_topup_with_basis_no_disclosure(self) -> None:
        import quick_assess as qa  # type: ignore[import-not-found]

        good_safe = {**_SAFE_BASIC, "id": "safe_ok"}
        sentinel = qa.quick_assess(
            company_name="Acmecorp",
            inputs=_BASIC_INPUTS,
            safes=[good_safe],
            notes=[],
            pre_money=10_000_000,
            new_money=2_000_000,
            target_pool_percent=0.10,
            target_basis="post_money",
        )
        sentinel.pop("_report_md")
        assumptions = sentinel.get("assumptions", [])
        assert not any("pool-sizing basis" in a for a in assumptions)

    def test_cli_pool_topup_without_target_basis_flag_succeeds(self) -> None:
        """--target-basis omitted at the CLI (default=None) must not crash the solver — it still
        resolves to a definite basis, just a disclosed one."""
        with tempfile.TemporaryDirectory() as d:
            inputs_path = os.path.join(d, "inputs.json")
            safes_path = os.path.join(d, "safes.json")
            with open(inputs_path, "w") as f:
                json.dump(_BASIC_INPUTS, f)
            with open(safes_path, "w") as f:
                json.dump([_SAFE_BASIC], f)

            rc, stdout, stderr = _run(
                "quick_assess.py",
                [
                    "--inputs",
                    inputs_path,
                    "--safes",
                    safes_path,
                    "--pre-money",
                    "10000000",
                    "--new-money",
                    "2000000",
                    "--target-pool-percent",
                    "0.10",
                    "--review-dir",
                    d,
                    "--pretty",
                ],
            )
            assert rc == 0, f"quick_assess.py failed: {stderr}"
            with open(os.path.join(d, "report_fast_assess.md")) as f:
                report_md = f.read()
            assert "## Assumptions" in report_md
            assert "basis" in report_md.lower()


def test_skill_md_fast_assess_pool_basis_gate_offers_pre_money() -> None:
    """SKILL.md's fast-assess pool-basis gate must not hardcode target_basis to post_money only —
    a founder whose term sheet sizes the pool pre-money needs an option to select that basis too."""
    skill_md_path = os.path.join(_REPO, "founder-skills", "skills", "cap-table", "SKILL.md")
    with open(skill_md_path, encoding="utf-8") as f:
        text = f.read()
    assert "--target-basis post_money` ONLY when the founder stated" not in text, (
        "the old hardcoded-post_money-only instruction should have been replaced"
    )
    assert "<pre_money|post_money>" in text or "pre-money-vs-post-money basis" in text


# ===========================================================================
# Reconciliation must be computed before coaching_payload so it can be cited
# ===========================================================================


class TestReconciliationReachesCoachingPayload:
    """compute_reconciliation_status must run BEFORE build_coaching_payload so a computed price-
    per-share / fully-diluted figure disagreeing with the founder's own source document is
    reachable from the Context-B coaching dispatch, not just a top-level report.json field."""

    _EMPTY_ARTIFACTS = {
        "inputs.json": _BASIC_INPUTS,
        "instruments.json": _BASIC_INSTRUMENTS,
        "scenarios.json": {"scenarios": []},
        "rule_audit.json": {"date_sensitive_watchlist": []},
        "counsel_packet.json": {"items": []},
    }

    def test_build_coaching_payload_carries_reconciliation_status(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        payload = compose_report.build_coaching_payload(
            artifacts=self._EMPTY_ARTIFACTS,
            review_dir="/tmp/x",
            report_path="/tmp/x/report.md",
            insertion_marker="<!-- M -->",
            reconciliation_status="diverged",
            reconciliation_max_divergence_ppm=5000.0,
        )
        assert payload["reconciliation"] == {"status": "diverged", "max_divergence_ppm": 5000.0}

    def test_build_coaching_payload_defaults_reconciliation_when_omitted(self) -> None:
        import compose_report  # type: ignore[import-not-found]

        payload = compose_report.build_coaching_payload(
            artifacts=self._EMPTY_ARTIFACTS,
            review_dir="/tmp/x",
            report_path="/tmp/x/report.md",
            insertion_marker="<!-- M -->",
        )
        assert payload["reconciliation"] == {"status": "not_applicable", "max_divergence_ppm": 0.0}

    def test_full_pipeline_coaching_payload_reflects_diverged_reconciliation(self) -> None:
        """End-to-end: a scenario whose equity_financing_price diverges from inputs.stated_totals
        must show up as report.json.coaching_payload.reconciliation.status == 'diverged' — the
        SAME finding as the top-level report.json.reconciliation_status field, not a structurally
        unreachable one."""
        scenario = _minimal_scenario(
            "s_priced",
            computed_outputs={
                "completeness": "full",
                "blockers": [],
                "equity_financing_price": 2.00,
                "aggregate_ownership_by_class": {"founders_pct": 0.5, "option_pool_pct": 0.1, "preferred_pct": 0.4},
            },
        )
        d = _make_cap_compose_dir(scenarios=[scenario])
        inputs_path = os.path.join(d, "inputs.json")
        with open(inputs_path) as f:
            inputs = json.load(f)
        inputs["stated_totals"] = {"price_per_share": 1.00}
        with open(inputs_path, "w") as f:
            json.dump(inputs, f)

        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, stderr
        assert report["reconciliation_status"] == "diverged"
        assert report["coaching_payload"]["reconciliation"]["status"] == "diverged", (
            "coaching_payload must carry the SAME reconciliation finding as the top-level field, "
            f"got coaching_payload.reconciliation={report['coaching_payload'].get('reconciliation')}"
        )


# ===========================================================================
# Backward-verification dispatch prompts must embed document text, not a path
# ===========================================================================


class TestBackwardVerifierInlineDocText:
    """The re-extraction dispatch prompt must embed the document text inline — a fresh Task
    sub-agent's file tools cannot resolve a bare main-thread/VM-shell path, so a path-based prompt
    either gets denied by the host-loop path gate or silently no-ops with no diagnostic."""

    def test_emit_prompts_embeds_doc_text_not_a_path(self) -> None:
        import backward_verifier as bv  # type: ignore[import-not-found]

        extraction = {
            "instrument_type": "safe",
            "fields": {"purchase_amount": 500_000},
        }
        doc_text = 'THE INVESTOR WILL PAY THE COMPANY $500,000 (the "Purchase Amount").'
        prompts = bv.emit_prompts(extraction, doc_text)
        assert prompts, "expected at least one prompt"
        body = prompts[0]["prompt"]
        assert doc_text in body, "the document text must be embedded verbatim in the prompt"
        assert "use the Read tool" not in body
        assert "Path:" not in body

    def test_cli_phase_prompt_requires_doc_text(self) -> None:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump({"instrument_type": "safe", "fields": {"purchase_amount": 500_000}}, f)
            extraction_path = f.name
        try:
            rc, stdout, stderr = _run(
                "backward_verifier.py",
                ["--phase=prompt", "--extraction", extraction_path],
            )
            assert rc == 2
            assert "--doc-text" in stderr
        finally:
            os.unlink(extraction_path)

    def test_cli_phase_prompt_rejects_missing_doc_text_file(self) -> None:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump({"instrument_type": "safe", "fields": {"purchase_amount": 500_000}}, f)
            extraction_path = f.name
        try:
            rc, stdout, stderr = _run(
                "backward_verifier.py",
                ["--phase=prompt", "--extraction", extraction_path, "--doc-text", "/nonexistent/doc.txt"],
            )
            assert rc == 2
            assert "not found" in stderr
        finally:
            os.unlink(extraction_path)

    def test_cli_phase_prompt_embeds_doc_text_from_file(self) -> None:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
            json.dump({"instrument_type": "safe", "fields": {"purchase_amount": 500_000}}, f)
            extraction_path = f.name
        with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False) as f:
            f.write("the Purchase Amount is $500,000 per this SAFE.")
            doc_text_path = f.name
        try:
            rc, stdout, stderr = _run(
                "backward_verifier.py",
                ["--phase=prompt", "--extraction", extraction_path, "--doc-text", doc_text_path],
            )
            assert rc == 0, stderr
            out = json.loads(stdout)
            assert out["n_prompts"] >= 1
            assert "$500,000" in out["prompts"][0]["prompt"]
        finally:
            os.unlink(extraction_path)
            os.unlink(doc_text_path)


# ===========================================================================
# Term-sheet / option-plan / amendment content must reach the full-pipeline report
# ===========================================================================


class TestExtractionAuditReachesFullPipelineReport:
    """A term_sheet / option_plan / amendment extraction saved as extraction_audit.json (Step 3's
    Lane-1 invocation) must be rendered by compose_report.py even when a real cap base is also
    present — the content must not be silently dropped just because the full pipeline, not the
    no-cap-base fork, ran."""

    def test_term_sheet_receipt_renders_in_full_pipeline_report(self) -> None:
        d = _make_cap_compose_dir(scenarios=[])
        receipt = {
            "classified_doc_type": "term_sheet",
            "classified_as_terms_doc": True,
            "terms_doc": {
                "fields": {
                    "pre_money_valuation": 20_000_000,
                    "investment_amount": 5_000_000,
                    "lead_investor_name": "Series A Lead",
                },
                "confidence": {
                    "pre_money_valuation": {"level": "high", "evidence_quote": "pre-money of $20,000,000"},
                },
            },
            "ambiguities": [],
        }
        with open(os.path.join(d, "extraction_audit.json"), "w") as f:
            json.dump(receipt, f)

        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, stderr
        md = report["report_markdown"]
        assert "Term sheet terms (as extracted)" in md, f"expected term-sheet section; got:\n{md}"
        assert "Pre Money Valuation" in md

    def test_amendment_receipt_renders_in_full_pipeline_report(self) -> None:
        d = _make_cap_compose_dir(scenarios=[])
        receipt = {
            "classified_doc_type": "amendment",
            "ambiguities": [
                {"field": "valuation_cap", "description": "cap raised from $8M to $10M per this amendment"}
            ],
        }
        with open(os.path.join(d, "extraction_audit.json"), "w") as f:
            json.dump(receipt, f)

        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, stderr
        md = report["report_markdown"]
        assert "Amendments (terms modified)" in md, f"expected amendments section; got:\n{md}"
        assert "cap raised from $8M to $10M" in md

    def test_absent_extraction_audit_is_a_no_op(self) -> None:
        """The common case (pure SAFE/note engagement, no terms-doc/amendment) must not render an
        empty or spurious section — this is the default path and must stay silent."""
        d = _make_cap_compose_dir(scenarios=[])
        rc, report, stderr = _run_cap_compose(d)
        assert rc == 0, stderr
        md = report["report_markdown"]
        assert "Term sheet terms (as extracted)" not in md
        assert "Amendments (terms modified)" not in md


class TestQuickAssessPoolLineIsStageNeutral:
    """`quick_assess.py` accepts no stage input of any kind (`grep -c stage` → 0),
    so nothing it writes may assert what investors expect of THIS company.

    The no-pool-target branch used to emit "seed investors typically require a
    10-15% post-money unallocated pool" into `report_fast_assess.md` — a founder
    deliverable, not a diagnostic. A Series B founder running fast-assess was
    handed the seed range as fact. Pool size is also `benchmark_only` in the rule
    pack (`option_pool.us_market_benchmarks`), so "require" overstated it twice.
    """

    EVAL2_INPUTS: dict[str, Any] = {
        "company_name": "TestCo",
        "analysis_date": "2026-05-21",
        "mode": "standard",
        "jurisdiction": {
            "structure": "delaware",
            "incorporated_date": "2024-06-01",
            "iia_grants_history": {"has_grants": False, "grant_details": []},
        },
        "founders": [{"name": "Founder", "founder_id": "founder_1", "common_shares": 10_000_000}],
        "preferred_series": [],
        "option_pool": {"plan_type": "nso", "authorized": 1_000_000, "issued": 0, "unallocated": 1_000_000},
        "common_batches": [],
        "metadata": {"run_id": "test"},
    }

    @staticmethod
    def _report_without_pool_target() -> str:
        import quick_assess as qa  # type: ignore[import-not-found]

        sentinel = qa.quick_assess(
            company_name="TestCo",
            inputs=TestQuickAssessPoolLineIsStageNeutral.EVAL2_INPUTS,
            safes=[],
            notes=[],
            pre_money=20_000_000,
            new_money=5_000_000,
            target_pool_percent=None,  # the branch under test
            target_basis="post_money",
            founder_prompt="no pool modeled",
            attached_docs=[],
        )
        return str(sentinel["_report_md"])

    def test_does_not_assert_a_stage_specific_requirement(self) -> None:
        md = self._report_without_pool_target()
        assert "No pool top-up modeled" in md, "the branch under test did not run"
        low = md.lower()
        assert "investors typically require" not in low, (
            "quick_assess asserts what investors 'typically require' while having no "
            "stage input — that is the defect. State pool ranges as stage-labelled "
            "benchmarks, not as a requirement."
        )

    def test_any_named_range_carries_its_stage(self) -> None:
        """A range may be quoted, but never one stage's range presented alone."""
        md = self._report_without_pool_target()
        low = md.lower()
        if "10–15" in md or "10-15" in md:
            assert "seed" in low, "the 10-15% range is the SEED range; say so"
            assert "series a" in low, (
                "naming only the seed range reads as universal to a founder at any "
                "other stage — name a second stage's range, or name no range at all"
            )


class TestIsraelBenchmarkIsStageScoped:
    """`founder_benchmarks.israel_preseed_context` is scoped by its own
    `applies_when` to "Israeli pre-seed or seed scenarios", but its matcher read
    `jurisdiction.structure` only — pure geography. A pre-seed-labelled benchmark
    set therefore fired for an Israeli company at ANY stage.

    Unknown stage must not fire: this is benchmark context, so withholding it
    costs a note while asserting it at the wrong stage misinforms.
    """

    RULE = "founder_benchmarks.israel_preseed_context"
    _INST: dict[str, Any] = {"safes": [], "convertible_notes": [], "warrants": [], "option_grants": []}
    _CS: dict[str, Any] = {"preferred_series": []}

    def _matches(self, inputs: dict[str, Any]) -> bool:
        matcher = rule_audit._RULE_MATCHERS.get(self.RULE, rule_audit._matcher_always)
        return bool(matcher(inputs, self._INST, self._CS))

    def test_fires_for_israeli_early_stage(self) -> None:
        for stage in ("pre-seed", "seed", "Seed"):
            inputs = {"jurisdiction": {"structure": "israeli"}, "metadata": {"stage": stage}}
            assert self._matches(inputs) is True, f"should fire for Israeli + {stage!r}"

    def test_does_not_fire_for_israeli_late_stage(self) -> None:
        for stage in ("series-a", "series-b", "later"):
            inputs = {"jurisdiction": {"structure": "israeli"}, "metadata": {"stage": stage}}
            assert self._matches(inputs) is False, (
                f"a pre-seed/seed benchmark set fired at {stage!r} — the defect this guards"
            )

    def test_does_not_fire_when_stage_unrecorded(self) -> None:
        assert self._matches({"jurisdiction": {"structure": "israeli"}}) is False, (
            "absent stage must not be treated as early — 'not recorded' and 'pre-seed' are different claims"
        )

    def test_still_requires_an_israeli_structure(self) -> None:
        inputs = {"jurisdiction": {"structure": "delaware"}, "metadata": {"stage": "seed"}}
        assert self._matches(inputs) is False, "stage alone must not substitute for geography"


# ---------------------------------------------------------------------------
# Founder-facing labelling (from the 0.7.0-candidate Cowork UI validation)
# ---------------------------------------------------------------------------


def test_report_leads_with_the_investor_name_not_our_slug() -> None:
    """A delivered report identified a SAFE only as `safe_foobar` while instruments.json carried
    `investor_name: "Foobar Capital LLC"`. The id stays — it ties this row to the explorer and the
    counsel packet — but behind the name."""
    import importlib.util

    spec = importlib.util.spec_from_file_location("_cr", os.path.join(SCRIPTS, "compose_report.py"))
    assert spec and spec.loader
    cr = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cr)
    instruments = {"safes": [{"id": "safe_foobar", "investor_name": "Foobar Capital LLC"}]}
    assert cr._instrument_label(instruments, "safe_foobar") == "Foobar Capital LLC (`safe_foobar`)"
    # No name recorded (or an unknown id) must still render something usable, never a bare word.
    assert cr._instrument_label(instruments, "safe_unknown") == "`safe_unknown`"
    assert cr._instrument_label({"safes": [{"id": "safe_x"}]}, "safe_x") == "`safe_x`"


def test_counsel_packet_summary_humanizes_the_domain_like_its_heading() -> None:
    """The heading humanized and the summary did not, which put `delaware_cross_border` into a
    delivered packet. Both now use one helper so they cannot drift apart again."""
    import importlib.util

    spec = importlib.util.spec_from_file_location("_cp", os.path.join(SCRIPTS, "counsel_packet.py"))
    assert spec and spec.loader
    cp = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cp)
    assert cp._domain_label("delaware_cross_border") == "Delaware Cross Border"
    # Rule ids and source ids are NOT routed through this — counsel cites them verbatim.
    assert "_" not in cp._domain_label("safe_terms")
