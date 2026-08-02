"""Lane-3 freeform deterministic-mapping tests (Phase 2).

Drives `freeform_mapper.map_freeform` — the pure function that maps a
SPREADSHEET_STRUCTURE_DETECTION block set + cell grid into schema-valid
inputs/instruments proposals plus an explicit blocker list. No live agent, no
LLM: a fixed (blocks, grid) input must map deterministically.
"""

from __future__ import annotations

import json
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

_REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SCRIPTS = os.path.join(_REPO, "founder-skills", "skills", "cap-table", "scripts")
sys.path.insert(0, SCRIPTS)

import cap_state as cap_state_mod  # type: ignore[import-not-found]  # noqa: E402
import freeform_mapper as fm  # type: ignore[import-not-found]  # noqa: E402
import pytest  # noqa: E402

RUN = "testrun01"


def _grid(sheets: dict[str, list[list]]) -> dict:
    """Build a --mode=grid-shaped payload from {sheet_name: rows}."""
    return {
        "ok": True,
        "mode": "grid",
        "sheets": {name: {"dimensions": "", "rows": rows, "merged_ranges": []} for name, rows in sheets.items()},
    }


def _meta_inputs() -> dict:
    """A Step-2 inputs.json carrying company meta only (no equity)."""
    return {
        "company_name": "Cadence",
        "analysis_date": "2026-06-19",
        "mode": "standard",
        "metadata": {"run_id": RUN, "schema_version": "v0.5.0-inputs"},
    }


def _blockers_for(result: dict, field: str) -> list[dict]:
    return [b for b in result["blockers"] if b["field"] == field]


# --- founders ---------------------------------------------------------------


def test_founders_block_happy() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B3",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000], ["Bob", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    founders = r["inputs"]["founders"]
    assert [(f["name"], f["common_shares"]) for f in founders] == [("Alice", 5000000), ("Bob", 5000000)]
    # company meta carried forward untouched
    assert r["inputs"]["company_name"] == "Cadence"
    assert r["inputs"]["metadata"]["schema_version"] == "v0.5.0-inputs"


def test_freeform_emit_stamps_provenance_deterministic() -> None:
    # The mapper genuinely produced the equity from the sheet → stamp cap_base_provenance so a downstream
    # consumer knows this base was deterministically mapped (not model-reconstructed).
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B3",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000], ["Bob", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["inputs"]["metadata"]["cap_base_provenance"] == "deterministic_mapped"


def test_freeform_all_preexisting_equity_no_deterministic_provenance() -> None:
    # Accumulator-gated: the equity is entirely PRE-EXISTING (the mapper produced nothing this call), so it
    # must NOT claim deterministic_mapped — otherwise a model-built base inherits a false provenance.
    existing = _meta_inputs()
    existing["founders"] = [{"name": "Alice", "common_shares": 5000000}]
    r = fm.map_freeform([], _grid({"Cap": [["x"]]}), existing_inputs=existing, run_id=RUN)
    assert r["inputs"].get("founders")  # equity present (from existing)
    assert (r["inputs"].get("metadata") or {}).get("cap_base_provenance") != "deterministic_mapped"


def test_freeform_emit_stamps_cap_base_confirmed() -> None:
    # Lane-3 carve-out for Issue B (default-to-assumed): the sheet IS the source of truth, so the
    # freeform emit must stamp metadata.cap_base_source="confirmed" — otherwise the cap_state
    # default-to-assumed warn fires spuriously on every sheet-sourced base.
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B3",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000], ["Bob", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["inputs"]["metadata"]["cap_base_source"] == "confirmed"


def test_blank_rows_in_range_skipped() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    # row 3 (index 2) fully blank (merged-cell / spacer) -> skipped, not a record
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000], [None, None], ["Bob", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert [f["name"] for f in r["inputs"]["founders"]] == ["Alice", "Bob"]


# --- safes ------------------------------------------------------------------


def test_safes_block_happy_form_discount_id_sentinel() -> None:
    blocks = [
        {
            "block_type": "safes_block",
            "sheet": "SAFEs",
            "cell_range": "A2:D2",
            "column_role_map": {"A": "investor_name", "B": "amount", "C": "post_money_cap", "D": "discount"},
        }
    ]
    grid = _grid({"SAFEs": [["Investor", "Amount", "Cap", "Disc"], ["Acme Ventures", 500000, 20000000, 0.20]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    s = r["instruments"]["safes"][0]
    assert s["id"] == "safe_000"
    assert s["investor_name"] == "Acme Ventures"
    assert s["purchase_amount"] == 500000.0
    assert s["post_money_valuation_cap"] == 20000000.0
    assert s["discount_multiplier"] == 0.8  # 0.20 rate -> 0.80 multiplier
    assert s["form"] == "cap_plus_discount"
    assert s["issuance_date"] == "1900-01-01"  # sentinel: no issue_date column
    assert s["extraction_confidence"] == "medium"
    assert s["source_document"].startswith("freeform:")


def test_two_safes_blocks_no_id_collision() -> None:
    blocks = [
        {
            "block_type": "safes_block",
            "sheet": "S",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "amount"},
        },
        {
            "block_type": "safes_block",
            "sheet": "S",
            "cell_range": "A5:B5",
            "column_role_map": {"A": "investor_name", "B": "amount"},
        },
    ]
    grid = _grid(
        {
            "S": [
                ["I", "A"],
                ["First", 100000],
                [None, None],
                [None, None],
                ["Second", 200000],
            ]
        }
    )
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    ids = [s["id"] for s in r["instruments"]["safes"]]
    assert ids == ["safe_000", "safe_001"]  # per-array, not per-block -> no collision


def test_discount_rate_forms_both_normalize() -> None:
    for raw, expect in [(20, 0.8), (0.20, 0.8)]:
        blocks = [
            {
                "block_type": "safes_block",
                "sheet": "S",
                "cell_range": "A2:C2",
                "column_role_map": {"A": "investor_name", "B": "amount", "C": "discount"},
            }
        ]
        grid = _grid({"S": [["I", "A", "D"], ["X", 100000, raw]]})
        r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
        assert r["instruments"]["safes"][0]["discount_multiplier"] == expect


# --- notes (interest_rate_type blocker + answer) ----------------------------


def test_notes_missing_interest_rate_type_blocks() -> None:
    blocks = [
        {
            "block_type": "notes_block",
            "sheet": "N",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "principal"},
        }
    ]
    grid = _grid({"N": [["I", "P"], ["Lender", 250000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "interest_rate_type"), "missing required enum must block"
    assert r["instruments"]["convertible_notes"] == []  # not emitted while blocked


def test_notes_answer_completes_block() -> None:
    blocks = [
        {
            "block_type": "notes_block",
            "sheet": "N",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "principal"},
        }
    ]
    grid = _grid({"N": [["I", "P"], ["Lender", 250000]]})
    r = fm.map_freeform(
        blocks,
        grid,
        existing_inputs=_meta_inputs(),
        answers={"0.interest_rate_type": "fixed_numeric_simple"},
        run_id=RUN,
    )
    assert r["blockers"] == []
    n = r["instruments"]["convertible_notes"][0]
    assert n["id"] == "note_000"
    assert n["interest_rate_type"] == "fixed_numeric_simple"
    assert n["principal"] == 250000.0


def test_notes_answer_rejects_non_enum() -> None:
    blocks = [
        {
            "block_type": "notes_block",
            "sheet": "N",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "principal"},
        }
    ]
    grid = _grid({"N": [["I", "P"], ["Lender", 250000]]})
    r = fm.map_freeform(
        blocks, grid, existing_inputs=_meta_inputs(), answers={"0.interest_rate_type": "weekly_compound"}, run_id=RUN
    )
    assert _blockers_for(r, "interest_rate_type"), "a non-enum answer must still block"


# --- preferred --------------------------------------------------------------


def test_preferred_missing_oip_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:C2",
            "column_role_map": {"A": "series_name", "B": "shares", "C": "issue_date"},
        }
    ]
    grid = _grid({"P": [["S", "Sh", "D"], ["Series A", 2000000, "2025-01-01"]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "original_issue_price"), "no price column -> blocker, never fabricated"


def test_preferred_cp_defaults_to_oip() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:D2",
            "column_role_map": {"A": "series_name", "B": "shares", "C": "issue_price", "D": "issue_date"},
        }
    ]
    grid = _grid({"P": [["S", "Sh", "Px", "D"], ["Series A", 2000000, 1.25, "2025-01-01"]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    ps = r["inputs"]["preferred_series"][0]
    assert ps["original_issue_price"] == 1.25
    assert ps["original_conversion_price"] == 1.25  # default 1:1 at fresh issuance
    assert ps["current_conversion_price"] == 1.25


def test_preferred_dup_series_name_conflict_keeps_existing() -> None:
    existing = _meta_inputs()
    existing["preferred_series"] = [
        {
            "series_name": "Series A",
            "shares": 1,
            "original_issue_price": 9.99,
            "original_conversion_price": 9.99,
            "current_conversion_price": 9.99,
            "issuance_date": "2020-01-01",
        }
    ]
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:D2",
            "column_role_map": {"A": "series_name", "B": "shares", "C": "issue_price", "D": "issue_date"},
        }
    ]
    grid = _grid({"P": [["S", "Sh", "Px", "D"], ["Series A", 2000000, 1.25, "2025-01-01"]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=existing, run_id=RUN)
    # conflict -> existing kept (no clobber), surfaced as a blocker
    assert any(b["block_type"] == "preferred_series_block" and "conflict" in b["reason"].lower() for b in r["blockers"])
    kept = [p for p in r["inputs"]["preferred_series"] if p["series_name"] == "Series A"]
    assert len(kept) == 1 and kept[0]["original_issue_price"] == 9.99


# --- P4: numeric --answer for original_issue_price --------------------------


def test_preferred_answer_resolves_oip() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:C2",
            "column_role_map": {"A": "series_name", "B": "shares", "C": "issue_date"},
        }
    ]
    grid = _grid({"P": [["S", "Sh", "D"], ["Series A", 2000000, "2025-01-01"]]})
    r = fm.map_freeform(
        blocks, grid, existing_inputs=_meta_inputs(), answers={"0.original_issue_price": "1.175"}, run_id=RUN
    )
    assert r["blockers"] == [], r["blockers"]
    ps = r["inputs"]["preferred_series"][0]
    assert ps["original_issue_price"] == 1.175
    assert ps["original_conversion_price"] == 1.175  # defaults forward from the answered OIP
    assert ps["current_conversion_price"] == 1.175


def test_preferred_answer_non_numeric_oip_reblocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:C2",
            "column_role_map": {"A": "series_name", "B": "shares", "C": "issue_date"},
        }
    ]
    grid = _grid({"P": [["S", "Sh", "D"], ["Series A", 2000000, "2025-01-01"]]})
    r = fm.map_freeform(
        blocks, grid, existing_inputs=_meta_inputs(), answers={"0.original_issue_price": "not-a-number"}, run_id=RUN
    )
    assert _blockers_for(r, "original_issue_price"), "a non-numeric OIP answer must re-block, never fabricate"


def test_answerable_blocker_fields_is_typed_map() -> None:
    # Contract shape (P4): a flat list would make numeric/bool fields indistinguishable from enum
    # fields; the typed map is what lets extract_cap_table.py warn on an unanswerable field name.
    rm = fm._load_role_map()
    fields = rm["answerable_blocker_fields"]
    assert isinstance(fields, dict)
    assert fields["interest_rate_type"] == "enum"
    assert fields["plan_type"] == "enum"
    assert fields["original_issue_price"] == "number"
    assert fields["pricing_unknown"] == "bool"


# --- P5 mapper hook: pricing_unknown -----------------------------------------


def test_preferred_pricing_unknown_answer_emits_sentinel() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "series_name", "B": "shares"},
        }
    ]
    grid = _grid({"P": [["S", "Sh"], ["Series A", 2000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), answers={"0.pricing_unknown": "true"}, run_id=RUN)
    assert r["blockers"] == [], "no issue_price column, but pricing_unknown must suppress the OIP blocker"
    ps = r["inputs"]["preferred_series"][0]
    assert ps["original_issue_price"] == 1.0
    assert ps["original_conversion_price"] == 1.0
    assert ps["current_conversion_price"] == 1.0
    assert ps["anti_dilution_protection"] == "none"
    assert ps["pricing_unknown"] is True


def test_preferred_pricing_unknown_false_string_still_blocks() -> None:
    # Coercion must be strict: only the literal 'true' is truthy. "false"/garbage must NOT
    # silently suppress the price gate.
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "series_name", "B": "shares"},
        }
    ]
    grid = _grid({"P": [["S", "Sh"], ["Series A", 2000000]]})
    r = fm.map_freeform(
        blocks, grid, existing_inputs=_meta_inputs(), answers={"0.pricing_unknown": "false"}, run_id=RUN
    )
    assert _blockers_for(r, "original_issue_price")


# --- option_pool ------------------------------------------------------------


def test_option_pool_unallocated_computed() -> None:
    blocks = [
        {
            "block_type": "option_pool_block",
            "sheet": "O",
            "cell_range": "A2:C2",
            "column_role_map": {"A": "plan_type", "B": "authorized", "C": "issued"},
        }
    ]
    grid = _grid({"O": [["Plan", "Auth", "Iss"], ["iso", 1000000, 250000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    pool = r["inputs"]["option_pool"]
    assert pool["plan_type"] == "iso"
    assert pool["authorized"] == 1000000
    assert pool["issued"] == 250000
    assert pool["unallocated"] == 750000  # computed authorized - issued


def test_option_pool_non_enum_plan_type_blocks() -> None:
    blocks = [
        {
            "block_type": "option_pool_block",
            "sheet": "O",
            "cell_range": "A2:C2",
            "column_role_map": {"A": "plan_type", "B": "authorized", "C": "issued"},
        }
    ]
    grid = _grid({"O": [["Plan", "Auth", "Iss"], ["ESOP", 1000000, 250000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "plan_type"), "non-enum plan_type must block"


# --- contract violations ----------------------------------------------------


def test_unknown_role_blocks() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "magic_unknown_role"},
        }
    ]
    grid = _grid({"Cap": [["N", "X"], ["Alice", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "magic_unknown_role"), "an off-contract role must block, never silently skip"


def test_warrants_and_grants_hard_block() -> None:
    for bt in ("warrants_block", "options_grants_block"):
        blocks = [{"block_type": bt, "sheet": "W", "cell_range": "A2:B2", "column_role_map": {"A": "x"}}]
        grid = _grid({"W": [["a", "b"], ["c", "d"]]})
        r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
        assert any(b["block_type"] == bt for b in r["blockers"]), f"{bt} must hard-block, not silently drop"


def test_ignore_block_types_no_op() -> None:
    blocks = [{"block_type": "noise", "sheet": "X", "cell_range": "A1:A1", "column_role_map": {}}]
    grid = _grid({"X": [[None]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    assert r["instruments"]["safes"] == []


# --- cell_range forms + determinism ----------------------------------------


def test_sheet_qualified_range_equivalent_to_bare() -> None:
    grid = _grid({"Cap": [["N", "S"], ["Alice", 5000000]]})
    rm = {"A": "holder_name", "B": "shares"}
    bare = fm.map_freeform(
        [{"block_type": "founders_block", "sheet": "Cap", "cell_range": "A2:B2", "column_role_map": rm}],
        grid,
        existing_inputs=_meta_inputs(),
        run_id=RUN,
    )
    qual = fm.map_freeform(
        [{"block_type": "founders_block", "sheet": "Cap", "cell_range": "Cap!A2:B2", "column_role_map": rm}],
        grid,
        existing_inputs=_meta_inputs(),
        run_id=RUN,
    )
    assert bare["inputs"]["founders"] == qual["inputs"]["founders"]


def test_run_twice_byte_identical() -> None:
    blocks = [
        {
            "block_type": "safes_block",
            "sheet": "S",
            "cell_range": "A2:D2",
            "column_role_map": {"A": "investor_name", "B": "amount", "C": "post_money_cap", "D": "discount"},
        }
    ]
    grid = _grid({"S": [["I", "A", "C", "D"], ["Acme", 500000, 20000000, 0.20]]})
    a = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    b = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert json.dumps(a, sort_keys=False) == json.dumps(b, sort_keys=False)


def test_blockers_sorted_deterministically() -> None:
    # two blocked blocks; blockers must be a stable sorted list (not set ordering)
    blocks = [
        {
            "block_type": "notes_block",
            "sheet": "N",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "principal"},
        },
        {
            "block_type": "option_pool_block",
            "sheet": "O",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "authorized", "B": "issued"},
        },
    ]
    grid = _grid({"N": [["I", "P"], ["L", 1000]], "O": [["A", "Iss"], [100, 10]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    keys = [(b["block_index"], b["field"]) for b in r["blockers"]]
    assert keys == sorted(keys)


# --- Phase 0d: cap_state no-equity-base guard --------------------------------

_A_SAFE = {
    "id": "safe_000",
    "investor_name": "Acme",
    "purchase_amount": 500000.0,
    "issuance_date": "2025-01-01",
    "form": "yc_postmoney_cap",
    "post_money_valuation_cap": 20000000.0,
    "extraction_confidence": "medium",
}


def _instruments(safes: list[dict]) -> dict:
    return {
        "safes": safes,
        "convertible_notes": [],
        "warrants": [],
        "option_grants": [],
        "metadata": {"run_id": RUN, "schema_version": "v0.5.0-instruments"},
    }


def test_cap_state_blocks_no_equity_base_with_instruments() -> None:
    """Absent founders AND absent option_pool + instruments → loud error, not silent zeros."""
    inputs = _meta_inputs()  # company meta only — no founders, no option_pool
    with pytest.raises(cap_state_mod.CapStateInvariantError, match="E_NO_EQUITY_BASE"):
        cap_state_mod.build_cap_state(inputs, _instruments([_A_SAFE]))


def test_cap_state_present_but_zero_pool_does_not_trip_guard() -> None:
    """A present-but-zero option_pool (founders []) must NOT trip the guard (regression)."""
    inputs = _meta_inputs()
    inputs["founders"] = []
    inputs["option_pool"] = {"plan_type": "nso", "authorized": 0, "issued": 0, "unallocated": 0}
    cs = cap_state_mod.build_cap_state(inputs, _instruments([_A_SAFE]))  # must not raise
    assert cs is not None


def test_cap_state_preferred_only_base_does_not_trip_guard() -> None:
    """A preferred-only equity base (no founders, no option_pool, no common_batches) is a valid
    non-zero base and must NOT trip E_NO_EQUITY_BASE."""
    inputs = _meta_inputs()
    inputs["founders"] = []
    inputs["preferred_series"] = [
        {
            "series_name": "Series Seed",
            "shares": 2_000_000,
            "original_issue_price": 1.0,
            "original_conversion_price": 1.0,
            "current_conversion_price": 1.0,
            "issuance_date": "2025-03-01",
        }
    ]
    cs = cap_state_mod.build_cap_state(inputs, _instruments([_A_SAFE]))  # must not raise
    assert cs["as_converted_totals"]["fully_diluted_shares"] > 0


# --- Phase 3: --mode=freeform-emit CLI wrapper (end-to-end) -----------------


def test_freeform_emit_cli_writes_schema_valid_artifacts(tmp_path: Path) -> None:
    from openpyxl import Workbook  # type: ignore[import-untyped]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Cap Table"
    rows: list[list[Any]] = [
        ["Name", "Shares"],
        ["Alice", 5000000],
        ["Bob", 5000000],
        [None, None],
        ["Investor", "Amt", "Cap"],
        ["Acme", 500000, 20000000],
    ]
    for row in rows:
        ws.append(row)
    xlsx = tmp_path / "cap.xlsx"
    wb.save(xlsx)
    (tmp_path / "inputs.json").write_text(
        json.dumps(
            {
                "company_name": "Cadence",
                "analysis_date": "2026-06-19",
                "mode": "standard",
                "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
            }
        )
    )
    blocks = json.dumps(
        {
            "blocks": [
                {
                    "block_type": "founders_block",
                    "sheet": "Cap Table",
                    "cell_range": "A2:B3",
                    "column_role_map": {"A": "holder_name", "B": "shares"},
                },
                {
                    "block_type": "safes_block",
                    "sheet": "Cap Table",
                    "cell_range": "A6:C6",
                    "column_role_map": {"A": "investor_name", "B": "amount", "C": "post_money_cap"},
                },
            ]
        }
    )
    proc = subprocess.run(
        [
            sys.executable,
            os.path.join(SCRIPTS, "extract_cap_table.py"),
            "--mode=freeform-emit",
            "--xlsx",
            str(xlsx),
            "--dir",
            str(tmp_path),
            "--run-id",
            "R1",
        ],
        input=blocks,
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0, proc.stderr
    receipt = json.loads(proc.stdout)
    assert receipt["ok"] is True
    inputs = json.loads((tmp_path / "inputs.json").read_text())
    instruments = json.loads((tmp_path / "instruments.json").read_text())
    assert [f["name"] for f in inputs["founders"]] == ["Alice", "Bob"]
    assert inputs["company_name"] == "Cadence"  # meta preserved
    assert instruments["safes"][0]["id"] == "safe_000"
    assert instruments["safes"][0]["post_money_valuation_cap"] == 20000000.0


def test_freeform_emit_cli_blocker_is_a_gate_not_error(tmp_path: Path) -> None:
    from openpyxl import Workbook  # type: ignore[import-untyped]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "N"
    ws.append(["Investor", "Principal"])
    ws.append(["Lender", 250000])
    wb.save(tmp_path / "n.xlsx")
    (tmp_path / "inputs.json").write_text(
        json.dumps(
            {
                "company_name": "Cadence",
                "analysis_date": "2026-06-19",
                "mode": "standard",
                "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
            }
        )
    )
    blocks = json.dumps(
        {
            "blocks": [
                {
                    "block_type": "notes_block",
                    "sheet": "N",
                    "cell_range": "A2:B2",
                    "column_role_map": {"A": "investor_name", "B": "principal"},
                }
            ]
        }
    )
    proc = subprocess.run(
        [
            sys.executable,
            os.path.join(SCRIPTS, "extract_cap_table.py"),
            "--mode=freeform-emit",
            "--xlsx",
            str(tmp_path / "n.xlsx"),
            "--dir",
            str(tmp_path),
            "--run-id",
            "R1",
        ],
        input=blocks,
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0  # a gate, not an error
    receipt = json.loads(proc.stdout)
    assert receipt["ok"] is False
    assert any(b["field"] == "interest_rate_type" for b in receipt["blockers"])
    assert not (tmp_path / "instruments.json").exists()  # nothing written while blocked


# --- Part 2: next_action branches on blocker kind ---------------------------
# An off-contract block_type/role is NOT founder-answerable — telling the agent to
# "ask the founder / re-run with --answer" is wrong advice. The producer must steer
# an off-contract blocker to a re-dispatch instead.


def _emit_cli(tmp_path: Path, sheet_rows: list, blocks: dict) -> dict[str, Any]:
    """Run --mode=freeform-emit over a one-sheet xlsx + blocks; return the receipt."""
    from openpyxl import Workbook  # type: ignore[import-untyped]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    for row in sheet_rows:
        ws.append(row)
    wb.save(tmp_path / "s.xlsx")
    (tmp_path / "inputs.json").write_text(
        json.dumps(
            {
                "company_name": "Cadence",
                "analysis_date": "2026-06-19",
                "mode": "standard",
                "metadata": {"run_id": "R1", "schema_version": "v0.5.0-inputs"},
            }
        )
    )
    proc = subprocess.run(
        [
            sys.executable,
            os.path.join(SCRIPTS, "extract_cap_table.py"),
            "--mode=freeform-emit",
            "--xlsx",
            str(tmp_path / "s.xlsx"),
            "--dir",
            str(tmp_path),
            "--run-id",
            "R1",
        ],
        input=json.dumps(blocks),
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0, proc.stderr
    return json.loads(proc.stdout)  # type: ignore[no-any-return]


def test_offcontract_blocktype_next_action_says_redispatch(tmp_path: Path) -> None:
    receipt = _emit_cli(
        tmp_path,
        [["Exit", "Proceeds"], [1_000_000, 500_000]],
        {
            "blocks": [
                {
                    "block_type": "waterfall_scenarios",  # off-contract: not in the closed vocabulary
                    "sheet": "S",
                    "cell_range": "A2:B2",
                    "column_role_map": {"A": "exit", "B": "proceeds"},
                }
            ]
        },
    )
    assert receipt["ok"] is False
    na = receipt["next_action"].lower()
    # off-contract → steer to re-dispatch, not to the founder/--answer path
    assert "re-dispatch" in na
    assert "spreadsheet_structure_detection" in na


def test_answerable_blocker_keeps_founder_next_action(tmp_path: Path) -> None:
    receipt = _emit_cli(
        tmp_path,
        [["Investor", "Principal"], ["Lender", 250_000]],
        {
            "blocks": [
                {
                    "block_type": "notes_block",  # on-contract; missing interest_rate_type is founder-answerable
                    "sheet": "S",
                    "cell_range": "A2:B2",
                    "column_role_map": {"A": "investor_name", "B": "principal"},
                }
            ]
        },
    )
    assert receipt["ok"] is False
    na = receipt["next_action"].lower()
    # purely founder-answerable → keep the --answer guidance, do NOT tell it to re-dispatch
    assert "--answer" in na
    assert "re-dispatch" not in na


# --- Silent-empty guards (real-doc finding): the mapper must NEVER report success with an empty/
#     partially-dropped cap base. A live freeform run wrote 0 founders/0 safes as ok:true because the
#     structure sub-agent emitted row_range/columns instead of cell_range/column_role_map. -----------


def test_emit_wrong_field_schema_blocks_not_silent_empty() -> None:
    # (1a) Model drift: block keyed row_range/columns instead of cell_range/column_role_map.
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "row_range": [2, 3],
            "columns": {"B": "holder_name", "C": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000], ["Bob", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"], "wrong field schema must raise a blocker, not silently map nothing"
    assert {"cell_range", "column_role_map"} & {b["field"] for b in r["blockers"]}
    assert "founders" not in r["inputs"]  # nothing silently written


def test_emit_valid_schema_blank_rows_global_emit_blocker() -> None:
    # (1b) Valid field schema, but cell_range points at blank rows → 0 records → must fail loud.
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:C3",
            "column_role_map": {"B": "holder_name", "C": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], [None, None, None], [None, None, None]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert any(b["field"] == "emit" for b in r["blockers"]), "0 records from an equity block must fail loud (1b)"


def test_emit_keep_existing_founders_no_false_emit_blocker() -> None:
    # (1b) false-fire guard: founders parsed from the sheet but dropped at merge (existing wins) still
    # count as "mapped this call" (accumulators are pre-merge) → NO global emit blocker.
    existing = _meta_inputs()
    existing["founders"] = [{"name": "Alice", "common_shares": 5000000}]
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=existing, run_id=RUN)
    assert not any(b["field"] == "emit" for b in r["blockers"])


def test_emit_mixed_partial_drop_warns_not_silent() -> None:
    # MR-2: a populated founders block + a safes block whose range is blank. 1b can't fire (founders
    # mapped), so the dropped safes block must surface as a WARNING, never a silent drop.
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        },
        {
            "block_type": "safes_block",
            "sheet": "Cap",
            "cell_range": "A5:B6",
            "column_role_map": {"A": "investor_name", "B": "amount"},
        },
    ]
    grid = _grid(
        {"Cap": [["Name", "Shares"], ["Alice", 5000000], [None, None], [None, None], [None, None], [None, None]]}
    )
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["inputs"].get("founders"), "founders should map"
    assert any("safes_block" in w and "0 data rows" in w for w in r["warnings"]), r["warnings"]


def test_emit_empty_does_not_stamp_cap_base_confirmed() -> None:
    # F3: nothing maps (wrong schema) → must NOT stamp cap_base_source=confirmed on an empty base.
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "Cap",
            "row_range": [2, 2],
            "columns": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Alice", 5000000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert (r["inputs"].get("metadata") or {}).get("cap_base_source") != "confirmed"


def test_safe_missing_issue_date_warns_on_sentinel() -> None:
    # Transparency symmetry (R-4): the preferred path warns when it applies _DATE_SENTINEL for a missing
    # issue date; the SAFE path applied it silently. A missing date must surface a warning on every path.
    blocks = [
        {
            "block_type": "safes_block",
            "sheet": "S",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "amount"},
        }
    ]
    grid = _grid({"S": [["Investor", "Amount"], ["Acme Ventures", 500000]]})  # no issue_date column
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["instruments"]["safes"][0]["issuance_date"] == "1900-01-01"  # sentinel applied
    assert any("issuance_date" in w and "Acme Ventures" in w for w in r["warnings"]), r["warnings"]


def test_note_missing_issue_date_warns_on_sentinel() -> None:
    blocks = [
        {
            "block_type": "notes_block",
            "sheet": "N",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "investor_name", "B": "principal"},
        }
    ]
    grid = _grid({"N": [["Investor", "Principal"], ["Lender", 250000]]})  # no issue_date column
    r = fm.map_freeform(
        blocks,
        grid,
        existing_inputs=_meta_inputs(),
        answers={"0.interest_rate_type": "fixed_numeric_simple"},
        run_id=RUN,
    )
    assert r["instruments"]["convertible_notes"][0]["issuance_date"] == "1900-01-01"
    assert any("issuance_date" in w and "Lender" in w for w in r["warnings"]), r["warnings"]


def test_wrong_field_schema_next_action_says_redispatch(tmp_path: Path) -> None:
    # CLI: the schema/empty blocker steers to a re-dispatch with the right field names — NOT the founder.
    receipt = _emit_cli(
        tmp_path,
        [["Name", "Shares"], ["Alice", 5000000]],
        {
            "blocks": [
                {
                    "block_type": "founders_block",
                    "sheet": "S",
                    "row_range": [2, 2],
                    "columns": {"A": "holder_name", "B": "shares"},
                }
            ]
        },
    )
    assert receipt["ok"] is False
    na = receipt["next_action"].lower()
    assert "cell_range" in na and "column_role_map" in na
    assert "re-dispatch" in na
    assert "--answer" not in na


def test_offcontract_role_next_action_says_redispatch(tmp_path: Path) -> None:
    # Off-contract ROLE (not block_type): the block_type is valid (founders_block) but a
    # column-role value is off-contract. This must hit the "off-contract" reason discriminator,
    # NOT the field=="block_type" one — so it isolates that branch.
    receipt = _emit_cli(
        tmp_path,
        [["Name", "Shares"], ["Alice", 1_000_000]],
        {
            "blocks": [
                {
                    "block_type": "founders_block",
                    "sheet": "S",
                    "cell_range": "A2:B2",
                    "column_role_map": {"A": "holder_name", "B": "bogus_role"},
                }
            ]
        },
    )
    assert receipt["ok"] is False
    # no block_type-field blocker here → only the reason clause can trigger re-dispatch
    assert all(b["field"] != "block_type" for b in receipt["blockers"])
    assert any("off-contract" in str(b.get("reason", "")) for b in receipt["blockers"])
    na = receipt["next_action"].lower()
    assert "re-dispatch" in na and "spreadsheet_structure_detection" in na


def test_mixed_offcontract_and_answerable_next_action(tmp_path: Path) -> None:
    # An off-contract block AND a founder-answerable blocker in the same run: steer to
    # re-dispatch (off-contract wins) but still surface the --answer path for the rest.
    receipt = _emit_cli(
        tmp_path,
        [["Exit", "Proceeds"], [1_000_000, 500_000], ["Investor", "Principal"], ["Lender", 250_000]],
        {
            "blocks": [
                {  # off-contract block_type
                    "block_type": "waterfall_scenarios",
                    "sheet": "S",
                    "cell_range": "A2:B2",
                    "column_role_map": {"A": "exit", "B": "proceeds"},
                },
                {  # on-contract but missing required interest_rate_type → founder-answerable
                    "block_type": "notes_block",
                    "sheet": "S",
                    "cell_range": "A4:B4",
                    "column_role_map": {"A": "investor_name", "B": "principal"},
                },
            ]
        },
    )
    assert receipt["ok"] is False
    fields = {b["field"] for b in receipt["blockers"]}
    assert "block_type" in fields  # the off-contract block
    assert "interest_rate_type" in fields  # the founder-answerable one
    na = receipt["next_action"].lower()
    assert "re-dispatch" in na  # off-contract steers the whole message to re-dispatch
    assert "--answer" in na  # ...while still pointing at the answerable path for the rest


# --- Part 3: drift guard — the contract vocabulary must stay surfaced in the
# sub-agent system prompt (the hand-maintained copy the sub-agent actually reads).


def test_role_map_vocab_present_in_agent_prompt() -> None:
    """Every closed-contract term — block_type KEYS, role KEYS (what the sub-agent emits,
    NOT role values or enum values), ignore types, hard-block keys — must appear word-bounded
    in agents/cap-table.md. Guards against adding a term to freeform-role-map.json without
    surfacing it in the SPREADSHEET_STRUCTURE_DETECTION prompt."""
    rm = fm._load_role_map()
    terms: set[str] = set(rm["block_types"].keys())
    for bt in rm["block_types"].values():
        terms |= set(bt["roles"].keys())  # role KEYS only
    terms |= set(rm["ignore_block_types"])
    terms |= set(rm["hard_block_block_types"].keys())

    agent_md = os.path.join(_REPO, "founder-skills", "agents", "cap-table.md")
    with open(agent_md, encoding="utf-8") as f:
        text = f.read()
    missing = [t for t in sorted(terms) if not re.search(r"\b" + re.escape(t) + r"\b", text)]
    assert not missing, f"closed-contract terms absent from agents/cap-table.md: {missing}"

    # P6: the optional block-level contract keys (freeform-role-map.json's `optional_block_keys`)
    # are NOT block_type/role/ignore/hard-block vocabulary, so the loop above never sees them —
    # extend the guard explicitly, or a detection payload could emit `role_constants`/`aggregate`
    # with the sub-agent never having been told either key exists (the exact drift class this
    # test exists to catch).
    for extra_term in ("role_constants", "aggregate", "stated_block_total"):
        assert re.search(r"\b" + re.escape(extra_term) + r"\b", text), (
            f"P6 contract term {extra_term!r} absent from agents/cap-table.md"
        )

    # `row_label` is a role KEY (preferred_series_block.roles), so it's already covered by the
    # `terms` loop above via `bt["roles"].keys()` -- verify that inclusion explicitly so this
    # guard doesn't silently stop catching row_label drift if the role map ever restructures.
    assert "row_label" in terms
    assert re.search(r"\brow_label\b", text), "P6 root-cause-fix term 'row_label' absent from agents/cap-table.md"


# --- L1-A: transposed / mis-mapped-orientation blocks must FAIL LOUD (not crash, not silent) ----------
# Scope (honest, per the reliability plan): L1-A closes the crash (text in a numeric role) and the
# type-incoherent / field-label-in-name mis-map. The section-label transpose (name col = arbitrary text
# like "RowA"/"Founders", data all-numeric) is a documented RESIDUAL owned by L1-B (correct mapping),
# NOT claimed here.


class TestTransposeFailLoud:
    def _founders(self, rows: list[list]) -> dict[str, Any]:
        blocks = [
            {
                "block_type": "founders_block",
                "sheet": "Cap",
                "cell_range": f"A1:B{len(rows)}",
                "column_role_map": {"A": "holder_name", "B": "shares"},
            }
        ]
        return fm.map_freeform(blocks, _grid({"Cap": rows}), existing_inputs=_meta_inputs(), run_id=RUN)  # type: ignore[no-any-return]

    def test_text_in_numeric_role_blocks_not_crashes(self) -> None:
        # Was an uncaught ValueError at _i(float("Bob")). Must become a blocker.
        r = self._founders([["Alice", "Bob"], ["Carol", "Dave"]])  # shares col entirely text
        assert not (r.get("inputs", {}).get("founders"))
        assert _blockers_for(r, "orientation"), "transposed (text-in-numeric) block must fail loud"

    def test_name_col_all_numeric_blocks(self) -> None:
        r = self._founders([[100, 5000000], [200, 4000000]])  # name col numeric
        assert _blockers_for(r, "orientation")

    def test_field_label_in_name_blocks(self) -> None:
        # holders-as-columns transpose: name col holds field labels
        r = self._founders([["Shares", 5000000], ["Total", 4000000]])
        assert _blockers_for(r, "orientation")

    def test_legit_two_founder_block_still_maps(self) -> None:
        r = self._founders([["Alice", 5000000], ["Bob", 4000000]])
        assert len(r["inputs"]["founders"]) == 2
        assert not r["blockers"]

    def test_entity_names_not_false_blocked(self) -> None:
        # whole-cell match, not substring: entity/number-bearing names must pass
        r = self._founders([["Class A Holdings LLC", 5000000], ["500 Startups", 4000000], ["7 Stars Ventures", 1000]])
        assert len(r["inputs"]["founders"]) == 3
        assert not r["blockers"]


# ---------------------------------------------------------------------------
# stated_totals
# ---------------------------------------------------------------------------

_MINIMAL_FOUNDERS_BLOCKS = [
    {
        "block_type": "founders_block",
        "sheet": "Cap",
        "cell_range": "A2:B3",
        "column_role_map": {"A": "holder_name", "B": "shares"},
    }
]
_MINIMAL_GRID = _grid({"Cap": [["Name", "Shares"], ["Alice", 3_000_000], ["Bob", 881_559]]})


def test_map_freeform_writes_stated_totals_when_provided() -> None:
    out = fm.map_freeform(
        _MINIMAL_FOUNDERS_BLOCKS,
        _MINIMAL_GRID,
        existing_inputs={},
        answers={},
        run_id="t",
        stated_total=3_881_559,
    )
    assert out["inputs"]["stated_totals"] == {"fully_diluted": 3_881_559, "source": "freeform_grid"}


def test_map_freeform_omits_stated_totals_when_absent() -> None:
    out = fm.map_freeform(
        _MINIMAL_FOUNDERS_BLOCKS,
        _MINIMAL_GRID,
        existing_inputs={},
        answers={},
        run_id="t",
    )
    assert "stated_totals" not in out["inputs"]


def test_freeform_emit_threads_stated_total_to_inputs(tmp_path: Path) -> None:
    """--mode=freeform-emit passes stated_total from the detection JSON to map_freeform,
    so inputs.stated_totals is populated when the detection payload carries stated_total.
    The value is the sheet's own printed grand FD total (non-circular: it comes from the
    source, not from the skill's own row sum)."""
    # _emit_cli creates a single sheet named "S"; reference it in the blocks.
    detection = {
        "blocks": [
            {
                "block_type": "founders_block",
                "sheet": "S",
                "cell_range": "A2:B3",
                "column_role_map": {"A": "holder_name", "B": "shares"},
            }
        ],
        "stated_total": 3_881_559,
    }
    receipt = _emit_cli(
        tmp_path,
        [["Name", "Shares"], ["Alice", 3_000_000], ["Bob", 881_559]],
        detection,
    )
    assert receipt["ok"] is True, receipt
    inputs = json.loads((tmp_path / "inputs.json").read_text())
    assert inputs["stated_totals"] == {"fully_diluted": 3_881_559, "source": "freeform_grid"}


# ---------------------------------------------------------------------------
# P6: wide-matrix (holder x class) support — role_constants + aggregate
# ---------------------------------------------------------------------------


def test_wide_matrix_holder_by_class_aggregates_per_series_and_founders() -> None:
    """The core P6 scenario: one sheet, holders as rows, one column per share class.
    founders_block covers holder_name + the common column only; each preferred column
    becomes its own preferred_series_block with role_constants.series_name (from the
    header) + aggregate: sum_by_constant. A preferred-only holder (blank common cell)
    must not block the founders_block. The sheet also prints a per-column Total row
    (row 6): each preferred block maps the holder-name column to `row_label` (so that
    row is recognized and excluded from the sum) AND declares `stated_block_total`
    (the column's own printed Total) as a cross-foot -- both mechanisms line up on the
    same correct series totals."""
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "M",
            "cell_range": "A2:B5",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        },
        {
            "block_type": "preferred_series_block",
            "sheet": "M",
            "cell_range": "A2:C6",
            "column_role_map": {"A": "row_label", "C": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
            "stated_block_total": 1_500_000,
        },
        {
            "block_type": "preferred_series_block",
            "sheet": "M",
            "cell_range": "A2:D6",
            "column_role_map": {"A": "row_label", "D": "shares"},
            "role_constants": {"series_name": "Series A", "issue_price": 5.0},
            "aggregate": "sum_by_constant",
            "stated_block_total": 2_000_000,
        },
    ]
    grid = _grid(
        {
            "M": [
                ["Holder", "Ordinary", "Seed", "Series A"],
                ["Alice", 3_000_000, None, None],
                ["Bob", 3_000_000, None, None],
                ["VC Fund 1", None, 1_000_000, None],
                ["VC Fund 2", None, 500_000, 2_000_000],
                ["Total", 6_000_000, 1_500_000, 2_000_000],
            ]
        }
    )
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]

    founders = {f["name"]: f["common_shares"] for f in r["inputs"]["founders"]}
    assert founders == {"Alice": 3_000_000, "Bob": 3_000_000}  # preferred-only holders skipped

    series = {p["series_name"]: p for p in r["inputs"]["preferred_series"]}
    assert series["Seed"]["shares"] == 1_500_000  # 1,000,000 + 500,000; Total row excluded + cross-footed
    assert series["Seed"]["original_issue_price"] == 2.5
    assert series["Series A"]["shares"] == 2_000_000  # Total row excluded + cross-footed
    assert series["Series A"]["original_issue_price"] == 5.0

    assert any(
        "skipped" in w and "blank common column" in w and "VC Fund 1" in w and "VC Fund 2" in w for w in r["warnings"]
    ), r["warnings"]
    assert any("Total" in w and "1,500,000" in w for w in r["warnings"]), r["warnings"]
    assert any("Total" in w and "2,000,000" in w for w in r["warnings"]), r["warnings"]


def test_founders_skip_blank_common_column_warns_not_blocks() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "F",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"F": [["Holder", "Ordinary"], ["Alice", 3_000_000], ["VC Fund", None], ["Bob", 0]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], "a blank common cell for a preferred-only holder must not hard-block founders_block"
    names = [f["name"] for f in r["inputs"]["founders"]]
    assert names == ["Alice", "Bob"]  # VC Fund skipped; Bob's explicit 0 preserved (_i(0) == 0)
    assert r["inputs"]["founders"][1]["common_shares"] == 0
    assert any("skipped" in w and "blank common column" in w and "VC Fund" in w for w in r["warnings"]), r["warnings"]


def test_role_constants_off_contract_role_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "role_constants": {"totally_bogus_role": "x"},
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "totally_bogus_role")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


def test_role_constants_non_dict_shape_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "role_constants": "series_name=Seed",  # wrong shape: must be an object
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "role_constants")


def test_aggregate_only_allowed_on_preferred_series_block() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "F",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "shares"},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"F": [["N", "S"], ["Alice", 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "aggregate")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


def test_aggregate_unknown_value_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "role_constants": {"series_name": "Seed"},
            "aggregate": "average_by_constant",
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "aggregate")


def test_block_rows_role_constants_overlay_after_blank_skip() -> None:
    """The exact P6 ordering trap: role_constants must be overlaid AFTER the blank-row skip,
    judged on column-sourced values only, or every blank spacer row inside the range becomes
    a bogus non-blank record."""
    block = {
        "block_type": "preferred_series_block",
        "sheet": "P",
        "cell_range": "A2:A5",
        "column_role_map": {"A": "shares"},
        "role_constants": {"series_name": "Seed"},
    }
    grid = _grid({"P": [["Sh"], [1_000_000], [None], [500_000], [None]]})
    rows = fm._block_rows(block, grid)
    assert len(rows) == 2, "the two blank spacer rows must be skipped entirely, not emitted as 0-share records"
    assert all(row["shares"] is not None for row in rows)
    assert all(row["series_name"] == "Seed" for row in rows)  # constant still stamped on real rows


def test_block_rows_column_value_wins_over_role_constant() -> None:
    block = {
        "block_type": "preferred_series_block",
        "sheet": "P",
        "cell_range": "A2:B2",
        "column_role_map": {"A": "shares", "B": "series_name"},
        "role_constants": {"series_name": "Seed"},
    }
    grid = _grid({"P": [["Sh", "S"], [1_000_000, "Series A-1"]]})
    rows = fm._block_rows(block, grid)
    assert rows == [{"shares": 1_000_000, "series_name": "Series A-1"}]  # column value wins


def test_aggregate_sums_across_blank_spacer_rows() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A5",
            "column_role_map": {"A": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000], [None], [500_000], [None]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    assert len(r["inputs"]["preferred_series"]) == 1  # collapsed to ONE record, dup-guard suppressed
    assert r["inputs"]["preferred_series"][0]["shares"] == 1_500_000  # blanks contribute nothing


def test_orientation_guard_tolerates_1column_aggregated_block() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A3",
            "column_role_map": {"A": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000], [500_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert not _blockers_for(r, "orientation"), "a 1-column aggregated shares block must not false-trip orientation"
    assert r["blockers"] == []
    assert r["inputs"]["preferred_series"][0]["shares"] == 1_500_000


def test_aggregate_subtotal_row_skipped_via_row_label_fixed() -> None:
    """P6 root-cause fix (was: test_aggregate_subtotal_row_double_counts_documented_limitation,
    a documented limitation). An aggregated preferred_series_block that maps the holder-name
    column to `row_label` (alongside `shares`) now recognizes a numeric 'Total' row sitting
    inside the same cell_range by its whole-cell label and excludes it from the sum, emitting
    a warning that names the skipped label and its share value -- instead of silently
    double-counting it."""
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "row_label", "B": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"P": [["Label", "Sh"], ["Fund A", 500_000], ["Fund B", 500_000], ["Total", 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]
    assert r["inputs"]["preferred_series"][0]["shares"] == 1_000_000  # 500k + 500k; Total row excluded
    assert any("Total" in w and "1,000,000" in w for w in r["warnings"]), r["warnings"]


def test_aggregate_unlabeled_total_row_still_double_counts_residual_limitation() -> None:
    """RESIDUAL LIMITATION (documented in agents/cap-table.md): the `row_label` discriminator
    only catches a Total row that actually CARRIES a recognizable label. A blank-label total
    row (distinguished in the source only by bold formatting/borders that never reach the
    extracted grid) still passes the row_label check -- `raw.get("row_label")` is blank, not a
    stoplist string -- and is summed like a real holder. No `stated_block_total` was supplied
    here to catch it via cross-foot either, so this is the honest remaining gap."""
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "row_label", "B": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"P": [["Label", "Sh"], ["Fund A", 500_000], ["Fund B", 500_000], [None, 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]
    # Real total is 1,000,000; the unlabeled subtotal row inside the range still doubles it.
    assert r["inputs"]["preferred_series"][0]["shares"] == 2_000_000


# --- stated_block_total -------------------------------------------------------------------


def test_stated_block_total_matches_clean() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "row_label", "B": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
            "stated_block_total": 1_000_000,
        }
    ]
    grid = _grid({"P": [["Label", "Sh"], ["Fund A", 500_000], ["Fund B", 500_000], ["Total", 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]
    assert r["inputs"]["preferred_series"][0]["shares"] == 1_000_000


def test_stated_block_total_mismatch_blocks() -> None:
    """The exact scenario stated_block_total is designed to catch: an UNLABELED total row (the
    row_label residual from the test above) left inside cell_range. The cross-foot against the
    sheet's own stated total catches what row_label alone could not."""
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "row_label", "B": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
            "stated_block_total": 1_000_000,
        }
    ]
    grid = _grid({"P": [["Label", "Sh"], ["Fund A", 500_000], ["Fund B", 500_000], [None, 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "stated_block_total")
    assert not r["inputs"].get("preferred_series")


def test_stated_block_total_on_nonaggregate_preferred_block_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "stated_block_total": 1_000_000,  # aggregate not set -- off-contract
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "stated_block_total")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


def test_stated_block_total_on_founders_block_blocks() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "F",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "shares"},
            "stated_block_total": 1_000_000,
        }
    ]
    grid = _grid({"F": [["N", "S"], ["Alice", 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "stated_block_total")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


def test_stated_block_total_non_numeric_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
            "stated_block_total": "not-a-number",
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "stated_block_total")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


# --- row_label: orientation-guard trap, whole-cell semantics, label-only rows, per-row path ----


def test_row_label_total_does_not_trip_orientation_guard() -> None:
    """Pin the trap: `row_label` is deliberately NOT in `_NAME_ROLES` (see that set's comment
    in freeform_mapper.py), so a 'Total' value in a row_label column must never hard-block the
    whole block via _orientation_blocker -- the row-level stoplist skip (scoped to
    preferred_series_block, downstream of the orientation check) is what handles it instead."""
    rows = [
        {"row_label": "Fund A", "shares": 500_000},
        {"row_label": "Fund B", "shares": 500_000},
        {"row_label": "Total", "shares": 1_000_000},
    ]
    assert fm._orientation_blocker(rows) is None


def test_row_label_whole_cell_semantics_entity_name_not_dropped() -> None:
    """Whole-cell match only: 'Total Ventures Fund I' is a real fund name and must NOT be
    skipped, while an exact 'Total' cell in the same block IS skipped-with-warning."""
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "row_label", "B": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid(
        {
            "P": [
                ["Label", "Sh"],
                ["Total Ventures Fund I", 500_000],
                ["Fund B", 500_000],
                ["Total", 1_000_000],
            ]
        }
    )
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]
    # "Total Ventures Fund I" counted (500k); exact "Total" excluded -> 500k + 500k = 1,000,000.
    assert r["inputs"]["preferred_series"][0]["shares"] == 1_000_000
    assert any("Total Ventures Fund I" not in w and "Total" in w and "1,000,000" in w for w in r["warnings"]), r[
        "warnings"
    ]


def test_row_label_only_row_contributes_nothing_no_warning() -> None:
    """A holder-name-present, blank-shares row contributes nothing to the sum and, since it
    carries no share value to exclude, produces NO warning (pure noise otherwise). A
    fully-blank-except-label row is caught by the amended _block_rows blank-skip."""
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:B4",
            "column_role_map": {"A": "row_label", "B": "shares"},
            "role_constants": {"series_name": "Seed", "issue_price": 2.5},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"P": [["Label", "Sh"], ["Fund A", 500_000], ["Fund B (no shares)", None], [None, None]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]
    assert r["inputs"]["preferred_series"][0]["shares"] == 500_000
    assert not any("skipped" in w for w in r["warnings"]), r["warnings"]


def test_row_label_stoplist_skip_on_per_row_path_not_dup_blocker() -> None:
    """The row-level stoplist skip covers the per-row (non-aggregated) path too, not just the
    aggregated path. Without the fix, the 'Total' row here would reach the per-row loop as a
    second series_name='Seed' record and trip the duplicate-series-name conflict blocker; with
    the fix it's skipped-with-warning before that loop ever sees it."""
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:C3",
            "column_role_map": {"A": "row_label", "B": "series_name", "C": "shares"},
            "role_constants": {"issue_price": 2.5},
        }
    ]
    grid = _grid({"P": [["Label", "Series", "Sh"], ["Fund A", "Seed", 500_000], ["Total", "Seed", 500_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == [], r["blockers"]
    assert len(r["inputs"]["preferred_series"]) == 1
    assert r["inputs"]["preferred_series"][0]["shares"] == 500_000
    assert any("Total" in w and "skipped" in w for w in r["warnings"]), r["warnings"]


def test_row_label_off_contract_on_founders_block_blocks() -> None:
    blocks = [
        {
            "block_type": "founders_block",
            "sheet": "F",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "row_label", "B": "shares"},
        }
    ]
    grid = _grid({"F": [["N", "S"], ["Alice", 1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "row_label")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


def test_row_label_off_contract_on_safes_block_blocks() -> None:
    blocks = [
        {
            "block_type": "safes_block",
            "sheet": "S",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "row_label", "B": "amount"},
        }
    ]
    grid = _grid({"S": [["N", "Amt"], ["Investor 1", 100_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "row_label")
    assert any("off-contract" in b["reason"] for b in r["blockers"])


def test_aggregate_missing_series_name_blocks() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "aggregate": "sum_by_constant",  # no role_constants.series_name and no series_name column
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "series_name")


def test_aggregate_pricing_unknown_answer() -> None:
    blocks = [
        {
            "block_type": "preferred_series_block",
            "sheet": "P",
            "cell_range": "A2:A2",
            "column_role_map": {"A": "shares"},
            "role_constants": {"series_name": "Seed"},
            "aggregate": "sum_by_constant",
        }
    ]
    grid = _grid({"P": [["Sh"], [1_000_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), answers={"0.pricing_unknown": "true"}, run_id=RUN)
    assert r["blockers"] == []
    ps = r["inputs"]["preferred_series"][0]
    assert ps["original_issue_price"] == 1.0
    assert ps["anti_dilution_protection"] == "none"
    assert ps["pricing_unknown"] is True


def test_common_holders_block_maps_to_common_batches() -> None:
    # Non-founder common/ordinary holders (angels, ex-employees, nominee trusts) get their own block
    # → inputs.common_batches, instead of being misclassified into founders.
    blocks = [
        {
            "block_type": "common_holders_block",
            "sheet": "Cap",
            "cell_range": "A2:B3",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Angel One", 100_000], ["Ex Employee", 50_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert r["blockers"] == []
    cb = r["inputs"]["common_batches"]
    assert [b.get("holder_name") for b in cb] == ["Angel One", "Ex Employee"]
    assert cb[0]["shares"] == 100_000
    assert all(b["holder_id"] for b in cb)  # auto-assigned, unique
    assert len({b["holder_id"] for b in cb}) == 2
    assert not r["inputs"].get("founders")  # NOT misclassified as founders


def test_common_holders_block_missing_shares_blocks() -> None:
    blocks = [
        {
            "block_type": "common_holders_block",
            "sheet": "Cap",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Angel One", None]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=_meta_inputs(), run_id=RUN)
    assert _blockers_for(r, "shares")  # a blank share count blocks, never silently drops


def test_common_holders_block_merges_with_existing_common_batches() -> None:
    existing = _meta_inputs()
    existing["common_batches"] = [{"holder_id": "seed_1", "shares": 10_000, "issuance_date": "2023-01-01"}]
    blocks = [
        {
            "block_type": "common_holders_block",
            "sheet": "Cap",
            "cell_range": "A2:B2",
            "column_role_map": {"A": "holder_name", "B": "shares"},
        }
    ]
    grid = _grid({"Cap": [["Name", "Shares"], ["Angel One", 100_000]]})
    r = fm.map_freeform(blocks, grid, existing_inputs=existing, run_id=RUN)
    assert r["blockers"] == []
    cb = r["inputs"]["common_batches"]
    assert [b["holder_id"] for b in cb][0] == "seed_1"  # existing kept
    assert any(b.get("holder_name") == "Angel One" for b in cb)  # new appended
    assert len({b["holder_id"] for b in cb}) == 2  # no id collision
