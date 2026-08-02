#!/usr/bin/env python3
"""Generate the Acmecorp .xlsx fixtures from the canonical numbers in README.md.

Hand-typing an 18-month model into a spreadsheet is how the deck and the model
drift apart. Everything here derives from the constants below; change them here
and in README.md together.

    python3 tools/landing-capture/fixtures/build_xlsx.py
"""

from __future__ import annotations

import pathlib

import openpyxl

HERE = pathlib.Path(__file__).parent

# --- canonical constants (mirror README.md) ---------------------------------
START_MRR = 42_000
CUSTOMERS = 140
ARPU = 300
GROWTH = 0.11
# Growth decays ~2.5% relative per month. A model that assumes 11% MoM forever is
# exactly the thing a reviewer should flag, and 11% compounded for 18 months puts
# month 18 at $248K — which contradicted the deck. Decay makes the plan realistic
# AND lets the deck state a figure the model actually reaches.
GROWTH_DECAY = 0.975
CHURN = 0.022
GROSS_MARGIN = 0.78
OPEX_START = 118_000
OPEX_GROWTH = 0.04  # modest hiring plan
CASH = 980_000
MONTHS = 18
# Month 1 of the plan. Bump these when the fixture is regenerated so the model
# reads as current rather than six months stale.
START_YEAR = 2026
START_MONTH = 8

FOUNDER_1 = 4_600_000
FOUNDER_2 = 3_400_000
POOL_TOTAL = 1_000_000
POOL_ISSUED = 600_000


def build_model() -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Assumptions"
    for row in [
        ("Parameter", "Value"),
        ("Company", "Acmecorp"),
        ("Starting MRR", START_MRR),
        ("Customers", CUSTOMERS),
        ("ARPU", ARPU),
        ("Monthly Growth Rate", GROWTH),
        ("Monthly Logo Churn", CHURN),
        ("Gross Margin", GROSS_MARGIN),
        ("Starting Monthly OpEx", OPEX_START),
        ("Monthly OpEx Growth", OPEX_GROWTH),
        ("Cash in Bank", CASH),
        ("Net Revenue Retention", 1.04),
        ("CAC", 2_600),
        ("Headcount", 9),
    ]:
        ws.append(row)

    rev = wb.create_sheet("Revenue")
    rev.append(("Month", "Customers", "ARPU", "MRR"))
    pnl = wb.create_sheet("P&L")
    pnl.append(("Month", "Revenue", "COGS", "Gross Profit", "OpEx", "Net Income"))
    cash = wb.create_sheet("Cash")
    cash.append(("Month", "Opening Cash", "Net Income", "Closing Cash"))

    mrr = float(START_MRR)
    custs = float(CUSTOMERS)
    opex = float(OPEX_START)
    balance = float(CASH)

    # Month 1 is "now". The first version of this started the labels at 2026-02
    # regardless of when the fixture was built, which made a reviewer stop and ask
    # what date the cash balance was as of — a fair question, and a wasted gate.
    for i in range(MONTHS):
        month_index = START_MONTH - 1 + i
        year = START_YEAR + month_index // 12
        month = month_index % 12 + 1
        label = f"{year}-{month:02d}"

        rev.append((label, round(custs), ARPU, round(mrr)))

        cogs = mrr * (1 - GROSS_MARGIN)
        gross = mrr - cogs
        net = gross - opex
        pnl.append((label, round(mrr), round(cogs), round(gross), round(opex), round(net)))

        opening = balance
        balance = opening + net
        cash.append((label, round(opening), round(net), round(balance)))

        custs = custs * (1 + GROWTH * (GROWTH_DECAY**i))
        mrr = custs * ARPU
        opex = opex * (1 + OPEX_GROWTH)

    out = HERE / "acmecorp-model.xlsx"
    wb.save(out)
    print(f"wrote {out}")
    print(f"  month {MONTHS} MRR = ${round(mrr / (1 + GROWTH * (GROWTH_DECAY ** (MONTHS - 1)))):,}")


def build_cap_table() -> None:
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary Cap Table"
    ws.append(())
    ws.append(("", "Acmecorp Summary Cap Table"))
    ws.append(("", "As of 2026-08-01 • synthetic fixture"))
    ws.append(())
    ws.append(
        (
            "Stakeholder",
            "Security",
            "Shares Issued and Outstanding",
            "Fully Diluted Shares",
        )
    )
    ws.append(("Founder 1 (CEO)", "Common", FOUNDER_1, FOUNDER_1))
    ws.append(("Founder 2 (CTO)", "Common", FOUNDER_2, FOUNDER_2))
    ws.append(("Option Pool — issued", "Options", POOL_ISSUED, POOL_ISSUED))
    ws.append(
        (
            "Option Pool — available",
            "Options",
            0,
            POOL_TOTAL - POOL_ISSUED,
        )
    )
    ws.append(("Total", "", FOUNDER_1 + FOUNDER_2 + POOL_ISSUED, FOUNDER_1 + FOUNDER_2 + POOL_TOTAL))

    led = wb.create_sheet("Convertible Ledger")
    led.append(())
    led.append(("", "Acmecorp Convertible Ledger"))
    led.append(("", "As of 2026-08-01 • synthetic fixture"))
    led.append(())
    led.append(
        (
            "Security ID",
            "Stakeholder Name",
            "Instrument",
            "Principal",
            "Post-Money Valuation Cap",
            "Discount Rate",
            "Issue Date",
        )
    )
    led.append(("SAFE-1", "Foobar Capital LLC", "SAFE", 500_000, 8_000_000, "", "2025-03-04"))
    led.append(("SAFE-2", "Northgate Angels LLC", "SAFE", 250_000, 10_000_000, "80%", "2025-09-18"))

    out = HERE / "acmecorp-cap-table.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    build_model()
    build_cap_table()
